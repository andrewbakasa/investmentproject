   
from investments_appraisal.excel_model import ExcelReport
from excel_response import ExcelResponse
import numpy as np
from openpyxl import Workbook, load_workbook 
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle,PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator

from openpyxl.utils import cell
import math
#from investments_appraisal.base_report import BaseBusinessReport



from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


class BeefExcelReport(ExcelReport):
    project_title = 'PERI-URBAN CATTLE FARMING OUTPUTS'
    track_inputs = { 
                    'timing':{
                        'base_period':{},
                        'construction_year_end':{},
                        'operation_start_year':{},
                        'operation_end':{},
                        'number_of_months_in_a_year':{},
                        },

                    'flags':{
                        'years':{},
                        },

                    'financing':{
                        'repayment_starts':{},
                        'num_of_installments':{},
                        'grace_period':{},
                        'real_interest_rate':{},
                        'risk_premium':{},
                        'equity':{},
                        'senior_debt':{},
                        }, 

                    'depreciation':{
                        'economic_life_of_machinery':{},
                        'economic_life_of_buildings':{},
                        'tax_life_of_machinery':{},
                        'tax_life_of_buildings':{},
                        'tax_life_of_soft_capital_costs':{},
                        },

                    'working_capital':{
                        'accounts_receivable':{},
                        'accounts_payable':{},
                        'cash_balance':{}
                        },
                    'taxes':{
                        'import_duty':{},
                        'sales_tax':{},
                        'corporate_income_tax':{}
                        },
                    
                    'macroeconomic_parameters':{
                        'header': {},
                        'discount_rate_equity': {}, 
                        'domestic_inflation_rate': {},
                        'us_inflation_rate': {}, 
                        'exchange_rate': {},
                        'dividend_payout_ratio': {},
                        'num_of_shares': {},
                        },

                     'feedlot_design_parameters':{
                        'header': {},
                        'num_of_feedlots': {}, 
                        'length': {},
                        'width': {},
                        'sqm': {}, 
                        'pen_area': {}, 
                        'sqm_per_cattle': {},
                        'total_cattle_per_pen_per_cycle': {},
                        'num_of_months_per_cycle': {},
                        'cattle_per_pen_per_year': {},
                        },

                     'prices':{
                        'header': {},
                        'base_price':{},
                        'change_in_price': {}, 
                        },    

                    'production_inventory':{
                        'header': {},
                        'cattle_per_pen_per_year': {}, 
                        'cattle_survival_rate': {},
                        'thousand': {}, 
                        'dressed_weight_at_selling': {},
                        'num_of_feedlot_targeted_for_construction': {},
                        'cum_pens_under_cattle': {},
                        'cum_pens_under_harvesting': {},
                        'production_quantity': {},
                        'closing_inventory': {},
                        'investment_roll_out_flag': {},
                        'total_pen_constructed': {},
                        },

                     'investment_cost':{
                        'header': {},
                        'total_land_for_pens': {}, 
                        'cost_of_land_per_sqm': {},
                        'cost_of_machinery_per_pen': {}, 
                        'cost_of_building_per_sqm': {},
                        'total_pens': {},
                        'cost_of_pen_construction': {},
                        'senior_debt_dynamic_parameter': {},
                        'initial_pens_employed': {},
                        'pen_cattle_density': {},
                        'pen_length': {},
                        'pen_width': {},
                        'pen_height': {},
                        'total_land_required': {},
                        'cost_of_pens_constructed': {},
                        'cost_of_land': {},
                        'investment_cost_of_land': {},
                        'cif_cost_of_machinery': {},
                        'investment_cost_of_buildings': {},
                        'investment_costs_over_run_factor': {},
                        }, 
                    'cattle_business_options': {'header': {},},
                    'sens':{
                        'header': {},
                        'domestic_inflation_rate': {}, 
                        'us_inflation_rate': {},
                        'change_in_price': {}, 
                        'base_price':{},
                        'accounts_receivable': {},
                        'accounts_payable': {},
                        'cash_balance': {},
                        'exchange_rate': {},
                        'senior_debt': {},
                        'total_cattle_per_pen_per_cycle': {},
                        'cattle_survival_rate': {},
                        'cattle_price_per_unit': {},
                        'initial_pens_employed': {}
                        },



                     'outputs':{
                        'header': {},
                        'model_selection': {},
                        },   


                    }
    _production_inventory = {
        'cattle_per_pen_per_year': {'title': 'Cattle per pen per year','value': 0, 'units': 'NUMBER'},
        'cattle_survival_rate': {'title': 'Cattle Survival rate','value': 1, 'units': 'PERCENT'},#duplicate
        'thousand': {'title': 'Thousand','value': 1000, 'units': 'NUMBER'}, 
        'dressed_weight_at_selling': {'title': 'Dressed Weight @ selling','value': 211.7 , 'units': 'KG'},
        'num_of_feedlot_targeted_for_construction': {'title': 'No. of feedlot(pens) targeted for construction','value': None, 'units': 'NUMBER'},
        'cum_pens_under_cattle': {'title': 'Cumulative pens under Cattle','value': None , 'units': 'NUMBER'}, 
        'cum_pens_under_harvesting': {'title': 'Cumulative pens under harvesting','value': None, 'units': 'NUMBER'},
        'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'TONS'},
        'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'TONS'},
        'investment_roll_out_flag': {'title': 'Investment Roll out Flag','value': None, 'units': 'NUMBER'},
        'total_pen_constructed': {'title': 'Total Pen Constructed','value':None , 'units': 'NUMBER'},
    }	
    def __init__(self, modelx):
        ExcelReport.__init__(modelx)
        self.sister_model =modelx

    def _add_feedlotDesignParameters_section(self, w_sheet, row_index,total_wsheet_cols):

        #upate user model decription
        self.sister_model.cattle_business_options['my_option']['description']=self.user_model_decription
        
        #overide all add tis to class
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Pen/Feedlot Design Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['num_of_feedlots','length','width','sqm','pen_area',
                   'sqm_per_cattle','total_cattle_per_pen_per_cycle',
                   'num_of_months_per_cycle','cattle_per_pen_per_year']
        
        for item in itemlist:
            if item in self.sister_model.feedlot_design_parameters:
                _unit =self.sister_model.feedlot_design_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.feedlot_design_parameters[item]['title'], self.sister_model.feedlot_design_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')

                  

                self.track_inputs['feedlot_design_parameters'][item]['row'] = row_index
                self.track_inputs['feedlot_design_parameters'][item]['unit'] = _unit
                self.track_inputs['feedlot_design_parameters'][item]['col'] = 6#redundant
                self.track_inputs['feedlot_design_parameters'][item]['value'] = self.sister_model.feedlot_design_parameters[item]['value'] 
                
                row_index +=1
                #skipp empty spaces 
                if item in ['sqm','pen_area','sqm_per_cattle','total_cattle_per_pen_per_cycle',]:
                    row_index +=1
                   
        
        #fill this place by retrievint from trackinputs-----
        self._update_feedlot_linkedcells(w_sheet)
        #-----------------------------------------------------
    
        #----leave blank for now: update later in Outputs
        self.track_inputs['cattle_business_options']['header']['row'] = row_index
        row_index +=1
        #------------
        #update here ..........................
        #-----------------------------------------------------------
        for item in self.sister_model.cattle_business_options.keys():
            w_sheet['%s%s'%(colHeader, row_index)].value = self.sister_model.cattle_business_options[item]['description']
            w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
            self.track_inputs['cattle_business_options'][item]['row'] = row_index
            self.track_inputs['cattle_business_options'][item]['heading'] = self.sister_model.cattle_business_options[item]['heading'] 
            self.track_inputs['cattle_business_options'][item]['description'] = self.sister_model.cattle_business_options[item]['description']            
            row_index +=1
        
        #---------------hide text for business options------------------------------------
        option_total= len(self.sister_model.cattle_business_options.keys())
        start_index= row_index - option_total-1
        self._hide_rows(w_sheet,start_index, row_index) 
        #----------------------------------------------- 
       
        return row_index

    def _write_output_sheet(self, wb,request):
        total_wsheet_cols = 17
        #---------------------------Worksheet 1---------------------------
        output_sheet = wb['Outputs']	
        col = get_column_letter(2)

        row_index = 2
        output_sheet['%s%s'%(col, row_index)].value = self.project_title# 'PERI-URBAN CATTLE FARMING OUTPUTS'# dynamic heading
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
          
        self._set_thick_bottom_border_range( output_sheet,  row_index, 1, total_wsheet_cols)
        
        row_index += 2
        output_sheet['%s%s'%(col, row_index)].value = 'Scenario Analysis '
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
        
        col = get_column_letter(6)
        output_sheet['%s%s'%(col, row_index)].value = 'Active Scenario'
        output_sheet['%s%s'%(col, row_index)].alignment = Alignment(horizontal='left', vertical='top',)
        output_sheet['%s%s'%(col, row_index)].font = Font(name='Calibri',bold=False, italic=True, sz=10.0, color='FF0070C0', scheme='minor')
        
        col = get_column_letter(7)
        _unit= "NOTE"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet['%s%s'%(col, row_index)].value = cell_display_formuale
        output_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'

       
        self._set_thick_bottom_border_range( output_sheet,  row_index, 1, total_wsheet_cols)
        col = get_column_letter(1)
        row_index+=2
        
        #-------------------Data Validation---------------------------
       
        i=0
        values_ =''
        for item in self.sister_model.cattle_business_options.keys():
            if i ==len(self.sister_model.cattle_business_options.keys())-1:
                values_ += self.sister_model.cattle_business_options[item]['heading'] 
            else:
               values_ += self.sister_model.cattle_business_options[item]['heading'] + ','     
            i+=1
        dv = DataValidation(type="list", formula1='"'+ values_ + '"', allow_blank=True)
        
        user_name=request.user
        # Optionally set a custom prompt message
        dv.prompt = f'{user_name} Please select from the list'
        dv.promptTitle = 'List Selection'
        
        # Add the data-validation object to the worksheet
        output_sheet.add_data_validation(dv)
        # Create some cells, and add them to the data-validation object
        model_sel_row= row_index
        model_sel_col= "F"
        model_sel_cell= "$" + model_sel_col + "$" + str(model_sel_row)
        model_cell = output_sheet[model_sel_cell]
        self.track_inputs['outputs']['model_selection']['row']= model_sel_row 
        self.track_inputs['outputs']['model_selection']['col']= model_sel_col
        self.track_inputs['outputs']['model_selection']['cell']= model_sel_cell

        model_cell.value = values_.split(",")[0]
        model_cell.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        model_cell.font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
        model_cell.alignment = Alignment(horizontal='left', vertical='top',)
        dv.add(model_cell)
        
       
       #----switch

       
        col = get_column_letter(3)
        output_sheet['%s%s'%(col, row_index)].value = 'A.'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(4)
        output_sheet['%s%s'%(col, row_index)].value = 'Alternative Investment Scenario'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(7)
        _unit= "SWITCH"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet['%s%s'%(col, row_index)].value = cell_display_formuale
        output_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'
       #------------------

        #####################333update input------------------------
        colHeader = get_column_letter(5)
        select_val ="="
        list_= self.sister_model.cattle_business_options.keys()
        #print(list_)
        closin_brackets=''
        for  i in range(len(list_)-1):
            closin_brackets+=')'

        i = 0
        input_row_index= self.track_inputs['cattle_business_options']['header']['row']
        for item in list_:
            #item =list_[i]
            if i ==len(list_)-1:
                select_val += colHeader + str(input_row_index +i+1) + closin_brackets
            else:
                cell_x= model_sel_cell
                select_val += 'IF(Outputs!' + cell_x + '="'+ self.sister_model.cattle_business_options[item]['heading'] +'",' + colHeader + str(input_row_index +i+1) + ','
            i+=1
        #---------------------------------------------------------------
         
        wb["Inputs"]['%s%s'%(colHeader, input_row_index)].value = select_val
        wb["Inputs"]['%s%s'%(colHeader, row_index)].font = self.description_font
        ##########################################3


      
        row_index+=2

        self._hide_empty_cells( output_sheet)
        #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':10})
        rangelist.append({'start':1, 'end': 2 , 'dim':5})
        rangelist.append({'start':2, 'end': 5, 'dim':2})
        #rangelist.append({'start':2, 'end': 3, 'dim':5})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':5, 'dim':55})
        indexlist.append({'index':6, 'dim':30})
        indexlist.append({'index':7, 'dim':15})

        #set dims----
        self._set_column_dim(output_sheet,rangelist,indexlist)
        #------------------------------------
        start_row = row_index
        #end_row = start_row +10
        
        # ------------Financial Section------------------ 
        end_row =self._add_output_financial(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row, end_row,'Financial Analysis Output',)
        
        #--------User View of Selected Project Option-------------------------------------------------------
        start_row= row_index
        source_row_ =self.track_inputs['cattle_business_options']['header']['row']
        wb['Outputs']['%s%s'%("I", start_row)].value = '=Inputs!$E$'+ str(source_row_)
        col2_ =column_index_from_string("I")
        wb['Outputs']['I' + str(start_row)].alignment = Alignment(wrap_text=True,vertical='top')
        wb['Outputs']['I' + str(start_row)].font = Font(name='Bookman Old Style',bold=False, sz=20.0, color='FF0070C0', scheme='minor')
        

        #description of selected project
        wb['Outputs'].merge_cells(start_row=start_row, start_column=col2_, end_row=end_row, end_column=col2_ + 6)
        #----------------------------------------------------------------- 
       
              
        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------Funding Section------------------ 
        end_row =self._add_output_funding(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'Funding','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------General Section------------------ 
        end_row =self._add_output_general(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'General','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        #self._remove_borders(output_sheet)
        # remove gridlines
        output_sheet.sheet_view.showGridLines = False
        output_sheet.sheet_view.showRowColHeaders = False
 

    def _write_output_sheet_orginal(self, wb, request):
        total_wsheet_cols = 17
        #---------------------------Worksheet 1---------------------------
        output_sheet = wb['Outputs']	
        col = get_column_letter(2)

        row_index = 2
        output_sheet['%s%s'%(col, row_index)].value = self.project_title# 'PERI-URBAN CATTLE FARMING OUTPUTS'# dynamic heading
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
          
        self._set_thick_bottom_border_range( output_sheet,  row_index, 1, total_wsheet_cols)
        
        row_index += 2
        output_sheet['%s%s'%(col, row_index)].value = 'Scenario Analysis '
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
        
        col = get_column_letter(6)
        output_sheet['%s%s'%(col, row_index)].value = 'Active Scenario'
        output_sheet['%s%s'%(col, row_index)].alignment = Alignment(horizontal='left', vertical='top',)
        output_sheet['%s%s'%(col, row_index)].font = Font(name='Calibri',bold=False, italic=True, sz=10.0, color='FF0070C0', scheme='minor')
        
        col = get_column_letter(7)
        _unit= "NOTE"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet['%s%s'%(col, row_index)].value = cell_display_formuale
        output_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'

       
        self._set_thick_bottom_border_range( output_sheet,  row_index, 1, total_wsheet_cols)
        col = get_column_letter(1)
        row_index+=2
        
        #-------------------Data Validation---------------------------
        i=0
        values_ =''
        for item in self.sister_model.cattle_business_options.keys():
            if i ==len(self.sister_model.cattle_business_options.keys())-1:
                values_ += self.sister_model.cattle_business_options[item]['heading'] 
            else:
               values_ += self.sister_model.cattle_business_options[item]['heading'] + ','     
            i+=1
        dv = DataValidation(type="list", formula1='"'+ values_ + '"', allow_blank=True)
        
        user_name=request.user
        # Optionally set a custom prompt message
        dv.prompt = f'{user_name} Please select from the list'
        dv.promptTitle = 'List Selection'
        
        # Add the data-validation object to the worksheet
        output_sheet.add_data_validation(dv)
        # Create some cells, and add them to the data-validation object
        model_sel_row= row_index
        model_sel_col= "F"
        model_sel_cell= "$" + model_sel_col + "$" + str(model_sel_row)
        model_cell = output_sheet[model_sel_cell]
        self.track_inputs['outputs']['model_selection']['row']= model_sel_row 
        self.track_inputs['outputs']['model_selection']['col']= model_sel_col
        self.track_inputs['outputs']['model_selection']['cell']= model_sel_cell

        model_cell.value = values_.split(",")[0]
        model_cell.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        model_cell.font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
        model_cell.alignment = Alignment(horizontal='left', vertical='top',)
        dv.add(model_cell)
        
       
       #----switch

       
        col = get_column_letter(3)
        output_sheet['%s%s'%(col, row_index)].value = 'A.'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(4)
        output_sheet['%s%s'%(col, row_index)].value = 'Alternative Investment Scenario'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(7)
        _unit= "SWITCH"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet['%s%s'%(col, row_index)].value = cell_display_formuale
        output_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'
       #------------------

        #####################333update input------------------------
        colHeader = get_column_letter(5)
        select_val ="="
        list_= self.sister_model.cattle_business_options.keys()
        #print(list_)
        closin_brackets=''
        for  i in range(len(list_)-1):
            closin_brackets+=')'

        i = 0
        input_row_index= self.track_inputs['cattle_business_options']['header']['row']
        for item in list_:
            #item =list_[i]
            if i ==len(list_)-1:
                select_val += colHeader + str(input_row_index +i+1) + closin_brackets
            else:
                cell_x= model_sel_cell
                select_val += 'IF(Outputs!' + cell_x + '="'+ self.sister_model.cattle_business_options[item]['heading'] +'",' + colHeader + str(input_row_index +i+1) + ','
            i+=1
        #---------------------------------------------------------------
         
        wb["Inputs"]['%s%s'%(colHeader, row_index)].value = select_val
        wb["Inputs"]['%s%s'%(colHeader, row_index)].font = self.description_font
        ##########################################3


      
        row_index+=2

        self._hide_empty_cells( output_sheet)
        #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':10})
        rangelist.append({'start':1, 'end': 2 , 'dim':5})
        rangelist.append({'start':2, 'end': 5, 'dim':2})
        #rangelist.append({'start':2, 'end': 3, 'dim':5})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':5, 'dim':55})
        indexlist.append({'index':6, 'dim':30})
        indexlist.append({'index':7, 'dim':15})

        #set dims----
        self._set_column_dim(output_sheet,rangelist,indexlist)
        #------------------------------------
        start_row = row_index
        #end_row = start_row +10
        
        # ------------Financial Section------------------ 
        end_row =self._add_output_financial(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row, end_row,'Financial Analysis Output',)
        
        #--------User View of Selected Project Option-------------------------------------------------------
        start_row= row_index
        source_row_ =self.track_inputs['cattle_business_options']['header']['row']
        wb['Outputs']['%s%s'%("I", start_row)].value = '=Inputs!$E$'+ str(source_row_)
        col2_ =column_index_from_string("I")
        wb['Outputs']['I' + str(start_row)].alignment = Alignment(wrap_text=True,vertical='top')
        wb['Outputs']['I' + str(start_row)].font = Font(name='Bookman Old Style',bold=False, sz=20.0, color='FF0070C0', scheme='minor')
        

        #description of selected project
        wb['Outputs'].merge_cells(start_row=start_row, start_column=col2_, end_row=end_row, end_column=col2_ + 6)
        #----------------------------------------------------------------- 
       
              
        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------Funding Section------------------ 
        end_row =self._add_output_funding(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'Funding','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------General Section------------------ 
        end_row =self._add_output_general(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'General','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        #self._remove_borders(output_sheet)
        # remove gridlines
        output_sheet.sheet_view.showGridLines = False
        output_sheet.sheet_view.showRowColHeaders = False
 



    def _get_investment_paramater_formulae(self, input_row_index):

        # retrive model values
        if 'outputs' in  self.track_inputs:
            if 'model_selection' in  self.track_inputs['outputs']:
                model_sel_cell  = self.track_inputs['outputs']['model_selection']['cell']

                start_col_index = 9
                formulae_part ="="

                list_= self.sister_model.cattle_business_options.keys()
                #print(list_)
                closin_brackets=''
                for  i in range(len(list_)-1):
                    closin_brackets+=')'

                i = 0
               
                for item in list_:
                    #item =list_[i]
                    colHeader = get_column_letter(start_col_index + i )
                    if i ==len(list_)-1:
                        # last part of formalae
                        formulae_part += colHeader + str(input_row_index ) + closin_brackets
                    else:
                        formulae_part += 'IF(Outputs!' + model_sel_cell + '="'+ self.sister_model.cattle_business_options[item]['heading'] +'",' + colHeader + str(input_row_index) + ','
                    i+=1
                #---------------------------------------------------------------
                        
        return formulae_part
   
    def _get_input_memory_details(self, input_parameter, search_start=0):
        #change_in_prices only ins prices
        # excludes sens
        #in future only unique names allowed
        intemlist = ['timing','prices','flags','financing','depreciation',
                    'working_capital','taxes','macroeconomic_parameters', 
                    'feedlot_design_parameters', 'production_inventory','cost_real', 'investment_cost']
        flag_found = False
        input_parameter_header = None
        input_parameter_val  =  None
        input_parameter_unit =  None
        item_instance = None
        
        # return the first instance found
        # in future retun all instances as list
        for iter in range(search_start, len(intemlist)):
            item =intemlist[iter]
            if input_parameter in self.track_inputs[item]:
                input_parameter_row = self.track_inputs[item][input_parameter]['row']# col F==>6
                # add value
                input_parameter_val = self.track_inputs[item][input_parameter]['value']
                # add value
                input_parameter_unit = self.track_inputs[item][input_parameter]['unit']
                #found and loop out
                colE = get_column_letter(5)
                input_parameter_header = '=Inputs!$' + colE + '$' + str(input_parameter_row)
                flag_found= True
                # early return
                item_instance= item
                #set pointer to next element in future search
                next_element = iter +1
                return item_instance, input_parameter_header,input_parameter_val,input_parameter_unit, flag_found, next_element
        next_element =len(intemlist)
        return item_instance, input_parameter_header,input_parameter_val,input_parameter_unit, flag_found, next_element
      
    
    def _write_input_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        output_sheet = wb.active
        output_sheet.title = 'Outputs'
        output_sheet = wb['Outputs']

        #inputs
        input_sheet = wb.create_sheet()
        input_sheet.title = 'Inputs'

        #Calculations
        calc_sheet = wb.create_sheet()
        calc_sheet.title = 'Calc'

        #Cash Flow
        cash_flow_sheet = wb.create_sheet()
        cash_flow_sheet.title = 'CF'
        
        user='Andrew'
        # Writing the first row of the csv	
        col = get_column_letter(1)
        input_sheet['%s%s'%(col, 1)].value = 'Input Sheet ' #% (user)
        input_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_modelspecification_section(self.sister_model, input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Timing Section =========================
        row_index +=2
        row_index = self._add_timingassumptions_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Prices Section =========================
        row_index +=2
        row_index = self._add_prices_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
        #==========Production and Inventory Section =========================
        row_index +=2
        row_index = self._add_prodInventory_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Investment Cost Section =========================
        row_index +=2
        row_index = self._add_investmentCost_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Depreciation Section =========================
        row_index +=2
        row_index = self._add_depreciation_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Costs Section =========================
        row_index +=2
        row_index = self._add_costs_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
      
        #==========WorkingCapital Section =========================
        row_index +=2
        row_index = self._add_workingCapital_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

         #==========Financing Section =========================
        row_index +=2
        row_index = self._add_financing_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Flags Section =========================
        row_index +=2
        row_index = self._add_flags_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Taxes Section =========================
        row_index +=2
        row_index = self._add_taxes_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Macroeconomic Parameters Section =========================
        row_index +=2
        row_index = self._add_macroeconomicParameters_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
         #==========Feedlot Design Parameter Section =========================
        row_index +=2
        row_index = self._add_feedlotDesignParameters_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

       
        # hide---------------
        self._hide_empty_cells(input_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        self._set_column_dim(input_sheet,rangelist,indexlist)
    
    def _update_sensitivity(self, sens_sheet,modelx):

        #==========purchase_price_section =========================
        row_index = 0
        if 'sens_parameters' in self.track_inputs:
            if 'endpoint' in self.track_inputs['sens_parameters']:
                row_index =self.track_inputs['sens_parameters']['endpoint']['row'] 

       
        #ddddd
        item_list =[
               {'title':"Cattle Purchase Price Per Unit", "track_input_item": "sens_cattle_price_per_unit",'base_val':self.sister_model.cattle_price_per_unit, 'number_format':"NUMBER"},
               {'title':"Base Price",                      "track_input_item": "sens_base_price",          'base_val':self.sister_model.base_price,             'number_format':"NUMBER"},
               {'title':"Change in Price [Beef]",                 "track_input_item": "sens_change_in_price",      'base_val':self.sister_model.change_in_price,       'number_format':"PERCENT"},

               {'title':"Domestic Inflation Rate",           "track_input_item": "sens_domestic_inflation_rate", 'base_val':self.sister_model.domestic_inflation_rate, 'number_format':"PERCENT"},
               {'title':"Exchange Rate", "track_input_item": "sens_exchange_rate"                      ,         'base_val':self.sister_model.exchange_rate,            'number_format':"NUMBER"},
               
               {'title':"Initial Pens Employed",               "track_input_item": "sens_initial_pens_employed" , 'base_val':self.sister_model.initial_pens_employed, 'number_format':"INTEGER"},
               
               {'title':"Senior Debt",               "track_input_item": "sens_senior_debt",           'base_val':self.sister_model.senior_debt,            'number_format':"PERCENT"},
               {'title':"Cattle Survival rate",      "track_input_item": "sens_cattle_survival_rate" , 'base_val':self.sister_model.cattle_survival_rate,   'number_format':"PERCENT"},
               
               {'title':"Cattle Feed Price per kg",      "track_input_item": "sens_cattle_feed_price_per_kg" , 'base_val':self.sister_model.cattle_feed_price_per_kg,   'number_format':"SMALL_NUMBER"},
               
               {'title':"Us Inflation rate",         "track_input_item": "sens_us_inflation_rate",     'base_val':self.sister_model.us_inflation_rate,      'number_format':"PERCENT"},

               
               {'title':"Accounts Receivable",            "track_input_item":  "sens_accounts_receivable",            'base_val':self.sister_model.accounts_receivable,            'number_format':"PERCENT"},
               {'title':"Accounts Payable",                "track_input_item": "sens_accounts_payable",               'base_val':self.sister_model.accounts_payable,               'number_format':"PERCENT"},
               {'title':"Cash Balance",                   "track_input_item":  "sens_cash_balance",                   'base_val':self.sister_model.cash_balance,                   'number_format':"PERCENT"},
               {'title':"Total Cattle per pen per cycle", "track_input_item":  "sens_total_cattle_per_pen_per_cycle", 'base_val':self.sister_model.cattle_pen_cycle, 'number_format':"INTEGER"},
               
               {'title':"Discount Rate Equity",            "track_input_item":  "sens_discount_rate_equity",            'base_val':self.sister_model.discount_rate_equity,            'number_format':"PERCENT"},
               {'title':"Dressed Weight @ Selling",                "track_input_item": "sens_dressed_weight_at_selling",               'base_val':self.sister_model.dressed_weight_at_selling,               'number_format':"NUMBER"},
               {'title':"Other Indirect Costs",                   "track_input_item":  "sens_other_indirect_costs",                   'base_val':self.sister_model.other_indirect_costs,                   'number_format':"NUMBER"},
               {'title':"Annual Change in price of Domestic inputs", "track_input_item":  "sens_annual_change_in_price_of_domestic_inputs", 'base_val':self.sister_model.annual_change_in_price_of_domestic_inputs, 'number_format':"PERCENT"},
               {'title':"Annual Change in price of Imported inputs", "track_input_item":  "sens_annual_change_in_price_of_imported_inputs", 'base_val':self.sister_model.annual_change_in_price_of_imported_inputs, 'number_format':"PERCENT"},
          
       
               {'title':"Monthly wage per worker",                    "track_input_item":  "sens_monthly_wage_per_worker",            'base_val':self.sister_model.monthly_wage_per_worker,            'number_format':"SMALL_NUMBER"},
               {'title':"Monthly wage per supervisor",                "track_input_item":    "sens_monthly_wage_per_supervisor",               'base_val':self.sister_model.monthly_wage_per_supervisor,               'number_format':"SMALL_NUMBER"},
               {'title':"Annual increase salaries workers",             "track_input_item":  "sens_annual_increase_salaries_workers",                   'base_val':self.sister_model.annual_increase_salaries_workers,                   'number_format':"PERCENT"},
               {'title':"Annual increase salaries Supervisors and Technicians", "track_input_item":  "sens_annual_increase_salaries_supervisors_technicians", 'base_val':self.sister_model.annual_increase_salaries_supervisors_technicians, 'number_format':"PERCENT"},
               {'title':"Number of workers per Supervisor", "track_input_item":  "sens_num_workers_per_supervisor", 'base_val':self.sister_model.num_workers_per_supervisor, 'number_format':"INTEGER"},
               {'title':"Investments Cost Overun", "track_input_item":  "sens_inv_cost_over_run", 'base_val':self.sister_model.inv_cost_over_run, 'number_format':"SMALL_NUMBER"},
           
            
            ] 

        parameter_unit_dict ={}
        for item_x in item_list:
            c =item_x['track_input_item']
            title =item_x['title']
            var= c.split("sens_",1)[1].strip()
            parameter_unit_dict[var]={'number_format':item_x['number_format'],'title':title}
        setattr(modelx,'parameter_unit_dict', parameter_unit_dict)
      
        # order one list using sequence of another list:::: 
        if hasattr(modelx,'para_list_by_grad'):
            #chop vars with no gradient.....
            item_list= self._sort_by_another_list(item_list,modelx.para_list_by_grad)[:len(modelx.para_list_by_grad)]
        #
        iter_i =0 
        for item in item_list:# Order starting with most senstive parameter
            iter_i +=1
            #skipp 2 rows
            row_index += 2
            #-----set independent variables---------
            if 'base_val' in item:
               val = item['base_val']
            else:
                val= 10
            #--------------------------------------
            #general parameter range
            list_, indexof_base ,indexof_npv_zero = modelx._sen_parameter_generator(val,item['track_input_item']) 
            #print(list_, index)
            #---------------Parameter Formating----------------------------------
            if 'number_format' in item:
                if item['number_format']=="PERCENT":
                    number_format ='0.0%'
                elif item['number_format']=="NUMBER":
                    #decimal

                    number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                elif item['number_format']=="SMALL_NUMBER":
                    #decimal
                    
                    number_format ='_(* #,##0.0##_);_(* \(#,##0.0##\);_(* "-"??_);_(@_)'
                elif item['number_format']=="INTEGER":
                    #integer---
                    number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                else:
                    number_format ='General' 
            else:
                number_format ='General'
            #--------------------------------------------------------------------     
            header_title =str(iter_i) +'. ' + item['title']
            row_index = self._add_sensitivity_x_section(modelx, sens_sheet, row_index, 15, 
                                     header_title,  item['track_input_item'],list_,indexof_base, indexof_npv_zero, number_format)
  
    def _populate_productionInventory_section(self, w_sheet, item ):
                 
        if item=='num_of_feedlot_targeted_for_construction':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            if found_state:
                span_ =self._get_span()
                if 'sens' in self.track_inputs:
                    if 'initial_pens_employed' in self.track_inputs['sens']:
                        _row = self.track_inputs['sens']['initial_pens_employed']['row'] 
                        _col = self.track_inputs['sens']['initial_pens_employed']['col']
                        col = get_column_letter(_col)
                        sens_val= "Sens!$" + col + "$" + str(_row)
                #bbbbbbbbbbb          
                r_index2, _, found_state2= self._retrieve_cell_row_colm('flags','OP')
                if found_state2:
                    #start at one less writing columns
                    startpoint = '$'+ get_column_letter(8) + '$'+ str(r_index2)# D07
                 
                    start_col_index = 9
                    first_slice_point = get_column_letter(start_col_index+1) + str(r_index)
                    second_slice_point = get_column_letter(start_col_index +1 + int(span_)) + str(r_index) # D39

                    #--------------------------------
                    w_sheet['%s%s'%(get_column_letter(start_col_index), r_index)].value ='=' + sens_val  
                    w_sheet['%s%s'%(get_column_letter(start_col_index), r_index)].style = 'Linkedcell'
                    w_sheet['%s%s'%(get_column_letter(start_col_index), r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 

                    #---------------------------------
                    
                    #loop each cell of this row D15:G15
                    for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                        for cell in cellObj:
                            new_column_letter = cell.column # J
                            prev_col=column_index_from_string(new_column_letter)-1
                            prev_letter= get_column_letter(prev_col)
                            prev_targeted_cell = prev_letter + str(r_index2)
                            current_targeted_cell = new_column_letter + str(r_index2)                        
                            w_sheet['%s%s'%(new_column_letter, cell.row)].value ='=0'   #'=IF(AND(SUM(' + startpoint + ':' + prev_targeted_cell +')=0,' + current_targeted_cell +'>0),'+ sens_val + ',0)' 
                            w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                            w_sheet['%s%s'%(new_column_letter, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
                    
       
        if item=='cum_pens_under_cattle':
            r_feedlot_constr, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            startpoint_feedlot_const_row = '$'+ get_column_letter(9) + '$'+ str(r_feedlot_constr)
         
            r_op, _, found_state2= self._retrieve_cell_row_colm('flags','OP')
            if found_state2:
                #start at one less writing columns
                startpoint = '$'+ get_column_letter(9) + '$'+ str(r_op)# D07

            if 'sens' in self.track_inputs:
                if 'initial_pens_employed' in self.track_inputs['sens']:
                    _row = self.track_inputs['sens']['initial_pens_employed']['row'] 
                    _col = self.track_inputs['sens']['initial_pens_employed']['col']
                    col = get_column_letter(_col)
                    sens_val= "Sens!$" + col + "$" + str(_row)

            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_cattle')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                       
                        new_column_letter = cell.column # J
                        prev_col=column_index_from_string(new_column_letter)-1
                        prev_letter= get_column_letter(prev_col)
                        prev_row = cell.row-1

                        endpoint_feedlot_const_row = '$'+ new_column_letter + '$'+ str(r_feedlot_constr)
                        current_op_cell = new_column_letter + str(r_op)  
                        answ = '=IF((SUM(' + startpoint + ':' + current_op_cell +')>0),SUM(' + startpoint_feedlot_const_row + ':' + endpoint_feedlot_const_row +'),0)'
                        w_sheet['%s%s'%(new_column_letter, cell.row)] = answ 
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                        
        if item=='cum_pens_under_harvesting':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_harvesting')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        new_column_letter = cell.column # J
                        prev_row = cell.row-1 
                        current_constr_cell = new_column_letter + str(prev_row)
                    
                    
                        w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=' + current_constr_cell 
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                        #w_s 
        if item=='production_quantity':
            #=A*B*C*D/E * F
            #= cattle_per_pen_per_year * cattle_survival_rate * dressed_weight_at_selling * cum_pens_under_cattle/ thousand * OP
            formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'production_inventory', 'para': 'cattle_per_pen_per_year','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'production_inventory', 'para': 'cattle_survival_rate','cell_type': 'single'},
                {'value':'*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'production_inventory', 'para': 'dressed_weight_at_selling','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'D', 'var_type': 'variable', 'header': 'production_inventory', 'para': 'cum_pens_under_harvesting','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'E', 'var_type': 'variable', 'header': 'production_inventory', 'para': 'thousand','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'F', 'var_type': 'variable', 'header': 'flags', 'para': 'OP','cell_type': 'cell_range'},
            ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'production_inventory',   'production_quantity')
  
        if item=='total_pen_constructed':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','total_pen_constructed')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet['%s%s'%(get_column_letter(c_index), r_index)].value = "=Sum(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = 'Calculation'
        
        if item=='investment_roll_out_flag':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            if found_state:
                span_ =self._get_span()
               
                r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
                if found_state2:
                    startpoint = '$'+ get_column_letter(8) + '$'+ str(r_index2)# D07
                    first_slice_point = get_column_letter(9) + str(r_index)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                    #loop each cell of this row D15:G15
                    #----------->
                    for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                        for cell in cellObj:
                            new_column_letter = cell.column # J
                            prev_col=column_index_from_string(new_column_letter)-1
                            prev_letter= get_column_letter(prev_col)
                            prev_targeted_cell = prev_letter + str(r_index2)
                            current_targeted_cell = new_column_letter + str(r_index2)                        
                            w_sheet['%s%s'%(new_column_letter, cell.row)].value ='=IF(AND(SUM(' + startpoint + ':' + prev_targeted_cell +')=0,' + current_targeted_cell +'>0),1,0)' 
                            w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Input'
                            w_sheet['%s%s'%(new_column_letter, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
                
        if item=='closing_inventory':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','closing_inventory')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        new_column_letter = cell.column # J
                        prev_row = cell.row-1                     
                        w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=0' 
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                        #w_s  
    def _populate_costs_section(self, w_sheet, item ):
                   
        if item=='cum_pens':
            # 
            cum_pens_row_index, _, found_state_pens= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_cattle')
           
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','cum_pens')

            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(cum_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + str(pens_targeted_cell) 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              

        if item=='num_of_workers':

            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','num_of_workers_per_pen')
            num_of_workers_per_pen_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            
            targeted_pens_row_index, _, found_state_pens= self._retrieve_cell_row_colm('cost_real','cum_pens')
           
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','num_of_workers')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(targeted_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] ='=CEILING('+ pens_targeted_cell +'*'+ num_of_workers_per_pen_cell +',1)' 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              
            
        if item=='num_of_supervisors_technicians':

            workers_r_index, c_index, found_state_workers= self._retrieve_cell_row_colm('cost_real','num_of_workers')
          
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','num_of_supervisors_technicians')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        num_of_workers_per_pen_cell = cell.column + '$' + str(workers_r_index ) if found_state_workers else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] ='=IF(' +  num_of_workers_per_pen_cell +' >0,INT(' + num_of_workers_per_pen_cell+ '/5),0)' 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              

        
        
        if item=='average_num_of_workers':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','average_num_of_workers')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('cost_real','num_of_workers')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet['%s%s'%(get_column_letter(c_index), r_index)].value = "=AVERAGE(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = 'Calculation'
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              
        
        if item=='average_num_of_supervisors':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','average_num_of_supervisors')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('cost_real','num_of_supervisors_technicians')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet['%s%s'%(get_column_letter(c_index), r_index)].value = "=AVERAGE(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = 'Calculation'
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
  
    def _add_investmentCost_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Investment Cost (Real)' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
        
        #keep track : to insert datatimeline:---> 2021, 2022....2040 
        if 'header' in self.track_inputs['investment_cost']:
            self.track_inputs['investment_cost']['header']['row'] = row_index
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['total_land_for_pens','cost_of_land_per_sqm','cost_of_machinery_per_pen', 
                   'cost_of_building_per_sqm','total_pens','cost_of_pen_construction', 
                   'senior_debt_dynamic_parameter','initial_pens_employed','pen_cattle_density',
                   'pen_length','pen_width','pen_height',
                   'total_land_required','cost_of_pens_constructed','cost_of_land',
                   'investment_cost_of_land','cif_cost_of_machinery','investment_cost_of_buildings',
                   'investment_costs_over_run_factor']
   
        #-----------------------------------
        for item in itemlist:
            if item in self.sister_model.investment_cost:
                _unit =self.sister_model.investment_cost[item]['units']
                _unit = _unit.upper()
                # print(item,_unit,self._get_number_formats(_unit))
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.investment_cost[item]['title'], self.sister_model.investment_cost[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['investment_cost'][item]['row'] = row_index
                self.track_inputs['investment_cost'][item]['unit'] = _unit
                self.track_inputs['investment_cost'][item]['col'] = 6# redundand
                self.track_inputs['investment_cost'][item]['value'] = self.sister_model.investment_cost[item]['value'] 

                if item in self.sister_model.investment_parameter_options:
                    #automaticaaly retrive related
                    self._populate_investment_parameters_in_excelrow(w_sheet, self.sister_model.investment_parameter_options[item],row_index, item)

                row_index +=1
                #skipp lines
                if item in ['senior_debt_dynamic_parameter','pen_height', 'total_land_required',]:
                    row_index +=1 
       
        # options headers
        self._populate_investment_parameters_header_rows(w_sheet)
        #cost_of_pens_constructed
        return row_index

    def _populate_investment_parameters_header_rows(self, w_sheet):

         #-----------------------------------------------------------------
        r_index, _, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')

        if found_state:
            headers=[]
            for item in self.sister_model.cattle_business_options.keys():
                headers.append(self.sister_model.cattle_business_options[item]['heading'])
            #self._populate_investment_parameters_in_excelrow(w_sheet, headers, r_index)
        
            total_len = len(headers)-1
            first_slice_point = get_column_letter(9) + str(r_index)# D09
            second_slice_point = get_column_letter(9 + int(total_len)) + str(r_index) # D16
            #loop each cell of this row D09:D16
            #----------->
            i=0
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = headers[i] 
                    w_sheet['%s%s'%(cell.column, cell.row)].fill = PatternFill(fill_type="solid", fgColor="70c4f4")
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
                    w_sheet['%s%s'%(cell.column, cell.row)].alignment = Alignment(wrap_text=True,horizontal='left', vertical='top',)
                    i += 1
    
    
    
    def _link_feedlot_design_parameters_to_model_selected(self, w_sheet):
       
        # similar pointer but different naming
        #correct this in future
        feedlot_para = ['num_of_feedlots','length', 'width',]#  'sqm_per_cattle' ]
        investment_cost_para = ['total_pens','pen_length', 'pen_width',]# 'pen_cattle_density']
        # take min to be safe

        list_len= min(len(feedlot_para), len(investment_cost_para))
        for iter in range(list_len):
            # source data
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost',investment_cost_para[iter])
            item_cell = '=$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            #copy to destination cells  
            r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters' ,feedlot_para[iter])
            if found_state:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = item_cell
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style= 'Linkedcell'
        
               

    def _update_investment_parameters(self, w_sheet):
        para_list= ['cost_of_land_per_sqm', 'cost_of_machinery_per_pen', 
            'cost_of_building_per_sqm', 'cost_of_pen_construction',
            'senior_debt_dynamic_parameter', 'initial_pens_employed', 'pen_cattle_density', 
            'pen_length', 'pen_width', 'pen_height',  'total_land_required',
        ]

        for item in  para_list:
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost',item)
            if found_state:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = self._get_investment_paramater_formulae(r_index)
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style= 'Calculation'
                if item in ['senior_debt_dynamic_parameter']:
                    w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format= '0.0%'

        
        # Total PEns different
        r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_pens')
        if found_state:
            span_ =self._get_span()
            r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            if found_state2: 
                first_slice_point = get_column_letter(9) + str(r_index2)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].value ="=SUM(" + first_slice_point + ":" + second_slice_point +")"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = 'Calculation'
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                 

    
    def _add_prodInventory_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Production and Inventory' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if 'header' in self.track_inputs['production_inventory']:
            self.track_inputs['production_inventory']['header']['row'] = row_index
           
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['cattle_per_pen_per_year','cattle_survival_rate','thousand', 
                   'dressed_weight_at_selling','num_of_feedlot_targeted_for_construction','cum_pens_under_cattle', 
                   'cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'investment_roll_out_flag','total_pen_constructed']
        #-----------------------------------
        for item in itemlist:
            if item in self._production_inventory:
                _unit =self._production_inventory[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self._production_inventory[item]['title'], self._production_inventory[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['production_inventory'][item]['row'] = row_index
                self.track_inputs['production_inventory'][item]['unit'] = _unit
                self.track_inputs['production_inventory'][item]['col'] =6 #redundand
                self.track_inputs['production_inventory'][item]['value'] = self._production_inventory[item]['value']
                
                row_index +=1 
   

       

        return row_index 
    def _populate_investmentcost_section(self, w_sheet, item ):
        r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','thousand')
        thousand_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
              
        if item=='cost_of_pens_constructed':
            # 
            targeted_pens_row_index, c_index, found_state_pens= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_pen_construction')
            cost_of_pen_construction_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            # 
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_pens_constructed')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(targeted_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + str(pens_targeted_cell) + '*' + cost_of_pen_construction_cell \
                                                                      + '/' + str(thousand_cell)
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General
        if item=='cost_of_land':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_required')
            total_land_required_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           

            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')
            total_land_for_pens_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_land_per_sqm')
            cost_of_land_per_sqm_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            roll_out_r_index,_, found_state_roll_out= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_land')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        investment_roll_out_flag_cell = cell.column + '$' + str(roll_out_r_index ) if found_state_roll_out else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ total_land_required_cell +'>0, MAX(' \
                                                                        + total_land_required_cell + ',' \
                                                                        + total_land_for_pens_cell + ')*' + cost_of_land_per_sqm_cell \
                                                                        + '*'+  investment_roll_out_flag_cell   + '/' \
                                                                        + thousand_cell +',0)'
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        
                        #w_s 
        if item=='investment_cost_of_land':
            cost_of_land_r_index, _, found_state_cost_of_land= self._retrieve_cell_row_colm('investment_cost','cost_of_land')
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','investment_cost_of_land')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
               
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        cost_of_land_cell = cell.column + '$' + str(cost_of_land_r_index ) if found_state_cost_of_land else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_land_cell
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        
        
        if item=='cif_cost_of_machinery':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_machinery_per_pen')
            cost_of_machinery_per_pen_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            pens_targeted_r_index, c_index, found_state_pens_targeted= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cif_cost_of_machinery')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
               
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_for_construction_cell = cell.column + '$' + str(pens_targeted_r_index ) if found_state_pens_targeted else '0'
           
                        w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_machinery_per_pen_cell +'*' \
                                                                        + pens_targeted_for_construction_cell + '/' + thousand_cell  
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
        
        if item=='investment_cost_of_buildings':            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_building_per_sqm')
            cost_of_building_per_sqm_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')
            total_land_for_pens_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            
            roll_out_r_index,_, found_state_roll_out= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','investment_cost_of_buildings')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        investment_roll_out_flag_cell = cell.column + '$' + str(roll_out_r_index ) if found_state_roll_out else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] ='='+ cost_of_building_per_sqm_cell + '*' \
                                                                    + total_land_for_pens_cell \
                                                                    + '*' + investment_roll_out_flag_cell  + '/' + thousand_cell  
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
 
    def _write_sens_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        
        sens_sheet = wb.create_sheet()
        sens_sheet.title = 'Sens'
        
      
        # Writing the first row of the csv	
        col = get_column_letter(1)
        sens_sheet['%s%s'%(col, 1)].value = 'Sensitivity Analysis sheet'
        sens_sheet['%s%s'%(col, 1)].style= 'Heading1'
    
        span_ = 6 
        total_wsheet_cols = 9 + int(span_)
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(sens_sheet, row_index,total_wsheet_cols,6)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_sensitivity_section(wb, sens_sheet, row_index,total_wsheet_cols)

        if not 'sens_parameters' in self.track_inputs:
            self.track_inputs['sens_parameters']= {}# insert inner

        if not ('endpoint' in self.track_inputs['sens_parameters']):
            self.track_inputs['sens_parameters']['endpoint'] = {}
  
        if 'endpoint' in self.track_inputs['sens_parameters']:
            self.track_inputs['sens_parameters']['endpoint']['row'] = row_index
        #======================================+++++++++++++++

       
        # hide---------------
        self._hide_empty_cells(sens_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':14, 'end': 16, 'dim':2})
        rangelist.append({'start':3, 'end': 13, 'dim':10})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':1, 'dim':2})
        indexlist.append({'index':2, 'dim':40})

        #set dims----
        self._set_column_dim(sens_sheet,rangelist,indexlist)
   
    
    def _update_feedlot_linkedcells(self, w_sheet,):
        
        # cattle_survival_rate
        # r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','cattle_survival_rate')
        # cattle_survival_rate = get_column_letter(c_index) + str(r_index ) if found_state else None
        # num_of_months_per_cycle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','num_of_months_per_cycle')
        num_of_months_per_cycle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # num_of_feedlots
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','num_of_feedlots')
        num_of_feedlots_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # length
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','length')
        length_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # width
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','width')
        width_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # pen_area
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','pen_area')
        pen_area_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # total_cattle_per_pen_per_cycle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','total_cattle_per_pen_per_cycle')
        cattle_per_cycle = get_column_letter(c_index) + str(r_index ) if found_state else .00001
        
        # SQM per cattle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','sqm_per_cattle')
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + pen_area_cell + '/' + cattle_per_cycle
            w_sheet['%s%s'%(c, r_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')  
        
        
        cal_val= '='  + str(pen_area_cell) + '/' + cattle_per_cycle
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','sqm_per_cattle', cal_val, "Calculated")

        # SQM Calculation
        cal_val = '=' + num_of_feedlots_cell + '*' + length_cell + '*' + width_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','sqm', cal_val, 'Calculated')
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','sqm')
        sqm_cell = None
        if found_state:
            sqm_cell = get_column_letter(c_index) + str(r_index ) 
          

        # PEN Calculation
        cal_val = '=' + length_cell + '*' + width_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','pen_area',cal_val, 'Calculated')
        
        # number_of_months_in_a_year
        r_index , c_index, found_state= self._retrieve_cell_row_colm('timing','number_of_months_in_a_year')
        if found_state:
            number_of_months_in_a_year= get_column_letter(c_index) + str(r_index)
        else:
            number_of_months_in_a_year = 12   
        
        # Cattle_per_pen_per_year Calculation
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','cattle_per_pen_per_year')
        cattle_per_pen_per_year = get_column_letter(c_index) + str(r_index ) if found_state else None
        
        cal_val= '=' + cattle_per_cycle + '*' + str(number_of_months_in_a_year) +   '/' + num_of_months_per_cycle_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','cattle_per_pen_per_year', cal_val, "Calculated")

        # total_land_for_pens
        self._set_linkedcell(w_sheet, 'investment_cost','total_land_for_pens', sqm_cell)
       
       
        self._set_linkedcell(w_sheet, 'production_inventory','cattle_per_pen_per_year', cattle_per_pen_per_year)
        
        self._set_linkedcell(w_sheet, 'cost_real','cattle_per_pen_per_year', cattle_per_pen_per_year)
        #derived from Sens!
        #self._set_linkedcell(w_sheet, 'production_inventory','cattle_survival_rate', cattle_survival_rate, )
        """ 
        itemlist =['cattle_per_pen_per_year','cattle_survival_rate','thousand', 
                   'dressed_weight_at_selling','num_of_feedlot_targeted_for_construction','cum_pens_under_cattle', 
                   'cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'investment_roll_out_flag','total_pen_constructed']
         """
    
    def _add_sensitivity_section(self, wb, w_sheet, row_index,total_wsheet_cols):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Sensitivity Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(2)
        colValue = get_column_letter(3)
        colUnits = get_column_letter(4)
        row_index +=1
       
        # Domestic Inflation rate	5%	%
        # US Inflation rate	3%	%
        # Change in real price of Fish	-  	%
        # Investment Costs Over-run Factor	-  	%
        # Accounts receivable [% of Gross Sales]	10%	%
        # Acccounts payable (% of total input cost)	10%	%
        # Cash balance  (% of gross sales)	10%	%
        # Cattle Survival rate	100%	%
        # Exchange Rate (LC/USD - year 2018)	1.4	#
        # Senior Debt (% of Investment Costs) 	70%	%
        # Total pens 	5 	#
        # Base price of beef in 2020 per ton 	4,500	LC
        # Total Cattle in one pen per cycle 	20 	#
        # Cattle purchase price per unit 	300	LC
        # Scale 	Moderate Scale
        itemlist =['domestic_inflation_rate','us_inflation_rate','change_in_price', 'accounts_receivable',
                   'accounts_payable','cash_balance','exchange_rate', 'senior_debt','base_price',
                   'total_cattle_per_pen_per_cycle' ,'cattle_survival_rate', 'cattle_price_per_unit',
                   'initial_pens_employed']
        update_exclusion_list= []#[{'item_instance': 'investment_cost','item':'initial_pens_employed'},]
        exclude_item= {}
        exclude_item['item_instance']='investment_cost'
        exclude_item['item']='initial_pens_employed'
        update_exclusion_list.append(exclude_item)
        for item in itemlist:
            # each item in list can be retrieved more than one
            #loop until nothing is found
            iter = 0 # for each ele start at zero
            continue_search = True # default initialisation
            tag_ = 0
            while continue_search:
                item_instance, header_,val_, _unit, flag_found, iter=self._get_input_memory_details(item,iter)
                continue_search = flag_found
            
                if flag_found:
                    tag_ +=1# increament # of counts
                    cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                    self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                            header_, val_, cell_display_val, self._get_number_formats(_unit), _unit=='BLANK')
                    
                    #dynamic declare variables
                    if not (item in self.track_inputs['sens']):
                       self.track_inputs['sens'][item] ={}

                    self.track_inputs['sens'][item]['row'] = row_index
                    self.track_inputs['sens'][item]['unit'] = _unit
                    self.track_inputs['sens'][item]['col'] = 3
                    #self.track_inputs['sens'][item]['value'] = row_index
                    
                    #----dynamic update of Inputvariable:
                    #  price, us infaltion , change_in_price
                    if item_instance in self.track_inputs  :
                        if item in self.track_inputs[item_instance] and self._check_item_dict_in_list(item_instance,item, update_exclusion_list):
                            _row = self.track_inputs[item_instance][item]['row'] 
                            _col = self.track_inputs[item_instance][item]['col']
                            col = get_column_letter(_col)
                            wb["Inputs"]['%s%s'%(col, _row)].value = "=Sens!$" + colValue + "$" + str(row_index)
                            #print(col, _row,"=Sens!$" + colValue + "$" + str(row_index))
                            #if not set change to linkedcell type
                            wb["Inputs"]['%s%s'%(col, _row)].style = 'Linkedcell'
                            wb["Inputs"]['%s%s'%(col, _row)].number_format= self._get_number_formats(_unit)

            #iter for writing the sens only {not related to how many instance are found per iteration}
            row_index +=1

            if tag_>1:
                pass
                #print(item, tag_)

        return row_index 
    def _populate_output_general(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
       
        #1 Transer cell to Outputs (Group Loop)
        array_dict= [
            {'target': 'land_under_feedLot',                'source': 'total_land_for_pens',             's_header': 'investment_cost',          's_wksheet': 'Inputs'},
            {'target': 'total_feedLots',                    'source': 'total_pen_constructed',           's_header': 'production_inventory',     's_wksheet': 'Inputs'},
            {'target': 'pen_area',                          'source': 'pen_area',                        's_header':  'feedlot_design_parameters','s_wksheet': 'Inputs'},
            {'target': 'total_cattle_in_one_pen_per_cycle', 'source': 'total_cattle_per_pen_per_cycle',  's_header': 'feedlot_design_parameters', 's_wksheet': 'Inputs'},
            {'target': 'cattle_per_pen_per_year',           'source': 'cattle_per_pen_per_year',         's_header': 'feedlot_design_parameters', 's_wksheet': 'Inputs'},
        
            {'target': 'base_price',                     'source': 'base_price',                       's_header':  'prices','s_wksheet': 'Inputs'},
            {'target': 'tonnes_produced_per_pen_per_year',   'source': 'tonnes_produced_per_pen_per_year',  's_header': 'cost_real',        's_wksheet': 'Inputs'},
            ]
        
        for i in array_dict:
            self._transfer_cell(wksheet, i['s_header'],'output_general',
                               i['source'], number_format, i['s_wksheet'], i['target'],'Output2')
        #2. Transer cell to Outputs (Group Loop)
        # in cellrange at specific cell
        array_dict= [
            {'target': 'annual_revenue_2yr',          'source': 'gross_sales_revenue',  's_header': 'cal_production_sales_nominal',  's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':1 },# second year
            {'target': 'gross_profit_2yr',        'source': 'gross_profit',         's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':1 },# second year
            {'target': 'net_profit_2yr',          'source': 'net_after_tax_income', 's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':1 },# second year
            {'target': 'gross_profit_margin_2yr', 'source': 'gross_profit_margin',  's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format':'0.0%',         'cell_pos':1 },# second year
            {'target': 'net_profit_margin_2yr',   'source': 'net_profit_margin',    's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format':'0.0%',         'cell_pos':1 },# second year
        ]
        for item in array_dict:
            self._transfer_cell(wksheet, item['s_header'],'output_general', item['source'], item['number_format'], 
                               item['s_wksheet'], item['target'],'Output2',item['cell_pos'])
        
        #3. total_cattle_per_year
        #  = cattle_per_pen_per_year  * total_feedLots
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'cattle_per_pen_per_year'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "B", 'var_type': 'variable', 'header': 'output_general', 'para': 'total_feedLots'},
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'total_cattle_per_year',None,'Output2')
        
        
        #3. total_cattle_per_week
        #  =total_cattle_per_year  /52
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'total_cattle_per_year'},
                {'value': '/52', 'var_type': 'const', 'header': '', 'para': '',},
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'total_cattle_per_week',None,'Output2')
         

        #3. annual_productivity_tons
        #  =tonnes_produced_per_pen_per_year  * total_feedLots
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'tonnes_produced_per_pen_per_year'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'B', 'var_type': 'variable','header': 'output_general', 'para': 'total_feedLots'},
               
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'annual_productivity_tons',None,'Output2')
                            
    def _add_output_general(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_general' in self.track_inputs:
            self.track_inputs['output_general']= {}# insert inner

        if not ('header' in self.track_inputs['output_general']):
            self.track_inputs['output_general']['header'] = {}
  
        if 'header' in self.track_inputs['output_general']:
            self.track_inputs['output_general']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= [
                'land_under_feedLot',  'total_feedLots',    'pen_area', 'total_cattle_in_one_pen_per_cycle','cattle_per_pen_per_year',
                'total_cattle_per_year',   'total_cattle_per_week',   'base_price','tonnes_produced_per_pen_per_year', 'annual_productivity_tons',
                'annual_revenue_2yr', 'gross_profit_2yr',   'net_profit_2yr',  'gross_profit_margin_2yr',   
                'net_profit_margin_2yr' 
                ]
        ic_parameters = {
            'land_under_feedLot': {'title': "Land Under FeedLot",'value': None, 'units': 'SQM'},
            'total_feedLots': {'title': "Total FeedLots",'value': None, 'units': 'NUMBER'},
            'pen_area': {'title': "Pen-Area",'value': None, 'units': 'SQM'},
            'total_cattle_in_one_pen_per_cycle': {'title': "Total Cattle in one pen per cycle",'value': None, 'units': 'NUMBER'},
            'cattle_per_pen_per_year': {'title': "Total Cattle per pen per year",'value': None, 'units': 'NUMBER'},
            'total_cattle_per_year': {'title': "Total Cattle per year",'value': None, 'units': 'NUMBER'},
            'total_cattle_per_week': {'title': "Total Cattle per week",'value': None, 'units': 'NUMBER'},

            'base_price': {'title': "Base price of beef per ton",'value': None, 'units': 'LC'},
            'tonnes_produced_per_pen_per_year': {'title': 'Beef Tonnes produced per pen per year','value': None, 'units': 'TONS'},
            'annual_productivity_tons': {'title': 'Annual Beef Productivity Tons','value': None, 'units': 'TONS'},
            'annual_revenue_2yr': {'title': "Annual Revenue (2nd Yr) ",'value': None, 'units': 'T_LC'},

            'gross_profit_2yr': {'title': "Gross Profit (2nd Yr)",'value': None, 'units': 'T_LC'},
            'net_profit_2yr': {'title': 'Net Profit (2nd Yr)','value': None, 'units': 'T_LC'},
            'gross_profit_margin_2yr': {'title': "Gross Profit Margin  (2nd Yr)",'value': None, 'units': 'PERCENT'},
            'net_profit_margin_2yr': {'title': "Net Profit Margin (2nd Yr)",'value': None, 'units': 'PERCENT'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                                   
                if not (item in self.track_inputs['output_general']):
                    self.track_inputs['output_general'][item] = {}

                self.track_inputs['output_general'][item]['row'] = row_index
                self.track_inputs['output_general'][item]['unit'] = _unit
                self.track_inputs['output_general'][item]['col'] = 6
                self.track_inputs['output_general'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

                if item in ['annual_revenue_2yr','annual_productivity_tons', 
                        'total_cattle_in_one_pen_per_cycle', 'total_cattle_per_week'] :
                    row_index +=1
                    
        return row_index         
    def _update_sens_section(self, w_sheet, item ):
        #dynamically change initial pens employed
        #eeeeeee
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        if item=='initial_pens_employed':
            self._transfer_cell(w_sheet,'investment_cost','sens','initial_pens_employed',number_format_integer, 'Inputs')
            
        if item=='senior_debt':
            self._transfer_cell(w_sheet,'investment_cost','sens','senior_debt_dynamic_parameter', '0.0%', 'Inputs','senior_debt')
        
        if item=='total_cattle_per_pen_per_cycle':
            self._transfer_cell(w_sheet,'investment_cost','sens','pen_cattle_density',number_format_integer,'Inputs','total_cattle_per_pen_per_cycle')
  
                                 
    def _add_costs_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Costs (Real)' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
   
        #-----------------------------------
        if not 'cost_real' in self.track_inputs:
            self.track_inputs['cost_real']= {}# insert inner
        
        self._add_sub_heading(w_sheet,'Number of workers and supervisors',row_index)
        row_index +=1

        for item in self.sister_model.cost_real.keys():
            if item in self.sister_model.cost_real:
                _unit =self.sister_model.cost_real[item]['units']
                _unit = _unit.upper() if _unit != None else 'BLANK'
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.cost_real[item]['title'], self.sister_model.cost_real[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not item in self.track_inputs['cost_real']:
                    self.track_inputs['cost_real'][item]= {}# insert inner

                self.track_inputs['cost_real'][item]['row'] = row_index
                self.track_inputs['cost_real'][item]['unit'] = _unit
                self.track_inputs['cost_real'][item]['col'] = 6# redundand
                self.track_inputs['cost_real'][item]['value'] = self.sister_model.cost_real[item]['value']

                #---------------Headers--------------------
                if item =='average_num_of_supervisors':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Monthly Wages and Salaries',row_index)

                elif item =='monthly_wage_per_supervisor':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Annual Increase in real salaries',row_index)

                elif item =='annual_increase_salaries_supervisors_technicians':
                   row_index +=1 
                   self._add_sub_heading(w_sheet,'Cost of inputs per ton of Beef',row_index)

                elif item =='cost_of_electricity_per_pen_per_annum':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Annual Price Change',row_index)

                #skip line after
                if item in ['cattle_feed_price_per_kg', 'dressed_weight_at_selling' , 
                            'total_input_costs_per_pen', 'tonnes_produced_per_pen_per_year']:
                   row_index +=1  
                # new line feed                
                row_index +=1
        #fill this place by retrievint from trackinputs-----
        self._update_cost_real_linkedcells(w_sheet)
        #----------------------------------------------------- 
        return row_index


    def _update_cost_real_linkedcells(self, w_sheet,):
        
        
        #  cattle_price_per_unit derived cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_price_per_unit')
        cattle_price_per_unit_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'


        #  cattle_per_pen_per_year derivd cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_per_pen_per_year')
        cattle_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # qnty_of_feed_per_cattle# derived cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','qnty_of_feed_per_cattle')
        qnty_of_feed_per_cattle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # cattle_feed_price_per_kg
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_feed_price_per_kg')
        cattle_feed_price_per_kg_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        

        # live_weight_at_selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # dressed_weight_percent
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_percent')
        dressed_weight_percent_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # cattle_survival_rate
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_survival_rate')
        cattle_survival_rate = get_column_letter(c_index) + str(r_index ) if found_state else None
        

        
        # daily_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','daily_weight_gain')
        daily_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        
        # days_in_feed_lot
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','days_in_feed_lot')
        days_in_feed_lot_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        # live_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_gain')
        live_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + daily_weight_gain_cell + '*' + days_in_feed_lot_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
        
        # weight_at_purchase
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','weight_at_purchase')
        weight_at_purchase_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        
        
        # Live Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + weight_at_purchase_cell + '+' + live_weight_gain_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
       
       

        # food_conversion_ratio
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','food_conversion_ratio')
        food_conversion_ratio_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
       
       
        #qnty_of_feed_per_cattle# recalculated here
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','qnty_of_feed_per_cattle')
        qnty_of_feed_per_cattle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + food_conversion_ratio_cell + '*' + daily_weight_gain_cell + '*' + days_in_feed_lot_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
       
        # Commercial Feed Per Pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','commercial_feed_per_pen')
        commercial_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + '*' + qnty_of_feed_per_cattle_cell
            #w_sheet['%s%s'%(c, r_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
        
       
       
        # Dressed Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_at_selling')
        dressed_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + dressed_weight_percent_cell + '*' + live_weight_at_selling_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
      
     
     
        
        # Cost of cattle per pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_per_pen')
        cost_of_cattle_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + '*' + cattle_price_per_unit_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
         # cost_of_cattle_feed_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_feed_per_pen')
        cost_of_cattle_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_feed_price_per_kg_cell + '*' + commercial_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # total_input_costs_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','total_input_costs_per_pen')
        total_input_costs_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cost_of_cattle_per_pen_cell + '+' + cost_of_cattle_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # tonnes_produced_per_pen_per_year
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','tonnes_produced_per_pen_per_year')
        tonnes_produced_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + \
                                                        '*' + dressed_weight_at_selling_cell + \
                                                        '*' +  cattle_survival_rate + '/1000' 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
 
         
        # cost_of_domestic_inputs_per_ton_of_beef_produced
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_domestic_inputs_per_ton_of_beef_produced')
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + total_input_costs_per_pen_cell + \
                                                        '/' + tonnes_produced_per_pen_per_year_cell 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
        # dressed_weight_at_selling linkedcell @ prosuxt
        self._set_linkedcell(w_sheet, 'production_inventory','dressed_weight_at_selling', dressed_weight_at_selling_cell) 
  
        
        #  cattle_price_per_unit derived cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_price_per_unit')
        cattle_price_per_unit_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'


        #  cattle_per_pen_per_year derivd cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_per_pen_per_year')
        cattle_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # qnty_of_feed_per_cattle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','qnty_of_feed_per_cattle')
        qnty_of_feed_per_cattle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # cattle_feed_price_per_kg
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_feed_price_per_kg')
        cattle_feed_price_per_kg_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        

        # live_weight_at_selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # dressed_weight_percent
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_percent')
        dressed_weight_percent_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # cattle_survival_rate
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_survival_rate')
        cattle_survival_rate = get_column_letter(c_index) + str(r_index ) if found_state else None
        

        
        # daily_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','daily_weight_gain')
        daily_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        
        # days_in_feed_lot
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','days_in_feed_lot')
        days_in_feed_lot_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        # live_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_gain')
        live_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + daily_weight_gain_cell + '*' + days_in_feed_lot_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
        
        # weight_at_purchase
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','weight_at_purchase')
        weight_at_purchase_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        
        
        # Live Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + weight_at_purchase_cell + '+' + live_weight_gain_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
       
       
        # Commercial Feed Per Pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','commercial_feed_per_pen')
        commercial_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + '*' + qnty_of_feed_per_cattle_cell
            #w_sheet['%s%s'%(c, r_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
        
       
       
        # Dressed Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_at_selling')
        dressed_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + dressed_weight_percent_cell + '*' + live_weight_at_selling_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        

         # Cost of cattle per pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_per_pen')
        cost_of_cattle_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + '*' + cattle_price_per_unit_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
         # cost_of_cattle_feed_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_feed_per_pen')
        cost_of_cattle_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_feed_price_per_kg_cell + '*' + commercial_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # total_input_costs_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','total_input_costs_per_pen')
        total_input_costs_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cost_of_cattle_per_pen_cell + '+' + cost_of_cattle_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # tonnes_produced_per_pen_per_year
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','tonnes_produced_per_pen_per_year')
        tonnes_produced_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + cattle_per_pen_per_year_cell + \
                                                        '*' + dressed_weight_at_selling_cell + \
                                                        '*' +  cattle_survival_rate + '/1000' 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
 
         
        # cost_of_domestic_inputs_per_ton_of_beef_produced
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_domestic_inputs_per_ton_of_beef_produced')
        if found_state:
            c= get_column_letter(c_index)
            w_sheet['%s%s'%(c, r_index)].value ='=' + total_input_costs_per_pen_cell + \
                                                        '/' + tonnes_produced_per_pen_per_year_cell 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
        # dressed_weight_at_selling linkedcell @ prosuxt
        self._set_linkedcell(w_sheet, 'production_inventory','dressed_weight_at_selling', dressed_weight_at_selling_cell) 
        
         
    def _add_cal_purchases_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'PURCHASES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_purchases_nominal' in self.track_inputs:
            self.track_inputs['calc_purchases_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['calc_purchases_nominal']):
            self.track_inputs['calc_purchases_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['calc_purchases_nominal']:
            self.track_inputs['calc_purchases_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_purchases_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #==========Sub-headings=========================
        self._add_sub_heading(w_sheet,'Imported Inputs',row_index)

        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['annual_change_in_price_of_imported_inputs','cost_of_imported_inputs_per_ton_of_beef_produced',
                   'real_cif_cost_of_imported_inputs_per_ton',
                   'us_price_index', 'nominal_exchange_rate','import_duty',
                   'nominal_cif_cost_of_imported_inputs_per_ton_usd',
                   'nominal_cif_cost_of_imported_inputs_per_ton_lc', 
                   'cost_of_imported_inputs_including_import_duty_per_ton', 'production_quantity',
                   'total_cost_of_imported_inputs_including_import_duty', 
                   'annual_change_in_price_of_domestic_inputs', 'cost_of_domestic_inputs_per_ton_of_beef_produced', 
                   'real_cost_of_domestic_inputs_per_ton',
                   'domestic_price_index','nominal_cost_of_domestic_inputs_per_ton',
                   'total_cost_of_domestic_inputs','total_cost_of_inputs_per_ton_nominal','total_input_cost_nominal']
        
        ic_parameters = {
            #Imported Inputs	
            'annual_change_in_price_of_imported_inputs': {'title': 'Annual change in price of imported inputs','value': None, 'units': 'PERCENT'},
            'cost_of_imported_inputs_per_ton_of_beef_produced': {'title': 'Cost of imported inputs per ton of Beef produced','value': None, 'units': 'USD'},
            'real_cif_cost_of_imported_inputs_per_ton': {'title': 'Real CIF Cost of imported inputs per ton','value': None, 'units': 'USD'},# linked cell
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},# linked cell
            'nominal_exchange_rate': {'title': 'Nominal Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
            'import_duty': {'title': 'Import duty','value': None, 'units': 'PERCENT'},

            'nominal_cif_cost_of_imported_inputs_per_ton_usd': {'title': 'Nominal CIF Cost of imported inputs per ton','value': None, 'units': 'USD'},
            'nominal_cif_cost_of_imported_inputs_per_ton_lc': {'title': 'Nominal CIF Cost of imported inputs per ton','value': None, 'units': 'LC'},
            'cost_of_imported_inputs_including_import_duty_per_ton': {'title': 'Cost of imported inputs including import duty per ton','value': None, 'units': 'LC'},
            
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'}, 
            'total_cost_of_imported_inputs_including_import_duty': {'title': 'Total cost of imported inputs including import duty','value': None, 'units': 'PERCENT'},
          
            #Domestic Inputs
            'annual_change_in_price_of_domestic_inputs': {'title': 'Annual change in the price of domestic inputs','value': None, 'units': 'PERCENT'},
            'cost_of_domestic_inputs_per_ton_of_beef_produced': {'title': 'Cost of domestic inputs per ton of Beef produced','value': None, 'units': 'LC'},
            'real_cost_of_domestic_inputs_per_ton': {'title': 'Real Cost of domestic inputs per ton','value': None, 'units': 'LC'},
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'nominal_cost_of_domestic_inputs_per_ton': {'title': 'Nominal Cost of domestic inputs per ton','value': None, 'units': 'LC'},
            'total_cost_of_domestic_inputs': {'title': 'Total cost of domestic inputs','value': None, 'units': 'T_LC'},

            'total_cost_of_inputs_per_ton_nominal': {'title': 'Total cost of inputs per ton (nominal)','value': None, 'units': 'LC'},

            'total_input_cost_nominal': {'title': 'Total Input Cost (nominal)','value': None, 'units': 'T_LC'},
            
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_purchases_nominal']):
                    self.track_inputs['calc_purchases_nominal'][item] = {}

                self.track_inputs['calc_purchases_nominal'][item]['row'] = row_index
                self.track_inputs['calc_purchases_nominal'][item]['unit'] = _unit
                self.track_inputs['calc_purchases_nominal'][item]['col'] = 6 #redundant
                self.track_inputs['calc_purchases_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['import_duty','cost_of_imported_inputs_including_import_duty_per_ton',
                            'total_cost_of_imported_inputs_including_import_duty',
                            'total_cost_of_domestic_inputs', 'total_cost_of_inputs_per_ton_nominal',
                            'total_input_cost_nominal'] :
                    row_index +=1 
                    
                    if item =='total_cost_of_imported_inputs_including_import_duty':
                        self._add_sub_heading(w_sheet,'Domestic Inputs',row_index)
                        row_index +=1 
        
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_purchases_nominal(w_sheet)
       
        return row_index 
       
      
    def _populate_cal_purchases_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #0 
        #copy price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','calc_purchases_nominal',
                    'domestic_price_index',number_format) 

        #1. copy import change in prices
        self._transfer_value_tocellrange(w_sheet, 'cost_real','calc_purchases_nominal', 
                 'annual_change_in_price_of_imported_inputs',  "Inputs", None,  '0%')                
        
        #2. copy cost of imported inputs
        self._transfer_cell(w_sheet,'cost_real','calc_purchases_nominal','cost_of_imported_inputs_per_ton_of_beef_produced',number_format, 'Inputs')
        
        
        #3-----REal CIF cost_cost_of_imported_inputs_per_ton--------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_purchases_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # costs of inputs
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_per_ton_of_beef_produced')
        cost_of_imports_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('calc_purchases_nominal','annual_change_in_price_of_imported_inputs')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cif_cost_of_imported_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_cif_cell = prev_letter + str(cell.row) 
                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + cost_of_imports_cell +',' \
                                                               + prev_real_cif_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        
        #4. copy import change in prices
        self._transfer_cell_range(w_sheet, 'calc_inflation_price_index','calc_purchases_nominal', 
                 'us_price_index',   number_format)                
       
        #5. copy norminal
        self._transfer_cell_range(w_sheet, 'calc_inflation_price_index','calc_purchases_nominal', 
                 'nominal_exchange_rate',   number_format)                
       
        #6. copy cost of imported inputs
        self._transfer_cell(w_sheet,'taxes','calc_purchases_nominal','import_duty','0%', 'Inputs')
        
        
        #7-----Nominal CIF cost of imported inputs per ton (usd)--------------
        #real_cif_cost
        real_cif_cost_r_index, c_index, found_state_real_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cif_cost_of_imported_inputs_per_ton')
        
        #us_price_index
        us_price_index_r_index, c_index, found_state_us_price_index= self._retrieve_cell_row_colm('calc_purchases_nominal','us_price_index')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_usd')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    real_cif_cost_cell = cell.column + '$' \
                        + str(real_cif_cost_r_index) if found_state_real_cif_cost else '0'

                    us_price_index_cell = cell.column + '$' \
                        + str(us_price_index_r_index) if found_state_us_price_index else '0'

                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ real_cif_cost_cell + '*' + us_price_index_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #8-----Nominal CIF cost of imported inputs per ton (LC)--------------
        #nominal_cif_cost
        nominal_cif_cost_r_index, c_index, found_state_nominal_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_usd')
        
        #nominal_exchange_rate
        nominal_exchange_rate_r_index, c_index, found_state_nominal_exchange_rate= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_exchange_rate')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_lc')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    nominal_cif_cost_cell = cell.column + '$' \
                        + str(nominal_cif_cost_r_index) if found_state_nominal_cif_cost else '0'

                    nominal_exchange_rate_cell = cell.column + '$' \
                        + str(nominal_exchange_rate_r_index) if found_state_nominal_exchange_rate else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ nominal_cif_cost_cell + '*' + nominal_exchange_rate_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
        #9-----Nominal CIF cost of imported inputs per ton (LC)--------------
        #nominal_cif_cost
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','import_duty')
        import_duty_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        #nominal_cif_cost
        nominal_cif_cost_r_index, _, found_state_nominal_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_lc')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    nominal_cif_cost_cell = cell.column + '$' \
                        + str(nominal_cif_cost_r_index) if found_state_nominal_cif_cost else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + nominal_cif_cost_cell + '*(1+'+ import_duty_cell + ')'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
        #10. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','calc_purchases_nominal',
                    'production_quantity',number_format, 'Inputs') 

        
        
        #11 -----Nominal CIF cost of imported inputs per ton (LC)--------------
        #cost_of_imported_inputs
        cost_of_imported_inputs_r_index, _, found_state_cost_of_imported_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
       
        #prod_qnty
        prod_qnty_r_index, _, found_state_prod_qnty= self._retrieve_cell_row_colm('calc_purchases_nominal','production_quantity')
        
        #total_cost_of_imported--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_imported_inputs_including_import_duty')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cost_of_imported_inputs_cell = cell.column + '$' \
                                          + str(cost_of_imported_inputs_r_index) if found_state_cost_of_imported_inputs else '0'
                   
                    prod_qnty_cell = cell.column + '$' \
                                          + str(prod_qnty_r_index) if found_state_prod_qnty else '0'                      
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + cost_of_imported_inputs_cell + '*'+ prod_qnty_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
       
        #12. copy domestic change in prices
        self._transfer_value_tocellrange(w_sheet, 'cost_real','calc_purchases_nominal', 
                 'annual_change_in_price_of_domestic_inputs',  "Inputs", None,  '0%')                
        

        #13. copy cost of domestic inputs
        self._transfer_cell(w_sheet,'cost_real','calc_purchases_nominal','cost_of_domestic_inputs_per_ton_of_beef_produced',number_format, 'Inputs')
        
        #14.Real cost of Domestic inputs_per_ton--------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_purchases_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # costs of inputs
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_domestic_inputs_per_ton_of_beef_produced')
        cost_of_domestic_input_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('calc_purchases_nominal','annual_change_in_price_of_domestic_inputs')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cost_of_domestic_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_cif_cell = prev_letter + str(cell.row) 
                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + cost_of_domestic_input_cell +',' \
                                                               + prev_real_cif_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        
        #15-----Nominal cost of domestic inputs per ton (usd)--------------
        #real_cost_domestic
        real_cost_domestic_r_index, c_index, found_state_real_cost_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cost_of_domestic_inputs_per_ton')
        
        #domestic_price_index
        domestic_price_index_r_index, c_index, found_state_domestic_price_index= self._retrieve_cell_row_colm('calc_purchases_nominal','domestic_price_index')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    real_cost_domestic_cell = cell.column + '$' \
                        + str(real_cost_domestic_r_index) if found_state_real_cost_domestic else '0'

                    domestic_price_index_cell = cell.column + '$' \
                        + str(domestic_price_index_r_index) if found_state_domestic_price_index else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ real_cost_domestic_cell + '*' + domestic_price_index_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'

        #16-----total_cost_of_domestic_inputs--------------
        #real_cost_domestic
        prod_qnty_r_index, c_index, found_state_prod_qnty= self._retrieve_cell_row_colm('calc_purchases_nominal','production_quantity')
        
        #nominal_cost_of_domestic
        nominal_cost_of_domestic_r_index, c_index, found_state_nominal_cost_of_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_domestic_inputs')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    nominal_cost_of_domestic_cell = cell.column + '$' \
                        + str(nominal_cost_of_domestic_r_index) if found_state_nominal_cost_of_domestic else '0'

                    prod_qnty_cell = cell.column + '$' \
                        + str(prod_qnty_r_index) if found_state_prod_qnty else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ prod_qnty_cell + '*' + nominal_cost_of_domestic_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'

        #17.   total_cost_of_inputs_per_ton_nominal--------------
        #cost_of_imported_inputs
        cost_of_imported_inputs_r_index, c_index, found_state_cost_of_imported_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
        
        #nominal_cost_of_domestic
        nominal_cost_of_domestic_r_index, c_index, found_state_nominal_cost_of_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_inputs_per_ton_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    cost_of_imported_inputs_cell = cell.column + '$' \
                        + str(cost_of_imported_inputs_r_index) if found_state_cost_of_imported_inputs else '0'

                    nominal_cost_of_domestic_cell = cell.column + '$' \
                        + str(nominal_cost_of_domestic_r_index) if found_state_nominal_cost_of_domestic else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_imported_inputs_cell + '+' + nominal_cost_of_domestic_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #18.   total_cost_of_inputs_per_ton_nominal--------------
        #total_cost_of_imported
        total_cost_of_imported_r_index, c_index, found_state_total_cost_of_imported= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_imported_inputs_including_import_duty')
        
        #nominal_cost_of_domestic
        total_cost_of_domestic_inputs_r_index, c_index, found_state_total_cost_of_domestic_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_domestic_inputs')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_input_cost_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    total_cost_of_imported_cell = cell.column + '$' \
                        + str(total_cost_of_imported_r_index) if found_state_total_cost_of_imported else '0'

                    total_cost_of_domestic_inputs_cell = cell.column + '$' \
                        + str(total_cost_of_domestic_inputs_r_index) if found_state_total_cost_of_domestic_inputs else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_cost_of_imported_cell + '+' + total_cost_of_domestic_inputs_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        list_= [
            {
             'header': 'calc_purchases_nominal', 
             'para': 'total_cost_of_imported_inputs_including_import_duty', 
             'action': '+'
            }, 
            {
             'header': 'calc_purchases_nominal', 
             'para': 'total_cost_of_domestic_inputs', 
             'action': '+'
            }
        ]
        self._sum_oflist_cell_range(w_sheet,list_,'calc_purchases_nominal','total_input_cost_nominal')
        
    def _write_calculation_sheet(self, wb):
        #---------------------------Worksheet 3---------------------------
       
        calculation_sheet = wb['Calc']

        # Writing the first row of the csv	
        col = get_column_letter(1)
        calculation_sheet['%s%s'%(col, 1)].value = 'Calculation Sheet ' #% (user)
        calculation_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
       
        #==========Inflation Exchange Rates _Price_indices_section =========================
        row_index +=2
        row_index = self._add_cal_inflation_price_indices_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
      
        #==========Investment( Real) =========================
        row_index +=2
        row_index = self._add_cal_investment_cost_real_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Investment (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_investment_cost_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Production And Sales (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_production_sales_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Purchases (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_purchases_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Working Capital (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_working_capital_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Labour Cost(Nominal) =========================
        row_index +=2
        row_index = self._add_cal_labour_cost_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Calc Of Unit Production(Nominal) =========================
        row_index +=2
        row_index = self._add_cal_unit_of_production_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        

        #==========Finished Product Inventory Valuation (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_finished_product_inventory_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Loan Schedule (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_loan_schedule_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Residual (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_residual_values_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Depreciation For Tax Purposes (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_depreciation_for_tax_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Income Tax Statement (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_income_tax_statement_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
        # hide---------------
        self._hide_empty_cells(calculation_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        self._set_column_dim(calculation_sheet,rangelist,indexlist)
    
    def _add_cal_production_sales_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'PRODUCTION AND SALES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'cal_production_sales_nominal' in self.track_inputs:
            self.track_inputs['cal_production_sales_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_production_sales_nominal']):
            self.track_inputs['cal_production_sales_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_production_sales_nominal']:
            self.track_inputs['cal_production_sales_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_production_sales_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist= ['cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'sales_quantity', 'domestic_price_index','change_in_price', 'base_price',
                   'real_price', 'nominal_price', 'net_sales_revenue', 
                   'sales_tax', 'sales_tax_paid', 'gross_sales_revenue']
                
        ic_parameters = {
            # PI
            'cum_pens_under_harvesting': {'title': 'Cumulative pens under harvesting','value': None, 'units': 'NUMBER'},
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'TONS'},
            'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'TONS'},# linked cell
            'sales_quantity': {'title': 'Sales quantity','value': None, 'units': 'TONS'},# linked cell

            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'change_in_price': {'title': 'Change in real price of Beef','value': None, 'units': 'PERCENT'},
            'base_price': {'title': 'Base price of beef per ton','value': None, 'units': 'LC'},
            'real_price': {'title': 'Real Price of Beef per ton','value': None, 'units': 'LC'},
            'nominal_price': {'title': 'Nominal Price of Beef per ton','value': None, 'units': 'LC'},
            
            'net_sales_revenue': {'title': 'Net sales revenue','value': None, 'units': 'T_LC'},
            'sales_tax': {'title': 'Sales tax','value': None, 'units': 'PERCENT'},
            'sales_tax_paid': {'title': 'Sales tax paid','value': None, 'units': 'T_LC'},
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            
        }
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_production_sales_nominal']):
                    self.track_inputs['cal_production_sales_nominal'][item] = {}

                self.track_inputs['cal_production_sales_nominal'][item]['row'] = row_index
                self.track_inputs['cal_production_sales_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_production_sales_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_production_sales_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['sales_quantity','nominal_price',] :
                    row_index +=1 

        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_production_sales_nominal(w_sheet)
       
        return row_index 
    def _add_cal_unit_of_production_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'UNIT COST OF PRODUCTION'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_unit_of_production' in self.track_inputs:
            self.track_inputs['cal_unit_of_production']= {}# insert inner

        if not ('header' in self.track_inputs['cal_unit_of_production']):
            self.track_inputs['cal_unit_of_production']['header'] = {}
  
        if 'header' in self.track_inputs['cal_unit_of_production']:
            self.track_inputs['cal_unit_of_production']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_unit_of_production', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist= [
            'production_quantity',
            'total_direct_labour_cost',
            'direct_labour_cost_per_ton_of_beef_produced',
            
            'OP',
            'total_cost_of_inputs_per_ton_nominal',
            'total_unit_cost_of_production_per_ton_nominal',
            ]

            
        ic_parameters = {
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'},
            'total_direct_labour_cost': {'title': 'Total Direct Labour Cost','value': None, 'units': 'T_LC'},
            'direct_labour_cost_per_ton_of_beef_produced': {'title': 'Direct Labour cost per ton of beef produced','value': None, 'units': 'LC'},
            
            'OP': {'title': 'Operation period','value': None, 'units': 'FLAGE'},# linked cell
            'total_cost_of_inputs_per_ton_nominal': {'title': 'Total cost of inputs per ton (nominal)','value': None, 'units': 'LC'},
            'total_unit_cost_of_production_per_ton_nominal': {'title': 'Total unit cost of production per ton (nominal)','value': None, 'units': 'LC'},
        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_unit_of_production']):
                    self.track_inputs['cal_unit_of_production'][item] = {}

                self.track_inputs['cal_unit_of_production'][item]['row'] = row_index
                self.track_inputs['cal_unit_of_production'][item]['unit'] = _unit
                self.track_inputs['cal_unit_of_production'][item]['col'] = 6#redundant
                self.track_inputs['cal_unit_of_production'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['direct_labour_cost_per_ton_of_beef_produced'] :
                    row_index +=1 

                    
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_unit_of_production(w_sheet)
       
        return row_index
    
    def _populate_cal_unit_of_production(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        
        
        #1. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_unit_of_production',
                    'production_quantity',number_format, 'Inputs') 

        
        #2. total_direct_labour_cost
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_unit_of_production',
                    'total_direct_labour_cost',number_format,)
        #
        
        #3. Direct Labor cost per ton of beef produced
        list_= [{'header': 'cal_unit_of_production', 'para': 'total_direct_labour_cost', 'action': '+' },
                {'header': 'cal_unit_of_production', 'para': 'production_quantity', 'action': '/' },]
        const_list_= [{'header': 'calc_investment_cost_real', 'para': 'million_to_thousand','action': '*' }] 
        self._sum_oflist_cell_range(w_sheet,list_,'cal_unit_of_production','direct_labour_cost_per_ton_of_beef_produced',const_list_)
  
        
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_unit_of_production', 'para': 'production_quantity','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_unit_of_production', 'para': 'total_direct_labour_cost','cell_type': 'cell_range'},
                {'value':'*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'calc_investment_cost_real', 'para': 'million_to_thousand','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_unit_of_production',
                                            'direct_labour_cost_per_ton_of_beef_produced')
  
    

        #4. copy . OP
        self._transfer_cell_range(w_sheet,'flags','cal_unit_of_production',
                    'OP',number_format_integer, "Inputs")
   

        #5. copy . OP
        self._transfer_cell_range(w_sheet,'calc_purchases_nominal','cal_unit_of_production',
                    'total_cost_of_inputs_per_ton_nominal',number_format_integer)

 
        #6. cost_of_electricity nominal
        listA= [{'header': 'cal_unit_of_production', 'para': 'direct_labour_cost_per_ton_of_beef_produced', 'action': '+' }, 
                {'header': 'cal_unit_of_production', 'para': 'total_cost_of_inputs_per_ton_nominal', 'action': '+' },]
        
        listB= [{'header': 'cal_unit_of_production', 'para': 'OP', 'action': '+' }, ]
        
        self._productof_sumof_list_cell_range(w_sheet,listA, listB,'cal_unit_of_production','total_unit_cost_of_production_per_ton_nominal')
        
    
    def _populate_cal_production_sales_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #copy pens
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'cum_pens_under_harvesting',number_format, 'Inputs') 
        
        #copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'production_quantity',number_format, 'Inputs') 

        #copy production inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'closing_inventory',number_format, 'Inputs') 

        # sales qnty      
        prod_qnty_r_index, _, found_state_prod_qnty= self._retrieve_cell_row_colm('cal_production_sales_nominal','production_quantity')
        closing_inv_r_index, _, found_state_closing_inv= self._retrieve_cell_row_colm('cal_production_sales_nominal','closing_inventory')
      
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_quantity')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    new_column_letter = cell.column # J
                    prev_col=column_index_from_string(new_column_letter)-1
                    prev_letter= get_column_letter(prev_col)
                    
                    prod_qnty_cell = cell.column + '$' + str(prod_qnty_r_index) if found_state_prod_qnty else '0'
                    closing_inv_prev_cell = prev_letter + '$' + str(closing_inv_r_index) if found_state_closing_inv else '0'
                    closing_inv_curr_cell = cell.column + '$' + str(closing_inv_r_index) if found_state_closing_inv else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = "=" + prod_qnty_cell + '+' \
                                                            + closing_inv_prev_cell +  '-' \
                                                            + closing_inv_curr_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        


        #---test
        self._transfer_value_tocellrange(w_sheet, 'prices', 
                        'cal_production_sales_nominal', 'change_in_price',  "Inputs", None,
                        '0%')                                
        
        #------Price-------------------------------------------
        
        self._transfer_cell(w_sheet,'prices','cal_production_sales_nominal','base_price',number_format, 'Inputs')        
        #------saless tax-------------------------------------------
        self._transfer_cell(w_sheet,'taxes','cal_production_sales_nominal','sales_tax','0%', 'Inputs')

        
        #copy price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_production_sales_nominal',
                    'domestic_price_index',number_format) 

        
        #------Real Price-------------------------------------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_production_sales_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # base price
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','base_price')
        base_price_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','change_in_price')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','real_price')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_price_cell = prev_letter + str(cell.row) 
                    #=IF(I89=Inputs!$F$32 ;$F$98, H99*(1+I97))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + base_price_cell +',' \
                                                               + prev_real_price_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #------Nominal Price-------------------------------------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_production_sales_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #real price
        real_price_r_index, _, found_state_real_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','real_price')
        
        #domestic
        domestic_price_index_r_index, _, found_state_domestic_price_index= self._retrieve_cell_row_colm('cal_production_sales_nominal','domestic_price_index')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','nominal_price')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    domestic_price_index_cell = cell.column + '$' \
                        + str(domestic_price_index_r_index) if found_state_domestic_price_index else '0'

                   
                    #----
                    real_price_cell = cell.column + '$' \
                        + str(real_price_r_index) if found_state_real_price else '0'

                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_price_cell = prev_letter + str(cell.row) 
                    #=IF(I89=Inputs!$F$32 ;$F$98, H99*(1+I97))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + real_price_cell +',' \
                                                               + real_price_cell + '*'+ domestic_price_index_cell+')'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                                
        #------NET SALES REVENUE-------------------------------------------
        #sales qnty
        sales_qnty_r_index, _, found_state_sales_qnty= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_quantity')
        #nominal price
        nominal_price_r_index, _, found_state_nominal_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','nominal_price')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #----
                    sales_qnty_cell = cell.column + '$' \
                        + str(sales_qnty_r_index) if found_state_sales_qnty else '0'
                    #----
                    nominal_price_cell = cell.column + '$' \
                        + str(nominal_price_r_index) if found_state_nominal_price else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ sales_qnty_cell + '*' \
                                                               + nominal_price_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=J100*J94/1000
        
        #------LESS SALES TAX-------------------------------------------
        #sales qnty
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax')
        #----
        sales_tax_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
                    
        #nominal price
        net_sales_revenue_r_index, _, found_state_net_sales_revenue= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax_paid')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                    net_sales_revenue_cell = cell.column + '$' \
                        + str(net_sales_revenue_r_index) if found_state_net_sales_revenue else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ sales_tax_cell + '*' \
                                                               + net_sales_revenue_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=$F$103*I102
        #------GROSS SALES REVENUE-------------------------------------------
        #net sales reveneu
        net_sales_revenue_r_index, _, found_state_net_sales_revenue= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
     
        #sales_tax_paid
        sales_tax_paid_r_index, _, found_state_sales_tax_paid= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax_paid')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','gross_sales_revenue')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                    net_sales_revenue_cell = cell.column + '$' \
                        + str(net_sales_revenue_r_index) if found_state_net_sales_revenue else '0'
                    
                    sales_tax_paid_cell = cell.column + '$' \
                        + str(sales_tax_paid_r_index) if found_state_sales_tax_paid else '0'    
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ net_sales_revenue_cell + '+' \
                                                               + sales_tax_paid_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=$F$103*I102

    def _add_cal_investment_cost_real_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'INVESTMENT COST (Real)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_investment_cost_real' in self.track_inputs:
            self.track_inputs['calc_investment_cost_real']= {}# insert inner

        if not ('header' in self.track_inputs['calc_investment_cost_real']):
            self.track_inputs['calc_investment_cost_real']['header'] = {}
  
        if 'header' in self.track_inputs['calc_investment_cost_real']:
            self.track_inputs['calc_investment_cost_real']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_investment_cost_real', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        #==========Sub-heading=========================
        self._add_sub_heading(w_sheet,'Feedlots(real)',row_index)
        row_index +=1

        itemlist= ['cost_of_pens_constructed','investment_cost_of_pens',
                   'million_to_thousand', 'investment_cost_of_land','investment_cost_of_land_real',
                   'investment_costs_over_run_factor','exchange_rate',
                   'cif_cost_of_machinery', 'cif_cost_of_machinery_usd','cif_cost_of_machinery_lc', 
                   'import_duty', 'total_cost_machinery', 'investment_cost_of_buildings', 
                   'investment_cost_of_buildings_real', 'equity', 'senior_debt',
                   'total_investment_cost','total_local_investment',
                   'equity_towards_investment','senior_debt_towards_investment']
        
               
        ic_parameters = {
            # Feedlots
            'cost_of_pens_constructed': {'title': 'Cost of feedlots constructed','value': None, 'units': 'T_LC'},
            'investment_cost_of_pens': {'title': 'Investment cost of Feedlots (real)','value': None, 'units': 'T_LC'},# linked cell
            
            # Land
            'million_to_thousand': {'title': 'Million to thousand conversion','value': None, 'units': 'Thousand'}, 
            'investment_cost_of_land': {'title': 'Investment cost of land','value': None, 'units': 'T_LC'},# linked cell
            'investment_cost_of_land_real': {'title': 'Investment cost of land (real)','value': None, 'units': 'T_LC'},
           
            # Machinery
            'investment_costs_over_run_factor': {'title': 'Investment Costs Over-run Factor','value': None, 'units': 'PERCENT'},
            'exchange_rate': {'title': 'Exchange Rate (LC/USD - base year)','value': None, 'units': 'NUMBER'},
            'cif_cost_of_machinery': {'title': 'CIF cost of Machinery (FC$ 000)','value': None, 'units': 'T_USD'},
            'cif_cost_of_machinery_usd': {'title': 'CIF cost of Machinery','value': None, 'units': 'T_USD'},
            'cif_cost_of_machinery_lc': {'title': 'CIF cost of Machinery','value': None, 'units': 'T_LC'},
            'import_duty': {'title': 'Import duty','value': None, 'units': 'PERCENT'},
            'total_cost_machinery': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            
            # Buildings
            'investment_cost_of_buildings': {'title': 'Investment cost of buildings (LC$ 000)','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_real': {'title': 'Investment cost of buildings (real)','value': None, 'units': 'T_LC'},

            # Financing Parameters
            'equity': {'title': 'Equity (% of Investment Costs)','value': None, 'units': 'PERCENT'},
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': None, 'units': 'PERCENT'},

             # Financing Parameters
            'total_investment_cost': {'title': 'Total Investment Cost (real)','value': None, 'units': 'T_LC'},
            'total_local_investment': {'title': 'Total Local Investment Cost (real)','value': None, 'units': 'T_LC'},
            'equity_towards_investment': {'title': 'Equity Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
            'senior_debt_towards_investment': {'title': 'Senior Debt Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_investment_cost_real']):
                    self.track_inputs['calc_investment_cost_real'][item] = {}

                self.track_inputs['calc_investment_cost_real'][item]['row'] = row_index
                self.track_inputs['calc_investment_cost_real'][item]['unit'] = _unit
                self.track_inputs['calc_investment_cost_real'][item]['col'] = 6#redundant
                self.track_inputs['calc_investment_cost_real'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['investment_cost_of_pens','investment_cost_of_land_real',
                            'total_cost_machinery','investment_cost_of_buildings_real','senior_debt'] :
                    row_index +=1 

                    if item =='investment_cost_of_pens':
                        #from preceeding then this----->Almost deleted this without documentation 
                        self._add_sub_heading(w_sheet,'Land(real)',row_index)
                        row_index +=1

                    if item =='investment_cost_of_land_real':
                        #from preceeding then this----->Almost deleted this without documentation                    
                        self._add_sub_heading(w_sheet,'Machinery(real)',row_index)
                        row_index +=1  

                    if item =='total_cost_machinery':
                        #from preceeding then this----->Almost deleted this without documentation
                        self._add_sub_heading(w_sheet,'Buildings(real)',row_index)
                        row_index +=1 

                    if item =='investment_cost_of_buildings_real':
                        #from preceeding then this----->Almost deleted this without documentation                     
                        self._add_sub_heading(w_sheet,'Financing Parameters',row_index)
                        row_index +=1  

                    if item =='senior_debt':
                        self._add_sub_heading(w_sheet,'Financing',row_index)
                        row_index +=1              

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_investment_cost_real(w_sheet)
       
        return row_index 

    def _write_cashflow_sheet(self, wb):
        #---------------------------Worksheet 3---------------------------
       
        cashflow_sheet = wb['CF']

        # Writing the first row of the csv	
        col = get_column_letter(1)
        cashflow_sheet['%s%s'%(col, 1)].value = 'Cash Flow Sheet ' #% (user)
        cashflow_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(cashflow_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
       
        #==========Cash Flow Nominal =========================
        row_index +=2
        row_index = self._add_cf_nominal_section(cashflow_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
      
        #==========Cash Flow (Real) =========================
        row_index +=2
        row_index = self._add_cf_real_section(cashflow_sheet, row_index,total_wsheet_cols)
        self._update_cashflow_linkedcells(wb)
       
        #======================================+++++++++++++++
        
       
       
        # hide---------------
        self._hide_empty_cells(cashflow_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        self._set_column_dim(cashflow_sheet,rangelist,indexlist)

	
    def _populate_cal_investment_cost_real(self, w_sheet):
         
        #*************************************************Pens*****************************************************
        #------Cost of pens constructed
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'cost_of_pens_constructed',number_format, 'Inputs')        
        #------------Real cost of Pens
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','cost_of_pens_constructed')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_pens_cell +'/' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        #**********************************Land Real*************************************
        #---Land 
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_cost_of_land',number_format, 'Inputs')
        #-----Land (real)-------------------
        self._transfer_cell_range( w_sheet=w_sheet, source_header='calc_investment_cost_real', 
                        target_header='calc_investment_cost_real', source_para='investment_cost_of_land',
                        cell_number_format =number_format, source_wksheet=None, 
                        target_para = 'investment_cost_of_land_real',cell_style='Calculation')                  
        
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        source_cell =1000 # self.modelspec_cell_ref["M_CONVERSION"]

        if 'model_specs' in self.track_inputs:
            if 'M_CONVERSION' in self.track_inputs['model_specs']:     
                row= self.track_inputs['model_specs']['M_CONVERSION']['row']
                col= self.track_inputs['model_specs']['M_CONVERSION']['col']
                source_cell = '=Inputs!' + get_column_letter(6) + str(row)
        
        
        self._transfer_value_tocell(w_sheet, source_cell,'calc_investment_cost_real', 'million_to_thousand',number_format)
       
        #***********************************Trancefer Cells from Input *********************************
        #----Machinery
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'cif_cost_of_machinery',number_format, 'Inputs')
       
        #------Cost overun LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_costs_over_run_factor','0%', 'Inputs')             
        #------Exchange LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'macroeconomic_parameters','calc_investment_cost_real',
                            'exchange_rate',number_format, 'Inputs')

        #------Import LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'taxes','calc_investment_cost_real',
                    'import_duty','0%', 'Inputs')                      
        #------Equity LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_real','equity','0%', 'Inputs')        
        #------Senior LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_real','senior_debt','0%', 'Inputs')


        #------CIF Cost of Machinery USD-------------------------------------------
        #cost overun
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_costs_over_run_factor')
        cost_overun_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        cif_r_index, _, found_state_cif= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_usd')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_cell = cell.column + '$' + str(cif_r_index) if found_state_cif else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_cell +'*(1+'+ cost_overun_cell + ')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #------CIF Cost of Machinery-- LC-----------------------------------------
        #cost overun
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','exchange_rate')
        exchange_rate_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        cif_usd_r_index, _, found_state_cif_usd= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_usd')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_lc')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_usd_cell = cell.column + '$' + str(cif_usd_r_index) if found_state_cif_usd else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_usd_cell +'*'+ exchange_rate_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #------To cost of machiner in LC [importy duty added]-----------------------------------------
        #import duty
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','import_duty')
        import_duty_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif LC
        cif_lc_r_index, _, found_state_cif_lc= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_lc')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_cost_machinery')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_lc_cell = cell.column + '$' + str(cif_lc_r_index) if found_state_cif_lc else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_lc_cell +'*(1+'+ import_duty_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        # Buidingd Real
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_cost_of_buildings',number_format, 'Inputs')
        
        #Building deprecaite over years: land doesnt
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
    
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_buildings_cell +'/' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        

       
         #------Total Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_land_real')
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_real','total_cost_machinery')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_machinery_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #------Total Local Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_land_real')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_local_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Equity toward Total Investment-------------------------
        #equity
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','equity')
        equity_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','equity_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ equity_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Senoir Debt Contributiom Total Investment-------------------------
        #senior_debt
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','senior_debt')
        senior_debt_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','senior_debt_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ senior_debt_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
    def _add_cf_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'CASH FLOW (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cf_nominal' in self.track_inputs:
            self.track_inputs['cf_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cf_nominal']):
            self.track_inputs['cf_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cf_nominal']:
            self.track_inputs['cf_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cf_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'RECEIPTS',row_index)
        row_index +=1

        itemlist= [
                'gross_sales_revenue','change_in_AR','residual_value_of_land_nominal',
                'residual_value_of_machinery_nominal','residual_value_of_buildings_nominal','total_cash_inflow',
                'pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal',

                'total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost',
                'cost_of_electricity_nominal','other_indirect_costs_nominal',

                'change_in_AP','change_in_CB','total_cash_outflow','net_cash_flow_before_taxes','sales_tax_paid',
                'income_tax_payment','net_after_tax_cash_flow_before_financing',
                'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
                'net_after_tax_cash_flow_after_financing',

                'discount_rate_equity','domestic_inflation_rate','nominal_interest_rate',

                'net_after_tax_cash_flow_before_financing_2','total_loan_repayment_outflow_local_2',
                'pv_annual_net_cash_flows', 'pv_annual_debt_repayment',

                'ADSCR','LLCR',
                'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
                'minimum_LLCR','maximum_LLCR','average_LLCR',
            ]

        ic_parameters = {
            # RECEIPTS
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'T_LC'},
            
            # Liquidation value
            'residual_value_of_land_nominal': {'title': 'Residual value of land (nominal)','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_nominal': {'title': "Residual value of machinery (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_nominal': {'title': 'Residual value of buildings (nominal)','value': None, 'units': 'T_LC'},
            'total_cash_inflow': {'title': 'TOTAL CASH INFLOW (+)','value': None, 'units': 'T_LC'},

            #EXPENDITURES
            #--------     Investment
            'pens_nominal': {'title': 'Pens (nominal)','value': None, 'units': 'T_LC'},
            'land_nominal': {'title': "Land (nominal)",'value': None, 'units': 'T_LC'},
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
            'buildings_nominal': {'title': "Buildings (nominal)",'value': None, 'units': 'T_LC'},
             #--------     Operating costs
           
            #Operating costs	
            'total_cost_of_imported_inputs_including_import_duty': {'title': "Total cost of imported inputs including import duty",'value': None, 'units': 'T_LC'},
            'total_cost_of_domestic_inputs': {'title': "Total cost of domestic inputs",'value': None, 'units': 'T_LC'},
            'total_labour_cost': {'title': 'Total Labour Cost','value': None, 'units': 'T_LC'},
            'cost_of_electricity_nominal': {'title': "Cost Of Electricity (nominal)",'value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': "Other Indirect Costs (nominal)",'value': None, 'units': 'T_LC'},

            	
            

            'change_in_AP': {'title': "Change in A/P",'value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},

            'total_cash_outflow': {'title': "TOTAL CASH OUTFLOW (-)",'value': None, 'units': 'T_LC'},

            'net_cash_flow_before_taxes': {'title': "NET CASH FLOW BEFORE TAXES",'value': None, 'units': 'T_LC'},

            'sales_tax_paid': {'title': 'Sales tax paid','value': None, 'units': 'T_LC'},
            'income_tax_payment': {'title': "Income tax payment",'value': None, 'units': 'T_LC'},

            
            'net_after_tax_cash_flow_before_financing': {'title': ' Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)','value': None, 'units': 'T_LC'},
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},

            'total_loan_repayment_outflow_local': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},
            'net_after_tax_cash_flow_after_financing': {'title': 'Net after tax cash flow after financing (OWNER PERSPECTIVE)','value': None, 'units': 'T_LC'},

            
            'discount_rate_equity': {'title': 'Discount Rate Equity','value': None, 'units': 'PERCENT'},
            'domestic_inflation_rate': {'title': "Domestic Inflation rate",'value': None, 'units': 'PERCENT'},
            'nominal_interest_rate': {'title': 'Nominal Interest Rate','value': None, 'units': 'PERCENT'},

            'net_after_tax_cash_flow_before_financing_2': {'title': "Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)",'value': None, 'units': 'T_LC'},
            'total_loan_repayment_outflow_local_2': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},

            'pv_annual_net_cash_flows': {'title': 'PV Annual Net Cash Flows (NCF)','value': None, 'units': 'T_LC'},
            'pv_annual_debt_repayment': {'title': "PV Annual Debt Repayment",'value': None, 'units': 'T_LC'},

            'ADSCR': {'title': "ADSCR",'value': None, 'units': 'NUMBER'},
            'LLCR': {'title': "LLCR",'value': None, 'units': 'NUMBER'},
            
             #Summary of ADSCR
            'minimum_ADSCR': {'title': "Minimum ADSCR",'value': None, 'units': 'NUMBER'},
            'maximum_ADSCR': {'title': "Maximum ADSCR",'value': None, 'units': 'NUMBER'},
            'average_ADSCR': {'title': "Average ADSCR",'value': None, 'units': 'NUMBER'},


             #Summary of LLCR
            'minimum_LLCR': {'title': "Minimum LLCR",'value': None, 'units': 'NUMBER'},
            'maximum_LLCR': {'title': "Maximum LLCR",'value': None, 'units': 'NUMBER'},
            'average_LLCR': {'title': "Average LLCR",'value': None, 'units': 'NUMBER'},
            

        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cf_nominal']):
                    self.track_inputs['cf_nominal'][item] = {}

                self.track_inputs['cf_nominal'][item]['row'] = row_index
                self.track_inputs['cf_nominal'][item]['unit'] = _unit
                self.track_inputs['cf_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cf_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['change_in_AR','total_cash_inflow','buildings_nominal',
                    'other_indirect_costs_nominal','pv_annual_debt_repayment','change_in_CB',
                    'total_cash_outflow','net_cash_flow_before_taxes',
                    'income_tax_payment','net_after_tax_cash_flow_before_financing',
                     
                    'total_loan_repayment_outflow_local','net_after_tax_cash_flow_after_financing',
                    'total_loan_repayment_outflow_local_2',
           
                    'LLCR','average_ADSCR'] :
                    row_index +=1 
                    # double line
                    if item=='net_after_tax_cash_flow_after_financing':
                        row_index +=1

                    if item=='total_cash_inflow':
                        self._add_sub_heading(w_sheet,' EXPENDITURES',row_index)
                        row_index +=1
                    
                        self._add_sub_heading2(w_sheet,'Investment',row_index)
                        row_index +=1

                    if item=='change_in_AR':
                        self._add_sub_heading2(w_sheet,'Liquidation value',row_index)
                        row_index +=1
                    
                    if item=='buildings_nominal':
                        self._add_sub_heading2(w_sheet,'Operating costs',row_index)
                        row_index +=1
                    
                    if item=='LLCR':
                        self._add_sub_heading2(w_sheet,'Summary of ADSCR',row_index)
                        row_index +=1
                    
                    if item=='average_ADSCR':
                        self._add_sub_heading2(w_sheet,'Summary of LLCR',row_index)
                        row_index +=1    
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cf_nominal(w_sheet)
       
        return row_index
    def _populate_cf_nominal(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #1. gross_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_working_capital_nominal','cf_nominal', 'gross_sales_revenue',number_format,'Calc') 
        
        #2. change_in_AR
        self._transfer_cell_range(w_sheet,'cal_working_capital_nominal','cf_nominal', 'change_in_AR',number_format,'Calc') 

        #3. residual_value_of_land_nominal
        for item in ['residual_value_of_land_nominal','residual_value_of_machinery_nominal',
                     'residual_value_of_buildings_nominal']:
            self._transfer_cell_range(w_sheet,'cal_residual_values','cf_nominal', item,number_format,'Calc') 
        
       
        #7. sum  total_cash_inflow
        formalue_string =[
                {'header': 'cf_nominal', 'para': 'gross_sales_revenue'},
                {'header': 'cf_nominal', 'para': 'residual_value_of_buildings_nominal'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_nominal', 'total_cash_inflow')
       
        #8. investment_cost nominal : pens, building, machinery & land
        source_string =[
            {'source': 'investment_cost_of_pens',             'target': 'pens_nominal',                'header':'calc_investment_cost_nominal'},
            {'source': 'investment_cost_of_land_nominal',     'target': 'land_nominal',                'header':'calc_investment_cost_nominal'},
            {'source': 'total_cost_machinery_nominal',        'target': 'total_cost_machinery_nominal','header':'calc_investment_cost_nominal'},
            {'source': 'investment_cost_of_buildings_nominal', 'target': 'buildings_nominal',          'header':'calc_investment_cost_nominal'},
        ]
        for i in source_string: 
            self._transfer_cell_range(w_sheet, i['header'],'cf_nominal',  i['source'], number_format,'Calc', i['target']) 
    
        
        #9. investment_cost nominal : inputs, labour, electricity, indirect , AP CB
        source_string =[
            {'source': 'total_cost_of_imported_inputs_including_import_duty', 'target': 'total_cost_of_imported_inputs_including_import_duty', 'header':'calc_purchases_nominal' },
            {'source': 'total_cost_of_domestic_inputs',                       'target': 'total_cost_of_domestic_inputs',                       'header':'calc_purchases_nominal'},
            {'source': 'total_labour_cost',                                   'target': 'total_labour_cost',                                   'header':'cal_labour_cost_nominal'},
            {'source': 'cost_of_electricity_per_pen_per_annum_nominal',       'target': 'cost_of_electricity_nominal',                         'header':'cal_labour_cost_nominal'},
            {'source': 'other_indirect_costs_nominal',                        'target': 'other_indirect_costs_nominal',                        'header':'cal_labour_cost_nominal'},
            {'source': 'change_in_AP',                                        'target': 'change_in_AP',                                        'header':'cal_working_capital_nominal'},
            {'source': 'change_in_CB',                                        'target': 'change_in_CB',                                        'header':'cal_working_capital_nominal'},
        ]
        for i in source_string: 
            self._transfer_cell_range(w_sheet, i['header'],'cf_nominal',  i['source'], number_format,'Calc', i['target']) 
    
        #10. sum  total_cash_outflow
        formalue_string =[
                {'header': 'cf_nominal', 'para': 'pens_nominal'},
                {'header': 'cf_nominal', 'para': 'change_in_CB'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_nominal', 'total_cash_outflow')

        #11. sum  net_cash_flow_before_taxes
        #=I16-I34
        #=A-B
        #=total_cash_inflow - total_cash_outflow 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'total_cash_inflow','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_cash_outflow','cell_type': 'cell_range'},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_cash_flow_before_taxes')
       
        #12. sales_tax_paid  
        self._transfer_cell_range(w_sheet, 'cal_production_sales_nominal','cf_nominal',  'sales_tax_paid', number_format,'Calc') 
       
        #13. income_tax_payment  
        self._transfer_cell_range(w_sheet, 'cal_income_tax_statement','cf_nominal',  'income_tax_payment', number_format,'Calc') 
        
        #14. sum  net_after_tax_cash_flow_before_financing
        #=I36-I38-I39
        #=A-B-C
        #=net_cash_flow_before_taxes - sales_tax_paid - income_tax_payment
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_cash_flow_before_taxes','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'sales_tax_paid','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_nominal', 'para': 'income_tax_payment','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_after_tax_cash_flow_before_financing')
       
        #15. debt_cash_inflow_in_local_currency  
        self._transfer_cell_range(w_sheet, 'cal_loan_schedule','cf_nominal',  'debt_cash_inflow_in_local_currency', number_format,'Calc') 
       
        #16. total_loan_repayment_outflow_local  
        self._transfer_cell_range(w_sheet, 'cal_loan_schedule','cf_nominal',  'total_loan_repayment_outflow_local', number_format,'Calc') 
       
            
        #17. sum  net_after_tax_cash_flow_after_financing
        #=I41+I43-I44
        #=A+B-C
        #=net_after_tax_cash_flow_before_financing + debt_cash_inflow_in_local_currency - total_loan_repayment_outflow_local
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'debt_cash_inflow_in_local_currency','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_after_tax_cash_flow_after_financing')
       
        
        #18. discount_rate_equity
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_nominal','discount_rate_equity','0%', 'Inputs')
        
        #18. domestic_inflation_rate
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_nominal','domestic_inflation_rate','0%', 'Inputs')
      
        #20. nominal_interest_rate
        #=F49+(1+F49)*F50
        #=A+(1+A)*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'discount_rate_equity'},
                {'value': '+(1+', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'domestic_inflation_rate'},
              
            ]
        self._cell_formalue_fromstring(w_sheet, formalue_string,'cf_nominal', 'nominal_interest_rate',"0.0%")
        
        #21. net_after_tax_cash_flow_before_financing_2 
        #lb, ub, foundstate= self._loan_principal_repayment_periods(self.myworboook['Inputs'])
        lb, ub, foundstate= self.sister_model._get_loan_principal_repayment_bounds()
        LUB= {}
        if  foundstate:
            LUB= {'lb': 9 + lb, 'ub':9 + ub}
        self._transfer_cell_range(w_sheet, 'cf_nominal','cf_nominal','net_after_tax_cash_flow_before_financing', 
                     number_format, None,'net_after_tax_cash_flow_before_financing_2',"Linkedcell", False, LUB) 
        
        self._transfer_cell_range(w_sheet, 'cf_nominal','cf_nominal','total_loan_repayment_outflow_local', 
                     number_format, None,'total_loan_repayment_outflow_local_2',"Linkedcell", False, LUB) 
        
       
        #22. pv_annual_net_cash_flows
        #=NPV($F$51;K52:O52)+J52
        #=A+(1+A)*B
        formalue_string =[
                {'value':'A', 'var_type': 'rate', 'header': 'cf_nominal', 'para': 'nominal_interest_rate',},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing_2'},
            ]
        self._npv_cell_range(w_sheet, formalue_string,'cf_nominal', 'pv_annual_net_cash_flows',LUB)
       
        #23. pv_annual_debt_repayment
        #=NPV($F$51;K52:O52)+J52
        formalue_string =[
                {'value':'A', 'var_type': 'rate', 'header': 'cf_nominal', 'para': 'nominal_interest_rate',},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local_2'},
            ]
        self._npv_cell_range(w_sheet, formalue_string,'cf_nominal', 'pv_annual_debt_repayment',LUB)
        
        #24. ADSCR
        #=MAX(IFERROR(J52/J53;0);0)
        #=A/B
        #=net_after_tax_cash_flow_before_financing_2 / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing_2','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local_2','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'ADSCR',None,LUB)
       
        #24. LLCR
        #=MAX(IFERROR(J55/J56;0);0)
        #=A/B
        #=pv_annual_net_cash_flows / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'pv_annual_net_cash_flows','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'pv_annual_debt_repayment','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'LLCR',None,LUB)
       
        
        #24. LLCR
        #=MAX(IFERROR(J55/J56;0);0)
        #=A/B
        #=pv_annual_net_cash_flows / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'pv_annual_net_cash_flows','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'pv_annual_debt_repayment','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'LLCR',None,LUB)
        
        #25. minimum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'minimum_ADSCR','Calculation',    'MINA',LUB)
        #26. maximum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'maximum_ADSCR','Calculation',    'MAXA',LUB)
        #27. average_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'average_ADSCR','Calculation',    'AVERAGEA',LUB)
       

        #28. minimum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'minimum_LLCR','Calculation',    'MINA',LUB)
        #29. maximum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'maximum_LLCR','Calculation',    'MAXA',LUB)
        #30. average_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'average_LLCR','Calculation',  'AVERAGEA', LUB)
       
        itemlist= [
                'gross_sales_revenue','change_in_AR','residual_value_of_land_nominal',
                'residual_value_of_machinery_nominal','residual_value_of_buildings_nominal','total_cash_inflow',
                
                'pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal',

                'total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost',
                'cost_of_electricity_nominal','other_indirect_costs_nominal',

                'change_in_AP','change_in_CB','total_cash_outflow',
                
                'net_cash_flow_before_taxes','sales_tax_paid',
                'income_tax_payment','net_after_tax_cash_flow_before_financing',
                'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
                'net_after_tax_cash_flow_after_financing',

                'discount_rate_equity','domestic_inflation_rate','nominal_interest_rate',

                'net_after_tax_cash_flow_before_financing_2','total_loan_repayment_outflow_local_2',
                'pv_annual_net_cash_flows', 'pv_annual_debt_repayment',

                'ADSCR','LLCR',
                'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
                'minimum_LLCR','maximum_LLCR','average_LLCR',
            ]
    def _add_cal_investment_cost_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'INVESTMENT COST (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_investment_cost_nominal' in self.track_inputs:
            self.track_inputs['calc_investment_cost_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['calc_investment_cost_nominal']):
            self.track_inputs['calc_investment_cost_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['calc_investment_cost_nominal']:
            self.track_inputs['calc_investment_cost_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_investment_cost_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['domestic_price_index','cost_of_pens_constructed','investment_cost_of_pens',
                   'million_to_thousand', 'investment_cost_of_land','investment_cost_of_land_nominal',
                   'investment_costs_over_run_factor',
                   'total_cost_machinery', 'total_cost_machinery_nominal', 'investment_cost_of_buildings', 
                   'investment_cost_of_buildings_nominal', 'equity', 'senior_debt',
                   'total_investment_cost','total_local_investment',
                   'equity_towards_investment','senior_debt_towards_investment']
        
           
        ic_parameters = {
            # PI
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            
            # Feedlots
            'cost_of_pens_constructed': {'title': 'Cost of feedlots constructed','value': None, 'units': 'T_LC'},
            'investment_cost_of_pens': {'title': 'Investment cost of Feedlots (nominal)','value': None, 'units': 'T_LC'},# linked cell
            
            # Land
            'investment_cost_of_land': {'title': 'Investment cost of land','value': None, 'units': 'T_LC'},# linked cell
            'investment_cost_of_land_nominal': {'title': 'Investment cost of land (nominal)','value': None, 'units': 'T_LC'},
           
            # Machinery
            'total_cost_machinery': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
         
            # Buildings
            'investment_cost_of_buildings': {'title': 'Investment cost of buildings (LC$ 000)','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_nominal': {'title': 'Buildings (nominal)','value': None, 'units': 'T_LC'},
            # Financing Parameters
            'equity': {'title': 'Equity (% of Investment Costs)','value': None, 'units': 'PERCENT'},
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': None, 'units': 'PERCENT'},

             # Financing Parameters
            'total_investment_cost': {'title': 'Total Investment Cost (nominal)','value': None, 'units': 'T_LC'},
            'equity_towards_investment': {'title': 'Equity Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
            'senior_debt_towards_investment': {'title': 'Senior Debt Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_investment_cost_nominal']):
                    self.track_inputs['calc_investment_cost_nominal'][item] = {}

                self.track_inputs['calc_investment_cost_nominal'][item]['row'] = row_index
                self.track_inputs['calc_investment_cost_nominal'][item]['unit'] = _unit
                self.track_inputs['calc_investment_cost_nominal'][item]['col'] = 6#redundant
                self.track_inputs['calc_investment_cost_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['domestic_price_index','investment_cost_of_pens','investment_cost_of_land_nominal',
                            'total_cost_machinery_nominal','investment_cost_of_buildings_nominal','senior_debt'] :
                    row_index +=1 

                     #==========Sub-headings=========================
                    if item =='domestic_price_index':
                        self._add_sub_heading(w_sheet,'Feedlots(nominal)',row_index)
                        row_index +=1


                    if item =='investment_cost_of_pens':
                        self._add_sub_heading(w_sheet,'Land(nominal)',row_index)
                        row_index +=1

                    if item =='investment_cost_of_land_nominal':
                        self._add_sub_heading(w_sheet,'Machinery(nominal)',row_index)
                        row_index +=1  

                    if item =='total_cost_machinery_nominal':
                        self._add_sub_heading(w_sheet,'Buildings(nominal)',row_index)
                        row_index +=1 

                    if item =='investment_cost_of_buildings_nominal':
                        self._add_sub_heading(w_sheet,'Financing Parameters',row_index)
                        row_index +=1  

                    if item =='senior_debt':
                        self._add_sub_heading(w_sheet,'Financing',row_index)
                        row_index +=1              

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_investment_cost_nominal(w_sheet)
       
        return row_index 

    def _add_cal_residual_values_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'RESIDUAL VALUES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_residual_values' in self.track_inputs:
            self.track_inputs['cal_residual_values']= {}# insert inner

        if not ('header' in self.track_inputs['cal_residual_values']):
            self.track_inputs['cal_residual_values']['header'] = {}
  
        if 'header' in self.track_inputs['cal_residual_values']:
            self.track_inputs['cal_residual_values']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_residual_values', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Initial total values',row_index)
        row_index +=1

        itemlist= [
            'investment_cost_of_land_real','total_cost_of_land_real','total_cost_machinery_real',
            'total_cost_machinery','investment_cost_of_buildings_real','total_cost_buildings',

            'economic_life_of_machinery','economic_life_of_buildings',

            'OP','depreciation_machinery','depreciation_buildings',

            'residual_value_of_land_real','residual_value_of_machinery_real', 'residual_value_of_buildings_real',

            'domestic_price_index','residual_value_of_land_nominal','residual_value_of_machinery_nominal','residual_value_of_buildings_nominal'
           
            ]
        
        ic_parameters = {	
            'investment_cost_of_land_real': {'title': 'Investment cost of land (real)','value': None, 'units': 'T_LC'},
            'total_cost_of_land_real': {'title': 'Total cost of land (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery_real': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_real': {'title': "Investment cost of buildings (real)",'value': None, 'units': 'T_LC'},
            'total_cost_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},

            'economic_life_of_machinery': {'title': 'Economic life of machinery','value': None, 'units': 'YEARS'},
            'economic_life_of_buildings': {'title': 'Economic life of buildings','value': None, 'units': 'YEARS'},

            'OP': {'title': "Operation period",'value': None, 'units': 'FLAG'},
            'depreciation_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'depreciation_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},
            
            'residual_value_of_land_real': {'title': 'Residual value of land (real)','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_real': {'title': "Residual value of machinery (real)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_real': {'title': "Residual value of buildings (real)",'value': None, 'units': 'T_LC'},
            
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'residual_value_of_land_nominal': {'title': "Residual value of land (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_nominal': {'title': "Residual value of machinery (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_nominal': {'title': "Residual value of buildings (nominal)",'value': None, 'units': 'T_LC'},
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_residual_values']):
                    self.track_inputs['cal_residual_values'][item] = {}

                self.track_inputs['cal_residual_values'][item]['row'] = row_index
                self.track_inputs['cal_residual_values'][item]['unit'] = _unit
                self.track_inputs['cal_residual_values'][item]['col'] = 6#redundant
                self.track_inputs['cal_residual_values'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['total_cost_buildings','economic_life_of_buildings','depreciation_buildings',
                            'residual_value_of_buildings_real'] :
                    row_index +=1 

                    if item=='total_cost_buildings':
                        self._add_sub_heading(w_sheet,'Economic service life',row_index)
                        row_index +=1

                    if item=='economic_life_of_buildings':
                        self._add_sub_heading(w_sheet,'Depreciation (Real)',row_index)
                        row_index +=1  
                    if item=='depreciation_buildings':
                        self._add_sub_heading(w_sheet,'Residual value (Real)',row_index)
                        row_index +=1

                    if item=='residual_value_of_buildings_real':
                        self._add_sub_heading(w_sheet,'Residual value (Nominal)',row_index)
                        row_index +=1         
              
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_residual_values(w_sheet)
       
        return row_index
    
   
    def _populate_cal_residual_values(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. investment_cost_of_land_real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','cal_residual_values','investment_cost_of_land', 
                           number_format, None,'investment_cost_of_land_real')

        #2. sum cell total_cost_of_land_real
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','investment_cost_of_land_real',
                                          number_format, None,'total_cost_of_land_real','Calculation')        
         
        #3. total_cost_machinery_real
        #==I42+I27
        #=A+B
        #=investment_cost_of_pens +  total_cost_machinery
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'calc_investment_cost_real', 'para': 'investment_cost_of_pens','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'calc_investment_cost_real', 'para': 'total_cost_machinery','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'total_cost_machinery_real')

        #4. sum cell total_cost_machinery
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','total_cost_machinery_real',
                                          number_format, None,'total_cost_machinery','Calculation')        
         
        
        #5. investment_cost_of_land_real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','cal_residual_values','investment_cost_of_buildings_real', 
                           number_format, None,'investment_cost_of_buildings_real')

        #6. sum cell total_cost_buildings
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','investment_cost_of_buildings_real',
                                          number_format, None,'total_cost_buildings','Calculation')        
     
        
        #7. economic_life_of_machinery
        self._transfer_cell(w_sheet,'depreciation','cal_residual_values','economic_life_of_machinery',number_format_integer, 'Inputs')
        
        #8. economic_life_of_buildings
        self._transfer_cell(w_sheet,'depreciation','cal_residual_values','economic_life_of_buildings',number_format_integer, 'Inputs')
       
        #9. OP
        self._transfer_cell_range(w_sheet,'flags','cal_residual_values','OP', number_format_integer, 'Inputs','OP',"Calculation")

        #10. depreciation_machinery
        #=$F$257*J265/$F$261
        #=A*B/C
        #=total_cost_machinery*OP / economic_life_of_machinery
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'total_cost_machinery','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'OP','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'economic_life_of_machinery','cell_type': 'single'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'depreciation_machinery')
        
        #11. depreciation_buildings
        #=$F$259*I265/$F$262
        #=A*B/C
        #=*total_cost_buildings*OP /  economic_life_of_buildings
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'total_cost_buildings','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'OP','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'economic_life_of_buildings','cell_type': 'single'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'depreciation_buildings')
 
        
      

        #12. domestic_price_index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_residual_values','domestic_price_index', number_format, )

       
        #13. residual_value_of_land_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_of_land_real',
                        number_format_integer, None,'residual_value_of_land_real',"Calculation")
       
        #14. residual_value_of_machinery_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_machinery',
                        number_format_integer, None,'residual_value_of_machinery_real',"Calculation")
        
        #15. residual_value_of_buildings_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_buildings',
                        number_format_integer, None,'residual_value_of_buildings_real',"Calculation")
         
        #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_land_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_land_nominal',
                                            number_format,0)
        
          #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_machinery_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_machinery_nominal',
                                            number_format,0)

          #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_buildings_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_buildings_nominal',
                                            number_format,0)                                    
       
    
    def _populate_cal_investment_cost_nominal(self, w_sheet):
         
        #*************************************************Pens*****************************************************
        #copy price index
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','calc_investment_cost_nominal',
                    'domestic_price_index',number_format) 

        #------Cost of pens constructed
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'cost_of_pens_constructed',number_format,)        
        #------------Nominal cost of Pens
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','cost_of_pens_constructed')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_pens_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        #**********************************Land Real*************************************
        #---Land 
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'investment_cost_of_land',number_format)
        
        
         #------------Nominal Land cost
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_land_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        
       
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
          #***********************************Trancefer Cells from Input *********************************
        #----Machinery
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'total_cost_machinery',number_format, )
        
        #------To cost of machiner in LC [importy duty added]-----------------------------------------
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_machinery_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        

       

                  
        #------Equity LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_nominal','equity','0%', 'Inputs')        
        #------Senior LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_nominal','senior_debt','0%', 'Inputs')


      
        # Buiding Real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'investment_cost_of_buildings',number_format)
        
        #Building deprecaite over years: land doesnt
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_buildings_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        

       
         #------Total Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery_nominal')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_machinery_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #------Total Local Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_local_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Equity toward Total Investment-------------------------
        #equity
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity')
        equity_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ equity_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Senoir Debt Contributiom Total Investment-------------------------
        #senior_debt
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','senior_debt')
        senior_debt_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','senior_debt_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ senior_debt_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
    
    
    def _add_cf_real_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'CASH FLOW  (Real)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cf_real' in self.track_inputs:
            self.track_inputs['cf_real']= {}# insert inner

        if not ('header' in self.track_inputs['cf_real']):
            self.track_inputs['cf_real']['header'] = {}
  
        if 'header' in self.track_inputs['cf_real']:
            self.track_inputs['cf_real']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cf_real', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       
        itemlist= [
            'domestic_price_index','gross_sales_revenue', 'change_in_AR',

            'residual_value_of_land_real','residual_value_of_machinery_real','residual_value_of_buildings_real',
            'total_cash_inflow',

            'pens_real','land_real','machinery_real','buildings_real',
            'total_cost_of_imported','total_cost_of_domestic_inputs','total_labour_cost',
            'cost_of_electricity_real','other_indirect_costs_real',

            'change_in_AP', 'change_in_CB','total_cash_outflow', 'net_cash_flow_before_taxes',
            'sales_tax_paid','income_tax_paid',
            'net_after_tax_cash_flow_before_financing','debt_cash_inflow_in_local_currency',

            'total_loan_repayment_outflow_local', 'net_after_tax_cash_flow_after_financing',

            'discount_rate_equity', 'NPV','IRR','MIRR'
            ]

       		
        ic_parameters = {
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            #RECEIPTS
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'PERCENT'},
            
             # Liquidation value
            'residual_value_of_land_real': {'title': 'Land ','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_real': {'title': "Machinery ",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_real': {'title': 'Buildings','value': None, 'units': 'T_LC'},
            'total_cash_inflow': {'title': 'TOTAL CASH INFLOW (+)','value': None, 'units': 'T_LC'},


            #EXPENDITURES
            #--------     Investment
            'pens_real': {'title': 'Pens','value': None, 'units': 'T_LC'},
            'land_real': {'title': "Land",'value': None, 'units': 'T_LC'},
            'machinery_real': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'buildings_real': {'title': "Buildings ",'value': None, 'units': 'T_LC'},
           
			
             #Operating costs	
            'total_cost_of_imported': {'title': "Imported inputs purchased",'value': None, 'units': 'T_LC'},
            'total_cost_of_domestic_inputs': {'title': "Domestic inputs purchased",'value': None, 'units': 'T_LC'},
            'total_labour_cost': {'title': 'Labour','value': None, 'units': 'T_LC'},
            'cost_of_electricity_real': {'title': "Cost Of Electricity",'value': None, 'units': 'T_LC'},
            'other_indirect_costs_real': {'title': "Other Indirect Costs",'value': None, 'units': 'T_LC'},

          
            'change_in_AP': {'title': "Change in A/P",'value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},

            'total_cash_outflow': {'title': "TOTAL CASH OUTFLOW (-)",'value': None, 'units': 'T_LC'},

            'net_cash_flow_before_taxes': {'title': "NET CASH FLOW BEFORE TAXES",'value': None, 'units': 'T_LC'},

            'sales_tax_paid': {'title': 'Sales tax','value': None, 'units': 'T_LC'},
            'income_tax_paid': {'title': "Income tax",'value': None, 'units': 'T_LC'},

            
            'net_after_tax_cash_flow_before_financing': {'title': 'Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)','value': None, 'units': 'T_LC'},
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},

            'total_loan_repayment_outflow_local': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},
            'net_after_tax_cash_flow_after_financing': {'title': 'Net after tax cash flow after financing (OWNER PERSPECTIVE)','value': None, 'units': 'T_LC'},

             'discount_rate_equity': {'title': "Discount Rate Equity",'value': None, 'units': 'PERCENT'},

            'NPV': {'title': "NPV",'value': None, 'units': 'T_LC'},

            'IRR': {'title': 'IRR','value': None, 'units': 'PERCENT'},
            'MIRR': {'title': "MIRR",'value': None, 'units': 'PERCENT'},


            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
               
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
               
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cf_real']):
                    self.track_inputs['cf_real'][item] = {}

                self.track_inputs['cf_real'][item]['row'] = row_index
                self.track_inputs['cf_real'][item]['unit'] = _unit
                self.track_inputs['cf_real'][item]['col'] = 6#redundant
                self.track_inputs['cf_real'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['total_cash_inflow','change_in_AR','buildings_real',
                             'domestic_price_index','total_loan_repayment_outflow_local',
                              'other_indirect_costs_real','change_in_CB', 'total_cash_outflow', 
                              'net_cash_flow_before_taxes','income_tax_paid', 
                              'net_after_tax_cash_flow_before_financing','net_after_tax_cash_flow_after_financing', ] :
                    row_index +=1 

                    if item=='domestic_price_index':
                        self._add_sub_heading(w_sheet,'RECEIPTS',row_index)
                        row_index +=1

                    if item=='total_cash_inflow':
                        self._add_sub_heading(w_sheet,'EXPENDITURES',row_index)
                        row_index +=1 
                        
                        self._add_sub_heading2(w_sheet,'Investment',row_index)
                        row_index +=1       


                    if item=='change_in_AR':
                        self._add_sub_heading2(w_sheet,'Liquidation value',row_index)
                        row_index +=1       

                  
                    if item=='buildings_real':
                        self._add_sub_heading2(w_sheet,'Operation Cost',row_index)
                        row_index +=1       

                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cf_real(w_sheet)
       
        return row_index
       
    
    def _populate_cf_real(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        number_format_percent ='0%'
        
       
        #1. domestic_price_index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cf_real', 'domestic_price_index',number_format,'Calc') 

        #2. gross_sales_revenue
        #=A/B
        #=gross_sales_revenue /domestic_price_index 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'gross_sales_revenue','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'gross_sales_revenue',number_format)

   
        #3. change_in_AR
        #=A/B
        #=gross_sales_revenue /domestic_price_index 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'change_in_AR','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'change_in_AR',number_format)

        
         
        #4 Group Sum Residual Values
        source= ['residual_value_of_land_nominal','residual_value_of_machinery_nominal','residual_value_of_buildings_nominal']
        target = ['residual_value_of_land_real','residual_value_of_machinery_real','residual_value_of_buildings_real',]
        for i in range(len(source)):
            # LOOP
            #=A/B
            #=pens_nominal /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)
      
       
        #5. sum  total_cash_inflow
        formalue_string =[
                {'header': 'cf_real', 'para': 'gross_sales_revenue'},
                {'header': 'cf_real', 'para': 'residual_value_of_buildings_real'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_real', 'total_cash_inflow')
       
        #6 Group Sum Investments
        source= ['pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal']
        target = ['pens_real','land_real','machinery_real','buildings_real',]
        for i in range(len(source)):
            #7. pens_real
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #6 Group Sum Operating costs
        source= ['total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost', 'cost_of_electricity_nominal','other_indirect_costs_nominal',  'change_in_AP','change_in_CB',]
        target = ['total_cost_of_imported',                             'total_cost_of_domestic_inputs','total_labour_cost', 'cost_of_electricity_real',   'other_indirect_costs_real',   'change_in_AP', 'change_in_CB']
        for i in range(len(source)):
            #7. 
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #7. sum  total_cash_outflow
        formalue_string =[
                {'header': 'cf_real', 'para': 'pens_real'},
                {'header': 'cf_real', 'para': 'change_in_CB'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_real', 'total_cash_outflow')

                
        #8 Group Sum Other Operating costs
        source= ['sales_tax_paid',  'income_tax_payment',  'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local']
        target =['sales_tax_paid','income_tax_paid',       'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local']
        for i in range(len(source)):
            #7. 
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #9. sum  net_cash_flow_before_taxes
        #=I16-I34
        #=A-B
        #=total_cash_inflow - total_cash_outflow 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'total_cash_inflow','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'total_cash_outflow','cell_type': 'cell_range'},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_cash_flow_before_taxes')
       

        #10. sum  net_after_tax_cash_flow_before_financing
        #=I36-I38-I39
        #=A-B-C
        #=net_cash_flow_before_taxes - sales_tax_paid - income_tax_payment
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_cash_flow_before_taxes','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'sales_tax_paid','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_real', 'para': 'income_tax_paid','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_after_tax_cash_flow_before_financing')
       

        #11. sum  net_after_tax_cash_flow_after_financing
        #=I41+I43-I44
        #=A+B-C
        #=net_after_tax_cash_flow_before_financing + debt_cash_inflow_in_local_currency - total_loan_repayment_outflow_local
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_before_financing','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'debt_cash_inflow_in_local_currency','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_real', 'para': 'total_loan_repayment_outflow_local','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_after_tax_cash_flow_after_financing')
       
        #12. discount_rate_equity
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_real','discount_rate_equity','0%', 'Inputs')
      
        #12. NPV
        formalue_string =[
            {'value':'A', 'var_type': 'rate', 'header': 'cf_real', 'para': 'discount_rate_equity',},
            {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._npv_cell(w_sheet, formalue_string,'cf_real', 'NPV',None)

        #13. IRR
        formalue_string =[
            {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._irr_cell(w_sheet, formalue_string,'cf_real', 'IRR',number_format_percent)

        #14. MIRR
        formalue_string =[
            {'value':'A', 'var_type': 'rate', 'header': 'cf_real', 'para': 'discount_rate_equity',},
            {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._mirr_cell(w_sheet, formalue_string,'cf_real', 'MIRR',number_format_percent)
 
    def _add_cal_labour_cost_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'LABOUR COSTS AND OTHER INDIRECT COSTS (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_labour_cost_nominal' in self.track_inputs:
            self.track_inputs['cal_labour_cost_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_labour_cost_nominal']):
            self.track_inputs['cal_labour_cost_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_labour_cost_nominal']:
            self.track_inputs['cal_labour_cost_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_labour_cost_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        self._add_sub_heading(w_sheet,'DIRECT LABOUR',row_index)
        row_index +=1

        itemlist= ['num_of_workers','annual_increase_salaries_workers','monthly_wage_per_worker',
            'number_of_months_in_a_year', 
            
            'real_yearly_wage_rate','domestic_price_index',
            'nominal_yearly_wage_rate','total_direct_labour_cost',

            'num_of_supervisors_technicians','annual_increase_salaries_supervisors_technicians',
            'monthly_wage_per_supervisor','real_yearly_salary_rate','nominal_yearly_salary_rate',
            'total_indirect_labour_cost',

            'total_labour_cost',

            'OP','cost_of_electricity_per_pen_per_annum',
            'cost_of_electricity_per_pen_per_annum_nominal',
            'other_indirect_costs',
            'other_indirect_costs_nominal',
            ]
            
        ic_parameters = {
            # DIRECT LABOUR
            'num_of_workers': {'title': 'Number of workers','value': None, 'units': 'NUMBER'},
            'annual_increase_salaries_workers': {'title': 'Annual increase in real salaries of workers','value': None, 'units': 'PERCENT'},
            'monthly_wage_per_worker': {'title': 'Monthly wage for workers','value': None, 'units': 'T_LC'},
            'number_of_months_in_a_year': {'title': 'Number of months in a year','value': None, 'units': 'NUMBER'},# linked cell
            'real_yearly_wage_rate': {'title': 'Real yearly wage rate','value': None, 'units': 'T_LC'},

            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'nominal_yearly_wage_rate': {'title': 'Nominal yearly wage rate','value': None, 'units': 'T_LC'},
            'total_direct_labour_cost': {'title': 'Total Direct Labour Cost','value': None, 'units': 'T_LC'},
            
            #INDIRECT LABOUR
            'num_of_supervisors_technicians': {'title': 'Number of supervisors & technicians','value': None, 'units': 'NUMBER'},
            'annual_increase_salaries_supervisors_technicians': {'title': 'Annual increase in real salaries of supervisors & technicians','value': None, 'units': 'PERCENT'},
            'monthly_wage_per_supervisor': {'title': 'Monthly wage for supervisors & technicians','value': None, 'units': 'T_LC'},
            'real_yearly_salary_rate': {'title': 'Real yearly salary rate','value': None, 'units': 'T_LC'},
            'nominal_yearly_salary_rate': {'title': 'Nominal yearly salary rate','value': None, 'units': 'T_LC'},
            'total_indirect_labour_cost': {'title': 'Total Indirect Labour Cost','value': None, 'units': 'T_LC'},
            
            'total_labour_cost': {'title': 'Total Labour Cost','value': None, 'units': 'T_LC'},


            'OP': {'title': 'Operation period','value': None, 'units': 'FLAG'},
            #'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'cost_of_electricity_per_pen_per_annum': {'title': 'Cost Of Electricity per pen per annum','value': None, 'units': 'LC'},
            'cost_of_electricity_per_pen_per_annum_nominal': {'title': 'Cost of electricity (nominal)','value': None, 'units': 'T_LC'},
            'other_indirect_costs': {'title': 'Other Indirect Costs','value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': 'Other Indirect Costs (nominal)','value': None, 'units': 'T_LC'},


        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_labour_cost_nominal']):
                    self.track_inputs['cal_labour_cost_nominal'][item] = {}

                self.track_inputs['cal_labour_cost_nominal'][item]['row'] = row_index
                self.track_inputs['cal_labour_cost_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_labour_cost_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_labour_cost_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['total_direct_labour_cost','real_yearly_wage_rate', 
                              'total_indirect_labour_cost', 'total_labour_cost' ] :
                    row_index +=1 

                    if item =='total_direct_labour_cost':
                        self._add_sub_heading(w_sheet,'INDIRECT LABOUR',row_index)
                        row_index +=1


        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_labour_cost_nominal(w_sheet)
       
        return row_index
   
    def _populate_cal_labour_cost_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        #o.copy domestic price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_labour_cost_nominal',
                    'domestic_price_index',number_format) 

        #1. num_of_workers
        self._transfer_cell_range(w_sheet,'cost_real','cal_labour_cost_nominal',
                    'num_of_workers',number_format, "Inputs")
        #2. copy annual_increase_salaries_workers
        self._transfer_value_tocellrange(w_sheet, 'cost_real', 'cal_labour_cost_nominal', 
                                  'annual_increase_salaries_workers',  "Inputs", None,    '0%')   
        
        #3. monthly_wage_per_worker
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','monthly_wage_per_worker',number_format, 'Inputs')
        

        #4. number_of_months_in_a_year
        self._transfer_cell(w_sheet,'timing','cal_labour_cost_nominal','number_of_months_in_a_year',number_format_integer, 'Inputs')
       
        
        #5... Real labour cost
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_labour_cost_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','monthly_wage_per_worker')
        monthly_wage_per_worker_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','number_of_months_in_a_year')
        number_of_months_in_a_year_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #annual_increase_salaries_workers
        annual_increase_salaries_workers_r_index, c_index, found_state_annual_increase_salaries_workers= self._retrieve_cell_row_colm('cal_labour_cost_nominal','annual_increase_salaries_workers')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','real_yearly_wage_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    annual_increase_salaries_workers_cell = cell.column + '$' \
                        + str(annual_increase_salaries_workers_r_index) if found_state_annual_increase_salaries_workers else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_yearly_rate_cell = prev_letter + str(cell.row) 
                    #=IF(K$153=Inputs!$F$32;$F$158*$F$159;J160*(1+K157))
                    year_salary =f'{monthly_wage_per_worker_cell} * {number_of_months_in_a_year_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + year_salary +',' \
                                                               + prev_yearly_rate_cell + '*(1+'+ annual_increase_salaries_workers_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        #6. nominal_yearly_wage_rate

        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'real_yearly_wage_rate',  'action': '+'  }, 
                { 'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','nominal_yearly_wage_rate')

        
        #7. total_direct_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'num_of_workers', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'nominal_yearly_wage_rate', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_direct_labour_cost')
        
        #8. num_of_supervisors_technicians
        self._transfer_cell_range(w_sheet,'cost_real','cal_labour_cost_nominal',
                    'num_of_supervisors_technicians',number_format, "Inputs")

        #9. copy annual_increase_salaries_supervisors_technicians
        self._transfer_value_tocellrange(w_sheet, 'cost_real', 'cal_labour_cost_nominal', 
                                  'annual_increase_salaries_supervisors_technicians',  "Inputs", None,    '0%')   
        
        #10. monthly_wage_per_supervisor
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal',
                                 'monthly_wage_per_supervisor',number_format, 'Inputs')
        

        #11... Real labour cost
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','monthly_wage_per_supervisor')
        monthly_wage_per_supervisor_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','number_of_months_in_a_year')
        number_of_months_in_a_year_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #annual_increase_salaries_workers
        annual_increase_salaries_supervisors_technicians_r_index, c_index, found_state_annual_increase_salaries_supervisors_technicians= self._retrieve_cell_row_colm('cal_labour_cost_nominal','annual_increase_salaries_supervisors_technicians')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','real_yearly_salary_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    annual_increase_salaries_supervisors_technicians_cell = cell.column + '$' \
                        + str(annual_increase_salaries_supervisors_technicians_r_index) if found_state_annual_increase_salaries_supervisors_technicians else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_yearly_rate_cell = prev_letter + str(cell.row) 
                    #=IF(K$153=Inputs!$F$32;$F$158*$F$159;J160*(1+K157))
                    year_salary =f'{monthly_wage_per_supervisor_cell} * {number_of_months_in_a_year_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + year_salary +',' \
                                                               + prev_yearly_rate_cell + '*(1+'+ annual_increase_salaries_supervisors_technicians_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #12. nominal_yearly_salary_rate

        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'real_yearly_salary_rate',  'action': '+'  }, 
                { 'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','nominal_yearly_salary_rate')
        
        #13. total_indirect_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'num_of_supervisors_technicians', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'nominal_yearly_salary_rate', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_indirect_labour_cost')
        
        
        #14. total_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'total_direct_labour_cost', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'total_indirect_labour_cost', 'action': '+' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_labour_cost')
        
        #15. OP
        self._transfer_cell_range(w_sheet,'flags','cal_labour_cost_nominal',
                    'OP',number_format_integer, "Inputs")

        #16. cost_of_electricity per pen
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','cost_of_electricity_per_pen_per_annum',number_format, 'Inputs')
        
        #17. cost_of_electricity nominal
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'OP', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' },
                {'header': 'cal_production_sales_nominal', 'para': 'cum_pens_under_harvesting', 'action': '*' }]
        
        const_list_= [{'header': 'cal_labour_cost_nominal', 'para': 'cost_of_electricity_per_pen_per_annum','action': '*' }, 
                {'header': 'calc_investment_cost_real', 'para': 'million_to_thousand' ,'action': '/'}]   

        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','cost_of_electricity_per_pen_per_annum_nominal',const_list_)
        
        #18. other_indirect_costs
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','other_indirect_costs',number_format, 'Inputs')
      
        
        #19. other_indirect_costs nominal
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'OP', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        const_list_= [{'header': 'cal_labour_cost_nominal', 'para': 'other_indirect_costs','action': '*' },
                      {'header': 'calc_investment_cost_real', 'para': 'million_to_thousand' ,'action': '/'}] 
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','other_indirect_costs_nominal',const_list_)
    
     
      
    