import copy
import datetime
from io import BytesIO
import math
from django.http import HttpResponse
import numpy as np
from common.utils import stringfy_list
from investments_appraisal.base_model import BaseModel
from beefapp.beef_excel_model import BeefExcelReport
from openpyxl import Workbook, load_workbook 
from datetime import date
import numpy_financial as npf
from investments_appraisal.fin import irr2_
from investments_appraisal.whatif import get_alternative_sd,get_data_table_sensitivity_in_parrallel, get_sensitivity_gradient
from django.utils.encoding import smart_str, force_str
class BeefBusinessReport(BaseModel):
    
    myworboook = None
    project_title = 'BEEF FARMING OUTPUTS'
    description= 'BEEF FARMING OUTPUTS'
    #no input from this : to keep value ONLY
    #keep database table of these paramere
    investment_parameter_options ={
        'cost_of_land_per_sqm':     [1.00,	 2.00,	 3.00 ,	 5.00 ,	 5.00,	 5.00],  
        'cost_of_machinery_per_pen':[1000, 	2500, 	5000, 	5000, 	5000, 	5000], 
        'cost_of_building_per_sqm': [10.00, 10.00, 	 8.00, 	 5.00, 	 2.00, 	 1.00], 
        'cost_of_pen_construction': [ 1200,  1200, 	 1500, 	 2000, 	 2000 ,	 3000],
        'senior_debt_dynamic_parameter': [.70, .70, .70, .70, .70, .70],
        'initial_pens_employed':  [1 ,	 5, 	 20, 	 50, 	 100, 	 1000], 
        'pen_cattle_density':   [20, 	 20, 	 20, 	 20, 	 20, 	 20] , 
        'pen_length':           [8.00,	8.00,	8.00,	8.00,	8.00,	8.00],
        'pen_width':            [20.00,	20.00,	20.00,	20.00,	20.00,	20.00], 
        'pen_height':           [2.00,	2.00,	2.00,	2.00,	2.00,	2.00],
        'total_land_required':  [1000, 	 1500, 	 5000, 	 10000, 	 50000,	 100000 ]
    }

    
    def __init__(
        self, model_specifications=None, timing_assumptions=None, 
                 prices= None, working_capital=None,
                 financing=None, taxes=None, 
                 macroeconomic_parameters=None,  depreciation=None, 
                 cost_real=None, investment_cost= None, feedlot_design_parameters= None, 
                 cattle_business_options= None, investment_parameter_options= None):
        #initialise
        BaseModel.__init__(self, model_specifications=model_specifications, timing_assumptions=timing_assumptions, 
                 prices= prices, working_capital=working_capital,
                 financing=financing, taxes=taxes, 
                 macroeconomic_parameters=macroeconomic_parameters,  depreciation=depreciation )
        self.cost_real = cost_real
        self.feedlot_design_parameters = feedlot_design_parameters
        self.investment_cost = investment_cost
        self.cattle_business_options = cattle_business_options
        self.investment_parameter_options = investment_parameter_options
        
        #create sister model that also refer back to this model
        self.excel_obj = BeefExcelReport(self)
        #setattr(self.excel_obj,'description' ,self.description)	

         
        for i in self.cattle_business_options.keys():
            self.excel_obj.track_inputs['cattle_business_options'][i]={}# {"name": Jonh, peter}
        
        #initial calculate this values
        self._cal_metrix()
     
    def _set_model_description(self,user_model_decription):
         #addition
        setattr(self.excel_obj,'user_model_decription' ,user_model_decription)

    
    def _set_parameters_simulation(self,num_reps = 10):
        from numpy.random import default_rng
        rg = default_rng(4470)
        from scipy.stats import norm
        
        from scipy import rand, stats

    
        base_scenario_inputs = {
            """     remove this""" 
            # 'initial_pens_employed':np.array([50]),

            'dressed_weight_at_selling': rg.uniform(0.8*self.dressed_weight_at_selling, get_alternative_sd(self.dressed_weight_at_selling,.3,30), num_reps),
            'cattle_price_per_unit': stats.triang.rvs(loc=.4*self.cattle_price_per_unit,scale= get_alternative_sd(self.cattle_price_per_unit,.9, 20),c=.5,size= num_reps),
            'cattle_survival_rate': stats.triang.rvs(loc=1.-.1,scale=.1,c=1, size= num_reps),# step function with prob
            #'base_price': stats.pearson3.rvs(loc=.8*self.base_price,   scale=get_alternative_sd(self.base_price, .25, 1000),skew=1 ,size=num_reps), 
            
            'cattle_feed_price_per_kg': stats.triang.rvs(loc=.25*self.cattle_feed_price_per_kg,scale=5*self.cattle_feed_price_per_kg,c=.1, size= num_reps),# step function with prob
            
            'base_price': stats.triang.rvs(loc=.5*self.base_price,scale=get_alternative_sd(self.base_price, .5, 1000),c=.5, size= num_reps),# step function with prob
            'discount_rate_equity': rg.uniform(.5*self.discount_rate_equity, get_alternative_sd(self.discount_rate_equity, 1.2, 0.2), num_reps),
            'domestic_inflation_rate': rg.uniform(.5*self.domestic_inflation_rate, get_alternative_sd(self.domestic_inflation_rate, 1.2, 0.1), num_reps),
            'exchange_rate': rg.uniform(1, 5, num_reps),
            'annual_change_in_price_of_domestic_inputs': rg.normal(self.annual_change_in_price_of_domestic_inputs, get_alternative_sd(self.annual_change_in_price_of_domestic_inputs,3,.3), num_reps),
        
            
            'accounts_payable': rg.normal(self.accounts_payable, .3* self.accounts_payable, num_reps),       
            'accounts_receivable': rg.normal(self.accounts_receivable, .3* self.accounts_receivable, num_reps),
            'change_in_price': rg.normal(self.change_in_price, get_alternative_sd(self.change_in_price,3,.3), num_reps),
            'cash_balance': rg.normal(self.cash_balance, .3* self.cash_balance, num_reps),
            'other_indirect_costs': rg.uniform(self.other_indirect_costs, 50* self.other_indirect_costs, num_reps),
            'senior_debt': np.array([1, 0.7, 0.3 ,.1, .0]),
        
            
            'monthly_wage_per_worker':  rg.uniform(self.monthly_wage_per_worker, 5* self.monthly_wage_per_worker, num_reps),
            'monthly_wage_per_supervisor':  rg.uniform(self.monthly_wage_per_supervisor, 5* self.monthly_wage_per_supervisor, num_reps),
            'annual_increase_salaries_workers':  rg.uniform(self.annual_increase_salaries_workers, get_alternative_sd(self.annual_increase_salaries_workers,.4,.2), num_reps),
            'annual_increase_salaries_supervisors_technicians':  rg.uniform(self.annual_increase_salaries_supervisors_technicians, get_alternative_sd(self.annual_increase_salaries_supervisors_technicians,.4,.2), num_reps),
            
            #'num_workers_per_supervisor': np.array([5]),      
        }
        setattr(self, 'base_scenario_inputs', base_scenario_inputs)

 
    def _model_outputs_dict(self,):
        #run with dynamic vars
        self._simulate_metrix()
        dict_ ={}

        dict_['npv']=self.model_npv if hasattr(self, 'model_npv') else 0
        dict_['irr']=self.irr if hasattr(self, 'irr') else 0
        dict_['mirr']=self.mirr if hasattr(self, 'mirr') else 0
        dict_['adscr_av']=self.adscr_av if hasattr(self, 'adscr_av') else 0
        dict_['llcr_av']=self.llcr_av if hasattr(self, 'llcr_av') else 0
        return dict_

    def _model_datatable_outputs_dict(self,):
        #run with dynamic vars
        self._simulate_metrix()
        dict_ ={}
        dict_['npv']=self.model_npv if hasattr(self, 'model_npv') else 0
        dict_['adscr-3']=self.adscr_3 if hasattr(self, 'adscr_3') else 0
        dict_['adscr-4']=self.adscr_4 if hasattr(self, 'adscr_4') else 0
        dict_['adscr-5']=self.adscr_5 if hasattr(self, 'adscr_5') else 0
        dict_['llcr-2']=self.llcr_2 if hasattr(self, 'llcr_2') else 0
        dict_['llcr-3']=self.llcr_3 if hasattr(self, 'llcr_3') else 0
        dict_['llcr-4']=self.llcr_4 if hasattr(self, 'llcr_4') else 0
        dict_['llcr-5']=self.llcr_5 if hasattr(self, 'llcr_5') else 0
        dict_['irr']=self.irr if hasattr(self, 'irr') else 0
        dict_['mirr']=self.mirr if hasattr(self, 'mirr') else 0
          #For testing 09 March 2022
        dict_['net_cash_real']=self.model_net_cash_real if hasattr(self, 'model_net_cash_real') else 0
       
        return dict_

  
    def _outputs_npv(self,):
        #run with dynamic vars
        self._simulate_metrix()
        
        return self.model_npv
     
    def _metric_priceIndex(self):       
        
        if not hasattr(self, 'accounts_receivable'):
            accounts_receivable= self.working_capital['accounts_receivable']['value']       
            setattr(self, 'accounts_receivable', accounts_receivable)

        accounts_payable= self.working_capital['accounts_payable']['value']
        setattr(self, 'accounts_payable', accounts_payable)
        cash_balance= self.working_capital['cash_balance']['value']
        setattr(self, 'cash_balance', cash_balance)
        
        #Base price of beef per ton
        base_price=self.prices['base_price']
        #dynamic var    
        setattr(self, 'base_price', base_price)
        #Change in real price of Beef
        change_in_price=self.prices['change_in_price']#['value'] 
        #dynamic var    
        setattr(self, 'change_in_price', change_in_price)

        #Cattle parameters
        num_feedlots=self.investment_parameter_options['initial_pens_employed'][0] 
        #print(f'num_feedlots: {num_feedlots}')
        setattr(self, 'initial_pens_employed', num_feedlots)
        fedlot_length_m= self.investment_parameter_options['pen_length'][0]
        setattr(self, 'pen_length', fedlot_length_m)
        fedlot_width_m=self.investment_parameter_options['pen_width'][0]
        setattr(self, 'pen_width', fedlot_width_m)     
        cattle_pen_cycle=self.investment_parameter_options['pen_cattle_density'][0]
        setattr(self, 'cattle_pen_cycle', cattle_pen_cycle)        
        months_per_cycle=self.feedlot_design_parameters['num_of_months_per_cycle']['value']
        setattr(self, 'months_per_cycle', months_per_cycle)
        
      
       
        cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
        setattr(self, 'cattle_per_pen_per_year', cattle_per_pen_per_year)
       

        cattle_survival_rate=self.cost_real['cattle_survival_rate']['value']
        setattr(self, 'cattle_survival_rate', cattle_survival_rate)
        
       
        
        cost_of_electricity_per_pen_per_annum=self.cost_real['cost_of_electricity_per_pen_per_annum']['value']
        setattr(self, 'cost_of_electricity_per_pen_per_annum', cost_of_electricity_per_pen_per_annum)

    
        #Economic parameter:
        real_interest_rate=self.financing['real_interest_rate']['value']
        setattr(self, 'real_interest_rate', real_interest_rate)
        risk_premium=self.financing['risk_premium']['value']
        setattr(self, 'risk_premium', risk_premium)
        num_of_installments=self.financing['num_of_installments']['value']
        setattr(self, 'num_of_installments', num_of_installments)

        #repayment_starts=self.financing['repayment_starts']['value']
        repayment_starts=self.timing_assumptions['base_period']['value'] + self.financing['grace_period']['value']
        setattr(self, 'repayment_starts', repayment_starts)
        
        grace_period=self.financing['grace_period']['value']
        setattr(self, 'grace_period', grace_period)
       
        #inputs costs
        annual_change_in_price_of_imported_inputs=self.cost_real['annual_change_in_price_of_imported_inputs']['value']
        setattr(self, 'annual_change_in_price_of_imported_inputs', annual_change_in_price_of_imported_inputs)

        annual_change_in_price_of_domestic_inputs=self.cost_real['annual_change_in_price_of_domestic_inputs']['value']
        setattr(self, 'annual_change_in_price_of_domestic_inputs', annual_change_in_price_of_domestic_inputs)

        cost_of_imported_inputs_per_ton_of_beef_produced=self.cost_real['cost_of_imported_inputs_per_ton_of_beef_produced']['value']
        setattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced', cost_of_imported_inputs_per_ton_of_beef_produced)
     
        # #---cattle inputs---------------------
        food_conversion_ratio=self.cost_real['food_conversion_ratio']['value']
        setattr(self, 'food_conversion_ratio', food_conversion_ratio)

        daily_weight_gain=self.cost_real['daily_weight_gain']['value']
        setattr(self, 'daily_weight_gain', daily_weight_gain)
        
        days_in_feed_lot=self.cost_real['days_in_feed_lot']['value']
        setattr(self, 'days_in_feed_lot', days_in_feed_lot)

        weight_at_purchase=self.cost_real['weight_at_purchase']['value']
        setattr(self, 'weight_at_purchase', weight_at_purchase)
        
        live_weight_gain=self.cost_real['live_weight_gain']['value']
        setattr(self, 'live_weight_gain', live_weight_gain)

        live_weight_at_selling= self.weight_at_purchase + self.live_weight_gain
        setattr(self, 'live_weight_at_selling', live_weight_at_selling)
        
        dressed_weight_percent=self.cost_real['dressed_weight_percent']['value']
        setattr(self, 'dressed_weight_percent', dressed_weight_percent) 
        
        
        dressed_weight_at_selling=self.dressed_weight_percent*self.live_weight_at_selling
        setattr(self, 'dressed_weight_at_selling', dressed_weight_at_selling)

        
        
        #qnty_of_feed_per_cattle=self.cost_real['qnty_of_feed_per_cattle']['value']
        qnty_of_feed_per_cattle=self.live_weight_gain*self.food_conversion_ratio
        setattr(self, 'qnty_of_feed_per_cattle', qnty_of_feed_per_cattle)
        
        commercial_feed_per_pen=self.cost_real['commercial_feed_per_pen']['value']
        setattr(self, 'commercial_feed_per_pen', commercial_feed_per_pen)

              
       
        cattle_price_per_unit=self.cost_real['cattle_price_per_unit']['value']
        setattr(self, 'cattle_price_per_unit', cattle_price_per_unit)

        cattle_feed_price_per_kg=self.cost_real['cattle_feed_price_per_kg']['value']
        setattr(self, 'cattle_feed_price_per_kg', cattle_feed_price_per_kg)
        
        cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs_per_ton_of_beef_as_raw_parameters()
        setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        other_indirect_costs=self.cost_real['other_indirect_costs']['value']
        setattr(self, 'other_indirect_costs', other_indirect_costs)
 

         #1.----------------Time Period----------------------
        base_period =self.timing_assumptions['base_period']['value']
        setattr(self, 'base_period', base_period)
        start_year =self.timing_assumptions['operation_start_year']['value']
        setattr(self, 'start_year', start_year)
        end_year =self.timing_assumptions['operation_end']['value']
        setattr(self, 'end_year', end_year)
        
        construction_year_end =self.timing_assumptions['construction_year_end']['value']
        setattr(self, 'construction_year_end', construction_year_end)
        
        #end period = ende_year+ 1(deconstruction)+ 1(pythonic)
        tp_list =[x for x in range(int(self.base_period),int(self.end_year)+1+1)]
        #was giving error
        setattr(self, 'tp_list', tp_list)
        #print('tp_list: len', len(tp_list))
        
         #2. -------------Price Index-----------------------------
    
        dividend_payout_ratio=self.macroeconomic_parameters['dividend_payout_ratio']['value']
        setattr(self, 'dividend_payout_ratio', dividend_payout_ratio)
        num_of_shares=self.macroeconomic_parameters['num_of_shares']['value']
        setattr(self, 'num_of_shares', num_of_shares)
        domestic_inflation_rate=self.macroeconomic_parameters['domestic_inflation_rate']['value']
        setattr(self, 'domestic_inflation_rate', domestic_inflation_rate)
        discount_rate_equity=self.macroeconomic_parameters['discount_rate_equity']['value']
        setattr(self, 'discount_rate_equity', discount_rate_equity)
        exchange_rate=self.macroeconomic_parameters['exchange_rate']['value']
        setattr(self, 'exchange_rate', exchange_rate)
        us_inflation_rate=self.macroeconomic_parameters['us_inflation_rate']['value']
        setattr(self, 'us_inflation_rate', us_inflation_rate)    
     

        senior_debt= self.financing['senior_debt']['value']
        setattr(self, 'senior_debt', senior_debt)

        equity=self.financing['equity']['value']
        setattr(self, 'equity', equity)

        sales_tax= self.taxes['sales_tax']['value']
        setattr(self, 'sales_tax', sales_tax)       
       

        corporate_income_tax= self.taxes['corporate_income_tax']['value']        
        setattr(self, 'corporate_income_tax', corporate_income_tax)

       
      
        #--2.a Dometstic price index
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', dpi_list)
        
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', uspi_list)
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', rpi_list)
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)
        #print('ner_list', ner_list)


        #Salaries & wages
        mwpw=self.cost_real['monthly_wage_per_worker']['value']
        setattr(self, 'monthly_wage_per_worker', mwpw)

        mwpst=self.cost_real['monthly_wage_per_supervisor']['value']
        setattr(self, 'monthly_wage_per_supervisor', mwpst)
        #Annual Increase in real salaries
        #Annual Increase Salaries  
        aisw=self.cost_real['annual_increase_salaries_workers']['value']
        setattr(self, 'annual_increase_salaries_workers', self.clean_inf_array(aisw))
        aispt=self.cost_real['annual_increase_salaries_supervisors_technicians']['value']
        setattr(self, 'annual_increase_salaries_supervisors_technicians', self.clean_inf_array(aispt))

        nowps= 5 # num of workers per supervisor
        setattr(self, 'num_workers_per_supervisor', nowps)

        nowpp=self.cost_real['num_of_workers_per_pen']['value']
        setattr(self, 'num_of_workers_per_pen', nowpp)

    def _get_qnty_of_feed_per_cattle(self):          
        total=self.food_conversion_ratio*self.daily_weight_gain*self.days_in_feed_lot
        return total
    def _get_cost_of_domestic_inputs_per_ton_of_beef(self):
          
        #1==A*B*C
        cost_of_cattle_feed_per_pen=self.cattle_per_pen_per_year*self.qnty_of_feed_per_cattle*self.cattle_feed_price_per_kg
             
        #1= A*D
        cost_of_input_cattle_per_pen =self.cattle_per_pen_per_year*self.cattle_price_per_unit

        #3 = A*B*C + A*D
        total_input_cost_per_pen_per_year =cost_of_cattle_feed_per_pen + cost_of_input_cattle_per_pen
        
        #4 = 1000*A*E*F----------------------
        tonnes_produced_per_pen_per_year =self.cattle_per_pen_per_year*self.dressed_weight_at_selling*self.cattle_survival_rate/1000
        #5- (A*B*C + A*D) / (A*E*F/1000)
        #=A(B*C+D)*1000/(A*E*F)
        # (B*C+D)*1000/(E*F)
        #(qnty_feed_kg_per_cattle*price_feed_per_kg + cattle_price)/ (1000*dressed_weight *survival_rate)
        # but===> qnty_feed =daily_weight_gain * days_in_feed_lot * food_conversion_ratio

        #
        input_cost_per_ton_of_beef=total_input_cost_per_pen_per_year/tonnes_produced_per_pen_per_year if  tonnes_produced_per_pen_per_year !=0 else 0
        return input_cost_per_ton_of_beef

    def _get_cost_of_domestic_inputs_per_ton_of_beef_as_raw_parameters(self):
        #(qnty_feed*price_feed + unit_cattle_cost)/ (dressed_weight *survival_rate)
        # but===> qnty_feed =daily_weight_gain * days_in_feed_lot * food_conversion_ratio
       
        qnty_feed=self.daily_weight_gain * self.days_in_feed_lot * self.food_conversion_ratio

        cost_numerator = qnty_feed * self.cattle_feed_price_per_kg + self.cattle_price_per_unit
        cost_denomenator= self.dressed_weight_at_selling * self.cattle_survival_rate
        total_cost = cost_numerator*1000/(cost_denomenator) if cost_denomenator !=0 else 0
        return total_cost





    def _metric_labourCost(self):
        """ 
        **********important****************** 
        should cone before unit cost of production
        """
        #6. LABOUR COSTS AND OTHER INDIRECT COSTS (Nominal)
        #Direct Labour
        #aaaaa
        #Salaries
        mwpw=self.monthly_wage_per_worker

        mwpst=self.monthly_wage_per_supervisor
        #Annual Increase in real salaries
        #Annual Increase Salaries  
        aisw=self.annual_increase_salaries_workers
        aispt=self.annual_increase_salaries_supervisors_technicians

        nowps= self.num_workers_per_supervisor

        #Number of workers per pen
       
        nowpp= self.num_of_workers_per_pen
        
        #Numner of workers = worker per per * pens per period
        now_list=[math.ceil(nowpp*x) for x in self.cum_feedlots_list]
        setattr(self, 'now_list', now_list)
        # Number of technicians and supervisors
        #self.cost_real['num_of_supervisors_technicians']['value']
        #Num of supervisors and techniciam
       
        nost_list=[math.floor(x/nowps) if nowps!=0 else 0 for x in now_list]
        setattr(self, 'nost_list', nost_list)

        #Real yearly wage rate        
        rywr_list= list(self._accumulate_salary_rate(self.tp_list,mwpw,self.clean_inf_array(aisw)))

        setattr(self, 'rywr_list', self.clean_inf_array(rywr_list))

        #Nominal yearly wage rate        
        nywr_list =[i*j for i,j in zip(self.dpi_list,rywr_list)]
        setattr(self, 'nywr_list', self.clean_inf_array(nywr_list))

        #Total Direct Labour Cost
        tdlc_list= [i*j for i,j in zip(now_list,nywr_list)]
        setattr(self, 'tdlc_list', self.clean_inf_array(tdlc_list))
        
        #Real yearly salary rate
        rysr_list= list(self._accumulate_salary_rate(self.tp_list,mwpst,aispt))
        setattr(self, 'rysr_list', self.clean_inf_array(rysr_list))
       
        #Nominal yearly salary rate        
        nysr_list =[i*j for i,j in zip(self.dpi_list,rysr_list)]
        setattr(self, 'nysr_list', self.clean_inf_array(nysr_list))
        
        #Total Indirect Labour Cost
        total_indirect_labour_list= [i*j for i,j in zip(nost_list,nysr_list)]
        setattr(self, 'total_indirect_labour_list', total_indirect_labour_list)

        #Total  Labour Cost
        tlc_list= [i+j for i,j in zip(tdlc_list,total_indirect_labour_list)]
        setattr(self, 'tlc_list', self.clean_inf_array(tlc_list))

        # Operation period		flag
        # Cost Of Electricity per pen per annum	 100.0 	LC
        coepppa= self.cost_of_electricity_per_pen_per_annum
        cost_of_electricity_list =[x*y*z*coepppa/1000 for x,y,z in zip(self.cum_feedlots_list,self.op_list,self.dpi_list)]
        setattr(self, 'cost_of_electricity_list', cost_of_electricity_list)
        # Cost of electricity (nominal)		000's LC
        # Other Indirect Costs	 1.0 	000's LC
        # Other Indirect Costs (nominal)		000's LC
        other_indirect_costs_list =[x*y*self.other_indirect_costs/1000 for x,y in zip(self.op_list,self.dpi_list)]
        setattr(self, 'other_indirect_costs_list', other_indirect_costs_list)
      
    def _metric_investmentCost(self):
        initial_pens_employed= self.investment_parameter_options['initial_pens_employed'][0]
        #dynamic var
        setattr(self, 'initial_pens_employed', initial_pens_employed)
        initial_pens_constructed_list=[self.initial_pens_employed if x==0 else 0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'initial_pens_constructed_list', initial_pens_constructed_list)
        #print('initial_pens_constructed_list', initial_pens_constructed_list)
        
        # Cost of unit pen constructed 
        cost_of_unit_pen= self.investment_cost['cost_of_pen_construction']['value']
        #dynamic var for sensitivity
        setattr(self, 'cost_of_unit_pen', cost_of_unit_pen)
        
        #1.2 Investment cost of Feedlots (nominal)
  
        # Cost of feedlots constructed       
        #Formulae = [No. of feedlot(pens) targeted for construction] * [Cost of Unit Pen Construction]/1000         
        cost_feedlots_nominal_list= [x*y*self.cost_of_unit_pen/1000  for x, y in zip(self.initial_pens_constructed_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_nominal_list', self.clean_inf_array(cost_feedlots_nominal_list))
     
        cost_feedlots_real_list= [0 if y==0 else  x/y  for x, y in zip(self.cost_feedlots_nominal_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_real_list', self.clean_inf_array(cost_feedlots_real_list))

        #1.3 Investment cost of land
              
     
        feedlot_sqm=self.initial_pens_employed*self.pen_length*self.pen_width 

        
        # sqm_per_cattle = feedlot_sqm/self.cattle_pen_cycle  if self.cattle_pen_cycle!=0 else 0
        # cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 

    

        land_for_pens= feedlot_sqm      
        #first option choosen by default
        option_land_requirement = self.investment_parameter_options['total_land_required'][0]
        setattr(self, 'option_land_requirement', option_land_requirement)

        land_required= max(land_for_pens,option_land_requirement)
        
        
        cost_of_land_per_sqm=self.investment_parameter_options['cost_of_land_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_land_per_sqm', cost_of_land_per_sqm)
         
        #Investment Roll out Flag
        # use initail pen under/ targeted for construction
        inv_roll= list(self._accumulate_investment_rollout_flag(self.initial_pens_constructed_list))
        setattr(self, 'invest_rollout_flag_list', inv_roll)
        

        cost_of_land_real_list= [x*land_required*self.cost_of_land_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_land_real_list', cost_of_land_real_list)
        
        cost_of_land_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_land_real_list)]
        setattr(self, 'cost_of_land_nominal_list', cost_of_land_nominal_list)
    
    
       #1.4 Investment cost of Machinery

        #Import duty
        import_duty=self.taxes['import_duty']['value']
        #dynamic variable
        setattr(self, 'import_duty', import_duty)



        inv_cost_over_run= self.investment_cost['investment_costs_over_run_factor']['value']
        setattr(self, 'inv_cost_over_run', inv_cost_over_run)
        

        #print('whatsinside?????', self.investment_parameter_options)
        cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
        cif_machinery_list= [x*self.cost_of_machinery_per_pen/1000  for x in self.initial_pens_constructed_list]
        setattr(self, 'cif_machinery_list', cif_machinery_list)      

        cif_machinery_with_overrun_list= [x*(1+self.inv_cost_over_run)  for x in self.cif_machinery_list]
        setattr(self, 'cif_machinery_with_overrun_list', cif_machinery_with_overrun_list)
       
        cif_machinery_with_overrun_lc_list= [x*y  for x,y  in zip(self.ner_list,self.cif_machinery_with_overrun_list)]
        setattr(self, 'cif_machinery_with_overrun_lc_list', cif_machinery_with_overrun_lc_list)
        
        #Total Cost of Machinery including import duty (real)
        total_machinery_incl_import_duty_real_list= [x*(1+self.import_duty)  for x  in cif_machinery_with_overrun_lc_list]
        setattr(self, 'total_machinery_incl_import_duty_real_list', total_machinery_incl_import_duty_real_list)
       
        total_machinery_incl_import_duty_nominal_list= [x*y  for x,y  in zip(self.dpi_list,total_machinery_incl_import_duty_real_list)]
        setattr(self, 'total_machinery_incl_import_duty_nominal_list', total_machinery_incl_import_duty_nominal_list)
       

        # Investment building
        cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        #cost of building as a function of pens constructed
        cost_of_building_real_list= [x*land_for_pens*self.cost_of_building_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_building_real_list', cost_of_building_real_list)

        cost_of_building_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_building_real_list)]
        setattr(self, 'cost_of_building_nominal_list', self.clean_inf_array(cost_of_building_nominal_list))
    
        #aaaaaaa
         
        # Total Investment Cost (nominal)		000's LC
        # Equity Contribution towards Total Investment Costs		000's LC
        # Senior Debt Contribution towards Total Investment Costs		000's LC
        total_invest_cost= np.zeros(len(self.tp_list))
       
        total_invest_cost= np.array(self.cost_feedlots_nominal_list) + \
                           np.array(self.cost_of_land_nominal_list)+ \
                          np.array(self.total_machinery_incl_import_duty_nominal_list) + \
                          np.array(self.cost_of_building_nominal_list)
        
       
       
      

        equity_contr_investments= self.equity*self.clean_inf_array(total_invest_cost)
        senior_debt_contr_investments= self.senior_debt*self.clean_inf_array(total_invest_cost)
        setattr(self, 'senior_debt_contr_investments', self.clean_inf_array(senior_debt_contr_investments))
    
    def _simulation_priceIndex(self):

        
        if not hasattr(self, 'accounts_receivable'):
            accounts_receivable= self.working_capital['accounts_receivable']['value']       
            setattr(self, 'accounts_receivable', accounts_receivable)

        if not hasattr(self, 'accounts_payable'):   
            accounts_payable= self.working_capital['accounts_payable']['value']
            setattr(self, 'accounts_payable', accounts_payable)
        
        if not hasattr(self, 'cash_balance'):
            cash_balance= self.working_capital['cash_balance']['value']
            setattr(self, 'cash_balance', cash_balance)
        
        if not hasattr(self, 'base_price'):    
            #Base price of beef per ton
            base_price=self.prices['base_price']
            #dynamic var    
            setattr(self, 'base_price', base_price)
        
        if not hasattr(self, 'change_in_price'):
            #Change in real price of Beef
            change_in_price=self.prices['change_in_price']#['value'] 
            #dynamic var    
            setattr(self, 'change_in_price', change_in_price)

        if not hasattr(self, 'initial_pens_employed'):    
            #Cattle parameters
            num_feedlots=self.investment_parameter_options['initial_pens_employed'][0] 
            setattr(self, 'initial_pens_employed', num_feedlots)
            #print(f'num_feedlots: {num_feedlots}')
        else:
            num_feedlots=self.investment_parameter_options['initial_pens_employed'][0] 
            #print(f'2..... num_feedlots: {num_feedlots}')   
        if not hasattr(self, 'pen_length'):    
            fedlot_length_m= self.investment_parameter_options['pen_length'][0]
            setattr(self, 'pen_length', fedlot_length_m)
            
        if not hasattr(self, 'pen_width'):    
            fedlot_width_m=self.investment_parameter_options['pen_width'][0]
            setattr(self, 'pen_width', fedlot_width_m)     
            
        if not hasattr(self, 'cattle_pen_cycle'):    
            cattle_pen_cycle=self.investment_parameter_options['pen_cattle_density'][0]
            setattr(self, 'cattle_pen_cycle', cattle_pen_cycle)        
        
        if not hasattr(self, 'months_per_cycle'):    
            months_per_cycle=self.feedlot_design_parameters['num_of_months_per_cycle']['value']
            setattr(self, 'months_per_cycle', months_per_cycle)
            
        if not hasattr(self, 'cattle_per_pen_per_year'):    
            cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
            setattr(self, 'cattle_per_pen_per_year', cattle_per_pen_per_year)
        
        if not hasattr(self, 'cattle_survival_rate'):
            cattle_survival_rate=self.cost_real['cattle_survival_rate']['value']
            setattr(self, 'cattle_survival_rate', cattle_survival_rate)
            
        if not hasattr(self, 'dressed_weight_at_selling'):
            dressed_weight_at_selling=self.cost_real['dressed_weight_at_selling']['value']
            setattr(self, 'dressed_weight_at_selling', dressed_weight_at_selling)

        if not hasattr(self, 'cost_of_electricity_per_pen_per_annum'):
            cost_of_electricity_per_pen_per_annum=self.cost_real['cost_of_electricity_per_pen_per_annum']['value']
            setattr(self, 'cost_of_electricity_per_pen_per_annum', cost_of_electricity_per_pen_per_annum)

        if not hasattr(self, 'real_interest_rate'):
            #Economic parameter:
            real_interest_rate=self.financing['real_interest_rate']['value']
            setattr(self, 'real_interest_rate', real_interest_rate)
            
        if not hasattr(self, 'risk_premium'):
            risk_premium=self.financing['risk_premium']['value']
            setattr(self, 'risk_premium', risk_premium)

        if not hasattr(self, 'num_of_installments'):    
            num_of_installments=self.financing['num_of_installments']['value']
            setattr(self, 'num_of_installments', num_of_installments)

        if not hasattr(self, 'repayment_starts'):   
            #repayment_starts=self.financing['repayment_starts']['value']
            repayment_starts=self.timing_assumptions['base_period']['value'] + self.financing['grace_period']['value']
    
            setattr(self, 'repayment_starts', repayment_starts)
            
        if not hasattr(self, 'grace_period'):   
            grace_period=self.financing['grace_period']['value']
            setattr(self, 'grace_period', grace_period)
        
        if not hasattr(self, 'annual_change_in_price_of_imported_inputs'):    
            #inputs costs
            annual_change_in_price_of_imported_inputs=self.cost_real['annual_change_in_price_of_imported_inputs']['value']
            setattr(self, 'annual_change_in_price_of_imported_inputs', annual_change_in_price_of_imported_inputs)

        if not hasattr(self, 'annual_change_in_price_of_domestic_inputs'):    
            annual_change_in_price_of_domestic_inputs=self.cost_real['annual_change_in_price_of_domestic_inputs']['value']
            setattr(self, 'annual_change_in_price_of_domestic_inputs', annual_change_in_price_of_domestic_inputs)

        if not hasattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced'):    
            cost_of_imported_inputs_per_ton_of_beef_produced=self.cost_real['cost_of_imported_inputs_per_ton_of_beef_produced']['value']
            setattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced', cost_of_imported_inputs_per_ton_of_beef_produced)

        if not hasattr(self, 'food_conversion_ratio'):
            #---cattle inputs---------------------
            food_conversion_ratio=self.cost_real['food_conversion_ratio']['value']
            setattr(self, 'food_conversion_ratio', food_conversion_ratio)

        if not hasattr(self, 'daily_weight_gain'):    
            daily_weight_gain=self.cost_real['daily_weight_gain']['value']
            setattr(self, 'daily_weight_gain', daily_weight_gain)
            
        if not hasattr(self, 'days_in_feed_lot'):    
            days_in_feed_lot=self.cost_real['days_in_feed_lot']['value']
            setattr(self, 'days_in_feed_lot', days_in_feed_lot)

        if not hasattr(self, 'weight_at_purchase'):
            weight_at_purchase=self.cost_real['weight_at_purchase']['value']
            setattr(self, 'weight_at_purchase', weight_at_purchase)
        
        if not hasattr(self, 'live_weight_gain'):
            live_weight_gain=self.daily_weight_gain*self.days_in_feed_lot
            setattr(self, 'live_weight_gain', live_weight_gain)

        if not hasattr(self, 'qnty_of_feed_per_cattle'):    
            qnty_of_feed_per_cattle=self.live_weight_gain*self.food_conversion_ratio#self._get_qnty_of_feed_per_cattle()
            setattr(self, 'qnty_of_feed_per_cattle', qnty_of_feed_per_cattle)
            
        if not hasattr(self, 'commercial_feed_per_pen'):    
            commercial_feed_per_pen=self.cost_real['commercial_feed_per_pen']['value']
            setattr(self, 'commercial_feed_per_pen', commercial_feed_per_pen)

        if not hasattr(self, 'dressed_weight_percent'):    
            dressed_weight_percent=self.cost_real['dressed_weight_percent']['value']
            setattr(self, 'dressed_weight_percent', dressed_weight_percent) 

        if not hasattr(self, 'cattle_price_per_unit'):
            cattle_price_per_unit=self.cost_real['cattle_price_per_unit']['value']
            setattr(self, 'cattle_price_per_unit', cattle_price_per_unit)
        
        if not hasattr(self, 'cattle_feed_price_per_kg'):
            cattle_feed_price_per_kg=self.cost_real['cattle_feed_price_per_kg']['value']
            setattr(self, 'cattle_feed_price_per_kg', cattle_feed_price_per_kg)
            
        #if not hasattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced'):    
        cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs_per_ton_of_beef_as_raw_parameters()
        setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        if not hasattr(self, 'other_indirect_costs'):    
            other_indirect_costs=self.cost_real['other_indirect_costs']['value']
            setattr(self, 'other_indirect_costs', other_indirect_costs)
    

         #1.----------------Time Period----------------------
        if not hasattr(self, 'base_period'):
            base_period =self.timing_assumptions['base_period']['value']    
            setattr(self, 'base_period', base_period)

        if not hasattr(self, 'start_year'):
            start_year =self.timing_assumptions['operation_start_year']['value']
            setattr(self, 'start_year', start_year)
        
        if not hasattr(self, 'end_year'):      
            end_year =self.timing_assumptions['operation_end']['value']
            setattr(self, 'end_year', end_year)
        
        if not hasattr(self, 'construction_year_end'):           
            construction_year_end =self.timing_assumptions['construction_year_end']['value']
            setattr(self, 'construction_year_end', construction_year_end)
        
        

        if not hasattr(self, 'tp_list'):
            tp_list =[x for x in range(int(self.base_period),int(self.end_year)+1)]       
            setattr(self, 'tp_list', tp_list)

       
            #Economic Parameter
        if not hasattr(self, 'dividend_payout_ratio'):    
            dividend_payout_ratio=self.macroeconomic_parameters['dividend_payout_ratio']['value']
            setattr(self, 'dividend_payout_ratio', dividend_payout_ratio)
            
        if not hasattr(self, 'num_of_shares'):    
            num_of_shares=self.macroeconomic_parameters['num_of_shares']['value']
            setattr(self, 'num_of_shares', num_of_shares)
        
        if not hasattr(self, 'domestic_inflation_rate'):    
            domestic_inflation_rate=self.macroeconomic_parameters['domestic_inflation_rate']['value']
            domestic_inflation_rate=round(domestic_inflation_rate,5)
            setattr(self, 'domestic_inflation_rate', domestic_inflation_rate)
            
        if not hasattr(self, 'discount_rate_equity'):    
            discount_rate_equity=self.macroeconomic_parameters['discount_rate_equity']['value']
            setattr(self, 'discount_rate_equity', discount_rate_equity)
            
        if not hasattr(self, 'exchange_rate'):    
            exchange_rate=self.macroeconomic_parameters['exchange_rate']['value']
            setattr(self, 'exchange_rate', exchange_rate)
           
        if not hasattr(self, 'us_inflation_rate'):    
            us_inflation_rate=self.macroeconomic_parameters['us_inflation_rate']['value']
            setattr(self, 'us_inflation_rate', us_inflation_rate)    
     
        if not hasattr(self, 'senior_debt'):
            senior_debt= self.financing['senior_debt']['value']
            setattr(self, 'senior_debt', senior_debt)

        if not hasattr(self, 'equity'):    
            equity=self.financing['equity']['value']
            setattr(self, 'equity', equity)

        if not hasattr(self, 'sales_tax'):    
            sales_tax= self.taxes['sales_tax']['value']
            setattr(self, 'sales_tax', sales_tax)       
        
        if not hasattr(self, 'corporate_income_tax'):
            corporate_income_tax= self.taxes['corporate_income_tax']['value']        
            setattr(self, 'corporate_income_tax', corporate_income_tax)

       
         #Salaries & wages
        if not hasattr(self, 'monthly_wage_per_worker'): 
            mwpw=self.cost_real['monthly_wage_per_worker']['value']
            setattr(self, 'monthly_wage_per_worker', mwpw)
        
        if not hasattr(self, 'monthly_wage_per_supervisor'):
            mwpst=self.cost_real['monthly_wage_per_supervisor']['value']
            setattr(self, 'monthly_wage_per_supervisor', mwpst)
        
        if not hasattr(self, 'annual_increase_salaries_workers'): 
            aisw=self.cost_real['annual_increase_salaries_workers']['value']
            setattr(self, 'annual_increase_salaries_workers', aisw)
        
        if not hasattr(self, 'annual_increase_salaries_supervisors_technicians'):
            aispt=self.cost_real['annual_increase_salaries_supervisors_technicians']['value']
            setattr(self, 'annual_increase_salaries_supervisors_technicians', aispt)

        if not hasattr(self, 'num_workers_per_supervisor'):
            nowps= 5 # num of workers per supervisor
            setattr(self, 'num_workers_per_supervisor', nowps)
        
        if not hasattr(self, 'num_of_workers_per_pen'):
            nowpp=self.cost_real['num_of_workers_per_pen']['value']
            setattr(self, 'num_of_workers_per_pen', nowpp)
        #--2.a Dometstic price index
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', self.clean_inf_array(dpi_list))
        #print('dpi_list',dpi_list )
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', self.clean_inf_array(uspi_list))
        # print('us_inflation_rate',us_inflation_rate )
        # print('uspi_list',uspi_list )
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', rpi_list)
        #print('rpi_list',rpi_list )
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)
       # print('ner_list', ner_list)           



        
       
        #1.----------------Time Period----------------------          
        cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs_per_ton_of_beef_as_raw_parameters()
        setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        #--2.a Dometstic price index [possibility of dynamic inflation]
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', self.clean_inf_array(dpi_list))
        
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', self.clean_inf_array(uspi_list))
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', self.clean_inf_array(rpi_list))
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)
    
    def _simulation_investmentCost(self):

        
        # if not hasattr(self, 'initial_pens_employed'):
        #     initial_pens_employed= self.investment_parameter_options['initial_pens_employed'][0]
        #     #dynamic var
        #     setattr(self, 'initial_pens_employed', initial_pens_employed)
       
        if not hasattr(self, 'cost_of_unit_pen'):
            # Cost of unit pen constructed 
            cost_of_unit_pen= self.investment_cost['cost_of_pen_construction']['value']
            setattr(self, 'cost_of_unit_pen', cost_of_unit_pen)
        
           
        if not hasattr(self, 'option_land_requirement'):
            option_land_requirement = self.investment_parameter_options['total_land_required'][0]
            setattr(self, 'option_land_requirement', option_land_requirement)

        
        if not hasattr(self, 'cost_of_land_per_sqm'):
            cost_of_land_per_sqm=self.investment_parameter_options['cost_of_land_per_sqm'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_land_per_sqm', cost_of_land_per_sqm)
         
        if not hasattr(self, 'import_duty'):
            #Import duty
            import_duty=self.taxes['import_duty']['value']
            #dynamic variable
            setattr(self, 'import_duty', import_duty)
            #print('import_duty',import_duty)


        if not hasattr(self, 'inv_cost_over_run'):
            inv_cost_over_run= self.investment_cost['investment_costs_over_run_factor']['value']
            setattr(self, 'inv_cost_over_run', inv_cost_over_run)
        

        if not hasattr(self, 'cost_of_machinery_per_pen'):
            cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
     
        if not hasattr(self, 'cost_of_building_per_sqm'):
            # Investment building
            cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        
        
        
        
        
        
        
        
        
        
        
        #------------------Up is new edit pliz-------------------------------
        initial_pens_constructed_list=[self.initial_pens_employed if x==0 else 0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'initial_pens_constructed_list', initial_pens_constructed_list)
        #print('initial_pens_constructed_list', initial_pens_constructed_list)  
        #1.2 Investment cost of Feedlots (nominal)
  
        # Cost of feedlots constructed       
        #Formulae = [No. of feedlot(pens) targeted for construction] * [Cost of Unit Pen Construction]/1000         
        cost_feedlots_nominal_list= [x*y*self.cost_of_unit_pen/1000  for x, y in zip(self.initial_pens_constructed_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_nominal_list', cost_feedlots_nominal_list)

        cost_feedlots_real_list= [0 if y==0 else  x/y for x, y in zip(self.cost_feedlots_nominal_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_real_list', cost_feedlots_real_list)
     
       
        #1.3 Investment cost of land        
         
        feedlot_sqm=self.initial_pens_employed*self.pen_length*self.pen_width 

        
        sqm_per_cattle = feedlot_sqm/self.cattle_pen_cycle  if self.cattle_pen_cycle!=0 else 0
        cattle_pen_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
        
        land_for_pens= feedlot_sqm

        land_required= max(land_for_pens,self.option_land_requirement)


          
        #Investment Roll out Flag
        # use initail pen under/ targeted for construction
        inv_roll= list(self._accumulate_investment_rollout_flag(self.initial_pens_constructed_list))
        setattr(self, 'invest_rollout_flag_list', inv_roll)
        

        cost_of_land_real_list= [x*land_required*self.cost_of_land_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_land_real_list', cost_of_land_real_list)
        
        cost_of_land_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_land_real_list)]
        setattr(self, 'cost_of_land_nominal_list', cost_of_land_nominal_list)
    
    
       #1.4 Investment cost of Machinery     
      
        cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
        cif_machinery_list= [x*self.cost_of_machinery_per_pen/1000  for x in self.initial_pens_constructed_list]
        setattr(self, 'cif_machinery_list', cif_machinery_list)    
        #print('Inspect?   cif_machinery_list',cif_machinery_list)  

        cif_machinery_with_overrun_list= [x*(1+self.inv_cost_over_run)  for x in self.cif_machinery_list]
        setattr(self, 'cif_machinery_with_overrun_list', cif_machinery_with_overrun_list)
       
        cif_machinery_with_overrun_lc_list= [x*y  for x,y  in zip(self.ner_list,self.cif_machinery_with_overrun_list)]
        setattr(self, 'cif_machinery_with_overrun_lc_list', cif_machinery_with_overrun_lc_list)
        
        #Total Cost of Machinery including import duty (real)
        total_machinery_incl_import_duty_real_list= [x*(1+self.import_duty)  for x  in cif_machinery_with_overrun_lc_list]
        setattr(self, 'total_machinery_incl_import_duty_real_list', total_machinery_incl_import_duty_real_list)
       
        total_machinery_incl_import_duty_nominal_list= [x*y  for x,y  in zip(self.dpi_list,total_machinery_incl_import_duty_real_list)]
        setattr(self, 'total_machinery_incl_import_duty_nominal_list', total_machinery_incl_import_duty_nominal_list)
       

        # Investment  of buildings
        cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        #cost of building as a function of pens constructed
        cost_of_building_real_list= [x*land_for_pens*self.cost_of_building_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_building_real_list', cost_of_building_real_list)

        cost_of_building_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_building_real_list)]
        setattr(self, 'cost_of_building_nominal_list', cost_of_building_nominal_list)
    
        #aaaaaaa
         
        # Total Investment Cost (nominal)		000's LC
        # Equity Contribution towards Total Investment Costs		000's LC
        # Senior Debt Contribution towards Total Investment Costs		000's LC
        total_invest_cost= np.zeros(len(self.tp_list))
       
        total_invest_cost= np.array(self.cost_feedlots_nominal_list) + \
                           np.array(self.cost_of_land_nominal_list)+ \
                          np.array(self.total_machinery_incl_import_duty_nominal_list) + \
                          np.array(self.cost_of_building_nominal_list)
        
        setattr(self, 'total_invest_cost', self.clean_inf_array(total_invest_cost))
    
        # senior_debt=self.financing['senior_debt']['value']
        # equity=self.financing['equity']['value']
     

        equity_contr_investments= self.equity*self.clean_inf_array(total_invest_cost)
        senior_debt_contr_investments= self.senior_debt*self.clean_inf_array(total_invest_cost)
        setattr(self, 'senior_debt_contr_investments', self.clean_inf_array(senior_debt_contr_investments))
    
        # there is no update of loans
        #print('inspect? senior_debt_contr_investments',senior_debt_contr_investments)

        self.investment_cost['investment_cost_of_buildings']['value']
        self.investment_cost['cif_cost_of_machinery']['value']
        self.investment_cost['investment_cost_of_land']['value'] 
        self.investment_parameter_options['cost_of_pen_construction'][0]
   
    def _metric_productionAndSales(self) :
        #4 Gross Sales Revenue**************************************************
        #4.a Productin Quantity--------------------------------------------
        
     
        #4.a.2 Cumulative pens under Cattle
        #cum_feedlots_list= list(self._accumulate_addition(self.initial_pens_constructed_list))
        cum_feedlots_list= list(self._accumulate_harvesting(self.initial_pens_constructed_list, self.op_list))
        #print('cum_feedlots_list' , cum_feedlots_list)
        

        setattr(self, 'cum_feedlots_list', cum_feedlots_list)
        #print('cum_feedlots_list', cum_feedlots_list)
        
        #4.a.3 Closing Inventory
        ci_list=[0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'ci_list', ci_list)

       
        # Production Quantity
        yearly_tons_production_per_pen_const= self.cattle_per_pen_per_year*self.cattle_survival_rate*self.dressed_weight_at_selling/1000
        #set dynamic variable
        # print('cattle_per_pen_per_year', self.cattle_per_pen_per_year)
        # print('cattle_survival_rate', self.cattle_survival_rate)
        # print('dressed_weight_at_selling', self.dressed_weight_at_selling)
        # print('yearly_tons_production_per_pen_const', yearly_tons_production_per_pen_const)
       
        prod_qnty_list=[yearly_tons_production_per_pen_const*x*y  for x,y in zip(self.op_list,cum_feedlots_list)]#dynamic later
        setattr(self, 'prod_qnty_list', self.clean_inf_array(prod_qnty_list))
        #print('prod_qnty_list', prod_qnty_list)

        #4.b Sales Revenue

        #4.b. Sales Qnty  
        sales_qnty_list= list(self._accumulate_sales_qnty(self.ci_list, self.prod_qnty_list))
        setattr(self, 'sales_qnty_list', self.clean_inf_array(sales_qnty_list))
        # print('sales_qnty_list', sales_qnty_list)
       
        
        #Real Price of Beef per ton        
        rpobpt_list=[float(1+self.change_in_price) * float(self.base_price) for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'rpobpt_list', self.clean_inf_array(rpobpt_list))
        
        
        #Nominal Price of Beef per ton
        npbpt_list=[round(x*y,5)  for x,y in zip(self.dpi_list,rpobpt_list)]
        setattr(self, 'npbpt_list', self.clean_inf_array(npbpt_list))
        

        #4.d. Net Sales Revenue 
        net_sales_revenue_list=[x*y/1000  for x,y in zip(sales_qnty_list,npbpt_list)]
        setattr(self, 'net_sales_revenue_list', self.clean_inf_array(net_sales_revenue_list))

        #4.e. Gross sales revenue  
        #Sales tax
       
        #sale tax paid
        st_list=[self.sales_tax * x for x in self.net_sales_revenue_list]
        setattr(self, 'st_list', self.clean_inf_array(st_list))

        
        # Gross sales revenue 
        gsr_list=[x+ y  for x,y in zip(self.net_sales_revenue_list,st_list)]
        setattr(self, 'gsr_list', self.clean_inf_array(gsr_list))


    def _metric_unitCost(self):
        # Production Quantity	
        # Total Direct Labour Cost	
        # Direct Labour cost per ton of beef produced
        direct_labour_per_ton_beef_produced=[0 if y==0 else round((x*1000/y),5)  for x,y in zip(self.tdlc_list,self.prod_qnty_list)]
        setattr(self, 'direct_labour_per_ton_beef_produced', self.clean_inf_array(direct_labour_per_ton_beef_produced))
            
        # Operation period	
        # Total cost of inputs per ton (nominal)	
        # Total unit cost of production per ton (nominal)
        total_unit_cost_production_per_ton_list=[(x+y)*z for x,y,z in zip(self.direct_labour_per_ton_beef_produced,
                                                      self.inputs_cost_per_ton_nominal_list,
                                                      self.op_list)]
        setattr(self, 'total_unit_cost_production_per_ton_list', self.clean_inf_array(total_unit_cost_production_per_ton_list))
        	
    def _cal_metrix(self):
        #1 Price Index
        self._metric_priceIndex()
        
        #2. Investments
        self._metric_investmentCost()

        #3. Residual Values
        self._metric_residualValues()
        
        #4. Production & Sales
        self._metric_productionAndSales()
        
        #6 Purchases
        #---unit_cost_used?????
        self._metric_purchases()
        
        #7. Working Capital
        self._metric_workingCapital()

        #5 Labour Costs
        self._metric_labourCost()
        
        
        #10 Unit Production Cost
        self._metric_unitCost()
        
       
        
        #8. LOAN SCHEDULE (Nominal)
        self._metric_loanSchedule()
        
        
        
        #9 Depreciation
        self._metric_depreciationForTaxPurposes()     
       

        #11
        self._metric_finishedProdictEvaluationFIFO()
        
       
        #12 INCOME TAX STATEMENT
        self._metric_IncomeTaxStatement()
        #13 Cash Flow
        self._metric_cashFlow()

    def _simulate_metrix(self):
        #1 Price Index
        #unit cost calculated here????
        self._simulation_priceIndex()
        
        #2. Investments
        self._simulation_investmentCost()
        
         #3. Residual Values
        self._metric_residualValues()
      
        #4. Production & Sales
        self._metric_productionAndSales()
        
       
       
        #6 Purchases
        #then used here????
        self._metric_purchases()
        
        #7. Working Capital
        self._metric_workingCapital()
        
        #5 Labour Costs
        self._metric_labourCost()
        
        #10 Unit Production Cost        
        self._metric_unitCost() 
       
        #8. LOAN SCHEDULE (Nominal)
        self._metric_loanSchedule()
         
        #9 Depreciation
        self._metric_depreciationForTaxPurposes()
        
       

        #11
        self._metric_finishedProdictEvaluationFIFO()
        
        
        #12 INCOME TAX STATEMENT
        self._metric_IncomeTaxStatement()
        
        #13 Cash Flow
        self._metric_cashFlow()
  
    def  _generate_data(self,output,request):
        # Create a workbook
        wb = Workbook()
        self.myworboook = wb
        #initialise styles
        self.excel_obj._initialise_named_styles(wb)
        
        self.excel_obj._write_input_sheet(wb,)
        #correct formating here---------
        self.excel_obj._write_output_sheet(wb, request)
        self.excel_obj._write_sens_sheet(wb, )
       
        self.excel_obj._write_calculation_sheet(wb, )
        self.excel_obj._write_cashflow_sheet(wb, )
        
        #update
        for item in self.excel_obj._production_inventory.keys():
            self.excel_obj._populate_productionInventory_section(wb["Inputs"], item)
        #---------------------------------------------------------------

        for item in self.investment_cost.keys():
            self.excel_obj._populate_investmentcost_section(wb["Inputs"], item)
        
	
        for item in self.cost_real.keys():
            self.excel_obj._populate_costs_section(wb["Inputs"], item)
        
        for item in ['initial_pens_employed', 'senior_debt', 'total_cattle_per_pen_per_cycle']:
            self.excel_obj._update_sens_section(wb["Sens"], item)
        
        #-------------
        self.excel_obj._update_investment_parameters(wb["Inputs"])

        #
        self.excel_obj._link_feedlot_design_parameters_to_model_selected(wb["Inputs"])
        self.excel_obj._update_sensitivity(wb["Sens"],self)
        #updated here again.... 02 Feb 2022
        #paid salary 110K
        # iam sure it will work
        # i love Programming
        # I love God
        # I love JESUS
        self.excel_obj._populate_output_sheet(wb["Outputs"])
        
        self.excel_obj._write_analytics_sheet(wb)
        # Write a formatted header row
        header = []
        comment = None

       
        # Write the spreadsheet
        wb.save(output)

               
    def _sens_sensitivity_parrallel_generator(self):
        input_list=[] 
        sens_input_dict={}
           
        model_clone = copy.deepcopy(self)
      
        params_list =[
                'domestic_inflation_rate',
                'us_inflation_rate',
                'change_in_price', 
                'accounts_receivable',
                'accounts_payable',
                'cash_balance',
                'exchange_rate', 
                'senior_debt',
                'base_price',
                'cattle_survival_rate',# fish_survival_rate
                'cattle_feed_price_per_kg',

                'cattle_price_per_unit',
                'initial_pens_employed',# initial tankers
                'discount_rate_equity',
                'dressed_weight_at_selling',# for fish
                'other_indirect_costs',
                'annual_change_in_price_of_domestic_inputs',
                'annual_change_in_price_of_imported_inputs',
                'total_cattle_per_pen_per_cycle',#verify why there is no attr

                'monthly_wage_per_worker',
                'monthly_wage_per_supervisor',
                'annual_increase_salaries_workers',  
                'annual_increase_salaries_supervisors_technicians',
                'num_workers_per_supervisor',
                'inv_cost_over_run'
                
            ]




        for var_x in params_list:      
        
            list_ =[]
            index = 0
            base_point=0
            if var_x  in ['cattle_survival_rate']:
                list_ = [1 ,0.99, 0.98, .95, .90, 0.85, 0.80, 0.70, .60, .50]
                base_point =model_clone.cattle_survival_rate 
                if not (model_clone.cattle_survival_rate in list_):
                    list_.append(base_point)
            elif var_x  in ['senior_debt']:
                list_ = [1 ,.95, 0.90, .90, .70, .60, .50, 0.40, .30, .20]
                base_point =model_clone.senior_debt  
                if not (model_clone.senior_debt in list_):
                    list_.append(base_point)
            elif var_x  in ['initial_pens_employed']:
                list_ = [1000 ,500, 200, 100, 50, 20, 10, 5, 1]
                base_point =model_clone.initial_pens_employed 
                if not (model_clone.initial_pens_employed in list_):
                    list_.append(base_point)
            elif var_x  in ['total_cattle_per_pen_per_cycle']:
                list_ = [100 ,50, 20, 10, 5]
                base_point =model_clone.cattle_pen_cycle
                if not (model_clone.cattle_pen_cycle in list_):
                    list_.append(base_point) 
                              
            elif var_x  in ['discount_rate_equity']:
                list_ = [.05,.1,.12,.15,.20,.3]
                base_point =model_clone.discount_rate_equity 
                if not (model_clone.discount_rate_equity in list_):
                    list_.append(base_point)   
            elif var_x  in ['dressed_weight_at_selling']:
                list_ = [90,120, 150,200, 300]
                base_point =model_clone.dressed_weight_at_selling
                if not (model_clone.dressed_weight_at_selling in list_):
                    list_.append(base_point)  
            elif var_x  in ['other_indirect_costs']:
                list_ = [100,1000,5000]
                base_point =model_clone.other_indirect_costs 
                if not (model_clone.other_indirect_costs in list_):
                    list_.append(base_point)     
            elif var_x  in ['inv_cost_over_run']:
                list_ = [.05,.1,.2,.4,.6, .9,1.]
                base_point =model_clone.inv_cost_over_run 
                if not (model_clone.inv_cost_over_run in list_):
                    list_.append(base_point)               
            else:
                if hasattr(self, var_x ):
                    base_point= getattr(self, var_x)

                if base_point==0:
                    list_ = [0.02 * i  for i in range(0,10)]
                else:
                    list_ = [0.2 * i * base_point  for i in range(1,10)]            
            
            list_.sort()
            input_list.append({'name':var_x, 'list':self.clean_inf_array(list_)})
           
            sens_input_dict[var_x]= {'list':self.clean_inf_array(list_),'base_val': self.clean_inf_array(base_point)}

        setattr(self, 'sens_input_dict', sens_input_dict) 
        #do once
        #before adding boundaries
        #get all parameters....
        dict_of_para= {}
        for i in params_list:
           if hasattr(self, i ):
                dict_of_para[i]=getattr(self,i) 
        """  
        print('....................  input_list  ........................')
        print(input_list)
        print('....................  dict_of_para  ........................')
        print(dict_of_para) 
        
        """
        sens_dict = get_data_table_sensitivity_in_parrallel(model_clone,input_list,dict_of_para)
        
        """ 
        print('....................  sens_dict  ........................')
        print(sens_dict)
        
        """
        setattr(self, 'sens_dict', sens_dict) 
        self.write_sens_data()  
        sens_grad =get_sensitivity_gradient(model_clone,input_list)
        sens_grad.fillna(0,inplace=True)
        sens_grad.sort_values(by=['abs_grad'], ascending=False) 
        print(sens_grad)
        setattr(self, 'sens_grad', sens_grad)
        mask_is_sensitivity=sens_grad['abs_grad'] >float(0.001)
        #list only which is sensitive
        para_list_by_grad= sens_grad[mask_is_sensitivity]['parameter'].tolist() 
        setattr(self, 'para_list_by_grad', para_list_by_grad)
        
        return sens_dict
 
    def spreadsheet(self, request) :
        
        if not hasattr(self, 'para_list_by_grad'):
            self._sens_sensitivity_parrallel_generator()
        # Return an excel spreadsheet
        output = BytesIO()
        self._generate_data(output,request)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=output.getvalue(),
        )
       
       
        wb = load_workbook(filename=BytesIO(output.getvalue()))
        #uselles for printing only...
        self.excel_obj._get_out_values_as_dict(wb)

        #-----------title name------------------
        date_now = datetime.datetime.now() #date.today()
        date_now = str(date_now.year) + '_' + str(date_now.month) \
                    + '_' + str(date_now.day)+ "_" + str(date_now.hour) \
                    + '_' + str(date_now.minute) + "_" +  str(date_now.second)
        title = "clearVision_solutions_auto-generated beefbusiness" + str(date_now)
        #-----------------------------

      
        response[
            "Content-Disposition"
        ] = "attachment; filename*=utf-8''%s.xlsx" % force_str(title) 
        response["Cache-Control"] = "no-cache, no-store"
        return response
      
    def _metric_residualValues(self):
        #Initial total values
        #1 Total cost of land (real)
        total_cost_land_real =np.sum(self.cost_of_land_real_list)         
        setattr(self, 'total_cost_land_real', total_cost_land_real)

        #2 Machinery (real)
        #---fuurther research if necessary to combine cost of feedlots into machinery
        #in calclaying thr residual value
        total_cost_machinery_real =np.sum(self.total_machinery_incl_import_duty_real_list) + \
                 np.sum(self.cost_feedlots_real_list)        
        setattr(self, 'total_cost_machinery_real', total_cost_machinery_real)

        #3 Building (real)
        total_cost_buidings_real =np.sum(self.cost_of_building_real_list)
        setattr(self, 'total_cost_buidings_real', total_cost_buidings_real)

        # Economic service life					
        # 			Economic life of machinery	 30 	Years
        # 			Economic life of buildings	 30 	Years
        economic_life_of_machinery= self.depreciation['economic_life_of_machinery']['value']
        setattr(self, 'economic_life_of_machinery', economic_life_of_machinery)
        economic_life_of_buildings= self.depreciation['economic_life_of_buildings']['value']
        setattr(self, 'economic_life_of_buildings', economic_life_of_buildings)
        tax_life_of_machinery=self.depreciation['tax_life_of_machinery']['value']
        setattr(self, 'tax_life_of_machinery', tax_life_of_machinery)
        tax_life_of_buildings=self.depreciation['tax_life_of_buildings']['value']
        setattr(self, 'tax_life_of_buildings', tax_life_of_buildings)
        tax_life_of_soft_capital_costs=self.depreciation['tax_life_of_soft_capital_costs']['value']
        setattr(self, 'tax_life_of_soft_capital_costs', tax_life_of_soft_capital_costs)



        # Depreciation (Real)			
		# 	Operation period

        #Formale=Total*Operation period/Economic Life
		# 	Machinery
        machinery_const=self.total_cost_machinery_real/self.economic_life_of_machinery if self.economic_life_of_machinery!=0 else 0
        depreciation_machinery_list=[machinery_const*x  for x in self.op_list]
        setattr(self, 'depreciation_machinery_list', depreciation_machinery_list)

        
		# 	Buildings
        buildings_const=self.total_cost_buidings_real/self.economic_life_of_buildings if self.economic_life_of_buildings!=0 else 0
        depreciation_buildings_list=[buildings_const*x  for x in self.op_list]
        setattr(self, 'depreciation_buildings_list', depreciation_buildings_list)

        # Residual value (Real)
        # #-----------------------------------------------------------------------------			
		# 	Residual value of land (real)
        residual_val_land_real_list=[0 for x in self.tp_list]
        residual_val_land_real_list[-1]=self.total_cost_land_real
        setattr(self, 'residual_val_land_real_list', residual_val_land_real_list)
        
        # 	Residual value of Machinery (real)
        residual_val_machinery_real_list=[0 for x in self.tp_list]
        residual_val_machinery_real_list[-1]=self.total_cost_machinery_real
        setattr(self, 'residual_val_machinery_real_list', residual_val_machinery_real_list)
        
        # 	Residual value of Buildings (real)
        residual_val_buildings_real_list=[0 for x in self.tp_list]
        residual_val_buildings_real_list[-1]=self.total_cost_buidings_real
        setattr(self, 'residual_val_buildings_real_list', residual_val_buildings_real_list)
		
         # Residual value (Nominal)	
         # #--------------------------------------------------------------------------------------------		
		
        # 	Residual value of land (Nominal)
        residual_val_land_nominal_list=[x*y for x,y in zip(self.dpi_list,self.residual_val_land_real_list)]
        setattr(self, 'residual_val_land_nominal_list', residual_val_land_nominal_list)
        
        # 	Residual value of Machinery (Nominal)
        residual_val_machinery_nominal_list=[x*y for x,y in zip(self.dpi_list, self.residual_val_machinery_real_list)]
        setattr(self, 'residual_val_machinery_nominal_list', residual_val_machinery_nominal_list)
        
        # 	Residual value of Buildings (Nominal)
        residual_val_buildings_nominal_list=[x*y for x,y in zip(self.dpi_list,self.residual_val_buildings_real_list)]
        setattr(self, 'residual_val_buildings_nominal_list', residual_val_buildings_nominal_list)
    
       
    def _metric_purchases(self):
        #7. PURCHASES (Nominal)
        
             
        #Real CIF Cost of imported inputs per ton
        rcifcoiipt_list= list(self._accumulate_yearly_rate(self.tp_list,
                                self.cost_of_imported_inputs_per_ton_of_beef_produced,
                                    self.annual_change_in_price_of_imported_inputs))
                                    
        setattr(self, 'rcifcoiipt_list', self.clean_inf_array(rcifcoiipt_list))

        #Nominal CIF Cost of imported inputs per ton
        ncifcoiipt_usd_list= [i*j for i,j in zip(self.uspi_list,rcifcoiipt_list)]
        setattr(self, 'ncifcoiipt_usd_list', self.clean_inf_array(ncifcoiipt_usd_list))

        #Nominal CIF Cost of imported inputs per ton
        ncifcoiipt_lc_list= [i*j for i,j in zip(self.ner_list,ncifcoiipt_usd_list)]
        setattr(self, 'ncifcoiipt_lc_list', self.clean_inf_array(ncifcoiipt_lc_list))
      
        #Cost of imported inputs including import duty per ton
        ncifcoiipt_lc_id_list= [x*(1+self.import_duty) for x in ncifcoiipt_lc_list]
        setattr(self, 'ncifcoiipt_lc_id_list', self.clean_inf_array(ncifcoiipt_lc_id_list))
        # print('ncifcoiipt_lc_list', ncifcoiipt_lc_list)
        # print('ncifcoiipt_lc_id_list', ncifcoiipt_lc_id_list)
        #Total cost of imported inputs including import duty
        tcoiiiid_list= [i*j/1000 for i,j in zip(ncifcoiipt_lc_id_list,self.prod_qnty_list)]
        setattr(self, 'tcoiiiid_list', self.clean_inf_array(tcoiiiid_list))
        # print('tcoiiiid_list', tcoiiiid_list)

        
        #Real Cost of domestic inputs per ton
        rcodipt_list= list(self._accumulate_yearly_rate(self.tp_list,
                         self.cost_of_domestic_inputs_per_ton_of_beef_produced,
                             self.annual_change_in_price_of_domestic_inputs))

        setattr(self, 'rcodipt_list', self.clean_inf_array(rcodipt_list))
        #print('Cost of domestic inputs per ton',self.cost_of_domestic_inputs_per_ton_of_beef_produced)
        #Nominal Cost of domestic inputs per ton
        ncodipt_list= [i*j for i,j in zip(self.dpi_list,rcodipt_list)]
        setattr(self, 'ncodipt_list', self.clean_inf_array(ncodipt_list))
     
        #Total cost of domestic inputs
        tcodi_list= [i*j/1000 for i,j in zip(ncodipt_list,self.prod_qnty_list)]
        setattr(self, 'tcodi_list', self.clean_inf_array(tcodi_list))
        # print('ncodipt_list',ncodipt_list)
        #print('Total cost of domestic inputs',tcodi_list)


        #Total cost of inputs per ton (nominal)
        inputs_cost_per_ton_nominal_list= [i+j for i,j in zip(ncifcoiipt_lc_id_list,ncodipt_list)]
        setattr(self, 'inputs_cost_per_ton_nominal_list', inputs_cost_per_ton_nominal_list)
        
        #Total Input Cost (nominal)
        tic_n_list= [i+j for i,j in zip(tcoiiiid_list,tcodi_list)]
        setattr(self, 'tic_n_list', tic_n_list)
        
       
    def _metric_cashFlow(self):
        
        #A. Total Cash Flow (+)
        #-------------------------------------------------------------------------------------
        #A. 1 RECEIPTS
        #----------------------------------------
        #A.1.1 Gross sales revenue
        #A.1.2 Change in A/R
        total_receipts_list =np.zeros(len(self.tp_list))     

        gross_revenus_sales_list=  np.array(self.gsr_list)
        change_in_AR_list=  np.array(self.ciar_list)
        #-----------------------------------------------------------------
        total_receipts_list =self.clean_inf_array(gross_revenus_sales_list) + self.clean_inf_array(change_in_AR_list) #+ c -d

        # print('total_receipts_list', total_receipts_list)

        #A.2  Liquidation Value
        total_liquidation_list =np.zeros(len(self.tp_list))
        total_liquidation_list= np.array(self.residual_val_buildings_nominal_list) + \
                                np.array(self.residual_val_machinery_nominal_list) + \
                                np.array(self.residual_val_land_nominal_list)
        
        # print('residual_val_buildings_nominal_list',self.residual_val_buildings_nominal_list)
        # print('residual_val_machinery_nominal_list',self.residual_val_machinery_nominal_list)
        # print('residual_val_land_nominal_list',self.residual_val_land_nominal_list)
        # print('total_liquidation_list',total_liquidation_list)


        #-----------------------------------------------
        #Total Cash_flow (+)
        total_cash_inflow =np.zeros(len(self.tp_list))
        total_cash_inflow= self.clean_inf_array(total_receipts_list) + self.clean_inf_array(total_liquidation_list)
        #****************************************************************************************
        #B. Total Cash Out Flow(+)
        #2.EXpenditure
        #--------------------------------------------------------------------------------------
        #  2.1 Investments
        total_invest_cost_list= np.zeros(len(self.tp_list))
        #  2.1.1 Investment cost of Feedlots (nominal)
        feedlot_cost_list= np.array(self.cost_feedlots_nominal_list)
       
        #  2.1.2 Investment Land (nominal)
        cost_of_land_nominal_list= np.array(self.cost_of_land_nominal_list)
       
        #  2.1.3 Investment Machinery (nominal)
        total_machinery_incl_import_duty_nominal_list= np.array(self.total_machinery_incl_import_duty_nominal_list)
       
        #  2.1.4 Investment Building (nominal)
        cost_of_building_nominal_list= np.array(self.cost_of_building_nominal_list)
       
        #Total Investment Cost
        total_invest_cost_list=feedlot_cost_list + cost_of_land_nominal_list + \
                             total_machinery_incl_import_duty_nominal_list +cost_of_building_nominal_list
        
        #----------------------------------------------------------------------------
        #  2.2 Operating cost
        #       2.2.1 Total cost of imported inputs including import duty
        imports_cost=  np.array(self.tcoiiiid_list)
        #print('imports_cost',imports_cost)
        #       2.2.2 Total cost of domestic inputs
        domestic_input_cost =  np.array(self.tcodi_list)
        #print('domestic_input_cost',domestic_input_cost)
      
        #       2.2.3 Total Labour Cost
        tlc_list=  np.array(self.tlc_list) 
        #print('tlc_list',tlc_list)
      
        #       2.2.4 Cost Of Electricity (nominal)
        coe=  np.array(self.cost_of_electricity_list)
        #print('Electricity',coe)
      
        #       2.2.5 Other Indirect Costs (nominal)
        oic=  np.array(self.other_indirect_costs_list)
        #print('Other Indirect Cost',oic)
        #print('other_indirect_costs',self.other_indirect_costs)
        #       Change in A/P
        ciap=  np.array(self.ciap_list)
        #print('AP',ciap)
	    #       Change in cash balance
       
        cicb=  np.array(self.cicb_list)
        #print('CB',cicb)

        #       3.Total Cash Outflow
        total_operating_cost_list= np.zeros(len(self.tp_list))
        total_cashoutflow_list= np.zeros(len(self.tp_list))
        
    
       
       
       
     
        
        total_operating_cost_list=imports_cost + domestic_input_cost + tlc_list + ciap +cicb +coe +oic
        #---------------------------------------------------------------------------------------------
        #print('total_operating_cost_list',total_operating_cost_list)
        
        #Total Outflow
        total_cashoutflow_list= self.clean_inf_array(total_operating_cost_list) + self.clean_inf_array(total_invest_cost_list)
        #**************************************************************************************************
        #print('total_cashoutflow_list',total_cashoutflow_list)

        #C. NET CASH FLOW BEFORE TAXES
        net_before_tax_before_financing= total_cash_inflow-total_cashoutflow_list
        #print('net_before_tax_before_financing',net_before_tax_before_financing)

        #*****************************************************************************
       
        #1.1 Sales tax paid
        sales_tax_paid=  np.array(self.st_list)
        #print('sales_tax_paid',sales_tax_paid)
        #1.2 Income tax payment
        income_tax_list=  np.array(self.income_tax_payment_list)
        #print('income_tax_list',income_tax_list)
        #D--- Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)
        net_cash_after_tax_before_financing= net_before_tax_before_financing- sales_tax_paid  -income_tax_list
       
        #print('net_cash_after_tax_before_financing',net_cash_after_tax_before_financing)
        #print('sales_tax_paid',sales_tax_paid)
        #print('income_tax_list',income_tax_list)
        #*******************************************************************

        #E---Net after tax cash flow after financing (OWNER PERSPECTIVE)
        
        #1.1 Debt cash Inflow in Local Currency
        debt_cash_inflow_lc=  np.array(self.debt_cash_inflow_lc_list)
        #1.2 Total Loan Repayment as an outflow in Local Currency
        total_loan_repayment_lc=  np.array(self.total_loan_repayment_ouflow_lc_list)
        #print('total_loan_repayment_lc',total_loan_repayment_lc)
        net_cash_after_tax_after_financing= self.clean_inf_array(net_cash_after_tax_before_financing)+ \
                                            debt_cash_inflow_lc - \
                                            total_loan_repayment_lc
        #print('net_cash_after_tax_after_financing',net_cash_after_tax_after_financing)
        #----------------------------------------------------------------------------------

        #PV Annual Net Cash Flows (NCF)
        #PV Annual Debt Repayment
        #Discount Rate Equity
        net_cash_after_tax_after_financing_real=[0 if y==0 else x/y  for x,y in zip(net_cash_after_tax_after_financing,self.dpi_list)]
        #print('net_cash_after_tax_after_financing_real',net_cash_after_tax_after_financing_real)
        #npv_= self.npv(net_cash_after_tax_after_financing_real,self.discount_rate_equity)


       
        #npv2_ =npf.npv(self.discount_rate_equity, net_cash_after_tax_after_financing_real[1:]) + net_cash_after_tax_after_financing_real[0];  
       
        npv2_=self._get_financial_npv(self.discount_rate_equity,net_cash_after_tax_after_financing_real)
        #print('real>>>>>', net_cash_after_tax_after_financing_real)
        #print('npv', npv2_)
        #print('discount_rate_equity', self.discount_rate_equity)
        # print('net_cash_after_tax_after_financing_real',net_cash_after_tax_after_financing_real)
        ##irr_ =npf.irr(self.clean_inf_array(net_cash_after_tax_after_financing_real))
        irr_ =irr2_(self.clean_inf_array(net_cash_after_tax_after_financing_real))
      
        #_assert_finite()
        # def _assert_finite(*arrays):
        # for a in arrays:
        # if not isfinite(a).all():
        #     raise LinAlgError("Array must not contain infs or NaNs")

        #print('irr', irr_)
        setattr(self, 'model_npv', npv2_)
        setattr(self, 'model_net_cash_real', stringfy_list(net_cash_after_tax_after_financing_real))
        setattr(self, 'irr', self.clean_inf_array(irr_))
        """ 
        check this error of float:::::>cannot convert float NaN to integer
        """
        mirr_ =npf.mirr(net_cash_after_tax_after_financing_real,self.discount_rate_equity,self.discount_rate_equity,)#finance_rate, reinvest_rate
        #print('mirr', round(mirr_,2))
        setattr(self, 'mirr', self.clean_inf_array(mirr_))
     
        # mirr= []
        """    
        myList = list(np.around(np.array(myList),2))
        myRoundedList =  [round(x,2) for x in mylist]
    
        """

        lb, up=self._get_loan_period_bounds()
        #print('LB --UP: >>>>>', lb, up,self.loan_principal_repayment_flag_list)
        
        # PV Annual Net Cash Flows (NCF)
       
        
        pv_ncf_list=[npf.npv(self.discount_rate_equity, net_cash_after_tax_before_financing[i+1:up+1]) + \
                        net_cash_after_tax_before_financing[i] for i in range(lb,up+1) ] 
        
        
        #print('pv_ncf_list',pv_ncf_list)
        setattr(self, 'pv_ncf_list', pv_ncf_list)
        # print('discount_rate_equity', self.discount_rate_equity)
        pv_annual_debt_repayment_list=[npf.npv(self.discount_rate_equity, self.total_loan_repayment_ouflow_lc_list[i+1:up+1]) + \
                        self.total_loan_repayment_ouflow_lc_list[i] for i in range(lb,up+1) ] 
          
        #print('pv_annual_debt_repayment_list',pv_annual_debt_repayment_list)
        setattr(self, 'pv_annual_debt_repayment_list', pv_annual_debt_repayment_list)
        # PV Annual Debt Repayment
        # ADSCR
        adscr = [0 if y==0 else x/y  for x,y in zip(net_cash_after_tax_before_financing[lb:up+1],
                                              self.total_loan_repayment_ouflow_lc_list[lb:up+1])]
        #print('adscr',adscr)
        setattr(self, 'adscr', self.clean_inf_array(adscr))
        # LLCR
        llcr =[0 if y==0 else x/y  for x,y in zip(pv_ncf_list, pv_annual_debt_repayment_list)]
        #print('llcr',llcr)
        setattr(self, 'llcr', self.clean_inf_array(llcr))
       
        # Summary of ADSCR
        # #--------------------------------------------		
		# Minimum ADSCR
        adscr_min=np.min(self.adscr)
        #print('adscr_min',adscr_min)
        setattr(self, 'adscr_min', self.clean_inf_array(adscr_min))
		
        # Maximum ADSCR
        adscr_max=np.max(self.adscr)
        #print('adscr_max',adscr_max)
        setattr(self, 'adscr_max', self.clean_inf_array(adscr_max))
       
        # Average ADSCR
        adscr_av=np.average(self.adscr)
        #print('adscr_av',adscr_av)
        setattr(self, 'adscr_av', self.clean_inf_array(adscr_av))

        for i in range(len(self.adscr)):
            setattr(self, 'adscr_'+ str(i+1), adscr[i])

       
        # Summary of LLCR	
        # #--------------------------------------------	
        # Minimum LLCR
        llcr_min=np.min(self.llcr)
        #print('llcr_min',llcr_min)
        setattr(self, 'llcr_min', llcr_min)
        
        # Maximum LLCR
        llcr_max=np.max(self.llcr)
        #print('llcr_max',llcr_max)
        setattr(self, 'llcr_max', llcr_max)
        
        # Average LLCR
        llcr_average=np.average(self.llcr)
        #print('llcr_av',llcr_average)
        setattr(self, 'llcr_av', llcr_average)
        
        for i in range(len(self.llcr)):
            setattr(self, 'llcr_'+ str(i+1), llcr[i])

        