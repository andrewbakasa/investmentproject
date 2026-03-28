import datetime
import io
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils.timezone import make_naive, is_aware

# OpenPyXL imports
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Project-specific imports
from trading.models import Investment, Investor
from django.contrib.auth.models import User
from common.utils import thousand_sep

def get_excel_investor_list(request, pk):
    obj_invest = get_object_or_404(Investment, pk=pk)
    
    # 1. Data gathering
    investors_qs = Investor.objects.filter(investment__id=pk).values()
    users_qs = User.objects.all().values('id', 'username', 'first_name', 'last_name', 'email')
    
    investors_df = pd.DataFrame(list(investors_qs))
    users_df = pd.DataFrame(list(users_qs))
    
    # Initialize empty DF with required columns to prevent crashes on empty results
    cols = [
        'Title', 'Motivation Statement', 'Amt Pledged', 'Created',
        'Status Changed', 'Status', 'Investor username', 'Investor F-Name',
        'Investor L-Name', 'Investor email', 'pending', 'verification',
        'accepted', 'engagement', 'rejected'
    ]
    df = pd.DataFrame(columns=cols)

    if not investors_df.empty and not users_df.empty:
        # Merge and Cleanup
        users_df.rename(columns={'id': 'user_id'}, inplace=True)
        df = pd.merge(investors_df, users_df, on='user_id', how="left")
        
        # Rename to match your report headers
        df = df.rename(columns={
            'name': 'Title', 
            'first_name': 'Investor F-Name',
            'email': 'Investor email', 
            'last_name': 'Investor L-Name',
            'username': 'Investor username', 
            'value': 'Amt Pledged',
            'application_status': 'Status', 
            'motivation': 'Motivation Statement',
            'date_created': 'Created', 
            'date_status_changed': 'Status Changed'
        })

        # --- FIX: Remove Timezones for Excel Compatibility ---
        for col in ['Created', 'Status Changed']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)

        # Calculate status columns
        status_types = ['pending', 'verification', 'accepted', 'engagement', 'rejected']
        for s in status_types:
            df[s] = df.apply(lambda x: x['Amt Pledged'] if x['Status'] == s else 0, axis=1)
        
        # Keep only the columns we want in the final Excel
        df = df[cols]

    # 2. Create the Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Investor List'     

    # Header Row (Row 1)
    heading = f"INVESTORS LIST WORTH [{thousand_sep(obj_invest.total_value)}] FOR [{obj_invest.name}]".upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    top_cell = ws.cell(row=1, column=1, value=heading)
    top_cell.font = Font(bold=True, size=14)
    top_cell.alignment = Alignment(horizontal='center')

    # Column Headers (Row 2)
    ws.append(list(df.columns))
    
    # Data Rows (Starting Row 3)
    for r in df.values.tolist():
        ws.append(r)
    
    # 3. Summary Formulas (Bottom of data)
    last_data_row = len(df) + 2
    summary_row = last_data_row + 1
    percent_row = last_data_row + 2
    
    # Indices for currency/numeric columns: C(3), K(11) through O(15)
    calc_cols = [3, 11, 12, 13, 14, 15] 
    
    for c in calc_cols:
        col_let = get_column_letter(c)
        # Sum Formula
        ws.cell(row=summary_row, column=c).value = f"=SUM({col_let}3:{col_let}{last_data_row})"
        # Percentage Formula (Relative to Amt Pledged Total in Col C)
        if c > 3:
            ws.cell(row=percent_row, column=c).value = f"={col_let}{summary_row}/C{summary_row}"
            ws.cell(row=percent_row, column=c).number_format = '0.00%'

    # 4. Styling
    setup_excel_styles(ws, df)

    # 5. Export
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"Investors_{obj_invest.name.replace(' ', '_')}_{timestamp}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def setup_excel_styles(ws, df):
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column width slightly
        ws.column_dimensions[get_column_letter(col_num)].width = 18
# import datetime
# import io
# import pandas as pd
# from django.http import HttpResponse
# from django.shortcuts import get_object_or_404
# from django.db.models import Q
# from django.utils.encoding import force_str

# # Standard OpenPyXL imports (Already in your environment)
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# # Project-specific imports
# from trading.models import Investment, Investor
# from django.contrib.auth.models import User
# from common.utils import thousand_sep

# def get_excel_investor_list(request, pk):
#     """
#     Generates an Excel export of investors for a specific investment.
#     Uses pure-python memory buffering to avoid C++ compiler dependencies.
#     """
#     obj_invest = get_object_or_404(Investment, pk=pk)
    
#     # Data gathering using Pandas (Requirement already satisfied in your env)
#     investors_df = pd.DataFrame(Investor.objects.filter(Q(investment__id=pk)).values())
#     users_df = pd.DataFrame(User.objects.all().values())
    
#     df = pd.DataFrame(columns=[
#         'Title', 'Motivation Statement', 'Amt Pledged', 'Created',
#         'Status Changed', 'Status', 'Investor username', 'Investor F-Name',
#         'Investor L-Name', 'Investor email', 'pending', 'verification',
#         'accepted', 'engagement', 'rejected'
#     ])  

#     if users_df.shape[0] > 0 and investors_df.shape[0] > 0:
#         users_df['user_id'] = users_df['id']
#         df = pd.merge(investors_df, users_df, on='user_id', how="left").drop([
#             'id_x', 'id_y', 'investment_id', 'user_id', 'password', 
#             'last_login', 'is_superuser', 'is_staff', 'is_active', 'date_joined'
#         ], axis=1).rename({
#             'name': 'Title', 'first_name': 'Investor F-Name',
#             'email': 'Investor email', 'last_name': 'Investor L-Name',
#             'username': 'Investor username', 'value': 'Amt Pledged',
#             'application_status': 'Status', 'motivation': 'Motivation Statement',
#             'date_created': 'Created', 'date_status_changed': 'Status Changed'
#         }, axis=1)   
        
#         for i in ['pending', 'verification', 'accepted', 'engagement', 'rejected']:
#             df[i] = df.apply(lambda x: get_column_state(x['Amt Pledged'], x['Status'], i), axis=1)     

#     # Create the Workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Investor List'    

#     # Header Row
#     heading_text = f"INVESTORS LIST WORTH [{thousand_sep(obj_invest.total_value)}] FOR [{obj_invest.name}]"
#     ws.append([heading_text.upper()])	
#     ws.append([colname for colname in df.columns])
    
#     # Data Rows
#     for index, row in df.iterrows():
#         ws.append([col for col in row])
    
#     # Totals and Percentages logic
#     total_ref = f'{get_column_letter(15)}1'
#     row_count = len(df)
    
#     # Summary formulas
#     sum_totals = [
#         f"=SUM({get_column_letter(c+1)}3:{get_column_letter(c+1)}{row_count + 2})" 
#         if c in [2, 10, 11, 12, 13, 14] else "" for c in range(15)
#     ]
#     ws.append(sum_totals)

#     as_percent = [
#         f"={get_column_letter(c+1)}{row_count + 3}/{total_ref}" 
#         if c in [2, 10, 11, 12, 13, 14] else "" for c in range(15)
#     ]
#     ws.append(as_percent)

#     # Styling and Formatting
#     # (Applying your existing formatting logic here...)
#     setup_excel_styles(ws, df)

#     # --- PRODUCTION EXPORT LOGIC (The Fix) ---
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     filename = f"Investors_{obj_invest.name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     response["Cache-Control"] = "no-cache, no-store"
    
#     return response

# def get_column_state(amt, status, output_status):
#     return amt if status == output_status else 0

# def setup_excel_styles(ws, df):
#     # This maintains your original visual look for the NRZ reports
#     bold_font = Font(bold=True, color='FFFFCC99', sz=12)
#     for col in range(1, 16):
#         cell = ws.cell(row=2, column=col)
#         cell.font = bold_font
#         ws.column_dimensions[get_column_letter(col)].width = 15

    