import json
from math import ceil
import math

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
import datetime

from django.db.models import Q
from common.views import landing_page
from investments_appraisal.models_landing_page import model_views_page, model_views_page_mentor

from common.decorators import unauthenticated_user, allowed_users, admin_only, \
     registered_user_only_with_client_routing


from django.template.loader import render_to_string

from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.core.mail import send_mail
from django.conf import settings
from common.models import ModifiedRecord, Vacancy
from django.views.decorators.csrf import csrf_exempt




from django.shortcuts import render
from plotly.offline import download_plotlyjs, plot
import plotly.offline as opy
import plotly.graph_objs as go
import chart_studio.plotly as py
import os
import numpy as np
from datetime import datetime
from plotly.graph_objs import Scatter
import plotly.express as px
import pandas as pd


from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
import datetime
from datetime import date
import os
from dateutil.relativedelta import relativedelta
from django.http import JsonResponse
from django.forms.models import model_to_dict
import decimal
import csv
from itertools import chain

from tempfile import NamedTemporaryFile
from excel_response import ExcelResponse
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle,PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle


import pandas as pd
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.shortcuts import render
from django.utils.decorators import method_decorator


import functools

from django.conf import settings
from django.views.generic import DetailView


from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout

from django.contrib import messages

from django.contrib.auth.decorators import login_required 
from django.contrib.auth.models import Group


from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from django.views import generic


from django.shortcuts import render
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from django.views import View

from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, render

from investments_appraisal.models import *

from investments_appraisal.forms import  ContactUsForm, FinancingForm, MacroeconomicParametersForm, NewsSubscribeForm, PurchasePlanForm, TimingAssumptionForm, TimingAssumptionFormUpdate, \
	 UserBusinessModelForm, PricesForm,DepreciationForm ,TaxesForm, UserModelFormUpdate, UserPreferenceForm, UserProfileForm, WorkingCapitalForm



from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models.expressions import F




@login_required(login_url="account_login")
@admin_only
def get_users_comments(request):
    queryset = ContactUs.objects.filter(deleted=False).order_by('-date_created')
    context = {
        "comments": queryset,
        "dashboard_page": "active",
        "total": queryset.count(),
    }

    return render(request, 'investments_appraisal/user_comments.html', context)


@login_required(login_url="account_login")
@admin_only
def get_users_last_login(request):
    queryset = UserProfile.objects.order_by('-last_login')
    context = {
        "users": queryset,
        "dashboard_page": "active",
        "total": queryset.count(),
    }

    return render(request, 'investments_appraisal/users_last_login.html', context)

login_required(login_url="account_login")
@admin_only  
def get_users_last_login_ajax(request):
    if request.method == 'GET' and request.is_ajax():

        queryset = UserProfile.objects.order_by('-last_login') 

        
        #%H:%M:%S
        item_object =data = [{'id': user.pk, 'last_login': user.last_login.strftime("%b. %d, %Y, %I:%M %p")} for user in queryset]# queryset_to_dict(queryset)
                
        return JsonResponse({'error': False, 'data': item_object})
    return JsonResponse({'status': 'Fail', 'message': 'Error, must be an Ajax call.'})


def search(request):
	return render(request, 'investments_appraisal/mentor/search.html',{})	
def display_models(request):
	userdata= UserModel.objects.filter(user=request.user)#.order_by('-date')
	#newsdata = UserModel.objects.all()
	# articles per page
	
	if not ('perpage' in request.session):
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	per_page=request.session['perpage']
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(userdata, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	
	current_page = obj_paginator.get_page(1)
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	
	
	context = {
		'obj_paginator':obj_paginator,
		'first_page':first_page,
		
		'current_page':current_page,
		'page_range':page_range
	}
	#  
	return render(request, 'investments_appraisal/pagenation.html',context)

def display_projects_ajax(request):
	models= ModelCategory.objects.all()#.order_by('-date')

	
	per_page=3
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(models, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	
	#
	if request.method == 'POST':
		#getting page number
		page_no = request.POST.get('page_no', None) 
		num_of_pages= int(obj_paginator.num_pages)
		totalrecords= int(obj_paginator.count)
		current_page = obj_paginator.get_page(page_no)    
		
		
		data={}	
		data["per_page"]=per_page
		data["page_no"]=page_no
		data["num_of_pages"]=num_of_pages
		data["totalrecords"]=totalrecords
			#data["has_previous"]=False
		if current_page.has_previous():
			data["has_previous"]=True  
			data["first"]=1 
			data["previous_page_number"]=current_page.previous_page_number() 
		
		data["current_page"]=current_page.number     
		
		#data["has_next"]=False
		if current_page.has_next():
			data["has_next"]=True  
			data["next_page_number"]=current_page.next_page_number()
		
		data["last"]=current_page.paginator.num_pages 
        

		if int(page_no)>int(num_of_pages):			
			data["results"]=[]
			return JsonResponse({"data":data})
	
		results=[] 

		#name 
		#description 
		#uniqueid  
		#date_created 

		for i in obj_paginator.page(page_no).object_list:
			
			cdate= i.date_created.ctime()
			item_object = model_to_dict(i)
			item_object['date_created']=f'new Date("{i.date_created.ctime()}")'
			item_object['date_created']=f'{i.date_created.ctime()}'
			#print(item_object)
			results.append(item_object)
		
		data["results"]=results
		
		
		return JsonResponse({"data":data})

def display_models_ajax_filter(request,slug, *args, **kwargs):
	userdata= UserModel.objects.filter(user=request.user).filter(name__icontains=slug)
	
	if not ('perpage' in request.session):
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	per_page=request.session['perpage']
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(userdata, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	
	#
	if request.method == 'POST':
		#getting page number
		page_no = request.POST.get('page_no', None) 
		num_of_pages= int(obj_paginator.num_pages)
		totalrecords= int(obj_paginator.count)
		
		current_page = obj_paginator.get_page(page_no)    
		
		
		data={}	
		data["per_page"]=per_page
		data["page_no"]=page_no
		data["num_of_pages"]=num_of_pages
		data["totalrecords"]=totalrecords
			#data["has_previous"]=False
		if current_page.has_previous():
			data["has_previous"]=True  
			data["first"]=1 
			data["previous_page_number"]=current_page.previous_page_number() 
		
		data["current_page"]=current_page.number     
		
		#data["has_next"]=False
		if current_page.has_next():
			data["has_next"]=True  
			data["next_page_number"]=current_page.next_page_number()
		
		data["last"]=current_page.paginator.num_pages 
        

		if int(page_no)>int(num_of_pages):			
			data["results"]=[]
			return JsonResponse({"data":data})
	
		results=[] 												 
		for i in obj_paginator.page(page_no).object_list:			
			cdate= i.date_created.ctime()
			item_object = model_to_dict(i)
			item_object['model_type']=i.model_type.name
			item_object['currency']=i.currency.symbol
			item_object['uniqueid']=i.model_type.uniqueid
			item_object['description']=i.model_type.description
			#item_object['date_created']='new Date("%s")' % i.date_created.ctime()
			item_object['date_created']=f'new Date("{i.date_created.ctime()}")'
			item_object['date_created']=f'{i.date_created.ctime()}'
			#print(item_object)
			results.append(item_object)
		
		data["results"]=results
		
		
		return JsonResponse({"data":data})

def display_models_ajax(request):
	userdata= UserModel.objects.filter(user=request.user)#.order_by('-date')
	
	if not ('perpage' in request.session):
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	per_page=request.session['perpage']
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(userdata, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	
	if request.method == 'POST':
		#getting page number
		page_no = request.POST.get('page_no', None) 
		num_of_pages= int(obj_paginator.num_pages)
		totalrecords= int(obj_paginator.count)
		current_page = obj_paginator.get_page(page_no)   
		data={}	
		data["per_page"]=per_page
		data["page_no"]=page_no
		data["num_of_pages"]=num_of_pages
		data["totalrecords"]=totalrecords
        
		#data["has_previous"]=False
		if current_page.has_previous():
			data["has_previous"]=True  
			data["first"]=1 
			data["previous_page_number"]=current_page.previous_page_number() 
		
		data["current_page"]=current_page.number     
		
		#data["has_next"]=False
		if current_page.has_next():
			data["has_next"]=True  
			data["next_page_number"]=current_page.next_page_number()
		
		data["last"]=current_page.paginator.num_pages 
        



		if int(page_no)>int(num_of_pages):			
			data["results"]=[]
			return JsonResponse({"data":data})
	
	                                          
		results=[] 												 
		for i in obj_paginator.page(page_no).object_list:
			
			
			cdate= i.date_created.ctime()
			item_object = model_to_dict(i)
			item_object['model_type']=i.model_type.name
			item_object['currency']=i.currency.symbol
			item_object['uniqueid']=i.model_type.uniqueid
			item_object['description']=i.model_type.description
			#item_object['date_created']='new Date("%s")' % i.date_created.ctime()
			item_object['date_created']=f'new Date("{i.date_created.ctime()}")'
			item_object['date_created']=f'{i.date_created.ctime()}'
			#print(item_object)
			results.append(item_object)
		
		data["results"]=results
		
		#print(data)
		return JsonResponse({"data":data})


@login_required(login_url="account_login")
def check_count_ajax(request, *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():
			count_=0
			iterations=0
			percent=0
			if ('count' in request.session):
				count_ = request.session['count'] 
				iterations =request.session['total_runs'] 
				percent =round((count_*100/ iterations))
				
			else:
				request.session['count'] =1
				request.session['total_runs'] =10000000000
				percent =0
				request.session.save()
			
			output = {'progress':percent}  
			return JsonResponse({'error': False, 'data': output})
			
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def create_usermodel_instance(request):
	
	if request.method == 'POST':
		#print(request.POST)
		data = request.POST#.get('subject', '')
		currency = get_object_or_404(Currency, id=data['currency'])
		modeltype=get_object_or_404(ModelCategory, id=data['model_type'])
		usermodel = UserModel(model_type=modeltype,name=data['name'],
							 currency=currency, user=request.user)

		form = UserBusinessModelForm(request.POST)
		if form.is_valid():
			form.save()
		
		if UserModel.objects.filter(Q(name = data['name']) & Q(user = request.user)).exists():
			record = UserModel.objects.filter(Q(name = data['name']) & Q(user = request.user)).first()
			dt= record.date_created.strftime('%d %b %Y')
			#print('Record Exist.....')
			#error......
			return HttpResponseRedirect(reverse('user_models',))
		#usermodel.save()
	
	return HttpResponseRedirect(reverse('user_models',))

@login_required(login_url="account_login")
def create_usermodel_ajax2(request,  _id, *args, **kwargs):
    if request.method == 'POST':
        if request.is_ajax():
            form = UserBusinessModelForm(request.POST)
            if form.is_valid():
                form.save()
            else:
                errors = form.errors
                return JsonResponse({'error': True, 'data': errors})  
            
            latest = UserModel.objects.latest('id').id
            record = UserModel.objects.get(pk=latest)
            item_object = model_to_dict(record)
            #print (item_object)
            return JsonResponse({'error': False, 'data': item_object})
        else:
            return JsonResponse({'error': True, 'data': "errors encontered"})
    else:
        error = {
            'message': 'Error, must be an Ajax call.'
        }
        return JsonResponse(error, content_type="application/json")  
@method_decorator(allowed_users(allowed_roles=['admin',  'editor','datacapture']), name='dispatch')  
class UserBusinessModelCreate(LoginRequiredMixin, generic.edit.CreateView):
	model = UserModel
	template_name = 'investments_appraisal/usermodel_edit.html'
	success_url = reverse_lazy('user_models')

	def get_initial(self):
		initial = super().get_initial()
		return initial

	def get_form_class(self):
		return UserBusinessModelForm

	def get_form_kwargs(self):
			kwargs = super().get_form_kwargs()
			kwargs.update({'user': self.request.user})
			return kwargs
	def get_context_data(self, **kwargs):
		context = super(UserBusinessModelCreate, self).get_context_data(**kwargs)
		context['menu'] = '___'
		context['submenu'] = "___" #self.type
		return context

#Any one can access
def index(request):
	#1. Projects Count
	projects_qs= ModelCategory.objects.all()
	if projects_qs:
		projects_count=projects_qs.count()
	else:
		projects_count=0
	
	#2. Models Count
	models_qs= UserModel.objects.all()
	if models_qs:
		models_count=models_qs.count()
	else:
		models_count=0

	
	#3. Downloads
	downloads_qs= Downloads.objects.all()
	if downloads_qs:
		downloads_count=downloads_qs.count()
	else:
		downloads_count=0
	
	popular_projects= ModelCategory.objects.order_by('-likes')[:3]

	context = {
		'popular_projects':popular_projects,
		'title' : 'Home',
		'models_count' : models_count,
		'projects_count' : projects_count,
		'downloads_count' : downloads_count,
	}
	

	return render(request, 'investments_appraisal/mentor/index.html', context)


@login_required(login_url="account_login")
def user_models(request):
	usermodels= UserModel.objects.filter(user=request.user)#.order_by('-date')
	form = UserBusinessModelForm(initial={'user': request.user})
	# articles per page
	
	if not ('perpage' in request.session):
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	per_page=request.session['perpage']
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(usermodels, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	current_page = obj_paginator.get_page(1)
	#print("...............")
	#print(current_page)  
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	context = {
		'obj_paginator':obj_paginator,
		'first_page':first_page,
		'current_page':current_page,
		'page_range':page_range,
		"user": request.user,
		"usermodels": usermodels,#dummy delete later
		'form': form
	}

	
	#  
	return render(request, 'investments_appraisal/user_models.html',context)


def home_page(request):
	
	return render(request, 'investments_appraisal/mentor/index.html', {})

def about(request):
	#1. Projects Count
	projects_qs= ModelCategory.objects.all()
	if projects_qs:
		projects_count=projects_qs.count()
	else:
		projects_count=0
	
	#2. Models Count
	models_qs= UserModel.objects.all()
	if models_qs:
		models_count=models_qs.count()
	else:
		models_count=0

	
	#3. Downloads
	downloads_qs= Downloads.objects.all()
	if downloads_qs:
		downloads_count=downloads_qs.count()
	else:
		downloads_count=0
	
	context = {
		'models_count' : models_count,
		'projects_count' : projects_count,
		'downloads_count' : downloads_count,
	}
	return render(request, 'investments_appraisal/mentor/about.html', context)
def contact(request):
	return render(request, 'investments_appraisal/mentor/contact.html', {})

def commingsoon(request) :
	context= {
		}
	return render(request, 'investments_appraisal/commingsoon.html', context)
def buy(request,type):
	email=''
	if request.user.is_authenticated:
		email= request.user.email
	if type=='free':
		context= {
              'user_plan':'free plan',
			  'email':	email,

		}
	elif type=='standard':
		context= {
 		'user_plan':'standard plan',
		 'email':	email,
		}
	elif type=='business':
		context= {
 		'user_plan':'business plan',
		 'email':	email,
		}
	elif type=='ultimate':
		context= {
 		'user_plan':'ultimate plan',
		 'email':	email,
		}	
	else:
		context= {
 			'user_plan':'other plan',
			 'email':	email,
		}

	return render(request, 'investments_appraisal/mentor/buy.html', context)
def projects(request):
	models= ModelCategory.objects.all()
	
	per_page=3
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(models, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	current_page = obj_paginator.get_page(1)
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	
	
	context = {
		'obj_paginator':obj_paginator,
		'first_page':first_page,
		
		'current_page':current_page,
		'page_range':page_range
	}
	
	return render(request, 'investments_appraisal/mentor/projects.html', context)

@login_required(login_url="account_login")
def models(request):
	usermodels= UserModel.objects.filter(user=request.user)#.order_by('-date')
	form = UserBusinessModelForm(initial={'user': request.user})
	# articles per page
	
	if not ('perpage' in request.session):
		#print('session[perpage] on set')
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			#print('userpref found')
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	else:
		#print('2. session[perpage] is...')
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			#print('2.userpref found')
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
		#print('2. session[perpage] =', request.session['perpage'])
	

	per_page=request.session['perpage']
	# Paginator in a view function to paginate a queryset
	# show 4 news per page
	obj_paginator = Paginator(usermodels, per_page)
	# list of objects on first page
	first_page = obj_paginator.page(1).object_list
	current_page = obj_paginator.get_page(1)
	# range iterator of page numbers
	page_range = obj_paginator.page_range

	context = {
		'obj_paginator':obj_paginator,
		'first_page':first_page,
		'current_page':current_page,
		'page_range':page_range,
		"user": request.user,
		"usermodels": usermodels,#dummy delete later
		'form': form
	}



	
	return render(request, 'investments_appraisal/mentor/models.html', context)
def experts(request):
	return render(request, 'investments_appraisal/mentor/experts.html', {})

def jobs(request):
	current_time = datetime.datetime.now()
	#only future events that are visible
	qs= Vacancy.objects.filter(due_date__gte=current_time,accessible=True)

	context = {
		'vacancies':qs,
		"total":qs.count()
	}
	return render(request, 'investments_appraisal/mentor/jobs.html', context)

def events(request):
	current_time = datetime.datetime.now()
	#only future events that are visible
	qs= Events.objects.filter(date__gte=current_time,accessible=True)

	context = {
		'events':qs,
	}

	return render(request, 'investments_appraisal/mentor/events.html', context)

def pricing(request):
	return render(request, 'investments_appraisal/mentor/pricing.html', {})

def project_details(request,id):
	#model= ModelCategory.objects.filter(pk=id)
	model = get_object_or_404(ModelCategory,pk=id)
	
	context ={'model':model}
	return render(request, 'investments_appraisal/mentor/project-details.html', context)

def vacancy_details(request,id):
	model = get_object_or_404(Vacancy,pk=id)
	
	context ={'model':model }
	return render(request, 'investments_appraisal/mentor/job-details.html', context)

@login_required(login_url="account_login")
def select_model_specs_page(request, model_id):
	#Divet to appropriate app
	# FISH:
	#BEEF
	usermodel = get_object_or_404(UserModel,pk=model_id)
	uniqueid=usermodel.model_type.uniqueid
	if uniqueid in model_views_page.keys():
		reverse_page= model_views_page[uniqueid]

	#print(f'reverse: {reverse_page}')
	
	return HttpResponseRedirect(reverse(reverse_page, args=(model_id,)))
   
@login_required(login_url="account_login")
def select_model_specs_page_mentor(request, model_id):
	#Divet to appropriate app
	# FISH:
	#BEEF
	usermodel = get_object_or_404(UserModel,pk=model_id)
	uniqueid=usermodel.model_type.uniqueid
	if uniqueid in model_views_page_mentor.keys():
		if 'ready' in model_views_page_mentor[uniqueid]:
			if model_views_page_mentor[uniqueid]['ready']== True:
				reverse_page= model_views_page_mentor[uniqueid]['view']
				#print(f'reverse: {reverse_page} {uniqueid}')
				return HttpResponseRedirect(reverse(reverse_page, args=(model_id,)))
			else:
				messages.success(request, "Project WIP")
				return HttpResponseRedirect(reverse('models'))
		else:
			#ready not define for the model
			messages.success(request, "parameter ready in in dict:Contact Admin")
			return HttpResponseRedirect(reverse('models'))
	else:
		# model app not yet available
		messages.success(request, "Project not yet availble")
		return HttpResponseRedirect(reverse('models'))
	    


	
@login_required(login_url="account_login")
def update_user_model(request):
	pass





#PRICES-------------------------------
@login_required(login_url="account_login")
def edit_model_prices_ajax(request,  model_id, price_id, *args, **kwargs):

	prices = get_object_or_404(Prices, pk=price_id)
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			title = request.POST['title']
			base_price = request.POST['base_price']
			change_in_price = request.POST['change_in_price']

			if base_price is None or base_price == 'None':
				base_price = 0.0
			else:
				base_price= decimal.Decimal(base_price)

			if change_in_price is None or change_in_price == 'None':
				change_in_price = 0.0
			else:
				change_in_price= decimal.Decimal(change_in_price)
			
			prices_instance = get_object_or_404(Prices, pk=price_id)
			
			
			try:
				prices_instance.title =title
				prices_instance.base_price =base_price
				prices_instance.change_in_price =change_in_price
				prices_instance.save()
				prices_item_object = model_to_dict(Prices.objects.get(pk=price_id))
				
				return JsonResponse({'error': False, 'data': prices_item_object})
			
			except (KeyError, Prices.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': prices_item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_prices_ajax(request, model_id,  *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			usermodel_id = request.POST['usermodel']
			title = request.POST['title']
			base_price = request.POST['base_price']
			change_in_price = request.POST['change_in_price']

			if base_price is None or base_price == 'None':
				base_price = 0.0
			else:
				base_price= decimal.Decimal(base_price)

			if change_in_price is None or change_in_price == 'None':
				change_in_price = 0.0
			else:
				change_in_price= decimal.Decimal(change_in_price)
			
			
			usermodel = get_object_or_404(UserModel,pk=usermodel_id)

			i = Prices.objects.create(usermodel=usermodel, title=title,
										base_price=base_price, change_in_price=change_in_price)
			i.save()
			#print('new model>>>', i)

			latest = Prices.objects.latest('id').id
			prices_instance = Prices.objects.get(pk=latest)
			prices_item_object = model_to_dict(prices_instance)
			prices_item_object['model_id']=usermodel_id
			prices_item_object['model_name']=i.usermodel.name
			#print("returned data", prices_item_object)
			
			messages.success(request, "Successfully added prices")
			
			return JsonResponse({'error': False, 'data': prices_item_object})
			
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")


# Timing Assumptions-------------------------------
@login_required(login_url="account_login")
def edit_model_timing_assumptions_ajax(request,  model_id, ta_id,  *args, **kwargs):

	ta_instance = get_object_or_404(TimingAssumption, pk=ta_id)
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			base_period = request.POST['base_period']
			construction_start_year = request.POST['construction_start_year']
			construction_len = request.POST['construction_len']

			construction_year_end = request.POST['construction_year_end']
			operation_start_year = request.POST['operation_start_year']
			operation_duration = request.POST['operation_duration']

			operation_end = request.POST['operation_end']
			number_of_months_in_a_year = request.POST['number_of_months_in_a_year']

			ta_instance = get_object_or_404(TimingAssumption, pk=ta_id)
			
			
			try:
				ta_instance.base_period =base_period
				ta_instance.construction_start_year =construction_start_year
				ta_instance.construction_len =construction_len

				ta_instance.construction_year_end =construction_year_end
				ta_instance.operation_start_year =operation_start_year
				ta_instance.operation_duration =operation_duration

				ta_instance.operation_end =operation_end
				ta_instance.number_of_months_in_a_year =number_of_months_in_a_year

				ta_instance.save()

				ta_item_object = model_to_dict(TimingAssumption.objects.get(pk=ta_id))
				
				return JsonResponse({'error': False, 'data': ta_item_object})
			
			except (KeyError, TimingAssumption.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': ta_item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_timing_assumptions_ajax(request, model_id,  *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			usermodel_id = request.POST['usermodel']
			base_period = request.POST['base_period']
			construction_start_year = request.POST['construction_start_year']
			construction_len = request.POST['construction_len']

			
			construction_year_end = request.POST['construction_year_end']
			operation_start_year = request.POST['operation_start_year']
			operation_duration = request.POST['operation_duration']

			operation_end = request.POST['operation_end']
			number_of_months_in_a_year = request.POST['number_of_months_in_a_year']

			#form = Prices(usermodel,title,base_price,change_in_price)
			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = TimingAssumption.objects.create(usermodel=usermodel, base_period=base_period,
					construction_start_year=construction_start_year, construction_len=construction_len,
					construction_year_end=construction_year_end, operation_start_year=operation_start_year,
					operation_duration=operation_duration, operation_end=operation_end,
					number_of_months_in_a_year=number_of_months_in_a_year)
			i.save()
			#print('new model>>>', i)

			latest = TimingAssumption.objects.latest('id').id
			ta_instance = TimingAssumption.objects.get(pk=latest)
			ta_item_object = model_to_dict(ta_instance)
			ta_item_object['model_id']=usermodel_id
			ta_item_object['model_name']=i.usermodel.name
			#print("returned data", ta_item_object)
				
			messages.success(request, "Successfully added timing assuptions")
			
			return JsonResponse({'error': False, 'data': ta_item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

#Depreciation-----------------------------------
@login_required(login_url="account_login")
def edit_model_depreciation_ajax(request, model_id, depr_id,   *args, **kwargs):

	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			economic_life_of_machinery = request.POST['economic_life_of_machinery']
			economic_life_of_buildings = request.POST['economic_life_of_buildings']
			tax_life_of_machinery = request.POST['tax_life_of_machinery']

			tax_life_of_buildings = request.POST['tax_life_of_buildings']
			tax_life_of_soft_capital_costs = request.POST['tax_life_of_soft_capital_costs']

			_instance = get_object_or_404(Depreciation, pk=depr_id)
			
			
			try:
				_instance.economic_life_of_machinery =economic_life_of_machinery
				_instance.economic_life_of_buildings =economic_life_of_buildings
				_instance.tax_life_of_machinery =tax_life_of_machinery

				_instance.tax_life_of_buildings =tax_life_of_buildings
				_instance.tax_life_of_soft_capital_costs =tax_life_of_soft_capital_costs
				
				_instance.save()

				_item_object = model_to_dict(Depreciation.objects.get(pk=depr_id))
				
				return JsonResponse({'error': False, 'data': _item_object})
			
			except (KeyError, Depreciation.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': _item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_depreciation_ajax(request, model_id,  *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			usermodel_id = request.POST['usermodel']
			economic_life_of_machinery = request.POST['economic_life_of_machinery']
			economic_life_of_buildings = request.POST['economic_life_of_buildings']
			tax_life_of_machinery = request.POST['tax_life_of_machinery']

			tax_life_of_buildings = request.POST['tax_life_of_buildings']
			tax_life_of_soft_capital_costs = request.POST['tax_life_of_soft_capital_costs']

			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = Depreciation.objects.create(usermodel=usermodel, economic_life_of_machinery=economic_life_of_machinery,
					economic_life_of_buildings=economic_life_of_buildings, tax_life_of_machinery=tax_life_of_machinery,
					tax_life_of_buildings=tax_life_of_buildings, tax_life_of_soft_capital_costs=tax_life_of_soft_capital_costs,)
			i.save()
			

			latest = Depreciation.objects.latest('id').id
			_instance = Depreciation.objects.get(pk=latest)
			item_object = model_to_dict(_instance)
			item_object['model_id']=usermodel_id
			item_object['model_name']=i.usermodel.name
			#print("returned data", item_object)
				
			messages.success(request, "Successfully added Depreciation")
			
			return JsonResponse({'error': False, 'data': item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

#Taxes---------------------------------------------
@login_required(login_url="account_login")
def edit_model_taxes_ajax(request, model_id, taxes_id, *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			
			import_duty = request.POST['import_duty']
			sales_tax = request.POST['sales_tax']
			corporate_income_tax = request.POST['corporate_income_tax']

			_instance = get_object_or_404(Taxes, pk=taxes_id)
			
			
			try:
				_instance.import_duty =import_duty
				_instance.sales_tax =sales_tax
				_instance.corporate_income_tax =corporate_income_tax

				_instance.save()

				_item_object = model_to_dict(Taxes.objects.get(pk=taxes_id))
				
				return JsonResponse({'error': False, 'data': _item_object})
			
			except (KeyError, Taxes.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': _item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_taxes_ajax(request, model_id,  *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():

			usermodel_id = request.POST['usermodel']
			import_duty = request.POST['import_duty']
			sales_tax = request.POST['sales_tax']
			corporate_income_tax = request.POST['corporate_income_tax']

			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = Taxes.objects.create(usermodel=usermodel, import_duty=import_duty,
					sales_tax=sales_tax, corporate_income_tax=corporate_income_tax,)
			i.save()
			

			latest = Taxes.objects.latest('id').id
			_instance = Taxes.objects.get(pk=latest)
			item_object = model_to_dict(_instance)
			item_object['model_id']=usermodel_id
			item_object['model_name']=i.usermodel.name
			#print("returned data", item_object)
				
			messages.success(request, "Successfully added Taxes")
			
			return JsonResponse({'error': False, 'data': item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

#Finacing--------------------------------------------------
@login_required(login_url="account_login")
def edit_model_financing_ajax(request, id, *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():			
			real_interest_rate = request.POST['real_interest_rate']
			risk_premium = request.POST['risk_premium']
			num_of_installments = request.POST['num_of_installments']
			grace_period = request.POST['grace_period']
			repayment_starts = request.POST['repayment_starts']
			equity = request.POST['equity']
			senior_debt = request.POST['senior_debt']

			if real_interest_rate is None or real_interest_rate == 'None':
				real_interest_rate = 0.0
			else:
				real_interest_rate= decimal.Decimal(real_interest_rate)

			if risk_premium is None or risk_premium == 'None':
				risk_premium = 0.0
			else:
				risk_premium= decimal.Decimal(risk_premium)
			
			
			if equity is None or equity == 'None':
				equity = 0.0
			else:
				equity= decimal.Decimal(equity)

			if senior_debt is None or senior_debt == 'None':
				senior_debt = 0.0
			else:
				senior_debt= decimal.Decimal(senior_debt)
			
		
		
			_instance = get_object_or_404(Financing, pk=id)
			
		
			try:
				_instance.real_interest_rate =real_interest_rate
				_instance.risk_premium =risk_premium
				_instance.num_of_installments =num_of_installments

				_instance.grace_period =grace_period
				_instance.repayment_starts =repayment_starts
				_instance.equity =equity
				_instance.senior_debt =senior_debt



				_instance.save()

				_item_object = model_to_dict(Financing.objects.get(pk=id))
				
				return JsonResponse({'error': False, 'data': _item_object})
			
			except (KeyError, Financing.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': _item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_financing_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():         
			usermodel_id = request.POST['usermodel']
			real_interest_rate = request.POST['real_interest_rate']
			risk_premium = request.POST['risk_premium']
			num_of_installments = request.POST['num_of_installments']
			grace_period = request.POST['grace_period']
			repayment_starts = request.POST['repayment_starts']
			equity = request.POST['equity']
			senior_debt = request.POST['senior_debt']

			if real_interest_rate is None or real_interest_rate == 'None':
				real_interest_rate = 0.0
			else:
				real_interest_rate= decimal.Decimal(real_interest_rate)

			if risk_premium is None or risk_premium == 'None':
				risk_premium = 0.0
			else:
				risk_premium= decimal.Decimal(risk_premium)
			
			
			if equity is None or equity == 'None':
				equity = 0.0
			else:
				equity= decimal.Decimal(equity)

			if senior_debt is None or senior_debt == 'None':
				senior_debt = 0.0
			else:
				senior_debt= decimal.Decimal(senior_debt)
			#print('Financing: usermodel_id....>', usermodel_id)
			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = Financing.objects.create(usermodel=usermodel, real_interest_rate=real_interest_rate,
					risk_premium=risk_premium, num_of_installments=num_of_installments,
					grace_period=grace_period, repayment_starts=repayment_starts,
					equity=equity, senior_debt=senior_debt,)
			i.save()
			

			latest = Financing.objects.latest('id').id
			_instance = Financing.objects.get(pk=latest)
			item_object = model_to_dict(_instance)
			item_object['model_id']=usermodel_id
			item_object['model_name']=i.usermodel.name
			#print("returned data", item_object)
				
			messages.success(request, "Successfully added Financing parameters")
			
			return JsonResponse({'error': False, 'data': item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

#Working Capital--------------------------------------------------
@login_required(login_url="account_login")
def edit_model_workingcapital_ajax(request, id, *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():

			accounts_receivable = request.POST['accounts_receivable']
			accounts_payable = request.POST['accounts_payable']
			cash_balance = request.POST['cash_balance']
			
			if accounts_receivable is None or accounts_receivable == 'None':
				accounts_receivable = 0.0
			else:
				accounts_receivable= decimal.Decimal(accounts_receivable)

			if accounts_payable is None or accounts_payable == 'None':
				accounts_payable = 0.0
			else:
				accounts_payable= decimal.Decimal(accounts_payable)			
			
			if cash_balance is None or cash_balance == 'None':
				cash_balance = 0.0
			else:
				cash_balance= decimal.Decimal(cash_balance)

		
		
			_instance = get_object_or_404(WorkingCapital, pk=id)
			
		
			try:
				_instance.accounts_receivable =accounts_receivable
				_instance.accounts_payable =accounts_payable
				_instance.cash_balance =cash_balance



				_instance.save()

				_item_object = model_to_dict(WorkingCapital.objects.get(pk=id))
				
				return JsonResponse({'error': False, 'data': _item_object})
			
			except (KeyError, WorkingCapital.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': _item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_workingcapital_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax(): 

			usermodel_id = request.POST['usermodel']
			accounts_receivable = request.POST['accounts_receivable']
			accounts_payable = request.POST['accounts_payable']
			cash_balance = request.POST['cash_balance']
			
			if accounts_receivable is None or accounts_receivable == 'None':
				accounts_receivable = 0.0
			else:
				accounts_receivable= decimal.Decimal(accounts_receivable)

			if accounts_payable is None or accounts_payable == 'None':
				accounts_payable = 0.0
			else:
				accounts_payable= decimal.Decimal(accounts_payable)			
			
			if cash_balance is None or cash_balance == 'None':
				cash_balance = 0.0
			else:
				cash_balance= decimal.Decimal(cash_balance)

			
			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = WorkingCapital.objects.create(usermodel=usermodel, accounts_receivable=accounts_receivable,
					accounts_payable=accounts_payable, cash_balance=cash_balance)
			i.save()
			

			latest = WorkingCapital.objects.latest('id').id
			_instance = WorkingCapital.objects.get(pk=latest)
			item_object = model_to_dict(_instance)
			item_object['model_id']=usermodel_id
			item_object['model_name']=i.usermodel.name
			#print("returned data", item_object)
				
			messages.success(request, "Successfully added WorkingCapital parameters")
			
			return JsonResponse({'error': False, 'data': item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

#Macroeconomic Parameters--------------------------------------------------
@login_required(login_url="account_login")
def edit_model_macroeconomicparameters_ajax(request, id, *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():

			discount_rate_equity = request.POST['discount_rate_equity']
			domestic_inflation_rate = request.POST['domestic_inflation_rate']
			us_inflation_rate = request.POST['us_inflation_rate']

			exchange_rate = request.POST['exchange_rate']
			dividend_payout_ratio = request.POST['dividend_payout_ratio']
			num_of_shares = request.POST['num_of_shares']
			investment_costs_over_run_factor = request.POST['investment_costs_over_run_factor']
			
			investment_costs_over_run_factor= decimal.Decimal(investment_costs_over_run_factor) if investment_costs_over_run_factor else 0.0
			discount_rate_equity= decimal.Decimal(discount_rate_equity) if discount_rate_equity else 0.0
			domestic_inflation_rate= decimal.Decimal(domestic_inflation_rate) if domestic_inflation_rate else 0.0
			us_inflation_rate= decimal.Decimal(us_inflation_rate) if us_inflation_rate else 0.0
			exchange_rate= decimal.Decimal(exchange_rate) if exchange_rate else 0.0
			dividend_payout_ratio= decimal.Decimal(dividend_payout_ratio) if dividend_payout_ratio else 0.0

			_instance = get_object_or_404(MacroeconomicParameters, pk=id)
			
		
			try:
				_instance.discount_rate_equity =discount_rate_equity
				_instance.domestic_inflation_rate =domestic_inflation_rate
				_instance.us_inflation_rate =us_inflation_rate
				_instance.exchange_rate =exchange_rate
				_instance.dividend_payout_ratio =dividend_payout_ratio
				_instance.num_of_shares =num_of_shares
				_instance.investment_costs_over_run_factor=investment_costs_over_run_factor




				_instance.save()

				_item_object = model_to_dict(MacroeconomicParameters.objects.get(pk=id))
				
				return JsonResponse({'error': False, 'data': _item_object})
			
			except (KeyError, MacroeconomicParameters.DoesNotExist):
				#errors = form.errors
				errors= {'__all__': ['No data changed']} 
				return JsonResponse({'erro': True, 'data': errors})
			else:
				return JsonResponse({'error': False, 'data': _item_object})
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def add_model_macroeconomicparameters_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax(): 

			usermodel_id = request.POST['usermodel']
			discount_rate_equity = request.POST['discount_rate_equity']
			domestic_inflation_rate = request.POST['domestic_inflation_rate']
			us_inflation_rate = request.POST['us_inflation_rate']

			exchange_rate = request.POST['exchange_rate']
			dividend_payout_ratio = request.POST['dividend_payout_ratio']
			num_of_shares = request.POST['num_of_shares']
			investment_costs_over_run_factor = request.POST['investment_costs_over_run_factor']
			
			investment_costs_over_run_factor= decimal.Decimal(investment_costs_over_run_factor) if investment_costs_over_run_factor else 0.0
			discount_rate_equity= decimal.Decimal(discount_rate_equity) if discount_rate_equity else 0.0
			domestic_inflation_rate= decimal.Decimal(domestic_inflation_rate) if domestic_inflation_rate else 0.0
			us_inflation_rate= decimal.Decimal(us_inflation_rate) if us_inflation_rate else 0.0
			exchange_rate= decimal.Decimal(exchange_rate) if exchange_rate else 0.0
			dividend_payout_ratio= decimal.Decimal(dividend_payout_ratio) if dividend_payout_ratio else 0.0


			usermodel = get_object_or_404(UserModel,pk=usermodel_id)
		
		
			i = MacroeconomicParameters.objects.create(usermodel=usermodel, discount_rate_equity=discount_rate_equity,
					domestic_inflation_rate=domestic_inflation_rate, us_inflation_rate=us_inflation_rate,
					exchange_rate=exchange_rate,dividend_payout_ratio=dividend_payout_ratio,
					num_of_shares=num_of_shares,investment_costs_over_run_factor=investment_costs_over_run_factor)
			i.save()
			

			latest = MacroeconomicParameters.objects.latest('id').id
			_instance = MacroeconomicParameters.objects.get(pk=latest)
			item_object = model_to_dict(_instance)
			item_object['model_id']=usermodel_id
			item_object['model_name']=i.usermodel.name
			#print("returned data", item_object)
				
			messages.success(request, "Successfully added Macroeconomic parameters")
			
			return JsonResponse({'error': False, 'data': item_object})
				
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")


@login_required(login_url="account_login")
def delete_bussiness_model(request, model_id):
	#print("did igo here")
	model_ = get_object_or_404(UserModel, pk=model_id)
	try:
		model_.delete()
	except:
		context = {
			'error_message' : "Unable to delete attachment!",
			'order_id' : model_id
		}
		return render(request, 'customers/order.html', context)
	else:
		return HttpResponseRedirect(reverse('order', args=(model_id,)))

@login_required(login_url="account_login")
@admin_only
def delete_comment_ajax(request,id, *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():
			
			model_ = get_object_or_404(ContactUs, pk=id)
			model_.deleted=True
			model_.save()
			#model_.delete()
			item_object = model_to_dict(model_)

			#total_pages= get_total_pages(request)
			data= {}
			#data['total_pages']=total_pages
			#data['deleted_page']=page_no
			
			data['model']=item_object
			
			messages.success(request, "Successfully deleted  record")
			
			return JsonResponse({'error': False, 'data': data})
			
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

@login_required(login_url="account_login")
def delete_bussiness_model_ajax(request, model_id, page_no, *args, **kwargs):
	if request.method == 'POST':
		#print(request.POST)
		if request.is_ajax():
			
			model_ = get_object_or_404(UserModel, pk=model_id)

			model_.delete()
			item_object = model_to_dict(model_)

			total_pages= get_total_pages(request)
			data= {}
			data['total_pages']=total_pages
			data['deleted_page']=page_no
			
			data['model']=item_object
			
			messages.success(request, "Successfully deleted  model")
			
			return JsonResponse({'error': False, 'data': data})
			
		else:
			return JsonResponse({'error': True, 'data': "errors encontered"})
	else:
		error = {
			'message': 'Error, must be an Ajax call.'
		}
		return JsonResponse(error, content_type="application/json")

login_required(login_url="account_login")
def create_usermodel_ajax(request):
	if request.method == 'POST':
		form = UserBusinessModelForm(data=request.POST or None)
	else:
		#first time call from loco failure list
		form = UserBusinessModelForm(initial={'user': request.user})
	return create_all_usermodel(request,form,'investments_appraisal/user_model_modal.html')

def get_total_pages(request):
	if not ('perpage' in request.session):
		obj= UserPreference.objects.filter(user=request.user).first()
		if obj:
			request.session['perpage']=obj.perpage
		else:# nothing in db
			request.session['perpage']= 3
	per_page=request.session['perpage']
	#per_page=request.session['perpage']
	recordcount= UserModel.objects.filter(user=request.user).count()
	#avoide error
	if per_page==0:
		per_page=6
	return math.ceil(recordcount/per_page)

login_required(login_url="account_login")
def create_all_usermodel(request,form,template_name):

	data = dict()
	errors= None
	if request.method == 'POST':
		errors=form.errors
		
		if form.is_valid():
			form.save()
			total_pages= get_total_pages(request)
			data['form_is_valid'] = True
			latest = UserModel.objects.latest('id').id
			record = UserModel.objects.get(pk=latest)
			item_object = model_to_dict(record)
			data['model'] = item_object
			data['total_pages']=total_pages
		else:
			data['form_is_valid'] = False
			# print('nofield errors: ',form.non_field_errors)
			# print('errors:::',form.errors)
	context = {
		'form':form
	}
	data['html_form'] = render_to_string(template_name,context,request=request)
	data['error']=errors
	return JsonResponse(data)


login_required(login_url="account_login")
def usermodel_update_ajax(request,pk):
	usermodel = get_object_or_404(UserModel,pk=pk)
	if request.method == 'POST':
		form = UserModelFormUpdate(request.POST or None, 
										instance=usermodel)
	else:
		#first time call from loco failure list
		form = UserModelFormUpdate( instance=usermodel)
	return save_all_usermodel(request,form,'investments_appraisal/user_model_edit_modal.html')



def save_all_usermodel(request,form,template_name):

	data = dict()
	errors= None
	# retrieve product
	pk = form.instance.id
	item_instance = get_object_or_404(UserModel,pk=pk)
	if request.method == 'POST':
		# retrieve product
		pk = form.instance.id
		item_instance = get_object_or_404(UserModel,pk=pk)
		#eject errors for modal form display
		errors=form.errors
		if form.is_valid():
			if item_instance.locked==True:
				data['locked']= 'Model record locked. Ask Admin for Help'
				data['form_is_valid'] = False
				data['html_form'] = render_to_string(template_name,{'form':form},request=request)
				return JsonResponse(data)
			else:
				
			
				form.save()
				pk = form.instance.id
				item_instance = get_object_or_404(UserModel,pk=pk)
				data['form_is_valid'] = True
				item_object = model_to_dict(item_instance)
				item_object['currency']=item_instance.currency.symbol
				item_object['model_type']=item_instance.model_type.name
				item_object['description']=item_instance.model_type.description
				#print(item_object)
				data['model'] = item_object
		else:
		
			data['form_is_valid'] = False

	context = {
		'form':form,
		'usermodel':item_instance,
	}

	data['html_form'] = render_to_string(template_name,context,request=request)
	data['error']= errors
	return JsonResponse(data)


def purchase_plan_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():
			form = PurchasePlanForm(request.POST or None)
			form.instance.user =request.user
			if form.is_valid():
				form.save()
				
			else:
				errors = form.errors
				return JsonResponse({'error': True, 'data': errors})  
			
			latest =  PurchasePlan.objects.latest('id').id
			record =  PurchasePlan.objects.get(pk=latest)
			item_object = model_to_dict(record)
			return JsonResponse({'error': False, 'data': item_object})
		else:
			return JsonResponse({'error': True, 'data': "Request not ajax"})
	else:
		return JsonResponse({'error': True, 'data': "Request not ajax"})


def contact_us_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():
			form = ContactUsForm(request.POST or None)
			form.instance.user =request.user
			if form.is_valid():
				form.save()
				
			else:
				errors = form.errors
				return JsonResponse({'error': True, 'data': errors})  
			
			latest =  ContactUs.objects.latest('id').id
			record =  ContactUs.objects.get(pk=latest)
			item_object = model_to_dict(record)
			return JsonResponse({'error': False, 'data': item_object})
		else:
			return JsonResponse({'error': True, 'data': "Request not ajax"})
	else:
		return JsonResponse({'error': True, 'data': "Request not ajax"})


def project_search_ajax(request,slug, *args, **kwargs):
	#
	if request.method == 'POST':
		#description= request.POST['description']
		modelsdata= ModelCategory.objects.filter(description__icontains=slug)
	
		if not ('perpage' in request.session):
			if request.user.is_authenticated:
				obj= UserPreference.objects.filter(user=request.user).first()
			    # anonymous user is not iterable
				if obj:
					request.session['perpage']=obj.perpage
				else:# nothing in db
					request.session['perpage']= 3
			else:
				request.session['perpage']= 3

		per_page=request.session['perpage']

		obj_paginator = Paginator(modelsdata, per_page)
		# list of objects on first page
		first_page = obj_paginator.page(1).object_list
		# range iterator of page numbers
		page_range = obj_paginator.page_range

		if request.is_ajax():
			#getting page number
			page_no = request.POST.get('page_no', None) 
			num_of_pages= int(obj_paginator.num_pages)
			totalrecords= int(obj_paginator.count)
			current_page = obj_paginator.get_page(page_no)    
		
		
			data={}	
			data["per_page"]=per_page
			data["page_no"]=page_no
			data["num_of_pages"]=num_of_pages
			data["totalrecords"]=totalrecords
				#data["has_previous"]=False
			if current_page.has_previous():
				data["has_previous"]=True  
				data["first"]=1 
				data["previous_page_number"]=current_page.previous_page_number() 
			
			data["current_page"]=current_page.number     
			
			#data["has_next"]=False
			if current_page.has_next():
				data["has_next"]=True  
				data["next_page_number"]=current_page.next_page_number()
			
			data["last"]=current_page.paginator.num_pages 
			

			
			if int(page_no)>int(num_of_pages):			
				data["results"]=[]
				return JsonResponse({"data":data})
			results=[] 												 
			for i in obj_paginator.page(page_no).object_list:			
				cdate= i.date_created.ctime()
				item_object = model_to_dict(i)
				item_object['date_created']=f'new Date("{i.date_created.ctime()}")'
				item_object['date_created']=f'{i.date_created.ctime()}'
				#print(item_object)
				results.append(item_object)
				
			data["results"]=results
		
		return JsonResponse({"data":data})


def news_subscribe_ajax(request,  *args, **kwargs):
	if request.method == 'POST':
		if request.is_ajax():
			form = NewsSubscribeForm(request.POST or None)
			form.instance.user =request.user
			if form.is_valid():
				form.save()
				
			else:
				errors = form.errors
				#print('error:' , errors)
				return JsonResponse({'error': True, 'data': errors})  
			
			latest =  NewsSubscribe.objects.latest('id').id
			record =  NewsSubscribe.objects.get(pk=latest)
			item_object = model_to_dict(record)
			#print (item_object)
			return JsonResponse({'error': False, 'data': item_object})
		else:
			return JsonResponse({'error': True, 'data': "Request not ajax"})
	else:
		return JsonResponse({'error': True, 'data': "Request not ajax"})

def update_model_likes_ajax(request,  id, *args, **kwargs):
	#obj = get_object_or_404(ModelCategory,id=id)
	obj = ModelCategory.objects.filter(pk=id)
	if request.method == 'POST':
		if request.is_ajax():
			obj.update(likes=F('likes') + 1)		
			#obj.save()
		else:
			return JsonResponse({'error': True, 'data': 'errors'})  
			
		record = ModelCategory.objects.get(pk=id)
		item_object = model_to_dict(record)
		#print (item_object)
		return JsonResponse({'error': False, 'data': item_object})
	else:
		return JsonResponse({'error': True, 'data': "Request not ajax"})

@login_required(login_url="account_login")
def update_user_profile_ajax(request,  *args, **kwargs):
	profile, created = UserProfile.objects.get_or_create(user=request.user)
	if request.method == 'POST':
		if request.is_ajax():
			form = UserProfileForm(request.POST or None,instance=profile)
			form.instance.user =request.user
			if form.is_valid():
				form.save()
				
			else:
				errors = form.errors
				return JsonResponse({'error': True, 'data': errors})  
			
			latest = UserProfile.objects.latest('id').id
			record = UserProfile.objects.get(pk=latest)
			item_object = model_to_dict(record)
			#print (item_object)
			return JsonResponse({'error': False, 'data': item_object})
		else:
			return JsonResponse({'error': True, 'data': "Request not ajax"})

	return JsonResponse({'error': True, 'data': "Request not ajax"})
	

@login_required(login_url="account_login")
def update_user_pref_ajax(request,  *args, **kwargs):
	pref, created = UserPreference.objects.get_or_create(user=request.user)
	if request.method == 'POST':
		if request.is_ajax():
			form = UserPreferenceForm(request.POST or None,instance=pref)
			form.instance.user =request.user
			if form.is_valid():
				form.save()
				obj= UserPreference.objects.filter(user=request.user).first()
				if obj:
					request.session['perpage']=obj.perpage
			else:
				errors = form.errors
				#print('error:' , errors)
				return JsonResponse({'error': True, 'data': errors})  
			
			latest = UserPreference.objects.latest('id').id
			record = UserPreference.objects.get(pk=latest)
			item_object = model_to_dict(record)
			#print (item_object)
			return JsonResponse({'error': False, 'data': item_object})
		else:
			return JsonResponse({'error': True, 'data': "Request not ajax"})
	else:
		form = UserPreferenceForm(instance=pref)
		context ={}
		context['form']=form
		return render(request, 'investments_appraisal/mentor/user_settings.html', context)

def upload_form_user_pref(request):
	pref, created = UserPreference.objects.get_or_create(user=request.user)
	form = UserPreferenceForm(instance=pref)
	context ={}
	context['form']=form
	return render(request, 'investments_appraisal/mentor/user_settings.html', context)

def upload_form_user_profile(request):
	pref, created = UserProfile.objects.get_or_create(user=request.user)
	form = UserProfileForm(instance=pref)
	context ={}
	personal_record= False
	if pref:
		if pref.age or pref.country or pref.sex or pref.profession or pref.aboutyou:
			personal_record=True

	context={
		'form':form,
		'personal_record':personal_record
	}
	return render(request, 'investments_appraisal/mentor/user_profile.html', context)

def update_user_pref(request):
	pref, created = UserPreference.objects.get_or_create(user=request.user)
	#print(pref)
	if request.method == 'POST':
		form = UserPreferenceForm(request.POST or None,instance=pref)
		form.instance.user =request.user
		if form.is_valid():
			form.save()
			obj= UserPreference.objects.filter(user=request.user).first()
			if obj:
				request.session['perpage']=obj.perpage

			return HttpResponseRedirect(reverse('index',)) 
		else:
			#return same page
			context ={}
			context['form']=form
			return render(request, 'investments_appraisal/mentor/user_settings.html', context)
			
	else:
		form = UserPreferenceForm(instance=pref)
		context ={}
		context['form']=form
		return render(request, 'investments_appraisal/mentor/user_settings.html', context)
