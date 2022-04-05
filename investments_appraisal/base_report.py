
from re import A, L
from django.utils import translation
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, force_str
from django.utils.html import escape
from django.utils.translation import gettext as _
from django.utils.formats import get_format, number_format
from django.utils.text import capfirst, get_text_list, format_lazy
from io import StringIO, BytesIO


from tempfile import NamedTemporaryFile
from excel_response import ExcelResponse
import numpy as np
from openpyxl import Workbook, load_workbook 
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle,PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator
import numpy_financial as npf
import math
class BaseBusinessReport:

    description_font = Font(name='Calibri',bold=False,  scheme='minor', sz=11.0)
    description_font2 = Font(name='Calibri',bold=False, color="FF3F3F76", scheme='minor', sz=12.0)
    #_get_number_formats_= {'PERCENT':'0%', 'NUMBER':'GENERAL','YEARS':'GENERAL','YEAR':'GENERAL', 'BLANK': 'GENERAL'}
    # Track all input for cell formulae refrences
    track_inputs = { 
                    'timing':{
                        'base_period':{},
                        'construction_year_end':{},
                        'operation_start_year':{},
                        'operation_end':{},
                        'number_of_months_in_a_year':{},
                        },

                    'flags':{
                        'years':{},
                        },

                    'financing':{
                        'repayment_starts':{},
                        'num_of_installments':{},
                        'grace_period':{},
                        'real_interest_rate':{},
                        'risk_premium':{},
                        'equity':{},
                        'senior_debt':{},
                        }, 

                    'depreciation':{
                        'economic_life_of_machinery':{},
                        'economic_life_of_buildings':{},
                        'tax_life_of_machinery':{},
                        'tax_life_of_buildings':{},
                        'tax_life_of_soft_capital_costs':{},
                        },

                    'working_capital':{
                        'accounts_receivable':{},
                        'accounts_payable':{},
                        'cash_balance':{}
                        },
                    'taxes':{
                        'import_duty':{},
                        'sales_tax':{},
                        'corporate_income_tax':{}
                        },
                    
                    'macroeconomic_parameters':{
                        'header': {},
                        'discount_rate_equity': {}, 
                        'domestic_inflation_rate': {},
                        'us_inflation_rate': {}, 
                        'exchange_rate': {},
                        'dividend_payout_ratio': {},
                        'num_of_shares': {},
                        },

                     'feedlot_design_parameters':{
                        'header': {},
                        'num_of_feedlots': {}, 
                        'length': {},
                        'width': {},
                        'sqm': {}, 
                        'pen_area': {}, 
                        'sqm_per_cattle': {},
                        'total_cattle_per_pen_per_cycle': {},
                        'num_of_months_per_cycle': {},
                        'cattle_per_pen_per_year': {},
                        },

                     'prices':{
                        'header': {},
                        'base_price':{},
                        'change_in_price': {}, 
                        },    

                    'production_inventory':{
                        'header': {},
                        'cattle_per_pen_per_year': {}, 
                        'cattle_survival_rate': {},
                        'thousand': {}, 
                        'dressed_weight_at_selling': {},
                        'num_of_feedlot_targeted_for_construction': {},
                        'cum_pens_under_cattle': {},
                        'cum_pens_under_harvesting': {},
                        'production_quantity': {},
                        'closing_inventory': {},
                        'investment_roll_out_flag': {},
                        'total_pen_constructed': {},
                        },

                     'investment_cost':{
                        'header': {},
                        'total_land_for_pens': {}, 
                        'cost_of_land_per_sqm': {},
                        'cost_of_machinery_per_pen': {}, 
                        'cost_of_building_per_sqm': {},
                        'total_pens': {},
                        'cost_of_pen_construction': {},
                        'senior_debt_dynamic_parameter': {},
                        'initial_pens_employed': {},
                        'pen_cattle_density': {},
                        'pen_length': {},
                        'pen_width': {},
                        'pen_height': {},
                        'total_land_required': {},
                        'cost_of_pens_constructed': {},
                        'cost_of_land': {},
                        'investment_cost_of_land': {},
                        'cif_cost_of_machinery': {},
                        'investment_cost_of_buildings': {},
                        'investment_costs_over_run_factor': {},
                        }, 
                    'cattle_business_options': {'header': {},},#fish_business_options
                    'sens':{
                        'header': {},
                        'domestic_inflation_rate': {}, 
                        'us_inflation_rate': {},
                        'change_in_price': {}, 
                        'base_price':{},
                        'accounts_receivable': {},
                        'accounts_payable': {},
                        'cash_balance': {},
                        'exchange_rate': {},
                        'senior_debt': {},
                        'total_cattle_per_pen_per_cycle': {},
                        'cattle_survival_rate': {},
                        'cattle_price_per_unit': {},
                        'initial_pens_employed': {}
                        },



                     'outputs':{
                        'header': {},
                        'model_selection': {},
                        },   


                    }
    # Model specifications


    
    model_specifications = [
        {'name':'LC','title':'Financial Statements Expressed in Local Currency','units':'LC'},
        {'name':'T_LC','title':'Financial Statements Expressed in thousands of Local Currency','units':"000's LC"},
        {'name':'M_LC','title':'Financial Statements Unit in Millions of Local Currency	Mil','units':'Mil LC'},
        {'name':'YEAR','title':'Year','units':'Year'},
        {'name':'YEARS','title':'Years','units':'Years'},
        {'name':'PERCENT','title':'Percentage','units':'%'},
        {'name':'T_TONS','title':'Thousand tons','units':"000's tons"},
        {'name':'KG','title':'Kilograms','units':'kg'},
        {'name':'TONS','title':'Tons','units':'tons'},
        {'name':'M_USD','title':'Million US Dollars','units':'Mil USD'},
        {'name':'T_USD','title':'Thousand US Dollars','units':"000's USD"},
        {'name':'USD','title':'US Dollars','units':'USD'},
        {'name':'NUMBER','title':'Number','units':'#'},
        {'name':'M_CONVERSION','title':'Million to thousand conversion','units':1000},
        {'name':'LITRES','title':'Litres','units':"Litres"},
        {'name':'FLAG','title':'Flag (1= true, 0= false)','units':'flag'},
        {'name':'SQM','title':'Square metre','units':'sqm'},
        {'name':'SWITCH','title':'Switch','units':"switch"},
        {'name':'NOTE','title':'Note','units':'Note'},
    ]    
  
  
    # TIMING ASSUMPTIONS
    timing_assumptions = {
        'base_period': {'title': 'Base Period','value': 2021}, 
        'construction_start_year': {'title': 'Construction start Year','value': 2021},
        'construction_len': {'title': 'Construction length','value': 1}, 
        'construction_year_end': {'title': 'Construction End Year','value': 2021},#one year
        'operation_start_year': {'title': 'Operation Start Year','value': 2022}, 
        'operation_duration': {'title': 'Operation Duration','value': 18},
        'operation_end': {'title': 'Operation End ','value': 2039}, 
        'number_of_months_in_a_year': {'title': 'Number of months in a year','value': 12},
            
        }
    

    #Prices
    prices = {
        'title': 'Net Sales Price:', 
        'base_price': 0, 
        'change_in_price': 0, 
    }
    #Working Capital
    working_capital = {
        'accounts_receivable': {'title': 'Accounts receivable [% of Gross Sales]','value': 0.10, 'units': 'PERCENT'}, 
        'accounts_payable': {'title': 'Acccounts payable (% of total input cost)','value': 0.10, 'units': 'PERCENT'}, 
        'cash_balance': {'title': 'Cash balance  (% of gross sales)','value': 0.10, 'units': 'PERCENT'},              
    }               
    #Financing	
    financing = {
        'real_interest_rate': {'title': 'Real interest rate','value': 6, 'units': 'PERCENT'}, 
        'risk_premium': {'title': 'Risk premium','value': 1, 'units': 'PERCENT'}, 
        'num_of_installments': {'title': 'No. of installments','value': 6, 'units': 'NUMBER'},  
        'grace_period': {'title': 'Grace period (years)','value': 0, 'units': 'YEARS'}, 
        'repayment_starts': {'title': 'Repayment starts year','value':2022, 'units': 'YEAR'}, 
        'equity': {'title': 'Equity (% of Investment Costs)','value': 30, 'units': 'PERCENT'}, 
        'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': 0.7, 'units': 'PERCENT'},
    }
    #Taxes
    taxes = {
        'import_duty': {'title': 'Import duty','value': 0.10, 'units': 'PERCENT'}, 
        'sales_tax': {'title': 'Sales tax','value': .05, 'units': 'PERCENT'}, 
        'corporate_income_tax': {'title': 'Corporate income tax','value': 0.25, 'units': 'PERCENT'}
    }
    #Macroeconomic Parameters
    macroeconomic_parameters = {
        'discount_rate_equity': {'title': 'Discount Rate Equity','value': 0.12, 'units': 'PERCENT'}, 
        'domestic_inflation_rate': {'title': 'Domestic Inflation rate','value': 0.05, 'units': 'PERCENT'},
        'us_inflation_rate': {'title': 'US Inflation rate','value': 0.03, 'units': 'PERCENT'}, 
        'exchange_rate': {'title': 'Exchange Rate (LC/USD - in Base Year )','value': 1.4, 'units': 'NUMBER'},
        'dividend_payout_ratio': {'title': 'Divident Payout Ratio','value': .50, 'units': 'PERCENT'},
        'num_of_shares': {'title': 'N# of shares','value': 10, 'units': 'NUMBER'},
    }
    #Depreciation
    depreciation = {
        'economic_life_of_machinery': {'title': 'Economic life of machinery','value': 20, 'units': 'YEARS'}, 
        'economic_life_of_buildings': {'title': 'Economic life of buildings','value': 32, 'units': 'YEARS'},
        'tax_life_of_machinery': {'title': 'Tax life of machinery','value': 10, 'units': 'YEARS'}, 
        'tax_life_of_buildings': {'title': 'Tax life of buildings','value': 30, 'units': 'YEARS'},
        'tax_life_of_soft_capital_costs': {'title': 'Tax life of soft capital costs','value': 3, 'units': 'YEARS'},
    }
   #Flags
    flags = {
        'YEAR': {'title': 'Years','units': 'Year'}, 
        'PS': {'title': 'Project start','units': 'flag'},
        'CP': {'title': 'Construction period flag','units':'flag'}, 
        'LPP': {'title': 'Loan principle repayment','units': 'flag'},
        'OP': {'title': 'Operating period','units': 'flag'},
        'RES': {'title': 'Residuals','units': 'flag'}, 
            
    }



    def __init__(self, model_specifications=None, timing_assumptions=None, 
                 prices= None, working_capital=None,
                 financing=None, taxes=None, 
                 macroeconomic_parameters=None,  depreciation=None, 
                 description_font=None, description_font2=None,):
        
        # self.is_bound = data is not None or files is not None # false/True
        self.model_specifications = model_specifications
        self.timing_assumptions = timing_assumptions
        self.prices = prices
        self.working_capital = working_capital
        self.financing = financing
        self.taxes = taxes
        self.macroeconomic_parameters = macroeconomic_parameters
        self.depreciation = depreciation
        self.financing = financing

        if description_font is not None:
            self.description_font= description_font

        if description_font2 is not None:
            self.description_font2= description_font2
        
        #initial calculate this values
        self._cal_metrix()
    def __str__(self):
        """
        String representation of Class object inputs var
        """
        return str(vars(self)) 
    def _get_number_formats(self, x):
        #print(isinstance(x,str),type(x)==str )
        if type(x)==str:
            if x.upper() =='PERCENT':
                return '0%'
            elif x.upper()=='NUMBER':
                return '_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
            else:
                return 'General' 
        else:
            return 'General'              
        
    def _cal_metrix(self):
        #1 Price Index
        self._metric_priceIndex()
        
        #2. Investments
        self._metric_investmentCost()

        #3. Residual Values
        self._metric_residualValues()
        
        #4. Production & Sales
        self._metric_productionAndSales()
        
       
        #5 Labour Costs
        self._metric_labourCost()
        
        #6 Purchases
        self._metric_purchases()
        
        #7. Working Capital
        self._metric_workingCapital()

        #8. LOAN SCHEDULE (Nominal)
        self._metric_loanSchedule()

        #9 Depreciation
        self._metric_depreciationForTaxPurposes()
        
        #10 Unit Production Cost
        self._metric_unitCost()

        #11
        self._metric_finishedProdictEvaluationFIFO()
        
        #12 INCOME TAX STATEMENT
        self._metric_IncomeTaxStatement()
        #13 Cash Flow
        self._metric_cashFlow()

    def _simulate_metrix(self):
        #1 Price Index
        self._simulation_priceIndex()
        
        #2. Investments
        self._simulation_investmentCost()
        
        #3. Residual Values
        self._metric_residualValues()

        #4. Production & Sales
        self._metric_productionAndSales()
        
       
        #5 Labour Costs
        self._metric_labourCost()
        
        #6 Purchases
        self._metric_purchases()
        
        #7. Working Capital
        self._metric_workingCapital()
        
        #8. LOAN SCHEDULE (Nominal)
        self._metric_loanSchedule()
         
        #9 Depreciation
        self._metric_depreciationForTaxPurposes()
        
        #10 Unit Production Cost
        self._metric_unitCost()

        #11
        self._metric_finishedProdictEvaluationFIFO()
        
        #12 INCOME TAX STATEMENT
        self._metric_IncomeTaxStatement()
        
        #13 Cash Flow
        self._metric_cashFlow()
    def _metric_priceIndex(self):
        
        
        if not hasattr(self, 'accounts_receivable'):
            accounts_receivable= self.working_capital['accounts_receivable']['value']       
            setattr(self, 'accounts_receivable', accounts_receivable)

        accounts_payable= self.working_capital['accounts_payable']['value']
        setattr(self, 'accounts_payable', accounts_payable)
        cash_balance= self.working_capital['cash_balance']['value']
        setattr(self, 'cash_balance', cash_balance)
        
        #Base price of beef per ton
        base_price=self.prices['base_price']
        #dynamic var    
        setattr(self, 'base_price', base_price)
        #Change in real price of Beef
        change_in_price=self.prices['change_in_price']#['value'] 
        #dynamic var    
        setattr(self, 'change_in_price', change_in_price)

        #Cattle parameters
        num_feedlots=self.investment_parameter_options['initial_pens_employed'][0] 
        setattr(self, 'initial_pens_employed', num_feedlots)
        fedlot_length_m= self.investment_parameter_options['pen_length'][0]
        setattr(self, 'pen_length', fedlot_length_m)
        fedlot_width_m=self.investment_parameter_options['pen_width'][0]
        setattr(self, 'pen_width', fedlot_width_m)     
        cattle_pen_cycle=self.investment_parameter_options['pen_cattle_density'][0]
        setattr(self, 'cattle_pen_cycle', cattle_pen_cycle)        
        months_per_cycle=self.feedlot_design_parameters['num_of_months_per_cycle']['value']
        setattr(self, 'months_per_cycle', months_per_cycle)
        
        cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
        setattr(self, 'cattle_per_pen_per_year', cattle_per_pen_per_year)
       

        cattle_survival_rate=self.cost_real['cattle_survival_rate']['value']
        setattr(self, 'cattle_survival_rate', cattle_survival_rate)
        
        dressed_weight_at_selling=self.cost_real['dressed_weight_at_selling']['value']
        setattr(self, 'dressed_weight_at_selling', dressed_weight_at_selling)

        
        cost_of_electricity_per_pen_per_annum=self.cost_real['cost_of_electricity_per_pen_per_annum']['value']
        setattr(self, 'cost_of_electricity_per_pen_per_annum', cost_of_electricity_per_pen_per_annum)

    
        #Economic parameter:
        real_interest_rate=self.financing['real_interest_rate']['value']
        setattr(self, 'real_interest_rate', real_interest_rate)
        risk_premium=self.financing['risk_premium']['value']
        setattr(self, 'risk_premium', risk_premium)
        num_of_installments=self.financing['num_of_installments']['value']
        setattr(self, 'num_of_installments', num_of_installments)
        repayment_starts=self.financing['repayment_starts']['value']
        setattr(self, 'repayment_starts', repayment_starts)
        
        grace_period=self.financing['grace_period']['value']
        setattr(self, 'grace_period', grace_period)
       
        #inputs costs
        annual_change_in_price_of_imported_inputs=self.cost_real['annual_change_in_price_of_imported_inputs']['value']
        setattr(self, 'annual_change_in_price_of_imported_inputs', annual_change_in_price_of_imported_inputs)

        annual_change_in_price_of_domestic_inputs=self.cost_real['annual_change_in_price_of_domestic_inputs']['value']
        setattr(self, 'annual_change_in_price_of_domestic_inputs', annual_change_in_price_of_domestic_inputs)

        cost_of_imported_inputs_per_ton_of_beef_produced=self.cost_real['cost_of_imported_inputs_per_ton_of_beef_produced']['value']
        setattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced', cost_of_imported_inputs_per_ton_of_beef_produced)
     
        # #---cattle inputs---------------------
        food_conversion_ratio=self.cost_real['food_conversion_ratio']['value']
        setattr(self, 'food_conversion_ratio', food_conversion_ratio)

        daily_weight_gain=self.cost_real['daily_weight_gain']['value']
        setattr(self, 'daily_weight_gain', daily_weight_gain)
        
        days_in_feed_lot=self.cost_real['days_in_feed_lot']['value']
        setattr(self, 'days_in_feed_lot', days_in_feed_lot)

        weight_at_purchase=self.cost_real['weight_at_purchase']['value']
        setattr(self, 'weight_at_purchase', weight_at_purchase)
        
        live_weight_gain=self.cost_real['live_weight_gain']['value']
        setattr(self, 'live_weight_gain', live_weight_gain)

        qnty_of_feed_per_cattle=self.cost_real['qnty_of_feed_per_cattle']['value']
        setattr(self, 'qnty_of_feed_per_cattle', qnty_of_feed_per_cattle)
        
        commercial_feed_per_pen=self.cost_real['commercial_feed_per_pen']['value']
        setattr(self, 'commercial_feed_per_pen', commercial_feed_per_pen)

        dressed_weight_percent=self.cost_real['dressed_weight_percent']['value']
        setattr(self, 'dressed_weight_percent', dressed_weight_percent)        
       
        cattle_price_per_unit=self.cost_real['cattle_price_per_unit']['value']
        setattr(self, 'cattle_price_per_unit', cattle_price_per_unit)

        cattle_feed_price_per_kg=self.cost_real['cattle_feed_price_per_kg']['value']
        setattr(self, 'cattle_feed_price_per_kg', cattle_feed_price_per_kg)
        
        cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs()
        setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        other_indirect_costs=self.cost_real['other_indirect_costs']['value']
        setattr(self, 'other_indirect_costs', other_indirect_costs)
 

         #1.----------------Time Period----------------------
        base_period =self.timing_assumptions['base_period']['value']
        setattr(self, 'base_period', base_period)
        start_year =self.timing_assumptions['operation_start_year']['value']
        setattr(self, 'start_year', start_year)
        end_year =self.timing_assumptions['operation_end']['value']
        setattr(self, 'end_year', end_year)
        
        construction_year_end =self.timing_assumptions['construction_year_end']['value']
        setattr(self, 'construction_year_end', construction_year_end)
        
        
        #end period = ende_year+ 1(deconstruction)+ 1(pythonic)
        tp_list =[x for x in range(int(self.base_period),int(self.end_year)+1+1)]
        #was giving error
        setattr(self, 'tp_list', tp_list)
        
         #2. -------------Price Index-----------------------------
    
        dividend_payout_ratio=self.macroeconomic_parameters['dividend_payout_ratio']['value']
        setattr(self, 'dividend_payout_ratio', dividend_payout_ratio)
        num_of_shares=self.macroeconomic_parameters['num_of_shares']['value']
        setattr(self, 'num_of_shares', num_of_shares)
        domestic_inflation_rate=self.macroeconomic_parameters['domestic_inflation_rate']['value']
        setattr(self, 'domestic_inflation_rate', domestic_inflation_rate)
        discount_rate_equity=self.macroeconomic_parameters['discount_rate_equity']['value']
        setattr(self, 'discount_rate_equity', discount_rate_equity)
        exchange_rate=self.macroeconomic_parameters['exchange_rate']['value']
        setattr(self, 'exchange_rate', exchange_rate)
        us_inflation_rate=self.macroeconomic_parameters['us_inflation_rate']['value']
        setattr(self, 'us_inflation_rate', us_inflation_rate)    
     

        senior_debt= self.financing['senior_debt']['value']
        setattr(self, 'senior_debt', senior_debt)

        equity=self.financing['equity']['value']
        setattr(self, 'equity', equity)

        sales_tax= self.taxes['sales_tax']['value']
        setattr(self, 'sales_tax', sales_tax)       
       

        corporate_income_tax= self.taxes['corporate_income_tax']['value']        
        setattr(self, 'corporate_income_tax', corporate_income_tax)

       
      
        #--2.a Dometstic price index
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', dpi_list)
        
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', uspi_list)
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', rpi_list)
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)
        #print('ner_list', ner_list)


        #Salaries & wages
        mwpw=self.cost_real['monthly_wage_per_worker']['value']
        setattr(self, 'monthly_wage_per_worker', mwpw)

        mwpst=self.cost_real['monthly_wage_per_supervisor']['value']
        setattr(self, 'monthly_wage_per_supervisor', mwpst)
        #Annual Increase in real salaries
        #Annual Increase Salaries  
        aisw=self.cost_real['annual_increase_salaries_workers']['value']
        setattr(self, 'annual_increase_salaries_workers', aisw)
        aispt=self.cost_real['annual_increase_salaries_supervisors_technicians']['value']
        setattr(self, 'annual_increase_salaries_supervisors_technicians', aispt)

        nowps= 5 # num of workers per supervisor
        setattr(self, 'num_workers_per_supervisor', nowps)

        nowpp=self.cost_real['num_of_workers_per_pen']['value']
        setattr(self, 'num_of_workers_per_pen', nowpp)

    def _get_cost_of_domestic_inputs(self):
          
        cost_of_cattle_feed_per_pen=self.cattle_per_pen_per_year*self.qnty_of_feed_per_cattle*self.cattle_feed_price_per_kg
             
        cost_of_input_cattle_per_pen =self.cattle_per_pen_per_year*self.cattle_price_per_unit

        total_input_cost_per_pen_per_year =cost_of_cattle_feed_per_pen + cost_of_input_cattle_per_pen

        tonnes_produced_per_pen_per_year =self.cattle_per_pen_per_year*self.dressed_weight_at_selling*self.cattle_survival_rate/1000
        
        annual_domestic_inputs_cost_per_pen=total_input_cost_per_pen_per_year/tonnes_produced_per_pen_per_year if  tonnes_produced_per_pen_per_year !=0 else 0
        return annual_domestic_inputs_cost_per_pen

    def _simulation_priceIndex(self):

        
        if not hasattr(self, 'accounts_receivable'):
            accounts_receivable= self.working_capital['accounts_receivable']['value']       
            setattr(self, 'accounts_receivable', accounts_receivable)

        if not hasattr(self, 'accounts_payable'):   
            accounts_payable= self.working_capital['accounts_payable']['value']
            setattr(self, 'accounts_payable', accounts_payable)
        
        if not hasattr(self, 'cash_balance'):
            cash_balance= self.working_capital['cash_balance']['value']
            setattr(self, 'cash_balance', cash_balance)
        
        if not hasattr(self, 'base_price'):    
            #Base price of beef per ton
            base_price=self.prices['base_price']
            #dynamic var    
            setattr(self, 'base_price', base_price)
        
        if not hasattr(self, 'change_in_price'):
            #Change in real price of Beef
            change_in_price=self.prices['change_in_price']#['value'] 
            #dynamic var    
            setattr(self, 'change_in_price', change_in_price)

        if not hasattr(self, 'initial_pens_employed'):    
            #Cattle parameters
            num_feedlots=self.investment_parameter_options['initial_pens_employed'][0] 
            setattr(self, 'initial_pens_employed', num_feedlots)
            
        if not hasattr(self, 'pen_length'):    
            fedlot_length_m= self.investment_parameter_options['pen_length'][0]
            setattr(self, 'pen_length', fedlot_length_m)
            
        if not hasattr(self, 'pen_width'):    
            fedlot_width_m=self.investment_parameter_options['pen_width'][0]
            setattr(self, 'pen_width', fedlot_width_m)     
            
        if not hasattr(self, 'cattle_pen_cycle'):    
            cattle_pen_cycle=self.investment_parameter_options['pen_cattle_density'][0]
            setattr(self, 'cattle_pen_cycle', cattle_pen_cycle)        
        
        if not hasattr(self, 'months_per_cycle'):    
            months_per_cycle=self.feedlot_design_parameters['num_of_months_per_cycle']['value']
            setattr(self, 'months_per_cycle', months_per_cycle)
            
        if not hasattr(self, 'cattle_per_pen_per_year'):    
            cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
            setattr(self, 'cattle_per_pen_per_year', cattle_per_pen_per_year)
        
        if not hasattr(self, 'cattle_survival_rate'):
            cattle_survival_rate=self.cost_real['cattle_survival_rate']['value']
            setattr(self, 'cattle_survival_rate', cattle_survival_rate)
            
        if not hasattr(self, 'dressed_weight_at_selling'):
            dressed_weight_at_selling=self.cost_real['dressed_weight_at_selling']['value']
            setattr(self, 'dressed_weight_at_selling', dressed_weight_at_selling)

        if not hasattr(self, 'cost_of_electricity_per_pen_per_annum'):
            cost_of_electricity_per_pen_per_annum=self.cost_real['cost_of_electricity_per_pen_per_annum']['value']
            setattr(self, 'cost_of_electricity_per_pen_per_annum', cost_of_electricity_per_pen_per_annum)

        if not hasattr(self, 'real_interest_rate'):
            #Economic parameter:
            real_interest_rate=self.financing['real_interest_rate']['value']
            setattr(self, 'real_interest_rate', real_interest_rate)
            
        if not hasattr(self, 'risk_premium'):
            risk_premium=self.financing['risk_premium']['value']
            setattr(self, 'risk_premium', risk_premium)

        if not hasattr(self, 'num_of_installments'):    
            num_of_installments=self.financing['num_of_installments']['value']
            setattr(self, 'num_of_installments', num_of_installments)

        if not hasattr(self, 'repayment_starts'):   
            repayment_starts=self.financing['repayment_starts']['value']
            setattr(self, 'repayment_starts', repayment_starts)
            
        if not hasattr(self, 'grace_period'):   
            grace_period=self.financing['grace_period']['value']
            setattr(self, 'grace_period', grace_period)
        
        if not hasattr(self, 'annual_change_in_price_of_imported_inputs'):    
            #inputs costs
            annual_change_in_price_of_imported_inputs=self.cost_real['annual_change_in_price_of_imported_inputs']['value']
            setattr(self, 'annual_change_in_price_of_imported_inputs', annual_change_in_price_of_imported_inputs)

        if not hasattr(self, 'annual_change_in_price_of_domestic_inputs'):    
            annual_change_in_price_of_domestic_inputs=self.cost_real['annual_change_in_price_of_domestic_inputs']['value']
            setattr(self, 'annual_change_in_price_of_domestic_inputs', annual_change_in_price_of_domestic_inputs)

        if not hasattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced'):    
            cost_of_imported_inputs_per_ton_of_beef_produced=self.cost_real['cost_of_imported_inputs_per_ton_of_beef_produced']['value']
            setattr(self, 'cost_of_imported_inputs_per_ton_of_beef_produced', cost_of_imported_inputs_per_ton_of_beef_produced)

        if not hasattr(self, 'food_conversion_ratio'):
            #---cattle inputs---------------------
            food_conversion_ratio=self.cost_real['food_conversion_ratio']['value']
            setattr(self, 'food_conversion_ratio', food_conversion_ratio)

        if not hasattr(self, 'daily_weight_gain'):    
            daily_weight_gain=self.cost_real['daily_weight_gain']['value']
            setattr(self, 'daily_weight_gain', daily_weight_gain)
            
        if not hasattr(self, 'days_in_feed_lot'):    
            days_in_feed_lot=self.cost_real['days_in_feed_lot']['value']
            setattr(self, 'days_in_feed_lot', days_in_feed_lot)

        if not hasattr(self, 'weight_at_purchase'):
            weight_at_purchase=self.cost_real['weight_at_purchase']['value']
            setattr(self, 'weight_at_purchase', weight_at_purchase)
        
        if not hasattr(self, 'live_weight_gain'):
            live_weight_gain=self.cost_real['live_weight_gain']['value']
            setattr(self, 'live_weight_gain', live_weight_gain)

        if not hasattr(self, 'qnty_of_feed_per_cattle'):    
            qnty_of_feed_per_cattle=self.cost_real['qnty_of_feed_per_cattle']['value']
            setattr(self, 'qnty_of_feed_per_cattle', qnty_of_feed_per_cattle)
            
        if not hasattr(self, 'commercial_feed_per_pen'):    
            commercial_feed_per_pen=self.cost_real['commercial_feed_per_pen']['value']
            setattr(self, 'commercial_feed_per_pen', commercial_feed_per_pen)

        if not hasattr(self, 'dressed_weight_percent'):    
            dressed_weight_percent=self.cost_real['dressed_weight_percent']['value']
            setattr(self, 'dressed_weight_percent', dressed_weight_percent) 

        if not hasattr(self, 'cattle_price_per_unit'):
            cattle_price_per_unit=self.cost_real['cattle_price_per_unit']['value']
            setattr(self, 'cattle_price_per_unit', cattle_price_per_unit)
        
        if not hasattr(self, 'cattle_feed_price_per_kg'):
            cattle_feed_price_per_kg=self.cost_real['cattle_feed_price_per_kg']['value']
            setattr(self, 'cattle_feed_price_per_kg', cattle_feed_price_per_kg)
            
        # if not hasattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced'):    
        #     cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs()
        #     setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        if not hasattr(self, 'other_indirect_costs'):    
            other_indirect_costs=self.cost_real['other_indirect_costs']['value']
            setattr(self, 'other_indirect_costs', other_indirect_costs)
    

         #1.----------------Time Period----------------------
        if not hasattr(self, 'base_period'):
            base_period =self.timing_assumptions['base_period']['value']    
            setattr(self, 'base_period', base_period)

        if not hasattr(self, 'start_year'):
            start_year =self.timing_assumptions['operation_start_year']['value']
            setattr(self, 'start_year', start_year)
        
        if not hasattr(self, 'end_year'):      
            end_year =self.timing_assumptions['operation_end']['value']
            setattr(self, 'end_year', end_year)
        
        if not hasattr(self, 'construction_year_end'):           
            construction_year_end =self.timing_assumptions['construction_year_end']['value']
            setattr(self, 'construction_year_end', construction_year_end)
        
        

        if not hasattr(self, 'tp_list'):
            tp_list =[x for x in range(int(self.base_period),int(self.end_year)+1)]       
            setattr(self, 'tp_list', tp_list)

       
            #Economic Parameter
        if not hasattr(self, 'dividend_payout_ratio'):    
            dividend_payout_ratio=self.macroeconomic_parameters['dividend_payout_ratio']['value']
            setattr(self, 'dividend_payout_ratio', dividend_payout_ratio)
            
        if not hasattr(self, 'num_of_shares'):    
            num_of_shares=self.macroeconomic_parameters['num_of_shares']['value']
            setattr(self, 'num_of_shares', num_of_shares)
        
        if not hasattr(self, 'domestic_inflation_rate'):    
            domestic_inflation_rate=self.macroeconomic_parameters['domestic_inflation_rate']['value']
            setattr(self, 'domestic_inflation_rate', domestic_inflation_rate)
            
        if not hasattr(self, 'discount_rate_equity'):    
            discount_rate_equity=self.macroeconomic_parameters['discount_rate_equity']['value']
            setattr(self, 'discount_rate_equity', discount_rate_equity)
            
        if not hasattr(self, 'exchange_rate'):    
            exchange_rate=self.macroeconomic_parameters['exchange_rate']['value']
            setattr(self, 'exchange_rate', exchange_rate)
           
        if not hasattr(self, 'us_inflation_rate'):    
            us_inflation_rate=self.macroeconomic_parameters['us_inflation_rate']['value']
            setattr(self, 'us_inflation_rate', us_inflation_rate)    
     
        if not hasattr(self, 'senior_debt'):
            senior_debt= self.financing['senior_debt']['value']
            setattr(self, 'senior_debt', senior_debt)

        if not hasattr(self, 'equity'):    
            equity=self.financing['equity']['value']
            setattr(self, 'equity', equity)

        if not hasattr(self, 'sales_tax'):    
            sales_tax= self.taxes['sales_tax']['value']
            setattr(self, 'sales_tax', sales_tax)       
        
        if not hasattr(self, 'corporate_income_tax'):
            corporate_income_tax= self.taxes['corporate_income_tax']['value']        
            setattr(self, 'corporate_income_tax', corporate_income_tax)

       
         #Salaries & wages
        if not hasattr(self, 'monthly_wage_per_worker'): 
            mwpw=self.cost_real['monthly_wage_per_worker']['value']
            setattr(self, 'monthly_wage_per_worker', mwpw)
        
        if not hasattr(self, 'monthly_wage_per_supervisor'):
            mwpst=self.cost_real['monthly_wage_per_supervisor']['value']
            setattr(self, 'monthly_wage_per_supervisor', mwpst)
        
        if not hasattr(self, 'annual_increase_salaries_workers'): 
            aisw=self.cost_real['annual_increase_salaries_workers']['value']
            setattr(self, 'annual_increase_salaries_workers', aisw)
        
        if not hasattr(self, 'annual_increase_salaries_supervisors_technicians'):
            aispt=self.cost_real['annual_increase_salaries_supervisors_technicians']['value']
            setattr(self, 'annual_increase_salaries_supervisors_technicians', aispt)

        if not hasattr(self, 'num_workers_per_supervisor'):
            nowps= 5 # num of workers per supervisor
            setattr(self, 'num_workers_per_supervisor', nowps)
        
        if not hasattr(self, 'num_of_workers_per_pen'):
            nowpp=self.cost_real['num_of_workers_per_pen']['value']
            setattr(self, 'num_of_workers_per_pen', nowpp)
        #--2.a Dometstic price index
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', dpi_list)
        #print('dpi_list',dpi_list )
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', uspi_list)
        # print('us_inflation_rate',us_inflation_rate )
        # print('uspi_list',uspi_list )
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', rpi_list)
        #print('rpi_list',rpi_list )
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)
       # print('ner_list', ner_list)           



        
       
        #1.----------------Time Period----------------------          
        cost_of_domestic_inputs_per_ton_of_beef_produced=self._get_cost_of_domestic_inputs()
        setattr(self, 'cost_of_domestic_inputs_per_ton_of_beef_produced', cost_of_domestic_inputs_per_ton_of_beef_produced)

        #--2.a Dometstic price index [possibility of dynamic inflation]
        dpi_list= list(self._accumulate_price_index(self.tp_list,self.domestic_inflation_rate))
        setattr(self, 'dpi_list', dpi_list)
        
        #2.b US Price Index
        uspi_list= list(self._accumulate_price_index(self.tp_list,self.us_inflation_rate))
        setattr(self, 'uspi_list', uspi_list)
       
        #4.a.3 Operating period         
        op_list=[1 if x >= self.start_year and x < self.end_year else 0 for x in self.tp_list]#dynamic later
        setattr(self, 'op_list', op_list)

        #2.c Relative Price Index
        rpi_list =[0 if j==0 else i/j for i,j in zip(self.dpi_list,uspi_list)]
        setattr(self, 'rpi_list', rpi_list)
        
        #3 Nominal Exchange Rate
        ner_list =[i*self.exchange_rate for i in self.rpi_list]
        setattr(self, 'ner_list', ner_list)

      
    def _metric_investmentCost(self):
        initial_pens_employed= self.investment_parameter_options['initial_pens_employed'][0]
        #dynamic var
        setattr(self, 'initial_pens_employed', initial_pens_employed)
        initial_pens_constructed_list=[self.initial_pens_employed if x==0 else 0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'initial_pens_constructed_list', initial_pens_constructed_list)
        #print('initial_pens_constructed_list', initial_pens_constructed_list)
        
        # Cost of unit pen constructed 
        cost_of_unit_pen= self.investment_cost['cost_of_pen_construction']['value']
        #dynamic var for sensitivity
        setattr(self, 'cost_of_unit_pen', cost_of_unit_pen)
        
        #1.2 Investment cost of Feedlots (nominal)
  
        # Cost of feedlots constructed       
        #Formulae = [No. of feedlot(pens) targeted for construction] * [Cost of Unit Pen Construction]/1000         
        cost_feedlots_nominal_list= [x*y*self.cost_of_unit_pen/1000  for x, y in zip(self.initial_pens_constructed_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_nominal_list', cost_feedlots_nominal_list)
     
        cost_feedlots_real_list= [0 if y==0 else  x/y  for x, y in zip(self.cost_feedlots_nominal_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_real_list', cost_feedlots_real_list)

        #1.3 Investment cost of land
              
     
        feedlot_sqm=self.initial_pens_employed*self.pen_length*self.pen_width 

        
        # sqm_per_cattle = feedlot_sqm/self.cattle_pen_cycle  if self.cattle_pen_cycle!=0 else 0
        # cattle_per_pen_per_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 

    

        land_for_pens= feedlot_sqm
        #self.cattle_business_options['']       
        #first option choosen by default
        option_land_requirement = self.investment_parameter_options['total_land_required'][0]
        setattr(self, 'option_land_requirement', option_land_requirement)

        land_required= max(land_for_pens,option_land_requirement)
        
        
        cost_of_land_per_sqm=self.investment_parameter_options['cost_of_land_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_land_per_sqm', cost_of_land_per_sqm)
         
        #Investment Roll out Flag
        # use initail pen under/ targeted for construction
        inv_roll= list(self._accumulate_investment_rollout_flag(self.initial_pens_constructed_list))
        setattr(self, 'invest_rollout_flag_list', inv_roll)
        

        cost_of_land_real_list= [x*land_required*self.cost_of_land_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_land_real_list', cost_of_land_real_list)
        
        cost_of_land_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_land_real_list)]
        setattr(self, 'cost_of_land_nominal_list', cost_of_land_nominal_list)
    
    
       #1.4 Investment cost of Machinery

        #Import duty
        import_duty=self.taxes['import_duty']['value']
        #dynamic variable
        setattr(self, 'import_duty', import_duty)



        inv_cost_over_run= self.investment_cost['investment_costs_over_run_factor']['value']
        setattr(self, 'inv_cost_over_run', inv_cost_over_run)
        

      
        cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
        cif_machinery_list= [x*self.cost_of_machinery_per_pen/1000  for x in self.initial_pens_constructed_list]
        setattr(self, 'cif_machinery_list', cif_machinery_list)      

        cif_machinery_with_overrun_list= [x*(1+self.inv_cost_over_run)  for x in self.cif_machinery_list]
        setattr(self, 'cif_machinery_with_overrun_list', cif_machinery_with_overrun_list)
       
        cif_machinery_with_overrun_lc_list= [x*y  for x,y  in zip(self.ner_list,self.cif_machinery_with_overrun_list)]
        setattr(self, 'cif_machinery_with_overrun_lc_list', cif_machinery_with_overrun_lc_list)
        
        #Total Cost of Machinery including import duty (real)
        total_machinery_incl_import_duty_real_list= [x*(1+self.import_duty)  for x  in cif_machinery_with_overrun_lc_list]
        setattr(self, 'total_machinery_incl_import_duty_real_list', total_machinery_incl_import_duty_real_list)
       
        total_machinery_incl_import_duty_nominal_list= [x*y  for x,y  in zip(self.dpi_list,total_machinery_incl_import_duty_real_list)]
        setattr(self, 'total_machinery_incl_import_duty_nominal_list', total_machinery_incl_import_duty_nominal_list)
       

        # Investment building
        cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        #cost of building as a function of pens constructed
        cost_of_building_real_list= [x*land_for_pens*self.cost_of_building_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_building_real_list', cost_of_building_real_list)

        cost_of_building_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_building_real_list)]
        setattr(self, 'cost_of_building_nominal_list', cost_of_building_nominal_list)
    
        #aaaaaaa
         
        # Total Investment Cost (nominal)		000's LC
        # Equity Contribution towards Total Investment Costs		000's LC
        # Senior Debt Contribution towards Total Investment Costs		000's LC
        total_invest_cost= np.zeros(len(self.tp_list))
       
        total_invest_cost= np.array(self.cost_feedlots_nominal_list) + \
                           np.array(self.cost_of_land_nominal_list)+ \
                          np.array(self.total_machinery_incl_import_duty_nominal_list) + \
                          np.array(self.cost_of_building_nominal_list)
        
       
       
      

        equity_contr_investments= self.equity*total_invest_cost
        senior_debt_contr_investments= self.senior_debt*total_invest_cost
        setattr(self, 'senior_debt_contr_investments', senior_debt_contr_investments)
    
      
      
    def _simulation_investmentCost(self):

        
        # if not hasattr(self, 'initial_pens_employed'):
        #     initial_pens_employed= self.investment_parameter_options['initial_pens_employed'][0]
        #     #dynamic var
        #     setattr(self, 'initial_pens_employed', initial_pens_employed)
       
        if not hasattr(self, 'cost_of_unit_pen'):
            # Cost of unit pen constructed 
            cost_of_unit_pen= self.investment_cost['cost_of_pen_construction']['value']
            setattr(self, 'cost_of_unit_pen', cost_of_unit_pen)
        
           
        if not hasattr(self, 'option_land_requirement'):
            option_land_requirement = self.investment_parameter_options['total_land_required'][0]
            setattr(self, 'option_land_requirement', option_land_requirement)

        
        if not hasattr(self, 'cost_of_land_per_sqm'):
            cost_of_land_per_sqm=self.investment_parameter_options['cost_of_land_per_sqm'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_land_per_sqm', cost_of_land_per_sqm)
         
        if not hasattr(self, 'import_duty'):
            #Import duty
            import_duty=self.taxes['import_duty']['value']
            #dynamic variable
            setattr(self, 'import_duty', import_duty)
            #print('import_duty',import_duty)


        if not hasattr(self, 'inv_cost_over_run'):
            inv_cost_over_run= self.investment_cost['investment_costs_over_run_factor']['value']
            setattr(self, 'inv_cost_over_run', inv_cost_over_run)
        

        if not hasattr(self, 'cost_of_machinery_per_pen'):
            cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
     
        if not hasattr(self, 'cost_of_building_per_sqm'):
            # Investment building
            cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
            #dynmic in simulation remove above
            setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        
        
        
        
        
        
        
        
        
        
        
        #------------------Up is new edit pliz-------------------------------
        initial_pens_constructed_list=[self.initial_pens_employed if x==0 else 0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'initial_pens_constructed_list', initial_pens_constructed_list)
        #print('initial_pens_constructed_list', initial_pens_constructed_list)  
        #1.2 Investment cost of Feedlots (nominal)
  
        # Cost of feedlots constructed       
        #Formulae = [No. of feedlot(pens) targeted for construction] * [Cost of Unit Pen Construction]/1000         
        cost_feedlots_nominal_list= [x*y*self.cost_of_unit_pen/1000  for x, y in zip(self.initial_pens_constructed_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_nominal_list', cost_feedlots_nominal_list)

        cost_feedlots_real_list= [0 if y==0 else  x/y for x, y in zip(self.cost_feedlots_nominal_list,self.dpi_list)]
        setattr(self, 'cost_feedlots_real_list', cost_feedlots_real_list)
     
       
        #1.3 Investment cost of land        
         
        feedlot_sqm=self.initial_pens_employed*self.pen_length*self.pen_width 

        
        sqm_per_cattle = feedlot_sqm/self.cattle_pen_cycle  if self.cattle_pen_cycle!=0 else 0
        cattle_pen_year=  self.cattle_pen_cycle*12/self.months_per_cycle  if self.months_per_cycle!=0 else 0 
        
        land_for_pens= feedlot_sqm

        land_required= max(land_for_pens,self.option_land_requirement)


          
        #Investment Roll out Flag
        # use initail pen under/ targeted for construction
        inv_roll= list(self._accumulate_investment_rollout_flag(self.initial_pens_constructed_list))
        setattr(self, 'invest_rollout_flag_list', inv_roll)
        

        cost_of_land_real_list= [x*land_required*self.cost_of_land_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_land_real_list', cost_of_land_real_list)
        
        cost_of_land_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_land_real_list)]
        setattr(self, 'cost_of_land_nominal_list', cost_of_land_nominal_list)
    
    
       #1.4 Investment cost of Machinery

      
        inv_cost_over_run= self.investment_cost['investment_costs_over_run_factor']['value']
        setattr(self, 'inv_cost_over_run', inv_cost_over_run)
        

      
        cost_of_machinery_per_pen=self.investment_parameter_options['cost_of_machinery_per_pen'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_machinery_per_pen', cost_of_machinery_per_pen)
      
        cif_machinery_list= [x*self.cost_of_machinery_per_pen/1000  for x in self.initial_pens_constructed_list]
        setattr(self, 'cif_machinery_list', cif_machinery_list)    
        #print('Inspect?   cif_machinery_list',cif_machinery_list)  

        cif_machinery_with_overrun_list= [x*(1+self.inv_cost_over_run)  for x in self.cif_machinery_list]
        setattr(self, 'cif_machinery_with_overrun_list', cif_machinery_with_overrun_list)
       
        cif_machinery_with_overrun_lc_list= [x*y  for x,y  in zip(self.ner_list,self.cif_machinery_with_overrun_list)]
        setattr(self, 'cif_machinery_with_overrun_lc_list', cif_machinery_with_overrun_lc_list)
        
        #Total Cost of Machinery including import duty (real)
        total_machinery_incl_import_duty_real_list= [x*(1+self.import_duty)  for x  in cif_machinery_with_overrun_lc_list]
        setattr(self, 'total_machinery_incl_import_duty_real_list', total_machinery_incl_import_duty_real_list)
       
        total_machinery_incl_import_duty_nominal_list= [x*y  for x,y  in zip(self.dpi_list,total_machinery_incl_import_duty_real_list)]
        setattr(self, 'total_machinery_incl_import_duty_nominal_list', total_machinery_incl_import_duty_nominal_list)
       

        # Investment  of buildings
        cost_of_building_per_sqm=self.investment_parameter_options['cost_of_building_per_sqm'][0]
        #dynmic in simulation remove above
        setattr(self, 'cost_of_building_per_sqm', cost_of_building_per_sqm)

        #cost of building as a function of pens constructed
        cost_of_building_real_list= [x*land_for_pens*self.cost_of_building_per_sqm/1000  for x in self.invest_rollout_flag_list]
        setattr(self, 'cost_of_building_real_list', cost_of_building_real_list)

        cost_of_building_nominal_list= [x*y  for x,y  in zip(self.dpi_list,cost_of_building_real_list)]
        setattr(self, 'cost_of_building_nominal_list', cost_of_building_nominal_list)
    
        #aaaaaaa
         
        # Total Investment Cost (nominal)		000's LC
        # Equity Contribution towards Total Investment Costs		000's LC
        # Senior Debt Contribution towards Total Investment Costs		000's LC
        total_invest_cost= np.zeros(len(self.tp_list))
       
        total_invest_cost= np.array(self.cost_feedlots_nominal_list) + \
                           np.array(self.cost_of_land_nominal_list)+ \
                          np.array(self.total_machinery_incl_import_duty_nominal_list) + \
                          np.array(self.cost_of_building_nominal_list)
        
        setattr(self, 'total_invest_cost', total_invest_cost)
    
        # senior_debt=self.financing['senior_debt']['value']
        # equity=self.financing['equity']['value']
     

        equity_contr_investments= self.equity*total_invest_cost
        senior_debt_contr_investments= self.senior_debt*total_invest_cost
        setattr(self, 'senior_debt_contr_investments', senior_debt_contr_investments)
    
        # there is no update of loans
        #print('inspect? senior_debt_contr_investments',senior_debt_contr_investments)

        self.investment_cost['investment_cost_of_buildings']['value']
        self.investment_cost['cif_cost_of_machinery']['value']
        self.investment_cost['investment_cost_of_land']['value'] 
        self.investment_parameter_options['cost_of_pen_construction'][0]
        # print("**************************************************************************")
        # print('TimePeriod:', self.tp_list)
        #print('roll-out flag: ', self.invest_rollout_flag_list)

        #print('cost_of_machinery_per_pen', cost_of_machinery_per_pen)
        # print('initial_pens_constructed_list', self.initial_pens_constructed_list)
        
        #print('Cost of feedlots investment: ', self.cost_feedlots_nominal_list)
        #print('num_feedlots', self.initial_pens_employed)
        # print('fedlot_length_m', self.pen_length)
        # print('fedlot_width_m', self.pen_width)
        #print('land_required', land_required)
      
    
    def _metric_residualValues(self):
        #Initial total values
        #1 Total cost of land (real)
        total_cost_land_real =np.sum(self.cost_of_land_real_list)         
        setattr(self, 'total_cost_land_real', total_cost_land_real)

        #2 Machinery (real)
        #---fuurther research if necessary to combine cost of feedlots into machinery
        #in calclaying thr residual value
        total_cost_machinery_real =np.sum(self.total_machinery_incl_import_duty_real_list) + \
                 np.sum(self.cost_feedlots_real_list)        
        setattr(self, 'total_cost_machinery_real', total_cost_machinery_real)

        #3 Building (real)
        total_cost_buidings_real =np.sum(self.cost_of_building_real_list)
        setattr(self, 'total_cost_buidings_real', total_cost_buidings_real)

        # Economic service life					
        # 			Economic life of machinery	 30 	Years
        # 			Economic life of buildings	 30 	Years
        economic_life_of_machinery= self.depreciation['economic_life_of_machinery']['value']
        setattr(self, 'economic_life_of_machinery', economic_life_of_machinery)
        economic_life_of_buildings= self.depreciation['economic_life_of_buildings']['value']
        setattr(self, 'economic_life_of_buildings', economic_life_of_buildings)
        tax_life_of_machinery=self.depreciation['tax_life_of_machinery']['value']
        setattr(self, 'tax_life_of_machinery', tax_life_of_machinery)
        tax_life_of_buildings=self.depreciation['tax_life_of_buildings']['value']
        setattr(self, 'tax_life_of_buildings', tax_life_of_buildings)
        tax_life_of_soft_capital_costs=self.depreciation['tax_life_of_soft_capital_costs']['value']
        setattr(self, 'tax_life_of_soft_capital_costs', tax_life_of_soft_capital_costs)



        # Depreciation (Real)			
		# 	Operation period

        #Formale=Total*Operation period/Economic Life
		# 	Machinery
        machinery_const=self.total_cost_machinery_real/self.economic_life_of_machinery if self.economic_life_of_machinery!=0 else 0
        depreciation_machinery_list=[machinery_const*x  for x in self.op_list]
        setattr(self, 'depreciation_machinery_list', depreciation_machinery_list)

        
		# 	Buildings
        buildings_const=self.total_cost_buidings_real/self.economic_life_of_buildings if self.economic_life_of_buildings!=0 else 0
        depreciation_buildings_list=[buildings_const*x  for x in self.op_list]
        setattr(self, 'depreciation_buildings_list', depreciation_buildings_list)

        # Residual value (Real)
        # #-----------------------------------------------------------------------------			
		# 	Residual value of land (real)
        residual_val_land_real_list=[0 for x in self.tp_list]
        residual_val_land_real_list[-1]=self.total_cost_land_real
        setattr(self, 'residual_val_land_real_list', residual_val_land_real_list)
        
        # 	Residual value of Machinery (real)
        residual_val_machinery_real_list=[0 for x in self.tp_list]
        residual_val_machinery_real_list[-1]=self.total_cost_machinery_real
        setattr(self, 'residual_val_machinery_real_list', residual_val_machinery_real_list)
        
        # 	Residual value of Buildings (real)
        residual_val_buildings_real_list=[0 for x in self.tp_list]
        residual_val_buildings_real_list[-1]=self.total_cost_buidings_real
        setattr(self, 'residual_val_buildings_real_list', residual_val_buildings_real_list)
		
         # Residual value (Nominal)	
         # #--------------------------------------------------------------------------------------------		
		
        # 	Residual value of land (Nominal)
        residual_val_land_nominal_list=[x*y for x,y in zip(self.dpi_list,self.residual_val_land_real_list)]
        setattr(self, 'residual_val_land_nominal_list', residual_val_land_nominal_list)
        
        # 	Residual value of Machinery (Nominal)
        residual_val_machinery_nominal_list=[x*y for x,y in zip(self.dpi_list, self.residual_val_machinery_real_list)]
        setattr(self, 'residual_val_machinery_nominal_list', residual_val_machinery_nominal_list)
        
        # 	Residual value of Buildings (Nominal)
        residual_val_buildings_nominal_list=[x*y for x,y in zip(self.dpi_list,self.residual_val_buildings_real_list)]
        setattr(self, 'residual_val_buildings_nominal_list', residual_val_buildings_nominal_list)
        

        
   
    def _metric_productionAndSales(self) :
        #4 Gross Sales Revenue**************************************************
        #4.a Productin Quantity--------------------------------------------
        
     
        #4.a.2 Cumulative pens under Cattle
        cum_feedlots_list= list(self._accumulate_addition(self.initial_pens_constructed_list))
        setattr(self, 'cum_feedlots_list', cum_feedlots_list)
        #print('cum_feedlots_list', cum_feedlots_list)
        
        #4.a.3 Closing Inventory
        ci_list=[0 for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'ci_list', ci_list)

       
        # Production Quantity
        yearly_tons_production_per_pen_const= self.cattle_per_pen_per_year*self.cattle_survival_rate*self.dressed_weight_at_selling/1000
        #set dynamic variable
        # print('cattle_per_pen_per_year', self.cattle_per_pen_per_year)
        # print('cattle_survival_rate', self.cattle_survival_rate)
        # print('dressed_weight_at_selling', self.dressed_weight_at_selling)
        # print('yearly_tons_production_per_pen_const', yearly_tons_production_per_pen_const)
       
        prod_qnty_list=[yearly_tons_production_per_pen_const*x*y  for x,y in zip(self.op_list,cum_feedlots_list)]#dynamic later
        setattr(self, 'prod_qnty_list', prod_qnty_list)
        #print('prod_qnty_list', prod_qnty_list)

        #4.b Sales Revenue

        #4.b. Sales Qnty  
        sales_qnty_list= list(self._accumulate_sales_qnty(self.ci_list, self.prod_qnty_list))
        setattr(self, 'sales_qnty_list', sales_qnty_list)

       
        
        #Real Price of Beef per ton        
        rpobpt_list=[float(1+self.change_in_price) * float(self.base_price) for x in range(len(self.tp_list))]#dynamic later
        setattr(self, 'rpobpt_list', rpobpt_list)
        
        
        #Nominal Price of Beef per ton
        npbpt_list=[x*y  for x,y in zip(self.dpi_list,rpobpt_list)]
        setattr(self, 'npbpt_list', npbpt_list)
        

        #4.d. Net Sales Revenue 
        net_sales_revenue_list=[x*y/1000  for x,y in zip(sales_qnty_list,npbpt_list)]
        setattr(self, 'net_sales_revenue_list', net_sales_revenue_list)

        #4.e. Gross sales revenue  
        #Sales tax
       
        #sale tax paid
        st_list=[float(self.sales_tax) * float(x) for x in self.net_sales_revenue_list]
        setattr(self, 'st_list', st_list)

        
        # Gross sales revenue 
        gsr_list=[x+ y  for x,y in zip(self.net_sales_revenue_list,st_list)]
        setattr(self, 'gsr_list', gsr_list)


    
    def _metric_labourCost(self):
        #6. LABOUR COSTS AND OTHER INDIRECT COSTS (Nominal)
        #Direct Labour
        #aaaaa
        #Salaries
        mwpw=self.monthly_wage_per_worker

        mwpst=self.monthly_wage_per_supervisor
        #Annual Increase in real salaries
        #Annual Increase Salaries  
        aisw=self.annual_increase_salaries_workers
        aispt=self.annual_increase_salaries_supervisors_technicians

        nowps= self.num_workers_per_supervisor

        #Number of workers per pen
       
        nowpp= self.num_of_workers_per_pen
        
        #Numner of workers = worker per per * pens per period
        now_list=[nowpp*x for x in self.cum_feedlots_list]
        setattr(self, 'now_list', now_list)
        # Number of technicians and supervisors
        #self.cost_real['num_of_supervisors_technicians']['value']
        #Num of supervisors and techniciam
       
        nost_list=[math.floor(x/nowps) if nowps!=0 else 0 for x in now_list]
        setattr(self, 'nost_list', nost_list)

        #Real yearly wage rate        
        rywr_list= list(self._accumulate_salary_rate(self.tp_list,mwpw,aisw))

        setattr(self, 'rywr_list', rywr_list)

        #Nominal yearly wage rate        
        nywr_list =[i*j for i,j in zip(self.dpi_list,rywr_list)]
        setattr(self, 'nywr_list', nywr_list)

        #Total Direct Labour Cost
        tdlc_list= [i*j for i,j in zip(now_list,nywr_list)]
        setattr(self, 'tdlc_list', tdlc_list)
        
        #Real yearly salary rate
        rysr_list= list(self._accumulate_salary_rate(self.tp_list,mwpst,aispt))
        setattr(self, 'rysr_list', rysr_list)
       
        #Nominal yearly salary rate        
        nysr_list =[i*j for i,j in zip(self.dpi_list,rysr_list)]
        setattr(self, 'nysr_list', nysr_list)
        
        #Total Indirect Labour Cost
        total_indirect_labour_list= [i*j for i,j in zip(nost_list,nysr_list)]
        setattr(self, 'total_indirect_labour_list', total_indirect_labour_list)

        #Total  Labour Cost
        tlc_list= [i+j for i,j in zip(tdlc_list,total_indirect_labour_list)]
        setattr(self, 'tlc_list', tlc_list)

        # Operation period		flag
        # Cost Of Electricity per pen per annum	 100.0 	LC
        coepppa= self.cost_of_electricity_per_pen_per_annum
        cost_of_electricity_list =[x*y*z*coepppa/1000 for x,y,z in zip(self.cum_feedlots_list,self.op_list,self.dpi_list)]
        setattr(self, 'cost_of_electricity_list', cost_of_electricity_list)
        # Cost of electricity (nominal)		000's LC
        # Other Indirect Costs	 1.0 	000's LC
        # Other Indirect Costs (nominal)		000's LC
        other_indirect_costs_list =[x*y*self.other_indirect_costs/1000 for x,y in zip(self.op_list,self.dpi_list)]
        setattr(self, 'other_indirect_costs_list', other_indirect_costs_list)
     
   
       
    def _metric_purchases(self):
        #7. PURCHASES (Nominal)
        
             
        #Real CIF Cost of imported inputs per ton
        rcifcoiipt_list= list(self._accumulate_yearly_rate(self.tp_list,
                                self.cost_of_imported_inputs_per_ton_of_beef_produced,
                                    self.annual_change_in_price_of_imported_inputs))
        setattr(self, 'rcifcoiipt_list', rcifcoiipt_list)

        #Nominal CIF Cost of imported inputs per ton
        ncifcoiipt_usd_list= [i*j for i,j in zip(self.uspi_list,rcifcoiipt_list)]
        setattr(self, 'ncifcoiipt_usd_list', ncifcoiipt_usd_list)

        #Nominal CIF Cost of imported inputs per ton
        ncifcoiipt_lc_list= [i*j for i,j in zip(self.ner_list,ncifcoiipt_usd_list)]
        setattr(self, 'ncifcoiipt_lc_list', ncifcoiipt_lc_list)
      
        #Cost of imported inputs including import duty per ton
        ncifcoiipt_lc_id_list= [x*(1+self.import_duty) for x in ncifcoiipt_lc_list]
        setattr(self, 'ncifcoiipt_lc_id_list', ncifcoiipt_lc_id_list)
        # print('ncifcoiipt_lc_list', ncifcoiipt_lc_list)
        # print('ncifcoiipt_lc_id_list', ncifcoiipt_lc_id_list)
        #Total cost of imported inputs including import duty
        tcoiiiid_list= [i*j/1000 for i,j in zip(ncifcoiipt_lc_id_list,self.prod_qnty_list)]
        setattr(self, 'tcoiiiid_list', tcoiiiid_list)
        # print('tcoiiiid_list', tcoiiiid_list)

        
        #Real Cost of domestic inputs per ton
        rcodipt_list= list(self._accumulate_yearly_rate(self.tp_list,
                         self.cost_of_domestic_inputs_per_ton_of_beef_produced,
                             self.annual_change_in_price_of_domestic_inputs))

        setattr(self, 'rcodipt_list', rcodipt_list)
        print('ncodipt_list',rcodipt_list)
        #Nominal Cost of domestic inputs per ton
        ncodipt_list= [i*j for i,j in zip(self.dpi_list,rcodipt_list)]
        setattr(self, 'ncodipt_list', ncodipt_list)
     
        #Total cost of domestic inputs
        tcodi_list= [i*j/1000 for i,j in zip(ncodipt_list,self.prod_qnty_list)]
        setattr(self, 'tcodi_list', tcodi_list)
        # print('ncodipt_list',ncodipt_list)
        print('Total cost of domestic inputs',tcodi_list)

        #Total cost of inputs per ton (nominal)
        inputs_cost_per_ton_nominal_list= [i+j for i,j in zip(ncifcoiipt_lc_id_list,ncodipt_list)]
        setattr(self, 'inputs_cost_per_ton_nominal_list', inputs_cost_per_ton_nominal_list)
        
        #Total Input Cost (nominal)
        tic_n_list= [i+j for i,j in zip(tcoiiiid_list,tcodi_list)]
        setattr(self, 'tic_n_list', tic_n_list)
      
    def _metric_workingCapital(self):
       

        #Accounts receivable [% of Gross Sales]
        arogs_list= [x*self.accounts_receivable for x in self.gsr_list]
        setattr(self, 'arogs_list', arogs_list)
     
        #Acccounts payable (% of total input cost)
        apotic_list= [x*self.accounts_payable for x in self.tic_n_list]
        setattr(self, 'apotic_list', apotic_list)
     
        #Cash balance  (% of gross sales)
        cbogs_list= [x*self.cash_balance for x in self.gsr_list]
        setattr(self, 'cbogs_list', cbogs_list)
     
        #Change in A/R
        ciar_list= list(self._accumulate_change_in(self.arogs_list))
        setattr(self, 'ciar_list', ciar_list)
        
        #Change in A/P
        ciap_list= list(self._accumulate_change_in(self.apotic_list))
        setattr(self, 'ciap_list', ciap_list)
        
        #Change in cash balance
        cicb_list= list(self._accumulate_change_in_CB(self.cbogs_list))
        setattr(self, 'cicb_list', cicb_list)
    def _metric_loanSchedule(self):
       
        
      
        #US inflation
        us_inflation_rate_list= [self.us_inflation_rate for x in self.tp_list]
        setattr(self, 'us_inflation_rate_list', us_inflation_rate_list)
        nominal_interest_rate_list= list(self._accumulate_nominal_interest_rate(self.us_inflation_rate_list,
                                    self.real_interest_rate,self.risk_premium))
        setattr(self, 'nominal_interest_rate_list', nominal_interest_rate_list)

        #Domestic
        domestic_inflation_rate_list= [self.domestic_inflation_rate for x in self.tp_list]
        setattr(self, 'domestic_inflation_rate_list', domestic_inflation_rate_list)
        nominal_interest_rate_lc_list= list(self._accumulate_nominal_interest_rate_lc(self.domestic_inflation_rate_list,
                                    self.discount_rate_equity))
        setattr(self, 'nominal_interest_rate_lc_list', nominal_interest_rate_lc_list)

        upper_bound= (self.repayment_starts + self.num_of_installments - self.grace_period-1)
        #upper bound questionable effect of grace period..........

        loan_principal_repayment_flag_list=[0 if x < self.start_year or x > upper_bound else 1 for x in self.tp_list]#dynamic later
        setattr(self, 'loan_principal_repayment_flag_list', loan_principal_repayment_flag_list)
        
        construction_period_flag_list=[0 if x > self.construction_year_end  else 1 for x in self.tp_list]#dynamic later
        setattr(self, 'construction_period_flag_list', construction_period_flag_list)
       
        # self.financing['equity']['value']

        # self.financing['senior_debt']['value']
      
        # OPENING INVENTORY			
		# 	Real interest rate
		# 	Risk premium
		# 	US Inflation rate
		# 	Nominal Interest Rate
		# 	Local Inflation rate
		# 	Nominal Interest Rate (Local)
			
        # SUPPLIERS CREDIT			
        #             US Price Index
        #             No. of installments
        #             Loan principle repayment

       
        # Loan Repayment Schedule			
		# 	Beginning Debt
           # 	Loan disbursement
       
        loan_disbursement_list= [x for x in self.senior_debt_contr_investments]
        setattr(self, 'loan_disbursement_list', loan_disbursement_list)
        total_loan_disbursed =np.sum(self.loan_disbursement_list)

        principal_paid_list= [total_loan_disbursed*x/self.num_of_installments if self.num_of_installments !=0 else 0 for x in self.loan_principal_repayment_flag_list]
        setattr(self, 'principal_paid_list', principal_paid_list)
        beggining_debt_list= list(self._accumulate_beginning_debt(self.loan_disbursement_list, 
                                                                self.principal_paid_list))
        
        setattr(self, 'beggining_debt_list', beggining_debt_list)
        # print('nominal_interest_rate_lc_list', nominal_interest_rate_lc_list)  
        # print('loan_disbursement_list', loan_disbursement_list) 
        # print('principal_paid_list', principal_paid_list) 
        # print('beggining_debt_list', beggining_debt_list) 


		# 	Interest accrued in year
        interest_accrued_list =[x*y for x,y in zip(self.beggining_debt_list,self.nominal_interest_rate_lc_list)]
        setattr(self, 'interest_accrued_list', interest_accrued_list)
		# 	Principal paid
		# 	Interest paid
        interest_paid_list =[x for x in interest_accrued_list]
        setattr(self, 'interest_paid_list', interest_paid_list)

		# 	Total Loan Repayment
        total_loan_repayment_list =[x+y for x,y in zip(self.interest_accrued_list,self.principal_paid_list)]
        setattr(self, 'total_loan_repayment_list', total_loan_repayment_list)

		# 	Outstanding debt at end of year
        outstanding_debt_list= np.array(self.beggining_debt_list) + \
                              np.array(self.loan_disbursement_list) + \
                              np.array(self.interest_accrued_list) - \
                              np.array(self.total_loan_repayment_list) 
        setattr(self, 'outstanding_debt_list', outstanding_debt_list)	
        #print('outstanding_debt_list', outstanding_debt_list)
		# 	Debt cash Inflow in Local Currency
        debt_cash_inflow_lc_list=[x for x in self.loan_disbursement_list]
        setattr(self, 'debt_cash_inflow_lc_list', debt_cash_inflow_lc_list)

		# 	Total Loan Repayment as an outflow in Local Currency
        total_loan_repayment_ouflow_lc_list=[x for x in self.total_loan_repayment_list]
        setattr(self, 'total_loan_repayment_ouflow_lc_list', total_loan_repayment_ouflow_lc_list)
        # 	Construction period flag
		# 	Operating period
		# 	Interest during Construction, Capitalized for Tax Purposes
        interest_during_construction_list= [x*y for x,y in zip(self.interest_accrued_list,self.construction_period_flag_list)]
        setattr(self, 'interest_during_construction_list', interest_during_construction_list)


       
        total_soft_capital_cost= np.sum(interest_during_construction_list)
        setattr(self, 'total_soft_capital_cost', total_soft_capital_cost)

        
        
        #Tax life of soft capital costs
        const_instalments= total_soft_capital_cost/self.tax_life_of_soft_capital_costs
        const_end= self.start_year + self.tax_life_of_soft_capital_costs-1
        soft_capital_cost_list= [const_instalments*x if (y>=self.start_year and y<=const_end) else 0 for x,y in zip(self.op_list,self.tp_list)]
        setattr(self, 'soft_capital_cost_list', soft_capital_cost_list)
       
        # 	Interest expense (for income statement)
        interest_expense_list= [x*y for x,y in zip(self.interest_accrued_list,self.op_list)]
        setattr(self, 'interest_expense_list', interest_expense_list)
       
       
    def _metric_depreciationForTaxPurposes(self):
        pass
    def _metric_unitCost(self):
        # Production Quantity	
        # Total Direct Labour Cost	
        # Direct Labour cost per ton of beef produced
        direct_labour_per_ton_beef_produced=[0 if y==0 else (x*1000/y)  for x,y in zip(self.tdlc_list,self.prod_qnty_list)]
        setattr(self, 'direct_labour_per_ton_beef_produced', direct_labour_per_ton_beef_produced)
            
        # Operation period	
        # Total cost of inputs per ton (nominal)	
        # Total unit cost of production per ton (nominal)
        total_unit_cost_production_per_ton_list=[(x+y)*z for x,y,z in zip(self.direct_labour_per_ton_beef_produced,
                                                      self.inputs_cost_per_ton_nominal_list,
                                                      self.op_list)]
        setattr(self, 'total_unit_cost_production_per_ton_list', total_unit_cost_production_per_ton_list)
        	

       
    def _metric_finishedProdictEvaluationFIFO(self):
        # OPENING INVENTORY			
        #     Closing Inventory
        closing_inventory_list= [0 for x in self.tp_list]
        setattr(self, 'closing_inventory_list', closing_inventory_list)

        #     Quantity from previous year (closing inventory)
        qnty_prev_closing_inventory_list= list(self._accumulate_previous(self.closing_inventory_list))
        setattr(self, 'qnty_prev_closing_inventory_list', qnty_prev_closing_inventory_list)
        #     Unit cost from previous year per ton (nominal)
        
        # cost_of_opening_inventory_list =list(self._accumulate_previous(self.total_unit_cost_per_ton))
        # setattr(self, 'cost_of_opening_inventory_list', cost_of_opening_inventory_list)
        # #     Cost of opening inventory (using previous year's unit cost FIFO method)
        
        #Formuale =qnty_prev_closing_inventory_list*cost_of_opening_inventory_list
        cost_of_opening_inventory_list=[x*y for x,y in zip(self.qnty_prev_closing_inventory_list,
                                                             self.qnty_prev_closing_inventory_list )]
        setattr(self, 'cost_of_opening_inventory_list', cost_of_opening_inventory_list)           
        # ADDITIONS			
        #             Production Quantity
        #             Total unit cost of production per ton (nominal)
                    
        # WITHDRAWALS			
        #             Sales quantity
        #             Quantity sold from this yr's production
        qnty_sold_this_year_production_list=[x-y for x,y in zip(self.prod_qnty_list,self.qnty_prev_closing_inventory_list )]
        setattr(self, 'qnty_sold_this_year_production_list', qnty_sold_this_year_production_list)
        #             Cost of the proportion of sales produced in current year
        cost_of_proportional_sales=[x*y/1000 for x,y in zip(self.total_unit_cost_production_per_ton_list,
                                                             self.qnty_sold_this_year_production_list )]
        #             COST OF GOODS SOLD (FOR INCOME STATEMENT)
        cost_of_goods_sold_list =np.array(cost_of_opening_inventory_list) + np.array(cost_of_proportional_sales)
        setattr(self, 'cost_of_goods_sold_list', cost_of_goods_sold_list)           
                
        # CLOSING INVENTORY			
        #             Quantity remained (unsold) in year
        qnty_remained_unsold_list= self.qnty_prev_closing_inventory_list + \
                              np.array(self.prod_qnty_list) - \
                              np.array(self.sales_qnty_list) 
        setattr(self, 'qnty_remained_unsold_list', qnty_remained_unsold_list)      
        #             Cost of closing inventory (using current year's unit cost)
        cost_of_closing_inventory_list=np.array(self.total_unit_cost_production_per_ton_list) * \
                              np.array(self.qnty_remained_unsold_list)  
        setattr(self, 'cost_of_closing_inventory_list', cost_of_closing_inventory_list)

        
        
        
     

    def _metric_IncomeTaxStatement(self):
        #Nominal
       
        #1. Depreciation expense			
		# 	Operation period
		# 	Machinery depreciation expense
		# 	Buildings depreciation expense
		# 	Soft Capital Costs (Interest during construction)
        
		# 	Annual Depreciation Expense
        annual_dep_exp_list =np.array(self.depreciation_machinery_list) + \
                            np.array(self.depreciation_buildings_list) + \
                             np.array(self.soft_capital_cost_list) 
        setattr(self, 'annual_dep_exp_list', annual_dep_exp_list)

      

       
        #1. REVENUE
        #---------------------------------------------------------------
        #1.1 Net sales revenue

        #2. LESS OPERATING EXPENSES
        #---------------------------------------------------------------
        #   2.1 Cost of goods sold
                                                            
        #cost_of_goods_sold_list =1
        # Annual Depreciation Expense
        # Total Indirect Labour Cost
        # Cost of electricity (nominal)
        # Other Indirect Costs (nominal)
        #2.1 Total Operating Expenses
        total_operating_expenses_list= np.array(self.cost_of_goods_sold_list) + \
                                  np.array(self.annual_dep_exp_list)  + \
                                  np.array(self.cost_of_electricity_list) +  \
                                  np.array(self.other_indirect_costs_list) +  \
                                  np.array(self.total_indirect_labour_list)    
        
        setattr(self, 'total_operating_expenses_list', total_operating_expenses_list)

        #2.2 INCOME FROM OPERATIONS
        income_from_operations_list =np.array(self.net_sales_revenue_list)-np.array(self.total_operating_expenses_list)
        setattr(self, 'income_from_operations_list', income_from_operations_list)
        # Interest paid

        #2.3 PRE-TAX INCOME
        pre_tax_income_list= np.array(self.income_from_operations_list)- np.array(self.interest_expense_list)
        setattr(self, 'pre_tax_income_list', list(pre_tax_income_list))
        # Cumulative losses
       
        cumulative_losses_list =list(self._accumulate_losess(self.pre_tax_income_list))
        setattr(self, 'cumulative_losses_list', cumulative_losses_list)
        # Taxable income (losses carried forward)
        taxable_income_list =list(self._accumulate_taxable_income(self.pre_tax_income_list,self.cumulative_losses_list))
        setattr(self, 'taxable_income_list', taxable_income_list)
        # Corporate income tax
        # Income tax payment
        income_tax_payment_list =[max(self.corporate_income_tax*x,0) for x in self.taxable_income_list]
        setattr(self, 'income_tax_payment_list', income_tax_payment_list)

        # NET AFTER-TAX INCOME
        next_after_tax_income_list =np.array(self.pre_tax_income_list)-np.array(self.income_tax_payment_list)
        setattr(self, 'next_after_tax_income_list', next_after_tax_income_list)
        #3. EQUITY
        # #----------------------------------------------------------------
       
     
    def _metric_cashFlow(self):
        
        #A. Total Cash Flow (+)
        #-------------------------------------------------------------------------------------
        #A. 1 RECEIPTS
        #----------------------------------------
        #A.1.1 Gross sales revenue
        #A.1.2 Change in A/R
        total_receipts_list =np.zeros(len(self.tp_list))     

        gross_revenus_sales_list=  np.array(self.gsr_list)
        change_in_AR_list=  np.array(self.ciar_list)
        #-----------------------------------------------------------------
        total_receipts_list =gross_revenus_sales_list + change_in_AR_list #+ c -d

        # print('total_receipts_list', total_receipts_list)

        #A.2  Liquidation Value
        total_liquidation_list =np.zeros(len(self.tp_list))
        total_liquidation_list= np.array(self.residual_val_buildings_nominal_list) + \
                                np.array(self.residual_val_machinery_nominal_list) + \
                                np.array(self.residual_val_land_nominal_list)
        
        # print('residual_val_buildings_nominal_list',self.residual_val_buildings_nominal_list)
        # print('residual_val_machinery_nominal_list',self.residual_val_machinery_nominal_list)
        # print('residual_val_land_nominal_list',self.residual_val_land_nominal_list)
        # print('total_liquidation_list',total_liquidation_list)


        #-----------------------------------------------
        #Total Cash_flow (+)
        total_cash_inflow =np.zeros(len(self.tp_list))
        total_cash_inflow= total_receipts_list + total_liquidation_list
        #****************************************************************************************
        #B. Total Cash Out Flow(+)
        #2.EXpenditure
        #--------------------------------------------------------------------------------------
        #  2.1 Investments
        total_invest_cost_list= np.zeros(len(self.tp_list))
        #  2.1.1 Investment cost of Feedlots (nominal)
        feedlot_cost_list= np.array(self.cost_feedlots_nominal_list)
       
        #  2.1.2 Investment Land (nominal)
        cost_of_land_nominal_list= np.array(self.cost_of_land_nominal_list)
       
        #  2.1.3 Investment Machinery (nominal)
        total_machinery_incl_import_duty_nominal_list= np.array(self.total_machinery_incl_import_duty_nominal_list)
       
        #  2.1.4 Investment Building (nominal)
        cost_of_building_nominal_list= np.array(self.cost_of_building_nominal_list)
       
        #Total Investment Cost
        total_invest_cost_list=feedlot_cost_list + cost_of_land_nominal_list + \
                             total_machinery_incl_import_duty_nominal_list +cost_of_building_nominal_list
        
        #----------------------------------------------------------------------------
        #  2.2 Operating cost
        #       2.2.1 Total cost of imported inputs including import duty
        imports_cost=  np.array(self.tcoiiiid_list)
        #print('imports_cost',imports_cost)
        #       2.2.2 Total cost of domestic inputs
        domestic_input_cost =  np.array(self.tcodi_list)
        #print('domestic_input_cost',domestic_input_cost)
      
        #       2.2.3 Total Labour Cost
        tlc_list=  np.array(self.tlc_list) 
        #print('tlc_list',tlc_list)
      
        #       2.2.4 Cost Of Electricity (nominal)
        coe=  np.array(self.cost_of_electricity_list)
        #print('Electricity',coe)
      
        #       2.2.5 Other Indirect Costs (nominal)
        oic=  np.array(self.other_indirect_costs_list)
        #print('Other Indirect Cost',oic)
        #print('other_indirect_costs',self.other_indirect_costs)
        #       Change in A/P
        ciap=  np.array(self.ciap_list)
        #print('AP',ciap)
	    #       Change in cash balance
       
        cicb=  np.array(self.cicb_list)
        #print('CB',cicb)

        #       3.Total Cash Outflow
        total_operating_cost_list= np.zeros(len(self.tp_list))
        total_cashoutflow_list= np.zeros(len(self.tp_list))
        
    
       
       
       
     
        
        total_operating_cost_list=imports_cost + domestic_input_cost + tlc_list + ciap +cicb +coe +oic
        #---------------------------------------------------------------------------------------------
        #print('total_operating_cost_list',total_operating_cost_list)
        
        #Total Outflow
        total_cashoutflow_list= total_operating_cost_list + total_invest_cost_list
        #**************************************************************************************************
        #print('total_cashoutflow_list',total_cashoutflow_list)

        #C. NET CASH FLOW BEFORE TAXES
        net_before_tax_before_financing= total_cash_inflow-total_cashoutflow_list
        #print('net_before_tax_before_financing',net_before_tax_before_financing)

        #*****************************************************************************
       
        #1.1 Sales tax paid
        sales_tax_paid=  np.array(self.st_list)
        #print('sales_tax_paid',sales_tax_paid)
        #1.2 Income tax payment
        income_tax_list=  np.array(self.income_tax_payment_list)
        #print('income_tax_list',income_tax_list)
        #D--- Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)
        net_cash_after_tax_before_financing= net_before_tax_before_financing-sales_tax_paid  -income_tax_list
        # print('net_cash_after_tax_before_financing',net_cash_after_tax_before_financing)
        #*******************************************************************

        #E---Net after tax cash flow after financing (OWNER PERSPECTIVE)
        
        #1.1 Debt cash Inflow in Local Currency
        debt_cash_inflow_lc=  np.array(self.debt_cash_inflow_lc_list)
        #1.2 Total Loan Repayment as an outflow in Local Currency
        total_loan_repayment_lc=  np.array(self.total_loan_repayment_ouflow_lc_list)
        #print('total_loan_repayment_lc',total_loan_repayment_lc)
        net_cash_after_tax_after_financing= net_cash_after_tax_before_financing+ \
                                            debt_cash_inflow_lc - \
                                            total_loan_repayment_lc
        # print('net_cash_after_tax_after_financing',net_cash_after_tax_after_financing)
        #----------------------------------------------------------------------------------

        #PV Annual Net Cash Flows (NCF)
        #PV Annual Debt Repayment
        #Discount Rate Equity
        net_cash_after_tax_after_financing_real=[0 if y==0 else x/y  for x,y in zip(net_cash_after_tax_after_financing,self.dpi_list)]
        #print('net_cash_after_tax_after_financing_real',net_cash_after_tax_after_financing_real)
        #npv_= self.npv(net_cash_after_tax_after_financing_real,self.discount_rate_equity)


       
        #npv2_ =npf.npv(self.discount_rate_equity, net_cash_after_tax_after_financing_real[1:]) + net_cash_after_tax_after_financing_real[0];  
       
        npv2_=self._get_financial_npv(self.discount_rate_equity,net_cash_after_tax_after_financing_real)
        
        #print('npv', npv2_)
        #print('discount_rate_equity', self.discount_rate_equity)
        # print('net_cash_after_tax_after_financing_real',net_cash_after_tax_after_financing_real)
        irr_ =npf.irr(net_cash_after_tax_after_financing_real)
        #print('irr', irr_)
        setattr(self, 'model_npv', npv2_)
        setattr(self, 'irr', irr_)
        """ 
        check this error of float:::::>cannot convert float NaN to integer
        """
        mirr_ =npf.mirr(net_cash_after_tax_after_financing_real,self.discount_rate_equity,self.discount_rate_equity,)#finance_rate, reinvest_rate
        #print('mirr', round(mirr_,2))
        setattr(self, 'mirr', mirr_)
     
        # mirr= []
        """    
        myList = list(np.around(np.array(myList),2))
        myRoundedList =  [round(x,2) for x in mylist]
    
        """

        lb, up=self._get_loan_period_bounds()
        #print('LB --UP: >>>>>', lb, up,self.loan_principal_repayment_flag_list)
        
        # PV Annual Net Cash Flows (NCF)
       
        
        pv_ncf_list=[npf.npv(self.discount_rate_equity, net_cash_after_tax_before_financing[i+1:up+1]) + \
                        net_cash_after_tax_before_financing[i] for i in range(lb,up+1) ] 
        
        
        #print('pv_ncf_list',pv_ncf_list)
        setattr(self, 'pv_ncf_list', pv_ncf_list)
        # print('discount_rate_equity', self.discount_rate_equity)
        pv_annual_debt_repayment_list=[npf.npv(self.discount_rate_equity, self.total_loan_repayment_ouflow_lc_list[i+1:up+1]) + \
                        self.total_loan_repayment_ouflow_lc_list[i] for i in range(lb,up+1) ] 
          
        #print('pv_annual_debt_repayment_list',pv_annual_debt_repayment_list)
        setattr(self, 'pv_annual_debt_repayment_list', pv_annual_debt_repayment_list)
        # PV Annual Debt Repayment
        # ADSCR
        adscr = [0 if y==0 else x/y  for x,y in zip(net_cash_after_tax_before_financing[lb:up+1],
                                              self.total_loan_repayment_ouflow_lc_list[lb:up+1])]
        #print('adscr',adscr)
        setattr(self, 'adscr', adscr)
        # LLCR
        llcr =[0 if y==0 else x/y  for x,y in zip(pv_ncf_list, pv_annual_debt_repayment_list)]
        #print('llcr',llcr)
        setattr(self, 'llcr', llcr)
       
        # Summary of ADSCR
        # #--------------------------------------------		
		# Minimum ADSCR
        adscr_min=np.min(adscr)
        #print('adscr_min',adscr_min)
        setattr(self, 'adscr_min', adscr_min)
		
        # Maximum ADSCR
        adscr_max=np.max(adscr)
        #print('adscr_max',adscr_max)
        setattr(self, 'adscr_max', adscr_max)
       
        # Average ADSCR
        adscr_av=np.average(adscr)
        #print('adscr_av',adscr_av)
        setattr(self, 'adscr_av', adscr_av)

        for i in range(len(adscr)):
            setattr(self, 'adscr_'+ str(i+1), adscr[i])

       
        # Summary of LLCR	
        # #--------------------------------------------	
        # Minimum LLCR
        llcr_min=np.min(llcr)
        #print('llcr_min',llcr_min)
        setattr(self, 'llcr_min', llcr_min)
        
        # Maximum LLCR
        llcr_max=np.max(llcr)
        #print('llcr_max',llcr_max)
        setattr(self, 'llcr_max', llcr_max)
        
        # Average LLCR
        llcr_average=np.average(llcr)
        #print('llcr_av',llcr_average)
        setattr(self, 'llcr_av', llcr_average)
        
        for i in range(len(llcr)):
            setattr(self, 'llcr_'+ str(i+1), llcr[i])


    def _get_financial_npv(self, rate_, cashflows_in):
        """ 
        Excel assumes that the first value is at t=1 (no initial wealth), 
        while numpy-financial assumes the first value to be at t=0. 
        Therefore, In numpy you start with an initial wealth of 4000 and your result is 
        basically the excel result * 1.06. 

        It may be preferable to split the projected cashflow into an initial investment 
        and expected future cashflows. In this case, the value of the initial cashflow 
        is zero and the initial investment is later added to the future cashflows net present value:
        """
        cashflows= cashflows_in.copy()
        initial_cashflow = cashflows[0]
        cashflows[0] = 0
        npv_ =np.round(npf.npv(rate_, cashflows) + initial_cashflow, 5)
        return npv_
    def _get_loan_period_bounds(self):
        i= self.loan_principal_repayment_flag_list.index(1)
        j= self.loan_principal_repayment_flag_list.index(0,i)
        return i, j-1
                 
    def _outputs_npv(self,):
        #run with dynamic vars
        self._simulate_metrix()
        
        return self.model_npv
    
    def _model_outputs_dict(self,):
        #run with dynamic vars
        self._simulate_metrix()
        dict_ ={}

        dict_['npv']=self.model_npv if hasattr(self, 'model_npv') else 0
        dict_['irr']=self.irr if hasattr(self, 'irr') else 0
        dict_['mirr']=self.mirr if hasattr(self, 'mirr') else 0
        dict_['adscr_av']=self.adscr_av if hasattr(self, 'adscr_av') else 0
        dict_['llcr_av']=self.llcr_av if hasattr(self, 'llcr_av') else 0
        return dict_

    def _model_datatable_outputs_dict(self,):
        #run with dynamic vars
        self._simulate_metrix()
        dict_ ={}
        dict_['npv']=self.model_npv if hasattr(self, 'model_npv') else 0
        dict_['adscr-3']=self.adscr_3 if hasattr(self, 'adscr_3') else 0
        dict_['adscr-4']=self.adscr_4 if hasattr(self, 'adscr_4') else 0
        dict_['adscr-5']=self.adscr_5 if hasattr(self, 'adscr_5') else 0
        dict_['llcr-2']=self.llcr_2 if hasattr(self, 'llcr_2') else 0
        dict_['llcr-3']=self.llcr_3 if hasattr(self, 'llcr_3') else 0
        dict_['llcr-4']=self.llcr_4 if hasattr(self, 'llcr_4') else 0
        dict_['llcr-5']=self.llcr_5 if hasattr(self, 'llcr_5') else 0
        dict_['irr']=self.irr if hasattr(self, 'irr') else 0
        dict_['mirr']=self.mirr if hasattr(self, 'mirr') else 0
        return dict_

    
    
    def npv(self, cfList, r):
        sum_pv = 0  # <-- variable used to sum result
        for i, pmt in enumerate(cfList, start=1):  # <-- use of enumerate allows you to do away with the counter variables.
            sum_pv += pmt / ((1 + r) ** i)  # <-- add pv of one of the cash flows to the sum variable

        return sum_pv  # 

    def pv_gen(cfList, r):
        #An alternate implementation would be to yield individual PVs from a PV generator
        for i, pmt in enumerate(cfList, start=1):
            yield pmt / ((1 + r) ** i)

    def pv2(self, cfList, r):
        return (sum(self.pv_gen(cfList, r)))

   
    def _accumulate_investment_rollout_flag(self,list_):
        next_value =0
        for i in range(len(list_)):
            #if nothing and the is vale
            if self._sum_list(list_,0,i)==0 and list_[i]>0:
                next_value= 1
            else:
                next_value =0
            yield next_value

    def _sum_list(self, list_, i, j):
        sum_ =0
        for p in range (i,j):
            sum_ += list_[p]
        return sum_

    def _accumulate_addition(self,list_):
        next_value =0
        for item in list_:
            next_value= next_value + item
            yield next_value
    
    def _accumulate_difference(self,list_):
        next_value =0
        for item in list_:
            next_value= next_value - item
            yield next_value
    
    def _accumulate_previous(self,list_):
        next_value =0
        for i in range(len(list_)):
            if i>0:
                next_value= list_[i-1]
            else:
                next_value =0
            yield next_value

    def _accumulate_beginning_debt(self,loan_disb_list, principal_paid_list):
        begin_debt= 0
        next_value=0
        for i in range(len(loan_disb_list)):
            outstanding_debt= begin_debt + loan_disb_list[i] - principal_paid_list[i]
            next_value=begin_debt
            #update
            begin_debt= outstanding_debt                 

            yield next_value
    def _accumulate_nominal_interest_rate(self, inflation_rate_list, real_interest_rate,risk_premium):
        #i+r+(1+i+r)*x
        #c+(1+c)*x
        const_=real_interest_rate +risk_premium 
        for i in range(len(inflation_rate_list)):
            next_value= const_+(1+const_)*(inflation_rate_list[i]) 
            yield next_value

    def _accumulate_nominal_interest_rate_lc(self, inflation_rate_list, discount_rate_equity):
        #i+r+(1+i+r)*x
        #c+(1+c)*x
        for i in range(len(inflation_rate_list)):
            next_value= discount_rate_equity+(1+discount_rate_equity)*(inflation_rate_list[i]) 
            yield next_value
    def _accumulate_losess(self, pre_tax_income_list):
        next_value =0
        for i in range(len(pre_tax_income_list)):
            next_value= min((pre_tax_income_list[i] + next_value),0) 
            yield next_value

    def _accumulate_taxable_income(self,pre_tax_income_list, cum_losses_list):
        next_value =0
        for i in range(len(pre_tax_income_list)):
            if i>0:
                next_value= max((pre_tax_income_list[i] + cum_losses_list[i-1]),0) 
            else:
                next_value= max(pre_tax_income_list[i], 0)
            yield next_value

      
    def _accumulate_change_in(self,list_):
        next_value =0
        for i in range(len(list_)):
            if i>0:
                next_value= list_[i-1] - list_[i]
            else:
                next_value =-list_[i]
            yield next_value
    def _accumulate_change_in_CB(self,list_):
        #next_value =0
        for i in range(len(list_)):
            if i>0:
                next_value= list_[i]-list_[i-1]
            else:
                next_value =list_[i]
            yield next_value
           
    def _accumulate_price_index(self,list_,inflation_rate):
        next_value =1
        for i in range(len(list_)):
            if i >0:
                next_value= next_value*(inflation_rate+1)
            yield next_value
    
    def _accumulate_salary_rate(self,list_,monthly_salary_rate, annual_incr):
        next_value =12*monthly_salary_rate
        for i in range(len(list_)):
            if i >0:
                #use prev val
                next_value= next_value*(1+annual_incr)
            yield next_value

    def _accumulate_yearly_rate(self,list_,base_cost, annual_incr):
        next_value =base_cost
        for i in range(len(list_)):
            if i >0:
                #use prev val
                next_value= next_value*(1+annual_incr)
            yield next_value
    def _accumulate_sales_qnty(self,list_, sales_qnty_list_):
        next_value=0
        i=int(0)
        k=int(0)
        for a,b in zip(list_, sales_qnty_list_):
            #print(i,k)
            if i > 0:
                next_value= list_[i-1] - list_[i]+ sales_qnty_list_[k]
            else:
                next_value=  sales_qnty_list_[k]- list_[i]
            i +=1
            k +=1
            yield next_value
    def _get_out_values_as_dict(self,wb):
        #list_dict =[]
        output_list =[]
        inputs_sheet = wb['Inputs']
        cal_sheet = wb['Calc']
        output_sheet = wb['Outputs']
        cf_sheet = wb['CF']
        sens_sheet = wb['Sens']

        list_dict =[
            {'wsheet':cal_sheet,'section':'cal_production_sales_nominal','item': 'gross_sales_revenue'},        
        ]
       
        for i in list_dict:
            dict_ ={}            
            dummy=self._get_excel_row_as_list(i['wsheet'],i['section'], i['item'])
            dict_['list']=dummy
            dict_['name']=i['item']
            output_list.append(dict_)
        #print(output_list)

    def _get_excel_row_as_list(self,w_sheet,section_, item):
        r_index, c_index, found_state= self._retrieve_cell_row_colm(section_, item)
        list_ =[]
        if found_state:
            span_ = self.timing_assumptions['operation_duration']['value'] if 'operation_duration' in self.timing_assumptions else 0
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                   value_ =w_sheet['%s%s'%(cell.column, cell.row)].value
                   #print(w_sheet['%s%s'%(cell.column, cell.row+1)].value)
                   list_.append(value_)
        return list_
        #=$F$103*I102
    def _set_Underline_row(self, wkSheet, row):
        
        # Let's create a style template for the header row
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.border = Border(bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="center", vertical="center")

        # Now let's apply this to all first row (header) cells
        header_row = wkSheet[row]
        for cell in header_row:
            cell.style = header

    def _set_thick_bottom_border_range(self, wkSheet, writing_row_index,
                            start_col_index, end_col_index):

        latest_row = writing_row_index

        col_start = get_column_letter(start_col_index)
        col_end = get_column_letter(end_col_index)

        first_slice_point = col_start + str(writing_row_index ) # D200
        second_slice_point = col_end + str(writing_row_index ) # F200

        thick_border_side = Side(border_style="thick", color='FF3F3F76')
        straight_thick_border = Border(
                                bottom=thick_border_side)

        #loop each cell of this row D20:Z20
        #-------->
        for cellObj in wkSheet[first_slice_point:second_slice_point]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
                # |
                # |
                # |
                # v
                cell.border = straight_thick_border

                
                latest_row = cell.row

        #------------------------------
        return latest_row


    def _set_thin_bottom_border_range(self, wkSheet, writing_row_index,
                            start_col_index, end_col_index,color_str='FFFFFF00'):
        

        latest_row = writing_row_index

        col_start = get_column_letter(start_col_index)
        col_end = get_column_letter(end_col_index)

        first_slice_point = col_start + str(writing_row_index ) # D200
        second_slice_point = col_end + str(writing_row_index ) # F200

        thick_border_side = Side(border_style="thin", color=color_str)
        straight_thick_border = Border(
                                bottom=thick_border_side)

        #loop each cell of this row D20:Z20
        #-------->
        for cellObj in wkSheet[first_slice_point:second_slice_point]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
                # |
                # |
                # |
                # v
                cell.border = straight_thick_border

                
                latest_row = cell.row

        #------------------------------
        return latest_row

    def _initialise_named_styles(self, wb):
        # Load a specific sheet by name
        #worksheet = workbook.sheet_by_name('Sheet1')

        # Load a specific sheet by index
        #worksheet = workbook.sheet_by_index(0)

        #ws = wb.worksheets[0]
        heading1 = NamedStyle(name="Heading1")
        heading1.font = Font(name='Cambria',bold=True, color='FF3F3F76', scheme='major', sz=15.0)	
        heading1.alignment = Alignment(horizontal='left', vertical='top',)
        heading1.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading1)
        
        
        heading2 = NamedStyle(name="Heading2")
        heading2.font = Font(name='Calibri',bold=True,  scheme='minor', sz=15.0)
        heading2.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading2.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading2)

        heading2_5 = NamedStyle(name="Heading2_5")
        heading2_5.font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
        heading2_5.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading2_5.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading2_5)
        
        heading3 = NamedStyle(name="Heading3")
        heading3.font = Font(name='Calibri',b=False,  scheme='minor', sz=11.0)
        heading3.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading3.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading3)
        
        heading4 = NamedStyle(name="Heading4")
        heading4.font = Font(name='Calibri',b=False,  scheme='minor', sz=10.0)
        heading4.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading4.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading4)

        thick_border_side = Side(border_style="thick")
        straight_thick_border = Border(
                                bottom=thick_border_side)

        thick = NamedStyle(name="Thick_Bottom_Border")
        thick.border = straight_thick_border


        #---add named_style:
        wb.add_named_style(thick)
        

        double_border_side = Side(border_style="double")
        single_border_side = Side(border_style="thin",color="FF7F7F7F")
        square_border = Border(top=double_border_side,
                                right=double_border_side,
                                bottom=double_border_side,
                                left=double_border_side)

        input_border = Border(outline=True,top=single_border_side,
                                right=single_border_side,
                                bottom=single_border_side,
                                left=single_border_side)
        
        input = NamedStyle(name="Input")
        input.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F76')
        input.border = input_border
        input.alignment = Alignment(vertical="top")
        input.fill = PatternFill(fgColor='FFFFCC99', patternType='solid', fill_type='solid')
        input.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        
        #---add named_style:
        wb.add_named_style(input)
        #patternType must be one of {'darkDown', 'lightDown', 'lightGrid', 'lightHorizontal', 'darkHorizontal', 'darkTrellis', 'solid', 'lightGray', 'lightUp', 'darkGray', 'mediumGray', 'lightVertical', 'lightTrellis', 'gray0625', 'darkUp', 'darkGrid', 'darkVertical', 'gray125'}

        calculation = NamedStyle(name="Calculation")
        calculation.font = Font(name='Calibri',bold=True, color='FFFA7D00',sz=12.0)
        calculation.border = input_border
        calculation.alignment = Alignment(vertical="top")
        calculation.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        calculation.number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(calculation)

        

        single_border_side_ouput = Side(border_style="thin",color="FF3F3F3F")
        output_border = Border(outline=True,top=single_border_side_ouput,
                                    right=single_border_side_ouput,
                                    bottom=single_border_side_ouput,
                                    left=single_border_side_ouput)

        output = NamedStyle(name="Output")
        output.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output.border = output_border
        output.alignment = Alignment(vertical="top")
        output.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output)

        output1 = NamedStyle(name="Output1")
        output1.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output1.border = output_border
        output1.alignment = Alignment(vertical="top")
        output1.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output1.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output1)

        output2 = NamedStyle(name="Output2")
        output2.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output2.border = output_border
        output2.alignment = Alignment(vertical="top")
        output2.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output2.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output2)

        output3 = NamedStyle(name="Output3")
        output3.font = Font(name='Calibri',bold=True, sz=12.0, color='00339966', scheme='minor')
        #output2.border = output_border
        output3.alignment = Alignment(vertical="top")
        output3.fill = PatternFill(fgColor='00CCCCFF', patternType='solid', fill_type='solid')
        output3.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output3)

        output4 = NamedStyle(name="Output4")
        output4.font = Font(name='Calibri',bold=True, sz=12.0, color='00000000', scheme='minor')
        #output2.border = output_border
        output4.alignment = Alignment(vertical="top")
        output4.fill = PatternFill(fgColor='0099CCFF', patternType='solid', fill_type='solid')
        output4.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output4)



        percent25 = NamedStyle(name="Percent25")
        percent25.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent25.border = output_border
        percent25.alignment = Alignment(vertical="top")
        percent25.fill = PatternFill(fgColor='FFFF0000', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent25)

        percent50 = NamedStyle(name="Percent50")
        percent50.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent50.border = output_border
        percent50.alignment = Alignment(vertical="top")
        percent50.fill = PatternFill(fgColor='FF0070C0', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent50)


        percent75 = NamedStyle(name="Percent75")
        percent75.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent75.border = output_border
        percent75.alignment = Alignment(vertical="top")
        percent75.fill = PatternFill(fgColor='FFFFFF00', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent75)


        percent100 = NamedStyle(name="Percent100")
        percent100.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent100.border = output_border
        percent100.alignment = Alignment(vertical="top")
        percent100.fill = PatternFill(fgColor='FF00B050', patternType='solid', fill_type='solid')
        #percent100.number_format ='0%'
        #---add named_style:
        wb.add_named_style(percent100)


        BlueItalicText = NamedStyle(name="Blue_Italic_Text")
        BlueItalicText.font = Font(name='Calibri',bold=False, italic=True, sz=12.0, color='FF0070C0', scheme='minor')
        BlueItalicText.alignment = Alignment(vertical="top")
        #percent100.number_format ='0%'
        #---add named_style:
        wb.add_named_style(BlueItalicText)


        double_border_side_linked = Side(border_style="double",color="FFFF8001")
        
        linked_border = Border(outline=True,
                                bottom=double_border_side_linked)
        
        linkedcell = NamedStyle(name="Linkedcell")
        linkedcell.font = Font(name='Calibri', bold=True, color='FFFA7D00',scheme='minor',sz=12.0)
        linkedcell.border = linked_border
        linkedcell.alignment = Alignment(vertical="top")
        linkedcell.fill = PatternFill(fgColor='00000000', bgColor='00000000')
        linkedcell.number_format ='#,##0_);\(#,##0\);"-  ";" "@'
        
        #---add named_style:
        wb.add_named_style(linkedcell)

        """  
        linked_percentcell = NamedStyle(name="Linked_percentcell")
        linked_percentcell.font = Font(name='Calibri', bold=True, color='FFFA7D00',scheme='minor',sz=12.0)
        linked_percentcell.border = linked_border
        linked_percentcell.alignment = Alignment(vertical="top")
        linked_percentcell.fill = PatternFill(fgColor='00000000', bgColor='00000000')
        linked_percentcell.number_format ='0%'
        
        #---add named_style:
        wb.add_named_style(linked_percentcell) 
        """
   
    def _set_border_all(self, wkSheet, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

        rows = wkSheet.iter_rows(cell_range)
        for row in rows:
            for cell in row:
                cell.border = border

    # def set_border(ws, cell_range):
    #     thin = Side(border_style="thin", color="000000")
    #     for row in ws[cell_range]:
    #         for cell in row:
    #             cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    def _set_border_yellow(self, wkSheet, cell_range):
        rows = list(wkSheet[cell_range])
        side = Side(border_style='thick', color="FFF00000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border            
  
    def _set_border(self, wkSheet, cell_range):
        rows = list(wkSheet[cell_range])
        side = Side(border_style='thick', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border            
    def _remove_borders(self, wkSheet):
        side = Side(border_style=None)
        no_border = Border(
                        left=side, 
                        right=side, 
                        top=side, 
                        bottom=side,
                    )
        # Loop through all cells in te worksheet
        for row in wkSheet:
            for cell in row:
                # Apply colorless and borderless styles
                #cell.fill = no_fill
                cell.border = no_border

    def _hide_empty_cells(self, wkSheet):
        from openpyxl.utils import cell
        max_column =wkSheet.max_column
        max_row =wkSheet.max_row

        last_column = cell.column_index_from_string('XFD')+1
        for idx in range(max_column+1, last_column):
            wkSheet.column_dimensions[get_column_letter(idx)].hidden =True
        
        # last_row = 16000000
        # for idx in range(max_row+1, last_row):
        # 	wkSheet.row_dimensions[idx].hidden =True	
    def _hide_rows(self, wkSheet, start, end):
        for idx in range(start, end):
            wkSheet.row_dimensions[idx].hidden =True
        
    def _set_column_dim(self,wkSheet, rangelist, indexlist):
            
        for i in rangelist:
            for idx in range(i['start'], i['end']):
                wkSheet.column_dimensions[get_column_letter(idx)].width=i['dim']
        for idj in indexlist:
            wkSheet.column_dimensions[get_column_letter(idj['index'] )].width=idj['dim']
    def __str__(self):
        """
        Print dictionary of object attributes but don't include the _initial_inputs dict.
        """
        return str({key: val for (key, val) in vars(self).items() if key[0] != '_'})
