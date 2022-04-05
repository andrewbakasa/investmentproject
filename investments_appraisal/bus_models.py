import copy
from django.utils import translation
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, force_str
from django.utils.html import escape
from django.utils.translation import gettext as _
from django.utils.formats import get_format, number_format
from django.utils.text import capfirst, get_text_list, format_lazy
from io import StringIO, BytesIO


from tempfile import NamedTemporaryFile
from excel_response import ExcelResponse
import numpy as np
from openpyxl import Workbook, load_workbook 
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle,PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator

import math
from investments_appraisal.base_report import BaseBusinessReport



from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
import datetime
from datetime import date

from investments_appraisal.whatif import get_para_data_table_sensitivity, get_data_table_sensitivity_in_parrallel, get_data_table_sensitivity_in_single_file, get_sensitivity_gradient

#from investments_appraisal.views.usermodels import get_data_table_sensitivity

""" 
protected members of a class can be accessed by other members within the class and are also available to their subclasses. 
No other entity can access these members. In order to do so, they can inherit the parent class. 
Python has a unique convention to make a member protected: Add a prefix _ (single underscore)
"""
	
class BeefBusinessReport(BaseBusinessReport):
    
    myworboook = None
    project_title = 'PERI-URBAN CATTLE FARMING OUTPUTS'
    #Costs (Real)					
    
    cost_real = {
        #Section A---------------
        # Number of workers and supervisors
        'num_of_workers_per_pen': {'title': 'Number of workers per pen','value': 1, 'units': 'NUMBER'}, 
        'cum_pens': {'title': 'Cumulative pens under harvesting','value': None, 'units': None}, 
        'num_of_workers': {'title': 'Number of workers','value': None, 'units': 'NUMBER'},
        'num_of_supervisors_technicians': {'title': 'Number of supervisors & technicians','value': None, 'units': 'NUMBER'},  
        'average_num_of_workers': {'title': 'Average n# of workers','value': 5.00, 'units': 'NUMBER'}, 
        'average_num_of_supervisors': {'title': 'Average n# of supervisors','value':None, 'units': 'NUMBER'}, 
        #Monthly Wages and salaries
        'total_pen_constructed': {'title': 'Total Pen Constructed','value': 5, 'units': 'NUMBER'}, 
        'monthly_wage_per_worker': {'title': 'Monthly wage for workers','value': .3, 'units': 'T_LC'},
        'monthly_wage_per_supervisor': {'title': 'Monthly wage for supervisors & technicians','value': .5, 'units': 'T_LC'},
        
        #Annual Increase in real salaries
        'annual_increase_salaries_workers': {'title': 'Annual increase in real salaries of workers','value': .03, 'units': 'PERCENT'},
        'annual_increase_salaries_supervisors_technicians': {'title': 'Annual increase in real salaries of supervisors & technicians','value': .03, 'units': 'PERCENT'},
       
        #Section B---------------
        
        # Inputs
        'cattle_per_pen_per_year': {'title': 'Cattle per pen per year','value': 80, 'units': 'NUMBER'},# duplicate key 3 
        'food_conversion_ratio': {'title': 'Food Conversion Ratio','value': 8, 'units': 'NUMBER'}, 
        'daily_weight_gain': {'title': 'Daily Weight Gain','value': 1.5, 'units': 'KG'},
        'days_in_feed_lot': {'title': 'Days in Feed lot','value': 90, 'units': 'NUMBER'},  
        'weight_at_purchase': {'title': 'Weight @ Purchase','value': 250, 'units': 'KG'}, 
        'live_weight_gain': {'title': 'Live Weight Gain','value':135, 'units': 'KG'}, 
        'live_weight_at_selling': {'title': 'Live Weight @ selling','value':385, 'units': 'KG'},
        'qnty_of_feed_per_cattle': {'title': 'Qnty of Feed per Cattle','value': 1080, 'units': 'KG'},
        'commercial_feed_per_pen': {'title': 'Commercial Feed per Pen','value': 86400 , 'units': 'KG'}, 
        'dressed_weight_percent': {'title': 'Dressed Weight %','value': .55, 'units': 'PERCENT'},
        'dressed_weight_at_selling': {'title': 'Dressed Weight @ selling','value': 211.75, 'units': 'KG'},
        
        #Section C
        #Input Prices
        'cattle_price_per_unit': {'title': 'Cattle price per unit','value': 300, 'units': 'LC'}, 
        'cattle_feed_price_per_kg': {'title': 'Cattle Feed price per kg','value': 0.27, 'units': 'LC'},
        'cost_of_cattle_per_pen': {'title': 'Cost of Cattle per Pen','value': 24000, 'units': 'LC'},# duplicate key 
        'cost_of_cattle_feed_per_pen': {'title': 'Cost of Cattle Feed per Pen','value': 23328, 'units': 'LC'},
        'total_input_costs_per_pen': {'title': 'Total Input costs per Pen','value': 47328, 'units': 'LC'},
        'cattle_per_pen_per_year': {'title': 'Cattle per pen per year','value':80, 'units': 'NUMBER'},# duplicate key 

        #'dressed_weight_of_cattle_at_selling': {'title': 'Dressed Weight of Cattle @ selling','value': 211.75, 'units': 'KG'},
        'cattle_survival_rate': {'title': 'Cattle Survival rate','value': 1 , 'units': 'PERCENT'}, #duplicate
        'tonnes_produced_per_pen_per_year': {'title': 'Tonnes Produced per pen per year','value': 16.94, 'units': 'NUMBER'},
        'cost_of_imported_inputs_per_ton_of_beef_produced': {'title': 'Cost of imported inputs per ton of Beef produced','value': 0, 'units': 'USD'},
        
        'cost_of_domestic_inputs_per_ton_of_beef_produced': {'title': 'Cost of domestic inputs per ton of Beef produced','value': 2794, 'units': 'LC'},#derived
        'cost_of_electricity_per_pen_per_annum': {'title': 'Cost Of Electricity per pen per annum','value': 100 , 'units': 'LC'}, 
        'annual_change_in_price_of_imported_inputs': {'title': 'Annual change in price of imported inputs','value': .01, 'units': 'PERCENT'},
        'annual_change_in_price_of_domestic_inputs': {'title': 'Annual change in price of domestic inputs','value': 0, 'units': 'PERCENT'},
        'other_indirect_costs': {'title': 'Other Indirect Costs','value': 200, 'units': 'LC'},
    }
    """
    Pen/Feedlot Design Parameters
    """	
    feedlot_design_parameters = {
        'num_of_feedlots': {'title': 'N# Of FeedLot','value': 5, 'units': 'NUMBER'}, 
        'length': {'title': 'Length in meters','value': 7.5, 'units': 'NUMBER'}, 
        'width': {'title': 'Width in meters','value': 20.0, 'units': 'NUMBER'},
        'sqm': {'title': 'SQM covered','value': 750, 'units': 'NUMBER'},  
        'pen_area': {'title': 'Pen-Area','value': 150, 'units': 'NUMBER'}, 
        'sqm_per_cattle': {'title': 'SQM per cattle','value':7.50, 'units': 'NUMBER'}, 
        'total_cattle_per_pen_per_cycle': {'title': 'Total Cattle in one pen per cycle','value': 20, 'units': 'NUMBER'}, 
        'num_of_months_per_cycle': {'title': 'N# of months per cycle','value': 3, 'units': 'NUMBER'},
        'cattle_per_pen_per_year': {'title': 'Total Cattle per pen per year','value': 80, 'units': 'NUMBER'},
    }

    """
    Investment Cost (Real)	
    """    
    investment_cost = {
        'total_land_for_pens': {'title': 'Total Land For Pens','value': 750, 'units': 'NUMBER'}, 
        
        'cost_of_land_per_sqm': {'title': 'Cost of Land per SQM','value': 2, 'units': 'NUMBER'},  
        'cost_of_machinery_per_pen': {'title': 'Cost of Machinery (Per Pen)','value': 2500, 'units': 'NUMBER'}, 
        'cost_of_building_per_sqm': {'title': 'Cost of building per SQM','value': 10, 'units': 'NUMBER'},   
        'total_pens': {'title': 'Total pens','value': 5, 'units': 'NUMBER'},
        'cost_of_pen_construction': {'title': 'Cost of Unit Pen Construction','value': 1200, 'units': 'NUMBER'},
         #for sensitivity analysis: linked sens page which then links Financing Section
        'senior_debt_dynamic_parameter': {'title': 'Senior Debt','value': .70, 'units': 'PERCENT'}, 

        'initial_pens_employed': {'title': 'Initial Pens Employed','value': 5, 'units': 'NUMBER'}, 
        'pen_cattle_density': {'title': 'Pen-Cattle Density (Cattle/Pen)','value':20, 'units': 'NUMBER'}, 
        'pen_length': {'title': 'Pen Length meters','value': 7.50, 'units': 'NUMBER'},
        'pen_width': {'title': 'Pen Width meters','value':20.00, 'units': 'NUMBER'}, 
        'pen_height': {'title': 'Pen Height meter','value': 2.00, 'units': 'NUMBER'},

        'total_land_required': {'title': 'Total Land Required (hectares)','value': 1500, 'units': 'NUMBER'},

        'cost_of_pens_constructed': {'title': 'Cost of pens constructed','value':None, 'units': 'T_LC'}, 
        'cost_of_land': {'title': 'Cost of Land','value': None, 'units': 'NUMBER'},
        'investment_cost_of_land': {'title': 'Investment cost of land','value':None, 'units': 'T_LC'}, 
        'cif_cost_of_machinery': {'title': 'CIF cost of Machinery','value': None, 'units': 'T_USD'},
        'investment_cost_of_buildings': {'title': 'Investment cost of buildings','value':None, 'units': 'T_LC'}, 
        'investment_costs_over_run_factor': {'title': 'Investment Costs Over-run Factor','value': 0, 'units': 'PERCENT'},
    }
    #no input from this : to keep value ONLY
    #keep database table of these paramere
    investment_parameter_options ={
        'cost_of_land_per_sqm':     [1.00,	 2.00,	 3.00 ,	 5.00 ,	 5.00,	 5.00],  
        'cost_of_machinery_per_pen':[1000, 	2500, 	5000, 	5000, 	5000, 	5000], 
        'cost_of_building_per_sqm': [10.00, 10.00, 	 8.00, 	 5.00, 	 2.00, 	 1.00], 
        'cost_of_pen_construction': [ 1200,  1200, 	 1500, 	 2000, 	 2000 ,	 3000],
        'senior_debt_dynamic_parameter': [.70, .70, .70, .70, .70, .70],
        'initial_pens_employed':  [1 ,	 5, 	 20, 	 50, 	 100, 	 1000], 
        'pen_cattle_density':   [20, 	 20, 	 20, 	 20, 	 20, 	 20] , 
        'pen_length':           [8.00,	8.00,	8.00,	8.00,	8.00,	8.00],
        'pen_width':            [20.00,	20.00,	20.00,	20.00,	20.00,	20.00], 
        'pen_height':           [2.00,	2.00,	2.00,	2.00,	2.00,	2.00],
        'total_land_required':  [1000, 	 1500, 	 5000, 	 10000, 	 50000,	 100000 ]
    }

    

    
    _production_inventory = {
        'cattle_per_pen_per_year': {'title': 'Cattle per pen per year','value': 0, 'units': 'NUMBER'},
        'cattle_survival_rate': {'title': 'Cattle Survival rate','value': 1, 'units': 'PERCENT'},#duplicate
        'thousand': {'title': 'Thousand','value': 1000, 'units': 'NUMBER'}, 
        'dressed_weight_at_selling': {'title': 'Dressed Weight @ selling','value': 211.7 , 'units': 'KG'},
        'num_of_feedlot_targeted_for_construction': {'title': 'No. of feedlot(pens) targeted for construction','value': None, 'units': 'NUMBER'},
        'cum_pens_under_cattle': {'title': 'Cumulative pens under Cattle','value': None , 'units': 'NUMBER'}, 
        'cum_pens_under_harvesting': {'title': 'Cumulative pens under harvesting','value': None, 'units': 'YEARS'},
        'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'TONS'},
        'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'TONS'},
        'investment_roll_out_flag': {'title': 'Investment Roll out Flag','value': None, 'units': 'NUMBER'},
        'total_pen_constructed': {'title': 'Total Pen Constructed','value':None , 'units': 'NUMBER'},
    }	
    
    cattle_business_options = {
        'minor_scale': {'heading': 'Minor Cattle Business' ,'description': 'Minor Cattle Business:  Take a salary based loan eqv $3.700, convert it into 100% EQUITY and payback in one year. Feasible anytime from now.'}, 
        'moderate_scale': {'heading': 'Moderate Cattle Business' ,'description': 'Moderate Cattle Business: Take a salary based loan of $25,000 and payback in 5 years. This can feasible if there is confidence in banking instituion.'},
        'bigger_scale': {'heading': 'Bigger Cattle Business' ,'description': 'Bigger Cattle Business: Heavy Investment vehicle required'},
        'larger_scale': {'heading': 'Larger Beef Business' ,'description': 'Larger Beef Business: Heavy Investment vehicle required'},
        'commercial_scale': {'heading': 'Commercial Beff Business' ,'description': 'Commercial Beef Business: Heavy Investment vehicle required'},
        'global_scale': {'heading': 'Global Cattle Business' ,'description': 'Global Cattle Business: Mass Investment vehicle required'},
    }

    def update(self, param_dict):
        """
        Update parameter values
        """
        for key in param_dict:
            setattr(self, key, param_dict[key])
        
    def __init__(
        self, model_specifications=None, timing_assumptions=None, 
                 prices= None, working_capital=None,
                 financing=None, taxes=None, 
                 macroeconomic_parameters=None,  depreciation=None, 
                 cost_real=None, investment_cost= None, feedlot_design_parameters= None, 
                 cattle_business_options= None, investment_parameter_options= None):
        #initialise
        BaseBusinessReport.__init__(self, model_specifications=model_specifications, timing_assumptions=timing_assumptions, 
                 prices= prices, working_capital=working_capital,
                 financing=financing, taxes=taxes, 
                 macroeconomic_parameters=macroeconomic_parameters,  depreciation=depreciation )
        self.cost_real = cost_real
        self.feedlot_design_parameters = feedlot_design_parameters
        self.investment_cost = investment_cost
        self.cattle_business_options = cattle_business_options
        self.investment_parameter_options = investment_parameter_options


        
        
        for i in self.cattle_business_options.keys():
            self.track_inputs['cattle_business_options'][i]={}# {"name": Jonh, peter}
        
        
       
            
    def _sens_sensitivity_parrallel_generator(self):
        input_list=[] 
        sens_input_dict={}
           
        model_clone = copy.deepcopy(self)
      
        params_list =[
                'domestic_inflation_rate',
                'us_inflation_rate',
                'change_in_price', 
                'accounts_receivable',
                'accounts_payable',
                'cash_balance',
                'exchange_rate', 
                'senior_debt',
                'base_price',
                'cattle_survival_rate',

                'cattle_price_per_unit',
                'initial_pens_employed',
                'discount_rate_equity',
                'dressed_weight_at_selling',
                'other_indirect_costs',
                'annual_change_in_price_of_domestic_inputs',
                'annual_change_in_price_of_imported_inputs',
                'total_cattle_per_pen_per_cycle',

                'monthly_wage_per_worker',
                'monthly_wage_per_supervisor',
                'annual_increase_salaries_workers',  
                'annual_increase_salaries_supervisors_technicians',
                'num_workers_per_supervisor',
                
            ]




        for var_x in params_list:      
        
            list_ =[]
            index = 0
            base_point=0
            if var_x  in ['cattle_survival_rate']:
                list_ = [1 ,0.99, 0.98, .95, .90, 0.85, 0.80, 0.70, .60, .50]
                base_point =model_clone.cattle_survival_rate 
                if not (model_clone.cattle_survival_rate in list_):
                    list_.append(base_point)
            elif var_x  in ['senior_debt']:
                list_ = [1 ,.95, 0.90, .90, .70, .60, .50, 0.40, .30, .20]
                base_point =model_clone.senior_debt  
                if not (model_clone.senior_debt in list_):
                    list_.append(base_point)
            elif var_x  in ['initial_pens_employed']:
                list_ = [1000 ,500, 200, 100, 50, 20, 10, 5, 1]
                base_point =model_clone.initial_pens_employed 
                if not (model_clone.initial_pens_employed in list_):
                    list_.append(base_point)
            elif var_x  in ['total_cattle_per_pen_per_cycle']:
                list_ = [100 ,50, 20, 10, 5]
                base_point =model_clone.cattle_pen_cycle
                if not (model_clone.cattle_pen_cycle in list_):
                    list_.append(base_point) 
                              
            elif var_x  in ['discount_rate_equity']:
                list_ = [.05,.1,.12,.15,.20,.3]
                base_point =model_clone.discount_rate_equity 
                if not (model_clone.discount_rate_equity in list_):
                    list_.append(base_point)   
            elif var_x  in ['dressed_weight_at_selling']:
                list_ = [90,120, 150,200, 300]
                base_point =model_clone.dressed_weight_at_selling
                if not (model_clone.dressed_weight_at_selling in list_):
                    list_.append(base_point)  
            elif var_x  in ['other_indirect_costs']:
                list_ = [100,1000,5000]
                base_point =model_clone.other_indirect_costs 
                if not (model_clone.other_indirect_costs in list_):
                    list_.append(base_point)                
            else:
                if hasattr(self, var_x ):
                    base_point= getattr(self, var_x)

                if base_point==0:
                    list_ = [0.02 * i  for i in range(0,10)]
                else:
                    list_ = [0.2 * i * base_point  for i in range(1,10)]            
            
            list_.sort()
            input_list.append({'name':var_x, 'list':list_})
           
            sens_input_dict[var_x]= {'list':list_,'base_val': base_point}

        setattr(self, 'sens_input_dict', sens_input_dict) 
        #do once
        #before adding boundaries
        sens_dict = get_data_table_sensitivity_in_parrallel(model_clone,input_list)
        
        setattr(self, 'sens_dict', sens_dict)   
        sens_grad =get_sensitivity_gradient(model_clone,input_list)
        sens_grad.fillna(0,inplace=True)
        sens_grad.sort_values(by=['abs_grad'], ascending=False) 
        print(sens_grad)
        setattr(self, 'sens_grad', sens_grad)
        mask_is_sensitivity=sens_grad['abs_grad'] >float(0.001)
        #list only which is sensitive
        para_list_by_grad= sens_grad[mask_is_sensitivity]['parameter'].tolist() 
        setattr(self, 'para_list_by_grad', para_list_by_grad)
        
        return sens_dict
    def spreadsheet(self) :
        
        if not hasattr(self, 'para_list_by_grad'):
            self._sens_sensitivity_parrallel_generator()
        # Return an excel spreadsheet
        output = BytesIO()
        self._generate_data(
            output
        )
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=output.getvalue(),
        )
       
       
        wb = load_workbook(filename=BytesIO(output.getvalue()))
        #wb = load_workbook(filename=BytesIO(output.getvalue()), data_only = True)
        #print(wb, wb.sheetnames)
        self._get_out_values_as_dict(wb)

        #-----------title name------------------
        date_now = datetime.datetime.now() #date.today()
        date_now = str(date_now.year) + '_' + str(date_now.month) \
                    + '_' + str(date_now.day)+ "_" + str(date_now.hour) \
                    + '_' + str(date_now.minute) + "_" +  str(date_now.second)
        title = "clearVision_solutions_auto-generated beefbusiness" + str(date_now)
        #-----------------------------

      
        response[
            "Content-Disposition"
        ] = "attachment; filename*=utf-8''%s.xlsx" % force_str(title) 
        response["Cache-Control"] = "no-cache, no-store"
        return response


	
    
    def  _generate_data(self,output):
        # Create a workbook
        wb = Workbook()
        self.myworboook = wb
        #initialise styles
        BaseBusinessReport._initialise_named_styles(self,wb)
        
        self._write_input_sheet(wb,)
        #correct formating here---------
        self._write_output_sheet(wb, )
        self._write_sens_sheet(wb, )
       
        self._write_calculation_sheet(wb, )
        self._write_cashflow_sheet(wb, )
        
        #update
        for item in self._production_inventory.keys():
            self._populate_productionInventory_section(wb["Inputs"], item)
        #---------------------------------------------------------------

        for item in self.investment_cost.keys():
            self._populate_investmentcost_section(wb["Inputs"], item)
        
	
        for item in self.cost_real.keys():
            self._populate_costs_section(wb["Inputs"], item)
        
        for item in ['initial_pens_employed', 'senior_debt', 'total_cattle_per_pen_per_cycle']:
            self._update_sens_section(wb["Sens"], item)
        
        #-------------
        self._update_investment_parameters(wb["Inputs"])

        #
        self._link_feedlot_design_parameters_to_model_selected(wb["Inputs"])
        self._update_sensitivity(wb["Sens"])
        #updated here again.... 02 Feb 2022
        #paid salary 110K
        # iam sure it will work
        # i love Programming
        # I love God
        # I love JESUS
        self._populate_output_sheet(wb["Outputs"])
        
        self._write_analytics_sheet(wb, )
        # Write a formatted header row
        header = []
        comment = None

       
        # Write the spreadsheet
        wb.save(output)

    
    def _write_output_sheet(self, wb):
        total_wsheet_cols = 17
        #---------------------------Worksheet 1---------------------------
        output_sheet = wb['Outputs']	
        col = get_column_letter(2)

        row_index = 2
        output_sheet.cell('%s%s'%(col, row_index)).value = self.project_title# 'PERI-URBAN CATTLE FARMING OUTPUTS'# dynamic heading
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
          
        BaseBusinessReport._set_thick_bottom_border_range(self, output_sheet,  row_index, 1, total_wsheet_cols)
        
        row_index += 2
        output_sheet.cell('%s%s'%(col, row_index)).value = 'Scenario Analysis '
        output_sheet['%s%s'%(col, row_index)].style= 'Heading1'
        
        col = get_column_letter(6)
        output_sheet.cell('%s%s'%(col, row_index)).value = 'Active Scenario'
        output_sheet['%s%s'%(col, row_index)].alignment = Alignment(horizontal='left', vertical='top',)
        output_sheet['%s%s'%(col, row_index)].font = Font(name='Calibri',bold=False, italic=True, sz=10.0, color='FF0070C0', scheme='minor')
        
        col = get_column_letter(7)
        _unit= "NOTE"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet.cell('%s%s'%(col, row_index)).value = cell_display_formuale
        output_sheet.cell('%s%s'%(col, row_index)).style= 'Explanatory Text'

       
        BaseBusinessReport._set_thick_bottom_border_range(self, output_sheet,  row_index, 1, total_wsheet_cols)
        col = get_column_letter(1)
        row_index+=2
        
        #-------------------Data Validation---------------------------
        i=0
        values_ =''
        for item in self.cattle_business_options.keys():
            if i ==len(self.cattle_business_options.keys())-1:
                values_ += self.cattle_business_options[item]['heading'] 
            else:
               values_ += self.cattle_business_options[item]['heading'] + ','     
            i+=1
        dv = DataValidation(type="list", formula1='"'+ values_ + '"', allow_blank=True)
        
        # Optionally set a custom prompt message
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'
        
        # Add the data-validation object to the worksheet
        output_sheet.add_data_validation(dv)
        # Create some cells, and add them to the data-validation object
        model_sel_row= row_index
        model_sel_col= "F"
        model_sel_cell= "$" + model_sel_col + "$" + str(model_sel_row)
        model_cell = output_sheet[model_sel_cell]
        self.track_inputs['outputs']['model_selection']['row']= model_sel_row 
        self.track_inputs['outputs']['model_selection']['col']= model_sel_col
        self.track_inputs['outputs']['model_selection']['cell']= model_sel_cell

        model_cell.value = values_.split(",")[0]
        model_cell.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        model_cell.font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
        model_cell.alignment = Alignment(horizontal='left', vertical='top',)
        dv.add(model_cell)
        
       
       #----switch

       
        col = get_column_letter(3)
        output_sheet.cell('%s%s'%(col, row_index)).value = 'A.'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(4)
        output_sheet.cell('%s%s'%(col, row_index)).value = 'Alternative Investment Scenario'
        output_sheet['%s%s'%(col, row_index)].style= 'Heading3'

        col = get_column_letter(7)
        _unit= "SWITCH"

        cell_display_formuale=""
        if self.modelspec_cell_ref:
            cell_display_formuale = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
        output_sheet.cell('%s%s'%(col, row_index)).value = cell_display_formuale
        output_sheet.cell('%s%s'%(col, row_index)).style= 'Explanatory Text'
       #------------------

        #####################333update input------------------------
        colHeader = get_column_letter(5)
        select_val ="="
        list_= self.cattle_business_options.keys()
        #print(list_)
        closin_brackets=''
        for  i in range(len(list_)-1):
            closin_brackets+=')'

        i = 0
        input_row_index= self.track_inputs['cattle_business_options']['header']['row']
        for item in list_:
            #item =list_[i]
            if i ==len(list_)-1:
                select_val += colHeader + str(input_row_index +i+1) + closin_brackets
            else:
                cell_x= model_sel_cell
                select_val += 'IF(Outputs!' + cell_x + '="'+ self.cattle_business_options[item]['heading'] +'",' + colHeader + str(input_row_index +i+1) + ','
            i+=1
        #---------------------------------------------------------------
         
        wb["Inputs"].cell('%s%s'%(colHeader, input_row_index)).value = select_val
        wb["Inputs"]['%s%s'%(colHeader, row_index)].font = self.description_font
        ##########################################3


      
        row_index+=2

        BaseBusinessReport._hide_empty_cells(self, output_sheet)
        #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':10})
        rangelist.append({'start':1, 'end': 2 , 'dim':5})
        rangelist.append({'start':2, 'end': 5, 'dim':2})
        #rangelist.append({'start':2, 'end': 3, 'dim':5})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':5, 'dim':55})
        indexlist.append({'index':6, 'dim':30})
        indexlist.append({'index':7, 'dim':15})

        #set dims----
        BaseBusinessReport._set_column_dim(self,output_sheet,rangelist,indexlist)
        #------------------------------------
        start_row = row_index
        #end_row = start_row +10
        
        # ------------Financial Section------------------ 
        end_row =self._add_output_financial(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row, end_row,'Financial Analysis Output',)
        
        #--------User View of Selected Project Option-------------------------------------------------------
        start_row= row_index
        source_row_ =self.track_inputs['cattle_business_options']['header']['row']
        wb['Outputs'].cell('%s%s'%("I", start_row)).value = '=Inputs!$E$'+ str(source_row_)
        col2_ =column_index_from_string("I")
        wb['Outputs']['I' + str(start_row)].alignment = Alignment(wrap_text=True,vertical='top')
        wb['Outputs']['I' + str(start_row)].font = Font(name='Bookman Old Style',bold=False, sz=20.0, color='FF0070C0', scheme='minor')
        

        #description of selected project
        wb['Outputs'].merge_cells(start_row=start_row, start_column=col2_, end_row=end_row, end_column=col2_ + 6)
        #----------------------------------------------------------------- 
       
              
        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------Funding Section------------------ 
        end_row =self._add_output_funding(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'Funding','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        start_row = end_row + 2
        #end_row = start_row +4
        row_index = start_row
        # ------------General Section------------------ 
        end_row =self._add_output_general(output_sheet, row_index)
        self._output_bordered_section(output_sheet,"B","H",start_row,end_row,'General','Heading1','Accent1')
        # ------------------------------Write cell------------------ 

        #self._remove_borders(output_sheet)
        # remove gridlines
        output_sheet.sheet_view.showGridLines = False
        output_sheet.sheet_view.showRowColHeaders = False
    
    def _populate_output_sheet(self, wksheet):
       self._populate_output_financial(wksheet)
       self._populate_output_fundings(wksheet)
       self._populate_output_general(wksheet)
    
   
  
    def _populate_output_financial(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        percent_format ='0%'

        itemlist= ['minimum_ADSCR','maximum_ADSCR','average_ADSCR',  'minimum_LLCR','maximum_LLCR','average_LLCR']
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_nominal','output_financial',i,number_format, 'CF',None,'Output2')

        itemlist= ['NPV', 'IRR', 'MIRR' ]
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_real','output_financial',i,number_format, 'CF',None,'Output2')
        
        itemlist= ['IRR', 'MIRR' ]
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_real','output_financial',i,percent_format, 'CF',None,'Output2')
              
          
    def _populate_output_fundings(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        itemlist= ['total_investment_cost','senior_debt','equity' ]
       
        #----------
        #set Cell_format= None to retian blue backcolot
        self._transfer_cell(wksheet, 'financing','output_funding', 'senior_debt', '0%', 
                               'Inputs', 'header', None, None ,6-9)
        #70% Senior Debt                       
        self._write_text_atcell(wksheet,'output_funding','header',"Senior Debt",7-9) 
        #1. total_investment_cost  
        self._transfer_range_formulae_tocell(wksheet, 'calc_investment_cost_nominal', 'output_funding',  'total_investment_cost' ,
               number_format , "Calc", None,'Output2' ) 

        #2. total_investment_cost  
        self._transfer_range_formulae_tocell(wksheet, 'calc_investment_cost_nominal', 'output_funding',  'senior_debt_towards_investment' ,
               number_format, "Calc", 'senior_debt','Output2' ) 


        #3. equity
        #=A-B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_funding', 'para': 'total_investment_cost'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "B", 'var_type': 'variable', 'header': 'output_funding', 'para': 'senior_debt'},
              
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_funding', 'equity',None,'Output2')
        


    def _populate_output_general(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
       
        #1 Transer cell to Outputs (Group Loop)
        array_dict= [
            {'target': 'land_under_feedLot',                'source': 'total_land_for_pens',             's_header': 'investment_cost',          's_wksheet': 'Inputs'},
            {'target': 'total_feedLots',                    'source': 'total_pen_constructed',           's_header': 'production_inventory',     's_wksheet': 'Inputs'},
            {'target': 'pen_area',                          'source': 'pen_area',                        's_header':  'feedlot_design_parameters','s_wksheet': 'Inputs'},
            {'target': 'total_cattle_in_one_pen_per_cycle', 'source': 'total_cattle_per_pen_per_cycle',  's_header': 'feedlot_design_parameters', 's_wksheet': 'Inputs'},
            {'target': 'cattle_per_pen_per_year',           'source': 'cattle_per_pen_per_year',         's_header': 'feedlot_design_parameters', 's_wksheet': 'Inputs'},
        
            {'target': 'base_price',                     'source': 'base_price',                       's_header':  'prices','s_wksheet': 'Inputs'},
            {'target': 'tonnes_produced_per_pen_per_year',   'source': 'tonnes_produced_per_pen_per_year',  's_header': 'cost_real',        's_wksheet': 'Inputs'},
            ]
        
        for i in array_dict:
            self._transfer_cell(wksheet, i['s_header'],'output_general',
                               i['source'], number_format, i['s_wksheet'], i['target'],'Output2')
        #2. Transer cell to Outputs (Group Loop)
        # in cellrange at specific cell
        array_dict= [
            {'target': 'annual_revenue',          'source': 'gross_sales_revenue',  's_header': 'cal_production_sales_nominal',  's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':0 },# first year
            {'target': 'gross_profit_2yr',        'source': 'gross_profit',         's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':1 },# second year
            {'target': 'net_profit_2yr',          'source': 'net_after_tax_income', 's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format': number_format, 'cell_pos':1 },# second year
            {'target': 'gross_profit_margin_2yr', 'source': 'gross_profit_margin',  's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format':'0.0%',         'cell_pos':1 },# second year
            {'target': 'net_profit_margin_2yr',   'source': 'net_profit_margin',    's_header': 'cal_income_tax_statement',      's_wksheet': 'Calc', 'number_format':'0.0%',         'cell_pos':1 },# second year
        ]
        for item in array_dict:
            self._transfer_cell(wksheet, item['s_header'],'output_general', item['source'], item['number_format'], 
                               item['s_wksheet'], item['target'],'Output2',item['cell_pos'])
        
        #3. total_cattle_per_year
        #  = cattle_per_pen_per_year  * total_feedLots
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'cattle_per_pen_per_year'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "B", 'var_type': 'variable', 'header': 'output_general', 'para': 'total_feedLots'},
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'total_cattle_per_year',None,'Output2')
        
        
        #3. total_cattle_per_week
        #  =total_cattle_per_year  /52
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'total_cattle_per_year'},
                {'value': '/52', 'var_type': 'const', 'header': '', 'para': '',},
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'total_cattle_per_week',None,'Output2')
         

        #3. annual_productivity_tons
        #  =tonnes_produced_per_pen_per_year  * total_feedLots
        #  =A*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_general', 'para': 'tonnes_produced_per_pen_per_year'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'B', 'var_type': 'variable','header': 'output_general', 'para': 'total_feedLots'},
               
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_general', 'annual_productivity_tons',None,'Output2')
                          
    def _add_output_general(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_general' in self.track_inputs:
            self.track_inputs['output_general']= {}# insert inner

        if not ('header' in self.track_inputs['output_general']):
            self.track_inputs['output_general']['header'] = {}
  
        if 'header' in self.track_inputs['output_general']:
            self.track_inputs['output_general']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= [
                'land_under_feedLot',  'total_feedLots',    'pen_area', 'total_cattle_in_one_pen_per_cycle','cattle_per_pen_per_year',
                'total_cattle_per_year',   'total_cattle_per_week',   'base_price','tonnes_produced_per_pen_per_year', 'annual_productivity_tons',
                'annual_revenue', 'gross_profit_2yr',   'net_profit_2yr',  'gross_profit_margin_2yr',   
                'net_profit_margin_2yr' 
                ]
        ic_parameters = {
            'land_under_feedLot': {'title': "Land Under FeedLot",'value': None, 'units': 'SQM'},
            'total_feedLots': {'title': "Total FeedLots",'value': None, 'units': 'NUMBER'},
            'pen_area': {'title': "Pen-Area",'value': None, 'units': 'SQM'},
            'total_cattle_in_one_pen_per_cycle': {'title': "Total Cattle in one pen per cycle",'value': None, 'units': 'NUMBER'},
            'cattle_per_pen_per_year': {'title': "Total Cattle per pen per year",'value': None, 'units': 'NUMBER'},
            'total_cattle_per_year': {'title': "Total Cattle per year",'value': None, 'units': 'NUMBER'},
            'total_cattle_per_week': {'title': "Total Cattle per week",'value': None, 'units': 'NUMBER'},

            'base_price': {'title': "Base price of beef per ton",'value': None, 'units': 'LC'},
            'tonnes_produced_per_pen_per_year': {'title': 'Beef Tonnes produced per pen per year','value': None, 'units': 'TONS'},
            'annual_productivity_tons': {'title': 'Annual Beef Productivity Tons','value': None, 'units': 'TONS'},
            'annual_revenue': {'title': "Annual Revenue",'value': None, 'units': 'T_LC'},

            'gross_profit_2yr': {'title': "Gross Profit (2nd Yr)",'value': None, 'units': 'T_LC'},
            'net_profit_2yr': {'title': 'Net Profit (2nd Yr)','value': None, 'units': 'T_LC'},
            'gross_profit_margin_2yr': {'title': "Gross Profit Margin  (2nd Yr)",'value': None, 'units': 'PERCENT'},
            'net_profit_margin_2yr': {'title': "Net Profit Margin (2nd Yr)",'value': None, 'units': 'PERCENT'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                                   
                if not (item in self.track_inputs['output_general']):
                    self.track_inputs['output_general'][item] = {}

                self.track_inputs['output_general'][item]['row'] = row_index
                self.track_inputs['output_general'][item]['unit'] = _unit
                self.track_inputs['output_general'][item]['col'] = 6
                self.track_inputs['output_general'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

                if item in ['annual_revenue','annual_productivity_tons', 
                        'total_cattle_in_one_pen_per_cycle', 'total_cattle_per_week'] :
                    row_index +=1
                    
        return row_index         
  
    def _add_output_funding(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_funding' in self.track_inputs:
            self.track_inputs['output_funding']= {}# insert inner

        if not ('header' in self.track_inputs['output_funding']):
            self.track_inputs['output_funding']['header'] = {}
  
        if 'header' in self.track_inputs['output_funding']:
            self.track_inputs['output_funding']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= ['total_investment_cost','senior_debt','equity' ]        
        ic_parameters = {
            'total_investment_cost': {'title': "Total Investment Cost",'value': None, 'units': 'T_LC'},
            'senior_debt': {'title': "Senior Debt",'value': None, 'units': 'T_LC'},
            'equity': {'title': " Equity",'value': None, 'units': 'T_LC'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit)
                
                                   
                if not (item in self.track_inputs['output_funding']):
                    self.track_inputs['output_funding'][item] = {}

                self.track_inputs['output_funding'][item]['row'] = row_index
                self.track_inputs['output_funding'][item]['unit'] = _unit
                self.track_inputs['output_funding'][item]['col'] = 6
                self.track_inputs['output_funding'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

        return row_index         
          
    def _add_output_financial(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_financial' in self.track_inputs:
            self.track_inputs['output_financial']= {}# insert inner

        if not ('header' in self.track_inputs['output_financial']):
            self.track_inputs['output_financial']['header'] = {}
  
        if 'header' in self.track_inputs['output_financial']:
            self.track_inputs['output_financial']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= [
            'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
            'minimum_LLCR','maximum_LLCR','average_LLCR',
            'NPV', 'IRR', 'MIRR'
            ]

        self._add_sub_heading(output_sheet,"A. Lender's Perspective",row_index)
        row_index +=1
        self._add_sub_heading2(output_sheet,"Debt Service Coverage Ratios",row_index)
        row_index +=1
        self._add_sub_heading3(output_sheet,"A.1. Annual Debt Service Coverage Ratio - ADSCR",row_index)
        row_index +=1
        ic_parameters = {
            # A. Lender's Perspective			
            #     Debt Service Coverage Ratios		
            #         A.1. Annual Debt Service Coverage Ratio - ADSCR	

            'minimum_ADSCR': {'title': "Minimum ADSCR",'value': None, 'units': 'NUMBER'},
            'maximum_ADSCR': {'title': "Maximum ADSCR",'value': None, 'units': 'NUMBER'},
            'average_ADSCR': {'title': "Average ADSCR",'value': None, 'units': 'NUMBER'},

             #        A.2. Loan Life Coverage Ratio - LLCR
            'minimum_LLCR': {'title': "Minimum LLCR",'value': None, 'units': 'NUMBER'},
            'maximum_LLCR': {'title': "Maximum LLCR",'value': None, 'units': 'NUMBER'},
            'average_LLCR': {'title': "Average LLCR",'value': None, 'units': 'NUMBER'},

            # B. Owner Perspective			
	        #       Financial Viability of the Project
            'NPV': {'title': "NPV",'value': None, 'units': 'T_LC'},
            'IRR': {'title': 'IRR','value': None, 'units': 'PERCENT'},
            'MIRR': {'title': "MIRR",'value': None, 'units': 'PERCENT'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
               
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
             
                                   
                if not (item in self.track_inputs['output_financial']):
                    self.track_inputs['output_financial'][item] = {}

                self.track_inputs['output_financial'][item]['row'] = row_index
                self.track_inputs['output_financial'][item]['unit'] = _unit
                self.track_inputs['output_financial'][item]['col'] = 6
                self.track_inputs['output_financial'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

                if item in ['average_ADSCR','average_LLCR'] :
                    row_index +=1
                    if item=='average_ADSCR':
                         self._add_sub_heading3(output_sheet,"A.2. Loan Life Coverage Ratio - LLCR",row_index)
                         row_index +=1
                    if item=='average_LLCR':
                        
                         self._add_sub_heading(output_sheet,"B. Owner Perspective",row_index)
                         row_index +=1 
                         self._add_sub_heading2(output_sheet,"Financial Viability of the Project",row_index)
                         row_index +=1  

        return row_index         
    def _output_bordered_section(self, w_sheet, col_letter_start, col_letter_end, 
                start_row, end_row, title_, heading_style='Heading1', accent_style='Accent1'):

        w_sheet.cell('%s%s'%(col_letter_start, start_row)).value = title_
        w_sheet.cell('%s%s'%(col_letter_start, start_row)).style = heading_style
        w_sheet[col_letter_start + str(start_row)].alignment = Alignment(wrap_text=False,vertical='top')

        col_s =column_index_from_string(col_letter_start)
        col_e= column_index_from_string(col_letter_end)
        #
        for i in range(col_s, col_e+1):
            coli = get_column_letter(i)
            #first row color:
            w_sheet['%s%s'%(coli, start_row)].style= accent_style

        _range=f"{col_letter_start}{start_row}:{col_letter_end}{end_row}"
        self._set_border(w_sheet,_range)
       
    def _add_legend_section(self, w_sheet, row_index,total_wsheet_cols, unit_start_col=7):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Legend' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        #===================================
        
        row_index +=2
        col = get_column_letter(unit_start_col)    
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Unit'
        w_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'

        col = get_column_letter(unit_start_col+2)    
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Input'
        w_sheet['%s%s'%(col, row_index)].style= 'Input'

        col = get_column_letter(unit_start_col+3)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Calculation' 
        w_sheet['%s%s'%(col, row_index)].style= 'Calculation'

        col = get_column_letter(unit_start_col+4)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Output'
        w_sheet['%s%s'%(col, row_index)].style= 'Output'

        col = get_column_letter(unit_start_col+5)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Linked Cell'
        w_sheet['%s%s'%(col, row_index)].style= 'Linkedcell'

        return row_index 
    
    def _add_modelspecification_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Model specification' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        row_index +=1
        self.modelspec_cell_ref ={}
        
        if not ('model_specs' in self.track_inputs):
            self.track_inputs['model_specs']= {} 
            
        for model_item in self.model_specifications:
            #print('model_item-->', model_item)
            w_sheet.cell('%s%s'%(colHeader, row_index)).value = model_item['title']
            w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font

            w_sheet.cell('%s%s'%(colValue, row_index)).value = model_item['units']
            w_sheet['%s%s'%(colValue, row_index)].style= 'Input'
            w_sheet['%s%s'%(colValue, row_index)].font = self.description_font2
            
            #keep record      
            self.modelspec_cell_ref[model_item['name']]= '=Inputs!$%s$%s' %(colValue,row_index)
            model_item_name  =model_item['name']
            # 7 lines not working yet
            if not (model_item_name in self.track_inputs['model_specs']):
                self.track_inputs['model_specs'][model_item_name]= {} 

            self.track_inputs['model_specs'][model_item_name]['row'] = row_index
            self.track_inputs['model_specs'][model_item_name]['col'] = 6
            self.track_inputs['model_specs'][model_item_name]['unit'] = model_item['units']
            self.track_inputs['model_specs'][model_item_name]['value'] =  model_item['title']

            row_index +=1

   
        return row_index 
    def _write_row_title_and_value(self, w_sheet, write_colHeader, write_colValue, 
        row_index, header_title,value, val_number_format= 'General'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet.cell('%s%s'%(write_colHeader, row_index)).value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font

        w_sheet.cell('%s%s'%(write_colValue, row_index)).value = value
        w_sheet['%s%s'%(write_colValue, row_index)].style= 'Input'
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2

    def _add_timingassumptions_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Timing Assumptions' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)# E
        colValue = get_column_letter(6) # F
        row_index +=1
       
        #********************************Base Period***************************************
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['base_period']['title'],self.timing_assumptions['base_period']['value'])     
       

        self.track_inputs['timing']['base_period']['row'] = row_index
        self.track_inputs['timing']['base_period']['col'] = 6
        self.track_inputs['timing']['base_period']['unit'] = 'YEAR'
        self.track_inputs['timing']['base_period']['value'] = self.timing_assumptions['base_period']['value']
        row_index +=2

        
       

        #******************************Construction Period********************************
        #----Start Year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['construction_start_year']['title'],self.timing_assumptions['construction_start_year']['value'])                   
        row_index +=1
        
        #---Length
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['construction_len']['title'],self.timing_assumptions['construction_len']['value'])
        row_index +=1

        #---End
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['construction_year_end']['title'],self.timing_assumptions['construction_year_end']['value'])
        self.track_inputs['timing']['construction_year_end']['row']= row_index 
        self.track_inputs['timing']['construction_year_end']['unit'] = 'YEAR'
        self.track_inputs['timing']['construction_year_end']['col'] = 6
        self.track_inputs['timing']['construction_year_end']['value'] = self.timing_assumptions['construction_year_end']['value']         
        row_index +=2

        

        #******************************Operation Period***********************************************
        #----Start Year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['operation_start_year']['title'],self.timing_assumptions['operation_start_year']['value'])        
        self.track_inputs['timing']['operation_start_year']['row'] = row_index 
        self.track_inputs['timing']['operation_start_year']['unit'] = 'YEAR'
        self.track_inputs['timing']['operation_start_year']['col'] = 6
        self.track_inputs['timing']['operation_start_year']['value'] = self.timing_assumptions['operation_start_year']['value'] 
                   
        row_index +=1
        #---duration
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['operation_duration']['title'],self.timing_assumptions['operation_duration']['value'])
        row_index +=1

        #---End
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['operation_end']['title'],self.timing_assumptions['operation_end']['value'])
        self.track_inputs['timing']['operation_end']['row'] = row_index 
        self.track_inputs['timing']['operation_end']['unit'] = 'YEAR'
        self.track_inputs['timing']['operation_end']['col'] = 6
        self.track_inputs['timing']['operation_end']['value'] = self.timing_assumptions['operation_end']['value'] 
                        
        row_index +=2

        #---Number of months in a year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.timing_assumptions['number_of_months_in_a_year']['title'],self.timing_assumptions['number_of_months_in_a_year']['value'])

        self.track_inputs['timing']['number_of_months_in_a_year']['row'] = row_index 
        self.track_inputs['timing']['number_of_months_in_a_year']['unit'] = 'YEAR'
        self.track_inputs['timing']['number_of_months_in_a_year']['col'] = 6
        self.track_inputs['timing']['number_of_months_in_a_year']['value'] = self.timing_assumptions['number_of_months_in_a_year']['value'] 
             

        return row_index 

    def _add_prices_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Prices' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================

        #eadin 2
        col = get_column_letter(2)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Net Sales Price: %s' % (commodity_title)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1


        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        
        # Base price of beef ROW:
        w_sheet.cell('%s%s'%(colHeader, row_index)).value = 'Base price of %s in %s per ton' % (commodity_title ,self.timing_assumptions['base_period']['value'])
        w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
        w_sheet.cell('%s%s'%(colValue, row_index)).value = self.prices['base_price']
        w_sheet['%s%s'%(colValue, row_index)].style= 'Input'
        if 'base_price' in self.track_inputs['prices']:
            self.track_inputs['prices']['base_price']['row'] = row_index
            self.track_inputs['prices']['base_price']['col'] = 6
            self.track_inputs['prices']['base_price']['cell'] =  "$" + colValue + "$" + str(row_index)
            self.track_inputs['prices']['base_price']['unit'] = "NUMBER"
            self.track_inputs['prices']['base_price']['value'] = self.prices['base_price']
         
       
        w_sheet.cell('%s%s'%(colUnits, row_index)).value  =self.modelspec_cell_ref['LC'] if  'LC' in self.modelspec_cell_ref else ''
        w_sheet['%s%s'%(colUnits, row_index)].style= 'Explanatory Text'  
        row_index +=1

        # Change in Price of Beef ROW:
        w_sheet.cell('%s%s'%(colHeader, row_index)).value = 'Change in Price of %s'  % (commodity_title)
        w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
        w_sheet.cell('%s%s'%(colValue, row_index)).value = self.prices['change_in_price']

        w_sheet['%s%s'%(colValue, row_index)].style= 'Linkedcell'
        w_sheet['%s%s'%(colValue, row_index)].number_format ='0%'

        if 'change_in_price' in self.track_inputs['prices']:
            self.track_inputs['prices']['change_in_price']['row'] = row_index
            self.track_inputs['prices']['change_in_price']['col'] = 6
            self.track_inputs['prices']['change_in_price']['cell'] =  "$" + colValue + "$" + str(row_index)
            self.track_inputs['prices']['change_in_price']['unit'] = "PERCENT"
            self.track_inputs['prices']['change_in_price']['value'] = self.prices['change_in_price']
         
        

        w_sheet.cell('%s%s'%(colUnits, row_index)).value = self.modelspec_cell_ref['PERCENT'] if  'PERCENT' in self.modelspec_cell_ref else ''
        w_sheet['%s%s'%(colUnits, row_index)].style= 'Explanatory Text'

        row_index +=1

        return row_index 

    def _get_span(self):
        span_ = self.timing_assumptions['operation_duration']['value'] if 'operation_duration' in self.timing_assumptions else 0
                
           
        """ 
        e.g
        2022-2032
        10 year operation
        1 year deconstrustion
        span is 10 + 1 = (oparation + 1 deconsctruion year)
        """ 
        if span_> 0:
            return span_ +1
        else:
            #if no diff return 0
            return span_ 
               
    def _populate_productionInventory_section(self, w_sheet, item ):

        if item=='num_of_feedlot_targeted_for_construction':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            if found_state:
                span_ =self._get_span()
                start_col_index = 9
                first_slice_point = get_column_letter(start_col_index +1) + str(r_index)# skipp first col
                second_slice_point = get_column_letter(start_col_index + int(span_)) + str(r_index) # D39
                

                if 'sens' in self.track_inputs:
                    if 'initial_pens_employed' in self.track_inputs['sens']:
                        _row = self.track_inputs['sens']['initial_pens_employed']['row'] 
                        _col = self.track_inputs['sens']['initial_pens_employed']['col']
                        col = get_column_letter(_col)
                        w_sheet.cell('%s%s'%(get_column_letter(start_col_index), r_index)).value = "=Sens!$" + col + "$" + str(_row)
                        w_sheet.cell('%s%s'%(get_column_letter(start_col_index), r_index)).style = 'Linkedcell'
                    
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                    prev_col=column_index_from_string(new_column_letter)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_cell = prev_letter + str(cell.row)
                   
                   
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '0'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
            
        if item=='cum_pens_under_cattle':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_cattle')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        
                        new_column_letter = cell.column # J
                        prev_col=column_index_from_string(new_column_letter)-1
                        prev_letter= get_column_letter(prev_col)
                        cum_time_minus_one_cell = prev_letter + str(cell.row)

                        prev_row = cell.row-1 
                        current_constr_cell = new_column_letter + str(prev_row)
                    
                    
                        w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=' + cum_time_minus_one_cell + '+' + current_constr_cell
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                        #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General
        if item=='cum_pens_under_harvesting':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_harvesting')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        new_column_letter = cell.column # J
                        prev_row = cell.row-1 
                        current_constr_cell = new_column_letter + str(prev_row)
                    
                    
                        w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=' + current_constr_cell 
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                        #w_s 
        if item=='production_quantity':
            #=A*B*C*D/E * F
            #= cattle_per_pen_per_year * cattle_survival_rate * dressed_weight_at_selling * cum_pens_under_cattle/ thousand * OP
            formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'production_inventory', 'para': 'cattle_per_pen_per_year','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'production_inventory', 'para': 'cattle_survival_rate','cell_type': 'single'},
                {'value':'*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'production_inventory', 'para': 'dressed_weight_at_selling','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'D', 'var_type': 'variable', 'header': 'production_inventory', 'para': 'cum_pens_under_harvesting','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'E', 'var_type': 'variable', 'header': 'production_inventory', 'para': 'thousand','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'F', 'var_type': 'variable', 'header': 'flags', 'para': 'OP','cell_type': 'cell_range'},
            ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'production_inventory',   'production_quantity')
  
        if item=='total_pen_constructed':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','total_pen_constructed')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).value = "=Sum(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).style = 'Calculation'
        
        if item=='investment_roll_out_flag':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            if found_state:
                span_ =self._get_span()
               
                r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
                if found_state2:
                    startpoint = '$'+ get_column_letter(8) + '$'+ str(r_index2)# D07
                    first_slice_point = get_column_letter(9) + str(r_index)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                    #loop each cell of this row D15:G15
                    #----------->
                    for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                        for cell in cellObj:
                            new_column_letter = cell.column # J
                            prev_col=column_index_from_string(new_column_letter)-1
                            prev_letter= get_column_letter(prev_col)
                            prev_targeted_cell = prev_letter + str(r_index2)
                            current_targeted_cell = new_column_letter + str(r_index2)                        
                            w_sheet['%s%s'%(new_column_letter, cell.row)].value ='=IF(AND(SUM(' + startpoint + ':' + prev_targeted_cell +')=0,' + current_targeted_cell +'>0),1,0)' 
                            w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Input'
                            w_sheet['%s%s'%(new_column_letter, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
                
        if item=='closing_inventory':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','closing_inventory')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        new_column_letter = cell.column # J
                        prev_row = cell.row-1                     
                        w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=0' 
                        w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Linkedcell'
                        #w_s  

    def _add_prodInventory_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Production and Inventory' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if 'header' in self.track_inputs['production_inventory']:
            self.track_inputs['production_inventory']['header']['row'] = row_index
           
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['cattle_per_pen_per_year','cattle_survival_rate','thousand', 
                   'dressed_weight_at_selling','num_of_feedlot_targeted_for_construction','cum_pens_under_cattle', 
                   'cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'investment_roll_out_flag','total_pen_constructed']
        #-----------------------------------
        for item in itemlist:
            if item in self._production_inventory:
                _unit =self._production_inventory[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self._production_inventory[item]['title'], self._production_inventory[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['production_inventory'][item]['row'] = row_index
                self.track_inputs['production_inventory'][item]['unit'] = _unit
                self.track_inputs['production_inventory'][item]['col'] =6 #redundand
                self.track_inputs['production_inventory'][item]['value'] = self._production_inventory[item]['value']
                
                row_index +=1 
   

       

        return row_index 
    def _populate_investmentcost_section(self, w_sheet, item ):
        r_index, c_index, found_state= self._retrieve_cell_row_colm('production_inventory','thousand')
        thousand_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
              
        if item=='cost_of_pens_constructed':
            # 
            targeted_pens_row_index, c_index, found_state_pens= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_pen_construction')
            cost_of_pen_construction_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            # 
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_pens_constructed')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(targeted_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + str(pens_targeted_cell) + '*' + cost_of_pen_construction_cell \
                                                                      + '/' + str(thousand_cell)
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General
        if item=='cost_of_land':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_required')
            total_land_required_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           

            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')
            total_land_for_pens_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_land_per_sqm')
            cost_of_land_per_sqm_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            roll_out_r_index,_, found_state_roll_out= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_land')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        investment_roll_out_flag_cell = cell.column + '$' + str(roll_out_r_index ) if found_state_roll_out else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ total_land_required_cell +'>0, MAX(' \
                                                                        + total_land_required_cell + ',' \
                                                                        + total_land_for_pens_cell + ')*' + cost_of_land_per_sqm_cell \
                                                                        + '*'+  investment_roll_out_flag_cell   + '/' \
                                                                        + thousand_cell +',0)'
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        
                        #w_s 
        if item=='investment_cost_of_land':
            cost_of_land_r_index, _, found_state_cost_of_land= self._retrieve_cell_row_colm('investment_cost','cost_of_land')
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','investment_cost_of_land')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
               
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        cost_of_land_cell = cell.column + '$' + str(cost_of_land_r_index ) if found_state_cost_of_land else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_land_cell
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        
        
        if item=='cif_cost_of_machinery':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_machinery_per_pen')
            cost_of_machinery_per_pen_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            pens_targeted_r_index, c_index, found_state_pens_targeted= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cif_cost_of_machinery')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
               
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_for_construction_cell = cell.column + '$' + str(pens_targeted_r_index ) if found_state_pens_targeted else '0'
           
                        w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_machinery_per_pen_cell +'*' \
                                                                        + pens_targeted_for_construction_cell + '/' + thousand_cell  
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
        
        if item=='investment_cost_of_buildings':            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','cost_of_building_per_sqm')
            cost_of_building_per_sqm_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')
            total_land_for_pens_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            
            roll_out_r_index,_, found_state_roll_out= self._retrieve_cell_row_colm('production_inventory','investment_roll_out_flag')
            
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','investment_cost_of_buildings')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        investment_roll_out_flag_cell = cell.column + '$' + str(roll_out_r_index ) if found_state_roll_out else '0'

                        w_sheet['%s%s'%(cell.column, cell.row)] ='='+ cost_of_building_per_sqm_cell + '*' \
                                                                    + total_land_for_pens_cell \
                                                                    + '*' + investment_roll_out_flag_cell  + '/' + thousand_cell  
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
        
         

    def _add_investmentCost_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Investment Cost (Real)' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
        
        #keep track : to insert datatimeline:---> 2021, 2022....2040 
        if 'header' in self.track_inputs['investment_cost']:
            self.track_inputs['investment_cost']['header']['row'] = row_index
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['total_land_for_pens','cost_of_land_per_sqm','cost_of_machinery_per_pen', 
                   'cost_of_building_per_sqm','total_pens','cost_of_pen_construction', 
                   'senior_debt_dynamic_parameter','initial_pens_employed','pen_cattle_density',
                   'pen_length','pen_width','pen_height',
                   'total_land_required','cost_of_pens_constructed','cost_of_land',
                   'investment_cost_of_land','cif_cost_of_machinery','investment_cost_of_buildings',
                   'investment_costs_over_run_factor']
   
        #-----------------------------------
        for item in itemlist:
            if item in self.investment_cost:
                _unit =self.investment_cost[item]['units']
                _unit = _unit.upper()
                # print(item,_unit,self._get_number_formats(_unit))
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.investment_cost[item]['title'], self.investment_cost[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['investment_cost'][item]['row'] = row_index
                self.track_inputs['investment_cost'][item]['unit'] = _unit
                self.track_inputs['investment_cost'][item]['col'] = 6# redundand
                self.track_inputs['investment_cost'][item]['value'] = self.investment_cost[item]['value'] 

                if item in self.investment_parameter_options:
                    #automaticaaly retrive related
                    self._populate_investment_parameters_in_excelrow(w_sheet, self.investment_parameter_options[item],row_index, item)

                row_index +=1
                #skipp lines
                if item in ['senior_debt_dynamic_parameter','pen_height', 'total_land_required',]:
                    row_index +=1 
       
        # options headers
        self._populate_investment_parameters_header_rows(w_sheet)
        #cost_of_pens_constructed
        return row_index

    def _populate_investment_parameters_header_rows(self, w_sheet):

         #-----------------------------------------------------------------
        r_index, _, found_state= self._retrieve_cell_row_colm('investment_cost','total_land_for_pens')
        if found_state:
            headers=[]
            for item in self.cattle_business_options.keys():
                headers.append(self.cattle_business_options[item]['heading'])
            #self._populate_investment_parameters_in_excelrow(w_sheet, headers, r_index)
        
            total_len = len(headers)-1
            first_slice_point = get_column_letter(9) + str(r_index)# D09
            second_slice_point = get_column_letter(9 + int(total_len)) + str(r_index) # D16
            #loop each cell of this row D09:D16
            #----------->
            i=0
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = headers[i] 
                    w_sheet['%s%s'%(cell.column, cell.row)].fill = PatternFill(fill_type="solid", fgColor="70c4f4")
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
                    w_sheet['%s%s'%(cell.column, cell.row)].alignment = Alignment(wrap_text=True,horizontal='left', vertical='top',)
                    i += 1
    
    
    
    def _link_feedlot_design_parameters_to_model_selected(self, w_sheet):
       
        # similar pointer but different naming
        #correct this in future
        feedlot_para = ['num_of_feedlots','length', 'width',]#  'sqm_per_cattle' ]
        investment_cost_para = ['total_pens','pen_length', 'pen_width',]# 'pen_cattle_density']
        # take min to be safe

        list_len= min(len(feedlot_para), len(investment_cost_para))
        for iter in range(list_len):
            # source data
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost',investment_cost_para[iter])
            item_cell = '=$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
           
            #copy to destination cells  
            r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters' ,feedlot_para[iter])
            if found_state:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = item_cell
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style= 'Linkedcell'
        
               

    def _update_investment_parameters(self, w_sheet):
        para_list= ['cost_of_land_per_sqm', 'cost_of_machinery_per_pen', 
            'cost_of_building_per_sqm', 'cost_of_pen_construction',
            'senior_debt_dynamic_parameter', 'initial_pens_employed', 'pen_cattle_density', 
            'pen_length', 'pen_width', 'pen_height',  'total_land_required',
        ]

        for item in  para_list:
            r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost',item)
            if found_state:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = self._get_investment_paramater_formulae(r_index)
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style= 'Calculation'
                if item in ['senior_debt_dynamic_parameter']:
                    w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format= '0.0%'

        
        # Total PEns different
        r_index, c_index, found_state= self._retrieve_cell_row_colm('investment_cost','total_pens')
        if found_state:
            span_ =self._get_span()
            r_index2, _, found_state2= self._retrieve_cell_row_colm('production_inventory','num_of_feedlot_targeted_for_construction')
            if found_state2: 
                first_slice_point = get_column_letter(9) + str(r_index2)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).value ="=SUM(" + first_slice_point + ":" + second_slice_point +")"
            w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).style = 'Calculation'
            w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                 

    def _get_investment_paramater_formulae(self, input_row_index):

        # retrive model values
        if 'outputs' in  self.track_inputs:
            if 'model_selection' in  self.track_inputs['outputs']:
                model_sel_cell  = self.track_inputs['outputs']['model_selection']['cell']

                start_col_index = 9
                formulae_part ="="

                list_= self.cattle_business_options.keys()
                #print(list_)
                closin_brackets=''
                for  i in range(len(list_)-1):
                    closin_brackets+=')'

                i = 0
               
                for item in list_:
                    #item =list_[i]
                    colHeader = get_column_letter(start_col_index + i )
                    if i ==len(list_)-1:
                        # last part of formalae
                        formulae_part += colHeader + str(input_row_index ) + closin_brackets
                    else:
                        formulae_part += 'IF(Outputs!' + model_sel_cell + '="'+ self.cattle_business_options[item]['heading'] +'",' + colHeader + str(input_row_index) + ','
                    i+=1
                #---------------------------------------------------------------
                        
        return formulae_part
    
    
    def _populate_investment_parameters_in_excelrow(self, w_sheet, options_arraylist, row_index, item):

        total_len = len(options_arraylist)-1
        first_slice_point = get_column_letter(9) + str(row_index)# D09
        second_slice_point = get_column_letter(9 + int(total_len)) + str(row_index) # D16
        #loop each cell of this row D09:D16
        #----------->
        i=0
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                w_sheet['%s%s'%(cell.column, cell.row)] = options_arraylist[i] 
                w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                if item in ['senior_debt_dynamic_parameter']:
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='0.0%'
                else:    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                i += 1

    def _populate_costs_section(self, w_sheet, item ):
                   
        if item=='cum_pens':
            # 
            cum_pens_row_index, _, found_state_pens= self._retrieve_cell_row_colm('production_inventory','cum_pens_under_cattle')
           
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','cum_pens')

            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(cum_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + str(pens_targeted_cell) 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              

        if item=='num_of_workers':

            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','num_of_workers_per_pen')
            num_of_workers_per_pen_cell = '$' + get_column_letter(c_index) + '$' + str(r_index ) if found_state else '0'
            
            targeted_pens_row_index, _, found_state_pens= self._retrieve_cell_row_colm('cost_real','cum_pens')
           
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','num_of_workers')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        pens_targeted_cell = cell.column + '$' + str(targeted_pens_row_index) if found_state_pens else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] ='=CEILING('+ pens_targeted_cell +'*'+ num_of_workers_per_pen_cell +',1)' 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              
            
        if item=='num_of_supervisors_technicians':

            workers_r_index, c_index, found_state_workers= self._retrieve_cell_row_colm('cost_real','num_of_workers')
          
            r_index, _, found_state= self._retrieve_cell_row_colm('cost_real','num_of_supervisors_technicians')
            if found_state:
                span_ =self._get_span()
                first_slice_point = get_column_letter(9) + str(r_index)# D07
                second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
                #loop each cell of this row D15:G15
                #----------->
                for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                    for cell in cellObj:
                        num_of_workers_per_pen_cell = cell.column + '$' + str(workers_r_index ) if found_state_workers else '0'
                        w_sheet['%s%s'%(cell.column, cell.row)] ='=IF(' +  num_of_workers_per_pen_cell +' >0,INT(' + num_of_workers_per_pen_cell+ '/5),0)' 
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              

        
        
        if item=='average_num_of_workers':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','average_num_of_workers')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('cost_real','num_of_workers')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).value = "=AVERAGE(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).style = 'Calculation'
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
              
        
        if item=='average_num_of_supervisors':
            r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','average_num_of_supervisors')
            if found_state:
                span_ =self._get_span()
                
                r_index2, _, found_state2= self._retrieve_cell_row_colm('cost_real','num_of_supervisors_technicians')
                if found_state2: 
                    first_slice_point = get_column_letter(9) + str(r_index2)# D07
                    second_slice_point = get_column_letter(9 + int(span_)) + str(r_index2) # D39
                    w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).value = "=AVERAGE(" + first_slice_point + ":" + second_slice_point +")"
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).style = 'Calculation'
                w_sheet.cell('%s%s'%(get_column_letter(c_index), r_index)).number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                       
                              
    def _add_costs_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Costs (Real)' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
   
        #-----------------------------------
        if not 'cost_real' in self.track_inputs:
            self.track_inputs['cost_real']= {}# insert inner
        
        self._add_sub_heading(w_sheet,'Number of workers and supervisors',row_index)
        row_index +=1

        for item in self.cost_real.keys():
            if item in self.cost_real:
                _unit =self.cost_real[item]['units']
                _unit = _unit.upper() if _unit != None else 'BLANK'
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.cost_real[item]['title'], self.cost_real[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not item in self.track_inputs['cost_real']:
                    self.track_inputs['cost_real'][item]= {}# insert inner

                self.track_inputs['cost_real'][item]['row'] = row_index
                self.track_inputs['cost_real'][item]['unit'] = _unit
                self.track_inputs['cost_real'][item]['col'] = 6# redundand
                self.track_inputs['cost_real'][item]['value'] = self.cost_real[item]['value']

                #---------------Headers--------------------
                if item =='average_num_of_supervisors':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Monthly Wages and Salaries',row_index)

                elif item =='monthly_wage_per_supervisor':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Annual Increase in real salaries',row_index)

                elif item =='annual_increase_salaries_supervisors_technicians':
                   row_index +=1 
                   self._add_sub_heading(w_sheet,'Cost of inputs per ton of Beef',row_index)

                elif item =='cost_of_electricity_per_pen_per_annum':
                   row_index +=1  
                   self._add_sub_heading(w_sheet,'Annual Price Change',row_index)

                #skip line after
                if item in ['cattle_feed_price_per_kg', 'dressed_weight_at_selling' , 
                            'total_input_costs_per_pen', 'tonnes_produced_per_pen_per_year']:
                   row_index +=1  
                # new line feed                
                row_index +=1
        #fill this place by retrievint from trackinputs-----
        self._update_cost_real_linkedcells(w_sheet)
        #----------------------------------------------------- 
        return row_index

    def _update_cost_real_linkedcells(self, w_sheet,):
        
        
        #  cattle_price_per_unit derived cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_price_per_unit')
        cattle_price_per_unit_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'


        #  cattle_per_pen_per_year derivd cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_per_pen_per_year')
        cattle_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # qnty_of_feed_per_cattle# derived cell
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','qnty_of_feed_per_cattle')
        qnty_of_feed_per_cattle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # cattle_feed_price_per_kg
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_feed_price_per_kg')
        cattle_feed_price_per_kg_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        

        # live_weight_at_selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # dressed_weight_percent
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_percent')
        dressed_weight_percent_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # cattle_survival_rate
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cattle_survival_rate')
        cattle_survival_rate = get_column_letter(c_index) + str(r_index ) if found_state else None
        

        
        # daily_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','daily_weight_gain')
        daily_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        
        # days_in_feed_lot
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','days_in_feed_lot')
        days_in_feed_lot_cell = get_column_letter(c_index) + str(r_index ) if found_state else None

        # live_weight_gain
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_gain')
        live_weight_gain_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + daily_weight_gain_cell + '*' + days_in_feed_lot_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
        
        # weight_at_purchase
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','weight_at_purchase')
        weight_at_purchase_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
        
        
        # Live Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','live_weight_at_selling')
        live_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + weight_at_purchase_cell + '+' + live_weight_gain_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
       
       

        # food_conversion_ratio
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','food_conversion_ratio')
        food_conversion_ratio_cell = get_column_letter(c_index) + str(r_index ) if found_state else None
       
       
        #qnty_of_feed_per_cattle# recalculated here
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','qnty_of_feed_per_cattle')
        qnty_of_feed_per_cattle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + food_conversion_ratio_cell + '*' + daily_weight_gain_cell + '*' + days_in_feed_lot_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
       
        # Commercial Feed Per Pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','commercial_feed_per_pen')
        commercial_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + cattle_per_pen_per_year_cell + '*' + qnty_of_feed_per_cattle_cell
            #w_sheet['%s%s'%(c, r_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)' 
        
       
       
        # Dressed Weight @ Selling
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','dressed_weight_at_selling')
        dressed_weight_at_selling_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + dressed_weight_percent_cell + '*' + live_weight_at_selling_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' 
        
      
     
     
        
        # Cost of cattle per pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_per_pen')
        cost_of_cattle_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + cattle_per_pen_per_year_cell + '*' + cattle_price_per_unit_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
         # cost_of_cattle_feed_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_cattle_feed_per_pen')
        cost_of_cattle_feed_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + cattle_feed_price_per_kg_cell + '*' + commercial_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # total_input_costs_per_pen
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','total_input_costs_per_pen')
        total_input_costs_per_pen_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + cost_of_cattle_per_pen_cell + '+' + cost_of_cattle_feed_per_pen_cell
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        # tonnes_produced_per_pen_per_year
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','tonnes_produced_per_pen_per_year')
        tonnes_produced_per_pen_per_year_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + cattle_per_pen_per_year_cell + \
                                                        '*' + dressed_weight_at_selling_cell + \
                                                        '*' +  cattle_survival_rate + '/1000' 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
 
         
        # cost_of_domestic_inputs_per_ton_of_beef_produced
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cost_real','cost_of_domestic_inputs_per_ton_of_beef_produced')
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + total_input_costs_per_pen_cell + \
                                                        '/' + tonnes_produced_per_pen_per_year_cell 
            w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            w_sheet['%s%s'%(c, r_index)].number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
        
        # dressed_weight_at_selling linkedcell @ prosuxt
        self._set_linkedcell(w_sheet, 'production_inventory','dressed_weight_at_selling', dressed_weight_at_selling_cell) 
        
           
  
    def _add_depreciation_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Depreciation' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['economic_life_of_machinery','economic_life_of_buildings',]
       
        #==========Sub-heading=========================
        # col = get_column_letter(2)
        # w_sheet.cell('%s%s'%(col, row_index)).value = 'Economic service life' 
        # w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        self._add_sub_heading(w_sheet,'Economic service life',row_index)
        row_index +=1
        #-----------------------------------
        for item in itemlist:
            if item in self.depreciation:
                _unit =self.depreciation[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.depreciation[item]['title'], self.depreciation[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['depreciation'][item]['row'] = row_index
                self.track_inputs['depreciation'][item]['unit'] = _unit
                self.track_inputs['depreciation'][item]['col'] = 6#redundant
                self.track_inputs['depreciation'][item]['value'] = self.depreciation[item]['value'] 
                row_index +=1  
        
        # Sub-Section B skipp rows
        row_index +=1 
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Recovery period for income tax' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #----------------------------------- 
        itemlist =['tax_life_of_machinery','tax_life_of_buildings','tax_life_of_soft_capital_costs']
        for item in itemlist:
            if item in self.depreciation:
                _unit =self.depreciation[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.depreciation[item]['title'], self.depreciation[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['depreciation'][item]['row'] = row_index
                self.track_inputs['depreciation'][item]['unit'] = _unit
                self.track_inputs['depreciation'][item]['col'] = 6# redundant
                self.track_inputs['depreciation'][item]['value'] = self.depreciation[item]['value'] 
                row_index +=1  

        return row_index  
    def _add_sub_heading3(self, w_sheet, title, row_index):
        col = get_column_letter(4)
        w_sheet.cell('%s%s'%(col, row_index)).value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading4'
      
    def _add_sub_heading2(self, w_sheet, title, row_index):
        col = get_column_letter(3)
        w_sheet.cell('%s%s'%(col, row_index)).value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading3'
      
    def _add_sub_heading(self, w_sheet, title, row_index):
        col = get_column_letter(2)
        w_sheet.cell('%s%s'%(col, row_index)).value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        
    def _add_workingCapital_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Working Capital' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

       

        itemlist =['accounts_receivable','accounts_payable','cash_balance']
        
        for item in itemlist:
            if item in self.working_capital:
                _unit =self.working_capital[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.working_capital[item]['title'], self.working_capital[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
               
                self.track_inputs['working_capital'][item]['row'] = row_index
                self.track_inputs['working_capital'][item]['unit'] =_unit
                self.track_inputs['working_capital'][item]['col'] =6# redundant
                self.track_inputs['working_capital'][item]['value'] = self.working_capital[item]['value'] 
                row_index +=1

        return row_index  
    
    def _write_row_title_and_value3(self, 
                                    w_sheet, write_colHeader, write_colValue, 
                                    write_colUnits,row_index, header_title , 
                                    input_value, ref_formular_unit, val_number_format= 'General', 
                                    blank_UNITS =False,  valfield_style='Explanatory Text'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet.cell('%s%s'%(write_colHeader, row_index)).value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font
        # if None skip
        if input_value != None:
            w_sheet.cell('%s%s'%(write_colValue, row_index)).value = input_value
            w_sheet['%s%s'%(write_colValue, row_index)].style= 'Input'
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
       
        if blank_UNITS==False:
            w_sheet.cell('%s%s'%(write_colUnits, row_index)).value = ref_formular_unit
            w_sheet['%s%s'%(write_colUnits, row_index)].style= valfield_style
        
    def _write_row_title_and_value4(self, 
                                    w_sheet, write_colHeader, write_colValue, 
                                    write_colUnits,row_index, header_title , 
                                    input_value, ref_formular_unit, val_number_format= 'General', 
                                    blank_UNITS =False,  valfield_style='Explanatory Text'):

        #print("inside _write_row_title_and_value4:  Numberformat>>>>>>>>>>>", val_number_format)
        #----Tis function write eader and value as excel row wit formatin
        w_sheet.cell('%s%s'%(write_colHeader, row_index)).value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font
        # if None skip
        if input_value != None:
            w_sheet.cell('%s%s'%(write_colValue, row_index)).value = input_value
        w_sheet['%s%s'%(write_colValue, row_index)].style= 'Output2'
        
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
       
        if blank_UNITS==False:
            w_sheet.cell('%s%s'%(write_colUnits, row_index)).value = ref_formular_unit
            w_sheet['%s%s'%(write_colUnits, row_index)].style= valfield_style
        
    
       
    def _add_financing_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Financing' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

       

        itemlist =['real_interest_rate','risk_premium','num_of_installments','grace_period','repayment_starts']
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Suppliers credit' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #-----------------------
        for item in itemlist:
            if item in self.financing:
                _unit =self.financing[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.financing[item]['title'], self.financing[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['financing'][item]['row'] = row_index
                self.track_inputs['financing'][item]['unit'] = _unit
                self.track_inputs['financing'][item]['col'] = 6#redundant
                self.track_inputs['financing'][item]['value'] = self.financing[item]['value'] 
                row_index +=1 
        row_index +=1
        
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Financing as a  % of Investment Costs' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #-----------------------
        for item in ['equity','senior_debt',]:
            if item in self.financing:
                _unit =self.financing[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                            self.financing[item]['title'], self.financing[item]['value'], cell_display_val, 
                            self._get_number_formats(_unit), _unit=='BLANK')

                self.track_inputs['financing'][item]['row'] = row_index
                self.track_inputs['financing'][item]['unit'] = _unit
                self.track_inputs['financing'][item]['col'] = 6#redundant
                self.track_inputs['financing'][item]['value'] = self.financing[item]['value']  

                if item=='equity':
                    w_sheet.cell('%s%s'%(colValue, row_index)).value ='=1-'+ colValue + str(row_index +1)
                    w_sheet['%s%s'%(colValue, row_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')  
                    
                #self._populate_flags_section(w_sheet, row_index, item, cell_ref_timing)
                row_index +=1      
            
       
        row_index +=1

        

        return row_index      
    

     
    def _write_row_title_and_value2(self, w_sheet, write_colHeader, write_colValue, 
        row_index, header_title,value, val_number_format= 'General', valfield_style='Explanatory Text'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet.cell('%s%s'%(write_colHeader, row_index)).value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font

        w_sheet.cell('%s%s'%(write_colValue, row_index)).value = value
        w_sheet['%s%s'%(write_colValue, row_index)].style= valfield_style
        #w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        #w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
    
    def _update_section_header_year(self, w_sheet, cat_section, flags_section_row_index,
                                   start_col_index, span_ ,source_w_sheet = None):
        #The function inserts years copied from Flas Section 
        #Do noting if no history is found
        if cat_section in self.track_inputs:
            if 'header' in self.track_inputs[cat_section]:
                start =start_col_index
                end = start_col_index + int(span_)
                if 'row' in self.track_inputs[cat_section]['header']:
                    header_row_= self.track_inputs[cat_section]['header']['row']
                    first_slice_point = get_column_letter(start) + str(header_row_)# D07
                    second_slice_point = get_column_letter(end) + str(header_row_) # D39
                    #loop each cell of this row D9:D37
                    #----------->
                    appended_worksheet = f'{source_w_sheet}!' if source_w_sheet != None else '' # =Input!AF10
                    for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                        for cell in cellObj:
                            #get column
                            new_column_letter = cell.column # J
                            w_sheet['%s%s'%(new_column_letter, cell.row)] =  '='+ appended_worksheet + '$'+ new_column_letter + '$'+ str(flags_section_row_index) 
                            w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Heading1'
                            w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General' 
                    
                    BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  header_row_, start, end)
        
    def _populate_flags_section(self, w_sheet, row_index, item, cell_ref_timing):
        if item=='YEAR':
            # set row----
            self.track_inputs['flags']['years']['row']=row_index
            span_ =self._get_span()
            start_col_index= 9
            first_slice_point = get_column_letter(start_col_index) + str(row_index)# D07
            second_slice_point = get_column_letter(start_col_index + int(span_)) + str(row_index) # D39
            
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                    prev_col=column_index_from_string(new_column_letter)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_cell = prev_letter + str(cell.row)
                   
                   
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF('+ prev_cell + ',' + prev_cell + '+1,' + cell_ref_timing['base_period'] + ')'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
            #
            #self._update_production_inventory_header_year(w_sheet, row_index, start_col_index, span_)
            self._update_section_header_year(w_sheet, 'production_inventory', row_index, start_col_index, span_)
            self._update_section_header_year(w_sheet, 'investment_cost', row_index, start_col_index, span_)
        if item=='PS':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '=' + cell_ref_timing['base_period'] + ',1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='CP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '>' + cell_ref_timing['construction_year_end'] + ',0,1)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='LPP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(OR(' + years_cell + '<' + cell_ref_timing['repayment_starts'] + \
                                                                  ',' + years_cell +'>('+ cell_ref_timing['repayment_starts'] + \
                                                                  '+' + cell_ref_timing['num_of_installments'] + '-' + cell_ref_timing['grace_period'] + '-1)),0,1)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='OP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(AND(' + years_cell + '>=$' + cell_ref_timing['operation_start_year'] + \
                                                                  ',' + years_cell +'<$'+ cell_ref_timing['operation_end']  + '),1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    
        if item=='RES':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '=$' + cell_ref_timing['operation_end']  + ',1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                  
  


    def _add_flags_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Flags' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(7)
        row_index +=1

        financing = {
            'real_interest_rate': {'title': 'Real interest rate','value': 6, 'units': 'PERCENT'}, 
            'risk_premium': {'title': 'Risk premium','value': 1, 'units': 'PERCENT'}, 
            'num_of_installments': {'title': 'No. of installments','value': 6, 'units': 'NUMBER'},  
            'grace_period': {'title': 'Grace period (years)','value': 0, 'units': 'YEARS'}, 
            'repayment_starts': {'title': 'Repayment starts year','value':2022, 'units': 'YEAR'}, 
            'equity': {'title': 'Equity (% of Investment Costs)','value': 30, 'units': 'PERCENT'}, 
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': .70, 'units': 'PERCENT'},  
                
        }
       
        # retrieve reference
        base_period_row = self.track_inputs['timing']['base_period']['row']
        base_period_col = self.track_inputs['timing']['base_period']['col']
        itemlist =['YEAR','PS','CP','LPP','OP','RES']

        base_period_ref = get_column_letter(base_period_col) + str(base_period_row) # F32

        construction_year_end_row = self.track_inputs['timing']['construction_year_end']['row']
        construction_year_end_ref = '$' + get_column_letter(base_period_col) + '$' + str(construction_year_end_row) # $F$32
  
        cell_ref_timing={}
        cell_ref_timing= {'base_period':base_period_ref, 'construction_year_end':construction_year_end_ref}
        cell_ref_timing['repayment_starts'] = get_column_letter(base_period_col) + str(self.track_inputs['financing']['repayment_starts']['row'])
        cell_ref_timing['num_of_installments'] =  get_column_letter(base_period_col) + str(self.track_inputs['financing']['num_of_installments']['row']) 
        cell_ref_timing['grace_period'] =  get_column_letter(base_period_col) + str(self.track_inputs['financing']['grace_period']['row'])         
        cell_ref_timing['operation_start_year'] = get_column_letter(base_period_col) + str(self.track_inputs['timing']['operation_start_year']['row'])
        cell_ref_timing['operation_end'] =  get_column_letter(base_period_col) + str(self.track_inputs['timing']['operation_end']['row']) 
      


        for item in self.flags.keys():#itemlist:
            if item in self.flags:
                flag_unit =self.flags[item]['units']
                flag_unit = flag_unit.upper()
                cell_display_val = self.modelspec_cell_ref[flag_unit] if  flag_unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value2(w_sheet, colHeader, colValue, row_index, 
                        self.flags[item]['title'],cell_display_val)

                #dynamic declare variables
                if not (item in self.track_inputs['flags']):
                    self.track_inputs['flags'][item] ={}

                self.track_inputs['flags'][item]['row'] = row_index
                self.track_inputs['flags'][item]['unit'] = flag_unit
                self.track_inputs['flags'][item]['col'] = 6
                self.track_inputs['flags'][item]['value'] = None      

                self._populate_flags_section(w_sheet, row_index, item, cell_ref_timing)
                 
                #skipp 2 line after YEAR LINE 
                if item == 'YEAR':
                    #update production & inventory year       
                    row_index +=2
                else:
                   row_index +=1  
       
        return row_index      

     
    def _add_taxes_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Taxes' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['import_duty','sales_tax','corporate_income_tax']
        
        for item in itemlist:
            if item in self.taxes:
                _unit =self.taxes[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.taxes[item]['title'], self.taxes[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['taxes'][item]['row'] = row_index
                self.track_inputs['taxes'][item]['unit'] = _unit
                self.track_inputs['taxes'][item]['col'] = 6#redundant
                self.track_inputs['taxes'][item]['value'] = self.taxes[item]['value'] 
                row_index +=1 

        return row_index      

    def _add_macroeconomicParameters_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Macroeconomic Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['discount_rate_equity','domestic_inflation_rate',
                    'us_inflation_rate','exchange_rate','dividend_payout_ratio','num_of_shares']
        
        for item in itemlist:
            if item in self.macroeconomic_parameters:
                _unit =self.macroeconomic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.macroeconomic_parameters[item]['title'], self.macroeconomic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                self.track_inputs['macroeconomic_parameters'][item]['row'] = row_index
                self.track_inputs['macroeconomic_parameters'][item]['unit'] = _unit
                self.track_inputs['macroeconomic_parameters'][item]['col'] = 6#redundant
                self.track_inputs['macroeconomic_parameters'][item]['value'] = self.macroeconomic_parameters[item]['value'] 
                row_index +=1  

        return row_index

    def _update_feedlot_linkedcells(self, w_sheet,):
        
        # cattle_survival_rate
        # r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','cattle_survival_rate')
        # cattle_survival_rate = get_column_letter(c_index) + str(r_index ) if found_state else None
        # num_of_months_per_cycle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','num_of_months_per_cycle')
        num_of_months_per_cycle_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # num_of_feedlots
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','num_of_feedlots')
        num_of_feedlots_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # length
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','length')
        length_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # width
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','width')
        width_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'
        
        # pen_area
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','pen_area')
        pen_area_cell = get_column_letter(c_index) + str(r_index ) if found_state else '0'

        # total_cattle_per_pen_per_cycle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','total_cattle_per_pen_per_cycle')
        cattle_per_cycle = get_column_letter(c_index) + str(r_index ) if found_state else .00001
        
        # SQM per cattle
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','sqm_per_cattle')
        if found_state:
            c= get_column_letter(c_index)
            w_sheet.cell('%s%s'%(c, r_index)).value ='=' + pen_area_cell + '/' + cattle_per_cycle
            w_sheet['%s%s'%(c, r_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')  
        
        
        cal_val= '='  + str(pen_area_cell) + '/' + cattle_per_cycle
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','sqm_per_cattle', cal_val, "Calculated")

        # SQM Calculation
        cal_val = '=' + num_of_feedlots_cell + '*' + length_cell + '*' + width_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','sqm', cal_val, 'Calculated')
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','sqm')
        sqm_cell = None
        if found_state:
            sqm_cell = get_column_letter(c_index) + str(r_index ) 
          

        # PEN Calculation
        cal_val = '=' + length_cell + '*' + width_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','pen_area',cal_val, 'Calculated')
        
        # number_of_months_in_a_year
        r_index , c_index, found_state= self._retrieve_cell_row_colm('timing','number_of_months_in_a_year')
        if found_state:
            number_of_months_in_a_year= get_column_letter(c_index) + str(r_index)
        else:
            number_of_months_in_a_year = 12   
        
        # Cattle_per_pen_per_year Calculation
        r_index, c_index, found_state= self._retrieve_cell_row_colm('feedlot_design_parameters','cattle_per_pen_per_year')
        cattle_per_pen_per_year = get_column_letter(c_index) + str(r_index ) if found_state else None
        
        cal_val= '=' + cattle_per_cycle + '*' + str(number_of_months_in_a_year) +   '/' + num_of_months_per_cycle_cell
        self._set_linkedcell(w_sheet, 'feedlot_design_parameters','cattle_per_pen_per_year', cal_val, "Calculated")

        # total_land_for_pens
        self._set_linkedcell(w_sheet, 'investment_cost','total_land_for_pens', sqm_cell)
       
       
        self._set_linkedcell(w_sheet, 'production_inventory','cattle_per_pen_per_year', cattle_per_pen_per_year)
        
        self._set_linkedcell(w_sheet, 'cost_real','cattle_per_pen_per_year', cattle_per_pen_per_year)
        #derived from Sens!
        #self._set_linkedcell(w_sheet, 'production_inventory','cattle_survival_rate', cattle_survival_rate, )
        """ 
        itemlist =['cattle_per_pen_per_year','cattle_survival_rate','thousand', 
                   'dressed_weight_at_selling','num_of_feedlot_targeted_for_construction','cum_pens_under_cattle', 
                   'cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'investment_roll_out_flag','total_pen_constructed']
         """
    def _set_linkedcell(self, w_sheet, header, subheader, 
                       val, linked_Style='Linkedcell', format_style= "",):
                       
        # retrieve tracked parmameters
        r_index, c_index, found_state= self._retrieve_cell_row_colm(header,subheader)
        #if found change details else leave the default valuew
        if found_state:
            c= get_column_letter(c_index)
            if val != None:
                w_sheet.cell('%s%s'%(c, r_index)).value ='=' +   val
            if  linked_Style =='Linkedcell':
                #linked  
                w_sheet.cell('%s%s'%(c, r_index)).style = linked_Style
            else:
                #calculated
                w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            
            if format_style != "":
                w_sheet.cell('%s%s'%(c, r_index)).number_format= self._get_number_formats(format_style) 
        
    def _add_feedlotDesignParameters_section(self, w_sheet, row_index,total_wsheet_cols):
        #overide all add tis to class
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Pen/Feedlot Design Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['num_of_feedlots','length','width','sqm','pen_area',
                   'sqm_per_cattle','total_cattle_per_pen_per_cycle',
                   'num_of_months_per_cycle','cattle_per_pen_per_year']
        
        for item in itemlist:
            if item in self.feedlot_design_parameters:
                _unit =self.feedlot_design_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.feedlot_design_parameters[item]['title'], self.feedlot_design_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')

                  

                self.track_inputs['feedlot_design_parameters'][item]['row'] = row_index
                self.track_inputs['feedlot_design_parameters'][item]['unit'] = _unit
                self.track_inputs['feedlot_design_parameters'][item]['col'] = 6#redundant
                self.track_inputs['feedlot_design_parameters'][item]['value'] = self.feedlot_design_parameters[item]['value'] 
                
                row_index +=1
                #skipp empty spaces 
                if item in ['sqm','pen_area','sqm_per_cattle','total_cattle_per_pen_per_cycle',]:
                    row_index +=1
                   
        
        #fill this place by retrievint from trackinputs-----
        self._update_feedlot_linkedcells(w_sheet)
        #-----------------------------------------------------
    
        #----leave blank for now: update later in Outputs
        self.track_inputs['cattle_business_options']['header']['row'] = row_index
        row_index +=1
        #-----------------------------------------------------------
        for item in self.cattle_business_options.keys():
            w_sheet.cell('%s%s'%(colHeader, row_index)).value = self.cattle_business_options[item]['description']
            w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
            self.track_inputs['cattle_business_options'][item]['row'] = row_index
            self.track_inputs['cattle_business_options'][item]['heading'] = self.cattle_business_options[item]['heading'] 
            self.track_inputs['cattle_business_options'][item]['description'] = self.cattle_business_options[item]['description']            
            row_index +=1
        
        #---------------hide text for business options------------------------------------
        option_total= len(self.cattle_business_options.keys())
        start_index= row_index - option_total-1
        BaseBusinessReport._hide_rows(self, w_sheet,start_index, row_index) 
        #----------------------------------------------- 
       
        return row_index
    def _retrieve_cell_row_colm(self, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:

                if 'row' in self.track_inputs[header_par][sub_header]:
                    _row_index = self.track_inputs[header_par][sub_header]['row'] 
                if 'col' in self.track_inputs[header_par][sub_header]:
                    _col_index = self.track_inputs[header_par][sub_header]['col']

                found_state=True
               
        return _row_index, _col_index, found_state  

    def _retrieve_cell_value(self, w_sheet, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:
                _row_index = self.track_inputs[header_par][sub_header]['row'] 
                _col_index = self.track_inputs[header_par][sub_header]['col']
                found_state=True
                col_Letter = get_column_letter(_col_index)
                return  w_sheet.cell('%s%s'%(col_Letter, _row_index)).value, found_state 
        return  _,  found_state

    def _retrieve_cell_formuale(self, w_sheet, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:
                _row_index = self.track_inputs[header_par][sub_header]['row'] 
                _col_index = self.track_inputs[header_par][sub_header]['col']
                found_state=True
                col_Letter = get_column_letter(_col_index)
                #A14
                return  col_Letter + str(_row_index ),found_state
        return  _, found_state     
                
    def _get_input_memory_details(self, input_parameter, search_start=0):
        #change_in_prices only ins prices
        # excludes sens
        #in future only unique names allowed
        intemlist = ['timing','prices','flags','financing','depreciation',
                    'working_capital','taxes','macroeconomic_parameters', 
                    'feedlot_design_parameters', 'production_inventory','cost_real', 'investment_cost']
        flag_found = False
        input_parameter_header = None
        input_parameter_val  =  None
        input_parameter_unit =  None
        item_instance = None
        
        # return the first instance found
        # in future retun all instances as list
        for iter in range(search_start, len(intemlist)):
            item =intemlist[iter]
            if input_parameter in self.track_inputs[item]:
                input_parameter_row = self.track_inputs[item][input_parameter]['row']# col F==>6
                # add value
                input_parameter_val = self.track_inputs[item][input_parameter]['value']
                # add value
                input_parameter_unit = self.track_inputs[item][input_parameter]['unit']
                #found and loop out
                colE = get_column_letter(5)
                input_parameter_header = '=Inputs!$' + colE + '$' + str(input_parameter_row)
                flag_found= True
                # early return
                item_instance= item
                #set pointer to next element in future search
                next_element = iter +1
                return item_instance, input_parameter_header,input_parameter_val,input_parameter_unit, flag_found, next_element
        next_element =len(intemlist)
        return item_instance, input_parameter_header,input_parameter_val,input_parameter_unit, flag_found, next_element

    def _check_item_dict_in_list(self, item_instance, item , exclusion_list):
        for list_item in exclusion_list:
            if item_instance ==list_item['item_instance'] and item == list_item['item']:
                return False# Exclude  
        return True# Procede
    def _add_sensitivity_section(self, wb, w_sheet, row_index,total_wsheet_cols):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Sensitivity Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(2)
        colValue = get_column_letter(3)
        colUnits = get_column_letter(4)
        row_index +=1
       
        # Domestic Inflation rate	5%	%
        # US Inflation rate	3%	%
        # Change in real price of Fish	-  	%
        # Investment Costs Over-run Factor	-  	%
        # Accounts receivable [% of Gross Sales]	10%	%
        # Acccounts payable (% of total input cost)	10%	%
        # Cash balance  (% of gross sales)	10%	%
        # Cattle Survival rate	100%	%
        # Exchange Rate (LC/USD - year 2018)	1.4	#
        # Senior Debt (% of Investment Costs) 	70%	%
        # Total pens 	5 	#
        # Base price of beef in 2020 per ton 	4,500	LC
        # Total Cattle in one pen per cycle 	20 	#
        # Cattle purchase price per unit 	300	LC
        # Scale 	Moderate Scale
        itemlist =['domestic_inflation_rate','us_inflation_rate','change_in_price', 'accounts_receivable',
                   'accounts_payable','cash_balance','exchange_rate', 'senior_debt','base_price',
                   'total_cattle_per_pen_per_cycle' ,'cattle_survival_rate', 'cattle_price_per_unit',
                   'initial_pens_employed']
        update_exclusion_list= []#[{'item_instance': 'investment_cost','item':'initial_pens_employed'},]
        exclude_item= {}
        exclude_item['item_instance']='investment_cost'
        exclude_item['item']='initial_pens_employed'
        update_exclusion_list.append(exclude_item)
        for item in itemlist:
            # each item in list can be retrieved more than one
            #loop until nothing is found
            iter = 0 # for each ele start at zero
            continue_search = True # default initialisation
            tag_ = 0
            while continue_search:
                item_instance, header_,val_, _unit, flag_found, iter=self._get_input_memory_details(item,iter)
                continue_search = flag_found
            
                if flag_found:
                    tag_ +=1# increament # of counts
                    cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                    self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                            header_, val_, cell_display_val, self._get_number_formats(_unit), _unit=='BLANK')
                    
                    #dynamic declare variables
                    if not (item in self.track_inputs['sens']):
                       self.track_inputs['sens'][item] ={}

                    self.track_inputs['sens'][item]['row'] = row_index
                    self.track_inputs['sens'][item]['unit'] = _unit
                    self.track_inputs['sens'][item]['col'] = 3
                    #self.track_inputs['sens'][item]['value'] = row_index
                    
                    #----dynamic update of Inputvariable:
                    #  price, us infaltion , change_in_price
                    if item_instance in self.track_inputs  :
                        if item in self.track_inputs[item_instance] and self._check_item_dict_in_list(item_instance,item, update_exclusion_list):
                            _row = self.track_inputs[item_instance][item]['row'] 
                            _col = self.track_inputs[item_instance][item]['col']
                            col = get_column_letter(_col)
                            wb["Inputs"].cell('%s%s'%(col, _row)).value = "=Sens!$" + colValue + "$" + str(row_index)
                            #print(col, _row,"=Sens!$" + colValue + "$" + str(row_index))
                            #if not set change to linkedcell type
                            wb["Inputs"].cell('%s%s'%(col, _row)).style = 'Linkedcell'
                            wb["Inputs"].cell('%s%s'%(col, _row)).number_format= self._get_number_formats(_unit)

            #iter for writing the sens only {not related to how many instance are found per iteration}
            row_index +=1

            if tag_>1:
                pass
                #print(item, tag_)

        return row_index 

    def _add_sensitivity_x_section(self, w_sheet, row_index,total_wsheet_cols, 
                                         header_title,  track_input_item, val_list, index_base_value_border, index_npv_zero_border,
                                         first_col_num_format='General'):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = header_title
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not track_input_item in self.track_inputs:
            self.track_inputs[track_input_item]= {}# insert inner

        if not ('header' in self.track_inputs[track_input_item]):
            self.track_inputs[track_input_item]['header'] = {}
  
        if 'header' in self.track_inputs[track_input_item]:
            self.track_inputs[track_input_item]['header']['row'] = row_index

      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index  +=1 # placeholders
        #===================================
      
        row_index +=1
        
         #aaaaa
        npv = self._get_cell_ref_asformulae('cf_real','NPV',"CF",None)
        irr = self._get_cell_ref_asformulae('cf_real','IRR',"CF",None)
        mirr = self._get_cell_ref_asformulae('cf_real','MIRR',"CF",None)
        
        lb, _, foundstate= self._get_loan_principal_repayment_bounds()
        if  not foundstate:
           lb =0

         
        ADSCR_3 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 3-1)
        ADSCR_4 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 4-1)
        ADSCR_5 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 5-1)
       
        LLCR_2 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 2-1)
        LLCR_3 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 3-1)
        LLCR_4 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 4-1)
        LLCR_5 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 5-1)
       
       
        data = []
        data.append(["","","",  npv, ADSCR_3, ADSCR_4,ADSCR_5,LLCR_2, LLCR_3, LLCR_4, LLCR_5, irr, mirr])
        #get  item_trcked, us list
        #rrrrrrrrrrrrrrrrrrrrrrrrrrrrr
        #convert sens_var ======> var
        sens_var= track_input_item.split("sens_",1)[1].strip()
        #i-way run sensitvity analysis 
        if hasattr(self,'sens_dict'):
            if sens_var in self.sens_dict.keys():
                df = self.sens_dict[sens_var]['df']
                x_npv_0 = self.sens_dict[sens_var]['x_npv_0']                
                print('Retrieving sensitivity...', sens_var)
            else:
                print('Getting sensitivity...', sens_var)
                df_sens,x_npv_0= get_para_data_table_sensitivity(self,sens_var,val_list)
                df= df_sens.copy()
        else:
            print('Getting sensitivity...', sens_var)
            df_sens, x_npv_0= get_para_data_table_sensitivity(self,sens_var,val_list)
            df= df_sens.copy()
        #print('Excell:DF list length:',len(val_list),len(df))
        for i in range(min(len(val_list),len(df))):
            
            data.append(["",        "",        val_list[i],
                        df.iloc[i]['npv'],     df.iloc[i]['adscr-3'],  df.iloc[i]['adscr-4'],
                        df.iloc[i]['adscr-5'], df.iloc[i]['llcr-2'],   df.iloc[i]['llcr-3'],
                        df.iloc[i]['llcr-4'],  df.iloc[i]['llcr-5'],   df.iloc[i]['irr'],
                        df.iloc[i]['mirr']]) 

        # add column headings. NB. these must be strings
        w_sheet.append(["","","","","","",""])
        w_sheet.append(["","","Values", 'NPV', 'ADSCR-3','ADSCR-4','ADSCR-5','LLCR-2','LLCR-3','LLCR-4','LLCR-5','IRR', 'MIRR'])
      
        #row_index +=2
        start_row = row_index
       
        for row in data:
            w_sheet.append(row)
            row_index +=1
        
        end_row= row_index 
       
       # format table
        first_slice_point = get_column_letter(3) + str(start_row + 1 )# skip heeder and first row 
        second_slice_point = get_column_letter(3 + int(10)) + str(end_row) # D39
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                
               # w_sheet['%s%s'%(cell.column, cell.row)] = row[i]
                if cell.column in ['L', 'M']:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
        
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='0.00%'

                    
                elif cell.column in ['C']:
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = first_col_num_format
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
        
                else:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
                    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                   
        
        #---Current Model Value Setborder Outline------------------------
        current_header_row_index = self.track_inputs[track_input_item]['header']['row'] 
        col_letter_start= 'C'
        col_letter_end= 'M'
        #Border of base Row
        if int(index_base_value_border) >0:
            border_start_row = current_header_row_index + index_base_value_border + 3
            border_start_end_row= border_start_row
            _range=f"{col_letter_start}{border_start_row}:{col_letter_end}{border_start_end_row}"
            self._set_border(w_sheet,_range)
        #--------------------------------------
        #print('index_npv_zero_border',index_npv_zero_border, index_base_value_border )
        if int(index_npv_zero_border) >0:
            #Border of NPV ==Zero
            border_start_row = current_header_row_index + index_npv_zero_border + 3
            border_start_end_row= border_start_row
            _range=f"{col_letter_start}{border_start_row}:{col_letter_end}{border_start_end_row}"
            self._set_border_yellow(w_sheet,_range)
        #--------------------------------------

        return row_index
    def _dummy(self,w_sheet, row_index):
        
        tab = Table(displayName="Table1", ref="C" + str(row_index+3) +":G"+ str(row_index+3 +5))
        for cellObj in w_sheet["D" + str(row_index+3+1) +":G"+ str(row_index+3 +5)]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
               
                latest_row = cell.row
                cell_ref = new_column_letter + str(cell.row)
                cell_row_val = "C" + str(cell.row)
                #Top row-------
                cell_orig= new_column_letter + str(19)
                #formulae_ ="=Product(C14,[@value])"
                w_sheet[ cell_ref] = '=TABLE(,C14)'#Translator(f'=Product($C$14,{cell_row_val})', origin=cell_orig).translate_formula(cell_ref)
                w_sheet[ cell_ref] ='{=TABLE(C14;)}'
        #------------------------------

      
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
      
        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unque through out defined names and all other table name. 
        '''
       
        from openpyxl.utils import range_boundaries 
        coord = list(range_boundaries(tab.ref))
        #print("coordinats:",coord)
        data = [["","",200,  '=TABLE(,C14)',  '=TABLE(,C14)',  '=TABLE(,C14)',  '=TABLE(,C14)']]
        for row in data:
            coord[-1] += 1  # <= No need to use max_row, just add one line!
            w_sheet.append(row)

        tab.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"

        #print(tab)
        """Change table range"""
        # ws.tables['MyTable'].ref = "A1:E5"
       
       
       
       
       
        # max_row = 10#sens_sheet.max_row
        # tab = w_sheet._tables["Table1"]
        # tab.ref = f"A1:{w_sheet.max_column}{w_sheet.max_row}"

        # for i in w_sheet.iter_rows(min_row=2, max_row=max_row):
        #     dynamic_cell = i[9]
        #     if dynamic_cell  == 'ice cream':
        #          data = [
        #             ['Apples', 10000, 5000, 8000, 6000],
        #             ['Pears',   2000, 3000, 4000, 5000],
        #             ['Bananas', 6000, 6000, 6500, 6000],
        #             ['Oranges',  500,  300,  200,  700],
        #         ]
        #         # perform some operation
    def _update_sensitivity(self, sens_sheet ):

        #==========purchase_price_section =========================
        row_index = 0
        if 'sens_parameters' in self.track_inputs:
            if 'endpoint' in self.track_inputs['sens_parameters']:
                row_index =self.track_inputs['sens_parameters']['endpoint']['row'] 

       
        #ddddd
        item_list =[
               {'title':"Cattle Purchase Price Per Unit", "track_input_item": "sens_cattle_price_per_unit",'base_val': self.cattle_price_per_unit, 'number_format':"NUMBER"},
               {'title':"Base Price",                      "track_input_item": "sens_base_price",          'base_val': self.base_price,             'number_format':"NUMBER"},
               {'title':"Change in Price [Beef]",                 "track_input_item": "sens_change_in_price",      'base_val': self.change_in_price,       'number_format':"PERCENT"},

               {'title':"Domestic Inflation Rate",           "track_input_item": "sens_domestic_inflation_rate", 'base_val': self.domestic_inflation_rate, 'number_format':"PERCENT"},
               {'title':"Exchange Rate", "track_input_item": "sens_exchange_rate"                      ,         'base_val': self.exchange_rate,            'number_format':"NUMBER"},
               
               {'title':"Initial Pens Employed",               "track_input_item": "sens_initial_pens_employed" , 'base_val': self.initial_pens_employed, 'number_format':"INTEGER"},
               
               {'title':"Senior Debt",               "track_input_item": "sens_senior_debt",           'base_val': self.senior_debt,            'number_format':"PERCENT"},
               {'title':"Cattle Survival rate",      "track_input_item": "sens_cattle_survival_rate" , 'base_val': self.cattle_survival_rate,   'number_format':"PERCENT"},
               {'title':"Us Inflation rate",         "track_input_item": "sens_us_inflation_rate",     'base_val': self.us_inflation_rate,      'number_format':"PERCENT"},

               
               {'title':"Accounts Receivable",            "track_input_item":  "sens_accounts_receivable",            'base_val': self.accounts_receivable,            'number_format':"PERCENT"},
               {'title':"Accounts Payable",                "track_input_item": "sens_accounts_payable",               'base_val': self.accounts_payable,               'number_format':"PERCENT"},
               {'title':"Cash Balance",                   "track_input_item":  "sens_cash_balance",                   'base_val': self.cash_balance,                   'number_format':"PERCENT"},
               {'title':"Total Cattle per pen per cycle", "track_input_item":  "sens_total_cattle_per_pen_per_cycle", 'base_val': self.cattle_pen_cycle, 'number_format':"INTEGER"},
               
               {'title':"Discount Rate Equity",            "track_input_item":  "sens_discount_rate_equity",            'base_val': self.discount_rate_equity,            'number_format':"PERCENT"},
               {'title':"Dressed Weight @ Selling",                "track_input_item": "sens_dressed_weight_at_selling",               'base_val': self.dressed_weight_at_selling,               'number_format':"NUMBER"},
               {'title':"Other Indirect Costs",                   "track_input_item":  "sens_other_indirect_costs",                   'base_val': self.other_indirect_costs,                   'number_format':"NUMBER"},
               {'title':"Annual Change in price of Domestic inputs", "track_input_item":  "sens_annual_change_in_price_of_domestic_inputs", 'base_val': self.annual_change_in_price_of_domestic_inputs, 'number_format':"PERCENT"},
               {'title':"Annual Change in price of Imported inputs", "track_input_item":  "sens_annual_change_in_price_of_imported_inputs", 'base_val': self.annual_change_in_price_of_imported_inputs, 'number_format':"PERCENT"},
          
       
               {'title':"Monthly wage per worker",                    "track_input_item":  "sens_monthly_wage_per_worker",            'base_val': self.monthly_wage_per_worker,            'number_format':"NUMBER"},
               {'title':"Monthly wage per supervisor",                "track_input_item":    "sens_monthly_wage_per_supervisor",               'base_val': self.monthly_wage_per_supervisor,               'number_format':"NUMBER"},
               {'title':"Annual increase salaries workers",             "track_input_item":  "sens_annual_increase_salaries_workers",                   'base_val': self.annual_increase_salaries_workers,                   'number_format':"PERCENT"},
               {'title':"Annual increase salaries Supervisors and Technicians", "track_input_item":  "sens_annual_increase_salaries_supervisors_technicians", 'base_val': self.annual_increase_salaries_supervisors_technicians, 'number_format':"PERCENT"},
               {'title':"Number of workers per Supervisor", "track_input_item":  "sens_num_workers_per_supervisor", 'base_val': self.num_workers_per_supervisor, 'number_format':"INTEGER"},
          
            
            ] 

        parameter_unit_dict ={}
        for item_x in item_list:
            c =item_x['track_input_item']
            title =item_x['title']
            var= c.split("sens_",1)[1].strip()
            parameter_unit_dict[var]={'number_format':item_x['number_format'],'title':title}
        setattr(self,'parameter_unit_dict', parameter_unit_dict)
      
        # order one list using sequence of another list:::: 
        if hasattr(self,'para_list_by_grad'):
            #chop vars with no gradient.....
            item_list= self._sort_by_another_list(item_list,self.para_list_by_grad)[:len(self.para_list_by_grad)]
        #
        iter_i =0 
        for item in item_list:# Order starting with most senstive parameter
            iter_i +=1
            #skipp 2 rows
            row_index += 2
            #-----set independent variables---------
            if 'base_val' in item:
               val = item['base_val']
            else:
                val= 10
            #--------------------------------------
            #general parameter range
            list_, indexof_base ,indexof_npv_zero = self._sen_parameter_generator(val,item['track_input_item']) 
            #print(list_, index)
            #---------------Parameter Formating----------------------------------
            if 'number_format' in item:
                if item['number_format']=="PERCENT":
                    number_format ='0.0%'
                elif item['number_format']=="NUMBER":
                    #decimal
                    number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                elif item['number_format']=="INTEGER":
                    #integer---
                    number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                else:
                    number_format ='General' 
            else:
                number_format ='General'
            #--------------------------------------------------------------------     
            header_title =str(iter_i) +'. ' + item['title']
            row_index = self._add_sensitivity_x_section(sens_sheet, row_index, 15, 
                                     header_title,  item['track_input_item'],list_,indexof_base, indexof_npv_zero, number_format)
    def _sort_by_another_list(self,test_dict, orderlist):
        ordering_list, first_list=self._get_order_index_as_list(test_dict,orderlist)
        #print('indexlist', indexlist)          
        zipped_lists = zip(ordering_list, first_list) 
        #print('zipped:', zipped_lists[0])
        sorted_zipped_lists = sorted(zipped_lists)
        #sorted_zipped_lists = sorted(test_dict)
         
        test_list = [element for _, element in sorted_zipped_lists]
        new_dict_list =[]
        #[4,5,7,10,1,3,5,2,6,8,9] 
     
        for item in test_list:
            #item =test_list[i]
            item_dict_ = test_dict[item]
            new_dict_list.append(item_dict_)
          
        return  new_dict_list   
    def _get_order_index_as_list(self, test_list_, list_in_desired_order):       
        
        new_list=[]
        first_list=[]  
        for i in range(len(test_list_)):
            #get item as sens_*******
            c= test_list_[i]['track_input_item']
            # strip tail
            varr= c.split("sens_",1)[1].strip()
            if varr in list_in_desired_order:
                ind_ =list_in_desired_order.index(varr)
                #testlist:  [a,b,c, f]
                #order_list [c,d,a]
                #new_list   [2,5,1,5]
                #get index
                new_list.append(ind_)
                first_list.append(i)
            else:
                #append at the end
                new_list.append(len(test_list_))
                 #new_list   [1,2,3,4]
                first_list.append(i)
        
        return new_list, first_list

        
    def _sen_parameter_generator(self, point_x, var_x):
        list_ =[]
        if hasattr(self,'sens_input_dict'):
            sens_var= var_x.split("sens_",1)[1].strip()
            #print('underesss ',sens_var)
            #print(self.sens_input_dict)
           
            #print(sens_var, ' is in sens_input_dict ???', sens_var in self.sens_input_dict.keys())
            if sens_var in self.sens_input_dict.keys():
                list_=self.sens_input_dict[sens_var]['list']
       
       
        #check x_npv_0 in included-------
        x_npv_0=None
        included=-1
        sens_var= var_x.split("sens_",1)[1].strip()
        if hasattr(self,'sens_dict'):
            if sens_var in self.sens_dict.keys():
                if 'x_npv_0' in self.sens_dict[sens_var].keys():
                    x_npv_0 = self.sens_dict[sens_var]['x_npv_0'] 
                if 'included' in self.sens_dict[sens_var].keys():
                    included = self.sens_dict[sens_var]['included'] 
        
        index_npv_zero=-1
        if x_npv_0!=None and included!=-1:
            if not (x_npv_0 in list_):
                    list_.append(x_npv_0)
            
            if x_npv_0!=point_x:
                index_npv_zero = list_.index(x_npv_0) + 1
                
        list_.sort()
        if hasattr(self,'sens_dict'):
            if sens_var in self.sens_dict.keys():
                list_= self.sens_dict[sens_var]['df'][sens_var].tolist()
                #print(list_)
                if x_npv_0!=None and included!=-1:
                    index_npv_zero = list_.index(x_npv_0) + 1
                
        #get index of base value
        if  (point_x in list_):
            index_of_base_value = list_.index(point_x) + 1
        else:
            index_of_base_value=-1

        # 
       
        return list_, index_of_base_value,index_npv_zero
    def colnum_string(n):
        # Similar to column_index_from_string 
        # 5 returns E
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string 

    def _update_sens_section(self, w_sheet, item ):
        #dynamically change initial pens employed
        #eeeeeee
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        if item=='initial_pens_employed':
            self._transfer_cell(w_sheet,'investment_cost','sens','initial_pens_employed',number_format_integer, 'Inputs')
            
        if item=='senior_debt':
            self._transfer_cell(w_sheet,'investment_cost','sens','senior_debt_dynamic_parameter', '0.0%', 'Inputs','senior_debt')
        
        if item=='total_cattle_per_pen_per_cycle':
            self._transfer_cell(w_sheet,'investment_cost','sens','pen_cattle_density',number_format_integer,'Inputs','total_cattle_per_pen_per_cycle')
  
  
    def _add_cal_investment_cost_real_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'INVESTMENT COST (Real)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_investment_cost_real' in self.track_inputs:
            self.track_inputs['calc_investment_cost_real']= {}# insert inner

        if not ('header' in self.track_inputs['calc_investment_cost_real']):
            self.track_inputs['calc_investment_cost_real']['header'] = {}
  
        if 'header' in self.track_inputs['calc_investment_cost_real']:
            self.track_inputs['calc_investment_cost_real']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_investment_cost_real', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        #==========Sub-heading=========================
        self._add_sub_heading(w_sheet,'Feedlots(real)',row_index)
        row_index +=1

        itemlist= ['cost_of_pens_constructed','investment_cost_of_pens',
                   'million_to_thousand', 'investment_cost_of_land','investment_cost_of_land_real',
                   'investment_costs_over_run_factor','exchange_rate',
                   'cif_cost_of_machinery', 'cif_cost_of_machinery_usd','cif_cost_of_machinery_lc', 
                   'import_duty', 'total_cost_machinery', 'investment_cost_of_buildings', 
                   'investment_cost_of_buildings_real', 'equity', 'senior_debt',
                   'total_investment_cost','total_local_investment',
                   'equity_towards_investment','senior_debt_towards_investment']
        
               
        ic_parameters = {
            # Feedlots
            'cost_of_pens_constructed': {'title': 'Cost of feedlots constructed','value': None, 'units': 'T_LC'},
            'investment_cost_of_pens': {'title': 'Investment cost of Feedlots (real)','value': None, 'units': 'T_LC'},# linked cell
            
            # Land
            'million_to_thousand': {'title': 'Million to thousand conversion','value': None, 'units': 'Thousand'}, 
            'investment_cost_of_land': {'title': 'Investment cost of land','value': None, 'units': 'T_LC'},# linked cell
            'investment_cost_of_land_real': {'title': 'Investment cost of land (real)','value': None, 'units': 'T_LC'},
           
            # Machinery
            'investment_costs_over_run_factor': {'title': 'Investment Costs Over-run Factor','value': None, 'units': 'PERCENT'},
            'exchange_rate': {'title': 'Exchange Rate (LC/USD - base year)','value': None, 'units': 'NUMBER'},
            'cif_cost_of_machinery': {'title': 'CIF cost of Machinery (FC$ 000)','value': None, 'units': 'T_USD'},
            'cif_cost_of_machinery_usd': {'title': 'CIF cost of Machinery','value': None, 'units': 'T_USD'},
            'cif_cost_of_machinery_lc': {'title': 'CIF cost of Machinery','value': None, 'units': 'T_LC'},
            'import_duty': {'title': 'Import duty','value': None, 'units': 'PERCENT'},
            'total_cost_machinery': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            
            # Buildings
            'investment_cost_of_buildings': {'title': 'Investment cost of buildings (LC$ 000)','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_real': {'title': 'Investment cost of buildings (real)','value': None, 'units': 'T_LC'},

            # Financing Parameters
            'equity': {'title': 'Equity (% of Investment Costs)','value': None, 'units': 'PERCENT'},
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': None, 'units': 'PERCENT'},

             # Financing Parameters
            'total_investment_cost': {'title': 'Total Investment Cost (real)','value': None, 'units': 'T_LC'},
            'total_local_investment': {'title': 'Total Local Investment Cost (real)','value': None, 'units': 'T_LC'},
            'equity_towards_investment': {'title': 'Equity Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
            'senior_debt_towards_investment': {'title': 'Senior Debt Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_investment_cost_real']):
                    self.track_inputs['calc_investment_cost_real'][item] = {}

                self.track_inputs['calc_investment_cost_real'][item]['row'] = row_index
                self.track_inputs['calc_investment_cost_real'][item]['unit'] = _unit
                self.track_inputs['calc_investment_cost_real'][item]['col'] = 6#redundant
                self.track_inputs['calc_investment_cost_real'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['investment_cost_of_pens','investment_cost_of_land_real',
                            'total_cost_machinery','investment_cost_of_buildings_real','senior_debt'] :
                    row_index +=1 

                    if item =='investment_cost_of_pens':
                        self._add_sub_heading(w_sheet,'Land(real)',row_index)
                        row_index +=1

                    if item =='investment_cost_of_land_real':
                        self._add_sub_heading(w_sheet,'Machinery(real)',row_index)
                        row_index +=1  

                    if item =='total_cost_machinery':
                        self._add_sub_heading(w_sheet,'Buildings(real)',row_index)
                        row_index +=1 

                    if item =='investment_cost_of_buildings_real':
                        self._add_sub_heading(w_sheet,'Financing Parameters',row_index)
                        row_index +=1  

                    if item =='senior_debt':
                        self._add_sub_heading(w_sheet,'Financing',row_index)
                        row_index +=1              

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_investment_cost_real(w_sheet)
       
        return row_index 


    def _populate_cal_investment_cost_real(self, w_sheet):
         
        #*************************************************Pens*****************************************************
        #------Cost of pens constructed
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'cost_of_pens_constructed',number_format, 'Inputs')        
        #------------Real cost of Pens
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','cost_of_pens_constructed')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_pens_cell +'/' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        #**********************************Land Real*************************************
        #---Land 
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_cost_of_land',number_format, 'Inputs')
        #-----Land (real)-------------------
        self._transfer_cell_range( w_sheet=w_sheet, source_header='calc_investment_cost_real', 
                        target_header='calc_investment_cost_real', source_para='investment_cost_of_land',
                        cell_number_format =number_format, source_wksheet=None, 
                        target_para = 'investment_cost_of_land_real',cell_style='Calculation')                  
        
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        source_cell =1000 # self.modelspec_cell_ref["M_CONVERSION"]

        if 'model_specs' in self.track_inputs:
            if 'M_CONVERSION' in self.track_inputs['model_specs']:     
                row= self.track_inputs['model_specs']['M_CONVERSION']['row']
                col= self.track_inputs['model_specs']['M_CONVERSION']['col']
                source_cell = '=Inputs!' + get_column_letter(6) + str(row)
        
        
        self._transfer_value_tocell(w_sheet, source_cell,'calc_investment_cost_real', 'million_to_thousand',number_format)
       
        #***********************************Trancefer Cells from Input *********************************
        #----Machinery
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'cif_cost_of_machinery',number_format, 'Inputs')
       
        #------Cost overun LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_costs_over_run_factor','0%', 'Inputs')             
        #------Exchange LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'macroeconomic_parameters','calc_investment_cost_real',
                            'exchange_rate',number_format, 'Inputs')

        #------Import LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'taxes','calc_investment_cost_real',
                    'import_duty','0%', 'Inputs')                      
        #------Equity LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_real','equity','0%', 'Inputs')        
        #------Senior LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_real','senior_debt','0%', 'Inputs')


        #------CIF Cost of Machinery USD-------------------------------------------
        #cost overun
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_costs_over_run_factor')
        cost_overun_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        cif_r_index, _, found_state_cif= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_usd')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_cell = cell.column + '$' + str(cif_r_index) if found_state_cif else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_cell +'*(1+'+ cost_overun_cell + ')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #------CIF Cost of Machinery-- LC-----------------------------------------
        #cost overun
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','exchange_rate')
        exchange_rate_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        cif_usd_r_index, _, found_state_cif_usd= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_usd')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_lc')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_usd_cell = cell.column + '$' + str(cif_usd_r_index) if found_state_cif_usd else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_usd_cell +'*'+ exchange_rate_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #------To cost of machiner in LC [importy duty added]-----------------------------------------
        #import duty
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','import_duty')
        import_duty_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif LC
        cif_lc_r_index, _, found_state_cif_lc= self._retrieve_cell_row_colm('calc_investment_cost_real','cif_cost_of_machinery_lc')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_cost_machinery')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cif_lc_cell = cell.column + '$' + str(cif_lc_r_index) if found_state_cif_lc else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cif_lc_cell +'*(1+'+ import_duty_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        # Buidingd Real
        self._transfer_cell_range(w_sheet,'investment_cost','calc_investment_cost_real',
                    'investment_cost_of_buildings',number_format, 'Inputs')
        
        #Building deprecaite over years: land doesnt
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
    
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_buildings_cell +'/' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        

       
         #------Total Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_land_real')
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_real','total_cost_machinery')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_machinery_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #------Total Local Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_land_real')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_real','investment_cost_of_buildings_real')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','total_local_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Equity toward Total Investment-------------------------
        #equity
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','equity')
        equity_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','equity_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ equity_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Senoir Debt Contributiom Total Investment-------------------------
        #senior_debt
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','senior_debt')
        senior_debt_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_real','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_real','senior_debt_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ senior_debt_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
    def _add_cal_production_sales_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'PRODUCTION AND SALES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'cal_production_sales_nominal' in self.track_inputs:
            self.track_inputs['cal_production_sales_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_production_sales_nominal']):
            self.track_inputs['cal_production_sales_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_production_sales_nominal']:
            self.track_inputs['cal_production_sales_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_production_sales_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist= ['cum_pens_under_harvesting','production_quantity','closing_inventory',
                   'sales_quantity', 'domestic_price_index','change_in_price', 'base_price',
                   'real_price', 'nominal_price', 'net_sales_revenue', 
                   'sales_tax', 'sales_tax_paid', 'gross_sales_revenue']
                
        ic_parameters = {
            # PI
            'cum_pens_under_harvesting': {'title': 'Cumulative pens under harvesting','value': None, 'units': 'NUMBER'},
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'TONS'},
            'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'TONS'},# linked cell
            'sales_quantity': {'title': 'Sales quantity','value': None, 'units': 'TONS'},# linked cell

            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'change_in_price': {'title': 'Change in real price of Beef','value': None, 'units': 'PERCENT'},
            'base_price': {'title': 'Base price of beef per ton','value': None, 'units': 'LC'},
            'real_price': {'title': 'Real Price of Beef per ton','value': None, 'units': 'LC'},
            'nominal_price': {'title': 'Nominal Price of Beef per ton','value': None, 'units': 'LC'},
            
            'net_sales_revenue': {'title': 'Net sales revenue','value': None, 'units': 'T_LC'},
            'sales_tax': {'title': 'Sales tax','value': None, 'units': 'PERCENT'},
            'sales_tax_paid': {'title': 'Sales tax paid','value': None, 'units': 'T_LC'},
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            
        }
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_production_sales_nominal']):
                    self.track_inputs['cal_production_sales_nominal'][item] = {}

                self.track_inputs['cal_production_sales_nominal'][item]['row'] = row_index
                self.track_inputs['cal_production_sales_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_production_sales_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_production_sales_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['sales_quantity','nominal_price',] :
                    row_index +=1 

        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_production_sales_nominal(w_sheet)
       
        return row_index 
    
    def _populate_cal_production_sales_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #copy pens
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'cum_pens_under_harvesting',number_format, 'Inputs') 
        
        #copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'production_quantity',number_format, 'Inputs') 

        #copy production inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_production_sales_nominal',
                    'closing_inventory',number_format, 'Inputs') 

        # sales qnty      
        prod_qnty_r_index, _, found_state_prod_qnty= self._retrieve_cell_row_colm('cal_production_sales_nominal','production_quantity')
        closing_inv_r_index, _, found_state_closing_inv= self._retrieve_cell_row_colm('cal_production_sales_nominal','closing_inventory')
      
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_quantity')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    new_column_letter = cell.column # J
                    prev_col=column_index_from_string(new_column_letter)-1
                    prev_letter= get_column_letter(prev_col)
                    
                    prod_qnty_cell = cell.column + '$' + str(prod_qnty_r_index) if found_state_prod_qnty else '0'
                    closing_inv_prev_cell = prev_letter + '$' + str(closing_inv_r_index) if found_state_closing_inv else '0'
                    closing_inv_curr_cell = cell.column + '$' + str(closing_inv_r_index) if found_state_closing_inv else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = "=" + prod_qnty_cell + '+' \
                                                            + closing_inv_prev_cell +  '-' \
                                                            + closing_inv_curr_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        


        #---test
        self._transfer_value_tocellrange(w_sheet, 'prices', 
                        'cal_production_sales_nominal', 'change_in_price',  "Inputs", None,
                        '0%')                                
        
        #------Price-------------------------------------------
        
        self._transfer_cell(w_sheet,'prices','cal_production_sales_nominal','base_price',number_format, 'Inputs')        
        #------saless tax-------------------------------------------
        self._transfer_cell(w_sheet,'taxes','cal_production_sales_nominal','sales_tax','0%', 'Inputs')

        
        #copy price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_production_sales_nominal',
                    'domestic_price_index',number_format) 

        
        #------Real Price-------------------------------------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_production_sales_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # base price
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','base_price')
        base_price_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','change_in_price')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','real_price')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_price_cell = prev_letter + str(cell.row) 
                    #=IF(I89=Inputs!$F$32 ;$F$98, H99*(1+I97))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + base_price_cell +',' \
                                                               + prev_real_price_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #------Nominal Price-------------------------------------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_production_sales_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #real price
        real_price_r_index, _, found_state_real_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','real_price')
        
        #domestic
        domestic_price_index_r_index, _, found_state_domestic_price_index= self._retrieve_cell_row_colm('cal_production_sales_nominal','domestic_price_index')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','nominal_price')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    domestic_price_index_cell = cell.column + '$' \
                        + str(domestic_price_index_r_index) if found_state_domestic_price_index else '0'

                   
                    #----
                    real_price_cell = cell.column + '$' \
                        + str(real_price_r_index) if found_state_real_price else '0'

                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_price_cell = prev_letter + str(cell.row) 
                    #=IF(I89=Inputs!$F$32 ;$F$98, H99*(1+I97))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + real_price_cell +',' \
                                                               + real_price_cell + '*'+ domestic_price_index_cell+')'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                                
        #------NET SALES REVENUE-------------------------------------------
        #sales qnty
        sales_qnty_r_index, _, found_state_sales_qnty= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_quantity')
        #nominal price
        nominal_price_r_index, _, found_state_nominal_price= self._retrieve_cell_row_colm('cal_production_sales_nominal','nominal_price')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #----
                    sales_qnty_cell = cell.column + '$' \
                        + str(sales_qnty_r_index) if found_state_sales_qnty else '0'
                    #----
                    nominal_price_cell = cell.column + '$' \
                        + str(nominal_price_r_index) if found_state_nominal_price else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ sales_qnty_cell + '*' \
                                                               + nominal_price_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=J100*J94/1000
        
        #------LESS SALES TAX-------------------------------------------
        #sales qnty
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax')
        #----
        sales_tax_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
                    
        #nominal price
        net_sales_revenue_r_index, _, found_state_net_sales_revenue= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax_paid')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                    net_sales_revenue_cell = cell.column + '$' \
                        + str(net_sales_revenue_r_index) if found_state_net_sales_revenue else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ sales_tax_cell + '*' \
                                                               + net_sales_revenue_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=$F$103*I102
        #------GROSS SALES REVENUE-------------------------------------------
        #net sales reveneu
        net_sales_revenue_r_index, _, found_state_net_sales_revenue= self._retrieve_cell_row_colm('cal_production_sales_nominal','net_sales_revenue')
     
        #sales_tax_paid
        sales_tax_paid_r_index, _, found_state_sales_tax_paid= self._retrieve_cell_row_colm('cal_production_sales_nominal','sales_tax_paid')
     
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_production_sales_nominal','gross_sales_revenue')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                    net_sales_revenue_cell = cell.column + '$' \
                        + str(net_sales_revenue_r_index) if found_state_net_sales_revenue else '0'
                    
                    sales_tax_paid_cell = cell.column + '$' \
                        + str(sales_tax_paid_r_index) if found_state_sales_tax_paid else '0'    
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ net_sales_revenue_cell + '+' \
                                                               + sales_tax_paid_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        #=$F$103*I102
      
    def _add_cal_purchases_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'PURCHASES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_purchases_nominal' in self.track_inputs:
            self.track_inputs['calc_purchases_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['calc_purchases_nominal']):
            self.track_inputs['calc_purchases_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['calc_purchases_nominal']:
            self.track_inputs['calc_purchases_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_purchases_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #==========Sub-headings=========================
        self._add_sub_heading(w_sheet,'Imported Inputs',row_index)

        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['annual_change_in_price_of_imported_inputs','cost_of_imported_inputs_per_ton_of_beef_produced',
                   'real_cif_cost_of_imported_inputs_per_ton',
                   'us_price_index', 'nominal_exchange_rate','import_duty',
                   'nominal_cif_cost_of_imported_inputs_per_ton_usd',
                   'nominal_cif_cost_of_imported_inputs_per_ton_lc', 
                   'cost_of_imported_inputs_including_import_duty_per_ton', 'production_quantity',
                   'total_cost_of_imported_inputs_including_import_duty', 
                   'annual_change_in_price_of_domestic_inputs', 'cost_of_domestic_inputs_per_ton_of_beef_produced', 
                   'real_cost_of_domestic_inputs_per_ton',
                   'domestic_price_index','nominal_cost_of_domestic_inputs_per_ton',
                   'total_cost_of_domestic_inputs','total_cost_of_inputs_per_ton_nominal','total_input_cost_nominal']
        
        ic_parameters = {
            #Imported Inputs	
            'annual_change_in_price_of_imported_inputs': {'title': 'Annual change in price of imported inputs','value': None, 'units': 'PERCENT'},
            'cost_of_imported_inputs_per_ton_of_beef_produced': {'title': 'Cost of imported inputs per ton of Beef produced','value': None, 'units': 'USD'},
            'real_cif_cost_of_imported_inputs_per_ton': {'title': 'Real CIF Cost of imported inputs per ton','value': None, 'units': 'USD'},# linked cell
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},# linked cell
            'nominal_exchange_rate': {'title': 'Nominal Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
            'import_duty': {'title': 'Import duty','value': None, 'units': 'PERCENT'},

            'nominal_cif_cost_of_imported_inputs_per_ton_usd': {'title': 'Nominal CIF Cost of imported inputs per ton','value': None, 'units': 'USD'},
            'nominal_cif_cost_of_imported_inputs_per_ton_lc': {'title': 'Nominal CIF Cost of imported inputs per ton','value': None, 'units': 'LC'},
            'cost_of_imported_inputs_including_import_duty_per_ton': {'title': 'Cost of imported inputs including import duty per ton','value': None, 'units': 'LC'},
            
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'}, 
            'total_cost_of_imported_inputs_including_import_duty': {'title': 'Total cost of imported inputs including import duty','value': None, 'units': 'PERCENT'},
          
            #Domestic Inputs
            'annual_change_in_price_of_domestic_inputs': {'title': 'Annual change in the price of domestic inputs','value': None, 'units': 'PERCENT'},
            'cost_of_domestic_inputs_per_ton_of_beef_produced': {'title': 'Cost of domestic inputs per ton of Beef produced','value': None, 'units': 'LC'},
            'real_cost_of_domestic_inputs_per_ton': {'title': 'Real Cost of domestic inputs per ton','value': None, 'units': 'LC'},
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'nominal_cost_of_domestic_inputs_per_ton': {'title': 'Nominal Cost of domestic inputs per ton','value': None, 'units': 'LC'},
            'total_cost_of_domestic_inputs': {'title': 'Total cost of domestic inputs','value': None, 'units': 'T_LC'},

            'total_cost_of_inputs_per_ton_nominal': {'title': 'Total cost of inputs per ton (nominal)','value': None, 'units': 'LC'},

            'total_input_cost_nominal': {'title': 'Total Input Cost (nominal)','value': None, 'units': 'T_LC'},
            
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_purchases_nominal']):
                    self.track_inputs['calc_purchases_nominal'][item] = {}

                self.track_inputs['calc_purchases_nominal'][item]['row'] = row_index
                self.track_inputs['calc_purchases_nominal'][item]['unit'] = _unit
                self.track_inputs['calc_purchases_nominal'][item]['col'] = 6 #redundant
                self.track_inputs['calc_purchases_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['import_duty','cost_of_imported_inputs_including_import_duty_per_ton',
                            'total_cost_of_imported_inputs_including_import_duty',
                            'total_cost_of_domestic_inputs', 'total_cost_of_inputs_per_ton_nominal',
                            'total_input_cost_nominal'] :
                    row_index +=1 
                    
                    if item =='total_cost_of_imported_inputs_including_import_duty':
                        self._add_sub_heading(w_sheet,'Domestic Inputs',row_index)
                        row_index +=1 
        
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_purchases_nominal(w_sheet)
       
        return row_index 

    
      
    def _populate_cal_purchases_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #0 
        #copy price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','calc_purchases_nominal',
                    'domestic_price_index',number_format) 

        #1. copy import change in prices
        self._transfer_value_tocellrange(w_sheet, 'cost_real','calc_purchases_nominal', 
                 'annual_change_in_price_of_imported_inputs',  "Inputs", None,  '0%')                
        
        #2. copy cost of imported inputs
        self._transfer_cell(w_sheet,'cost_real','calc_purchases_nominal','cost_of_imported_inputs_per_ton_of_beef_produced',number_format, 'Inputs')
        
        
        #3-----REal CIF cost_cost_of_imported_inputs_per_ton--------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_purchases_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # costs of inputs
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_per_ton_of_beef_produced')
        cost_of_imports_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('calc_purchases_nominal','annual_change_in_price_of_imported_inputs')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cif_cost_of_imported_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_cif_cell = prev_letter + str(cell.row) 
                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + cost_of_imports_cell +',' \
                                                               + prev_real_cif_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        
        #4. copy import change in prices
        self._transfer_cell_range(w_sheet, 'calc_inflation_price_index','calc_purchases_nominal', 
                 'us_price_index',   number_format)                
       
        #5. copy norminal
        self._transfer_cell_range(w_sheet, 'calc_inflation_price_index','calc_purchases_nominal', 
                 'nominal_exchange_rate',   number_format)                
       
        #6. copy cost of imported inputs
        self._transfer_cell(w_sheet,'taxes','calc_purchases_nominal','import_duty','0%', 'Inputs')
        
        
        #7-----Nominal CIF cost of imported inputs per ton (usd)--------------
        #real_cif_cost
        real_cif_cost_r_index, c_index, found_state_real_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cif_cost_of_imported_inputs_per_ton')
        
        #us_price_index
        us_price_index_r_index, c_index, found_state_us_price_index= self._retrieve_cell_row_colm('calc_purchases_nominal','us_price_index')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_usd')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    real_cif_cost_cell = cell.column + '$' \
                        + str(real_cif_cost_r_index) if found_state_real_cif_cost else '0'

                    us_price_index_cell = cell.column + '$' \
                        + str(us_price_index_r_index) if found_state_us_price_index else '0'

                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ real_cif_cost_cell + '*' + us_price_index_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #8-----Nominal CIF cost of imported inputs per ton (LC)--------------
        #nominal_cif_cost
        nominal_cif_cost_r_index, c_index, found_state_nominal_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_usd')
        
        #nominal_exchange_rate
        nominal_exchange_rate_r_index, c_index, found_state_nominal_exchange_rate= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_exchange_rate')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_lc')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    nominal_cif_cost_cell = cell.column + '$' \
                        + str(nominal_cif_cost_r_index) if found_state_nominal_cif_cost else '0'

                    nominal_exchange_rate_cell = cell.column + '$' \
                        + str(nominal_exchange_rate_r_index) if found_state_nominal_exchange_rate else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ nominal_cif_cost_cell + '*' + nominal_exchange_rate_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
        #9-----Nominal CIF cost of imported inputs per ton (LC)--------------
        #nominal_cif_cost
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','import_duty')
        import_duty_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        #nominal_cif_cost
        nominal_cif_cost_r_index, _, found_state_nominal_cif_cost= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cif_cost_of_imported_inputs_per_ton_lc')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    nominal_cif_cost_cell = cell.column + '$' \
                        + str(nominal_cif_cost_r_index) if found_state_nominal_cif_cost else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + nominal_cif_cost_cell + '*(1+'+ import_duty_cell + ')'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
        #10. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','calc_purchases_nominal',
                    'production_quantity',number_format, 'Inputs') 

        
        
        #11 -----Nominal CIF cost of imported inputs per ton (LC)--------------
        #cost_of_imported_inputs
        cost_of_imported_inputs_r_index, _, found_state_cost_of_imported_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
       
        #prod_qnty
        prod_qnty_r_index, _, found_state_prod_qnty= self._retrieve_cell_row_colm('calc_purchases_nominal','production_quantity')
        
        #total_cost_of_imported--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_imported_inputs_including_import_duty')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    cost_of_imported_inputs_cell = cell.column + '$' \
                                          + str(cost_of_imported_inputs_r_index) if found_state_cost_of_imported_inputs else '0'
                   
                    prod_qnty_cell = cell.column + '$' \
                                          + str(prod_qnty_r_index) if found_state_prod_qnty else '0'                      
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + cost_of_imported_inputs_cell + '*'+ prod_qnty_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        
       
        #12. copy domestic change in prices
        self._transfer_value_tocellrange(w_sheet, 'cost_real','calc_purchases_nominal', 
                 'annual_change_in_price_of_domestic_inputs',  "Inputs", None,  '0%')                
        

        #13. copy cost of domestic inputs
        self._transfer_cell(w_sheet,'cost_real','calc_purchases_nominal','cost_of_domestic_inputs_per_ton_of_beef_produced',number_format, 'Inputs')
        
        #14.Real cost of Domestic inputs_per_ton--------------
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_purchases_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # costs of inputs
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_domestic_inputs_per_ton_of_beef_produced')
        cost_of_domestic_input_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #change in price
        change_in_price_r_index, c_index, found_state_change_in_price= self._retrieve_cell_row_colm('calc_purchases_nominal','annual_change_in_price_of_domestic_inputs')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cost_of_domestic_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    change_in_price_cell = cell.column + '$' \
                        + str(change_in_price_r_index) if found_state_change_in_price else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_real_cif_cell = prev_letter + str(cell.row) 
                    #=IF(I107=Inputs!$F$32; $F$111;H112*(1+I110)
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + cost_of_domestic_input_cell +',' \
                                                               + prev_real_cif_cell + '*(1+'+ change_in_price_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        
        #15-----Nominal cost of domestic inputs per ton (usd)--------------
        #real_cost_domestic
        real_cost_domestic_r_index, c_index, found_state_real_cost_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','real_cost_of_domestic_inputs_per_ton')
        
        #domestic_price_index
        domestic_price_index_r_index, c_index, found_state_domestic_price_index= self._retrieve_cell_row_colm('calc_purchases_nominal','domestic_price_index')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    real_cost_domestic_cell = cell.column + '$' \
                        + str(real_cost_domestic_r_index) if found_state_real_cost_domestic else '0'

                    domestic_price_index_cell = cell.column + '$' \
                        + str(domestic_price_index_r_index) if found_state_domestic_price_index else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ real_cost_domestic_cell + '*' + domestic_price_index_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'

        #16-----total_cost_of_domestic_inputs--------------
        #real_cost_domestic
        prod_qnty_r_index, c_index, found_state_prod_qnty= self._retrieve_cell_row_colm('calc_purchases_nominal','production_quantity')
        
        #nominal_cost_of_domestic
        nominal_cost_of_domestic_r_index, c_index, found_state_nominal_cost_of_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_domestic_inputs')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    nominal_cost_of_domestic_cell = cell.column + '$' \
                        + str(nominal_cost_of_domestic_r_index) if found_state_nominal_cost_of_domestic else '0'

                    prod_qnty_cell = cell.column + '$' \
                        + str(prod_qnty_r_index) if found_state_prod_qnty else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ prod_qnty_cell + '*' + nominal_cost_of_domestic_cell + '/1000'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'

        #17.   total_cost_of_inputs_per_ton_nominal--------------
        #cost_of_imported_inputs
        cost_of_imported_inputs_r_index, c_index, found_state_cost_of_imported_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','cost_of_imported_inputs_including_import_duty_per_ton')
        
        #nominal_cost_of_domestic
        nominal_cost_of_domestic_r_index, c_index, found_state_nominal_cost_of_domestic= self._retrieve_cell_row_colm('calc_purchases_nominal','nominal_cost_of_domestic_inputs_per_ton')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_inputs_per_ton_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    cost_of_imported_inputs_cell = cell.column + '$' \
                        + str(cost_of_imported_inputs_r_index) if found_state_cost_of_imported_inputs else '0'

                    nominal_cost_of_domestic_cell = cell.column + '$' \
                        + str(nominal_cost_of_domestic_r_index) if found_state_nominal_cost_of_domestic else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ cost_of_imported_inputs_cell + '+' + nominal_cost_of_domestic_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #18.   total_cost_of_inputs_per_ton_nominal--------------
        #total_cost_of_imported
        total_cost_of_imported_r_index, c_index, found_state_total_cost_of_imported= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_imported_inputs_including_import_duty')
        
        #nominal_cost_of_domestic
        total_cost_of_domestic_inputs_r_index, c_index, found_state_total_cost_of_domestic_inputs= self._retrieve_cell_row_colm('calc_purchases_nominal','total_cost_of_domestic_inputs')
        
        #nominal_cif_cost--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_purchases_nominal','total_input_cost_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    total_cost_of_imported_cell = cell.column + '$' \
                        + str(total_cost_of_imported_r_index) if found_state_total_cost_of_imported else '0'

                    total_cost_of_domestic_inputs_cell = cell.column + '$' \
                        + str(total_cost_of_domestic_inputs_r_index) if found_state_total_cost_of_domestic_inputs else '0'

                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_cost_of_imported_cell + '+' + total_cost_of_domestic_inputs_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        list_= [
            {
             'header': 'calc_purchases_nominal', 
             'para': 'total_cost_of_imported_inputs_including_import_duty', 
             'action': '+'
            }, 
            {
             'header': 'calc_purchases_nominal', 
             'para': 'total_cost_of_domestic_inputs', 
             'action': '+'
            }
        ]
        self._sum_oflist_cell_range(w_sheet,list_,'calc_purchases_nominal','total_input_cost_nominal')
        
 

    def _populate_cal_working_capital_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #1. copy accounts_receivable
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','accounts_receivable','0%', 'Inputs')
         
        #2. gross_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_working_capital_nominal',
                    'gross_sales_revenue',number_format) 
        
        #3. Cal accounts_receivable
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_receivable',
                                                   'gross_sales_revenue','accounts_receivable_val')

        #4. copy accounts_payable
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','accounts_payable','0%', 'Inputs')
        
        
        #5. copy total_input_cost_nominal
        self._transfer_cell_range(w_sheet,'calc_purchases_nominal','cal_working_capital_nominal',
                    'total_input_cost_nominal',number_format) 
 
        
        #6. Cal accounts_payable
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_payable',
                                                   'total_input_cost_nominal','accounts_payable_val')


        #7. copy cash_balance
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','cash_balance','0%', 'Inputs')
        
        #8. Cal cash_balance
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','cash_balance',
                                                   'gross_sales_revenue','cash_balance_val')

        #9. Change in Accounts Receivable
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_receivable_val','change_in_AR',True)
        
        #10. Change in accounts_payable
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_payable_val','change_in_AP', True)
        
        #11 Change in cash_balance
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','cash_balance_val','change_in_CB')
    
    def _populate_cal_labour_cost_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        #o.copy domestic price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_labour_cost_nominal',
                    'domestic_price_index',number_format) 

        #1. num_of_workers
        self._transfer_cell_range(w_sheet,'cost_real','cal_labour_cost_nominal',
                    'num_of_workers',number_format, "Inputs")
        #2. copy annual_increase_salaries_workers
        self._transfer_value_tocellrange(w_sheet, 'cost_real', 'cal_labour_cost_nominal', 
                                  'annual_increase_salaries_workers',  "Inputs", None,    '0%')   
        
        #3. monthly_wage_per_worker
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','monthly_wage_per_worker',number_format, 'Inputs')
        

        #4. number_of_months_in_a_year
        self._transfer_cell(w_sheet,'timing','cal_labour_cost_nominal','number_of_months_in_a_year',number_format_integer, 'Inputs')
       
        
        #5... Real labour cost
        # year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('cal_labour_cost_nominal','header')        
        
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','monthly_wage_per_worker')
        monthly_wage_per_worker_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','number_of_months_in_a_year')
        number_of_months_in_a_year_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #annual_increase_salaries_workers
        annual_increase_salaries_workers_r_index, c_index, found_state_annual_increase_salaries_workers= self._retrieve_cell_row_colm('cal_labour_cost_nominal','annual_increase_salaries_workers')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','real_yearly_wage_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    annual_increase_salaries_workers_cell = cell.column + '$' \
                        + str(annual_increase_salaries_workers_r_index) if found_state_annual_increase_salaries_workers else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_yearly_rate_cell = prev_letter + str(cell.row) 
                    #=IF(K$153=Inputs!$F$32;$F$158*$F$159;J160*(1+K157))
                    year_salary =f'{monthly_wage_per_worker_cell} * {number_of_months_in_a_year_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + year_salary +',' \
                                                               + prev_yearly_rate_cell + '*(1+'+ annual_increase_salaries_workers_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                 
        #6. nominal_yearly_wage_rate

        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'real_yearly_wage_rate',  'action': '+'  }, 
                { 'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','nominal_yearly_wage_rate')

        
        #7. total_direct_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'num_of_workers', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'nominal_yearly_wage_rate', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_direct_labour_cost')
        
        #8. num_of_supervisors_technicians
        self._transfer_cell_range(w_sheet,'cost_real','cal_labour_cost_nominal',
                    'num_of_supervisors_technicians',number_format, "Inputs")

        #9. copy annual_increase_salaries_supervisors_technicians
        self._transfer_value_tocellrange(w_sheet, 'cost_real', 'cal_labour_cost_nominal', 
                                  'annual_increase_salaries_supervisors_technicians',  "Inputs", None,    '0%')   
        
        #10. monthly_wage_per_supervisor
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal',
                                 'monthly_wage_per_supervisor',number_format, 'Inputs')
        

        #11... Real labour cost
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','monthly_wage_per_supervisor')
        monthly_wage_per_supervisor_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # monthly_wage_per_worker
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','number_of_months_in_a_year')
        number_of_months_in_a_year_cell =  get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        #annual_increase_salaries_workers
        annual_increase_salaries_supervisors_technicians_r_index, c_index, found_state_annual_increase_salaries_supervisors_technicians= self._retrieve_cell_row_colm('cal_labour_cost_nominal','annual_increase_salaries_supervisors_technicians')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_labour_cost_nominal','real_yearly_salary_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    annual_increase_salaries_supervisors_technicians_cell = cell.column + '$' \
                        + str(annual_increase_salaries_supervisors_technicians_r_index) if found_state_annual_increase_salaries_supervisors_technicians else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_yearly_rate_cell = prev_letter + str(cell.row) 
                    #=IF(K$153=Inputs!$F$32;$F$158*$F$159;J160*(1+K157))
                    year_salary =f'{monthly_wage_per_supervisor_cell} * {number_of_months_in_a_year_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',' \
                                                               + year_salary +',' \
                                                               + prev_yearly_rate_cell + '*(1+'+ annual_increase_salaries_supervisors_technicians_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #12. nominal_yearly_salary_rate

        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'real_yearly_salary_rate',  'action': '+'  }, 
                { 'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','nominal_yearly_salary_rate')
        
        #13. total_indirect_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'num_of_supervisors_technicians', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'nominal_yearly_salary_rate', 'action': '*' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_indirect_labour_cost')
        
        
        #14. total_labour_cost
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'total_direct_labour_cost', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'total_indirect_labour_cost', 'action': '+' }]
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','total_labour_cost')
        
        #15. OP
        self._transfer_cell_range(w_sheet,'flags','cal_labour_cost_nominal',
                    'OP',number_format_integer, "Inputs")

        #16. cost_of_electricity per pen
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','cost_of_electricity_per_pen_per_annum',number_format, 'Inputs')
        
        #17. cost_of_electricity nominal
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'OP', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' },
                {'header': 'cal_production_sales_nominal', 'para': 'cum_pens_under_harvesting', 'action': '*' }]
        
        const_list_= [{'header': 'cal_labour_cost_nominal', 'para': 'cost_of_electricity_per_pen_per_annum','action': '*' }, 
                {'header': 'calc_investment_cost_real', 'para': 'million_to_thousand' ,'action': '/'}]   

        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','cost_of_electricity_per_pen_per_annum_nominal',const_list_)
        
        #18. other_indirect_costs
        self._transfer_cell(w_sheet,'cost_real','cal_labour_cost_nominal','other_indirect_costs',number_format, 'Inputs')
      
        
        #19. other_indirect_costs nominal
        list_= [{'header': 'cal_labour_cost_nominal', 'para': 'OP', 'action': '+' }, 
                {'header': 'cal_labour_cost_nominal', 'para': 'domestic_price_index', 'action': '*' }]
        const_list_= [{'header': 'cal_labour_cost_nominal', 'para': 'other_indirect_costs','action': '*' },
                      {'header': 'calc_investment_cost_real', 'para': 'million_to_thousand' ,'action': '/'}] 
        self._sum_oflist_cell_range(w_sheet,list_,'cal_labour_cost_nominal','other_indirect_costs_nominal',const_list_)
    
    
    def _populate_cal_unit_of_production(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        
        
        #1. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_unit_of_production',
                    'production_quantity',number_format, 'Inputs') 

        
        #2. total_direct_labour_cost
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_unit_of_production',
                    'total_direct_labour_cost',number_format,)
        #
        
        #3. Direct Labor cost per ton of beef produced
        list_= [{'header': 'cal_unit_of_production', 'para': 'total_direct_labour_cost', 'action': '+' },
                {'header': 'cal_unit_of_production', 'para': 'production_quantity', 'action': '/' },]
        const_list_= [{'header': 'calc_investment_cost_real', 'para': 'million_to_thousand','action': '*' }] 
        self._sum_oflist_cell_range(w_sheet,list_,'cal_unit_of_production','direct_labour_cost_per_ton_of_beef_produced',const_list_)
  
        
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_unit_of_production', 'para': 'production_quantity','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_unit_of_production', 'para': 'total_direct_labour_cost','cell_type': 'cell_range'},
                {'value':'*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'calc_investment_cost_real', 'para': 'million_to_thousand','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_unit_of_production',
                                            'direct_labour_cost_per_ton_of_beef_produced')
  
    

        #4. copy . OP
        self._transfer_cell_range(w_sheet,'flags','cal_unit_of_production',
                    'OP',number_format_integer, "Inputs")
   

        #5. copy . OP
        self._transfer_cell_range(w_sheet,'calc_purchases_nominal','cal_unit_of_production',
                    'total_cost_of_inputs_per_ton_nominal',number_format_integer)

 
        #6. cost_of_electricity nominal
        listA= [{'header': 'cal_unit_of_production', 'para': 'direct_labour_cost_per_ton_of_beef_produced', 'action': '+' }, 
                {'header': 'cal_unit_of_production', 'para': 'total_cost_of_inputs_per_ton_nominal', 'action': '+' },]
        
        listB= [{'header': 'cal_unit_of_production', 'para': 'OP', 'action': '+' }, ]
        
        self._productof_sumof_list_cell_range(w_sheet,listA, listB,'cal_unit_of_production','total_unit_cost_of_production_per_ton_nominal')
        
    
    def _add_cal_finished_product_inventory_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'FINISHED PRODUCT INVENTORY VALUATION USING FIFO METHOD (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_finished_product_inventory' in self.track_inputs:
            self.track_inputs['cal_finished_product_inventory']= {}# insert inner

        if not ('header' in self.track_inputs['cal_finished_product_inventory']):
            self.track_inputs['cal_finished_product_inventory']['header'] = {}
  
        if 'header' in self.track_inputs['cal_finished_product_inventory']:
            self.track_inputs['cal_finished_product_inventory']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_finished_product_inventory', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'OPENING INVENTORY',row_index)
        row_index +=1

        itemlist= [
            'closing_inventory', 'prev_closing_inventory',
            'prev_total_unit_cost_of_production_per_ton_nominal',
            'cost_of_opening_inventory_fifo',

            'production_quantity',
            'total_unit_cost_of_production_per_ton_nominal',
            
            'sales_quantity', 'quantity_sold_from_this_yr_production',
            'cost_of_proportion_sales_produced','cost_of_goods_sold',

            'quantity_remained_unsold','cost_of_closing_inventory'
            ]

        ic_parameters = {
            #OPENING INVENTORY
            'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'T_TONS'},
            'prev_closing_inventory': {'title': 'Quantity from previous year (closing inventory)','value': None, 'units': 'T_TONS'},
            'prev_total_unit_cost_of_production_per_ton_nominal': {'title': 'Unit cost from previous year per ton (nominal)','value': None, 'units': 'LC'},
            'cost_of_opening_inventory_fifo': {'title': "Cost of opening inventory (using previous year's unit cost FIFO method)",'value': None, 'units': 'T_LC'},# linked cell
             #ADDITIONS
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'},
            'total_unit_cost_of_production_per_ton_nominal': {'title': 'Total unit cost of production per ton (nominal)','value': None, 'units': 'LC'},
             #WITHDRAWALS
            'sales_quantity': {'title': 'Sales quantity','value': None, 'units': 'TONS'},
            'quantity_sold_from_this_yr_production': {'title': "Quantity sold from this yr's production",'value': None, 'units': 'TONS'},
            'cost_of_proportion_sales_produced': {'title': 'Cost of the proportion of sales produced in current year','value': None, 'units': 'T_LC'},
            'cost_of_goods_sold': {'title': "COST OF GOODS SOLD (FOR INCOME STATEMENT)",'value': None, 'units': 'T_LC'},
             #CLOSING INVENTORY
            'quantity_remained_unsold': {'title': 'Quantity remained (unsold) in year','value': None, 'units': 'T_TONS'},
            'cost_of_closing_inventory': {'title': "Cost of closing inventory (using current year's unit cost)",'value': None, 'units': 'T_LC'},
           
        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_finished_product_inventory']):
                    self.track_inputs['cal_finished_product_inventory'][item] = {}

                self.track_inputs['cal_finished_product_inventory'][item]['row'] = row_index
                self.track_inputs['cal_finished_product_inventory'][item]['unit'] = _unit
                self.track_inputs['cal_finished_product_inventory'][item]['col'] = 6#redundant
                self.track_inputs['cal_finished_product_inventory'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['cost_of_opening_inventory_fifo','total_unit_cost_of_production_per_ton_nominal','cost_of_goods_sold'] :
                    row_index +=1 

                    if item=='cost_of_opening_inventory_fifo':
                        self._add_sub_heading(w_sheet,'ADDITIONS',row_index)
                        row_index +=1

                    if item=='total_unit_cost_of_production_per_ton_nominal':
                        self._add_sub_heading(w_sheet,'WITHDRAWALS',row_index)
                        row_index +=1    

                    if item=='cost_of_goods_sold':
                        self._add_sub_heading(w_sheet,'CLOSING INVENTORY',row_index)
                        row_index +=1 
                    
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_finished_product_inventory(w_sheet)
       
        return row_index
    
    
     
    def _add_cal_loan_schedule_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'LOAN SCHEDULE (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_loan_schedule' in self.track_inputs:
            self.track_inputs['cal_loan_schedule']= {}# insert inner

        if not ('header' in self.track_inputs['cal_loan_schedule']):
            self.track_inputs['cal_loan_schedule']['header'] = {}
  
        if 'header' in self.track_inputs['cal_loan_schedule']:
            self.track_inputs['cal_loan_schedule']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_loan_schedule', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'OPENING INVENTORY',row_index)
        row_index +=1

        itemlist= [
            'real_interest_rate','risk_premium','us_inflation_rate','nominal_interest_rate',
            'domestic_inflation_rate','nominal_interest_rate_local',
            
            'us_price_index','num_of_installments','LPP',

            'beginning_debt','loan_disbursement','interest_accrued_in_year',
            'principal_paid','interest_paid','total_loan_repayment','outstanding_debt_at_end_of_year',
            
            'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
             
             'CP', 'OP','interest_during_construction','interest_expense'
            ]

        ic_parameters = {
            'real_interest_rate': {'title': 'Real interest rate','value': None, 'units': 'PERCENT'},
            'risk_premium': {'title': 'Risk premium','value': None, 'units': 'PERCENT'},
            'us_inflation_rate': {'title': 'US Inflation rate','value': None, 'units': 'PERCENT'},
            'nominal_interest_rate': {'title': "Nominal Interest Rate",'value': None, 'units': 'PERCENT'},# linked cell
            'domestic_inflation_rate': {'title': 'Local Inflation rate','value': None, 'units': 'PERCENT'},
            #updated in CF
            'nominal_interest_rate_local': {'title': 'Nominal Interest Rate (Local)','value': None, 'units': 'PERCENT'},# CashFlow derived
             
            
            #SUPPLIERS CREDIT
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},
            'num_of_installments': {'title': "No. of installments",'value': None, 'units': 'NUMBER'},
            'LPP': {'title': 'Loan principle repayment','value': None, 'units': 'FLAG'},
            #Loan Repayment Schedule
            'beginning_debt': {'title': "Beginning Debt",'value': None, 'units': 'T_LC'},
            'loan_disbursement': {'title': 'Loan disbursement','value': None, 'units': 'T_LC'},
            'interest_accrued_in_year': {'title': "Interest accrued in year",'value': None, 'units': 'T_LC'},
            'principal_paid': {'title': "Principal paid",'value': None, 'units': 'T_LC'},
            'interest_paid': {'title': 'Interest paid','value': None, 'units': 'T_LC'},
            'total_loan_repayment': {'title': "Total Loan Repayment",'value': None, 'units': 'T_LC'},
            'outstanding_debt_at_end_of_year': {'title': "Outstanding debt at end of year",'value': None, 'units': 'T_LC'},

          
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},
            'total_loan_repayment_outflow_local': {'title': 'Total Loan Repayment as an outflow in Local Currency','value': None, 'units': 'T_LC'},
            
            'CP': {'title': "Construction period flag",'value': None, 'units': 'FLAG'},
            'OP': {'title': "Operating period",'value': None, 'units': 'FLAG'},
            'interest_during_construction': {'title': 'Interest during Construction, Capitalized for Tax Purposes','value': None, 'units': 'T_LC'},
            'interest_expense': {'title': "Interest expense (for income statement)",'value': None, 'units': 'T_LC'},
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_loan_schedule']):
                    self.track_inputs['cal_loan_schedule'][item] = {}

                self.track_inputs['cal_loan_schedule'][item]['row'] = row_index
                self.track_inputs['cal_loan_schedule'][item]['unit'] = _unit
                self.track_inputs['cal_loan_schedule'][item]['col'] = 6#redundant
                self.track_inputs['cal_loan_schedule'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['nominal_interest_rate_local','LPP','outstanding_debt_at_end_of_year',
                             'total_loan_repayment_outflow_local','outstanding_debt_at_end_of_year'] :
                    row_index +=1 

                    if item=='nominal_interest_rate_local':
                        self._add_sub_heading(w_sheet,'SUPPLIERS CREDIT',row_index)
                        row_index +=1

                    if item=='LPP':
                        self._add_sub_heading(w_sheet,'Loan Repayment Schedule',row_index)
                        row_index +=1    

                  
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_loan_schedule(w_sheet)
       
        return row_index

       
    def _add_cal_residual_values_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'RESIDUAL VALUES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_residual_values' in self.track_inputs:
            self.track_inputs['cal_residual_values']= {}# insert inner

        if not ('header' in self.track_inputs['cal_residual_values']):
            self.track_inputs['cal_residual_values']['header'] = {}
  
        if 'header' in self.track_inputs['cal_residual_values']:
            self.track_inputs['cal_residual_values']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_residual_values', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Initial total values',row_index)
        row_index +=1

        itemlist= [
            'investment_cost_of_land_real','total_cost_of_land_real','total_cost_machinery_real',
            'total_cost_machinery','investment_cost_of_buildings_real','total_cost_buildings',

            'economic_life_of_machinery','economic_life_of_buildings',

            'OP','depreciation_machinery','depreciation_buildings',

            'residual_value_of_land_real','residual_value_of_machinery_real', 'residual_value_of_buildings_real',

            'domestic_price_index','residual_value_of_land_nominal','residual_value_of_machinery_nominal','residual_value_of_buildings_nominal'
           
            ]
        
        ic_parameters = {	
            'investment_cost_of_land_real': {'title': 'Investment cost of land (real)','value': None, 'units': 'T_LC'},
            'total_cost_of_land_real': {'title': 'Total cost of land (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery_real': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_real': {'title': "Investment cost of buildings (real)",'value': None, 'units': 'T_LC'},
            'total_cost_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},

            'economic_life_of_machinery': {'title': 'Economic life of machinery','value': None, 'units': 'YEARS'},
            'economic_life_of_buildings': {'title': 'Economic life of buildings','value': None, 'units': 'YEARS'},

            'OP': {'title': "Operation period",'value': None, 'units': 'FLAG'},
            'depreciation_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'depreciation_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},
            
            'residual_value_of_land_real': {'title': 'Residual value of land (real)','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_real': {'title': "Residual value of machinery (real)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_real': {'title': "Residual value of buildings (real)",'value': None, 'units': 'T_LC'},
            
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'residual_value_of_land_nominal': {'title': "Residual value of land (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_nominal': {'title': "Residual value of machinery (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_nominal': {'title': "Residual value of buildings (nominal)",'value': None, 'units': 'T_LC'},
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_residual_values']):
                    self.track_inputs['cal_residual_values'][item] = {}

                self.track_inputs['cal_residual_values'][item]['row'] = row_index
                self.track_inputs['cal_residual_values'][item]['unit'] = _unit
                self.track_inputs['cal_residual_values'][item]['col'] = 6#redundant
                self.track_inputs['cal_residual_values'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['total_cost_buildings','economic_life_of_buildings','depreciation_buildings',
                            'residual_value_of_buildings_real'] :
                    row_index +=1 

                    if item=='total_cost_buildings':
                        self._add_sub_heading(w_sheet,'Economic service life',row_index)
                        row_index +=1

                    if item=='economic_life_of_buildings':
                        self._add_sub_heading(w_sheet,'Depreciation (Real)',row_index)
                        row_index +=1  
                    if item=='depreciation_buildings':
                        self._add_sub_heading(w_sheet,'Residual value (Real)',row_index)
                        row_index +=1

                    if item=='residual_value_of_buildings_real':
                        self._add_sub_heading(w_sheet,'Residual value (Nominal)',row_index)
                        row_index +=1         
              
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_residual_values(w_sheet)
       
        return row_index
    
    def _add_cal_depreciation_for_tax_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'DEPRECIATION FOR TAX PURPOSES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_depreciation_for_tax' in self.track_inputs:
            self.track_inputs['cal_depreciation_for_tax']= {}# insert inner

        if not ('header' in self.track_inputs['cal_depreciation_for_tax']):
            self.track_inputs['cal_depreciation_for_tax']['header'] = {}
  
        if 'header' in self.track_inputs['cal_depreciation_for_tax']:
            self.track_inputs['cal_depreciation_for_tax']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_depreciation_for_tax', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Initial total values',row_index)
        row_index +=1

        itemlist= [
            'total_cost_machinery_nominal','total_machinery','buildings_nominal',
            'total_buildings','interest_during_construction','soft_capital_costs_interest_during_construction',

            'operation_start_year','tax_life_of_machinery','tax_life_of_buildings','tax_life_of_soft_capital_costs',

            'OP','machinery_depreciation_expense', 'buildings_depreciation_expense', 'soft_capital_costs_interest_during_construction_expense','annual_depreciation_expense'
            ]
        ic_parameters = {
            #Initial total values           
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
            'total_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'buildings_nominal': {'title': 'Buildings (nominal)','value': None, 'units': 'T_LC'},
            'total_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},# linked cell
            'interest_during_construction': {'title': 'Interest during Construction, Capitalized for Tax Purposes','value': None, 'units': 'T_LC'},
            'soft_capital_costs_interest_during_construction': {'title': 'Soft Capital Costs (Interest during construction)','value': None, 'units': 'T_LC'},
            #Recovery Period for income tax 
            'operation_start_year': {'title': 'Operation Start Year','value': None, 'units': 'YEAR'},
            'tax_life_of_machinery': {'title': "Tax life of machinery",'value': None, 'units': 'YEARS'},
            'tax_life_of_buildings': {'title': 'Tax life of buildings','value': None, 'units': 'YEARS'},
            'tax_life_of_soft_capital_costs': {'title': "Tax life of soft capital costs",'value': None, 'units': 'YEARS'},
            #Depreciation
            'OP': {'title': "Operation period",'value': None, 'units': 'FLAG'},
            'machinery_depreciation_expense': {'title': 'Machinery depreciation expense','value': None, 'units': 'T_LC'},
            'buildings_depreciation_expense': {'title': "Buildings depreciation expense",'value': None, 'units': 'T_LC'},
            'soft_capital_costs_interest_during_construction_expense': {'title': "Soft Capital Costs (Interest during construction)",'value': None, 'units': 'T_LC'},
            'annual_depreciation_expense': {'title': 'Annual Depreciation Expense','value': None, 'units': 'T_LC'},
           
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_depreciation_for_tax']):
                    self.track_inputs['cal_depreciation_for_tax'][item] = {}

                self.track_inputs['cal_depreciation_for_tax'][item]['row'] = row_index
                self.track_inputs['cal_depreciation_for_tax'][item]['unit'] = _unit
                self.track_inputs['cal_depreciation_for_tax'][item]['col'] = 6#redundant
                self.track_inputs['cal_depreciation_for_tax'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                #go to newline
                row_index +=1 
                if item in ['soft_capital_costs_interest_during_construction','tax_life_of_soft_capital_costs',] :
                    row_index +=1 

                    if item=='soft_capital_costs_interest_during_construction':
                        self._add_sub_heading(w_sheet,'Recovery period for income tax',row_index)
                        row_index +=1

                    if item=='tax_life_of_soft_capital_costs':
                        self._add_sub_heading(w_sheet,'Depreciation expense',row_index)
                        row_index +=1    

                      
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_depreciation_for_tax(w_sheet)
       
        return row_index
     
    
    def _add_cal_income_tax_statement_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'INCOME TAX STATEMENT (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_income_tax_statement' in self.track_inputs:
            self.track_inputs['cal_income_tax_statement']= {}# insert inner

        if not ('header' in self.track_inputs['cal_income_tax_statement']):
            self.track_inputs['cal_income_tax_statement']['header'] = {}
  
        if 'header' in self.track_inputs['cal_income_tax_statement']:
            self.track_inputs['cal_income_tax_statement']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_income_tax_statement', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Revenues',row_index)
        row_index +=1

        
      				
       
				

        itemlist= [
            'net_sales_revenue','cost_of_goods_sold','annual_depreciation_expense', 'total_indirect_labour_cost',
            'cost_of_electricity_per_pen_per_annum_nominal','other_indirect_costs_nominal','total_operating_expenses',
            
            'income_from_operations','interest_paid','pre_tax_income','cumulative_losses',
            'taxable_income_losses_carried_forward','corporate_income_tax','income_tax_payment','net_after_tax_income',

            'share_capital','dividend_payout_ratio','dividend_payouts','num_of_shares','earning_per_share',
            'retained_earnings', 'shareholder_equity' ,'gross_profit',  'gross_profit_margin',
            'net_profit_margin',

            'debt_to_equity_ratio',
            ]

        ic_parameters = {
            # Revenues
            'net_sales_revenue': {'title': 'Net sales revenue','value': None, 'units': 'T_LC'},
            
             #Operating expenses
            'cost_of_goods_sold': {'title': 'Cost of goods sold','value': None, 'units': 'T_LC'},
            'annual_depreciation_expense': {'title': 'Annual Depreciation Expense','value': None, 'units': 'T_LC'},
            'total_indirect_labour_cost': {'title': "Total Indirect Labour Cost",'value': None, 'units': 'T_LC'},# linked cell
            'cost_of_electricity_per_pen_per_annum_nominal': {'title': 'Cost of electricity (nominal)','value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': 'Other Indirect Costs (nominal)','value': None, 'units': 'T_LC'},
            'total_operating_expenses': {'title': "Total Operating Expenses",'value': None, 'units': 'T_LC'},
            
            'income_from_operations': {'title': 'INCOME FROM OPERATIONS','value': None, 'units': 'T_LC'},
            
            'interest_paid': {'title': "Interest paid",'value': None, 'units': 'T_LC'},
            
            'pre_tax_income': {'title': 'PRE-TAX INCOME','value': None, 'units': 'T_LC'},
            
            'cumulative_losses': {'title': "Cumulative losses",'value': None, 'units': 'T_LC'},
            
            'taxable_income_losses_carried_forward': {'title': "Taxable income (losses carried forward)",'value': None, 'units': 'T_LC'},
            
            'corporate_income_tax': {'title': 'Corporate income tax','value': None, 'units': 'PERCENT'},
            'income_tax_payment': {'title': "Income tax payment",'value': None, 'units': 'T_LC'},
            'net_after_tax_income': {'title': "NET AFTER-TAX INCOME",'value': None, 'units': 'T_LC'},

          	
            # EQUITY
            'share_capital': {'title': "Share Capital",'value': None, 'units': 'T_LC'},
            'dividend_payout_ratio': {'title': 'Dividend payout ratio','value': None, 'units': 'PERCENT'},
            'dividend_payouts': {'title': 'Dividend payouts','value': None, 'units': 'PERCENT'},
            'num_of_shares': {'title': "N# of Shares",'value': None, 'units': 'NUMBER'},
            'earning_per_share': {'title': "EPS (Earning Per share)",'value': None, 'units': 'T_LC'},
            
            'retained_earnings': {'title': 'Retained Earnings','value': None, 'units': 'T_LC'},
            'shareholder_equity': {'title': "Shareholder Equity",'value': None, 'units': 'T_LC'},
            'gross_profit': {'title': 'GROSS PROFIT   (Sales-COGS)','value': None, 'units': 'T_LC'},
            'gross_profit_margin': {'title': "GROSS PROFIT MARGIN   [(Sales-COGS)/Sales]",'value': None, 'units': 'PERCENT'},
            'net_profit_margin': {'title': 'NET PROFIT MARGIN','value': None, 'units': 'PERCENT'},
            
            'debt_to_equity_ratio': {'title': "Debt-to-Equity Ratio",'value': None, 'units': 'PERCENT'},
          
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_income_tax_statement']):
                    self.track_inputs['cal_income_tax_statement'][item] = {}

                self.track_inputs['cal_income_tax_statement'][item]['row'] = row_index
                self.track_inputs['cal_income_tax_statement'][item]['unit'] = _unit
                self.track_inputs['cal_income_tax_statement'][item]['col'] = 6#redundant
                self.track_inputs['cal_income_tax_statement'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['net_sales_revenue','LPP','total_operating_expenses','income_from_operations',
                            'net_after_tax_income','interest_paid','pre_tax_income',
                            'taxable_income_losses_carried_forward','income_tax_payment',
                             'earning_per_share','shareholder_equity',
                            'net_profit_margin'] :
                    row_index +=1 

                    if item=='net_sales_revenue':
                        self._add_sub_heading(w_sheet,'Operating expenses',row_index)
                        row_index +=1

                    if item=='net_after_tax_income':
                        self._add_sub_heading(w_sheet,'EQUITY',row_index)
                        row_index +=1    

                  
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_income_tax_statement(w_sheet)
       
        return row_index
  
    def _populate_cal_income_tax_statement(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. net_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_income_tax_statement','net_sales_revenue',number_format)

        #2. cost_of_goods_sold
        self._transfer_cell_range(w_sheet,'cal_finished_product_inventory','cal_income_tax_statement','cost_of_goods_sold',number_format)
        
        #3. annual_depreciation_expense
        self._transfer_cell_range(w_sheet,'cal_depreciation_for_tax','cal_income_tax_statement','annual_depreciation_expense',number_format)

        #4. total_indirect_labour_cost
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','total_indirect_labour_cost',number_format)
       
        #5. cost_of_electricity_per_pen_per_annum_nominal
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','cost_of_electricity_per_pen_per_annum_nominal',number_format)        
        #6. other_indirect_costs_nominal
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','other_indirect_costs_nominal',number_format)

        #7. sum  total_operating_expenses
        formalue_string =[
                {'header': 'cal_income_tax_statement', 'para': 'cost_of_goods_sold'},
                {'header': 'cal_income_tax_statement', 'para': 'other_indirect_costs_nominal'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'total_operating_expenses')
        
        #8. income_from_operations
        #==I305-I313
        #=A-B
        #=net_sales_revenue - total_operating_expenses 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'total_operating_expenses','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'income_from_operations')

        #9. interest_paid
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_income_tax_statement','interest_expense',
                                   number_format, None,'interest_paid')
        
        #10. pre_tax_income
        #==I315-I317
        #=A-B
        #=net_sales_revenue - total_operating_expenses 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'income_from_operations','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'interest_paid','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'pre_tax_income')

        #11.cumulative_losses
        pre_tax_income_r_index, _, found_state_pre_tax_income= self._retrieve_cell_row_colm('cal_income_tax_statement','pre_tax_income')
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','cumulative_losses')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    prev= cell.column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_cum_tax_cell = prev_letter + '$' + str(cell.row)
                    pre_tax_income_cell = cell.column + '$' + str(pre_tax_income_r_index) if found_state_pre_tax_income else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MIN({prev_cum_tax_cell}+{pre_tax_income_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #11.taxable_income_losses_carried_forward
        pre_tax_income_r_index, _, found_state_pre_tax_income= self._retrieve_cell_row_colm('cal_income_tax_statement','pre_tax_income')
        cumulative_losses_r_index, _, found_state_cumulative_losses= self._retrieve_cell_row_colm('cal_income_tax_statement','cumulative_losses')
   
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','taxable_income_losses_carried_forward')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    prev= cell.column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    pre_tax_income_cell = cell.column + '$' + str(pre_tax_income_r_index) if found_state_pre_tax_income else '0'
                    last_year_cum_looses_cell = prev_letter + '$' + str(cumulative_losses_r_index) if found_state_cumulative_losses else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MAX({last_year_cum_looses_cell}+{pre_tax_income_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #12. taxes
        self._transfer_cell(w_sheet,'taxes','cal_income_tax_statement','corporate_income_tax','0%', 'Inputs')
        
        #13.income_tax_payment
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','corporate_income_tax')
        corporate_income_tax_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
                 
        taxable_income_losses_r_index, _, found_state_taxable_income_losses= self._retrieve_cell_row_colm('cal_income_tax_statement','taxable_income_losses_carried_forward')
   
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','income_tax_payment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    taxable_income_losses_cell = cell.column + '$' + str(taxable_income_losses_r_index) if found_state_taxable_income_losses else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MAX({corporate_income_tax_cell}*{taxable_income_losses_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        
        #14. net_after_tax_income
        #==I315-I317
        #=A-B
        #=pre_tax_income - income_tax_payment 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'pre_tax_income','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'income_tax_payment','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'net_after_tax_income')

        #15. 'share_capital'
        r_index, c_index, found_state = self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = '$' + get_column_letter(c_index) + '$' + str(r_index)
        
        year_header_r_index, _, found_state_year_header= self._retrieve_cell_row_colm('cal_income_tax_statement','header')
        equity_towards_investment_r_index, _, found_state_equity_towards_investment= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity_towards_investment')
        
        shareholder_equity_r_index, _, found_state_shareholder_equity= self._retrieve_cell_row_colm('cal_income_tax_statement','shareholder_equity')
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','share_capital')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    year_header_cell = cell.column + '$' + str(year_header_r_index) if found_state_year_header else '0'
                    equity_towards_investment_cell = cell.column + '$' + str(equity_towards_investment_r_index) if found_state_equity_towards_investment else '0'
                    
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_shareholder_equity_cell = prev_letter + '$' + str(shareholder_equity_r_index) if found_state_shareholder_equity else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=IF({year_header_cell}=Inputs!{base_period_cell},{equity_towards_investment_cell},{prev_shareholder_equity_cell})'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #16 'dividend_payout_ratio'
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cal_income_tax_statement','dividend_payout_ratio','0%', 'Inputs')
        
        #17 'dividend_payouts'
        #=I326*$F$330
        #=A*B
        #=dividend_payout_ratio * net_after_tax_income 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'dividend_payout_ratio','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'dividend_payouts')

        
        #18 'num_of_shares'
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cal_income_tax_statement','num_of_shares',number_format_integer, 'Inputs')
       
        #19 'earning_per_share'
        #=I331/$F$332
        #=A*B
        #=num_of_shares * dividend_payouts 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'MIN(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, 
                {'value': "C", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'dividend_payouts','cell_type': 'cell_range'},
                {'value': ';', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, 
             
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'num_of_shares','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'dividend_payouts','cell_type': 'cell_range'},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, 
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'earning_per_share')

        #20 'retained_earnings'
        #=I326-I331
        #=A-B
        #=net_after_tax_income - dividend_payouts 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'dividend_payouts','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'retained_earnings')

        #21 'shareholder_equity'
        #=I329+I335
        #=A+B
        #=share_capital + retained_earnings 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'share_capital','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'retained_earnings','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'shareholder_equity')

        itemlist= [
            'net_sales_revenue','cost_of_goods_sold','annual_depreciation_expense', 'total_indirect_labour_cost',
            'cost_of_electricity_per_pen_per_annum_nominal','other_indirect_costs_nominal','total_operating_expenses',
            
            'income_from_operations','interest_paid','pre_tax_income','cumulative_losses',
            'taxable_income_losses_carried_forward','corporate_income_tax','income_tax_payment','net_after_tax_income',

            'share_capital','dividend_payout_ratio','num_of_shares','earning_per_share',
            'retained_earnings', 'shareholder_equity' ,'gross_profit',  'gross_profit_margin',
            'net_profit_margin',

            'debt_to_equity_ratio',
            ]
        #22. 'gross_profit'
        #===I305-I308
        #=A-B
        #=net_sales_revenue - cost_of_goods_sold 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'cost_of_goods_sold','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'gross_profit')

        #23. 'gross_profit_margin'
        #=I338/I305
        #=A/B
        #=gross_profit / net_sales_revenue 
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'gross_profit','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'gross_profit_margin',"0.0%")
        
        #24. 'net_profit_margin'
        #=I326/I305
        #=A/B
        #=net_after_tax_income / net_sales_revenue 
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ]
        
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'net_profit_margin',"0.0%")
        
        #25. 'debt_to_equity_ratio'
        #=I326/I305
        #=A/B
        #=begining_debt / shareholder_equity
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'shareholder_equity','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ] 
        
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'debt_to_equity_ratio',"0.0%")

       
    
    def _add_cf_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'CASH FLOW (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cf_nominal' in self.track_inputs:
            self.track_inputs['cf_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cf_nominal']):
            self.track_inputs['cf_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cf_nominal']:
            self.track_inputs['cf_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cf_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'RECEIPTS',row_index)
        row_index +=1

        itemlist= [
                'gross_sales_revenue','change_in_AR','residual_value_of_land_nominal',
                'residual_value_of_machinery_nominal','residual_value_of_buildings_nominal','total_cash_inflow',
                'pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal',

                'total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost',
                'cost_of_electricity_nominal','other_indirect_costs_nominal',

                'change_in_AP','change_in_CB','total_cash_outflow','net_cash_flow_before_taxes','sales_tax_paid',
                'income_tax_payment','net_after_tax_cash_flow_before_financing',
                'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
                'net_after_tax_cash_flow_after_financing',

                'discount_rate_equity','domestic_inflation_rate','nominal_interest_rate',

                'net_after_tax_cash_flow_before_financing_2','total_loan_repayment_outflow_local_2',
                'pv_annual_net_cash_flows', 'pv_annual_debt_repayment',

                'ADSCR','LLCR',
                'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
                'minimum_LLCR','maximum_LLCR','average_LLCR',
            ]

        ic_parameters = {
            # RECEIPTS
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'T_LC'},
            
            # Liquidation value
            'residual_value_of_land_nominal': {'title': 'Residual value of land (nominal)','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_nominal': {'title': "Residual value of machinery (nominal)",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_nominal': {'title': 'Residual value of buildings (nominal)','value': None, 'units': 'T_LC'},
            'total_cash_inflow': {'title': 'TOTAL CASH INFLOW (+)','value': None, 'units': 'T_LC'},

            #EXPENDITURES
            #--------     Investment
            'pens_nominal': {'title': 'Pens (nominal)','value': None, 'units': 'T_LC'},
            'land_nominal': {'title': "Land (nominal)",'value': None, 'units': 'T_LC'},
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
            'buildings_nominal': {'title': "Buildings (nominal)",'value': None, 'units': 'T_LC'},
             #--------     Operating costs
           
            #Operating costs	
            'total_cost_of_imported_inputs_including_import_duty': {'title': "Total cost of imported inputs including import duty",'value': None, 'units': 'T_LC'},
            'total_cost_of_domestic_inputs': {'title': "Total cost of domestic inputs",'value': None, 'units': 'T_LC'},
            'total_labour_cost': {'title': 'Total Labour Cost','value': None, 'units': 'T_LC'},
            'cost_of_electricity_nominal': {'title': "Cost Of Electricity (nominal)",'value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': "Other Indirect Costs (nominal)",'value': None, 'units': 'T_LC'},

            	
            

            'change_in_AP': {'title': "Change in A/P",'value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},

            'total_cash_outflow': {'title': "TOTAL CASH OUTFLOW (-)",'value': None, 'units': 'T_LC'},

            'net_cash_flow_before_taxes': {'title': "NET CASH FLOW BEFORE TAXES",'value': None, 'units': 'T_LC'},

            'sales_tax_paid': {'title': 'Sales tax paid','value': None, 'units': 'T_LC'},
            'income_tax_payment': {'title': "Income tax payment",'value': None, 'units': 'T_LC'},

            
            'net_after_tax_cash_flow_before_financing': {'title': ' Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)','value': None, 'units': 'T_LC'},
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},

            'total_loan_repayment_outflow_local': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},
            'net_after_tax_cash_flow_after_financing': {'title': 'Net after tax cash flow after financing (OWNER PERSPECTIVE)','value': None, 'units': 'T_LC'},

            
            'discount_rate_equity': {'title': 'Discount Rate Equity','value': None, 'units': 'PERCENT'},
            'domestic_inflation_rate': {'title': "Domestic Inflation rate",'value': None, 'units': 'PERCENT'},
            'nominal_interest_rate': {'title': 'Nominal Interest Rate','value': None, 'units': 'PERCENT'},

            'net_after_tax_cash_flow_before_financing_2': {'title': "Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)",'value': None, 'units': 'T_LC'},
            'total_loan_repayment_outflow_local_2': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},

            'pv_annual_net_cash_flows': {'title': 'PV Annual Net Cash Flows (NCF)','value': None, 'units': 'T_LC'},
            'pv_annual_debt_repayment': {'title': "PV Annual Debt Repayment",'value': None, 'units': 'T_LC'},

            'ADSCR': {'title': "ADSCR",'value': None, 'units': 'NUMBER'},
            'LLCR': {'title': "LLCR",'value': None, 'units': 'NUMBER'},
            
             #Summary of ADSCR
            'minimum_ADSCR': {'title': "Minimum ADSCR",'value': None, 'units': 'NUMBER'},
            'maximum_ADSCR': {'title': "Maximum ADSCR",'value': None, 'units': 'NUMBER'},
            'average_ADSCR': {'title': "Average ADSCR",'value': None, 'units': 'NUMBER'},


             #Summary of LLCR
            'minimum_LLCR': {'title': "Minimum LLCR",'value': None, 'units': 'NUMBER'},
            'maximum_LLCR': {'title': "Maximum LLCR",'value': None, 'units': 'NUMBER'},
            'average_LLCR': {'title': "Average LLCR",'value': None, 'units': 'NUMBER'},
            

        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cf_nominal']):
                    self.track_inputs['cf_nominal'][item] = {}

                self.track_inputs['cf_nominal'][item]['row'] = row_index
                self.track_inputs['cf_nominal'][item]['unit'] = _unit
                self.track_inputs['cf_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cf_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['change_in_AR','total_cash_inflow','buildings_nominal',
                    'other_indirect_costs_nominal','pv_annual_debt_repayment','change_in_CB',
                    'total_cash_outflow','net_cash_flow_before_taxes',
                    'income_tax_payment','net_after_tax_cash_flow_before_financing',
                     
                    'total_loan_repayment_outflow_local','net_after_tax_cash_flow_after_financing',
                    'total_loan_repayment_outflow_local_2',
           
                    'LLCR','average_ADSCR'] :
                    row_index +=1 
                    # double line
                    if item=='net_after_tax_cash_flow_after_financing':
                        row_index +=1

                    if item=='total_cash_inflow':
                        self._add_sub_heading(w_sheet,' EXPENDITURES',row_index)
                        row_index +=1
                    
                        self._add_sub_heading2(w_sheet,'Investment',row_index)
                        row_index +=1

                    if item=='change_in_AR':
                        self._add_sub_heading2(w_sheet,'Liquidation value',row_index)
                        row_index +=1
                    
                    if item=='buildings_nominal':
                        self._add_sub_heading2(w_sheet,'Operating costs',row_index)
                        row_index +=1
                    
                    if item=='LLCR':
                        self._add_sub_heading2(w_sheet,'Summary of ADSCR',row_index)
                        row_index +=1
                    
                    if item=='average_ADSCR':
                        self._add_sub_heading2(w_sheet,'Summary of LLCR',row_index)
                        row_index +=1    
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cf_nominal(w_sheet)
       
        return row_index
    
    def _populate_cf_nominal(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #1. gross_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_working_capital_nominal','cf_nominal', 'gross_sales_revenue',number_format,'Calc') 
        
        #2. change_in_AR
        self._transfer_cell_range(w_sheet,'cal_working_capital_nominal','cf_nominal', 'change_in_AR',number_format,'Calc') 

        #3. residual_value_of_land_nominal
        for item in ['residual_value_of_land_nominal','residual_value_of_machinery_nominal',
                     'residual_value_of_buildings_nominal']:
            self._transfer_cell_range(w_sheet,'cal_residual_values','cf_nominal', item,number_format,'Calc') 
        
       
        #7. sum  total_cash_inflow
        formalue_string =[
                {'header': 'cf_nominal', 'para': 'gross_sales_revenue'},
                {'header': 'cf_nominal', 'para': 'residual_value_of_buildings_nominal'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_nominal', 'total_cash_inflow')
       
        #8. investment_cost nominal : pens, building, machinery & land
        source_string =[
            {'source': 'investment_cost_of_pens',             'target': 'pens_nominal',                'header':'calc_investment_cost_nominal'},
            {'source': 'investment_cost_of_land_nominal',     'target': 'land_nominal',                'header':'calc_investment_cost_nominal'},
            {'source': 'total_cost_machinery_nominal',        'target': 'total_cost_machinery_nominal','header':'calc_investment_cost_nominal'},
            {'source': 'investment_cost_of_buildings_nominal', 'target': 'buildings_nominal',          'header':'calc_investment_cost_nominal'},
        ]
        for i in source_string: 
            self._transfer_cell_range(w_sheet, i['header'],'cf_nominal',  i['source'], number_format,'Calc', i['target']) 
    
        
        #9. investment_cost nominal : inputs, labour, electricity, indirect , AP CB
        source_string =[
            {'source': 'total_cost_of_imported_inputs_including_import_duty', 'target': 'total_cost_of_imported_inputs_including_import_duty', 'header':'calc_purchases_nominal' },
            {'source': 'total_cost_of_domestic_inputs',                       'target': 'total_cost_of_domestic_inputs',                       'header':'calc_purchases_nominal'},
            {'source': 'total_labour_cost',                                   'target': 'total_labour_cost',                                   'header':'cal_labour_cost_nominal'},
            {'source': 'cost_of_electricity_per_pen_per_annum_nominal',       'target': 'cost_of_electricity_nominal',                         'header':'cal_labour_cost_nominal'},
            {'source': 'other_indirect_costs_nominal',                        'target': 'other_indirect_costs_nominal',                        'header':'cal_labour_cost_nominal'},
            {'source': 'change_in_AP',                                        'target': 'change_in_AP',                                        'header':'cal_working_capital_nominal'},
            {'source': 'change_in_CB',                                        'target': 'change_in_CB',                                        'header':'cal_working_capital_nominal'},
        ]
        for i in source_string: 
            self._transfer_cell_range(w_sheet, i['header'],'cf_nominal',  i['source'], number_format,'Calc', i['target']) 
    
        #10. sum  total_cash_outflow
        formalue_string =[
                {'header': 'cf_nominal', 'para': 'pens_nominal'},
                {'header': 'cf_nominal', 'para': 'change_in_CB'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_nominal', 'total_cash_outflow')

        #11. sum  net_cash_flow_before_taxes
        #=I16-I34
        #=A-B
        #=total_cash_inflow - total_cash_outflow 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'total_cash_inflow','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_cash_outflow','cell_type': 'cell_range'},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_cash_flow_before_taxes')
       
        #12. sales_tax_paid  
        self._transfer_cell_range(w_sheet, 'cal_production_sales_nominal','cf_nominal',  'sales_tax_paid', number_format,'Calc') 
       
        #13. income_tax_payment  
        self._transfer_cell_range(w_sheet, 'cal_income_tax_statement','cf_nominal',  'income_tax_payment', number_format,'Calc') 
        
        #14. sum  net_after_tax_cash_flow_before_financing
        #=I36-I38-I39
        #=A-B-C
        #=net_cash_flow_before_taxes - sales_tax_paid - income_tax_payment
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_cash_flow_before_taxes','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'sales_tax_paid','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_nominal', 'para': 'income_tax_payment','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_after_tax_cash_flow_before_financing')
       
        #15. debt_cash_inflow_in_local_currency  
        self._transfer_cell_range(w_sheet, 'cal_loan_schedule','cf_nominal',  'debt_cash_inflow_in_local_currency', number_format,'Calc') 
       
        #16. total_loan_repayment_outflow_local  
        self._transfer_cell_range(w_sheet, 'cal_loan_schedule','cf_nominal',  'total_loan_repayment_outflow_local', number_format,'Calc') 
       
            
        #17. sum  net_after_tax_cash_flow_after_financing
        #=I41+I43-I44
        #=A+B-C
        #=net_after_tax_cash_flow_before_financing + debt_cash_inflow_in_local_currency - total_loan_repayment_outflow_local
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'debt_cash_inflow_in_local_currency','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'net_after_tax_cash_flow_after_financing')
       
        
        #18. discount_rate_equity
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_nominal','discount_rate_equity','0%', 'Inputs')
        
        #18. domestic_inflation_rate
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_nominal','domestic_inflation_rate','0%', 'Inputs')
      
        #20. nominal_interest_rate
        #=F49+(1+F49)*F50
        #=A+(1+A)*B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'discount_rate_equity'},
                {'value': '+(1+', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'domestic_inflation_rate'},
              
            ]
        self._cell_formalue_fromstring(w_sheet, formalue_string,'cf_nominal', 'nominal_interest_rate',"0.0%")
        
        #21. net_after_tax_cash_flow_before_financing_2 
        #lb, ub, foundstate= self._loan_principal_repayment_periods(self.myworboook['Inputs'])
        lb, ub, foundstate= self._get_loan_principal_repayment_bounds()
        LUB= {}
        if  foundstate:
            LUB= {'lb': 9 + lb, 'ub':9 + ub}
        self._transfer_cell_range(w_sheet, 'cf_nominal','cf_nominal','net_after_tax_cash_flow_before_financing', 
                     number_format, None,'net_after_tax_cash_flow_before_financing_2',"Linkedcell", False, LUB) 
        
        self._transfer_cell_range(w_sheet, 'cf_nominal','cf_nominal','total_loan_repayment_outflow_local', 
                     number_format, None,'total_loan_repayment_outflow_local_2',"Linkedcell", False, LUB) 
        
       
        #22. pv_annual_net_cash_flows
        #=NPV($F$51;K52:O52)+J52
        #=A+(1+A)*B
        formalue_string =[
                {'value':'A', 'var_type': 'rate', 'header': 'cf_nominal', 'para': 'nominal_interest_rate',},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing_2'},
            ]
        self._npv_cell_range(w_sheet, formalue_string,'cf_nominal', 'pv_annual_net_cash_flows',LUB)
       
        #23. pv_annual_debt_repayment
        #=NPV($F$51;K52:O52)+J52
        formalue_string =[
                {'value':'A', 'var_type': 'rate', 'header': 'cf_nominal', 'para': 'nominal_interest_rate',},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local_2'},
            ]
        self._npv_cell_range(w_sheet, formalue_string,'cf_nominal', 'pv_annual_debt_repayment',LUB)
        
        #24. ADSCR
        #=MAX(IFERROR(J52/J53;0);0)
        #=A/B
        #=net_after_tax_cash_flow_before_financing_2 / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'net_after_tax_cash_flow_before_financing_2','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'total_loan_repayment_outflow_local_2','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'ADSCR',None,LUB)
       
        #24. LLCR
        #=MAX(IFERROR(J55/J56;0);0)
        #=A/B
        #=pv_annual_net_cash_flows / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'pv_annual_net_cash_flows','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'pv_annual_debt_repayment','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'LLCR',None,LUB)
       
        
        #24. LLCR
        #=MAX(IFERROR(J55/J56;0);0)
        #=A/B
        #=pv_annual_net_cash_flows / total_loan_repayment_outflow_local_2
        formalue_string =[
                {'value':'=MAX(IFERROR(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_nominal', 'para': 'pv_annual_net_cash_flows','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_nominal', 'para': 'pv_annual_debt_repayment','cell_type': 'cell_range'},
                {'value': ',0),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
           
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_nominal', 'LLCR',None,LUB)
        
        #25. minimum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'minimum_ADSCR','Calculation',    'MINA',LUB)
        #26. maximum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'maximum_ADSCR','Calculation',    'MAXA',LUB)
        #27. average_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'ADSCR' ,
               number_format , None, 'average_ADSCR','Calculation',    'AVERAGEA',LUB)
       

        #28. minimum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'minimum_LLCR','Calculation',    'MINA',LUB)
        #29. maximum_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'maximum_LLCR','Calculation',    'MAXA',LUB)
        #30. average_ADSCR  
        self._transfer_range_formulae_tocell(w_sheet, 'cf_nominal', 'cf_nominal',  'LLCR' ,
               number_format , None, 'average_LLCR','Calculation',  'AVERAGEA', LUB)
       
        itemlist= [
                'gross_sales_revenue','change_in_AR','residual_value_of_land_nominal',
                'residual_value_of_machinery_nominal','residual_value_of_buildings_nominal','total_cash_inflow',
                
                'pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal',

                'total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost',
                'cost_of_electricity_nominal','other_indirect_costs_nominal',

                'change_in_AP','change_in_CB','total_cash_outflow',
                
                'net_cash_flow_before_taxes','sales_tax_paid',
                'income_tax_payment','net_after_tax_cash_flow_before_financing',
                'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
                'net_after_tax_cash_flow_after_financing',

                'discount_rate_equity','domestic_inflation_rate','nominal_interest_rate',

                'net_after_tax_cash_flow_before_financing_2','total_loan_repayment_outflow_local_2',
                'pv_annual_net_cash_flows', 'pv_annual_debt_repayment',

                'ADSCR','LLCR',
                'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
                'minimum_LLCR','maximum_LLCR','average_LLCR',
            ]

    
    def _add_cf_real_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'CASH FLOW  (Real)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cf_real' in self.track_inputs:
            self.track_inputs['cf_real']= {}# insert inner

        if not ('header' in self.track_inputs['cf_real']):
            self.track_inputs['cf_real']['header'] = {}
  
        if 'header' in self.track_inputs['cf_real']:
            self.track_inputs['cf_real']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cf_real', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       
        itemlist= [
            'domestic_price_index','gross_sales_revenue', 'change_in_AR',

            'residual_value_of_land_real','residual_value_of_machinery_real','residual_value_of_buildings_real',
            'total_cash_inflow',

            'pens_real','land_real','machinery_real','buildings_real',
            'total_cost_of_imported','total_cost_of_domestic_inputs','total_labour_cost',
            'cost_of_electricity_real','other_indirect_costs_real',

            'change_in_AP', 'change_in_CB','total_cash_outflow', 'net_cash_flow_before_taxes',
            'sales_tax_paid','income_tax_paid',
            'net_after_tax_cash_flow_before_financing','debt_cash_inflow_in_local_currency',

            'total_loan_repayment_outflow_local', 'net_after_tax_cash_flow_after_financing',

            'discount_rate_equity', 'NPV','IRR','MIRR'
            ]

       		
        ic_parameters = {
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            #RECEIPTS
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'PERCENT'},
            
             # Liquidation value
            'residual_value_of_land_real': {'title': 'Land ','value': None, 'units': 'T_LC'},
            'residual_value_of_machinery_real': {'title': "Machinery ",'value': None, 'units': 'T_LC'},
            'residual_value_of_buildings_real': {'title': 'Buildings','value': None, 'units': 'T_LC'},
            'total_cash_inflow': {'title': 'TOTAL CASH INFLOW (+)','value': None, 'units': 'T_LC'},


            #EXPENDITURES
            #--------     Investment
            'pens_real': {'title': 'Pens','value': None, 'units': 'T_LC'},
            'land_real': {'title': "Land",'value': None, 'units': 'T_LC'},
            'machinery_real': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'buildings_real': {'title': "Buildings ",'value': None, 'units': 'T_LC'},
           
			
             #Operating costs	
            'total_cost_of_imported': {'title': "Imported inputs purchased",'value': None, 'units': 'T_LC'},
            'total_cost_of_domestic_inputs': {'title': "Domestic inputs purchased",'value': None, 'units': 'T_LC'},
            'total_labour_cost': {'title': 'Labour','value': None, 'units': 'T_LC'},
            'cost_of_electricity_real': {'title': "Cost Of Electricity",'value': None, 'units': 'T_LC'},
            'other_indirect_costs_real': {'title': "Other Indirect Costs",'value': None, 'units': 'T_LC'},

          
            'change_in_AP': {'title': "Change in A/P",'value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},

            'total_cash_outflow': {'title': "TOTAL CASH OUTFLOW (-)",'value': None, 'units': 'T_LC'},

            'net_cash_flow_before_taxes': {'title': "NET CASH FLOW BEFORE TAXES",'value': None, 'units': 'T_LC'},

            'sales_tax_paid': {'title': 'Sales tax','value': None, 'units': 'T_LC'},
            'income_tax_paid': {'title': "Income tax",'value': None, 'units': 'T_LC'},

            
            'net_after_tax_cash_flow_before_financing': {'title': 'Net after tax cash flow before financing (TOTAL INVESTMENT PERSPECTIVE)','value': None, 'units': 'T_LC'},
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},

            'total_loan_repayment_outflow_local': {'title': "Total Loan Repayment as an outflow in Local Currency",'value': None, 'units': 'T_LC'},
            'net_after_tax_cash_flow_after_financing': {'title': 'Net after tax cash flow after financing (OWNER PERSPECTIVE)','value': None, 'units': 'T_LC'},

             'discount_rate_equity': {'title': "Discount Rate Equity",'value': None, 'units': 'PERCENT'},

            'NPV': {'title': "NPV",'value': None, 'units': 'T_LC'},

            'IRR': {'title': 'IRR','value': None, 'units': 'PERCENT'},
            'MIRR': {'title': "MIRR",'value': None, 'units': 'PERCENT'},


            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
               
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
               
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cf_real']):
                    self.track_inputs['cf_real'][item] = {}

                self.track_inputs['cf_real'][item]['row'] = row_index
                self.track_inputs['cf_real'][item]['unit'] = _unit
                self.track_inputs['cf_real'][item]['col'] = 6#redundant
                self.track_inputs['cf_real'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['total_cash_inflow','change_in_AR','buildings_real',
                             'domestic_price_index','total_loan_repayment_outflow_local',
                              'other_indirect_costs_real','change_in_CB', 'total_cash_outflow', 
                              'net_cash_flow_before_taxes','income_tax_paid', 
                              'net_after_tax_cash_flow_before_financing','net_after_tax_cash_flow_after_financing', ] :
                    row_index +=1 

                    if item=='domestic_price_index':
                        self._add_sub_heading(w_sheet,'RECEIPTS',row_index)
                        row_index +=1

                    if item=='total_cash_inflow':
                        self._add_sub_heading(w_sheet,'EXPENDITURES',row_index)
                        row_index +=1 
                        
                        self._add_sub_heading2(w_sheet,'Investment',row_index)
                        row_index +=1       


                    if item=='change_in_AR':
                        self._add_sub_heading2(w_sheet,'Liquidation value',row_index)
                        row_index +=1       

                  
                    if item=='buildings_real':
                        self._add_sub_heading2(w_sheet,'Operation Cost',row_index)
                        row_index +=1       

                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cf_real(w_sheet)
       
        return row_index
      
    def _populate_cf_real(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        number_format_percent ='0%'
        
       
        #1. domestic_price_index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cf_real', 'domestic_price_index',number_format,'Calc') 

        #2. gross_sales_revenue
        #=A/B
        #=gross_sales_revenue /domestic_price_index 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'gross_sales_revenue','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'gross_sales_revenue',number_format)

   
        #3. change_in_AR
        #=A/B
        #=gross_sales_revenue /domestic_price_index 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': 'change_in_AR','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'change_in_AR',number_format)

        
         
        #4 Group Sum Residual Values
        source= ['residual_value_of_land_nominal','residual_value_of_machinery_nominal','residual_value_of_buildings_nominal']
        target = ['residual_value_of_land_real','residual_value_of_machinery_real','residual_value_of_buildings_real',]
        for i in range(len(source)):
            # LOOP
            #=A/B
            #=pens_nominal /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)
      
       
        #5. sum  total_cash_inflow
        formalue_string =[
                {'header': 'cf_real', 'para': 'gross_sales_revenue'},
                {'header': 'cf_real', 'para': 'residual_value_of_buildings_real'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_real', 'total_cash_inflow')
       
        #6 Group Sum Investments
        source= ['pens_nominal','land_nominal','total_cost_machinery_nominal','buildings_nominal']
        target = ['pens_real','land_real','machinery_real','buildings_real',]
        for i in range(len(source)):
            #7. pens_real
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #6 Group Sum Operating costs
        source= ['total_cost_of_imported_inputs_including_import_duty','total_cost_of_domestic_inputs','total_labour_cost', 'cost_of_electricity_nominal','other_indirect_costs_nominal',  'change_in_AP','change_in_CB',]
        target = ['total_cost_of_imported',                             'total_cost_of_domestic_inputs','total_labour_cost', 'cost_of_electricity_real',   'other_indirect_costs_real',   'change_in_AP', 'change_in_CB']
        for i in range(len(source)):
            #7. 
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #7. sum  total_cash_outflow
        formalue_string =[
                {'header': 'cf_real', 'para': 'pens_real'},
                {'header': 'cf_real', 'para': 'change_in_CB'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cf_real', 'total_cash_outflow')

                
        #8 Group Sum Other Operating costs
        source= ['sales_tax_paid',  'income_tax_payment',  'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local']
        target =['sales_tax_paid','income_tax_paid',       'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local']
        for i in range(len(source)):
            #7. 
            #=A/B
            #=A /domestic_price_index 
            formalue_string =[
                    {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': 'A', 'var_type': 'variable','header': 'cf_nominal', 'para': source[i],'cell_type': 'cell_range'},
                    {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                    {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'domestic_price_index','cell_type': 'cell_range'},
                
                ]
            self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', target[i],number_format)

        #9. sum  net_cash_flow_before_taxes
        #=I16-I34
        #=A-B
        #=total_cash_inflow - total_cash_outflow 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'total_cash_inflow','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'total_cash_outflow','cell_type': 'cell_range'},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_cash_flow_before_taxes')
       

        #10. sum  net_after_tax_cash_flow_before_financing
        #=I36-I38-I39
        #=A-B-C
        #=net_cash_flow_before_taxes - sales_tax_paid - income_tax_payment
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_cash_flow_before_taxes','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'sales_tax_paid','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_real', 'para': 'income_tax_paid','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_after_tax_cash_flow_before_financing')
       

        #11. sum  net_after_tax_cash_flow_after_financing
        #=I41+I43-I44
        #=A+B-C
        #=net_after_tax_cash_flow_before_financing + debt_cash_inflow_in_local_currency - total_loan_repayment_outflow_local
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_before_financing','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable','header': 'cf_real', 'para': 'debt_cash_inflow_in_local_currency','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable','header': 'cf_real', 'para': 'total_loan_repayment_outflow_local','cell_type': 'cell_range'},
          
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cf_real', 'net_after_tax_cash_flow_after_financing')
       
        #12. discount_rate_equity
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cf_real','discount_rate_equity','0%', 'Inputs')
      
        #12. NPV
        formalue_string =[
            {'value':'A', 'var_type': 'rate', 'header': 'cf_real', 'para': 'discount_rate_equity',},
            {'value': "B", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._npv_cell(w_sheet, formalue_string,'cf_real', 'NPV',None)

        #13. IRR
        formalue_string =[
            {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._irr_cell(w_sheet, formalue_string,'cf_real', 'IRR',number_format_percent)

        #14. MIRR
        formalue_string =[
            {'value':'A', 'var_type': 'rate', 'header': 'cf_real', 'para': 'discount_rate_equity',},
            {'value': "A", 'var_type': 'variable', 'header': 'cf_real', 'para': 'net_after_tax_cash_flow_after_financing'},
        ]
        self._mirr_cell(w_sheet, formalue_string,'cf_real', 'MIRR',number_format_percent)

    
        
     
    
    def _update_cashflow_linkedcells(self,wb):
        #6. Nominal inflation CF update
        # self._transfer_vcell_range(wb['Calc'],'cf_nominal','cal_loan_schedule','nominal_interest_rate',
        #                           '0%',"CF",'nominal_interest_rate_local')
        
        self._transfer_value_tocellrange(wb['Calc'], 'cf_nominal','cal_loan_schedule', 
                                  'nominal_interest_rate', "CF", 'nominal_interest_rate_local','0.0%')
                         

       
   
    def _populate_cal_loan_schedule(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. real_interest_rate
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','real_interest_rate','0%', 'Inputs')
        #2. risk_premium
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','risk_premium','0%', 'Inputs')        
        
        #3. copy us_inflation_rate
        self._transfer_value_tocellrange(w_sheet, 'macroeconomic_parameters','cal_loan_schedule', 
                    'us_inflation_rate',  "Inputs",  None,     '0%')

      
        #4. nominal_interest_rate
        #=$F$219+$F$220+(1+$F$219+$F$220)*I221
        #=A+B+(1+A+B)*C
        #=real_interest +  real_interest +(1 + real_interest  + real_interest)*usinfaltion
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'real_interest_rate','cell_type': 'single'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'risk_premium','cell_type': 'single'},
                {'value': '+(1+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'us_inflation_rate','cell_type': 'cell_range'},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'nominal_interest_rate',"0%")

        #5. Domestic inflation rate
        self._transfer_value_tocellrange(w_sheet, 'macroeconomic_parameters',  'cal_loan_schedule',
                                        'domestic_inflation_rate',  "Inputs", None,     '0%')
        #6. Nominal inflation CF update
        self._transfer_cell_range(w_sheet,'cf_nominal','cal_loan_schedule','nominal_interest_rate_local','0%',"CF")

        
        #7. us price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_loan_schedule', 'us_price_index',number_format,) 

        #8. num_of_installments
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','num_of_installments',number_format_integer, 'Inputs')
       
        #9. Loan Principal Payment Flags
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule',  'LPP',number_format_integer,"Inputs")

        #10. beginning_debt
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','outstanding_debt_at_end_of_year',
                                   None, None,'beginning_debt',"Calculation",True)

        #11. loan_disbursement
        self._transfer_cell_range(w_sheet,'calc_investment_cost_nominal','cal_loan_schedule','senior_debt_towards_investment',
                                   None, None,'loan_disbursement',"Calculation")

        
        #12. interest_accrued_in_year
        #=I234*I224
        #=A*B
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'nominal_interest_rate_local','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_accrued_in_year')

        #13. prinical_paid
        #=SUM($I$235:$Z$235)/$F$229*K230
        #=SUM(loan_disbursement/num_of_installments*LPP
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_loan_schedule','num_of_installments')
        num_of_installments_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        loan_disbursement_r_index, _, found_state_loan_disbursement= self._retrieve_cell_row_colm('cal_loan_schedule','loan_disbursement')
        LPP_r_index, _, found_state_LPP= self._retrieve_cell_row_colm('cal_loan_schedule','LPP')
        
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_loan_schedule','principal_paid')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    num_of_installments_cell
                    LPP_cell = cell.column + '$' + str(LPP_r_index) if found_state_LPP else '0'
                    #------------------disbursement------------------------------------
                    a= get_column_letter(9) + str(loan_disbursement_r_index)
                    b= get_column_letter(9 + int(span_)) + str(loan_disbursement_r_index)
                    loan_disbursement_cell = f'SUM({a}:{b})' if found_state_loan_disbursement else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF(' + num_of_installments_cell + '=0, 0,' + loan_disbursement_cell + \
                                                              '/'+ num_of_installments_cell  + "*" +  LPP_cell +")"
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #13. interest_paid
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','interest_accrued_in_year',
                                   number_format, None,'interest_paid',"Calculation")

        
        #14. total_loan_repayment
        #=I234*I224
        #=A+B
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'principal_paid','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'total_loan_repayment')
        
        #15. outstanding_debt_at_end_of_year
        # =I234  +I235  + I236 - I239
        # =A + B + C - D
        #=beginning_debt + loan_disbursement + interest_accrued_in_year -total_loan_repayment 
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'loan_disbursement','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'D', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'total_loan_repayment','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'outstanding_debt_at_end_of_year')
          
        #16. debt_cash_inflow_in_local_currency
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','loan_disbursement',
                                   number_format, None,'debt_cash_inflow_in_local_currency',"Calculation")

        #17. total_loan_repayment_outflow_local
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','total_loan_repayment',
                                   number_format, None,'total_loan_repayment_outflow_local',"Calculation")
        
        #18. CP
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule','CP',  number_format_integer, 'Inputs','CP',"Calculation")

        #19. OP
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule','OP', number_format_integer, 'Inputs','OP',"Calculation")

        #20. interest_during_construction
        # =(I236)*I246
        # = A * B 
        # = interest_accrued_in_year *  CP
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'CP','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_during_construction')
          
       
        #21. interest_expense
        #=(I236)*I247
        # = A * B 
        # = interest_accrued_in_year *  OP
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'OP','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_expense')
          
       
    
   
    def _populate_cal_residual_values(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. investment_cost_of_land_real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','cal_residual_values','investment_cost_of_land', 
                           number_format, None,'investment_cost_of_land_real')

        #2. sum cell total_cost_of_land_real
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','investment_cost_of_land_real',
                                          number_format, None,'total_cost_of_land_real','Calculation')        
         
        #3. total_cost_machinery_real
        #==I42+I27
        #=A+B
        #=investment_cost_of_pens +  total_cost_machinery
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'calc_investment_cost_real', 'para': 'investment_cost_of_pens','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'calc_investment_cost_real', 'para': 'total_cost_machinery','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'total_cost_machinery_real')

        #4. sum cell total_cost_machinery
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','total_cost_machinery_real',
                                          number_format, None,'total_cost_machinery','Calculation')        
         
        
        #5. investment_cost_of_land_real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','cal_residual_values','investment_cost_of_buildings_real', 
                           number_format, None,'investment_cost_of_buildings_real')

        #6. sum cell total_cost_buildings
        self._transfer_sumof_cell_range(w_sheet,'cal_residual_values','cal_residual_values','investment_cost_of_buildings_real',
                                          number_format, None,'total_cost_buildings','Calculation')        
     
        
        #7. economic_life_of_machinery
        self._transfer_cell(w_sheet,'depreciation','cal_residual_values','economic_life_of_machinery',number_format_integer, 'Inputs')
        
        #8. economic_life_of_buildings
        self._transfer_cell(w_sheet,'depreciation','cal_residual_values','economic_life_of_buildings',number_format_integer, 'Inputs')
       
        #9. OP
        self._transfer_cell_range(w_sheet,'flags','cal_residual_values','OP', number_format_integer, 'Inputs','OP',"Calculation")

        #10. depreciation_machinery
        #=$F$257*J265/$F$261
        #=A*B/C
        #=total_cost_machinery*OP / economic_life_of_machinery
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'total_cost_machinery','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'OP','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'economic_life_of_machinery','cell_type': 'single'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'depreciation_machinery')
        
        #11. depreciation_buildings
        #=$F$259*I265/$F$262
        #=A*B/C
        #=*total_cost_buildings*OP /  economic_life_of_buildings
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'total_cost_buildings','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'OP','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'economic_life_of_buildings','cell_type': 'single'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'depreciation_buildings')
 
        
      

        #12. domestic_price_index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_residual_values','domestic_price_index', number_format, )

       
        #13. residual_value_of_land_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_of_land_real',
                        number_format_integer, None,'residual_value_of_land_real',"Calculation")
       
        #14. residual_value_of_machinery_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_machinery',
                        number_format_integer, None,'residual_value_of_machinery_real',"Calculation")
        
        #15. residual_value_of_buildings_real
        self._transfer_cell_atcolumn(w_sheet,'cal_residual_values','cal_residual_values','total_cost_buildings',
                        number_format_integer, None,'residual_value_of_buildings_real',"Calculation")
         
        #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_land_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_land_nominal',
                                            number_format,0)
        
          #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_machinery_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_machinery_nominal',
                                            number_format,0)

          #16. residual_value_of_land_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_residual_values', 'para': 'residual_value_of_buildings_real','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_residual_values', 'para': 'domestic_price_index','cell_type': 'cell_range'},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_residual_values', 'residual_value_of_buildings_nominal',
                                            number_format,0)                                    
       
       
    def _populate_cal_depreciation_for_tax(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        
        #1. total_cost_machinery_nominal
        #==I74+I65
        #=A+B
        #=investment_cost_of_pens +  total_cost_machinery_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'calc_investment_cost_nominal', 'para': 'investment_cost_of_pens','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'calc_investment_cost_nominal', 'para': 'total_cost_machinery_nominal','cell_type': 'cell_range'},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'total_cost_machinery_nominal')

        #2. sum cell total_machinery
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','total_cost_machinery_nominal',
                                          number_format, None,'total_machinery','Calculation')        
       
        
        #3. buildings_nominal
        self._transfer_cell_range(w_sheet,'calc_investment_cost_nominal','cal_depreciation_for_tax',
                              'investment_cost_of_buildings_nominal',    number_format, None, 'buildings_nominal')

        
        #4. sum cell total_buildings
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','buildings_nominal',
                                          number_format, None,'total_buildings','Calculation')        
       
        #5. soft_capital_costs_interest_during_construction
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_depreciation_for_tax',
                              'interest_during_construction',    number_format,)

        
        #6. sum cell soft_capital_costs_interest_during_construction_expense
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','interest_during_construction',
                                          number_format, None,'soft_capital_costs_interest_during_construction','Calculation')        
       
        #7. operation_start_year
        self._transfer_cell(w_sheet,'timing','cal_depreciation_for_tax','operation_start_year','General', 'Inputs')
         
        #8. tax_life_of_machinery
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_machinery',number_format_integer, 'Inputs')
        
        #9. tax_life_of_buildings
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_buildings',number_format_integer, 'Inputs')
        
        #10. tax_life_of_soft_capital_costs
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_soft_capital_costs',number_format_integer, 'Inputs')
       
        #11. OP
        self._transfer_cell_range(w_sheet,'flags','cal_depreciation_for_tax','OP', number_format_integer, 'Inputs','OP',"Calculation")

        
        #12. machinery_depreciation_expense
        #=IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$292 -1);  ($F$284/$F$292)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_machinery  '-1),('   total_machinery '/'  tax_life_of_machinery ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_machinery','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'total_machinery','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'machinery_depreciation_expense')

        
        #13. buildings_depreciation_expense
        #==IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$293 -1);  ($F$286/$F$293)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_buildings  '-1),('   total_buildings '/'  tax_life_of_buildings ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_buildings','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'total_buildings','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'buildings_depreciation_expense')

        #14. soft_capital_costs_interest_during_construction_expense
        #==IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$293 -1);  ($F$286/$F$293)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_soft_capital_costs  '-1),('   soft_capital_costs_interest_during_construction '/'  tax_life_of_soft_capital_costs ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_soft_capital_costs','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'soft_capital_costs_interest_during_construction','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'soft_capital_costs_interest_during_construction_expense')

        
       
        #7. sum  annual_depreciation_expense
        formalue_string =[
                {'header': 'cal_depreciation_for_tax', 'para': 'machinery_depreciation_expense'},
                {'header': 'cal_depreciation_for_tax', 'para': 'soft_capital_costs_interest_during_construction_expense'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'annual_depreciation_expense')
        
       
    
    
    def _populate_cal_finished_product_inventory(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        #1. copy closing_inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'closing_inventory',number_format, 'Inputs') 

        #2. copy prev_closing_inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'closing_inventory',number_format, 'Inputs','prev_closing_inventory',"Linkedcell", True,) 

        #3. copy prev total_unit_cost_of_production_per_ton_nominal
        self._transfer_cell_range(w_sheet,'cal_unit_of_production','cal_finished_product_inventory',
                    'total_unit_cost_of_production_per_ton_nominal',number_format, None,
                        'prev_total_unit_cost_of_production_per_ton_nominal',"Linkedcell", True,) 

        #4. cost_of_opening_inventory_fifo
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'prev_total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',
                                            'cost_of_opening_inventory_fifo')

        #5. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'production_quantity',number_format, 'Inputs') 
        
        #6. copy total_unit_cost_of_production_per_ton_nominal
        self._transfer_cell_range(w_sheet,'cal_unit_of_production','cal_finished_product_inventory',
                    'total_unit_cost_of_production_per_ton_nominal',number_format, ) 


        #7. copy sales_quantity
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_finished_product_inventory',
                    'sales_quantity',number_format,) 

        
        #8. quantity_sold_from_this_yr_production
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'sales_quantity','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},  ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',   'quantity_sold_from_this_yr_production')
        #9.  cost_of_proportion_sales_produced
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'quantity_sold_from_this_yr_production','cell_type': 'cell_range'},
                {'value': '/1000', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',  'cost_of_proportion_sales_produced')
        #10.  cost_of_goods_sold
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'cost_of_opening_inventory_fifo','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'cost_of_proportion_sales_produced','cell_type': 'cell_range'},]
                
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',   'cost_of_goods_sold')
        
        #11.  quantity_remained_unsold
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'production_quantity','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'sales_quantity','cell_type': 'cell_range'},]
             
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory', 'quantity_remained_unsold')
        
        #11.  cost_of_closing_inventory
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'quantity_remained_unsold','cell_type': 'cell_range'}  ]
             
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory', 'cost_of_closing_inventory')
        
      
  
    def _add_cal_unit_of_production_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'UNIT COST OF PRODUCTION'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_unit_of_production' in self.track_inputs:
            self.track_inputs['cal_unit_of_production']= {}# insert inner

        if not ('header' in self.track_inputs['cal_unit_of_production']):
            self.track_inputs['cal_unit_of_production']['header'] = {}
  
        if 'header' in self.track_inputs['cal_unit_of_production']:
            self.track_inputs['cal_unit_of_production']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_unit_of_production', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist= [
            'production_quantity',
            'total_direct_labour_cost',
            'direct_labour_cost_per_ton_of_beef_produced',
            
            'OP',
            'total_cost_of_inputs_per_ton_nominal',
            'total_unit_cost_of_production_per_ton_nominal',
            ]

            
        ic_parameters = {
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'},
            'total_direct_labour_cost': {'title': 'Total Direct Labour Cost','value': None, 'units': 'T_LC'},
            'direct_labour_cost_per_ton_of_beef_produced': {'title': 'Direct Labour cost per ton of beef produced','value': None, 'units': 'LC'},
            
            'OP': {'title': 'Operation period','value': None, 'units': 'FLAGE'},# linked cell
            'total_cost_of_inputs_per_ton_nominal': {'title': 'Total cost of inputs per ton (nominal)','value': None, 'units': 'LC'},
            'total_unit_cost_of_production_per_ton_nominal': {'title': 'Total unit cost of production per ton (nominal)','value': None, 'units': 'LC'},
        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_unit_of_production']):
                    self.track_inputs['cal_unit_of_production'][item] = {}

                self.track_inputs['cal_unit_of_production'][item]['row'] = row_index
                self.track_inputs['cal_unit_of_production'][item]['unit'] = _unit
                self.track_inputs['cal_unit_of_production'][item]['col'] = 6#redundant
                self.track_inputs['cal_unit_of_production'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['direct_labour_cost_per_ton_of_beef_produced'] :
                    row_index +=1 

                    
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_unit_of_production(w_sheet)
       
        return row_index
    
    def _add_cal_labour_cost_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'LABOUR COSTS AND OTHER INDIRECT COSTS (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_labour_cost_nominal' in self.track_inputs:
            self.track_inputs['cal_labour_cost_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_labour_cost_nominal']):
            self.track_inputs['cal_labour_cost_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_labour_cost_nominal']:
            self.track_inputs['cal_labour_cost_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_labour_cost_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        self._add_sub_heading(w_sheet,'DIRECT LABOUR',row_index)
        row_index +=1

        itemlist= ['num_of_workers','annual_increase_salaries_workers','monthly_wage_per_worker',
            'number_of_months_in_a_year', 
            
            'real_yearly_wage_rate','domestic_price_index',
            'nominal_yearly_wage_rate','total_direct_labour_cost',

            'num_of_supervisors_technicians','annual_increase_salaries_supervisors_technicians',
            'monthly_wage_per_supervisor','real_yearly_salary_rate','nominal_yearly_salary_rate',
            'total_indirect_labour_cost',

            'total_labour_cost',

            'OP','cost_of_electricity_per_pen_per_annum',
            'cost_of_electricity_per_pen_per_annum_nominal',
            'other_indirect_costs',
            'other_indirect_costs_nominal',
            ]
            
        ic_parameters = {
            # DIRECT LABOUR
            'num_of_workers': {'title': 'Number of workers','value': None, 'units': 'NUMBER'},
            'annual_increase_salaries_workers': {'title': 'Annual increase in real salaries of workers','value': None, 'units': 'PERCENT'},
            'monthly_wage_per_worker': {'title': 'Monthly wage for workers','value': None, 'units': 'T_LC'},
            'number_of_months_in_a_year': {'title': 'Number of months in a year','value': None, 'units': 'NUMBER'},# linked cell
            'real_yearly_wage_rate': {'title': 'Real yearly wage rate','value': None, 'units': 'T_LC'},

            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'nominal_yearly_wage_rate': {'title': 'Nominal yearly wage rate','value': None, 'units': 'T_LC'},
            'total_direct_labour_cost': {'title': 'Total Direct Labour Cost','value': None, 'units': 'T_LC'},
            
            #INDIRECT LABOUR
            'num_of_supervisors_technicians': {'title': 'Number of supervisors & technicians','value': None, 'units': 'NUMBER'},
            'annual_increase_salaries_supervisors_technicians': {'title': 'Annual increase in real salaries of supervisors & technicians','value': None, 'units': 'PERCENT'},
            'monthly_wage_per_supervisor': {'title': 'Monthly wage for supervisors & technicians','value': None, 'units': 'T_LC'},
            'real_yearly_salary_rate': {'title': 'Real yearly salary rate','value': None, 'units': 'T_LC'},
            'nominal_yearly_salary_rate': {'title': 'Nominal yearly salary rate','value': None, 'units': 'T_LC'},
            'total_indirect_labour_cost': {'title': 'Total Indirect Labour Cost','value': None, 'units': 'T_LC'},
            
            'total_labour_cost': {'title': 'Total Labour Cost','value': None, 'units': 'T_LC'},


            'OP': {'title': 'Operation period','value': None, 'units': 'FLAG'},
            #'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            'cost_of_electricity_per_pen_per_annum': {'title': 'Cost Of Electricity per pen per annum','value': None, 'units': 'LC'},
            'cost_of_electricity_per_pen_per_annum_nominal': {'title': 'Cost of electricity (nominal)','value': None, 'units': 'T_LC'},
            'other_indirect_costs': {'title': 'Other Indirect Costs','value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': 'Other Indirect Costs (nominal)','value': None, 'units': 'T_LC'},


        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_labour_cost_nominal']):
                    self.track_inputs['cal_labour_cost_nominal'][item] = {}

                self.track_inputs['cal_labour_cost_nominal'][item]['row'] = row_index
                self.track_inputs['cal_labour_cost_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_labour_cost_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_labour_cost_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['total_direct_labour_cost','real_yearly_wage_rate', 
                              'total_indirect_labour_cost', 'total_labour_cost' ] :
                    row_index +=1 

                    if item =='total_direct_labour_cost':
                        self._add_sub_heading(w_sheet,'INDIRECT LABOUR',row_index)
                        row_index +=1


        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_labour_cost_nominal(w_sheet)
       
        return row_index
                     
             
    def _add_cal_working_capital_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'WORKING CAPITAL (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'cal_working_capital_nominal' in self.track_inputs:
            self.track_inputs['cal_working_capital_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_working_capital_nominal']):
            self.track_inputs['cal_working_capital_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_working_capital_nominal']:
            self.track_inputs['cal_working_capital_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_working_capital_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['accounts_receivable','gross_sales_revenue','accounts_receivable_val',
                   'accounts_payable', 'total_input_cost_nominal','accounts_payable_val',
                   'cash_balance','cash_balance_val',
                   'change_in_AR','change_in_AP','change_in_CB']
        ic_parameters = {
            # PI
            'accounts_receivable': {'title': 'Accounts receivable [% of Gross Sales]','value': None, 'units': 'PERCENT'},
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'accounts_receivable_val': {'title': 'Accounts receivable [% of Gross Sales]','value': None, 'units': 'T_LC'},# linked cell
            
            'accounts_payable': {'title': 'Acccounts payable (% of total input cost)','value': None, 'units': 'PERCENT'},# linked cell
            'total_input_cost_nominal': {'title': 'Total Input Cost (nominal)','value': None, 'units': 'T_LC'},
            'accounts_payable_val': {'title': 'Acccounts payable (% of total input cost)','value': None, 'units': 'T_LC'},
            
            'cash_balance': {'title': 'Cash balance  (% of gross sales)','value': None, 'units': 'PERCENT'},
            'cash_balance_val': {'title': 'Cash balance  (% of gross sales)','value': None, 'units': 'T_LC'},
            
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'T_LC'},
            'change_in_AP': {'title': 'Change in A/P','value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},
        }
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_working_capital_nominal']):
                    self.track_inputs['cal_working_capital_nominal'][item] = {}

                self.track_inputs['cal_working_capital_nominal'][item]['row'] = row_index
                self.track_inputs['cal_working_capital_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_working_capital_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_working_capital_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['accounts_receivable_val','accounts_payable_val','cash_balance_val'] :
                    row_index +=1 

        
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_working_capital_nominal(w_sheet)
       
        return row_index 

    def _add_cal_investment_cost_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'INVESTMENT COST (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_investment_cost_nominal' in self.track_inputs:
            self.track_inputs['calc_investment_cost_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['calc_investment_cost_nominal']):
            self.track_inputs['calc_investment_cost_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['calc_investment_cost_nominal']:
            self.track_inputs['calc_investment_cost_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_investment_cost_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['domestic_price_index','cost_of_pens_constructed','investment_cost_of_pens',
                   'million_to_thousand', 'investment_cost_of_land','investment_cost_of_land_nominal',
                   'investment_costs_over_run_factor',
                   'total_cost_machinery', 'total_cost_machinery_nominal', 'investment_cost_of_buildings', 
                   'investment_cost_of_buildings_nominal', 'equity', 'senior_debt',
                   'total_investment_cost','total_local_investment',
                   'equity_towards_investment','senior_debt_towards_investment']
        
           
        ic_parameters = {
            # PI
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'NUMBER'},
            
            # Feedlots
            'cost_of_pens_constructed': {'title': 'Cost of feedlots constructed','value': None, 'units': 'T_LC'},
            'investment_cost_of_pens': {'title': 'Investment cost of Feedlots (nominal)','value': None, 'units': 'T_LC'},# linked cell
            
            # Land
            'investment_cost_of_land': {'title': 'Investment cost of land','value': None, 'units': 'T_LC'},# linked cell
            'investment_cost_of_land_nominal': {'title': 'Investment cost of land (nominal)','value': None, 'units': 'T_LC'},
           
            # Machinery
            'total_cost_machinery': {'title': 'Total Cost of Machinery including import duty (real)','value': None, 'units': 'T_LC'},
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
         
            # Buildings
            'investment_cost_of_buildings': {'title': 'Investment cost of buildings (LC$ 000)','value': None, 'units': 'T_LC'},
            'investment_cost_of_buildings_nominal': {'title': 'Buildings (nominal)','value': None, 'units': 'T_LC'},
            # Financing Parameters
            'equity': {'title': 'Equity (% of Investment Costs)','value': None, 'units': 'PERCENT'},
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': None, 'units': 'PERCENT'},

             # Financing Parameters
            'total_investment_cost': {'title': 'Total Investment Cost (nominal)','value': None, 'units': 'T_LC'},
            'equity_towards_investment': {'title': 'Equity Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
            'senior_debt_towards_investment': {'title': 'Senior Debt Contribution towards Total Investment Costs','value': None, 'units': 'T_LC'},
        }

   

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_investment_cost_nominal']):
                    self.track_inputs['calc_investment_cost_nominal'][item] = {}

                self.track_inputs['calc_investment_cost_nominal'][item]['row'] = row_index
                self.track_inputs['calc_investment_cost_nominal'][item]['unit'] = _unit
                self.track_inputs['calc_investment_cost_nominal'][item]['col'] = 6#redundant
                self.track_inputs['calc_investment_cost_nominal'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['domestic_price_index','investment_cost_of_pens','investment_cost_of_land_nominal',
                            'total_cost_machinery_nominal','investment_cost_of_buildings_nominal','senior_debt'] :
                    row_index +=1 

                     #==========Sub-headings=========================
                    if item =='domestic_price_index':
                        self._add_sub_heading(w_sheet,'Feedlots(nominal)',row_index)
                        row_index +=1


                    if item =='investment_cost_of_pens':
                        self._add_sub_heading(w_sheet,'Land(nominal)',row_index)
                        row_index +=1

                    if item =='investment_cost_of_land_nominal':
                        self._add_sub_heading(w_sheet,'Machinery(nominal)',row_index)
                        row_index +=1  

                    if item =='total_cost_machinery_nominal':
                        self._add_sub_heading(w_sheet,'Buildings(nominal)',row_index)
                        row_index +=1 

                    if item =='investment_cost_of_buildings_nominal':
                        self._add_sub_heading(w_sheet,'Financing Parameters',row_index)
                        row_index +=1  

                    if item =='senior_debt':
                        self._add_sub_heading(w_sheet,'Financing',row_index)
                        row_index +=1              

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_investment_cost_nominal(w_sheet)
       
        return row_index 


    def _populate_cal_investment_cost_nominal(self, w_sheet):
         
        #*************************************************Pens*****************************************************
        #copy price index
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','calc_investment_cost_nominal',
                    'domestic_price_index',number_format) 

        #------Cost of pens constructed
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'cost_of_pens_constructed',number_format,)        
        #------------Nominal cost of Pens
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','cost_of_pens_constructed')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_pens_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        #**********************************Land Real*************************************
        #---Land 
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'investment_cost_of_land',number_format)
        
        
         #------------Nominal Land cost
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_land_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        
        
       
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
          #***********************************Trancefer Cells from Input *********************************
        #----Machinery
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'total_cost_machinery',number_format, )
        
        #------To cost of machiner in LC [importy duty added]-----------------------------------------
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_machinery_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        

       

                  
        #------Equity LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_nominal','equity','0%', 'Inputs')        
        #------Senior LinkedCell-------------------------------------------
        self._transfer_cell(w_sheet,'financing','calc_investment_cost_nominal','senior_debt','0%', 'Inputs')


      
        # Buiding Real
        self._transfer_cell_range(w_sheet,'calc_investment_cost_real','calc_investment_cost_nominal',
                    'investment_cost_of_buildings',number_format)
        
        #Building deprecaite over years: land doesnt
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings')
        infla_dom_pi_r_index, _, found_state_pi_domestic= self._retrieve_cell_row_colm('calc_investment_cost_nominal','domestic_price_index')
    
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                    price_index_domestic_cell = cell.column + '$' + str(infla_dom_pi_r_index) if found_state_pi_domestic else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ investment_cost_of_buildings_cell +'*' +  price_index_domestic_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        

       
         #------Total Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        invest_machinery_r_index, _, found_state_invest_machinery= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_cost_machinery_nominal')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_machinery_cell = cell.column + '$' + str(invest_machinery_r_index) if found_state_invest_machinery else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_machinery_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
         #------Total Local Investment cost-----------------------------------------       
        invest_pens_r_index, _, found_state_invest_pens= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_pens')
        invest_land_r_index, _, found_state_invest_land= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_land_nominal')
        invest_buildings_r_index, _, found_state_invest_buildings= self._retrieve_cell_row_colm('calc_investment_cost_nominal','investment_cost_of_buildings_nominal')
   
        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_local_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    investment_cost_of_pens_cell = cell.column + '$' + str(invest_pens_r_index) if found_state_invest_pens else '0'
                    investment_cost_of_land_cell = cell.column + '$' + str(invest_land_r_index) if found_state_invest_land else '0'
                    investment_cost_of_buildings_cell = cell.column + '$' + str(invest_buildings_r_index) if found_state_invest_buildings else '0'
                  
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=SUM('+ investment_cost_of_pens_cell + ',' \
                                                                     + investment_cost_of_land_cell + ',' \
                                                                     + investment_cost_of_buildings_cell +')' 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Equity toward Total Investment-------------------------
        #equity
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity')
        equity_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ equity_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
        
        #-----------------------------Senoir Debt Contributiom Total Investment-------------------------
        #senior_debt
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','senior_debt')
        senior_debt_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        #cif
        total_invest_r_index, _, found_state_total_invest= self._retrieve_cell_row_colm('calc_investment_cost_nominal','total_investment_cost')
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_investment_cost_nominal','senior_debt_towards_investment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    total_invest_cell = cell.column + '$' + str(total_invest_r_index) if found_state_total_invest else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ total_invest_cell +'*'+ senior_debt_cell  
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   

    def _sumofcolumn_range_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, number_format=None,last_column=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        """ 
        =SUM(I308:I312)
        =SUM( I  row_start :I   row_end)
        """
        row_start=None
        row_end = None
        for item in formalue_string:
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    if row_start ==None:
                        row_start =source_r_index
                    elif row_start > source_r_index:
                        row_start = source_r_index 

                    if row_end ==None:
                        row_end =source_r_index
                    elif row_end < source_r_index:
                        row_end = source_r_index 
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

            if last_column !=None and type(last_column)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(last_column)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    maths_formulae ='=SUM('+ cell.column + str(row_start) + ':' + cell.column +  str(row_end) +')'
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                      
    def _calculate_formalue_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, 
                        number_format=None,lower_upper_bound=None ):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        """ 
        #=IF(I187=0;0;I188*$F$30/I187)
        '=IF('+ I187 + '=0,0'+ I188 + '*' + $F$30 + '/' + I187 + ')'
        '=IF('+ A + '=0,0' + B + '*' + C + '/' + D + ')'

        ['=IF(', A , '=0,0' , B , '*' , C , '/' , D , ')']

        ['=IF(',  '=0,0' ,  '*' ,  '/'  ')'] [ A ,  B ,  C ,  D , ] 
        """
        #seperate string, variables
        variables_ ={}
        memory_={}
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='variable' and item['cell_type']== 'single':
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                variables_[item['value']]= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable' and item['cell_type']== 'cell_range':
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                memory_[item['value']]={}
                memory_[item['value']]['row']= source_r_index
                memory_[item['value']]['found_state'] = found_state_source

       
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39


            #change if possible
            #int or dict
            if lower_upper_bound !=None and type(lower_upper_bound)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(lower_upper_bound)) + str(r_index) # D39
            elif lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
           
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    maths_formulae =''
                    for item in formalue_string:
                        if item['var_type']=='const' :
                          maths_formulae += item['value']
                        elif item['var_type']=='variable' and item['cell_type']== 'single':
                            #retrieve from variabble
                            maths_formulae +=  variables_[item['value']] 
                        elif item['var_type']=='variable' and item['cell_type']== 'cell_range':
                            source_r_index = memory_[item['value']]['row']
                            found_state = memory_[item['value']]['found_state']
                            variables_[item['value']]= cell.column + '$' + str(source_r_index) if found_state else '0'
                            maths_formulae +=  variables_[item['value']] 
                        elif item['value'] in variables_:
                            #already calculated
                            #print(item['value'],variables_[item['value']]) 
                            maths_formulae +=  variables_[item['value']]      
                    #print(maths_formulae) 
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                   

         
    def _cell_formalue_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, number_format=None, 
                        cell_style='Calculation'):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
      
        #seperate string, variables
        variables_ ={}
        memory_={}
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='variable' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                #keep previous variable unspoilt
                #subsequent similar variable are skiipped
                if not(item['value'] in variables_):
                    variables_[item['value']]=  get_column_letter(c_index) + str(r_index) if found_state else '0'
           
        #write in target--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            maths_formulae =''
            for item in formalue_string:
                if item['var_type']=='const' :
                    maths_formulae += item['value']
                elif item['var_type']=='variable' :
                    #retrieve from variabble
                    maths_formulae +=  variables_[item['value']] 
                elif item['value'] in variables_:
                    #already calculated
                    #print(item['value'],variables_[item['value']]) 
                    maths_formulae +=  variables_[item['value']]      
            col= get_column_letter(c_index) 
            #print("formuales: ", maths_formulae)
            w_sheet['%s%s'%(col, r_index)] = maths_formulae
            w_sheet['%s%s'%(col, r_index)].style = cell_style
            w_sheet['%s%s'%(col, r_index)].number_format =number_format
                   
     
    def _sum_oflist_cell_range(self, w_sheet,
                        list_, target_header, target_para, const_list=[]):
        #----List
        mat_attr= []
        for item in list_:
            header_ =item['header']
            para_ =item['para']
            action_ =item['action']           
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_

            mat_attr.append(prop)
        #print(mat_attr)
        mul_const= ''
        div_const= ''

        for x in const_list:
            header_ =x['header']
            para_ =x['para']
            action_ = x['action']          
            #source
            r_index, c_index, found_state= self._retrieve_cell_row_colm(header_,para_)
            val_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
            if mul_const ==''  and action_ =="*":
               mul_const = val_cell
            elif action_ =="*":
               mul_const += '*' +  val_cell

            if div_const ==''  and action_ =="/" :#and val_cell!='0':# avoid zero divsion
               div_const = val_cell
            elif action_ =="/": #and val_cell !='0':# avoid zero divsion
               div_const += '*' +  val_cell    

        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    
                    maths_formulae =''
                    for i_val in mat_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if maths_formulae =='' and i_val['action'] in ["-"]:
                            maths_formulae = i_val['action'] + val
                        elif maths_formulae =='':  
                            maths_formulae = val
                        else:
                            maths_formulae += i_val['action'] + val        
                    
                    # preepend const if value
                    if not (mul_const==""):
                        maths_formulae = mul_const + "*" + maths_formulae    
                    # append division if value
                    if not (div_const==""):
                        maths_formulae = maths_formulae + "/" + div_const 

                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

     
    def _productof_sumof_list_cell_range(self, w_sheet,
                        listA, listB, target_header, target_para,):
        # mulplies cell range sum together
        #----List
        listA_attr= []
        listB_attr= []
        for item in listA:
            header_ =item['header']
            para_ =item['para']
            action_ =item['action']           
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_
            listA_attr.append(prop)
       
        for x in listB:
            header_ =x['header']
            para_ =x['para']
            action_ = x['action']          
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_
            
            listB_attr.append(prop)  
               
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    
                    mathsA_formulae =''
                    for i_val in listA_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if mathsA_formulae =='' and i_val['action'] in ["-"]:
                            mathsA_formulae = i_val['action'] + val
                        elif mathsA_formulae =='':  
                            mathsA_formulae = val
                        else:
                            mathsA_formulae += i_val['action'] + val   

                    mathsA_formulae = "(" + mathsA_formulae +")" if len(listA_attr)>1 else mathsA_formulae

                    mathsB_formulae =''
                    for i_val in listB_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if mathsB_formulae =='' and i_val['action'] in ["-"]:
                            mathsB_formulae = i_val['action'] + val
                        elif mathsB_formulae =='':  
                            mathsB_formulae = val
                        else:
                            mathsB_formulae += i_val['action'] + val

                    mathsB_formulae = "(" + mathsB_formulae +")" if len(listB_attr)>1 else mathsB_formulae
                    
                   
                   
                    # preepend const if value
                    if not (mathsA_formulae=="" or mathsB_formulae==""):
                         maths_formulae= mathsA_formulae + "*" + mathsB_formulae    
                    elif not (mathsA_formulae==""):
                        maths_formulae= mathsA_formulae 
                    elif not (mathsB_formulae==""):
                        maths_formulae= mathsB_formulae 

                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

   

    def _delta_diff_in_cell_range(self, w_sheet,
                                     source_header,
                                     source_para, target_para, 
                                     reverse_diff = False, target_header = None,):
       
        #source
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header,source_para)
        
        #write in target--------------
        target_header = source_header if target_header==None else target_header
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_source_cell = prev_letter + str(source_r_index)
                   
                    source_cell = cell.column + '$' \
                        + str(source_r_index) if found_state_source else '0'
                    if reverse_diff:
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + prev_source_cell +'-'+ source_cell
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + source_cell + '-' + prev_source_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

    def _product_value_and_cell_range(self, w_sheet, 
                          val_header , val_para,
                          source_para, target_para, source_header=None, target_header = None,):
        #value
        r_index, c_index, found_state= self._retrieve_cell_row_colm(val_header,val_para)
        val_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        #source
        source_header = val_header if source_header==None else source_header
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header,source_para)
        
        #write in target--------------
        target_header = source_header if target_header==None else target_header
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    source_cell = cell.column + '$' \
                        + str(source_r_index) if found_state_source else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + val_cell  + '*' + source_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'


     
    def _transfer_sumof_cell_range(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format=None, source_wksheet=None, 
                        target_para = None, cell_style='Linkedcell'):

        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header, source_para)
        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        
        if found_state:
            span_ =self._get_span()
           
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''

            a= get_column_letter(9) + str(source_r_index)
            b= get_column_letter(9 + int(span_)) + str(source_r_index)
            sumofsource_range = f'SUM({a}:{b})' if found_state_source else '0'
                   
            col= get_column_letter(c_index)
            w_sheet['%s%s'%(col, r_index)] = f'={sumofsource_range}'
            w_sheet['%s%s'%(col, r_index)].style = cell_style
            if cell_number_format !=None:
                w_sheet['%s%s'%(col, r_index)].number_format = cell_number_format

    def _get_loan_principal_repayment_bounds(self,):
        found_lb= False
        lbound= None
        ubound= None
        
        base_period= int(self.timing_assumptions['base_period']['value'])
        vara =   int(self.financing['repayment_starts']['value'])
        varb =  int(self.financing['repayment_starts']['value']) + int(self.financing['num_of_installments']['value'])  - int(self.financing['grace_period']['value'])-1
        operation_end= int(self.timing_assumptions['operation_end']['value'])
        
        list_ =[]
        continue_flag= True
        ONES_found= False
        period_iter=base_period
       
        while continue_flag:
            
            span_num = period_iter - base_period
            if not(period_iter < vara  or period_iter > varb):
                # Repayment period
                ONES_found = True
                found_lb= True
                if lbound==None:
                    #latch the first
                    lbound= span_num
                if ubound==None:
                    ubound=span_num 
                else:
                    #continue updating
                    ubound=span_num 
            else:
                #outside repayment perion
                #0 preceding 1 means repayment done
                if ONES_found == True:
                    continue_flag=False
            # premature exist no period found        
            if period_iter >operation_end:
                continue_flag=False

            period_iter += 1   
        return lbound, ubound, found_lb

    def _loan_principal_repayment_periods(self, w_sheet):

        r_index, _, found_state= self._retrieve_cell_row_colm('flags', 'LPP')
        found_lb= False
        lbound= None
        ubound= None
        if found_state:
            span_ =self._get_span()
            # default
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
           
           
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    col_num = column_index_from_string(cell.column)
                    cell_value =  w_sheet['%s%s'%(cell.column, cell.row)].value
                    #print("found_integer:....", cell_value)
                    if int(cell_value) ==1:
                        found_lb= True
                        if lbound==None:
                            #latch the first
                           lbound=col_num
                    if ubound==None:
                        ubound=col_num 
                    else:
                       #continue updating
                       ubound=col_num 
        return lbound, ubound, found_lb
    
    
    
    
   
    
    def _mirr_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        #=IFERROR(MIRR(I111:Z111;$F$113;$F$113);"N/A")
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=IFERROR(MIRR(' + first_source_cell + ':' + last_source_cell + ',' + rate_cell + ',' + rate_cell +'),"N/A")' 
                 
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format ='0%' #number_format


    def _irr_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            
            if item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=IFERROR(IRR(' + first_source_cell + ':' + last_source_cell + '),"N/A")' 
                  
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = '0%'#number_format

    def _npv_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    next_source_cell = get_column_letter(next_) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=NPV('+ rate_cell + ','+ next_source_cell + ':' + last_source_cell + ')+' + first_source_cell
                      
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = number_format

    def _npv_cell_range(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
              
       
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

            upperbound_=9 + int(span_)
            
            #change if possible
            if lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                    
                    upperbound_ = ub

                    span_= ub-lb




            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    #=NPV($F$51;K52:O52)+J52
                    ahead =  column_index_from_string(cell.column) + 1
                    next_col = column_index_from_string(cell.column)+ span_
                    curr_source_cell = cell.column + str(source_r_index) if found_state_source else '0'
                    ahead_source_cell = get_column_letter(ahead) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(next_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=NPV('+ rate_cell + ','+ ahead_source_cell + ':' + last_source_cell + ')+' + curr_source_cell
                    
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                               
       
    
    def _transfer_cell_range(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format=None, source_wksheet=None, 
                        target_para = None, cell_style='Linkedcell',shift_prev=False,lower_upper_bound= None):
        
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header, source_para)
        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)

        if found_state:
            span_ =self._get_span()
            # default
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #change if possible
            if lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07

            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:

                    source_cell = cell.column + '$' + str(source_r_index) if found_state_source else '0'
                    if shift_prev==True:
                        prev= cell.column
                        prev_col = column_index_from_string(cell.column)-1
                        prev_letter= get_column_letter(prev_col)
                        source_cell = prev_letter + '$' + str(source_r_index) if found_state_source else '0'

                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'={source_wksheet}{source_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = cell_style
                    if cell_number_format !=None:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = cell_number_format
       
    def _transfer_cell_atcolumn(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None,cell_style='Linkedcell',last_column=None):
        
        if last_column == None: 
            span_ =self._get_span()
            last_column = get_column_letter(9 + int(span_)) 
        
        #------Senior Debt LinkedCell-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        #print(source_header, source_para, "col:", c_index)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            w_sheet['%s%s'%(last_column, r_index)] = f'={source_wksheet}{source_cell}'
            w_sheet['%s%s'%(last_column, r_index)].style = cell_style
            w_sheet['%s%s'%(last_column, r_index)].number_format = cell_number_format
    
    def _get_cell_ref_asformulae(self,  source_header,     source_para , 
                                 source_wksheet=None,   from_col_pos=None):
      
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        if from_col_pos !=None and type(from_col_pos)==int:
            c_index = 9 + int(from_col_pos)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #print(f'={source_wksheet}{source_cell}')
            return f'={source_wksheet}{source_cell}'
        
        #nothing found
        return '0'   
    def _write_text_atcell(self, w_sheet, target_header, target_para, text_value,
                           target_col_pos=None, cell_number_format='General', cell_style=None, ):

        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if target_col_pos !=None and type(target_col_pos)==int:
            c_index = 9 + int(target_col_pos)
        if found_state:
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = text_value
            if cell_style !=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
    
    def _transfer_cell(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None, cell_style='Linkedcell', 
                        from_col_pos=None, target_col_pos=None):
         #------Senior Debt LinkedCell-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        #print(source_header, source_para, "col:", c_index)
        if from_col_pos !=None and type(from_col_pos)==int:
            c_index = 9 + int(from_col_pos)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if target_col_pos !=None and type(target_col_pos)==int:
            c_index = 9 + int(target_col_pos)
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = f'={source_wksheet}{source_cell}'
            #dont change formats if to retain former
            if cell_style !=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            if cell_number_format !=None:    
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
    
    def _transfer_range_formulae_tocell(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None,cell_style='Linkedcell',
                        formulae_str='Sum',lower_upper_bound=None):
        
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #change if possible
                #int or dict
            if lower_upper_bound !=None and type(lower_upper_bound)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(lower_upper_bound)) + str(r_index) # D39
            elif lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''        
            source_cell = f'{formulae_str}({source_wksheet}{first_slice_point}:{source_wksheet}{second_slice_point})'
            
        else:
            source_cell =  '0'
           

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            #print({source_cell})
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = f'={source_cell}'
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format

    def _transfer_value_tocell(self, w_sheet, source_cell, 
                        target_header, target_para,
                        cell_number_format=None,
                        cell_style='Linkedcell'):
        
        #if None=== set default as source paramter
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = source_cell
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            if cell_number_format!=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
            else:
                #print ('_transfer_value_tocell','Default formating')
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = 'General' 

    
    def _transfer_value_tocellrange(self, w_sheet, source_header, 
                        target_header, source_para,  source_wksheet=None,target_para=None,
                        cell_number_format=None,
                        cell_style='Linkedcell'):
        
        
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        target_para = source_para if target_para==None else target_para
        #if None=== set default as source paramter
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'={source_wksheet}{source_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = cell_style
                    if cell_number_format !=None:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = cell_number_format
       

                            
    def _add_cal_inflation_price_indices_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'INFLATION, EXCHANGE RATE AND PRICE INDICES' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_inflation_price_index' in self.track_inputs:
            self.track_inputs['calc_inflation_price_index']= {}# insert inner

        if not ('header' in self.track_inputs['calc_inflation_price_index']):
            self.track_inputs['calc_inflation_price_index']['header'] = {}
  
        if 'header' in self.track_inputs['calc_inflation_price_index']:
            self.track_inputs['calc_inflation_price_index']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_inflation_price_index', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        #==========Sub-heading=========================
      
        self._add_sub_heading(w_sheet,'Domestic',row_index)
        row_index +=1

        itemlist= ['domestic_inflation_rate','domestic_price_index',
                   'us_inflation_rate', 'us_price_index',
                   'relative_price_index', 'real_exchange_rate', 'nominal_exchange_rate']
        
        iep_parameters = {
            'domestic_inflation_rate': {'title': 'Domestic Inflation rate','value': None, 'units': 'PERCENT'},
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'PERCENT'},
            'us_inflation_rate': {'title': 'US Inflation rate','value': None, 'units': 'PERCENT'}, 
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},
            'relative_price_index': {'title': 'Relative Price Index','value': None, 'units': 'NUMBER'},
            'real_exchange_rate': {'title': 'Real Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
            'nominal_exchange_rate': {'title': 'Nominal Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
        }
       

        #-----------------------------------
        for item in itemlist:
            if item in iep_parameters:
                _unit =iep_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        iep_parameters[item]['title'], iep_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_inflation_price_index']):
                    self.track_inputs['calc_inflation_price_index'][item] = {}

                self.track_inputs['calc_inflation_price_index'][item]['row'] = row_index
                self.track_inputs['calc_inflation_price_index'][item]['unit'] = _unit
                self.track_inputs['calc_inflation_price_index'][item]['col'] = 6#redundant
                self.track_inputs['calc_inflation_price_index'][item]['value'] = None #self.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['domestic_price_index','us_price_index','relative_price_index'] :
                    row_index +=1 
                    if item =='domestic_price_index':
                        self._add_sub_heading(w_sheet,'Foreign',row_index)
                        row_index +=1

                    if item =='relative_price_index':
                        self._add_sub_heading(w_sheet,'Exchange Rate',row_index)
                        row_index +=1    

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_inflation_price_indices(w_sheet)
       
        return row_index 

    def _populate_cal_inflation_price_indices(self, w_sheet):
         
            
        self._transfer_value_tocellrange(w_sheet, 'sens', 
                        'calc_inflation_price_index', 'domestic_inflation_rate',  "Sens", None,
                        '0%')
                     
        #------Domestic Price Index-------------------------------------------
        # price index year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_inflation_price_index','header')        
        #header_cell------------------------
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # current dom inflation
        domestic_inflation_rate_r_index, c_index, found_state_domestic_inflation= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_inflation_rate')
        #domestic_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    domestic_inflation_rate_cell = cell.column + '$' \
                        + str(domestic_inflation_rate_r_index) if found_state_domestic_inflation else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_price_index_cell = prev_letter + str(cell.row) 
                    #=IF(I7=Inputs!$F$32;1;H11*(1+I10))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',1,' \
                                                               + prev_price_index_cell + '*(1+'+ domestic_inflation_rate_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
        
        #------US Inflation-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('sens','us_inflation_rate')
        us_inflation_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        
        us_inflation_rate_r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','us_inflation_rate')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(us_inflation_rate_r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(us_inflation_rate_r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=Sens!' + us_inflation_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
        
        
        #------US Price Index-------------------------------------------
        # price index year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_inflation_price_index','header')        
        #header_cell------------------------
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # current dom inflation
        us_inflation_rate_r_index, c_index, found_state_us_inflation= self._retrieve_cell_row_colm('calc_inflation_price_index','us_inflation_rate')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','us_price_index')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    us_inflation_rate_cell = cell.column + '$' \
                        + str(us_inflation_rate_r_index) if found_state_us_inflation else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_price_index_cell = prev_letter + str(cell.row) 
                    #=IF(I7=Inputs!$F$32;1;H11*(1+I10))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',1,' \
                                                               + prev_price_index_cell + '*(1+'+ us_inflation_rate_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         
        
        # Relative price index------------------------------
        us_price_index_r_index, c_index, found_state_us_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','us_price_index')
        #---------------
        domestic_price_index_r_index, c_index, found_state_dom_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index') 
        #-------------------        
        relative_price_index_r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','relative_price_index')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(relative_price_index_r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(relative_price_index_r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    us_price_index_cell = cell.column +  str(us_price_index_r_index) if found_state_us_pi else '0'
                    domestic_price_index_cell = cell.column +  str(domestic_price_index_r_index) if found_state_dom_pi else '0'
       
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ domestic_price_index_cell + '/' + us_price_index_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         

        # Real Exchange Rate------------------------------
        r_index, c_index, found_state_us_pi= self._retrieve_cell_row_colm('macroeconomic_parameters','exchange_rate')
        exchange_rate_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state_us_pi else '0'
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','real_exchange_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=Inputs!'+ exchange_rate_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         


        # Nominal Exchange Rate------------------------------
        relative_price_index_r_index, c_index, found_state_rel_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','relative_price_index')
        #---------------
        real_exchange_rate_index, c_index, found_state_real_x_rate= self._retrieve_cell_row_colm('calc_inflation_price_index','real_exchange_rate') 
        #-------------------        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','nominal_exchange_rate')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    relative_price_index_cell = cell.column +  str(relative_price_index_r_index) if found_state_rel_pi else '0'
                    real_eachange_rate_cell = cell.column +  str(real_exchange_rate_index) if found_state_real_x_rate else '0'
       
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ relative_price_index_cell + '*' + real_eachange_rate_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
                              
    def _write_analytics_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        
        analytics_sheet = wb.create_sheet()
        analytics_sheet.title = 'Analytics'
        
      
        # Writing the first row of the csv	
        col = get_column_letter(1)
        analytics_sheet.cell('%s%s'%(col, 1)).value = 'Analytic sheet'
        analytics_sheet['%s%s'%(col, 1)].style= 'Heading1'
    
        span_ = 6 
        total_wsheet_cols = 9 + int(span_)
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(analytics_sheet, row_index,total_wsheet_cols,6)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_analytics_gradient_section(wb, analytics_sheet, row_index,total_wsheet_cols)

       

       
        # hide---------------
        BaseBusinessReport._hide_empty_cells(self,analytics_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':14, 'end': 16, 'dim':2})
        rangelist.append({'start':3, 'end': 13, 'dim':10})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':1, 'dim':2})
        indexlist.append({'index':2, 'dim':40})
        
        #skipp one row
        analytics_sheet.append([])
        if hasattr(self,'npv_distribution'):
            npv_y= list(self.npv_distribution['y'])
            npv_x= list(self.npv_distribution['x'])
            npv_cdf= list(self.npv_distribution['cdf'])
            npv_y.insert(0,'PDF')
            npv_y.insert(0,'')
            npv_x.insert(0,'NPV')
            npv_x.insert(0,'')
            npv_cdf.insert(0,'CDF')
            npv_cdf.insert(0,'')
            data = [
                npv_x,
                npv_y,
                npv_cdf,
            ]      
        
            # write content of each row in 1st and 2nd
            # column of the active sheet respectively .
            for row in data:
                analytics_sheet.append(row)
    
            latest_row= row_index + 2 +1
            last_col = 3 + len(data[0])-3# three empty space subtract
            first_col =3
            heading_row_index=latest_row-1
            #6.-----------------------Draw Chart------------------------------
            self.drawChart(analytics_sheet,latest_row, latest_row,
                first_col, last_col,	latest_row +3, 
                heading_row_index, heading_row_index,
                "PDF","NPV")
            #-------------------------------------------------

        #set dims----
        BaseBusinessReport._set_column_dim(self,analytics_sheet,rangelist,indexlist)
    
    def _add_analytics_gradient_section(self, wb, w_sheet, row_index,total_wsheet_cols):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet.cell('%s%s'%(col, row_index)).value = 'Parameters Gradients' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        BaseBusinessReport._set_thick_bottom_border_range(self, w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
     
        row_index +=1


        data = []
        data.append(["",   'Parameter' ,   'NPV Gradient'  ,    '@npv_0','Direction', '', 'Financial comment Viable Project' ])
        #get  item_trcked, us list
        #rrrrrrrrrrrrrrrrrrrrrrrrrrrrr
        
        para_list=[]
        if hasattr(self, 'sens_grad'):
            df = self.sens_grad
            for i in range(len(df)):
                direction_symbol =""
                direction =0
                pos_neg_symbol =""
                parameter_ =df.iloc[i]['parameter']
                comments =''
                x_npv_0= round(df.iloc[i]['x_npv_0'],2)
                #print(self.parameter_unit_dict)
                if df.iloc[i]['gradient']>0:
                    direction_symbol ="↗"
                    direction =1
                    pos_neg_symbol ="+"
                    comments =f'{parameter_} should be at least {x_npv_0}'
                elif df.iloc[i]['gradient']<0:
                    direction_symbol ="↘"
                    direction =-1
                    pos_neg_symbol ="-"
                    comments =f'{parameter_} should not exceed {x_npv_0}'

                if hasattr(self, 'parameter_unit_dict'):
                    if parameter_ in self.parameter_unit_dict:
                        title =self.parameter_unit_dict[parameter_]['title']
                    else:
                        title =parameter_

                   
                    para_list.append({'name':parameter_, 
                                     'number_format': self.parameter_unit_dict[parameter_]['number_format'],
                                     'direction': direction})
                else:
                    para_list.append({'name':parameter_, 'number_format': 'NUMBER','direction': direction})
                    title =parameter_

                
                data.append(["",  title, df.iloc[i]['gradient'],  
                                   df.iloc[i]['x_npv_0'],direction_symbol, pos_neg_symbol, comments]) 
                
               
        #print(para_list)
        # add column headings. NB. these must be strings
        w_sheet.append(["","","","","",""])
        w_sheet.append(["","","","","",""])
      
        #row_index +=2
        start_row = row_index
        

       
        for row in data:
            w_sheet.append(row)
            row_index +=1
        
        end_row= row_index 
       
       # format table
        first_slice_point = get_column_letter(2) + str(start_row + 1 )# skip heeder and first row 
        second_slice_point = get_column_letter(2 + int(10)) + str(end_row) # D39
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                this_row= min(max(cell.row-(start_row+2),0),len(para_list)-1)
                
                   
                if cell.column in ['B']:

                    #italic
                    if cell.row-(start_row+2)>=0:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                        sz=11.0, color='FF0070C0', scheme='minor')   
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                        sz=11.0, color='FF0070C0', scheme='minor')
                
                elif cell.column in ['D'] and cell.row-(start_row+2)>=0:
                    
                    #print(this_row,len(para_list))
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = self._get_number_formats(para_list[this_row]['number_format'])
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                #exclude C
                else:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
                    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                   
                if cell.row-(start_row+2)>=0 and para_list[this_row]['direction']==1 and cell.column in ['B'] :

                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FFFFCC99', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FFFFCC99', scheme='minor')

                    #.fill = PatternFill(fgColor='FFFFCC99', patternType='solid', fill_type='solid')
                elif cell.row-(start_row+2)>=0 and para_list[this_row]['direction']==-1 and cell.column in ['B']:
                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FFF00000', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FFF00000', scheme='minor')                                                     
                 
                   #.fill = PatternFill(fgColor='70c4f4', patternType='solid', fill_type='solid')
       
                elif cell.row-(start_row+2)>=0 and cell.column in ['B']:
                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FF3F3F3F', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF3F3F3F', scheme='minor')       
                
                   #.fill = PatternFill(fgColor='d0ebfb', patternType='solid', fill_type='solid')
       
        return row_index 

    def _write_sens_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        
        sens_sheet = wb.create_sheet()
        sens_sheet.title = 'Sens'
        
      
        # Writing the first row of the csv	
        col = get_column_letter(1)
        sens_sheet.cell('%s%s'%(col, 1)).value = 'Sensitivity analysis sheet'
        sens_sheet['%s%s'%(col, 1)].style= 'Heading1'
    
        span_ = 6 
        total_wsheet_cols = 9 + int(span_)
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(sens_sheet, row_index,total_wsheet_cols,6)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_sensitivity_section(wb, sens_sheet, row_index,total_wsheet_cols)

        if not 'sens_parameters' in self.track_inputs:
            self.track_inputs['sens_parameters']= {}# insert inner

        if not ('endpoint' in self.track_inputs['sens_parameters']):
            self.track_inputs['sens_parameters']['endpoint'] = {}
  
        if 'endpoint' in self.track_inputs['sens_parameters']:
            self.track_inputs['sens_parameters']['endpoint']['row'] = row_index
        #======================================+++++++++++++++

       
        # hide---------------
        BaseBusinessReport._hide_empty_cells(self,sens_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':14, 'end': 16, 'dim':2})
        rangelist.append({'start':3, 'end': 13, 'dim':10})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':1, 'dim':2})
        indexlist.append({'index':2, 'dim':40})

        #set dims----
        BaseBusinessReport._set_column_dim(self,sens_sheet,rangelist,indexlist)
          
    def _write_input_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        output_sheet = wb.active
        output_sheet.title = 'Outputs'
        output_sheet = wb['Outputs']

        #inputs
        input_sheet = wb.create_sheet()
        input_sheet.title = 'Inputs'

        #Calculations
        calc_sheet = wb.create_sheet()
        calc_sheet.title = 'Calc'

        #Cash Flow
        cash_flow_sheet = wb.create_sheet()
        cash_flow_sheet.title = 'CF'
        
        user='Andrew'
        # Writing the first row of the csv	
        col = get_column_letter(1)
        input_sheet.cell('%s%s'%(col, 1)).value = 'Input Sheet ' #% (user)
        input_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_modelspecification_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Timing Section =========================
        row_index +=2
        row_index = self._add_timingassumptions_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Prices Section =========================
        row_index +=2
        row_index = self._add_prices_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
        #==========Production and Inventory Section =========================
        row_index +=2
        row_index = self._add_prodInventory_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Investment Cost Section =========================
        row_index +=2
        row_index = self._add_investmentCost_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Depreciation Section =========================
        row_index +=2
        row_index = self._add_depreciation_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Costs Section =========================
        row_index +=2
        row_index = self._add_costs_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
      
        #==========WorkingCapital Section =========================
        row_index +=2
        row_index = self._add_workingCapital_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

         #==========Financing Section =========================
        row_index +=2
        row_index = self._add_financing_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Flags Section =========================
        row_index +=2
        row_index = self._add_flags_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Taxes Section =========================
        row_index +=2
        row_index = self._add_taxes_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Macroeconomic Parameters Section =========================
        row_index +=2
        row_index = self._add_macroeconomicParameters_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
         #==========Feedlot Design Parameter Section =========================
        row_index +=2
        row_index = self._add_feedlotDesignParameters_section(input_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

       
        # hide---------------
        BaseBusinessReport._hide_empty_cells(self,input_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        BaseBusinessReport._set_column_dim(self,input_sheet,rangelist,indexlist)
    def _write_calculation_sheet(self, wb):
        #---------------------------Worksheet 3---------------------------
       
        calculation_sheet = wb['Calc']

        # Writing the first row of the csv	
        col = get_column_letter(1)
        calculation_sheet.cell('%s%s'%(col, 1)).value = 'Calculation Sheet ' #% (user)
        calculation_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
       
        #==========Inflation Exchange Rates _Price_indices_section =========================
        row_index +=2
        row_index = self._add_cal_inflation_price_indices_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
      
        #==========Investment( Real) =========================
        row_index +=2
        row_index = self._add_cal_investment_cost_real_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Investment (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_investment_cost_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Production And Sales (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_production_sales_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Purchases (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_purchases_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Working Capital (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_working_capital_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Labour Cost(Nominal) =========================
        row_index +=2
        row_index = self._add_cal_labour_cost_nominal_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Calc Of Unit Production(Nominal) =========================
        row_index +=2
        row_index = self._add_cal_unit_of_production_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        

        #==========Finished Product Inventory Valuation (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_finished_product_inventory_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Loan Schedule (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_loan_schedule_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Residual (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_residual_values_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        #==========Depreciation For Tax Purposes (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_depreciation_for_tax_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        #==========Income Tax Statement (Nominal) =========================
        row_index +=2
        row_index = self._add_cal_income_tax_statement_section(calculation_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
        # hide---------------
        BaseBusinessReport._hide_empty_cells(self,calculation_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        BaseBusinessReport._set_column_dim(self,calculation_sheet,rangelist,indexlist)
    def _write_cashflow_sheet(self, wb):
        #---------------------------Worksheet 3---------------------------
       
        cashflow_sheet = wb['CF']

        # Writing the first row of the csv	
        col = get_column_letter(1)
        cashflow_sheet.cell('%s%s'%(col, 1)).value = 'Cash Flow Sheet ' #% (user)
        cashflow_sheet['%s%s'%(col, 1)].style= 'Heading1'
    

        margin_offset = 4 # num of column left for grid ruling
        start_column= 1 + margin_offset
        col_idx = start_column
        unit_price_col = 2 + margin_offset
        fixed_source_col_index = unit_price_col
        item_col_index = 2 # margin_offset + 1 #not dynamic

        # months + middle-offset + ucost + item + margin offset
        total_wsheet_cols = 12 + 2 + 1 + 1 + margin_offset

        span_ =self._get_span()
        total_wsheet_cols = 9 + int(span_)



        last_col= total_wsheet_cols
        #----set underlying thick bottom border @ row 1, from A(1)---Q(17)

        last_col= total_wsheet_cols
        first_col = total_wsheet_cols -11
        col_first = get_column_letter(first_col)# max [0,first_col]
        col_last = get_column_letter(last_col)

        
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(cashflow_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++
        
        
       
        #==========Cash Flow Nominal =========================
        row_index +=2
        row_index = self._add_cf_nominal_section(cashflow_sheet, row_index,total_wsheet_cols)
        #======================================+++++++++++++++

        
      
        #==========Cash Flow (Real) =========================
        row_index +=2
        row_index = self._add_cf_real_section(cashflow_sheet, row_index,total_wsheet_cols)
        self._update_cashflow_linkedcells(wb)
       
        #======================================+++++++++++++++
        
       
       
        # hide---------------
        BaseBusinessReport._hide_empty_cells(self,cashflow_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':1, 'end': total_wsheet_cols, 'dim':13})
        rangelist.append({'start':1, 'end': margin_offset +1, 'dim':2})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':8, 'dim':2})
        indexlist.append({'index':5, 'dim':40})

        #set dims----
        BaseBusinessReport._set_column_dim(self,cashflow_sheet,rangelist,indexlist)

    
    @classmethod
    def get(cls, fmt, *args, **kwargs):

        #fmt = request.GET.get("format", None)
        if not fmt:
            pass
            # Return HTML page
            #return render(request, cls.template, context)
        elif fmt == "json":
            pass
          
        elif fmt in ("spreadsheetlist", "spreadsheettable", "spreadsheet"):            
            # Return an excel spreadsheet
            output = BytesIO()
            cls._generate_spreadsheet_data(
                output, *args, **kwargs
            )
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content=output.getvalue(),
            )
            title ="bbbbb"
            response[
                "Content-Disposition"
            ] = "attachment; filename*=utf-8''%s.xlsx" % urllib.parse.quote(
                force_str(title)
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt in ("csvlist", "csvtable", "csv"):
            pass

           
            # # Return CSV data to export the data
            # response = StreamingHttpResponse(
            #     content_type="text/csv; charset=%s" % settings.CSV_CHARSET,
            #     streaming_content=cls._generate_csv_data(
            #         request, scenario_list, *args, **kwargs
            #     ),
            # )
            # # Filename parameter is encoded as specified in rfc5987
            # if callable(cls.title):
            #     title = cls.title(request, *args, **kwargs)
            # else:
            #     title = cls.model._meta.verbose_name_plural if cls.model else cls.title
            # response[
            #     "Content-Disposition"
            # ] = "attachment; filename*=utf-8''%s.csv" % urllib.parse.quote(
            #     force_str(title)
            # )
            # response["Cache-Control"] = "no-cache, no-store"
            return response
        
        else:
            raise Http404("Unknown format type")

    @classmethod
    def _generate_spreadsheet_data(
        cls, output, *args, **kwargs
    ):
        # Create a workbook
        wb = Workbook(write_only=True)
        title= "vvvv"
        
        ws = wb.create_sheet(title=force_str(title)[:31])

        # Create a named style for the header row
        headerstyle = NamedStyle(name="headerstyle")
        headerstyle.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        wb.add_named_style(headerstyle)
        readlonlyheaderstyle = NamedStyle(name="readlonlyheaderstyle")
        readlonlyheaderstyle.fill = PatternFill(fill_type="solid", fgColor="d0ebfb")
        wb.add_named_style(readlonlyheaderstyle)

        
        # Write a formatted header row
        header = []
        comment = None

        # Write the spreadsheet
        wb.save(output)
	# 	wkSheet.row_dimensions[idx].hidden =True	


  
	    
    def drawChart(self, wkSheet,
                min_row,max_row,
                min_col,max_col,
                chart_row_index_pos, 
                cat_min_row, cat_max_row, 
                chart_title, x_axis_title):
        #-----------------------Draw Charts
        chart = BarChart()

        data = Reference(worksheet=wkSheet,
                        min_row=min_row ,
                        max_row=max_row,
                        min_col=min_col-1,
                        max_col=max_col,
                        )

        chart.add_data(data, from_rows=True, titles_from_data=True)
        cats = Reference(worksheet=wkSheet,
                        min_row=cat_min_row,
                        max_row=cat_max_row,
                        min_col=min_col,
                        max_col=max_col,
                        )
        chart.set_categories(cats)
        col_first = get_column_letter(min_col)# max [0,first_col]
        chart_draw_cell = col_first + str(chart_row_index_pos)# "F3"
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = chart_title

        # You can play with this by choosing any number between 1 and 48
        chart.style = 24
        wkSheet.add_chart(chart, chart_draw_cell)

        """ 
        # create data for plotting
        from openpyxl.chart import BubbleChart, Reference, Series
        xvalues = Reference(wkSheet, min_col = 1, min_row = 2, max_row = 5)
        yvalues = Reference(wkSheet, min_col = 2, min_row = 2, max_row = 5)
        size = Reference(wkSheet, min_col = 3, min_row = 2, max_row = 5)
        # create a 1st series of data
        series = Series(values = yvalues, xvalues = xvalues, zvalues = size, title ="2013")
        # add series data to the chart object
        chart.series.append(series)
        # add chart to the sheet the top-left corner of a chart
        # is anchored to cell E2
        sheet.add_chart(chart, "E2")   
        """
       

