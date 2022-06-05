from ast import Pass
import copy
from django.utils import translation
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, force_str
from django.utils.html import escape
from django.utils.translation import gettext as _
from django.utils.formats import get_format, number_format
from django.utils.text import capfirst, get_text_list, format_lazy
from io import StringIO, BytesIO


from tempfile import NamedTemporaryFile
from excel_response import ExcelResponse
import numpy as np
from openpyxl import Workbook, load_workbook 
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle,PatternFill
from openpyxl.chart import BarChart, Reference, Series, LineChart, BubbleChart
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator

from openpyxl.utils import cell
import math
#from investments_appraisal.base_report import BaseBusinessReport



from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
import datetime
from datetime import date

from investments_appraisal.whatif import get_para_data_table_sensitivity, get_data_table_sensitivity_in_parrallel, get_data_table_sensitivity_in_single_file, get_sensitivity_gradient

#from investments_appraisal.views.usermodels import get_data_table_sensitivity

""" 
protected members of a class can be accessed by other members within the class and are also available to their subclasses. 
No other entity can access these members. In order to do so, they can inherit the parent class. 
Python has a unique convention to make a member protected: Add a prefix _ (single underscore)
"""
	
class ExcelReport():
    
    description_font = Font(name='Calibri',bold=False,  scheme='minor', sz=11.0)
    description_font2 = Font(name='Calibri',bold=False, color="FF3F3F76", scheme='minor', sz=12.0)
    
 
     #Flags
    flags = {
        'YEAR': {'title': 'Years','units': 'Year'}, 
        'PS': {'title': 'Project start','units': 'flag'},
        'CP': {'title': 'Construction period flag','units':'flag'}, 
        'LPP': {'title': 'Loan principle repayment','units': 'flag'},
        'OP': {'title': 'Operating period','units': 'flag'},
        'RES': {'title': 'Residuals','units': 'flag'}, 
            
    }
    
 
    def update(self, param_dict):
        """
        Update parameter values
        """
        for key in param_dict:
            setattr(self, key, param_dict[key])
        
    def __init__(self,modelx=None):
        self.sister_model =modelx

    def _get_number_formats(self, x):
        #print(isinstance(x,str),type(x)==str )
        if type(x)==str:
            if x.upper() =='PERCENT':
                return '0%'
            elif x.upper()=='NUMBER':
                return '_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)' 
            elif x.upper()=='SMALL_NUMBER':
                return '_(* #,##0.0##_);_(* \(#,##0.0##\);_(* "-"??_);_(@_)'
            else:
                return 'General' 
        else:
            return 'General'              
  
      
    def _initialise_named_styles(self, wb):
        # Load a specific sheet by name
        #worksheet = workbook.sheet_by_name('Sheet1')

        # Load a specific sheet by index
        #worksheet = workbook.sheet_by_index(0)

        #ws = wb.worksheets[0]
        heading1 = NamedStyle(name="Heading1")
        heading1.font = Font(name='Cambria',bold=True, color='FF3F3F76', scheme='major', sz=15.0)	
        heading1.alignment = Alignment(horizontal='left', vertical='top',)
        heading1.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading1)
        
        
        heading2 = NamedStyle(name="Heading2")
        heading2.font = Font(name='Calibri',bold=True,  scheme='minor', sz=15.0)
        heading2.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading2.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading2)

        heading2_5 = NamedStyle(name="Heading2_5")
        heading2_5.font = Font(name='Calibri',bold=True,  scheme='minor', sz=11.0)
        heading2_5.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading2_5.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading2_5)
        
        heading3 = NamedStyle(name="Heading3")
        heading3.font = Font(name='Calibri',b=False,  scheme='minor', sz=11.0)
        heading3.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading3.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading3)
        
        heading4 = NamedStyle(name="Heading4")
        heading4.font = Font(name='Calibri',b=False,  scheme='minor', sz=10.0)
        heading4.alignment = Alignment(horizontal='left', vertical='top',)
        #heading1.fill = PatternFill(fgColor='00000000',  fill_type='solid')
        heading4.number_format ='General'
        #---add named_style:
        wb.add_named_style(heading4)

        thick_border_side = Side(border_style="thick")
        straight_thick_border = Border(
                                bottom=thick_border_side)

        thick = NamedStyle(name="Thick_Bottom_Border")
        thick.border = straight_thick_border


        #---add named_style:
        wb.add_named_style(thick)
        

        double_border_side = Side(border_style="double")
        single_border_side = Side(border_style="thin",color="FF7F7F7F")
        square_border = Border(top=double_border_side,
                                right=double_border_side,
                                bottom=double_border_side,
                                left=double_border_side)

        input_border = Border(outline=True,top=single_border_side,
                                right=single_border_side,
                                bottom=single_border_side,
                                left=single_border_side)
        
        input = NamedStyle(name="Input")
        input.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F76')
        input.border = input_border
        input.alignment = Alignment(vertical="top")
        input.fill = PatternFill(fgColor='FFFFCC99', patternType='solid', fill_type='solid')
        input.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        
        #---add named_style:
        wb.add_named_style(input)
        #patternType must be one of {'darkDown', 'lightDown', 'lightGrid', 'lightHorizontal', 'darkHorizontal', 'darkTrellis', 'solid', 'lightGray', 'lightUp', 'darkGray', 'mediumGray', 'lightVertical', 'lightTrellis', 'gray0625', 'darkUp', 'darkGrid', 'darkVertical', 'gray125'}

        calculation = NamedStyle(name="Calculation")
        calculation.font = Font(name='Calibri',bold=True, color='FFFA7D00',sz=12.0)
        calculation.border = input_border
        calculation.alignment = Alignment(vertical="top")
        calculation.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        calculation.number_format ='_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(calculation)

        

        single_border_side_ouput = Side(border_style="thin",color="FF3F3F3F")
        output_border = Border(outline=True,top=single_border_side_ouput,
                                    right=single_border_side_ouput,
                                    bottom=single_border_side_ouput,
                                    left=single_border_side_ouput)

        output = NamedStyle(name="Output")
        output.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output.border = output_border
        output.alignment = Alignment(vertical="top")
        output.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output)

        output1 = NamedStyle(name="Output1")
        output1.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output1.border = output_border
        output1.alignment = Alignment(vertical="top")
        output1.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output1.number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output1)

        output2 = NamedStyle(name="Output2")
        output2.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        output2.border = output_border
        output2.alignment = Alignment(vertical="top")
        output2.fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')
        output2.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output2)

        output3 = NamedStyle(name="Output3")
        output3.font = Font(name='Calibri',bold=True, sz=12.0, color='00339966', scheme='minor')
        #output2.border = output_border
        output3.alignment = Alignment(vertical="top")
        output3.fill = PatternFill(fgColor='00CCCCFF', patternType='solid', fill_type='solid')
        output3.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output3)

        output4 = NamedStyle(name="Output4")
        output4.font = Font(name='Calibri',bold=True, sz=12.0, color='00000000', scheme='minor')
        #output2.border = output_border
        output4.alignment = Alignment(vertical="top")
        output4.fill = PatternFill(fgColor='0099CCFF', patternType='solid', fill_type='solid')
        output4.number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'

        #---add named_style:
        wb.add_named_style(output4)



        percent25 = NamedStyle(name="Percent25")
        percent25.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent25.border = output_border
        percent25.alignment = Alignment(vertical="top")
        percent25.fill = PatternFill(fgColor='FFFF0000', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent25)

        percent50 = NamedStyle(name="Percent50")
        percent50.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent50.border = output_border
        percent50.alignment = Alignment(vertical="top")
        percent50.fill = PatternFill(fgColor='FF0070C0', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent50)


        percent75 = NamedStyle(name="Percent75")
        percent75.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent75.border = output_border
        percent75.alignment = Alignment(vertical="top")
        percent75.fill = PatternFill(fgColor='FFFFFF00', patternType='solid', fill_type='solid')
        #---add named_style:
        wb.add_named_style(percent75)


        percent100 = NamedStyle(name="Percent100")
        percent100.font = Font(name='Calibri',bold=True, sz=12.0, color='FF3F3F3F', scheme='minor')
        percent100.border = output_border
        percent100.alignment = Alignment(vertical="top")
        percent100.fill = PatternFill(fgColor='FF00B050', patternType='solid', fill_type='solid')
        #percent100.number_format ='0%'
        #---add named_style:
        wb.add_named_style(percent100)


        BlueItalicText = NamedStyle(name="Blue_Italic_Text")
        BlueItalicText.font = Font(name='Calibri',bold=False, italic=True, sz=12.0, color='FF0070C0', scheme='minor')
        BlueItalicText.alignment = Alignment(vertical="top")
        #percent100.number_format ='0%'
        #---add named_style:
        wb.add_named_style(BlueItalicText)


        double_border_side_linked = Side(border_style="double",color="FFFF8001")
        
        linked_border = Border(outline=True,
                                bottom=double_border_side_linked)
        
        linkedcell = NamedStyle(name="Linkedcell")
        linkedcell.font = Font(name='Calibri', bold=True, color='FFFA7D00',scheme='minor',sz=12.0)
        linkedcell.border = linked_border
        linkedcell.alignment = Alignment(vertical="top")
        linkedcell.fill = PatternFill(fgColor='00000000', bgColor='00000000')
        linkedcell.number_format ='#,##0_);\(#,##0\);"-  ";" "@'
        
        #---add named_style:
        wb.add_named_style(linkedcell)

        """  
        linked_percentcell = NamedStyle(name="Linked_percentcell")
        linked_percentcell.font = Font(name='Calibri', bold=True, color='FFFA7D00',scheme='minor',sz=12.0)
        linked_percentcell.border = linked_border
        linked_percentcell.alignment = Alignment(vertical="top")
        linked_percentcell.fill = PatternFill(fgColor='00000000', bgColor='00000000')
        linked_percentcell.number_format ='0%'
        
        #---add named_style:
        wb.add_named_style(linked_percentcell) 
        """
 
    

    def _set_thick_bottom_border_range(self, wkSheet, writing_row_index,
                            start_col_index, end_col_index):

        latest_row = writing_row_index

        col_start = get_column_letter(start_col_index)
        col_end = get_column_letter(end_col_index)

        first_slice_point = col_start + str(writing_row_index ) # D200
        second_slice_point = col_end + str(writing_row_index ) # F200

        thick_border_side = Side(border_style="thick", color='FF3F3F76')
        straight_thick_border = Border(
                                bottom=thick_border_side)

        #loop each cell of this row D20:Z20
        #-------->
        for cellObj in wkSheet[first_slice_point:second_slice_point]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
                # |
                # |
                # |
                # v
                cell.border = straight_thick_border

                
                latest_row = cell.row

        #------------------------------
        return latest_row

    
    def _hide_empty_cells(self, wkSheet,extra_cols=None):
        max_column =wkSheet.max_column

        if extra_cols != None:
            max_column +=extra_cols
        max_row =wkSheet.max_row

        last_column = cell.column_index_from_string('XFD')+1
        for idx in range(max_column+1, last_column):
            wkSheet.column_dimensions[get_column_letter(idx)].hidden =True
    
    def _set_column_dim(self,wkSheet, rangelist, indexlist):
        
        for i in rangelist:
            for idx in range(i['start'], i['end']):
                wkSheet.column_dimensions[get_column_letter(idx)].width=i['dim']
        for idj in indexlist:
            wkSheet.column_dimensions[get_column_letter(idj['index'] )].width=idj['dim']
    
    def _populate_output_sheet(self, wksheet):
       self._populate_output_financial(wksheet)
       self._populate_output_fundings(wksheet)
       self._populate_output_general(wksheet)
    
   
  
    def _populate_output_financial(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        percent_format ='0%'

        itemlist= ['minimum_ADSCR','maximum_ADSCR','average_ADSCR',  'minimum_LLCR','maximum_LLCR','average_LLCR']
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_nominal','output_financial',i,number_format, 'CF',None,'Output2')

        itemlist= ['NPV', 'IRR', 'MIRR' ]
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_real','output_financial',i,number_format, 'CF',None,'Output2')
        
        itemlist= ['IRR', 'MIRR' ]
        for i in itemlist:
            self._transfer_cell(wksheet,'cf_real','output_financial',i,percent_format, 'CF',None,'Output2')
              
          
    def _populate_output_fundings(self,wksheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

        itemlist= ['total_investment_cost','senior_debt','equity' ]
       
        #----------
        #set Cell_format= None to retian blue backcolot
        self._transfer_cell(wksheet, 'financing','output_funding', 'senior_debt', '0%', 
                               'Inputs', 'header', None, None ,6-9)
        #70% Senior Debt                       
        self._write_text_atcell(wksheet,'output_funding','header',"Senior Debt",7-9) 
        #1. total_investment_cost  
        self._transfer_range_formulae_tocell(wksheet, 'calc_investment_cost_nominal', 'output_funding',  'total_investment_cost' ,
               number_format , "Calc", None,'Output2' ) 

        #2. total_investment_cost  
        self._transfer_range_formulae_tocell(wksheet, 'calc_investment_cost_nominal', 'output_funding',  'senior_debt_towards_investment' ,
               number_format, "Calc", 'senior_debt','Output2' ) 


        #3. equity
        #=A-B
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': 'A', 'var_type': 'variable','header': 'output_funding', 'para': 'total_investment_cost'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '',},
                {'value': "B", 'var_type': 'variable', 'header': 'output_funding', 'para': 'senior_debt'},
              
            ]
        self._cell_formalue_fromstring(wksheet, formalue_string,'output_funding', 'equity',None,'Output2')
        


  
    def _add_output_funding(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_funding' in self.track_inputs:
            self.track_inputs['output_funding']= {}# insert inner

        if not ('header' in self.track_inputs['output_funding']):
            self.track_inputs['output_funding']['header'] = {}
  
        if 'header' in self.track_inputs['output_funding']:
            self.track_inputs['output_funding']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= ['total_investment_cost','senior_debt','equity' ]        
        ic_parameters = {
            'total_investment_cost': {'title': "Total Investment Cost",'value': None, 'units': 'T_LC'},
            'senior_debt': {'title': "Senior Debt",'value': None, 'units': 'T_LC'},
            'equity': {'title': " Equity",'value': None, 'units': 'T_LC'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()

                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                a=self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                #print(a)                   
                if not (item in self.track_inputs['output_funding']):
                    self.track_inputs['output_funding'][item] = {}

                self.track_inputs['output_funding'][item]['row'] = row_index
                self.track_inputs['output_funding'][item]['unit'] = _unit
                self.track_inputs['output_funding'][item]['col'] = 6
                self.track_inputs['output_funding'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

        return row_index         
          
    def _add_output_financial(self, output_sheet, row_index):
        #-----------------------------------
        #===================================
        
        if not 'output_financial' in self.track_inputs:
            self.track_inputs['output_financial']= {}# insert inner

        if not ('header' in self.track_inputs['output_financial']):
            self.track_inputs['output_financial']['header'] = {}
  
        if 'header' in self.track_inputs['output_financial']:
            self.track_inputs['output_financial']['header']['row'] = row_index
        
        row_index+=2
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)       
        
        itemlist= [
            'minimum_ADSCR','maximum_ADSCR','average_ADSCR',
            'minimum_LLCR','maximum_LLCR','average_LLCR',
            'NPV', 'IRR', 'MIRR'
            ]

        self._add_sub_heading(output_sheet,"A. Lender's Perspective",row_index)
        row_index +=1
        self._add_sub_heading2(output_sheet,"Debt Service Coverage Ratios",row_index)
        row_index +=1
        self._add_sub_heading3(output_sheet,"A.1. Annual Debt Service Coverage Ratio - ADSCR",row_index)
        row_index +=1
        ic_parameters = {
            # A. Lender's Perspective			
            #     Debt Service Coverage Ratios		
            #         A.1. Annual Debt Service Coverage Ratio - ADSCR	

            'minimum_ADSCR': {'title': "Minimum ADSCR",'value': None, 'units': 'NUMBER'},
            'maximum_ADSCR': {'title': "Maximum ADSCR",'value': None, 'units': 'NUMBER'},
            'average_ADSCR': {'title': "Average ADSCR",'value': None, 'units': 'NUMBER'},

             #        A.2. Loan Life Coverage Ratio - LLCR
            'minimum_LLCR': {'title': "Minimum LLCR",'value': None, 'units': 'NUMBER'},
            'maximum_LLCR': {'title': "Maximum LLCR",'value': None, 'units': 'NUMBER'},
            'average_LLCR': {'title': "Average LLCR",'value': None, 'units': 'NUMBER'},

            # B. Owner Perspective			
	        #       Financial Viability of the Project
            'NPV': {'title': "NPV",'value': None, 'units': 'T_LC'},
            'IRR': {'title': 'IRR','value': None, 'units': 'PERCENT'},
            'MIRR': {'title': "MIRR",'value': None, 'units': 'PERCENT'},
        }

        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
               
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                a= self._write_row_title_and_value4(output_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                                   
                if not (item in self.track_inputs['output_financial']):
                    self.track_inputs['output_financial'][item] = {}

                self.track_inputs['output_financial'][item]['row'] = row_index
                self.track_inputs['output_financial'][item]['unit'] = _unit
                self.track_inputs['output_financial'][item]['col'] = 6
                self.track_inputs['output_financial'][item]['value'] = None 
                
                #go to newline
                row_index +=1 

                if item in ['average_ADSCR','average_LLCR'] :
                    row_index +=1
                    if item=='average_ADSCR':
                         self._add_sub_heading3(output_sheet,"A.2. Loan Life Coverage Ratio - LLCR",row_index)
                         row_index +=1
                    if item=='average_LLCR':
                        
                         self._add_sub_heading(output_sheet,"B. Owner Perspective",row_index)
                         row_index +=1 
                         self._add_sub_heading2(output_sheet,"Financial Viability of the Project",row_index)
                         row_index +=1  

        return row_index         
    def _output_bordered_section(self, w_sheet, col_letter_start, col_letter_end, 
                start_row, end_row, title_, heading_style='Heading1', accent_style='Accent1'):

        w_sheet['%s%s'%(col_letter_start, start_row)].value = title_
        w_sheet['%s%s'%(col_letter_start, start_row)].style = heading_style
        w_sheet[col_letter_start + str(start_row)].alignment = Alignment(wrap_text=False,vertical='top')

        col_s =column_index_from_string(col_letter_start)
        col_e= column_index_from_string(col_letter_end)
        #
        for i in range(col_s, col_e+1):
            coli = get_column_letter(i)
            #first row color:
            w_sheet['%s%s'%(coli, start_row)].style= accent_style

        _range=f"{col_letter_start}{start_row}:{col_letter_end}{end_row}"
        self._set_border(w_sheet,_range)
       
    def _add_legend_section(self, w_sheet, row_index,total_wsheet_cols, unit_start_col=7):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Legend' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        #===================================
        
        row_index +=2
        col = get_column_letter(unit_start_col)    
        w_sheet['%s%s'%(col, row_index)].value = 'Unit'
        w_sheet['%s%s'%(col, row_index)].style= 'Explanatory Text'

        col = get_column_letter(unit_start_col+2)    
        w_sheet['%s%s'%(col, row_index)].value = 'Input'
        w_sheet['%s%s'%(col, row_index)].style= 'Input'

        col = get_column_letter(unit_start_col+3)
        w_sheet['%s%s'%(col, row_index)].value = 'Calculation' 
        w_sheet['%s%s'%(col, row_index)].style= 'Calculation'

        col = get_column_letter(unit_start_col+4)
        w_sheet['%s%s'%(col, row_index)].value = 'Output'
        w_sheet['%s%s'%(col, row_index)].style= 'Output'

        col = get_column_letter(unit_start_col+5)
        w_sheet['%s%s'%(col, row_index)].value = 'Linked Cell'
        w_sheet['%s%s'%(col, row_index)].style= 'Linkedcell'

        return row_index 
    
    def _add_modelspecification_section(self,modelx, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Model specification' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        row_index +=1
        self.modelspec_cell_ref ={}
        
        if not ('model_specs' in self.track_inputs):
            self.track_inputs['model_specs']= {} 
        #print('################################################3')    
        for model_item in self.sister_model.model_specifications:
            #print('model_item-->', model_item)

            w_sheet['%s%s'%(colHeader, row_index)].value = model_item['title']
            w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font

            w_sheet['%s%s'%(colValue, row_index)].value = model_item['units']
            w_sheet['%s%s'%(colValue, row_index)].style= 'Input'
            w_sheet['%s%s'%(colValue, row_index)].font = self.description_font2
            
            #keep record      
            self.modelspec_cell_ref[model_item['name']]= '=Inputs!$%s$%s' %(colValue,row_index)
            model_item_name  =model_item['name']
            # 7 lines not working yet
            if not (model_item_name in self.track_inputs['model_specs']):
                self.track_inputs['model_specs'][model_item_name]= {} 

            self.track_inputs['model_specs'][model_item_name]['row'] = row_index
            self.track_inputs['model_specs'][model_item_name]['col'] = 6
            self.track_inputs['model_specs'][model_item_name]['unit'] = model_item['units']
            self.track_inputs['model_specs'][model_item_name]['value'] =  model_item['title']

            row_index +=1

   
        return row_index 
    def _write_row_title_and_value(self, w_sheet, write_colHeader, write_colValue, 
        row_index, header_title,value, val_number_format= 'General'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet['%s%s'%(write_colHeader, row_index)].value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font

        w_sheet['%s%s'%(write_colValue, row_index)].value = value
        w_sheet['%s%s'%(write_colValue, row_index)].style= 'Input'
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2

    def _add_timingassumptions_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Timing Assumptions' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)# E
        colValue = get_column_letter(6) # F
        row_index +=1
       
        #********************************Base Period***************************************
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['base_period']['title'],self.sister_model.timing_assumptions['base_period']['value'])     
       

        self.track_inputs['timing']['base_period']['row'] = row_index
        self.track_inputs['timing']['base_period']['col'] = 6
        self.track_inputs['timing']['base_period']['unit'] = 'YEAR'
        self.track_inputs['timing']['base_period']['value'] = self.sister_model.timing_assumptions['base_period']['value']
        row_index +=2

        
       

        #******************************Construction Period********************************
        #----Start Year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['construction_start_year']['title'],self.sister_model.timing_assumptions['construction_start_year']['value'])                   
        row_index +=1
        
        #---Length
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['construction_len']['title'],self.sister_model.timing_assumptions['construction_len']['value'])
        row_index +=1

        #---End
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['construction_year_end']['title'],self.sister_model.timing_assumptions['construction_year_end']['value'])
        self.track_inputs['timing']['construction_year_end']['row']= row_index 
        self.track_inputs['timing']['construction_year_end']['unit'] = 'YEAR'
        self.track_inputs['timing']['construction_year_end']['col'] = 6
        self.track_inputs['timing']['construction_year_end']['value'] = self.sister_model.timing_assumptions['construction_year_end']['value']         
        row_index +=2

        

        #******************************Operation Period***********************************************
        #----Start Year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['operation_start_year']['title'],self.sister_model.timing_assumptions['operation_start_year']['value'])        
        self.track_inputs['timing']['operation_start_year']['row'] = row_index 
        self.track_inputs['timing']['operation_start_year']['unit'] = 'YEAR'
        self.track_inputs['timing']['operation_start_year']['col'] = 6
        self.track_inputs['timing']['operation_start_year']['value'] = self.sister_model.timing_assumptions['operation_start_year']['value'] 
                   
        row_index +=1
        #---duration
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['operation_duration']['title'],self.sister_model.timing_assumptions['operation_duration']['value'])
        row_index +=1

        #---End
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['operation_end']['title'],self.sister_model.timing_assumptions['operation_end']['value'])
        self.track_inputs['timing']['operation_end']['row'] = row_index 
        self.track_inputs['timing']['operation_end']['unit'] = 'YEAR'
        self.track_inputs['timing']['operation_end']['col'] = 6
        self.track_inputs['timing']['operation_end']['value'] = self.sister_model.timing_assumptions['operation_end']['value'] 
                        
        row_index +=2

        #---Number of months in a year
        self._write_row_title_and_value(w_sheet, colHeader, colValue, row_index, 
                  self.sister_model.timing_assumptions['number_of_months_in_a_year']['title'],self.sister_model.timing_assumptions['number_of_months_in_a_year']['value'])

        self.track_inputs['timing']['number_of_months_in_a_year']['row'] = row_index 
        self.track_inputs['timing']['number_of_months_in_a_year']['unit'] = 'YEAR'
        self.track_inputs['timing']['number_of_months_in_a_year']['col'] = 6
        self.track_inputs['timing']['number_of_months_in_a_year']['value'] = self.sister_model.timing_assumptions['number_of_months_in_a_year']['value'] 
             

        return row_index 

    def _add_prices_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Prices' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================

        #eadin 2
        col = get_column_letter(2)
        w_sheet['%s%s'%(col, row_index)].value = 'Net Sales Price: %s' % (commodity_title)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1


        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        
        # Base price of beef ROW:
        w_sheet['%s%s'%(colHeader, row_index)].value = 'Base price of %s in %s per ton' % (commodity_title ,self.sister_model.timing_assumptions['base_period']['value'])
        w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
        w_sheet['%s%s'%(colValue, row_index)].value = self.sister_model.prices['base_price']
        w_sheet['%s%s'%(colValue, row_index)].style= 'Input'
        if 'base_price' in self.track_inputs['prices']:
            self.track_inputs['prices']['base_price']['row'] = row_index
            self.track_inputs['prices']['base_price']['col'] = 6
            self.track_inputs['prices']['base_price']['cell'] =  "$" + colValue + "$" + str(row_index)
            self.track_inputs['prices']['base_price']['unit'] = "NUMBER"
            self.track_inputs['prices']['base_price']['value'] = self.sister_model.prices['base_price']
         
       
        w_sheet['%s%s'%(colUnits, row_index)].value  =self.modelspec_cell_ref['LC'] if  'LC' in self.modelspec_cell_ref else ''
        w_sheet['%s%s'%(colUnits, row_index)].style= 'Explanatory Text'  
        row_index +=1

        # Change in Price of Beef ROW:
        w_sheet['%s%s'%(colHeader, row_index)].value = 'Change in Price of %s'  % (commodity_title)
        w_sheet['%s%s'%(colHeader, row_index)].font = self.description_font
        w_sheet['%s%s'%(colValue, row_index)].value = self.sister_model.prices['change_in_price']

        w_sheet['%s%s'%(colValue, row_index)].style= 'Linkedcell'
        w_sheet['%s%s'%(colValue, row_index)].number_format ='0%'

        if 'change_in_price' in self.track_inputs['prices']:
            self.track_inputs['prices']['change_in_price']['row'] = row_index
            self.track_inputs['prices']['change_in_price']['col'] = 6
            self.track_inputs['prices']['change_in_price']['cell'] =  "$" + colValue + "$" + str(row_index)
            self.track_inputs['prices']['change_in_price']['unit'] = "PERCENT"
            self.track_inputs['prices']['change_in_price']['value'] = self.sister_model.prices['change_in_price']
         
        

        w_sheet['%s%s'%(colUnits, row_index)].value = self.modelspec_cell_ref['PERCENT'] if  'PERCENT' in self.modelspec_cell_ref else ''
        w_sheet['%s%s'%(colUnits, row_index)].style= 'Explanatory Text'

        row_index +=1

        return row_index 

    def _get_span(self):
        start=None
        end= None
        if 'base_period' in self.sister_model.timing_assumptions:
            start=self.sister_model.timing_assumptions['base_period']['value'] 
        
        if 'operation_end' in self.sister_model.timing_assumptions:
            end=self.sister_model.timing_assumptions['operation_end']['value'] 

        if start and end:
            span_ = end - start
            #print(f' span_ = end - start: {span_} = {end} - {start}')

        else:
            span_ = self.sister_model.timing_assumptions['operation_duration']['value'] if 'operation_duration' in self.sister_model.timing_assumptions else 0
                
           
        """ 
        e.g
        2022-2032
        10 year operation
        1 year deconstrustion
        span is 10 + 1 = (oparation + 1 deconsctruion year)
        """ 
        if span_> 0:
            return span_ #+1 excel already add plus 1
        else:
            #if no diff return 0
            return span_ 
               
       
         
  
    
    def _populate_investment_parameters_in_excelrow(self, w_sheet, options_arraylist, row_index, item):
   
        total_len = len(options_arraylist)-1
        first_slice_point = get_column_letter(9) + str(row_index)# D09
        second_slice_point = get_column_letter(9 + int(total_len)) + str(row_index) # D16
        #loop each cell of this row D09:D16
        #----------->
        i=0
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                w_sheet['%s%s'%(cell.column, cell.row)] = options_arraylist[i] 
                w_sheet['%s%s'%(cell.column, cell.row)].style = 'Input'
                if item in ['senior_debt_dynamic_parameter']:
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='0.0%'
                else:    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
                i += 1

                       
           
  
    def _add_depreciation_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Depreciation' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['economic_life_of_machinery','economic_life_of_buildings',]
       
        #==========Sub-heading=========================
        # col = get_column_letter(2)
        # w_sheet['%s%s'%(col, row_index)].value = 'Economic service life' 
        # w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        self._add_sub_heading(w_sheet,'Economic service life',row_index)
        row_index +=1
        #-----------------------------------
        for item in itemlist:
            if item in self.sister_model.depreciation:
                _unit =self.sister_model.depreciation[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.depreciation[item]['title'], self.sister_model.depreciation[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
               
                self.track_inputs['depreciation'][item]['row'] = row_index
                self.track_inputs['depreciation'][item]['unit'] = _unit
                self.track_inputs['depreciation'][item]['col'] = 6#redundant
                self.track_inputs['depreciation'][item]['value'] = self.sister_model.depreciation[item]['value'] 
                row_index +=1  
        
        # Sub-Section B skipp rows
        row_index +=1 
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet['%s%s'%(col, row_index)].value = 'Recovery period for income tax' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #----------------------------------- 
        itemlist =['tax_life_of_machinery','tax_life_of_buildings','tax_life_of_soft_capital_costs']
        for item in itemlist:
            if item in self.sister_model.depreciation:
                _unit =self.sister_model.depreciation[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.depreciation[item]['title'], self.sister_model.depreciation[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['depreciation'][item]['row'] = row_index
                self.track_inputs['depreciation'][item]['unit'] = _unit
                self.track_inputs['depreciation'][item]['col'] = 6# redundant
                self.track_inputs['depreciation'][item]['value'] = self.sister_model.depreciation[item]['value'] 
                row_index +=1  

        return row_index  
    def _add_sub_heading3(self, w_sheet, title, row_index):
        col = get_column_letter(4)
        w_sheet['%s%s'%(col, row_index)].value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading4'
      
    def _add_sub_heading2(self, w_sheet, title, row_index):
        col = get_column_letter(3)
        w_sheet['%s%s'%(col, row_index)].value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading3'
      
    def _add_sub_heading(self, w_sheet, title, row_index):
        col = get_column_letter(2)
        w_sheet['%s%s'%(col, row_index)].value = title 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        
    def _add_workingCapital_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Working Capital' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

       

        itemlist =['accounts_receivable','accounts_payable','cash_balance']
        
        for item in itemlist:
            if item in self.sister_model.working_capital:
                _unit =self.sister_model.working_capital[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.working_capital[item]['title'], self.sister_model.working_capital[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
               
                self.track_inputs['working_capital'][item]['row'] = row_index
                self.track_inputs['working_capital'][item]['unit'] =_unit
                self.track_inputs['working_capital'][item]['col'] =6# redundant
                self.track_inputs['working_capital'][item]['value'] = self.sister_model.working_capital[item]['value'] 
                row_index +=1

        return row_index  
    
    def _write_row_title_and_value3(self, 
                                    w_sheet, write_colHeader, write_colValue, 
                                    write_colUnits,row_index, header_title , 
                                    input_value, ref_formular_unit, val_number_format= 'General', 
                                    blank_UNITS =False,  valfield_style='Explanatory Text'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet['%s%s'%(write_colHeader, row_index)].value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font
        # if None skip
        if input_value != None:
            w_sheet['%s%s'%(write_colValue, row_index)].value = input_value
            w_sheet['%s%s'%(write_colValue, row_index)].style= 'Input'
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
       
        if blank_UNITS==False:
            w_sheet['%s%s'%(write_colUnits, row_index)].value = ref_formular_unit
            w_sheet['%s%s'%(write_colUnits, row_index)].style= valfield_style
        
    def _write_row_title_and_value4(self, 
                                    w_sheet, write_colHeader, write_colValue, 
                                    write_colUnits,row_index, header_title , 
                                    input_value, ref_formular_unit, val_number_format= 'General', 
                                    blank_UNITS = False,  
                                    valfield_style='Explanatory Text'):
        
      
        #----Tis function write eader and value as excel row wit formatin
        w_sheet['%s%s'%(write_colHeader, row_index)].value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font
        # if None skip
        if input_value != None:
            w_sheet['%s%s'%(write_colValue, row_index)].value = input_value
        w_sheet['%s%s'%(write_colValue, row_index)].style= 'Output2'
        
        w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
       
        if blank_UNITS == False:
            w_sheet['%s%s'%(write_colUnits, row_index)].value = ref_formular_unit
            w_sheet['%s%s'%(write_colUnits, row_index)].style= valfield_style
       
        return [w_sheet,write_colHeader, write_colValue, write_colUnits,row_index, header_title ,  
               input_value, ref_formular_unit, val_number_format,blank_UNITS]
    
       
    def _add_financing_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Financing' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

       

        itemlist =['real_interest_rate','risk_premium','num_of_installments','grace_period','repayment_starts']
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet['%s%s'%(col, row_index)].value = 'Suppliers credit' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #-----------------------
        for item in itemlist:
            if item in self.sister_model.financing:
                _unit =self.sister_model.financing[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.financing[item]['title'], self.sister_model.financing[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['financing'][item]['row'] = row_index
                self.track_inputs['financing'][item]['unit'] = _unit
                self.track_inputs['financing'][item]['col'] = 6#redundant
                self.track_inputs['financing'][item]['value'] = self.sister_model.financing[item]['value'] 
                row_index +=1 
        row_index +=1
        
        #==========Sub-heading=========================
        col = get_column_letter(2)
        w_sheet['%s%s'%(col, row_index)].value = 'Financing as a  % of Investment Costs' 
        w_sheet['%s%s'%(col, row_index)].style= 'Heading2'
        row_index +=1
        #-----------------------
        for item in ['equity','senior_debt',]:
            if item in self.sister_model.financing:
                _unit =self.sister_model.financing[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                            self.sister_model.financing[item]['title'], self.sister_model.financing[item]['value'], cell_display_val, 
                            self._get_number_formats(_unit), _unit=='BLANK')

                self.track_inputs['financing'][item]['row'] = row_index
                self.track_inputs['financing'][item]['unit'] = _unit
                self.track_inputs['financing'][item]['col'] = 6#redundant
                self.track_inputs['financing'][item]['value'] = self.sister_model.financing[item]['value']  

                if item=='equity':
                    w_sheet['%s%s'%(colValue, row_index)].value ='=1-'+ colValue + str(row_index +1)
                    w_sheet['%s%s'%(colValue, row_index)].fill = PatternFill(fgColor='FFF2F2F2', patternType='solid', fill_type='solid')  
                    
                #self._populate_flags_section(w_sheet, row_index, item, cell_ref_timing)
                row_index +=1      
            
       
        row_index +=1

        

        return row_index      
    

     
    def _write_row_title_and_value2(self, w_sheet, write_colHeader, write_colValue, 
        row_index, header_title,value, val_number_format= 'General', valfield_style='Explanatory Text'):

        #----Tis function write eader and value as excel row wit formatin
        w_sheet['%s%s'%(write_colHeader, row_index)].value = header_title
        w_sheet['%s%s'%(write_colHeader, row_index)].font = self.description_font

        w_sheet['%s%s'%(write_colValue, row_index)].value = value
        w_sheet['%s%s'%(write_colValue, row_index)].style= valfield_style
        #w_sheet['%s%s'%(write_colValue, row_index)].number_format= val_number_format
        #w_sheet['%s%s'%(write_colValue, row_index)].font = self.description_font2
    
    def _update_section_header_year(self, w_sheet, cat_section, flags_section_row_index,
                                   start_col_index, span_ ,source_w_sheet = None):
        #The function inserts years copied from Flas Section 
        #Do noting if no history is found
        if cat_section in self.track_inputs:
            if 'header' in self.track_inputs[cat_section]:
                start =start_col_index
                end = start_col_index + int(span_)
                if 'row' in self.track_inputs[cat_section]['header']:
                    header_row_= self.track_inputs[cat_section]['header']['row']
                    first_slice_point = get_column_letter(start) + str(header_row_)# D07
                    second_slice_point = get_column_letter(end) + str(header_row_) # D39
                    #loop each cell of this row D9:D37
                    #----------->
                    appended_worksheet = f'{source_w_sheet}!' if source_w_sheet != None else '' # =Input!AF10
                    for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                        for cell in cellObj:
                            #get column
                            new_column_letter = cell.column # J
                            w_sheet['%s%s'%(new_column_letter, cell.row)] =  '='+ appended_worksheet + '$'+ new_column_letter + '$'+ str(flags_section_row_index) 
                            w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Heading1'
                            w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General' 
                    
                    self._set_thick_bottom_border_range( w_sheet,  header_row_, start, end)
        
    def _populate_flags_section(self, w_sheet, row_index, item, cell_ref_timing):
        if item=='YEAR':
            # set row----
            self.track_inputs['flags']['years']['row']=row_index
            #columns is not set as it is not important
            span_ =self._get_span()
            start_col_index= 9
            first_slice_point = get_column_letter(start_col_index) + str(row_index)# D07
            second_slice_point = get_column_letter(start_col_index + int(span_)) + str(row_index) # D39
            
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                    prev_col=column_index_from_string(new_column_letter)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_cell = prev_letter + str(cell.row)
                   
                   
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF('+ prev_cell + ',' + prev_cell + '+1,' + cell_ref_timing['base_period'] + ')'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
            #
            #self._update_production_inventory_header_year(w_sheet, row_index, start_col_index, span_)
            self._update_section_header_year(w_sheet, 'production_inventory', row_index, start_col_index, span_)
            self._update_section_header_year(w_sheet, 'investment_cost', row_index, start_col_index, span_)
        if item=='PS':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '=' + cell_ref_timing['base_period'] + ',1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='CP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '>' + cell_ref_timing['construction_year_end'] + ',0,1)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='LPP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(OR(' + years_cell + '<' + cell_ref_timing['repayment_starts'] + \
                                                                  ',' + years_cell +'>('+ cell_ref_timing['repayment_starts'] + \
                                                                  '+' + cell_ref_timing['num_of_installments'] + '-1)),0,1)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    #w_sheet['%s%s'%(new_column_letter, cell.row)].number_format = 'General'
		
        if item=='OP':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(AND(' + years_cell + '>=$' + cell_ref_timing['operation_start_year'] + \
                                                                  ',' + years_cell +'<$'+ cell_ref_timing['operation_end']  + '),1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                    
        if item=='RES':
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(row_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(row_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            

            flags_years_row = self.track_inputs['flags']['years']['row']
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    new_column_letter = cell.column # J
                   #
                    years_cell = cell.column + str(flags_years_row)
                    w_sheet['%s%s'%(new_column_letter, cell.row)] =  '=IF(' + years_cell + '=$' + cell_ref_timing['operation_end']  + ',1,0)'
                    w_sheet['%s%s'%(new_column_letter, cell.row)].style = 'Output2'
                  
  


    def _add_flags_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Flags' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(7)
        row_index +=1

        financing = {
            'real_interest_rate': {'title': 'Real interest rate','value': 6, 'units': 'PERCENT'}, 
            'risk_premium': {'title': 'Risk premium','value': 1, 'units': 'PERCENT'}, 
            'num_of_installments': {'title': 'No. of installments','value': 6, 'units': 'NUMBER'},  
            'grace_period': {'title': 'Grace period (years)','value': 0, 'units': 'YEARS'}, 
            'repayment_starts': {'title': 'Repayment starts year','value':2022, 'units': 'YEAR'}, 
            'equity': {'title': 'Equity (% of Investment Costs)','value': 30, 'units': 'PERCENT'}, 
            'senior_debt': {'title': 'Senior Debt (% of Investment Costs)','value': .70, 'units': 'PERCENT'},  
                
        }
       
        # retrieve reference
        base_period_row = self.track_inputs['timing']['base_period']['row']
        base_period_col = self.track_inputs['timing']['base_period']['col']
        itemlist =['YEAR','PS','CP','LPP','OP','RES']

        base_period_ref = get_column_letter(base_period_col) + str(base_period_row) # F32

        construction_year_end_row = self.track_inputs['timing']['construction_year_end']['row']
        construction_year_end_ref = '$' + get_column_letter(base_period_col) + '$' + str(construction_year_end_row) # $F$32
  
        cell_ref_timing={}
        cell_ref_timing= {'base_period':base_period_ref, 'construction_year_end':construction_year_end_ref}
        cell_ref_timing['repayment_starts'] = get_column_letter(base_period_col) + str(self.track_inputs['financing']['repayment_starts']['row'])
        cell_ref_timing['num_of_installments'] =  get_column_letter(base_period_col) + str(self.track_inputs['financing']['num_of_installments']['row']) 
        cell_ref_timing['grace_period'] =  get_column_letter(base_period_col) + str(self.track_inputs['financing']['grace_period']['row'])         
        cell_ref_timing['operation_start_year'] = get_column_letter(base_period_col) + str(self.track_inputs['timing']['operation_start_year']['row'])
        cell_ref_timing['operation_end'] =  get_column_letter(base_period_col) + str(self.track_inputs['timing']['operation_end']['row']) 
      

        
        for item in self.flags.keys():#itemlist:
            if item in self.flags:
                flag_unit =self.flags[item]['units']
                flag_unit = flag_unit.upper()
                cell_display_val = self.modelspec_cell_ref[flag_unit] if  flag_unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value2(w_sheet, colHeader, colValue, row_index, 
                        self.flags[item]['title'],cell_display_val)

                #dynamic declare variables
                if not (item in self.track_inputs['flags']):
                    self.track_inputs['flags'][item] ={}

                self.track_inputs['flags'][item]['row'] = row_index
                self.track_inputs['flags'][item]['unit'] = flag_unit
                self.track_inputs['flags'][item]['col'] = 6
                self.track_inputs['flags'][item]['value'] = None      

                self._populate_flags_section(w_sheet, row_index, item, cell_ref_timing)
                 
                #skipp 2 line after YEAR LINE 
                if item == 'YEAR':
                    #update production & inventory year       
                    row_index +=2
                else:
                   row_index +=1  
       
        return row_index      

     
    def _add_taxes_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Taxes' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['import_duty','sales_tax','corporate_income_tax']
        
        for item in itemlist:
            if item in self.sister_model.taxes:
                _unit =self.sister_model.taxes[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.taxes[item]['title'], self.sister_model.taxes[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                self.track_inputs['taxes'][item]['row'] = row_index
                self.track_inputs['taxes'][item]['unit'] = _unit
                self.track_inputs['taxes'][item]['col'] = 6#redundant
                self.track_inputs['taxes'][item]['value'] = self.sister_model.taxes[item]['value'] 
                row_index +=1 

        return row_index      

    def _add_macroeconomicParameters_section(self, w_sheet, row_index,total_wsheet_cols):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Macroeconomic Parameters' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
       
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1

        itemlist =['discount_rate_equity','domestic_inflation_rate',
                    'us_inflation_rate','exchange_rate','dividend_payout_ratio','num_of_shares']
        
        for item in itemlist:
            if item in self.sister_model.macroeconomic_parameters:
                _unit =self.sister_model.macroeconomic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        self.sister_model.macroeconomic_parameters[item]['title'], self.sister_model.macroeconomic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                self.track_inputs['macroeconomic_parameters'][item]['row'] = row_index
                self.track_inputs['macroeconomic_parameters'][item]['unit'] = _unit
                self.track_inputs['macroeconomic_parameters'][item]['col'] = 6#redundant
                self.track_inputs['macroeconomic_parameters'][item]['value'] = self.sister_model.macroeconomic_parameters[item]['value'] 
                row_index +=1  

        return row_index
    def _set_linkedcell(self, w_sheet, header, subheader, 
                       val, linked_Style='Linkedcell', format_style= "",):
                       
        # retrieve tracked parmameters
        r_index, c_index, found_state= self._retrieve_cell_row_colm(header,subheader)
        #if found change details else leave the default valuew
        if found_state:
            c= get_column_letter(c_index)
            if val != None:
                w_sheet['%s%s'%(c, r_index)].value ='=' +   val
            if  linked_Style =='Linkedcell':
                #linked  
                w_sheet['%s%s'%(c, r_index)].style = linked_Style
            else:
                #calculated
                w_sheet['%s%s'%(c, r_index)].style = 'Calculation'
            
            if format_style != "":
                w_sheet['%s%s'%(c, r_index)].number_format= self._get_number_formats(format_style) 
        
  
    def _hide_rows(self, wkSheet, start, end):
        for idx in range(start, end):
            wkSheet.row_dimensions[idx].hidden =True

    def _retrieve_cell_row_colm(self, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:
                # return true if at least row/col is found
                #error cant return all in some instance especially on flags:year

                if 'row' in self.track_inputs[header_par][sub_header]:
                    _row_index = self.track_inputs[header_par][sub_header]['row']
                    found_state=True
                # else:
                #     found_state=False

                if 'col' in self.track_inputs[header_par][sub_header]:
                    _col_index = self.track_inputs[header_par][sub_header]['col']
                    found_state=True
                # else:
                #     found_state=False

                
               
        return _row_index, _col_index, found_state  

    def _retrieve_cell_value(self, w_sheet, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:
                _row_index = self.track_inputs[header_par][sub_header]['row'] 
                _col_index = self.track_inputs[header_par][sub_header]['col']
                found_state=True
                col_Letter = get_column_letter(_col_index)
                return  w_sheet['%s%s'%(col_Letter, _row_index)].value, found_state 
        return  _,  found_state

    def _retrieve_cell_formuale(self, w_sheet, header_par, sub_header):
        found_state= False
        _row_index = None 
        _col_index = None
        if header_par in self.track_inputs:
            if sub_header in self.track_inputs[header_par]:
                _row_index = self.track_inputs[header_par][sub_header]['row'] 
                _col_index = self.track_inputs[header_par][sub_header]['col']
                found_state=True
                col_Letter = get_column_letter(_col_index)
                #A14
                return  col_Letter + str(_row_index ),found_state
        return  _, found_state     
                
    def _check_item_dict_in_list(self, item_instance, item , exclusion_list):
        for list_item in exclusion_list:
            if item_instance ==list_item['item_instance'] and item == list_item['item']:
                return False# Exclude  
        return True# Procede
  
    def _add_sensitivity_x_section(self, modelx, w_sheet, row_index,total_wsheet_cols, 
                                         header_title,  track_input_item, val_list, index_base_value_border, index_npv_zero_border,
                                         first_col_num_format='General'):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = header_title
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not track_input_item in self.track_inputs:
            self.track_inputs[track_input_item]= {}# insert inner

        if not ('header' in self.track_inputs[track_input_item]):
            self.track_inputs[track_input_item]['header'] = {}
  
        if 'header' in self.track_inputs[track_input_item]:
            self.track_inputs[track_input_item]['header']['row'] = row_index

      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index  +=1 # placeholders
        #===================================
      
        row_index +=1
        
         #aaaaa
        npv = self._get_cell_ref_asformulae('cf_real','NPV',"CF",None)
        irr = self._get_cell_ref_asformulae('cf_real','IRR',"CF",None)
        mirr = self._get_cell_ref_asformulae('cf_real','MIRR',"CF",None)
        
        lb, _, foundstate= modelx._get_loan_principal_repayment_bounds()
        if  not foundstate:
           lb =0

         
        ADSCR_3 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 3-1)
        ADSCR_4 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 4-1)
        ADSCR_5 = self._get_cell_ref_asformulae('cf_nominal','ADSCR',"CF",9-9 + lb + 5-1)
       
        LLCR_2 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 2-1)
        LLCR_3 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 3-1)
        LLCR_4 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 4-1)
        LLCR_5 = self._get_cell_ref_asformulae('cf_nominal','LLCR',"CF",9-9 + lb + 5-1)
       
       
        data = []
        data.append(["","","",  npv, ADSCR_3, ADSCR_4,ADSCR_5,LLCR_2, LLCR_3, LLCR_4, LLCR_5, irr, mirr])
        #get  item_trcked, us list
        #rrrrrrrrrrrrrrrrrrrrrrrrrrrrr
        #convert sens_var ======> var
        sens_var= track_input_item.split("sens_",1)[1].strip()
        #i-way run sensitvity analysis 
        if hasattr(modelx,'sens_dict'):
            if sens_var in modelx.sens_dict.keys():
                df = modelx.sens_dict[sens_var]['df']
                x_npv_0 = modelx.sens_dict[sens_var]['x_npv_0']                
                #print('Retrieving sensitivity...', sens_var)
            else:
                #print('Getting sensitivity...', sens_var)
                df_sens,x_npv_0= get_para_data_table_sensitivity(modelx,sens_var,val_list)
                df= df_sens.copy()
        else:
            #print('Getting sensitivity...', sens_var)
            df_sens, x_npv_0= get_para_data_table_sensitivity(modelx,sens_var,val_list)
            df= df_sens.copy()
        #print('Excell:DF list length:',len(val_list),len(df))
        for i in range(min(len(val_list),len(df))):
            
            data.append(["",        "",        val_list[i],
                        df.iloc[i]['npv'],     df.iloc[i]['adscr-3'],  df.iloc[i]['adscr-4'],
                        df.iloc[i]['adscr-5'], df.iloc[i]['llcr-2'],   df.iloc[i]['llcr-3'],
                        df.iloc[i]['llcr-4'],  df.iloc[i]['llcr-5'],   df.iloc[i]['irr'],
                        df.iloc[i]['mirr']]) 

        # add column headings. NB. these must be strings
        w_sheet.append(["","","","","","",""])
        w_sheet.append(["","","Values", 'NPV', 'ADSCR-3','ADSCR-4','ADSCR-5','LLCR-2','LLCR-3','LLCR-4','LLCR-5','IRR', 'MIRR'])
      
        #row_index +=2
        start_row = row_index
       
        for row in data:
            w_sheet.append(row)
            row_index +=1
        
        end_row= row_index 
       
       # format table
        first_slice_point = get_column_letter(3) + str(start_row + 1 )# skip heeder and first row 
        second_slice_point = get_column_letter(3 + int(10)) + str(end_row) # D39
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                
               # w_sheet['%s%s'%(cell.column, cell.row)] = row[i]
                if cell.column in ['L', 'M']:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
        
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='0.00%'

                    
                elif cell.column in ['C']:
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = first_col_num_format
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
        
                else:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
                    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                   
        
        #---Current Model Value Setborder Outline------------------------
        current_header_row_index = self.track_inputs[track_input_item]['header']['row'] 
        col_letter_start= 'C'
        col_letter_end= 'M'
        #Border of base Row
        if int(index_base_value_border) >0:
            border_start_row = current_header_row_index + index_base_value_border + 3
            border_start_end_row= border_start_row
            _range=f"{col_letter_start}{border_start_row}:{col_letter_end}{border_start_end_row}"
            self._set_border(w_sheet,_range)
        #--------------------------------------
        #print('index_npv_zero_border',index_npv_zero_border, index_base_value_border )
        if int(index_npv_zero_border) >0:
            #Border of NPV ==Zero
            border_start_row = current_header_row_index + index_npv_zero_border + 3
            border_start_end_row= border_start_row
            _range=f"{col_letter_start}{border_start_row}:{col_letter_end}{border_start_end_row}"
            self._set_border_yellow(w_sheet,_range)
        #--------------------------------------

        return row_index
    def _dummy(self,w_sheet, row_index):
        
        tab = Table(displayName="Table1", ref="C" + str(row_index+3) +":G"+ str(row_index+3 +5))
        for cellObj in w_sheet["D" + str(row_index+3+1) +":G"+ str(row_index+3 +5)]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
               
                latest_row = cell.row
                cell_ref = new_column_letter + str(cell.row)
                cell_row_val = "C" + str(cell.row)
                #Top row-------
                cell_orig= new_column_letter + str(19)
                #formulae_ ="=Product(C14,[@value])"
                w_sheet[ cell_ref] = '=TABLE(,C14)'#Translator(f'=Product($C$14,{cell_row_val})', origin=cell_orig).translate_formula(cell_ref)
                w_sheet[ cell_ref] ='{=TABLE(C14;)}'
        #------------------------------

      
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
      
        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unque through out defined names and all other table name. 
        '''
       
        from openpyxl.utils import range_boundaries 
        coord = list(range_boundaries(tab.ref))
        #print("coordinats:",coord)
        data = [["","",200,  '=TABLE(,C14)',  '=TABLE(,C14)',  '=TABLE(,C14)',  '=TABLE(,C14)']]
        for row in data:
            coord[-1] += 1  # <= No need to use max_row, just add one line!
            w_sheet.append(row)

        tab.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"

        #print(tab)
        """Change table range"""
        # ws.tables['MyTable'].ref = "A1:E5"
       
       
       
       
       
        # max_row = 10#sens_sheet.max_row
        # tab = w_sheet._tables["Table1"]
        # tab.ref = f"A1:{w_sheet.max_column}{w_sheet.max_row}"

        # for i in w_sheet.iter_rows(min_row=2, max_row=max_row):
        #     dynamic_cell = i[9]
        #     if dynamic_cell  == 'ice cream':
        #          data = [
        #             ['Apples', 10000, 5000, 8000, 6000],
        #             ['Pears',   2000, 3000, 4000, 5000],
        #             ['Bananas', 6000, 6000, 6500, 6000],
        #             ['Oranges',  500,  300,  200,  700],
        #         ]
        #         # perform some operation
    def _sort_by_another_list(self,test_dict, orderlist):
        ordering_list, first_list=self._get_order_index_as_list(test_dict,orderlist)
        #print('indexlist', indexlist)          
        zipped_lists = zip(ordering_list, first_list) 
        #print('zipped:', zipped_lists[0])
        sorted_zipped_lists = sorted(zipped_lists)
        #sorted_zipped_lists = sorted(test_dict)
         
        test_list = [element for _, element in sorted_zipped_lists]
        new_dict_list =[]
        #[4,5,7,10,1,3,5,2,6,8,9] 
     
        for item in test_list:
            #item =test_list[i]
            item_dict_ = test_dict[item]
            new_dict_list.append(item_dict_)
          
        return  new_dict_list   
    def _get_order_index_as_list(self, test_list_, list_in_desired_order):       
        
        new_list=[]
        first_list=[]  
        for i in range(len(test_list_)):
            #get item as sens_*******
            c= test_list_[i]['track_input_item']
            # strip tail
            varr= c.split("sens_",1)[1].strip()
            if varr in list_in_desired_order:
                ind_ =list_in_desired_order.index(varr)
                #testlist:  [a,b,c, f]
                #order_list [c,d,a]
                #new_list   [2,5,1,5]
                #get index
                new_list.append(ind_)
                first_list.append(i)
            else:
                #append at the end
                new_list.append(len(test_list_))
                 #new_list   [1,2,3,4]
                first_list.append(i)
        
        return new_list, first_list

        
   
    def colnum_string(n):
        # Similar to column_index_from_string 
        # 5 returns E
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string 

   
  
    

    def _populate_cal_working_capital_nominal(self, w_sheet):
         
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
       
        #1. copy accounts_receivable
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','accounts_receivable','0%', 'Inputs')
         
        #2. gross_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_working_capital_nominal',
                    'gross_sales_revenue',number_format) 
        
        #3. Cal accounts_receivable
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_receivable',
                                                   'gross_sales_revenue','accounts_receivable_val')

        #4. copy accounts_payable
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','accounts_payable','0%', 'Inputs')
        
        
        #5. copy total_input_cost_nominal
        self._transfer_cell_range(w_sheet,'calc_purchases_nominal','cal_working_capital_nominal',
                    'total_input_cost_nominal',number_format) 
 
        
        #6. Cal accounts_payable
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_payable',
                                                   'total_input_cost_nominal','accounts_payable_val')


        #7. copy cash_balance
        self._transfer_cell(w_sheet,'working_capital','cal_working_capital_nominal','cash_balance','0%', 'Inputs')
        
        #8. Cal cash_balance
        self._product_value_and_cell_range(w_sheet, 'cal_working_capital_nominal','cash_balance',
                                                   'gross_sales_revenue','cash_balance_val')

        #9. Change in Accounts Receivable
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_receivable_val','change_in_AR',True)
        
        #10. Change in accounts_payable
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','accounts_payable_val','change_in_AP', True)
        
        #11 Change in cash_balance
        self._delta_diff_in_cell_range(w_sheet, 'cal_working_capital_nominal','cash_balance_val','change_in_CB')
    
   
    def _add_cal_finished_product_inventory_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'FINISHED PRODUCT INVENTORY VALUATION USING FIFO METHOD (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_finished_product_inventory' in self.track_inputs:
            self.track_inputs['cal_finished_product_inventory']= {}# insert inner

        if not ('header' in self.track_inputs['cal_finished_product_inventory']):
            self.track_inputs['cal_finished_product_inventory']['header'] = {}
  
        if 'header' in self.track_inputs['cal_finished_product_inventory']:
            self.track_inputs['cal_finished_product_inventory']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_finished_product_inventory', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'OPENING INVENTORY',row_index)
        row_index +=1

        itemlist= [
            'closing_inventory', 'prev_closing_inventory',
            'prev_total_unit_cost_of_production_per_ton_nominal',
            'cost_of_opening_inventory_fifo',

            'production_quantity',
            'total_unit_cost_of_production_per_ton_nominal',
            
            'sales_quantity', 'quantity_sold_from_this_yr_production',
            'cost_of_proportion_sales_produced','cost_of_goods_sold',

            'quantity_remained_unsold','cost_of_closing_inventory'
            ]

        ic_parameters = {
            #OPENING INVENTORY
            'closing_inventory': {'title': 'Closing Inventory','value': None, 'units': 'T_TONS'},
            'prev_closing_inventory': {'title': 'Quantity from previous year (closing inventory)','value': None, 'units': 'T_TONS'},
            'prev_total_unit_cost_of_production_per_ton_nominal': {'title': 'Unit cost from previous year per ton (nominal)','value': None, 'units': 'LC'},
            'cost_of_opening_inventory_fifo': {'title': "Cost of opening inventory (using previous year's unit cost FIFO method)",'value': None, 'units': 'T_LC'},# linked cell
             #ADDITIONS
            'production_quantity': {'title': 'Production Quantity','value': None, 'units': 'T_TONS'},
            'total_unit_cost_of_production_per_ton_nominal': {'title': 'Total unit cost of production per ton (nominal)','value': None, 'units': 'LC'},
             #WITHDRAWALS
            'sales_quantity': {'title': 'Sales quantity','value': None, 'units': 'TONS'},
            'quantity_sold_from_this_yr_production': {'title': "Quantity sold from this yr's production",'value': None, 'units': 'TONS'},
            'cost_of_proportion_sales_produced': {'title': 'Cost of the proportion of sales produced in current year','value': None, 'units': 'T_LC'},
            'cost_of_goods_sold': {'title': "COST OF GOODS SOLD (FOR INCOME STATEMENT)",'value': None, 'units': 'T_LC'},
             #CLOSING INVENTORY
            'quantity_remained_unsold': {'title': 'Quantity remained (unsold) in year','value': None, 'units': 'T_TONS'},
            'cost_of_closing_inventory': {'title': "Cost of closing inventory (using current year's unit cost)",'value': None, 'units': 'T_LC'},
           
        }

				
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_finished_product_inventory']):
                    self.track_inputs['cal_finished_product_inventory'][item] = {}

                self.track_inputs['cal_finished_product_inventory'][item]['row'] = row_index
                self.track_inputs['cal_finished_product_inventory'][item]['unit'] = _unit
                self.track_inputs['cal_finished_product_inventory'][item]['col'] = 6#redundant
                self.track_inputs['cal_finished_product_inventory'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['cost_of_opening_inventory_fifo','total_unit_cost_of_production_per_ton_nominal','cost_of_goods_sold'] :
                    row_index +=1 

                    if item=='cost_of_opening_inventory_fifo':
                        self._add_sub_heading(w_sheet,'ADDITIONS',row_index)
                        row_index +=1

                    if item=='total_unit_cost_of_production_per_ton_nominal':
                        self._add_sub_heading(w_sheet,'WITHDRAWALS',row_index)
                        row_index +=1    

                    if item=='cost_of_goods_sold':
                        self._add_sub_heading(w_sheet,'CLOSING INVENTORY',row_index)
                        row_index +=1 
                    
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_finished_product_inventory(w_sheet)
       
        return row_index
    
    
     
    def _add_cal_loan_schedule_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'LOAN SCHEDULE (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_loan_schedule' in self.track_inputs:
            self.track_inputs['cal_loan_schedule']= {}# insert inner

        if not ('header' in self.track_inputs['cal_loan_schedule']):
            self.track_inputs['cal_loan_schedule']['header'] = {}
  
        if 'header' in self.track_inputs['cal_loan_schedule']:
            self.track_inputs['cal_loan_schedule']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_loan_schedule', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'OPENING INVENTORY',row_index)
        row_index +=1

        itemlist= [
            'real_interest_rate','risk_premium','us_inflation_rate','nominal_interest_rate',
            'domestic_inflation_rate','nominal_interest_rate_local',
            
            'us_price_index','num_of_installments','LPP',

            'beginning_debt','loan_disbursement','interest_accrued_in_year',
            'principal_paid','interest_paid','total_loan_repayment','outstanding_debt_at_end_of_year',
            
            'debt_cash_inflow_in_local_currency','total_loan_repayment_outflow_local',
             
             'CP', 'OP','interest_during_construction','interest_expense'
            ]

        ic_parameters = {
            'real_interest_rate': {'title': 'Real interest rate','value': None, 'units': 'PERCENT'},
            'risk_premium': {'title': 'Risk premium','value': None, 'units': 'PERCENT'},
            'us_inflation_rate': {'title': 'US Inflation rate','value': None, 'units': 'PERCENT'},
            'nominal_interest_rate': {'title': "Nominal Interest Rate",'value': None, 'units': 'PERCENT'},# linked cell
            'domestic_inflation_rate': {'title': 'Local Inflation rate','value': None, 'units': 'PERCENT'},
            #updated in CF
            'nominal_interest_rate_local': {'title': 'Nominal Interest Rate (Local)','value': None, 'units': 'PERCENT'},# CashFlow derived
             
            
            #SUPPLIERS CREDIT
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},
            'num_of_installments': {'title': "No. of installments",'value': None, 'units': 'NUMBER'},
            'LPP': {'title': 'Loan principle repayment','value': None, 'units': 'FLAG'},
            #Loan Repayment Schedule
            'beginning_debt': {'title': "Beginning Debt",'value': None, 'units': 'T_LC'},
            'loan_disbursement': {'title': 'Loan disbursement','value': None, 'units': 'T_LC'},
            'interest_accrued_in_year': {'title': "Interest accrued in year",'value': None, 'units': 'T_LC'},
            'principal_paid': {'title': "Principal paid",'value': None, 'units': 'T_LC'},
            'interest_paid': {'title': 'Interest paid','value': None, 'units': 'T_LC'},
            'total_loan_repayment': {'title': "Total Loan Repayment",'value': None, 'units': 'T_LC'},
            'outstanding_debt_at_end_of_year': {'title': "Outstanding debt at end of year",'value': None, 'units': 'T_LC'},

          
            'debt_cash_inflow_in_local_currency': {'title': "Debt cash Inflow in Local Currency",'value': None, 'units': 'T_LC'},
            'total_loan_repayment_outflow_local': {'title': 'Total Loan Repayment as an outflow in Local Currency','value': None, 'units': 'T_LC'},
            
            'CP': {'title': "Construction period flag",'value': None, 'units': 'FLAG'},
            'OP': {'title': "Operating period",'value': None, 'units': 'FLAG'},
            'interest_during_construction': {'title': 'Interest during Construction, Capitalized for Tax Purposes','value': None, 'units': 'T_LC'},
            'interest_expense': {'title': "Interest expense (for income statement)",'value': None, 'units': 'T_LC'},
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_loan_schedule']):
                    self.track_inputs['cal_loan_schedule'][item] = {}

                self.track_inputs['cal_loan_schedule'][item]['row'] = row_index
                self.track_inputs['cal_loan_schedule'][item]['unit'] = _unit
                self.track_inputs['cal_loan_schedule'][item]['col'] = 6#redundant
                self.track_inputs['cal_loan_schedule'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['nominal_interest_rate_local','LPP','outstanding_debt_at_end_of_year',
                             'total_loan_repayment_outflow_local','outstanding_debt_at_end_of_year'] :
                    row_index +=1 

                    if item=='nominal_interest_rate_local':
                        self._add_sub_heading(w_sheet,'SUPPLIERS CREDIT',row_index)
                        row_index +=1

                    if item=='LPP':
                        self._add_sub_heading(w_sheet,'Loan Repayment Schedule',row_index)
                        row_index +=1    

                  
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_loan_schedule(w_sheet)
       
        return row_index

       
    
    def _add_cal_depreciation_for_tax_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'DEPRECIATION FOR TAX PURPOSES (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_depreciation_for_tax' in self.track_inputs:
            self.track_inputs['cal_depreciation_for_tax']= {}# insert inner

        if not ('header' in self.track_inputs['cal_depreciation_for_tax']):
            self.track_inputs['cal_depreciation_for_tax']['header'] = {}
  
        if 'header' in self.track_inputs['cal_depreciation_for_tax']:
            self.track_inputs['cal_depreciation_for_tax']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_depreciation_for_tax', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Initial total values',row_index)
        row_index +=1

        itemlist= [
            'total_cost_machinery_nominal','total_machinery','buildings_nominal',
            'total_buildings','interest_during_construction','soft_capital_costs_interest_during_construction',

            'operation_start_year','tax_life_of_machinery','tax_life_of_buildings','tax_life_of_soft_capital_costs',

            'OP','machinery_depreciation_expense', 'buildings_depreciation_expense', 'soft_capital_costs_interest_during_construction_expense','annual_depreciation_expense'
            ]
        ic_parameters = {
            #Initial total values           
            'total_cost_machinery_nominal': {'title': 'Total Cost of Machinery including import duty (nominal)','value': None, 'units': 'T_LC'},
            'total_machinery': {'title': 'Machinery','value': None, 'units': 'T_LC'},
            'buildings_nominal': {'title': 'Buildings (nominal)','value': None, 'units': 'T_LC'},
            'total_buildings': {'title': "Buildings",'value': None, 'units': 'T_LC'},# linked cell
            'interest_during_construction': {'title': 'Interest during Construction, Capitalized for Tax Purposes','value': None, 'units': 'T_LC'},
            'soft_capital_costs_interest_during_construction': {'title': 'Soft Capital Costs (Interest during construction)','value': None, 'units': 'T_LC'},
            #Recovery Period for income tax 
            'operation_start_year': {'title': 'Operation Start Year','value': None, 'units': 'YEAR'},
            'tax_life_of_machinery': {'title': "Tax life of machinery",'value': None, 'units': 'YEARS'},
            'tax_life_of_buildings': {'title': 'Tax life of buildings','value': None, 'units': 'YEARS'},
            'tax_life_of_soft_capital_costs': {'title': "Tax life of soft capital costs",'value': None, 'units': 'YEARS'},
            #Depreciation
            'OP': {'title': "Operation period",'value': None, 'units': 'FLAG'},
            'machinery_depreciation_expense': {'title': 'Machinery depreciation expense','value': None, 'units': 'T_LC'},
            'buildings_depreciation_expense': {'title': "Buildings depreciation expense",'value': None, 'units': 'T_LC'},
            'soft_capital_costs_interest_during_construction_expense': {'title': "Soft Capital Costs (Interest during construction)",'value': None, 'units': 'T_LC'},
            'annual_depreciation_expense': {'title': 'Annual Depreciation Expense','value': None, 'units': 'T_LC'},
           
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_depreciation_for_tax']):
                    self.track_inputs['cal_depreciation_for_tax'][item] = {}

                self.track_inputs['cal_depreciation_for_tax'][item]['row'] = row_index
                self.track_inputs['cal_depreciation_for_tax'][item]['unit'] = _unit
                self.track_inputs['cal_depreciation_for_tax'][item]['col'] = 6#redundant
                self.track_inputs['cal_depreciation_for_tax'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                #go to newline
                row_index +=1 
                if item in ['soft_capital_costs_interest_during_construction','tax_life_of_soft_capital_costs',] :
                    row_index +=1 

                    if item=='soft_capital_costs_interest_during_construction':
                        self._add_sub_heading(w_sheet,'Recovery period for income tax',row_index)
                        row_index +=1

                    if item=='tax_life_of_soft_capital_costs':
                        self._add_sub_heading(w_sheet,'Depreciation expense',row_index)
                        row_index +=1    

                      
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_depreciation_for_tax(w_sheet)
       
        return row_index
     
    
    def _add_cal_income_tax_statement_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'INCOME TAX STATEMENT (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'

        if not 'cal_income_tax_statement' in self.track_inputs:
            self.track_inputs['cal_income_tax_statement']= {}# insert inner

        if not ('header' in self.track_inputs['cal_income_tax_statement']):
            self.track_inputs['cal_income_tax_statement']['header'] = {}
  
        if 'header' in self.track_inputs['cal_income_tax_statement']:
            self.track_inputs['cal_income_tax_statement']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_income_tax_statement', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        self._add_sub_heading(w_sheet,'Revenues',row_index)
        row_index +=1

        
      				
       
				

        itemlist= [
            'net_sales_revenue','cost_of_goods_sold','annual_depreciation_expense', 'total_indirect_labour_cost',
            'cost_of_electricity_per_pen_per_annum_nominal','other_indirect_costs_nominal','total_operating_expenses',
            
            'income_from_operations','interest_paid','pre_tax_income','cumulative_losses',
            'taxable_income_losses_carried_forward','corporate_income_tax','income_tax_payment','net_after_tax_income',

            'share_capital','dividend_payout_ratio','dividend_payouts','num_of_shares','earning_per_share',
            'retained_earnings', 'shareholder_equity' ,'gross_profit',  'gross_profit_margin',
            'net_profit_margin',

            'debt_to_equity_ratio',
            ]

        ic_parameters = {
            # Revenues
            'net_sales_revenue': {'title': 'Net sales revenue','value': None, 'units': 'T_LC'},
            
             #Operating expenses
            'cost_of_goods_sold': {'title': 'Cost of goods sold','value': None, 'units': 'T_LC'},
            'annual_depreciation_expense': {'title': 'Annual Depreciation Expense','value': None, 'units': 'T_LC'},
            'total_indirect_labour_cost': {'title': "Total Indirect Labour Cost",'value': None, 'units': 'T_LC'},# linked cell
            'cost_of_electricity_per_pen_per_annum_nominal': {'title': 'Cost of electricity (nominal)','value': None, 'units': 'T_LC'},
            'other_indirect_costs_nominal': {'title': 'Other Indirect Costs (nominal)','value': None, 'units': 'T_LC'},
            'total_operating_expenses': {'title': "Total Operating Expenses",'value': None, 'units': 'T_LC'},
            
            'income_from_operations': {'title': 'INCOME FROM OPERATIONS','value': None, 'units': 'T_LC'},
            
            'interest_paid': {'title': "Interest paid",'value': None, 'units': 'T_LC'},
            
            'pre_tax_income': {'title': 'PRE-TAX INCOME','value': None, 'units': 'T_LC'},
            
            'cumulative_losses': {'title': "Cumulative losses",'value': None, 'units': 'T_LC'},
            
            'taxable_income_losses_carried_forward': {'title': "Taxable income (losses carried forward)",'value': None, 'units': 'T_LC'},
            
            'corporate_income_tax': {'title': 'Corporate income tax','value': None, 'units': 'PERCENT'},
            'income_tax_payment': {'title': "Income tax payment",'value': None, 'units': 'T_LC'},
            'net_after_tax_income': {'title': "NET AFTER-TAX INCOME",'value': None, 'units': 'T_LC'},

          	
            # EQUITY
            'share_capital': {'title': "Share Capital",'value': None, 'units': 'T_LC'},
            'dividend_payout_ratio': {'title': 'Dividend payout ratio','value': None, 'units': 'PERCENT'},
            'dividend_payouts': {'title': 'Dividend payouts','value': None, 'units': 'PERCENT'},
            'num_of_shares': {'title': "N# of Shares",'value': None, 'units': 'NUMBER'},
            'earning_per_share': {'title': "EPS (Earning Per share)",'value': None, 'units': 'T_LC'},
            
            'retained_earnings': {'title': 'Retained Earnings','value': None, 'units': 'T_LC'},
            'shareholder_equity': {'title': "Shareholder Equity",'value': None, 'units': 'T_LC'},
            'gross_profit': {'title': 'GROSS PROFIT   (Sales-COGS)','value': None, 'units': 'T_LC'},
            'gross_profit_margin': {'title': "GROSS PROFIT MARGIN   [(Sales-COGS)/Sales]",'value': None, 'units': 'PERCENT'},
            'net_profit_margin': {'title': 'NET PROFIT MARGIN','value': None, 'units': 'PERCENT'},
            
            'debt_to_equity_ratio': {'title': "Debt-to-Equity Ratio",'value': None, 'units': 'PERCENT'},
          
            
        }

        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_income_tax_statement']):
                    self.track_inputs['cal_income_tax_statement'][item] = {}

                self.track_inputs['cal_income_tax_statement'][item]['row'] = row_index
                self.track_inputs['cal_income_tax_statement'][item]['unit'] = _unit
                self.track_inputs['cal_income_tax_statement'][item]['col'] = 6#redundant
                self.track_inputs['cal_income_tax_statement'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
       
                
                #go to newline
                row_index +=1 
                if item in ['net_sales_revenue','LPP','total_operating_expenses','income_from_operations',
                            'net_after_tax_income','interest_paid','pre_tax_income',
                            'taxable_income_losses_carried_forward','income_tax_payment',
                             'earning_per_share','shareholder_equity',
                            'net_profit_margin'] :
                    row_index +=1 

                    if item=='net_sales_revenue':
                        self._add_sub_heading(w_sheet,'Operating expenses',row_index)
                        row_index +=1

                    if item=='net_after_tax_income':
                        self._add_sub_heading(w_sheet,'EQUITY',row_index)
                        row_index +=1    

                  
        
                       
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_income_tax_statement(w_sheet)
       
        return row_index
  
    def _populate_cal_income_tax_statement(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. net_sales_revenue
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_income_tax_statement','net_sales_revenue',number_format)

        #2. cost_of_goods_sold
        self._transfer_cell_range(w_sheet,'cal_finished_product_inventory','cal_income_tax_statement','cost_of_goods_sold',number_format)
        
        #3. annual_depreciation_expense
        self._transfer_cell_range(w_sheet,'cal_depreciation_for_tax','cal_income_tax_statement','annual_depreciation_expense',number_format)

        #4. total_indirect_labour_cost
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','total_indirect_labour_cost',number_format)
       
        #5. cost_of_electricity_per_pen_per_annum_nominal
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','cost_of_electricity_per_pen_per_annum_nominal',number_format)        
        #6. other_indirect_costs_nominal
        self._transfer_cell_range(w_sheet,'cal_labour_cost_nominal','cal_income_tax_statement','other_indirect_costs_nominal',number_format)

        #7. sum  total_operating_expenses
        formalue_string =[
                {'header': 'cal_income_tax_statement', 'para': 'cost_of_goods_sold'},
                {'header': 'cal_income_tax_statement', 'para': 'other_indirect_costs_nominal'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'total_operating_expenses')
        
        #8. income_from_operations
        #==I305-I313
        #=A-B
        #=net_sales_revenue - total_operating_expenses 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'total_operating_expenses','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'income_from_operations')

        #9. interest_paid
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_income_tax_statement','interest_expense',
                                   number_format, None,'interest_paid')
        
        #10. pre_tax_income
        #==I315-I317
        #=A-B
        #=net_sales_revenue - total_operating_expenses 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'income_from_operations','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'interest_paid','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'pre_tax_income')

        #11.cumulative_losses
        pre_tax_income_r_index, _, found_state_pre_tax_income= self._retrieve_cell_row_colm('cal_income_tax_statement','pre_tax_income')
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','cumulative_losses')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    prev= cell.column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_cum_tax_cell = prev_letter + '$' + str(cell.row)
                    pre_tax_income_cell = cell.column + '$' + str(pre_tax_income_r_index) if found_state_pre_tax_income else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MIN({prev_cum_tax_cell}+{pre_tax_income_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #11.taxable_income_losses_carried_forward
        pre_tax_income_r_index, _, found_state_pre_tax_income= self._retrieve_cell_row_colm('cal_income_tax_statement','pre_tax_income')
        cumulative_losses_r_index, _, found_state_cumulative_losses= self._retrieve_cell_row_colm('cal_income_tax_statement','cumulative_losses')
   
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','taxable_income_losses_carried_forward')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    prev= cell.column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    pre_tax_income_cell = cell.column + '$' + str(pre_tax_income_r_index) if found_state_pre_tax_income else '0'
                    last_year_cum_looses_cell = prev_letter + '$' + str(cumulative_losses_r_index) if found_state_cumulative_losses else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MAX({last_year_cum_looses_cell}+{pre_tax_income_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #12. taxes
        self._transfer_cell(w_sheet,'taxes','cal_income_tax_statement','corporate_income_tax','0%', 'Inputs')
        
        #13.income_tax_payment
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','corporate_income_tax')
        corporate_income_tax_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
                 
        taxable_income_losses_r_index, _, found_state_taxable_income_losses= self._retrieve_cell_row_colm('cal_income_tax_statement','taxable_income_losses_carried_forward')
   
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','income_tax_payment')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    taxable_income_losses_cell = cell.column + '$' + str(taxable_income_losses_r_index) if found_state_taxable_income_losses else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=MAX({corporate_income_tax_cell}*{taxable_income_losses_cell},0)'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        
        #14. net_after_tax_income
        #==I315-I317
        #=A-B
        #=pre_tax_income - income_tax_payment 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'pre_tax_income','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'income_tax_payment','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'net_after_tax_income')

        #15. 'share_capital'
        r_index, c_index, found_state = self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = '$' + get_column_letter(c_index) + '$' + str(r_index)
        
        year_header_r_index, _, found_state_year_header= self._retrieve_cell_row_colm('cal_income_tax_statement','header')
        equity_towards_investment_r_index, _, found_state_equity_towards_investment= self._retrieve_cell_row_colm('calc_investment_cost_nominal','equity_towards_investment')
        
        shareholder_equity_r_index, _, found_state_shareholder_equity= self._retrieve_cell_row_colm('cal_income_tax_statement','shareholder_equity')
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_income_tax_statement','share_capital')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    year_header_cell = cell.column + '$' + str(year_header_r_index) if found_state_year_header else '0'
                    equity_towards_investment_cell = cell.column + '$' + str(equity_towards_investment_r_index) if found_state_equity_towards_investment else '0'
                    
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_shareholder_equity_cell = prev_letter + '$' + str(shareholder_equity_r_index) if found_state_shareholder_equity else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'=IF({year_header_cell}=Inputs!{base_period_cell},{equity_towards_investment_cell},{prev_shareholder_equity_cell})'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #16 'dividend_payout_ratio'
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cal_income_tax_statement','dividend_payout_ratio','0%', 'Inputs')
        
        #17 'dividend_payouts'
        #=I326*$F$330
        #=A*B
        #=dividend_payout_ratio * net_after_tax_income 
        # formalue_string =[
        #         {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
        #         {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'dividend_payout_ratio','cell_type': 'single'},
        #         {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
        #         {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
               
        #     ]
        formalue_string =[
                {'value':'=MAX(MIN(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, 
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'dividend_payout_ratio','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': '),0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, 
            ]
           
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'dividend_payouts')
        #=MIN(J345;$F$349*J$345)
        
        #18 'num_of_shares'
        self._transfer_cell(w_sheet,'macroeconomic_parameters','cal_income_tax_statement','num_of_shares',number_format_integer, 'Inputs')
       
        #19 'earning_per_share'
        #=I331/$F$332
        #=A*B
        #=num_of_shares * dividend_payouts 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'num_of_shares','cell_type': 'single'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'dividend_payouts','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'earning_per_share')

        #20 'retained_earnings'
        #=I326-I331
        #=A-B
        #=net_after_tax_income - dividend_payouts 
        formalue_string =[
                {'value':'=MAX(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'dividend_payouts','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'retained_earnings')

        #21 'shareholder_equity'
        #=I329+I335
        #=A+B
        #=share_capital + retained_earnings 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'share_capital','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'retained_earnings','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'shareholder_equity')

        itemlist= [
            'net_sales_revenue','cost_of_goods_sold','annual_depreciation_expense', 'total_indirect_labour_cost',
            'cost_of_electricity_per_pen_per_annum_nominal','other_indirect_costs_nominal','total_operating_expenses',
            
            'income_from_operations','interest_paid','pre_tax_income','cumulative_losses',
            'taxable_income_losses_carried_forward','corporate_income_tax','income_tax_payment','net_after_tax_income',

            'share_capital','dividend_payout_ratio','num_of_shares','earning_per_share',
            'retained_earnings', 'shareholder_equity' ,'gross_profit',  'gross_profit_margin',
            'net_profit_margin',

            'debt_to_equity_ratio',
            ]
        #22. 'gross_profit'
        #===I305-I308
        #=A-B
        #=net_sales_revenue - cost_of_goods_sold 
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'cost_of_goods_sold','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'gross_profit')

        #23. 'gross_profit_margin'
        #=I338/I305
        #=A/B
        #=gross_profit / net_sales_revenue 
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'gross_profit','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'gross_profit_margin',"0.0%")
        
        #24. 'net_profit_margin'
        #=I326/I305
        #=A/B
        #=net_after_tax_income / net_sales_revenue 
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'net_sales_revenue','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_income_tax_statement', 'para': 'net_after_tax_income','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ]
        
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'net_profit_margin',"0.0%")
        
        #25. 'debt_to_equity_ratio'
        #=I326/I305
        #=A/B
        #=begining_debt / shareholder_equity
        formalue_string =[
                {'value':'=IF(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_income_tax_statement', 'para': 'shareholder_equity','cell_type': 'cell_range'},
                {'value': '=0,0,', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
             
            ] 
        
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_income_tax_statement', 'debt_to_equity_ratio',"0.0%")

       
    
    
    
        
     
    
    def _update_cashflow_linkedcells(self,wb):
        #6. Nominal inflation CF update
        # self._transfer_vcell_range(wb['Calc'],'cf_nominal','cal_loan_schedule','nominal_interest_rate',
        #                           '0%',"CF",'nominal_interest_rate_local')
        
        self._transfer_value_tocellrange(wb['Calc'], 'cf_nominal','cal_loan_schedule', 
                                  'nominal_interest_rate', "CF", 'nominal_interest_rate_local','0.0%')
                         

       
   
    def _populate_cal_loan_schedule(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        #1. real_interest_rate
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','real_interest_rate','0%', 'Inputs')
        #2. risk_premium
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','risk_premium','0%', 'Inputs')        
        
        #3. copy us_inflation_rate
        self._transfer_value_tocellrange(w_sheet, 'macroeconomic_parameters','cal_loan_schedule', 
                    'us_inflation_rate',  "Inputs",  None,     '0%')

      
        #4. nominal_interest_rate
        #=$F$219+$F$220+(1+$F$219+$F$220)*I221
        #=A+B+(1+A+B)*C
        #=real_interest +  real_interest +(1 + real_interest  + real_interest)*usinfaltion
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'real_interest_rate','cell_type': 'single'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'risk_premium','cell_type': 'single'},
                {'value': '+(1+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'us_inflation_rate','cell_type': 'cell_range'},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'nominal_interest_rate',"0%")

        #5. Domestic inflation rate
        self._transfer_value_tocellrange(w_sheet, 'macroeconomic_parameters',  'cal_loan_schedule',
                                        'domestic_inflation_rate',  "Inputs", None,     '0%')
        #6. Nominal inflation CF update
        self._transfer_cell_range(w_sheet,'cf_nominal','cal_loan_schedule','nominal_interest_rate_local','0%',"CF")

        
        #7. us price index
        self._transfer_cell_range(w_sheet,'calc_inflation_price_index','cal_loan_schedule', 'us_price_index',number_format,) 

        #8. num_of_installments
        self._transfer_cell(w_sheet,'financing','cal_loan_schedule','num_of_installments',number_format_integer, 'Inputs')
       
        #9. Loan Principal Payment Flags
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule',  'LPP',number_format_integer,"Inputs")

        #10. beginning_debt
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','outstanding_debt_at_end_of_year',
                                   None, None,'beginning_debt',"Calculation",True)

        #11. loan_disbursement
        self._transfer_cell_range(w_sheet,'calc_investment_cost_nominal','cal_loan_schedule','senior_debt_towards_investment',
                                   None, None,'loan_disbursement',"Calculation")

        
        #12. interest_accrued_in_year
        #=I234*I224
        #=A*B
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'nominal_interest_rate_local','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_accrued_in_year')

        #13. prinical_paid
        #=SUM($I$235:$Z$235)/$F$229*K230
        #=SUM(loan_disbursement/num_of_installments*LPP
        r_index, c_index, found_state= self._retrieve_cell_row_colm('cal_loan_schedule','num_of_installments')
        num_of_installments_cell = '$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        loan_disbursement_r_index, _, found_state_loan_disbursement= self._retrieve_cell_row_colm('cal_loan_schedule','loan_disbursement')
        LPP_r_index, _, found_state_LPP= self._retrieve_cell_row_colm('cal_loan_schedule','LPP')
        
        r_index, _, found_state= self._retrieve_cell_row_colm('cal_loan_schedule','principal_paid')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    num_of_installments_cell
                    LPP_cell = cell.column + '$' + str(LPP_r_index) if found_state_LPP else '0'
                    #------------------disbursement------------------------------------
                    a= get_column_letter(9) + str(loan_disbursement_r_index)
                    b= get_column_letter(9 + int(span_)) + str(loan_disbursement_r_index)
                    loan_disbursement_cell = f'SUM({a}:{b})' if found_state_loan_disbursement else '0'
                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF(' + num_of_installments_cell + '=0, 0,' + loan_disbursement_cell + \
                                                              '/'+ num_of_installments_cell  + "*" +  LPP_cell +")"
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
   
        #13. interest_paid
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','interest_accrued_in_year',
                                   number_format, None,'interest_paid',"Calculation")

        
        #14. total_loan_repayment
        #=I234*I224
        #=A+B
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'principal_paid','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'total_loan_repayment')
        
        #15. outstanding_debt_at_end_of_year
        # =I234  +I235  + I236 - I239
        # =A + B + C - D
        #=beginning_debt + loan_disbursement + interest_accrued_in_year -total_loan_repayment 
        #=Nominal Interest Rate (Local) *  Beginning Debt
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'beginning_debt','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'loan_disbursement','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'C', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'D', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'total_loan_repayment','cell_type': 'cell_range'},
               
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'outstanding_debt_at_end_of_year')
          
        #16. debt_cash_inflow_in_local_currency
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','loan_disbursement',
                                   number_format, None,'debt_cash_inflow_in_local_currency',"Calculation")

        #17. total_loan_repayment_outflow_local
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_loan_schedule','total_loan_repayment',
                                   number_format, None,'total_loan_repayment_outflow_local',"Calculation")
        
        #18. CP
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule','CP',  number_format_integer, 'Inputs','CP',"Calculation")

        #19. OP
        self._transfer_cell_range(w_sheet,'flags','cal_loan_schedule','OP', number_format_integer, 'Inputs','OP',"Calculation")

        #20. interest_during_construction
        # =(I236)*I246
        # = A * B 
        # = interest_accrued_in_year *  CP
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'CP','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_during_construction')
          
       
        #21. interest_expense
        #=(I236)*I247
        # = A * B 
        # = interest_accrued_in_year *  OP
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_loan_schedule', 'para': 'interest_accrued_in_year','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'B', 'var_type': 'variable', 'header': 'cal_loan_schedule', 'para': 'OP','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_loan_schedule', 'interest_expense')
          
       
    
    def _populate_cal_depreciation_for_tax(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
      
        
        #1. total_cost_machinery_nominal
        #==I74+I65
        #=A+B
        #=investment_cost_of_pens +  total_cost_machinery_nominal
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'calc_investment_cost_nominal', 'para': 'investment_cost_of_pens','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'calc_investment_cost_nominal', 'para': 'total_cost_machinery_nominal','cell_type': 'cell_range'},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'total_cost_machinery_nominal')

        #2. sum cell total_machinery
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','total_cost_machinery_nominal',
                                          number_format, None,'total_machinery','Calculation')        
       
        
        #3. buildings_nominal
        self._transfer_cell_range(w_sheet,'calc_investment_cost_nominal','cal_depreciation_for_tax',
                              'investment_cost_of_buildings_nominal',    number_format, None, 'buildings_nominal')

        
        #4. sum cell total_buildings
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','buildings_nominal',
                                          number_format, None,'total_buildings','Calculation')        
       
        #5. soft_capital_costs_interest_during_construction
        self._transfer_cell_range(w_sheet,'cal_loan_schedule','cal_depreciation_for_tax',
                              'interest_during_construction',    number_format,)

        
        #6. sum cell soft_capital_costs_interest_during_construction_expense
        self._transfer_sumof_cell_range(w_sheet,'cal_depreciation_for_tax','cal_depreciation_for_tax','interest_during_construction',
                                          number_format, None,'soft_capital_costs_interest_during_construction','Calculation')        
       
        #7. operation_start_year
        self._transfer_cell(w_sheet,'timing','cal_depreciation_for_tax','operation_start_year','General', 'Inputs')
         
        #8. tax_life_of_machinery
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_machinery',number_format_integer, 'Inputs')
        
        #9. tax_life_of_buildings
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_buildings',number_format_integer, 'Inputs')
        
        #10. tax_life_of_soft_capital_costs
        self._transfer_cell(w_sheet,'depreciation','cal_depreciation_for_tax','tax_life_of_soft_capital_costs',number_format_integer, 'Inputs')
       
        #11. OP
        self._transfer_cell_range(w_sheet,'flags','cal_depreciation_for_tax','OP', number_format_integer, 'Inputs','OP',"Calculation")

        
        #12. machinery_depreciation_expense
        #=IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$292 -1);  ($F$284/$F$292)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_machinery  '-1),('   total_machinery '/'  tax_life_of_machinery ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_machinery','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'total_machinery','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'machinery_depreciation_expense')

        
        #13. buildings_depreciation_expense
        #==IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$293 -1);  ($F$286/$F$293)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_buildings  '-1),('   total_buildings '/'  tax_life_of_buildings ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_buildings','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'total_buildings','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'buildings_depreciation_expense')

        #14. soft_capital_costs_interest_during_construction_expense
        #==IF(AND(I280>=Inputs!$F$38;calc!I280<=Inputs!$F$38 +calc!$F$293 -1);  ($F$286/$F$293)*I296;0)
        # '=IF(AND('  A  '>=Inputs!'  B  ','  A  '<=Inputs!' B  + C  '-1),('   D '/'  C ')*' E ',0)'
        # '=IF(AND('  YEAR_Header  '>=Inputs!'  operation_start_year  ','  YEAR_Header  '<=Inputs!' operation_start_year  + tax_life_of_soft_capital_costs  '-1),('   soft_capital_costs_interest_during_construction '/'  tax_life_of_soft_capital_costs ')*' OP ',0)'
       
        formalue_string =[
                {'value':'=IF(AND(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_depreciation_for_tax', 'para': 'header','cell_type': 'cell_range'},
                {'value': '>=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'timing', 'para': 'operation_start_year','cell_type': 'single'},
                {'value': ',', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "A", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '<=Inputs!', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'tax_life_of_soft_capital_costs','cell_type': 'single'},
                {'value': '-1),(', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "D", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'soft_capital_costs_interest_during_construction','cell_type': 'single'},
                {'value': '/', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': '', 'para': '','cell_type': ''},
                {'value': ')*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "E", 'var_type': 'variable', 'header': 'cal_depreciation_for_tax', 'para': 'OP','cell_type': 'cell_range'},
                {'value': ',0)', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
              
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'soft_capital_costs_interest_during_construction_expense')

        
       
        #7. sum  annual_depreciation_expense
        formalue_string =[
                {'header': 'cal_depreciation_for_tax', 'para': 'machinery_depreciation_expense'},
                {'header': 'cal_depreciation_for_tax', 'para': 'soft_capital_costs_interest_during_construction_expense'},
            ]
        self._sumofcolumn_range_fromstring(w_sheet,formalue_string,'cal_depreciation_for_tax', 'annual_depreciation_expense')
        
       
    
    
    def _populate_cal_finished_product_inventory(self, w_sheet):
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_integer ='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
        
        #1. copy closing_inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'closing_inventory',number_format, 'Inputs') 

        #2. copy prev_closing_inventory
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'closing_inventory',number_format, 'Inputs','prev_closing_inventory',"Linkedcell", True,) 

        #3. copy prev total_unit_cost_of_production_per_ton_nominal
        self._transfer_cell_range(w_sheet,'cal_unit_of_production','cal_finished_product_inventory',
                    'total_unit_cost_of_production_per_ton_nominal',number_format, None,
                        'prev_total_unit_cost_of_production_per_ton_nominal',"Linkedcell", True,) 

        #4. cost_of_opening_inventory_fifo
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'prev_total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                
            ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',
                                            'cost_of_opening_inventory_fifo')

        #5. copy production quantity
        self._transfer_cell_range(w_sheet,'production_inventory','cal_finished_product_inventory',
                    'production_quantity',number_format, 'Inputs') 
        
        #6. copy total_unit_cost_of_production_per_ton_nominal
        self._transfer_cell_range(w_sheet,'cal_unit_of_production','cal_finished_product_inventory',
                    'total_unit_cost_of_production_per_ton_nominal',number_format, ) 


        #7. copy sales_quantity
        self._transfer_cell_range(w_sheet,'cal_production_sales_nominal','cal_finished_product_inventory',
                    'sales_quantity',number_format,) 

        
        #8. quantity_sold_from_this_yr_production
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'sales_quantity','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},  ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',   'quantity_sold_from_this_yr_production')
        #9.  cost_of_proportion_sales_produced
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'quantity_sold_from_this_yr_production','cell_type': 'cell_range'},
                {'value': '/1000', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''}, ]
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',  'cost_of_proportion_sales_produced')
        #10.  cost_of_goods_sold
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'cost_of_opening_inventory_fifo','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'cost_of_proportion_sales_produced','cell_type': 'cell_range'},]
                
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory',   'cost_of_goods_sold')
        
        #11.  quantity_remained_unsold
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'prev_closing_inventory','cell_type': 'cell_range'},
                {'value': '+', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'production_quantity','cell_type': 'cell_range'},
                {'value': '-', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "C", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'sales_quantity','cell_type': 'cell_range'},]
             
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory', 'quantity_remained_unsold')
        
        #11.  cost_of_closing_inventory
        formalue_string =[
                {'value':'=', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': 'A', 'var_type': 'variable','header': 'cal_finished_product_inventory', 'para': 'total_unit_cost_of_production_per_ton_nominal','cell_type': 'cell_range'},
                {'value': '*', 'var_type': 'const', 'header': '', 'para': '','cell_type': ''},
                {'value': "B", 'var_type': 'variable', 'header': 'cal_finished_product_inventory', 'para': 'quantity_remained_unsold','cell_type': 'cell_range'}  ]
             
        self._calculate_formalue_fromstring(w_sheet,formalue_string,'cal_finished_product_inventory', 'cost_of_closing_inventory')
        
      
  
                     
             
    def _add_cal_working_capital_nominal_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'WORKING CAPITAL (Nominal)'
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'cal_working_capital_nominal' in self.track_inputs:
            self.track_inputs['cal_working_capital_nominal']= {}# insert inner

        if not ('header' in self.track_inputs['cal_working_capital_nominal']):
            self.track_inputs['cal_working_capital_nominal']['header'] = {}
  
        if 'header' in self.track_inputs['cal_working_capital_nominal']:
            self.track_inputs['cal_working_capital_nominal']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'cal_working_capital_nominal', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
       

        itemlist= ['accounts_receivable','gross_sales_revenue','accounts_receivable_val',
                   'accounts_payable', 'total_input_cost_nominal','accounts_payable_val',
                   'cash_balance','cash_balance_val',
                   'change_in_AR','change_in_AP','change_in_CB']
        ic_parameters = {
            # PI
            'accounts_receivable': {'title': 'Accounts receivable [% of Gross Sales]','value': None, 'units': 'PERCENT'},
            'gross_sales_revenue': {'title': 'Gross sales revenue','value': None, 'units': 'T_LC'},
            'accounts_receivable_val': {'title': 'Accounts receivable [% of Gross Sales]','value': None, 'units': 'T_LC'},# linked cell
            
            'accounts_payable': {'title': 'Acccounts payable (% of total input cost)','value': None, 'units': 'PERCENT'},# linked cell
            'total_input_cost_nominal': {'title': 'Total Input Cost (nominal)','value': None, 'units': 'T_LC'},
            'accounts_payable_val': {'title': 'Acccounts payable (% of total input cost)','value': None, 'units': 'T_LC'},
            
            'cash_balance': {'title': 'Cash balance  (% of gross sales)','value': None, 'units': 'PERCENT'},
            'cash_balance_val': {'title': 'Cash balance  (% of gross sales)','value': None, 'units': 'T_LC'},
            
            'change_in_AR': {'title': 'Change in A/R','value': None, 'units': 'T_LC'},
            'change_in_AP': {'title': 'Change in A/P','value': None, 'units': 'T_LC'},
            'change_in_CB': {'title': 'Change in cash balance','value': None, 'units': 'T_LC'},
        }
        #-----------------------------------
        for item in itemlist:
            #print(item)
            if item in ic_parameters.keys():
                #print(item)
                _unit =ic_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        ic_parameters[item]['title'], ic_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['cal_working_capital_nominal']):
                    self.track_inputs['cal_working_capital_nominal'][item] = {}

                self.track_inputs['cal_working_capital_nominal'][item]['row'] = row_index
                self.track_inputs['cal_working_capital_nominal'][item]['unit'] = _unit
                self.track_inputs['cal_working_capital_nominal'][item]['col'] = 6#redundant
                self.track_inputs['cal_working_capital_nominal'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['accounts_receivable_val','accounts_payable_val','cash_balance_val'] :
                    row_index +=1 

        
        # Sub-Section B skipp rows
        row_index +=1
        self._populate_cal_working_capital_nominal(w_sheet)
       
        return row_index 

   


    def _sumofcolumn_range_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, number_format=None,last_column=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        """ 
        =SUM(I308:I312)
        =SUM( I  row_start :I   row_end)
        """
        row_start=None
        row_end = None
        for item in formalue_string:
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    if row_start ==None:
                        row_start =source_r_index
                    elif row_start > source_r_index:
                        row_start = source_r_index 

                    if row_end ==None:
                        row_end =source_r_index
                    elif row_end < source_r_index:
                        row_end = source_r_index 
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

            if last_column !=None and type(last_column)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(last_column)) + str(r_index) # D39
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    maths_formulae ='=SUM('+ cell.column + str(row_start) + ':' + cell.column +  str(row_end) +')'
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                      
    def _calculate_formalue_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, 
                        number_format=None,lower_upper_bound=None ):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        """ 
        #=IF(I187=0;0;I188*$F$30/I187)
        '=IF('+ I187 + '=0,0'+ I188 + '*' + $F$30 + '/' + I187 + ')'
        '=IF('+ A + '=0,0' + B + '*' + C + '/' + D + ')'

        ['=IF(', A , '=0,0' , B , '*' , C , '/' , D , ')']

        ['=IF(',  '=0,0' ,  '*' ,  '/'  ')'] [ A ,  B ,  C ,  D , ] 
        """
        #seperate string, variables
        variables_ ={}
        memory_={}
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='variable' and item['cell_type']== 'single':
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                variables_[item['value']]= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable' and item['cell_type']== 'cell_range':
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                memory_[item['value']]={}
                memory_[item['value']]['row']= source_r_index
                memory_[item['value']]['found_state'] = found_state_source

       
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39


            #change if possible
            #int or dict
            if lower_upper_bound !=None and type(lower_upper_bound)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(lower_upper_bound)) + str(r_index) # D39
            elif lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
           
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    maths_formulae =''
                    for item in formalue_string:
                        if item['var_type']=='const' :
                          maths_formulae += item['value']
                        elif item['var_type']=='variable' and item['cell_type']== 'single':
                            #retrieve from variabble
                            maths_formulae +=  variables_[item['value']] 
                        elif item['var_type']=='variable' and item['cell_type']== 'cell_range':
                            source_r_index = memory_[item['value']]['row']
                            found_state = memory_[item['value']]['found_state']
                            variables_[item['value']]= cell.column + '$' + str(source_r_index) if found_state else '0'
                            maths_formulae +=  variables_[item['value']] 
                        elif item['value'] in variables_:
                            #already calculated
                            #print(item['value'],variables_[item['value']]) 
                            maths_formulae +=  variables_[item['value']]      
                    #print(maths_formulae) 
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                   

         
    def _cell_formalue_fromstring(self, w_sheet,
                        formalue_string, target_header, target_para, number_format=None, 
                        cell_style='Calculation'):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
      
        #seperate string, variables
        variables_ ={}
        memory_={}
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='variable' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                #keep previous variable unspoilt
                #subsequent similar variable are skiipped
                if not(item['value'] in variables_):
                    variables_[item['value']]=  get_column_letter(c_index) + str(r_index) if found_state else '0'
           
        #write in target--------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            maths_formulae =''
            for item in formalue_string:
                if item['var_type']=='const' :
                    maths_formulae += item['value']
                elif item['var_type']=='variable' :
                    #retrieve from variabble
                    maths_formulae +=  variables_[item['value']] 
                elif item['value'] in variables_:
                    #already calculated
                    #print(item['value'],variables_[item['value']]) 
                    maths_formulae +=  variables_[item['value']]      
            col= get_column_letter(c_index) 
            #print("formuales: ", maths_formulae)
            w_sheet['%s%s'%(col, r_index)] = maths_formulae
            w_sheet['%s%s'%(col, r_index)].style = cell_style
            w_sheet['%s%s'%(col, r_index)].number_format =number_format
                   
     
    def _sum_oflist_cell_range(self, w_sheet,
                        list_, target_header, target_para, const_list=[]):
        #----List
        mat_attr= []
        for item in list_:
            header_ =item['header']
            para_ =item['para']
            action_ =item['action']           
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_

            mat_attr.append(prop)
        #print(mat_attr)
        mul_const= ''
        div_const= ''

        for x in const_list:
            header_ =x['header']
            para_ =x['para']
            action_ = x['action']          
            #source
            r_index, c_index, found_state= self._retrieve_cell_row_colm(header_,para_)
            val_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
            if mul_const ==''  and action_ =="*":
               mul_const = val_cell
            elif action_ =="*":
               mul_const += '*' +  val_cell

            if div_const ==''  and action_ =="/" :#and val_cell!='0':# avoid zero divsion
               div_const = val_cell
            elif action_ =="/": #and val_cell !='0':# avoid zero divsion
               div_const += '*' +  val_cell    

        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    
                    maths_formulae =''
                    for i_val in mat_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if maths_formulae =='' and i_val['action'] in ["-"]:
                            maths_formulae = i_val['action'] + val
                        elif maths_formulae =='':  
                            maths_formulae = val
                        else:
                            maths_formulae += i_val['action'] + val        
                    
                    # preepend const if value
                    if not (mul_const==""):
                        maths_formulae = mul_const + "*" + maths_formulae    
                    # append division if value
                    if not (div_const==""):
                        maths_formulae = maths_formulae + "/" + div_const 

                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

     
    def _productof_sumof_list_cell_range(self, w_sheet,
                        listA, listB, target_header, target_para,):
        # mulplies cell range sum together
        #----List
        listA_attr= []
        listB_attr= []
        for item in listA:
            header_ =item['header']
            para_ =item['para']
            action_ =item['action']           
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_
            listA_attr.append(prop)
       
        for x in listB:
            header_ =x['header']
            para_ =x['para']
            action_ = x['action']          
            #source
            source_r_index, _, found_state_source= self._retrieve_cell_row_colm(header_,para_)
            prop={}
            prop['row'] = source_r_index
            prop['found_state'] = found_state_source
            prop['action'] = action_
            
            listB_attr.append(prop)  
               
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    
                    mathsA_formulae =''
                    for i_val in listA_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if mathsA_formulae =='' and i_val['action'] in ["-"]:
                            mathsA_formulae = i_val['action'] + val
                        elif mathsA_formulae =='':  
                            mathsA_formulae = val
                        else:
                            mathsA_formulae += i_val['action'] + val   

                    mathsA_formulae = "(" + mathsA_formulae +")" if len(listA_attr)>1 else mathsA_formulae

                    mathsB_formulae =''
                    for i_val in listB_attr:
                        val =  cell.column + '$' + str(i_val['row']) if  i_val['found_state'] else '0'
                        
                        #only append neg-- on first element
                        if mathsB_formulae =='' and i_val['action'] in ["-"]:
                            mathsB_formulae = i_val['action'] + val
                        elif mathsB_formulae =='':  
                            mathsB_formulae = val
                        else:
                            mathsB_formulae += i_val['action'] + val

                    mathsB_formulae = "(" + mathsB_formulae +")" if len(listB_attr)>1 else mathsB_formulae
                    
                   
                   
                    # preepend const if value
                    if not (mathsA_formulae=="" or mathsB_formulae==""):
                         maths_formulae= mathsA_formulae + "*" + mathsB_formulae    
                    elif not (mathsA_formulae==""):
                        maths_formulae= mathsA_formulae 
                    elif not (mathsB_formulae==""):
                        maths_formulae= mathsB_formulae 

                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

   

    def _delta_diff_in_cell_range(self, w_sheet,
                                     source_header,
                                     source_para, target_para, 
                                     reverse_diff = False, target_header = None,):
       
        #source
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header,source_para)
        
        #write in target--------------
        target_header = source_header if target_header==None else target_header
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_source_cell = prev_letter + str(source_r_index)
                   
                    source_cell = cell.column + '$' \
                        + str(source_r_index) if found_state_source else '0'
                    if reverse_diff:
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + prev_source_cell +'-'+ source_cell
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)] = '=' + source_cell + '-' + prev_source_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'

    def _product_value_and_cell_range(self, w_sheet, 
                          val_header , val_para,
                          source_para, target_para, source_header=None, target_header = None,):
        #value
        r_index, c_index, found_state= self._retrieve_cell_row_colm(val_header,val_para)
        val_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        #source
        source_header = val_header if source_header==None else source_header
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header,source_para)
        
        #write in target--------------
        target_header = source_header if target_header==None else target_header
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    source_cell = cell.column + '$' \
                        + str(source_r_index) if found_state_source else '0'
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=' + val_cell  + '*' + source_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'


     
    def _transfer_sumof_cell_range(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format=None, source_wksheet=None, 
                        target_para = None, cell_style='Linkedcell'):

        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header, source_para)
        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        
        if found_state:
            span_ =self._get_span()
           
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''

            a= get_column_letter(9) + str(source_r_index)
            b= get_column_letter(9 + int(span_)) + str(source_r_index)
            sumofsource_range = f'SUM({a}:{b})' if found_state_source else '0'
                   
            col= get_column_letter(c_index)
            w_sheet['%s%s'%(col, r_index)] = f'={sumofsource_range}'
            w_sheet['%s%s'%(col, r_index)].style = cell_style
            if cell_number_format !=None:
                w_sheet['%s%s'%(col, r_index)].number_format = cell_number_format

  

    def _loan_principal_repayment_periods(self, w_sheet):

        r_index, _, found_state= self._retrieve_cell_row_colm('flags', 'LPP')
        found_lb= False
        lbound= None
        ubound= None
        if found_state:
            span_ =self._get_span()
            # default
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
           
           
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    col_num = column_index_from_string(cell.column)
                    cell_value =  w_sheet['%s%s'%(cell.column, cell.row)].value
                    #print("found_integer:....", cell_value)
                    if int(cell_value) ==1:
                        found_lb= True
                        if lbound==None:
                            #latch the first
                           lbound=col_num
                    if ubound==None:
                        ubound=col_num 
                    else:
                       #continue updating
                       ubound=col_num 
        return lbound, ubound, found_lb
    
    
    
    
   
    
    def _mirr_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        #=IFERROR(MIRR(I111:Z111;$F$113;$F$113);"N/A")
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=IFERROR(MIRR(' + first_source_cell + ':' + last_source_cell + ',' + rate_cell + ',' + rate_cell +'),"N/A")' 
                 
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format ='0%' #number_format


    def _irr_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            
            if item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=IFERROR(IRR(' + first_source_cell + ':' + last_source_cell + '),"N/A")' 
                  
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = '0%'#number_format

    def _npv_cell(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        maths_formulae= '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
                if found_state_source:
                    span_ =self._get_span()
                    #first_slice_point = get_column_letter(9) + str(r_index)# D07
                    #second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

                    upperbound_=9 + int(span_)
                    lowerbound = 9
                    
                    #change if possible
                    if lower_upper_bound != None:
                        if  'lb' in lower_upper_bound :
                            lowerbound=lower_upper_bound['lb']
                            #first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                        if  'ub' in lower_upper_bound :
                            upperbound_=lower_upper_bound['ub']
                            #second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                           
                            span_= upperbound_-lowerbound

                       #=NPV($F$51;K52:O52)+J52
                    next_ =  lowerbound + 1
                    last_col = lowerbound + span_
                    first_source_cell = get_column_letter(lowerbound) + str(source_r_index) if found_state_source else '0'
                    next_source_cell = get_column_letter(next_) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(last_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=NPV('+ rate_cell + ','+ next_source_cell + ':' + last_source_cell + ')+' + first_source_cell
                      
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = maths_formulae
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = "Calculation"
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = number_format

    def _npv_cell_range(self, w_sheet, formalue_string, 
           target_header, target_para,  lower_upper_bound= None, number_format=None):
        
        if number_format==None:
            number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        
        #seperate string, variables
        rate_cell = '0'
        found_state_source= False
        for item in formalue_string:
            #collect single cell values only
            if item['var_type']=='rate' :
                r_index, c_index, found_state= self._retrieve_cell_row_colm(item['header'],item['para'])
                rate_cell= '$' + get_column_letter(c_index)+ '$' +str(r_index) if found_state else '0'
             #collect range cell values only
            elif item['var_type']=='variable': 
                source_r_index, _, found_state_source= self._retrieve_cell_row_colm(item['header'],item['para'])
              
       
        #write in target--------------
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39

            upperbound_=9 + int(span_)
            
            #change if possible
            if lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
                    
                    upperbound_ = ub

                    span_= ub-lb




            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   
                    #=NPV($F$51;K52:O52)+J52
                    ahead =  column_index_from_string(cell.column) + 1
                    next_col = column_index_from_string(cell.column)+ span_
                    curr_source_cell = cell.column + str(source_r_index) if found_state_source else '0'
                    ahead_source_cell = get_column_letter(ahead) + str(source_r_index) if found_state_source else '0'
                    last_source_cell = get_column_letter(next_col) + str(source_r_index) if found_state_source else '0'
                  
                    maths_formulae ='=NPV('+ rate_cell + ','+ ahead_source_cell + ':' + last_source_cell + ')+' + curr_source_cell
                    
                    w_sheet['%s%s'%(cell.column, cell.row)] = maths_formulae
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format =number_format
                               
       
    
    def _transfer_cell_range(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format=None, source_wksheet=None, 
                        target_para = None, cell_style='Linkedcell',shift_prev=False,lower_upper_bound= None):
        
        source_r_index, _, found_state_source= self._retrieve_cell_row_colm(source_header, source_para)
        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)

        if found_state:
            span_ =self._get_span()
            # default
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #change if possible
            if lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07

            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:

                    source_cell = cell.column + '$' + str(source_r_index) if found_state_source else '0'
                    if shift_prev==True:
                        prev= cell.column
                        prev_col = column_index_from_string(cell.column)-1
                        prev_letter= get_column_letter(prev_col)
                        source_cell = prev_letter + '$' + str(source_r_index) if found_state_source else '0'

                   
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'={source_wksheet}{source_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = cell_style
                    if cell_number_format !=None:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = cell_number_format
       
    def _transfer_cell_atcolumn(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None,cell_style='Linkedcell',last_column=None):
        
        if last_column == None: 
            span_ =self._get_span()
            last_column = get_column_letter(9 + int(span_)) 
        
        #------Senior Debt LinkedCell-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        #print(source_header, source_para, "col:", c_index)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            w_sheet['%s%s'%(last_column, r_index)] = f'={source_wksheet}{source_cell}'
            w_sheet['%s%s'%(last_column, r_index)].style = cell_style
            w_sheet['%s%s'%(last_column, r_index)].number_format = cell_number_format
    
    def _get_cell_ref_asformulae(self,  source_header,     source_para , 
                                 source_wksheet=None,   from_col_pos=None):
      
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        if from_col_pos !=None and type(from_col_pos)==int:
            c_index = 9 + int(from_col_pos)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #print(f'={source_wksheet}{source_cell}')
            return f'={source_wksheet}{source_cell}'
        
        #nothing found
        return '0'   
    def _write_text_atcell(self, w_sheet, target_header, target_para, text_value,
                           target_col_pos=None, cell_number_format='General', cell_style=None, ):

        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if target_col_pos !=None and type(target_col_pos)==int:
            c_index = 9 + int(target_col_pos)
        if found_state:
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = text_value
            if cell_style !=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
    
    def _transfer_cell(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None, cell_style='Linkedcell', 
                        from_col_pos=None, target_col_pos=None):
         #------Senior Debt LinkedCell-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        #print(source_header, source_para, "col:", c_index)
        if from_col_pos !=None and type(from_col_pos)==int:
            c_index = 9 + int(from_col_pos)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if target_col_pos !=None and type(target_col_pos)==int:
            c_index = 9 + int(target_col_pos)
        if found_state:
            #default == empty space
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = f'={source_wksheet}{source_cell}'
            #dont change formats if to retain former
            if cell_style !=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            if cell_number_format !=None:    
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
    
    def _transfer_range_formulae_tocell(self, w_sheet, source_header, target_header, 
                        source_para ,cell_number_format='General', 
                        source_wksheet=None, target_para = None,cell_style='Linkedcell',
                        formulae_str='Sum',lower_upper_bound=None):
        
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #change if possible
                #int or dict
            if lower_upper_bound !=None and type(lower_upper_bound)==int:
                #0,1,2,3,
                first_slice_point = get_column_letter(9 + int(span_)-int(lower_upper_bound)) + str(r_index) # D39
            elif lower_upper_bound != None:
                if  'lb' in lower_upper_bound :
                    lb=lower_upper_bound['lb']
                    first_slice_point = get_column_letter(int(lb)) + str(r_index)# D07
                if  'ub' in lower_upper_bound :
                    ub=lower_upper_bound['ub']
                    second_slice_point = get_column_letter(int(ub)) + str(r_index)# D07
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''        
            source_cell = f'{formulae_str}({source_wksheet}{first_slice_point}:{source_wksheet}{second_slice_point})'
            
        else:
            source_cell =  '0'
           

        #if None=== set default as source paramter
        target_para = source_para if target_para==None else target_para
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            #print({source_cell})
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = f'={source_cell}'
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format

    def _transfer_value_tocell(self, w_sheet, source_cell, 
                        target_header, target_para,
                        cell_number_format=None,
                        cell_style='Linkedcell'):
        
        #if None=== set default as source paramter
        r_index, c_index, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        if found_state:
            #default == empty space
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)] = source_cell
            w_sheet['%s%s'%(get_column_letter(c_index), r_index)].style = cell_style
            if cell_number_format!=None:
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = cell_number_format
            else:
                #print ('_transfer_value_tocell','Default formating')
                w_sheet['%s%s'%(get_column_letter(c_index), r_index)].number_format = 'General' 

    
    def _transfer_value_tocellrange(self, w_sheet, source_header, 
                        target_header, source_para,  source_wksheet=None,target_para=None,
                        cell_number_format=None,
                        cell_style='Linkedcell'):
        
        
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm(source_header, source_para)
        source_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state else '0'

        target_para = source_para if target_para==None else target_para
        #if None=== set default as source paramter
        r_index, _, found_state= self._retrieve_cell_row_colm(target_header,target_para)
        
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            source_wksheet = f'{source_wksheet}!' if source_wksheet !=None else ''
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = f'={source_wksheet}{source_cell}'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = cell_style
                    if cell_number_format !=None:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = cell_number_format
       

                            
    def _add_cal_inflation_price_indices_section(self, w_sheet, row_index,total_wsheet_cols, commodity_title='Beef'):
        #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'INFLATION, EXCHANGE RATE AND PRICE INDICES' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'


        if not 'calc_inflation_price_index' in self.track_inputs:
            self.track_inputs['calc_inflation_price_index']= {}# insert inner

        if not ('header' in self.track_inputs['calc_inflation_price_index']):
            self.track_inputs['calc_inflation_price_index']['header'] = {}
  
        if 'header' in self.track_inputs['calc_inflation_price_index']:
            self.track_inputs['calc_inflation_price_index']['header']['row'] = row_index


        flags_year_r_index, c_index, found_state= self._retrieve_cell_row_colm('flags','years')
        #print(f'Retrieving {flags_year_r_index}, {c_index}, {found_state}')
        if found_state:
            span_ =self._get_span()
            start_col_index= 9
        
            self._update_section_header_year(w_sheet, 'calc_inflation_price_index', flags_year_r_index, start_col_index, span_, "Inputs")
        
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
       
        #===================================
        row_index +=1
        #===================================
        colHeader = get_column_letter(5)
        colValue = get_column_letter(6)
        colUnits = get_column_letter(7)
        row_index +=1
        #==========Sub-heading=========================
      
        self._add_sub_heading(w_sheet,'Domestic',row_index)
        row_index +=1

        itemlist= ['domestic_inflation_rate','domestic_price_index',
                   'us_inflation_rate', 'us_price_index',
                   'relative_price_index', 'real_exchange_rate', 'nominal_exchange_rate']
        
        iep_parameters = {
            'domestic_inflation_rate': {'title': 'Domestic Inflation rate','value': None, 'units': 'PERCENT'},
            'domestic_price_index': {'title': 'Domestic Price Index','value': None, 'units': 'PERCENT'},
            'us_inflation_rate': {'title': 'US Inflation rate','value': None, 'units': 'PERCENT'}, 
            'us_price_index': {'title': 'US Price Index','value': None, 'units': 'NUMBER'},
            'relative_price_index': {'title': 'Relative Price Index','value': None, 'units': 'NUMBER'},
            'real_exchange_rate': {'title': 'Real Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
            'nominal_exchange_rate': {'title': 'Nominal Exchange Rate (LC/USD)','value': None, 'units': 'NUMBER'},
        }
       

        #-----------------------------------
        for item in itemlist:
            if item in iep_parameters:
                _unit =iep_parameters[item]['units']
                _unit = _unit.upper()
                cell_display_val = self.modelspec_cell_ref[_unit] if  _unit in self.modelspec_cell_ref else ''
                self._write_row_title_and_value3(w_sheet, colHeader, colValue, colUnits, row_index, 
                        iep_parameters[item]['title'], iep_parameters[item]['value'], cell_display_val, 
                        self._get_number_formats(_unit), _unit=='BLANK')
                
                if not (item in self.track_inputs['calc_inflation_price_index']):
                    self.track_inputs['calc_inflation_price_index'][item] = {}

                self.track_inputs['calc_inflation_price_index'][item]['row'] = row_index
                self.track_inputs['calc_inflation_price_index'][item]['unit'] = _unit
                self.track_inputs['calc_inflation_price_index'][item]['col'] = 6#redundant
                self.track_inputs['calc_inflation_price_index'][item]['value'] = None #self.sister_model.macroeconomic_parameters[item]['value'] 
                
                #go to newline
                row_index +=1 
                if item in ['domestic_price_index','us_price_index','relative_price_index'] :
                    row_index +=1 
                    if item =='domestic_price_index':
                        self._add_sub_heading(w_sheet,'Foreign',row_index)
                        row_index +=1

                    if item =='relative_price_index':
                        self._add_sub_heading(w_sheet,'Exchange Rate',row_index)
                        row_index +=1    

        
        # Sub-Section B skipp rows
        row_index +=1

        self._populate_cal_inflation_price_indices(w_sheet)
       
        return row_index 

    def _populate_cal_inflation_price_indices(self, w_sheet):
         
            
        self._transfer_value_tocellrange(w_sheet, 'sens', 
                        'calc_inflation_price_index', 'domestic_inflation_rate',  "Sens", None,
                        '0%')
                     
        #------Domestic Price Index-------------------------------------------
        # price index year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_inflation_price_index','header')        
        #header_cell------------------------
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # current dom inflation
        domestic_inflation_rate_r_index, c_index, found_state_domestic_inflation= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_inflation_rate')
        #domestic_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    domestic_inflation_rate_cell = cell.column + '$' \
                        + str(domestic_inflation_rate_r_index) if found_state_domestic_inflation else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_price_index_cell = prev_letter + str(cell.row) 
                    #=IF(I7=Inputs!$F$32;1;H11*(1+I10))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',1,' \
                                                               + prev_price_index_cell + '*(1+'+ domestic_inflation_rate_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
        
        #------US Inflation-------------------------------------------
        r_index, c_index, found_state= self._retrieve_cell_row_colm('sens','us_inflation_rate')
        us_inflation_cell = get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'

        
        us_inflation_rate_r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','us_inflation_rate')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(us_inflation_rate_r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(us_inflation_rate_r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=Sens!' + us_inflation_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
        
        
        #------US Price Index-------------------------------------------
        # price index year row
        header_r_index, c_index, found_state_header= self._retrieve_cell_row_colm('calc_inflation_price_index','header')        
        #header_cell------------------------
        # base period
        r_index, c_index, found_state= self._retrieve_cell_row_colm('timing','base_period')
        base_period_cell = 'Inputs!$'+ get_column_letter(c_index) + '$' + str(r_index) if found_state else '0'
        
        # current dom inflation
        us_inflation_rate_r_index, c_index, found_state_us_inflation= self._retrieve_cell_row_colm('calc_inflation_price_index','us_inflation_rate')
        #us_inflation_rate_cell--------------

        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','us_price_index')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    #
                    header_cell = cell.column + '$' + str(header_r_index) if found_state_header else '0'
       
                    #----
                    us_inflation_rate_cell = cell.column + '$' \
                        + str(us_inflation_rate_r_index) if found_state_us_inflation else '0'

                   
                    #get column
                    prev_col = column_index_from_string(cell.column)-1
                    prev_letter= get_column_letter(prev_col)
                    prev_price_index_cell = prev_letter + str(cell.row) 
                    #=IF(I7=Inputs!$F$32;1;H11*(1+I10))
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=IF('+ header_cell +'='+ base_period_cell + ',1,' \
                                                               + prev_price_index_cell + '*(1+'+ us_inflation_rate_cell+'))'
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         
        
        # Relative price index------------------------------
        us_price_index_r_index, c_index, found_state_us_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','us_price_index')
        #---------------
        domestic_price_index_r_index, c_index, found_state_dom_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','domestic_price_index') 
        #-------------------        
        relative_price_index_r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','relative_price_index')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(relative_price_index_r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(relative_price_index_r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    us_price_index_cell = cell.column +  str(us_price_index_r_index) if found_state_us_pi else '0'
                    domestic_price_index_cell = cell.column +  str(domestic_price_index_r_index) if found_state_dom_pi else '0'
       
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ domestic_price_index_cell + '/' + us_price_index_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         

        # Real Exchange Rate------------------------------
        r_index, c_index, found_state_us_pi= self._retrieve_cell_row_colm('macroeconomic_parameters','exchange_rate')
        exchange_rate_cell = '$' + get_column_letter(c_index) +  str(r_index) if found_state_us_pi else '0'
       
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','real_exchange_rate')
        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    w_sheet['%s%s'%(cell.column, cell.row)] = '=Inputs!'+ exchange_rate_cell 
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Linkedcell'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
         


        # Nominal Exchange Rate------------------------------
        relative_price_index_r_index, c_index, found_state_rel_pi= self._retrieve_cell_row_colm('calc_inflation_price_index','relative_price_index')
        #---------------
        real_exchange_rate_index, c_index, found_state_real_x_rate= self._retrieve_cell_row_colm('calc_inflation_price_index','real_exchange_rate') 
        #-------------------        
        r_index, c_index, found_state= self._retrieve_cell_row_colm('calc_inflation_price_index','nominal_exchange_rate')

        if found_state:
            span_ =self._get_span()
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                    relative_price_index_cell = cell.column +  str(relative_price_index_r_index) if found_state_rel_pi else '0'
                    real_eachange_rate_cell = cell.column +  str(real_exchange_rate_index) if found_state_real_x_rate else '0'
       
                    w_sheet['%s%s'%(cell.column, cell.row)] = '='+ relative_price_index_cell + '*' + real_eachange_rate_cell
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Calculation'
                    #w_sheet['%s%s'%(cell.column, cell.row)].number_format = '0%'
                              
    def _write_analytics_sheet(self, wb,):
        #---------------------------Worksheet 1---------------------------
        
        analytics_sheet = wb.create_sheet()
        analytics_sheet.title = 'Analytics'
        
      
        # Writing the first row of the csv	
        col = get_column_letter(1)
        analytics_sheet['%s%s'%(col, 1)].value = 'Analytic sheet'
        analytics_sheet['%s%s'%(col, 1)].style= 'Heading1'
    
        span_ = 6 
        total_wsheet_cols = 9 + int(span_)
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(analytics_sheet, row_index,total_wsheet_cols,6)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_analytics_gradient_section(wb, analytics_sheet, row_index,total_wsheet_cols)
        
        #add three row
        row_index += 3
       
        # hide---------------
        self._hide_empty_cells(analytics_sheet,2)

         #--set dim---
        rangelist= []
        rangelist.append({'start':14+2, 'end': 16+2, 'dim':2})
        rangelist.append({'start':3, 'end': 13+2, 'dim':10})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':1, 'dim':2})
        indexlist.append({'index':2, 'dim':40})
        
        
        
       
      
        if hasattr(self.sister_model,'npv_distribution'):
            for i in self.sister_model.employed_scenario_inputs.keys():
                Pass
            #skipp thre row
            analytics_sheet.append([])
            analytics_sheet.append(["","","","","",""])
            analytics_sheet.append(["","","","","",""])   
      

            col = get_column_letter(2)    
            analytics_sheet['%s%s'%(col, row_index)].value = 'MonteCarlo Simulation Graphs' #% (user)
            analytics_sheet['%s%s'%(col, row_index)].style= 'Heading1'

            npv_y= list(self.sister_model.npv_distribution['y'])
            npv_x_thousand = np.array(list(self.sister_model.npv_distribution['x']))/1000
            npv_x= list(np.round(npv_x_thousand,1))
            npv_cdf= list(self.sister_model.npv_distribution['cdf'])
            prob_npv_zero=self.sister_model.npv_distribution['prob_npv_zero']
             
            
            mean= round(self.sister_model.npv_distribution['mean']/1000,1)
            sd=round(math.sqrt(self.sister_model.npv_distribution['variance']/1000000),1)
           
            #-------Title plus blank
            npv_y.insert(0,'PDF')
            npv_y.insert(0,'')

            #-------Title plus blank
            npv_x.insert(0,'NPV')
            npv_x.insert(0,'')

            #-------Title plus blank
            npv_cdf.insert(0,'CDF')
            npv_cdf.insert(0,'')
            data = [
                npv_x,
                npv_y,
                npv_cdf,
            ]      
            #print(data)
            # write content of each row in 1st and 2nd
            # column of the active sheet respectively .
            for row in data:
                analytics_sheet.append(row)
            
            #last row for NPV-ZERO
            npvz= []
            npvz.append('')
            npvz.append('Prob. NPV >= Zero')
            for i in range(max(len(data[0])-1-1,0)):
                npvz.append(prob_npv_zero) 
            analytics_sheet.append(npvz)
            
            npvless= []
            npvless.append('')
            npvless.append('Prob. NPV less Zero')
            for i in range(max(len(data[0])-1-1,0)):
                npvless.append(1-prob_npv_zero) 
            analytics_sheet.append(npvless)
            
            #------------------------------------
            stride= max(len(data[0])-1-1,0)
            #set border
            self._set_chart_data_border(analytics_sheet, row_index, stride)
            
          
            
            latest_row= row_index  + 2  
            last_col = 3 + len(data[0])-3 # three empty space subtract
            first_col =3
            heading_row_index= row_index +1

            prob_npv_zero=round(prob_npv_zero*100)

            prob_npv_zero_stat= f'Prob NPV > 0 is {prob_npv_zero} %'
            #6.-----------------------Draw Chart------------------------------
            self.drawChart2(analytics_sheet, latest_row, latest_row,
                first_col, last_col,	latest_row + 4 + 5, 
                heading_row_index, heading_row_index,
                f"PDF {prob_npv_zero_stat}",f"NPV (000)  u:{mean}  sd: {sd} ")
            #-------------------------------------------------
            added_rows=6
            self._cal_mean_variance(analytics_sheet, row_index, added_rows, stride)

          
        #set dims----
        self._set_column_dim(analytics_sheet,rangelist,indexlist)

    def _cal_mean_variance(self ,w_sheet, base_index, added_rows, stride):
        row_index= base_index + added_rows
        row_i = row_index +1
        number_format ='_(* #,##0.0##_);_(* \(#,##0.0##\);_(* "-"??_);_(@_)'

        first_slice_point = get_column_letter(3) + str(row_i )# skip heeder and first row 
        second_slice_point = get_column_letter(3 + int(stride-1)) + str(row_i+2) # D39

        # last cell
        header_column = get_column_letter(2) 
        last_column = get_column_letter(3 + int(stride-1+1)) 
        last_column_less_one = get_column_letter(3 + int(stride-1))
        
        #Average Sum
        first_row_first_point = get_column_letter(3) + str(row_i )
        first_row_last_point = last_column_less_one + str(row_i )
        range_ = first_row_first_point + ':' + first_row_last_point
        denominator_ ='COUNTA(' + range_ + ')'
        w_sheet['%s%s'%(last_column, row_i)]= '=Sum('+ range_ + ')/' + denominator_ 
        w_sheet['%s%s'%(header_column, row_i)]= 'Mean' 
        w_sheet['%s%s'%(last_column, row_i)].number_format = number_format
        w_sheet['%s%s'%(last_column, row_i)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')     
      
        #Sum Of Difference ((COUNTA(C47:N47)-1))
        third_row_first_point = get_column_letter(3) + str(row_i+2 )
        third_row_last_point = last_column_less_one + str(row_i+2 )
        range_ = third_row_first_point + ':' + third_row_last_point
        denominator_ ='COUNTA(' + range_ + ')-1'
        w_sheet['%s%s'%(last_column, row_i + 2)]= '=Sum('+ range_ + ')/(' + denominator_ + ')'
        w_sheet['%s%s'%(header_column, row_i + 2)]= 'Variance'
        w_sheet['%s%s'%(last_column, row_i + 2)].number_format = number_format
        w_sheet['%s%s'%(last_column, row_i + 2)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')  

        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                #w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'  

                if  cell.row == row_i : #dict_['1']:  
                    npv_cell = cell.column + str(row_i -added_rows)  
                    pdf_cell = cell.column + str(row_i - added_rows +1)    
                    w_sheet['%s%s'%(cell.column, cell.row)]= '='+ npv_cell + '*' + pdf_cell
                elif (cell.row == row_i + 1):
                    #Diff
                    npv_cell = cell.column + str(cell.row-1)  
                    average_cell = last_column + str(cell.row-1 )    
                    w_sheet['%s%s'%(cell.column, cell.row)]= '='+ npv_cell + '-' + average_cell

                elif (cell.row == row_i + 2):
                    #summm sqr
                    prev_cell = cell.column + str(cell.row-1)     
                    w_sheet['%s%s'%(cell.column, cell.row)]= '='+ prev_cell + '^2' 

                w_sheet['%s%s'%(cell.column, cell.row)].number_format = number_format   

        #---------List Parameters used in Simulation::::::
        scenario_var =self.sister_model.employed_scenario_inputs
        next_toheader_column = get_column_letter(3) 
        list_of_paras= ''
        for i in scenario_var.keys():
            if len(list_of_paras)>0:
                list_of_paras += ', ' + i
            else:
                #first ele
                list_of_paras =  i
        #Write it 25 lines below
        w_sheet['%s%s'%(header_column, row_i + 2 + 17 )]= f'{len(scenario_var)} Parameters Employed'
        w_sheet['%s%s'%(next_toheader_column, row_i + 2 + 17 )]= list_of_paras

            
    def _set_chart_data_border(self ,w_sheet, row_index, stride):
        row_i= row_index+1
        # format table
        number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
        number_format_2 ='_(* #,##0.0###_);_(* \(#,##0.0###\);_(* "-"??_);_(@_)'
        first_slice_point = get_column_letter(3) + str(row_i )# skip heeder and first row 
        second_slice_point = get_column_letter(3 + int(stride-1)) + str(row_i  + 4) # D39
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'  

                if  cell.row > row_i:        
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = number_format_2
                else:
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = number_format
                w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                sz=11.0, color='FF0070C0', scheme='minor')  

               
    def _add_analytics_gradient_section(self,wb, w_sheet, row_index,total_wsheet_cols):
         #==========Lengend=========================
        col = get_column_letter(1)
        w_sheet['%s%s'%(col, row_index)].value = 'Parameters Gradients' #% (user)
        w_sheet['%s%s'%(col, row_index)].style= 'Heading1'
      
        self._set_thick_bottom_border_range( w_sheet,  row_index, 1, total_wsheet_cols)
        row_index +=1
        #===================================
     
        row_index +=1


        data = []
        data.append(["",   'Parameter' ,   'NPV Gradient'  ,    '@npv_0','Direction', '', 'Financial comment Viable Project' ])
        #get  item_trcked, us list
        #rrrrrrrrrrrrrrrrrrrrrrrrrrrrr
        
        para_list=[]
        if hasattr(self.sister_model, 'sens_grad'):
            df = self.sister_model.sens_grad
            for i in range(len(df)):
                direction_symbol =""
                direction =0
                pos_neg_symbol =""
                parameter_ =df.iloc[i]['parameter']
                comments =''
                x_npv_0= round(df.iloc[i]['x_npv_0'],2)
                #print(self.parameter_unit_dict)
                if df.iloc[i]['gradient']>0:
                    direction_symbol ="↗"
                    direction =1
                    pos_neg_symbol ="+"
                    comments =f'{parameter_} should be at least {x_npv_0}'
                elif df.iloc[i]['gradient']<0:
                    direction_symbol ="↘"
                    direction =-1
                    pos_neg_symbol ="-"
                    comments =f'{parameter_} should not exceed {x_npv_0}'

                if hasattr(self.sister_model, 'parameter_unit_dict'):
                    if parameter_ in self.sister_model.parameter_unit_dict:
                        title =self.sister_model.parameter_unit_dict[parameter_]['title']
                    else:
                        title =parameter_

                   
                    para_list.append({'name':parameter_, 
                                     'number_format': self.sister_model.parameter_unit_dict[parameter_]['number_format'],
                                     'direction': direction})
                else:
                    para_list.append({'name':parameter_, 'number_format': 'NUMBER','direction': direction})
                    title =parameter_

                
                data.append(["",  title, df.iloc[i]['gradient'],  
                                   df.iloc[i]['x_npv_0'],direction_symbol, pos_neg_symbol, comments]) 
                
               
        #print(para_list)
        # add column headings. NB. these must be strings
        w_sheet.append(["","","","","",""])
        w_sheet.append(["","","","","",""])
      
        #row_index +=2
        start_row = row_index
        

       
        for row in data:
            w_sheet.append(row)
            row_index +=1
        
        end_row= row_index 
       
       # format table
        first_slice_point = get_column_letter(2) + str(start_row + 1 )# skip heeder and first row 
        second_slice_point = get_column_letter(2 + int(10)) + str(end_row) # D39
        for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
            for cell in cellObj:
                this_row= min(max(cell.row-(start_row+2),0),len(para_list)-1)
                
                   
                if cell.column in ['B']:

                    #italic
                    if cell.row-(start_row+2)>=0:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                        sz=11.0, color='FF0070C0', scheme='minor')   
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                        sz=11.0, color='FF0070C0', scheme='minor')
                
                elif cell.column in ['D'] and cell.row-(start_row+2)>=0:
                    
                    #print(this_row,len(para_list))
                    w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format = self._get_number_formats(para_list[this_row]['number_format'])
                    w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                #exclude C
                else:
                    if not cell.row == start_row + 1:
                        w_sheet['%s%s'%(cell.column, cell.row)].style = 'Output4'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')
                    else:
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Arial',bold=True, 
                                                     sz=11.0, color='FF0070C0', scheme='minor')                                 
                    
                    w_sheet['%s%s'%(cell.column, cell.row)].number_format ='_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)'
                   
                if cell.row-(start_row+2)>=0 and para_list[this_row]['direction']==1 and cell.column in ['B'] :

                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FFFFCC99', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FFFFCC99', scheme='minor')

                    #.fill = PatternFill(fgColor='FFFFCC99', patternType='solid', fill_type='solid')
                elif cell.row-(start_row+2)>=0 and para_list[this_row]['direction']==-1 and cell.column in ['B']:
                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FFF00000', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FFF00000', scheme='minor')                                                     
                 
                   #.fill = PatternFill(fgColor='70c4f4', patternType='solid', fill_type='solid')
       
                elif cell.row-(start_row+2)>=0 and cell.column in ['B']:
                    if cell.column in ['B']:
                        w_sheet['%s%s'%(cell.column, cell.row)].number_format = 'General'
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, italic=True, 
                                                            sz=11.0, color='FF3F3F3F', scheme='minor')   
                    else:                     
                        w_sheet['%s%s'%(cell.column, cell.row)].font = Font(name='Calibri',bold=False, 
                                                     sz=11.0, color='FF3F3F3F', scheme='minor')       
                
                   #.fill = PatternFill(fgColor='d0ebfb', patternType='solid', fill_type='solid')
       
        return row_index 

    def _write_sens_sheet(self, wb):
        #---------------------------Worksheet 1---------------------------
        
        sens_sheet = wb.create_sheet()
        sens_sheet.title = 'Sens'
        
      
        # Writing the first row of the csv	
        col = get_column_letter(1)
        sens_sheet['%s%s'%(col, 1)].value = 'Sensitivity analysis sheet'
        sens_sheet['%s%s'%(col, 1)].style= 'Heading1'
    
        span_ = 6 
        total_wsheet_cols = 9 + int(span_)
        #==========Lengend Section =========================
        row_index =3
        row_index = self._add_legend_section(sens_sheet, row_index,total_wsheet_cols,6)
        #======================================+++++++++++++++
        
        #==========Model Specification Section =========================
        row_index +=2
        row_index = self._add_sensitivity_section(wb, sens_sheet, row_index,total_wsheet_cols)

        if not 'sens_parameters' in self.track_inputs:
            self.track_inputs['sens_parameters']= {}# insert inner

        if not ('endpoint' in self.track_inputs['sens_parameters']):
            self.track_inputs['sens_parameters']['endpoint'] = {}
  
        if 'endpoint' in self.track_inputs['sens_parameters']:
            self.track_inputs['sens_parameters']['endpoint']['row'] = row_index
        #======================================+++++++++++++++

       
        # hide---------------
        self._hide_empty_cells(sens_sheet)

         #--set dim---
        rangelist= []
        rangelist.append({'start':14, 'end': 16, 'dim':2})
        rangelist.append({'start':3, 'end': 13, 'dim':10})

        indexlist= []
        #indexlist.append({'index':margin_offset+1, 'dim':5})
        indexlist.append({'index':1, 'dim':2})
        indexlist.append({'index':2, 'dim':40})

        #set dims----
        self._set_column_dim(sens_sheet,rangelist,indexlist)
          
     
    def drawChart(self, wkSheet,
                min_row,max_row,
                min_col,max_col,
                chart_row_index_pos, 
                cat_min_row, cat_max_row, 
                chart_title, x_axis_title):
        #-----------------------Draw Charts
        chart = BarChart()

        data = Reference(worksheet=wkSheet,
                        min_row=min_row ,
                        max_row=max_row,
                        min_col=min_col-1,
                        max_col=max_col,
                        )

        chart.add_data(data, from_rows=True, titles_from_data=True)
        cats = Reference(worksheet=wkSheet,
                        min_row=cat_min_row,
                        max_row=cat_max_row,
                        min_col=min_col,
                        max_col=max_col,
                        )
        chart.set_categories(cats)
        col_first = get_column_letter(min_col)# max [0,first_col]
        chart_draw_cell = col_first + str(chart_row_index_pos)# "F3"
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = chart_title

        # You can play with this by choosing any number between 1 and 48
        chart.style = 24
        wkSheet.add_chart(chart, chart_draw_cell)

       
         
       
    def drawChart2(self, wkSheet,
                min_row,max_row,
                min_col,max_col,
                chart_row_index_pos, 
                cat_min_row, cat_max_row, 
                chart_title, x_axis_title):
        #-----------------------Draw Charts
        #chart2 = BarChart()

       
       
        col_first = get_column_letter(min_col)# max [0,first_col]
        chart_draw_cell = col_first + str(chart_row_index_pos)# "F3"
     
     
        # setup the chart
        chart = LineChart()
        #chart.drawing.name = 'This is my chart'

        # setup and append the first series
        values1 = Reference(wkSheet, min_col = min_col, max_col=max_col, min_row = min_row, max_row = max_row)
        series1 = Series(values1, title="PDF")
        chart.append(series1)

        # setup and append the second series
        values2 = Reference(wkSheet, min_col = min_col, max_col=max_col, min_row = min_row+1, max_row = max_row+1)
        series2 = Series(values2, title="CDF")
        chart.append(series2)

         # setup and append the third series
        values3 = Reference(wkSheet, min_col = min_col, max_col=max_col, min_row = min_row+2, max_row = max_row+2)
        series3 = Series(values3, title="NPV +")
        chart.append(series3)


          # setup and append the third series
        values4 = Reference(wkSheet, min_col = min_col, max_col=max_col, min_row = min_row+3, max_row = max_row+3)
        series4 = Series(values4, title="NPV -")
        #print(f'type of values4 4: {type(values4)}')
        chart.append(series4)

        cats = Reference(worksheet=wkSheet,
                        min_row=cat_min_row,
                        max_row=cat_max_row,
                        min_col=min_col,
                        max_col=max_col,
                        )
        chart.set_categories(cats)

        chart.x_axis.title = x_axis_title
        chart.y_axis.title = chart_title

        wkSheet.add_chart(chart, chart_draw_cell)
        #wkSheet.add_chart(chart2, chart_draw_cell)
        
       
          
    #----------------Check Additiona Methods---------------------------------------------
            
    def _get_out_values_as_dict(self,wb):
        #list_dict =[]
        output_list =[]
        inputs_sheet = wb['Inputs']
        cal_sheet = wb['Calc']
        output_sheet = wb['Outputs']
        cf_sheet = wb['CF']
        sens_sheet = wb['Sens']

        list_dict =[
            {'wsheet':cal_sheet,'section':'cal_production_sales_nominal','item': 'gross_sales_revenue'},        
        ]
       
        for i in list_dict:
            dict_ ={}            
            dummy=self._get_excel_row_as_list(i['wsheet'],i['section'], i['item'])
            dict_['list']=dummy
            dict_['name']=i['item']
            output_list.append(dict_)
        #print(output_list)

    def _get_excel_row_as_list(self,w_sheet,section_, item):
        r_index, c_index, found_state= self._retrieve_cell_row_colm(section_, item)
        list_ =[]
        if found_state:
            span_ = self.sister_model.timing_assumptions['operation_duration']['value'] if 'operation_duration' in self.sister_model.timing_assumptions else 0
            first_slice_point = get_column_letter(9) + str(r_index)# D07
            second_slice_point = get_column_letter(9 + int(span_)) + str(r_index) # D39
            #loop each cell of this row D15:G15
            #----------->
            for cellObj in w_sheet[first_slice_point:second_slice_point]:# along row axis
                for cell in cellObj:
                   #----
                   value_ =w_sheet['%s%s'%(cell.column, cell.row)].value
                   #print(w_sheet['%s%s'%(cell.column, cell.row+1)].value)
                   list_.append(value_)
        return list_
        #=$F$103*I102
    def _set_Underline_row(self, wkSheet, row):
        
        # Let's create a style template for the header row
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.border = Border(bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="center", vertical="center")

        # Now let's apply this to all first row (header) cells
        header_row = wkSheet[row]
        for cell in header_row:
            cell.style = header

 
    def _set_thin_bottom_border_range(self, wkSheet, writing_row_index,
                            start_col_index, end_col_index,color_str='FFFFFF00'):
        

        latest_row = writing_row_index

        col_start = get_column_letter(start_col_index)
        col_end = get_column_letter(end_col_index)

        first_slice_point = col_start + str(writing_row_index ) # D200
        second_slice_point = col_end + str(writing_row_index ) # F200

        thick_border_side = Side(border_style="thin", color=color_str)
        straight_thick_border = Border(
                                bottom=thick_border_side)

        #loop each cell of this row D20:Z20
        #-------->
        for cellObj in wkSheet[first_slice_point:second_slice_point]:
            for cell in cellObj:
                #get column
                new_column_letter = cell.column
                # |
                # |
                # |
                # v
                cell.border = straight_thick_border

                
                latest_row = cell.row

        #------------------------------
        return latest_row

   
    def _set_border_all(self, wkSheet, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

        rows = wkSheet.iter_rows(cell_range)
        for row in rows:
            for cell in row:
                cell.border = border

    # def set_border(ws, cell_range):
    #     thin = Side(border_style="thin", color="000000")
    #     for row in ws[cell_range]:
    #         for cell in row:
    #             cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    def _set_border_yellow(self, wkSheet, cell_range):
        rows = list(wkSheet[cell_range])
        side = Side(border_style='thick', color="FFF00000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border            
  
    def _set_border(self, wkSheet, cell_range):
        rows = list(wkSheet[cell_range])
        side = Side(border_style='thick', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border            
    def _remove_borders(self, wkSheet):
        side = Side(border_style=None)
        no_border = Border(
                        left=side, 
                        right=side, 
                        top=side, 
                        bottom=side,
                    )
        # Loop through all cells in te worksheet
        for row in wkSheet:
            for cell in row:
                # Apply colorless and borderless styles
                #cell.fill = no_fill
                cell.border = no_border

   