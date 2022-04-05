
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

ws = wb.create_sheet('New Sheet')

for number in range(1,100): #Generates 99 "ip" address in the Column A;
    ws['A{}'.format(number)].value= "192.168.1.{}".format(number)

data_val = DataValidation(type="list",formula1='=$A:$A') #You can change =$A:$A with a smaller range like =A1:A9
ws.add_data_validation(data_val)

data_val.add(ws["B1"]) #If you go to the cell B1 you will find a drop down list with all the values from the column A

wb.save('Test.xlsx')




from openpyxl import load_workbook
book = load_workbook('table.xlsx')
sheet = book.active

tables = sheet._tables
table_name = 'Table1'

def find_table(table_name, tables):
    for table in tables:
        if table.displayName == table_name:
            return table.ref


table_range = find_table(table_name, tables)




import openpyxl

# create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# populate some sample data    
worksheet["A1"] = "Fruit"
worksheet["B1"] = "Color"
worksheet["A2"] = "Apple"
worksheet["B2"] = "Red"
worksheet["A3"] = "Banana"
worksheet["B3"] = "Yellow"
worksheet["A4"] = "Coconut"
worksheet["B4"] = "Brown"

# define a table style
mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                      showRowStripes=True)
# create a table
table = openpyxl.worksheet.table.Table(ref='A1:B4',
                                       displayName='FruitColors',
                                       tableStyleInfo=mediumStyle)
# add the table to the worksheet
worksheet.add_table(table)

# save the workbook file
workbook.save('fruit.xlsx')



import pandas as pd

df = pd.DataFrame({'Col1': [1,2,3], 'Col2': list('abc')})

filename = 'so58326392.xlsx'
sheetname = 'mySheet'
with pd.ExcelWriter(filename) as writer:
    if not df.index.name:
        df.index.name = 'Index'
    df.to_excel(writer, sheet_name=sheetname)

import openpyxl
wb = openpyxl.load_workbook(filename = filename)
tab = openpyxl.worksheet.table.Table(displayName="df", ref=f'A1:{chr(len(df.columns)+64)}{len(df)+1}')
wb[sheetname].add_table(tab)
wb.save(filename)


import pandas as pd
import xlsxwriter as xl

df = pd.DataFrame({'Col1': [1,2,3], 'Col2': list('abc')})

filename = 'output.xlsx'
sheetname = 'Table'
tablename = 'TEST'

(rows, cols) = df.shape
data = df.to_dict('split')['data']
headers = []
for col in df.columns:
    headers.append({'header':col})

wb = xl.Workbook(filename)
ws = wb.add_worksheet()

ws.add_table(0, 0, rows, cols-1,
    {'name': tablename
    ,'data': data
    ,'columns': headers})

wb.close()


""" 
Worksheet Tables
Worksheet tables are references to groups of cells. This makes certain operations such as styling the cells in a table easier.

Creating a table
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
'''
ws.add_table(tab)
wb.save("table.xlsx")
Table names must be unique within a workbook. By default tables are created with a header from the first row and filters for all the columns and table headers and column headings must always contain strings.

Warning

In write-only mode you must add column headings to tables manually and the values must always be the same as the values of the corresponding cells (ee below for an example of how to do this), otherwise Excel may consider the file invalid and remove the table.

Styles are managed using the the TableStyleInfo object. This allows you to stripe rows or columns and apply the different colour schemes.

Working with Tables
ws.tables is a dictionary-like object of all the tables in a particular worksheet:

>>> ws.tables
{"Table1",  <openpyxl.worksheet.table.Table object>}
Get Table by name or range
>>> ws.tables["Table1"]
or
>>> ws.tables["A1:D10"]
Iterate through all tables in a worksheet
>>> for table in ws.tables.values():
>>>    print(table)
Get table name and range of all tables in a worksheet
Returns a list of table name and their ranges.

>>> ws.tables.items()
>>> [("Table1", "A1:D10")]
Delete a table
>>> del ws.tables["Table1"]
The number of tables in a worksheet
>>> len(ws.tables)
>>> 1
Manually adding column headings
In write-only mode you can either only add tables without headings:

>>> table.headerRowCount = False
Or initialise the column headings manually:

>>> headings = ["Fruit", "2011", "2012", "2013", "2014"] # all values must be strings
>>> table._initialise_columns()
>>> for column, value in zip(table.tableColumns, headings):
    column.name = value 
    
    
    """

""" 

Workbook Protection
To prevent other users from viewing hidden worksheets, adding, moving, deleting, or hiding worksheets, and renaming worksheets, you can protect the structure of your workbook with a password. The password can be set using the openpyxl.workbook.protection.WorkbookProtection.workbookPassword() property

>>> wb.security.workbookPassword = '...'
>>> wb.security.lockStructure = True
Similarly removing change tracking and change history from a shared workbook can be prevented by setting another password. This password can be set using the openpyxl.workbook.protection.WorkbookProtection.revisionsPassword() property

>>> wb.security.revisionsPassword = '...'
Other properties on the openpyxl.workbook.protection.WorkbookProtection object control exactly what restrictions are in place, but these will only be enforced if the appropriate password is set.

Specific setter functions are provided if you need to set the raw password value without using the default hashing algorithm - e.g.

hashed_password = ...
wb.security.set_workbook_password(hashed_password, already_hashed=True)
Worksheet Protection
Various aspects of a worksheet can also be locked by setting attributes on the openpyxl.worksheet.protection.SheetProtection object. Unlike workbook protection, sheet protection may be enabled with or without using a password. Sheet protection is enabled using the openpxyl.worksheet.protection.SheetProtection.sheet attribute or calling enable() or disable():

>>> ws = wb.active
>>> ws.protection.sheet = True
>>> ws.protection.enable()
>>> ws.protection.disable()
If no password is specified, users can disable configured sheet protection without specifying a password. Otherwise they must supply a password to change configured protections. The password is set using the openpxyl.worksheet.protection.SheetProtection.password() property

>>> ws = wb.active
>>> ws.protection.password = '...'

 """